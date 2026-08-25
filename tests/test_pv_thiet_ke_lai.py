"""Tham chiếu bị từ chối hai lần → AI thiết kế lại nhân vật → Excel cập nhật dàn + mọi cảnh.

Thiết kế nằm ở MỘT chỗ (sheet `characters`); khối khoá trong `img_prompt` của
mọi cảnh do mã chèn, nên đổi thiết kế là mọi cảnh đổi theo. Không gọi mạng.
"""

from __future__ import annotations

from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from core.prompt_visuals import DUOI_CHAN_DUNG
from ui_qt.trang_prompt_visuals import TrangPromptVisuals


def _xlsx(tmp_path):
    wb = Workbook()
    ws = wb.active; ws.title = "characters"
    ws.append(["id", "role", "name", "english_prompt", "sheet_prompt", "notes"])
    ws.append(["nv4", "the cat", "Puss", "orange cat; outfit at this stage: no clothes",
               "orange cat; outfit at this stage: no clothes" + DUOI_CHAN_DUNG + " Style: 3D", ""])
    ws.append(["nv4b", "the cat", "Puss", "orange cat; outfit at this stage: feathered hat, boots",
               "orange cat; outfit at this stage: feathered hat, boots" + DUOI_CHAN_DUNG + " Style: 3D", ""])
    ws.append(["nv1", "hero", "", "a young man", "a young man" + DUOI_CHAN_DUNG + " Style: 3D", ""])
    sc = wb.create_sheet("scenes")
    sc.append(["scene_id", "img_prompt", "video_prompt", "characters_used", "reference_files"])
    sc.append([1, "Wide shot of nv4b [reference image 1]\nREFERENCE IMAGES are attached, in this order:\n"
                  "- reference image 1 = nv4b, the the cat: orange cat; outfit at this stage: feathered hat, boots\n"
                  "- reference image 2 = nv1, the hero: a young man", "clip 1", "nv4b, nv1", '["nv4b.png", "nv1.png"]'])
    sc.append([2, "nv1 alone\nREFERENCE IMAGES are attached, in this order:\n- reference image 1 = nv1, the hero: a young man",
               "clip 2", "nv1", '["nv1.png"]'])
    p = tmp_path / "prompts.xlsx"; wb.save(str(p))
    return str(p)


def test_ap_thiet_ke_lai_vao_excel(tmp_path):
    p = _xlsx(tmp_path)
    ds, so_canh = TrangPromptVisuals._ap_thiet_ke_lai_xlsx(p, "nv4b", "grey cat with a white chest")
    assert sorted(i for i, _ in ds) == ["nv4", "nv4b"] and so_canh == 1
    wb = load_workbook(p, read_only=True, data_only=True)
    dan = {r[0]: r for r in wb["characters"].iter_rows(min_row=2, values_only=True)}
    assert dan["nv4"][3] == "grey cat with a white chest; outfit at this stage: no clothes"
    assert dan["nv4b"][4].startswith("grey cat with a white chest; outfit at this stage: feathered hat, boots" + DUOI_CHAN_DUNG)
    assert dan["nv4b"][4].endswith(" Style: 3D")
    assert dan["nv1"][3] == "a young man"
    canh = {r[0]: r for r in wb["scenes"].iter_rows(min_row=2, values_only=True)}
    assert "reference image 1 = nv4b, the the cat: grey cat with a white chest; outfit at this stage: feathered hat, boots" in canh[1][1]
    assert canh[2][1].endswith("a young man") and canh[1][2] == "clip 1"
    wb.close()


def test_hong_hai_lan_thi_goi_thiet_ke_lai_mot_lan():
    goi = []

    class _Trang:
        _tc_dang_cho = {"k": ("nv4b", "x.xlsx", "p")}
        _tc_da_thu_lai = {"nv4b"}
        _tc_da_thiet_ke_lai = set()
        _tc_thieu = []
        _app = SimpleNamespace(client=object())
        nhat_ky = []

        def _ghi(self, chu):
            self.nhat_ky.append(chu)

        def _thiet_ke_lai_tham_chieu(self, duong, ma_id, ly_do):
            goi.append((duong, ma_id, ly_do))

        def _bao_du_tham_chieu(self):
            pass

        _tham_chieu_hong = TrangPromptVisuals._tham_chieu_hong

    t = _Trang()
    t._tham_chieu_hong("k", SimpleNamespace(message="bị từ chối"))
    assert goi == [("x.xlsx", "nv4b", "bị từ chối")] and t._tc_thieu == []
    # Lần sau cùng nhân vật (id gốc nv4) không thiết kế lại nữa → báo thiếu.
    t._tc_dang_cho["k2"] = ("nv4", "x.xlsx", "p")
    t._tc_da_thu_lai.add("nv4")
    t._tham_chieu_hong("k2", SimpleNamespace(message="bị từ chối"))
    assert len(goi) == 1 and t._tc_thieu == ["nv4"]
