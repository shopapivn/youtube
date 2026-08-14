"""Nạp sẵn một khâu bằng file của khách — và từ chối cho ra hồn khi file sai.

Chủ dự án, 14/08/2026: *"có thể đẩy content sẵn có viết chỗ khác vào theo dạng
txt, hay là chỗ excel có thể tải file mẫu về điền và up lên"*.

Phần đáng canh nhất không phải chuyện chép file — mà là **lời từ chối**. Khâu
sau tiêu tiền thật: nạp nhầm một Excel thiếu cột `srt_start` thì tool không
chết ngay, nó chạy tiếp, tạo 111 tấm ảnh, rồi mới hỏng ở khâu dựng. Lúc đó
tiền đã đi. Nên file phải bị soi TRƯỚC, và câu từ chối phải nói rõ thiếu đúng
cái gì.

Không bài nào gọi mạng.
"""

from __future__ import annotations

import os

import pytest

from core.nap_san import (
    LoiNapSan, co_mau, kiem_file, kieu_file_cua_khau, nap_file, viet_mau,
)

KICH_BAN_THAT = ("Câu chuyện bắt đầu vào một buổi chiều. " * 20)


@pytest.fixture
def luot(tmp_path):
    d = tmp_path / "L99"
    d.mkdir()
    return str(d)


class TestKichBan:
    def test_nap_duoc_file_txt_binh_thuong(self, tmp_path, luot):
        nguon = tmp_path / "kich-ban-cua-toi.txt"
        nguon.write_text(KICH_BAN_THAT, encoding="utf-8")
        dich = nap_file(luot, "kich-ban", str(nguon))
        assert os.path.basename(dich) == "1-kich-ban.txt", \
            "phải đổi thành đúng tên khâu sau đi tìm"
        assert open(dich, encoding="utf-8").read() == KICH_BAN_THAT

    def test_tu_choi_file_qua_ngan(self, tmp_path, luot):
        nguon = tmp_path / "nham.txt"
        nguon.write_text("vài chữ", encoding="utf-8")
        with pytest.raises(LoiNapSan, match="quá ngắn"):
            nap_file(luot, "kich-ban", str(nguon))

    def test_tu_choi_dung_duoi(self, tmp_path, luot):
        nguon = tmp_path / "kich-ban.docx"
        nguon.write_text(KICH_BAN_THAT, encoding="utf-8")
        with pytest.raises(LoiNapSan, match="nhận file"):
            nap_file(luot, "kich-ban", str(nguon))

    def test_tu_choi_file_rong(self, tmp_path, luot):
        nguon = tmp_path / "rong.txt"
        nguon.write_text("", encoding="utf-8")
        with pytest.raises(LoiNapSan, match="rỗng"):
            nap_file(luot, "kich-ban", str(nguon))


class TestPhuDe:
    def test_nap_duoc_srt(self, tmp_path, luot):
        nguon = tmp_path / "a.srt"
        nguon.write_text("1\n00:00:01,000 --> 00:00:04,000\nXin chào\n",
                         encoding="utf-8")
        assert os.path.basename(nap_file(luot, "phu-de", str(nguon))) == "3-phu-de.srt"

    def test_tu_choi_file_khong_phai_phu_de(self, tmp_path, luot):
        nguon = tmp_path / "a.srt"
        nguon.write_text("chỉ là chữ thường thôi", encoding="utf-8")
        with pytest.raises(LoiNapSan, match="không giống phụ đề"):
            nap_file(luot, "phu-de", str(nguon))


openpyxl = pytest.importorskip("openpyxl")


class TestBangCanh:
    """Excel là chỗ dễ điền sai nhất, nên soi kỹ nhất."""

    def _sach(self, tmp_path, ten, cot, so_dong=2):
        from openpyxl import Workbook

        s = Workbook()
        t = s.active
        t.title = "scenes"
        t.append(list(cot))
        for i in range(so_dong):
            t.append([i + 1] + [""] * (len(cot) - 1))
        duong = tmp_path / ten
        s.save(duong)
        return str(duong)

    def test_file_mau_tai_ve_roi_nap_len_duoc_ngay(self, tmp_path, luot):
        """Vòng tròn khép kín: tải mẫu -> nạp lên. Hỏng chỗ này là tính năng vô dụng."""
        mau = str(tmp_path / "mau.xlsx")
        viet_mau("bang-canh", mau)
        dich = nap_file(luot, "bang-canh", mau)
        assert os.path.basename(dich) == "4-canh.xlsx"

    def test_file_mau_co_trang_huong_dan(self, tmp_path):
        from openpyxl import load_workbook

        mau = str(tmp_path / "mau.xlsx")
        viet_mau("bang-canh", mau)
        s = load_workbook(mau)
        assert "scenes" in s.sheetnames
        assert "huong-dan" in s.sheetnames, "khách không biết cột nào bắt buộc"
        assert s["scenes"].max_row >= 3, "phải có ít nhất hai dòng mẫu"

    def test_tu_choi_khi_thieu_trang_scenes(self, tmp_path, luot):
        from openpyxl import Workbook

        s = Workbook()
        s.active.title = "linh-tinh"
        duong = str(tmp_path / "sai.xlsx")
        s.save(duong)
        with pytest.raises(LoiNapSan, match="scenes"):
            nap_file(luot, "bang-canh", duong)

    def test_tu_choi_va_NOI_RO_cot_nao_thieu(self, tmp_path, luot):
        duong = self._sach(tmp_path, "thieu.xlsx",
                           ("scene_id", "duration", "img_prompt"))
        with pytest.raises(LoiNapSan, match="srt_start"):
            nap_file(luot, "bang-canh", duong)

    def test_tu_choi_khi_chua_co_dong_canh_nao(self, tmp_path, luot):
        duong = self._sach(tmp_path, "trong.xlsx",
                           ("scene_id", "srt_start", "duration", "img_prompt"),
                           so_dong=0)
        with pytest.raises(LoiNapSan, match="chưa có dòng"):
            nap_file(luot, "bang-canh", duong)

    def test_khong_doi_du_25_cot(self, tmp_path, luot):
        """Đòi đủ 25 cột là bắt khách điền hai chục ô tool không bao giờ nhìn."""
        duong = self._sach(tmp_path, "vua_du.xlsx",
                           ("scene_id", "srt_start", "duration", "img_prompt"))
        assert nap_file(luot, "bang-canh", duong)

    def test_khong_chep_gi_khi_file_sai(self, tmp_path, luot):
        """Soi trước rồi mới chép — không để lại một nửa."""
        duong = self._sach(tmp_path, "thieu.xlsx", ("scene_id",))
        with pytest.raises(LoiNapSan):
            nap_file(luot, "bang-canh", duong)
        assert not os.path.exists(os.path.join(luot, "4-canh.xlsx"))


class TestThuMuc:
    """Khâu ảnh/clip/bìa đẻ ra cả thư mục, nên nạp bằng thư mục."""

    def test_nap_ca_thu_muc_anh(self, tmp_path, luot):
        nguon = tmp_path / "anh-cua-toi"
        nguon.mkdir()
        for i in range(3):
            (nguon / "{0}.png".format(i + 1)).write_bytes(b"\x89PNG-gia-vo")
        dich = nap_file(luot, "anh", str(nguon))
        assert len(os.listdir(dich)) == 3

    def test_tu_choi_thu_muc_khong_co_anh(self, tmp_path, luot):
        nguon = tmp_path / "rong"
        nguon.mkdir()
        (nguon / "ghi-chu.txt").write_text("x", encoding="utf-8")
        with pytest.raises(LoiNapSan, match="không có file"):
            nap_file(luot, "anh", str(nguon))


class TestMau:
    def test_chi_hai_khau_co_mau(self):
        assert co_mau("kich-ban") and co_mau("bang-canh")
        assert not co_mau("giong-doc") and not co_mau("clip")

    def test_mau_kich_ban_noi_ro_cach_dung(self, tmp_path):
        duong = str(tmp_path / "m.txt")
        viet_mau("kich-ban", duong)
        chu = open(duong, encoding="utf-8").read()
        assert "Nạp file có sẵn" in chu, "phải chỉ đúng nút cần bấm"
        assert "UTF-8" in chu, "lưu sai mã là chữ có dấu thành ô vuông"

    def test_khau_khong_co_mau_thi_bao_ro(self, tmp_path):
        with pytest.raises(LoiNapSan):
            viet_mau("clip", str(tmp_path / "x"))


class TestKieuFile:
    def test_khau_thu_muc_khong_khai_duoi(self):
        assert kieu_file_cua_khau("anh")[1] == ()

    def test_khau_tep_khai_dung_duoi(self):
        assert ".xlsx" in kieu_file_cua_khau("bang-canh")[1]
        assert ".mp3" in kieu_file_cua_khau("giong-doc")[1]


def test_khong_thay_file_thi_bao_ro(luot):
    with pytest.raises(LoiNapSan, match="Không thấy"):
        kiem_file("kich-ban", os.path.join(luot, "khong-ton-tai.txt"))
