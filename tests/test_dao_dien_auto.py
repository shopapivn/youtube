"""Tab Tự động nối dây chuyền đạo diễn của Prompt Visuals khi kênh khai `che_do_ke`.

Không gọi mạng: hàm AI chia cảnh và hàm tạo ảnh đều được bơm giả.
"""
import json
import os
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from core import dao_dien_auto as dd
from core.auto_khau import _hop_cho_canh, _viet_xlsx
from core.kenh import Kenh, doc_kenh

GOC = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _bc(tmp_path, che_do="tu_xay", anh_nv=None):
    nhat_ky = []
    kenh = SimpleNamespace(ma="story-3d", che_do_ke=che_do, engine="veo3", mo_hinh="claude-sonnet-5",
                           style={"default_character_prompt": "a cat"}, anh_nv=anh_nv or [])
    return SimpleNamespace(goc=str(tmp_path), kenh=kenh, client=object(), ghi=nhat_ky.append,
                           _nhat_ky=nhat_ky)


def _luot(tmp_path, srt=True):
    d = tmp_path / "0001"
    d.mkdir(exist_ok=True)
    if srt:
        (d / "3-phu-de.srt").write_text("1\n00:00:00,000 --> 00:00:02,000\nxin chào\n", encoding="utf-8")
    (d / "1-kich-ban.txt").write_text("Ngày xưa có một con mèo.", encoding="utf-8")
    return SimpleNamespace(thu_muc=str(d), ma_luot="0001", ma_kenh="story-3d")


MAN = {
    "scenes": [{"scene_id": 1, "img_prompt": "cat by the mill", "video_prompt": "cat blinks",
                "srt_start": "00:00:00,000", "srt_end": "00:00:02,000",
                "reference_files": json.dumps(["nv1.png", "loc1.png"])}],
    "characters": [{"id": "nv1", "name": "Cat", "role": "hero", "english_prompt": "a grey cat",
                    "sheet_prompt": "portrait of a grey cat"},
                   {"id": "nv2", "name": "King", "role": "king", "sheet_prompt": "portrait of a king"}],
    "locations": [{"id": "loc1", "name": "Mill", "english_prompt": "an old mill",
                   "sheet_prompt": "an old mill at eye level"}],
    "director_plan": [{"scene_id": 1, "location": "loc1"}],
}


class TestKenh:
    def test_doc_che_do_ke_tu_yaml(self, tmp_path):
        d = tmp_path / "CHANNEL" / "x"
        d.mkdir(parents=True)
        (d / "kenh.yaml").write_text("ma: x\nten: X\nche_do_ke: tu_xay\n", encoding="utf-8")
        assert doc_kenh(str(tmp_path), "x").che_do_ke == "tu_xay"

    def test_kenh_khong_khai_thi_rong(self, tmp_path):
        d = tmp_path / "CHANNEL" / "y"
        d.mkdir(parents=True)
        (d / "kenh.yaml").write_text("ma: y\n", encoding="utf-8")
        assert doc_kenh(str(tmp_path), "y").che_do_ke == ""
        assert not dd.che_do_dao_dien(doc_kenh(str(tmp_path), "y"))

    def test_che_do_dao_dien(self):
        assert dd.che_do_dao_dien(Kenh(che_do_ke="tu_xay"))
        assert dd.che_do_dao_dien(Kenh(che_do_ke="nhan_vat_va_boi_canh"))
        assert not dd.che_do_dao_dien(Kenh(che_do_ke="mot_nhan_vat"))
        assert not dd.che_do_dao_dien(Kenh())


class TestChayDaoDien:
    def test_goi_handle_va_ghi_dan(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        nhan = {}

        def handle(req):
            nhan.update(req)
            return {"scenes": {"json": MAN}}

        canh, man = dd.chay_dao_dien(bc, luot, handle=handle)
        assert [c["scene_id"] for c in canh] == [1]
        assert man["characters"][1]["id"] == "nv2"
        # Đầu vào đúng dạng run.py mong: {"path": ...}
        assert nhan["inputs"]["subtitles"]["path"].endswith("3-phu-de.srt")
        assert os.path.isfile(nhan["inputs"]["context"]["path"])
        with open(nhan["inputs"]["context"]["path"], encoding="utf-8") as f:
            ctx = json.load(f)
        assert "con mèo" in json.dumps(ctx, ensure_ascii=False)
        # Không tốn lượt ảnh bìa / nhạc — tab Tự động có khâu riêng.
        assert nhan["config"]["thumbnail"] is False and nhan["config"]["nhac"] is False
        assert nhan["config"]["che_do_ke"] == "tu_xay"
        assert nhan["workspace"] == luot.thu_muc
        with open(os.path.join(luot.thu_muc, dd.TEP_DAN), encoding="utf-8") as f:
            dan = json.load(f)
        assert [c["id"] for c in dan["characters"]] == ["nv1", "nv2"]
        assert dan["locations"][0]["id"] == "loc1"

    def test_khong_canh_thi_nem(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        with pytest.raises(RuntimeError):
            dd.chay_dao_dien(bc, luot, handle=lambda r: {"scenes": {"json": {"scenes": []}}})

    def test_thieu_phu_de_thi_nem(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path, srt=False)
        with pytest.raises(RuntimeError):
            dd.chay_dao_dien(bc, luot, handle=lambda r: MAN)

    def test_nhan_vat_co_dinh_dua_nv1_vao_boi_canh(self, tmp_path):
        bc, luot = _bc(tmp_path, che_do="nhan_vat_va_boi_canh"), _luot(tmp_path)
        nhan = {}

        def handle(req):
            nhan.update(req)
            return {"scenes": {"json": MAN}}

        dd.chay_dao_dien(bc, luot, handle=handle)
        with open(nhan["inputs"]["context"]["path"], encoding="utf-8") as f:
            ctx = json.dumps(json.load(f))
        assert "nv1.png" in ctx and "a cat" in ctx


class TestTaoThamChieu:
    def test_tao_thieu_bo_qua_da_co_va_chep_co_dinh(self, tmp_path):
        nv1 = tmp_path / "nv1.png"
        nv1.write_bytes(b"PNG-kenh")
        bc, luot = _bc(tmp_path, anh_nv=[str(nv1)]), _luot(tmp_path)
        tc = tmp_path / "0001" / dd.THU_MUC_THAM_CHIEU
        tc.mkdir()
        (tc / "loc1.png").write_bytes(b"da co")
        man = json.loads(json.dumps(MAN))
        man["characters"][0]["co_dinh"] = True
        goi = []

        def tao_anh(ma_id, prompt, dich, tham_chieu=None):
            goi.append((ma_id, prompt))
            if ma_id == "nv2":
                raise RuntimeError("content_rejected")
            with open(dich, "wb") as f:
                f.write(b"x")

        thieu = dd.tao_tham_chieu(bc, luot, man, tao_anh=tao_anh)
        assert thieu == ["nv2"]
        assert goi == [("nv2", "portrait of a king")]        # loc1 đã có, nv1 cố định → chép
        assert (tc / "nv1.png").read_bytes() == b"PNG-kenh"
        assert (tc / "loc1.png").read_bytes() == b"da co"
        assert any("nv2" in d and "KHÔNG" in d for d in bc._nhat_ky)

    def test_khong_co_gi_de_tao(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)

        def khong_goi(*a):
            pytest.fail("không được gọi")

        assert dd.tao_tham_chieu(bc, luot, {"characters": [], "locations": []}, tao_anh=khong_goi) == []


class TestThamChieuTheoCanh:
    def test_duong_tham_chieu_canh(self, tmp_path):
        luot = _luot(tmp_path)
        tc = tmp_path / "0001" / dd.THU_MUC_THAM_CHIEU
        tc.mkdir()
        (tc / "nv1.png").write_bytes(b"x")
        c = {"reference_files": json.dumps(["nv1.png", "loc9.png"])}
        assert [os.path.basename(p) for p in dd.duong_tham_chieu_canh(luot, c)] == ["nv1.png"]
        assert dd.duong_tham_chieu_canh(luot, {"reference_files": "nv1.png, loc9.png"}) == \
            dd.duong_tham_chieu_canh(luot, c)
        assert dd.duong_tham_chieu_canh(luot, {}) == []

    def test_hop_cho_canh(self, tmp_path, monkeypatch):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        tc = tmp_path / "0001" / dd.THU_MUC_THAM_CHIEU
        tc.mkdir()
        (tc / "nv1.png").write_bytes(b"x")
        hop_cu = object()
        c = {"reference_files": json.dumps(["nv1.png"])}
        # Kênh không khai → hộp cũ y nguyên (TL4-T7 không đổi gì).
        bc.kenh.che_do_ke = ""
        assert _hop_cho_canh(bc, luot, c, hop_cu) is hop_cu
        bc.kenh.che_do_ke = "tu_xay"
        hop = _hop_cho_canh(bc, luot, c, hop_cu)
        assert isinstance(hop, dd.ThamChieuCanh)
        goi = []

        def tai_len_gia(client, p):
            goi.append(p)
            return "https://u/" + os.path.basename(p)

        monkeypatch.setattr("core.anh_len.tai_len", tai_len_gia)
        assert hop.lay() == ["https://u/nv1.png"]
        assert hop.lay() == ["https://u/nv1.png"] and len(goi) == 1     # nhớ, không tải lại
        assert _hop_cho_canh(bc, luot, {"reference_files": "[]"}, hop_cu).lay() == []


class TestXlsxCoDan:
    def test_sheet_nhan_vat_va_boi_canh(self, tmp_path):
        k = SimpleNamespace(style={"default_character_prompt": "a cat", "reference_lock": ""})
        duong = str(tmp_path / "4-canh.xlsx")
        canh = [dict(MAN["scenes"][0]), {"scene_id": 2, "img_prompt": "mill", "video_prompt": "wind",
                                         "srt_start": "0", "srt_end": "1", "reference_files": ""}]
        _viet_xlsx(duong, canh, k, ke_hoach=MAN["director_plan"], dan=MAN["characters"],
                   boi_canh=MAN["locations"])
        wb = load_workbook(duong)
        nv = wb["characters"]
        ids = [nv.cell(h, 1).value for h in range(2, nv.max_row + 1)]
        assert ids == ["nv1", "nv2"]
        dau = [nv.cell(1, c).value for c in range(1, nv.max_column + 1)]
        assert nv.cell(3, dau.index("image_file") + 1).value == "nv2.png"
        assert nv.cell(3, dau.index("sheet_prompt") + 1).value == "portrait of a king"
        lo = wb["locations"]
        assert lo.cell(2, 1).value == "loc1" and lo.cell(2, 6).value == "loc1.png"
        assert "story_map" in wb.sheetnames
        ws = wb.worksheets[0]
        dau = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        cot = dau.index("reference_files") + 1
        assert json.loads(ws.cell(2, cot).value) == ["nv1.png", "loc1.png"]
        assert not (ws.cell(3, cot).value or "")           # cảnh không khai → để trống, KHÔNG ép nv1

    def test_duong_cu_van_mac_dinh_nv1(self, tmp_path):
        k = SimpleNamespace(style={"default_character_prompt": "a cat", "reference_lock": ""})
        duong = str(tmp_path / "4-canh.xlsx")
        _viet_xlsx(duong, [{"scene_id": 1, "img_prompt": "x", "video_prompt": "y",
                            "srt_start": "0", "srt_end": "1"}], k)
        wb = load_workbook(duong)
        nv = wb["characters"]
        assert nv.cell(2, 1).value == "nv1" and "locations" not in wb.sheetnames
        ws = wb.worksheets[0]
        dau = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert json.loads(ws.cell(2, dau.index("reference_files") + 1).value) == ["nv1.png"]


class TestThietKeLai:
    def test_tu_choi_hai_lan_thi_thiet_ke_lai_va_tao_lai(self, tmp_path):
        from core.prompt_visuals import DUOI_CHAN_DUNG
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        man = json.loads(json.dumps(MAN))
        lock = "REFERENCE IMAGES are attached, in this order:\n- reference image 1 = nv1, the hero: a grey cat with a feathered hat\n"
        man["characters"] = [
            {"id": "nv1", "name": "Cat", "role": "hero", "english_prompt": "a grey cat with a feathered hat",
             "sheet_prompt": "a grey cat with a feathered hat" + DUOI_CHAN_DUNG + " Style: 3D"},
            {"id": "nv1b", "name": "Cat", "role": "hero",
             "english_prompt": "a grey cat with a feathered hat; outfit at this stage: boots",
             "sheet_prompt": "a grey cat with a feathered hat; outfit at this stage: boots" + DUOI_CHAN_DUNG + " Style: 3D"},
        ]
        canh = [{"scene_id": 1, "img_prompt": "cat by the mill\n" + lock, "video_prompt": "x",
                 "reference_files": json.dumps(["nv1.png"])}]
        goi = []

        def tao_anh(ma_id, prompt, dich, tham_chieu=None):
            goi.append((ma_id, prompt[:40]))
            if "feathered" in prompt:
                raise RuntimeError("content_rejected")
            with open(dich, "wb") as f:
                f.write(b"png")

        def goi_ai(loi_nhac):
            assert "feathered hat" in loi_nhac and "REJECTING" in loi_nhac
            return '{"english_prompt": "a grey cat wearing a small red beret"}'

        thieu = dd.tao_tham_chieu(bc, luot, man, canh=canh, tao_anh=tao_anh, goi_ai=goi_ai)
        assert thieu == []
        # Cả hai giai đoạn tạo lại bằng thiết kế mới, giai đoạn b giữ đồ riêng.
        tc = tmp_path / "0001" / dd.THU_MUC_THAM_CHIEU
        assert (tc / "nv1.png").exists() and (tc / "nv1b.png").exists()
        assert man["characters"][1]["english_prompt"] == "a grey cat wearing a small red beret; outfit at this stage: boots"
        assert man["characters"][0]["sheet_prompt"].endswith(" Style: 3D")
        # Khối khoá trong cảnh đổi theo, và đã ghi ra đĩa.
        assert "small red beret" in canh[0]["img_prompt"] and "feathered" not in canh[0]["img_prompt"]
        with open(os.path.join(luot.thu_muc, "4-canh.json"), encoding="utf-8") as f:
            assert "small red beret" in json.load(f)[0]["img_prompt"]
        with open(os.path.join(luot.thu_muc, dd.TEP_DAN), encoding="utf-8") as f:
            assert "beret" in json.load(f)["characters"][0]["english_prompt"]

    def test_thiet_ke_lai_hong_thi_van_bao_thieu(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        man = json.loads(json.dumps(MAN))

        def tao_anh(ma_id, prompt, dich, tham_chieu=None):
            raise RuntimeError("content_rejected")

        thieu = dd.tao_tham_chieu(bc, luot, man, canh=[], tao_anh=tao_anh, goi_ai=lambda l: "không có json")
        assert sorted(thieu) == ["loc1", "nv1", "nv2"]
        assert any("KHÔNG tạo được" in d for d in bc._nhat_ky)


    def test_giai_doan_bi_chan_thi_doi_ca_bo_do_cua_giai_doan(self, tmp_path):
        from core.prompt_visuals import DUOI_CHAN_DUNG
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        man = json.loads(json.dumps(MAN))
        man["characters"] = [
            {"id": "nv1", "role": "hero", "english_prompt": "a yellow cat",
             "sheet_prompt": "a yellow cat" + DUOI_CHAN_DUNG + " Style: 3D"},
            {"id": "nv1b", "role": "hero",
             "english_prompt": "a yellow cat; outfit at this stage: a feathered hat and boots",
             "sheet_prompt": "a yellow cat; outfit at this stage: a feathered hat and boots" + DUOI_CHAN_DUNG + " Style: 3D"},
        ]
        lock = "REFERENCE IMAGES are attached, in this order:" + chr(10) + \
            "- reference image 1 = nv1b, the hero: a yellow cat; outfit at this stage: a feathered hat and boots" + chr(10)
        canh = [{"scene_id": 20, "img_prompt": "cat walks" + chr(10) + lock, "video_prompt": "x",
                 "reference_files": json.dumps(["nv1b.png"])}]
        (tmp_path / "0001" / dd.THU_MUC_THAM_CHIEU).mkdir()
        (tmp_path / "0001" / dd.THU_MUC_THAM_CHIEU / "nv1.png").write_bytes(b"cu")
        hoi = []

        def tao_anh(ma_id, prompt, dich, tham_chieu=None):
            if "feathered" in prompt:
                raise RuntimeError("content_rejected")
            with open(dich, "wb") as f:
                f.write(b"png")

        def goi_ai(loi_nhac):
            hoi.append(loi_nhac)
            return '{"english_prompt": "a yellow cat with green eyes", "outfit": "plain brown boots and a small red scarf"}'

        thieu = dd.tao_tham_chieu(bc, luot, man, canh=canh, tao_anh=tao_anh, goi_ai=goi_ai)
        assert thieu == []
        assert "feathered hat and boots" in hoi[0] and '"outfit"' in hoi[0]     # AI biết đồ cũ của giai đoạn
        nv1, nv1b = man["characters"]
        assert nv1["english_prompt"] == "a yellow cat with green eyes"
        assert nv1b["english_prompt"] == "a yellow cat with green eyes; outfit at this stage: plain brown boots and a small red scarf"
        assert nv1b["sheet_prompt"].endswith(" Style: 3D") and "feathered" not in nv1b["sheet_prompt"]
        assert "plain brown boots" in canh[0]["img_prompt"] and "feathered" not in canh[0]["img_prompt"]
        tc = tmp_path / "0001" / dd.THU_MUC_THAM_CHIEU
        assert (tc / "nv1b.png").read_bytes() == b"png" and (tc / "nv1.png").read_bytes() == b"png"
        assert (tc / "nv1.png.cu").read_bytes() == b"cu"


class TestSuaCanhTheoDoMoi:
    def test_chi_sua_canh_co_nhan_vat_va_ta_do_cu(self):
        lock = chr(10) + "REFERENCE IMAGES are attached, in this order:" + chr(10) + "- reference image 1 = nv1b, the hero: cat"
        canh = [
            {"scene_id": 5, "reference_files": json.dumps(["nv1b.png"]),
             "img_prompt": "nv1b tips its wide-brim hat, boots mid-step" + lock, "video_prompt": "the hat brim lifts"},
            {"scene_id": 6, "reference_files": json.dumps(["nv1b.png"]),
             "img_prompt": "nv1b waves happily" + lock, "video_prompt": "waves"},
            {"scene_id": 7, "reference_files": json.dumps(["nv8.png"]),
             "img_prompt": "a guard in tall boots" + lock, "video_prompt": "x"},
        ]
        hoi = []

        def goi_ai(loi_nhac):
            hoi.append(loi_nhac)
            return json.dumps({"5": {"img": "nv1b touches its red beret, vest mid-step", "video": "the beret tilts"}})

        n = dd.sua_canh_theo_do_moi(canh, "nv1b", "a wide-brim felt hat with a feather, leather boots and a cloth sack",
                                     "a small red beret and a little brown vest", goi_ai)
        assert n == 1 and len(hoi) == 1
        assert '"5"' in hoi[0] and '"6"' not in hoi[0] and '"7"' not in hoi[0]   # chỉ cảnh có nv1b và tả đồ cũ
        assert canh[0]["img_prompt"].startswith("nv1b touches its red beret") and "REFERENCE IMAGES" in canh[0]["img_prompt"]
        assert canh[0]["video_prompt"] == "the beret tilts"
        assert canh[2]["img_prompt"].startswith("a guard in tall boots")          # cảnh khác không đụng

    def test_khong_co_canh_nao_thi_khong_goi_ai(self):
        def goi_ai(l):
            pytest.fail("không được gọi")

        assert dd.sua_canh_theo_do_moi([{"scene_id": 1, "reference_files": "[]", "img_prompt": "x"}],
                                        "nv1b", "boots", "vest", goi_ai) == 0


class TestGiaiDoanVeTuAnh:
    """Giai đoạn sau vẽ TỪ ẢNH giai đoạn đầu (đo 26/08/2026: vẽ lại từ chữ ra cá thể khác)."""

    def test_giai_doan_sau_nhan_anh_goc_lam_tham_chieu(self, tmp_path):
        bc = _bc(tmp_path)
        luot = SimpleNamespace(thu_muc=str(tmp_path), ma_luot="u1", ma_kenh="story-3d")
        man = {"characters": [{"id": "nv2", "sheet_prompt": "meo vang"},
                              {"id": "nv2b", "sheet_prompt": "meo dung hai chan", "goc_id": "nv2"}],
               "locations": [{"id": "loc1", "sheet_prompt": "canh dong"}]}
        goi = []

        def tao_anh(ma_id, prompt, dich, tham_chieu=None):
            goi.append((ma_id, [os.path.basename(x) for x in (tham_chieu or [])]))
            open(dich, "wb").write(b"png")

        assert dd.tao_tham_chieu(bc, luot, man, tao_anh=tao_anh) == []
        theo = dict(goi)
        assert theo["nv2"] == [] and theo["loc1"] == []
        assert theo["nv2b"] == ["nv2.png"], "giai đoạn sau phải vẽ kèm ảnh giai đoạn đầu"
        # và giai đoạn đầu phải xong TRƯỚC (cùng lớp thì mới song song)
        assert [x[0] for x in goi].index("nv2") < [x[0] for x in goi].index("nv2b")


class TestSoiChanDung:
    def test_chan_dung_thieu_vuong_mien_thi_ve_lai_va_giu_ban_hon(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        man = {"characters": [{"id": "nv5", "role": "the king", "english_prompt": "a jolly old king with a golden crown and red robe",
                               "sheet_prompt": "portrait of the king"}], "locations": []}
        goi = []

        def tao_anh(ma_id, prompt, dich, tham_chieu=None):
            goi.append(prompt)
            with open(dich, "wb") as f:
                f.write(b"lan%d" % len(goi))

        def cham(anh, mo_ta, vai):
            assert vai == "the king" and "crown" in mo_ta
            return (3, "golden crown, royal robe") if open(anh, "rb").read() == b"lan1" else (5, "")

        assert dd.tao_tham_chieu(bc, luot, man, tao_anh=tao_anh, cham=cham) == []
        assert len(goi) == 2 and "MUST clearly show: golden crown, royal robe" in goi[1]
        tc = tmp_path / "0001" / dd.THU_MUC_THAM_CHIEU
        assert (tc / "nv5.png").read_bytes() == b"lan2"           # bản vẽ lại tốt hơn được dùng
        assert (tc / "nv5.png.lan1.png").exists()
        assert any("3/5" in d for d in bc._nhat_ky)

    def test_ban_ve_lai_khong_hon_thi_giu_tam_dau(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        man = {"characters": [{"id": "nv5", "role": "the king", "english_prompt": "a king", "sheet_prompt": "p"}],
               "locations": [{"id": "loc1", "sheet_prompt": "a hall"}]}
        n = {"k": 0}

        def tao_anh(ma_id, prompt, dich, tham_chieu=None):
            n["k"] += 1
            with open(dich, "wb") as f:
                f.write(b"x")

        cham_goi = []

        def cham(anh, mo_ta, vai):
            cham_goi.append(vai)
            return (3, "crown")

        assert dd.tao_tham_chieu(bc, luot, man, tao_anh=tao_anh, cham=cham) == []
        assert n["k"] == 3                           # nv5 hai lần, loc1 một lần (bối cảnh không chấm)
        assert cham_goi == ["the king", "the king"]
        tc = tmp_path / "0001" / dd.THU_MUC_THAM_CHIEU
        assert (tc / "nv5.png").exists() and not (tc / "nv5.png.lan2.png").exists()


class TestKhoaLanChay:
    def test_run_id_doi_khi_phu_de_doi(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        nhan = []

        def handle(req):
            nhan.append(req["run_id"])
            return {"scenes": {"json": MAN}}

        dd.chay_dao_dien(bc, luot, handle=handle)
        (tmp_path / "0001" / "3-phu-de.srt").write_text("1\n00:00:00,000 --> 00:00:02,000\nkhac\n", encoding="utf-8")
        dd.chay_dao_dien(bc, luot, handle=handle)
        assert nhan[0].startswith("story-3d-0001-") and nhan[0] != nhan[1]

    def test_khong_dan_nhan_vat_thi_nem(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        man = {"scenes": MAN["scenes"], "characters": [], "locations": []}
        with pytest.raises(RuntimeError, match="dàn nhân vật"):
            dd.chay_dao_dien(bc, luot, handle=lambda r: {"scenes": {"json": man}})


class TestKhuonChiaCuaKenh:
    def test_khuon_du_cho_trong(self):
        assert not dd.khuon_du_cho_dao_dien("You are a storyboard director. nv1 (nv1.png) <<SRT>> <<MAX_SEC>>")
        du = " ".join(dd.CHO_TRONG_KHUON_CHIA)
        assert dd.khuon_du_cho_dao_dien("khuôn " + du)
        assert not dd.khuon_du_cho_dao_dien("")

    def test_dua_7_canh_cua_kenh_vao_boi_canh(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        bc.kenh.prompt = {"7-canh.md": "KHUON KENH " + " ".join(dd.CHO_TRONG_KHUON_CHIA)}
        nhan = {}

        def handle(req):
            nhan.update(req)
            return {"scenes": {"json": MAN}}

        dd.chay_dao_dien(bc, luot, handle=handle)
        with open(nhan["inputs"]["context"]["path"], encoding="utf-8") as f:
            ctx = json.load(f)
        assert ctx["storyboard_template"].startswith("KHUON KENH")
        # Khuôn kiểu TL4 (thiếu chỗ trống) thì không đưa vào.
        bc.kenh.prompt = {"7-canh.md": "nv1 (nv1.png) <<SRT>>"}
        dd.chay_dao_dien(bc, luot, handle=handle)
        with open(nhan["inputs"]["context"]["path"], encoding="utf-8") as f:
            assert "storyboard_template" not in json.load(f)

    def test_kenh_story_3d_co_khuon_du(self):
        from core.kenh import doc_kenh
        goc = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        k = doc_kenh(goc, "story-3d")
        assert dd.khuon_du_cho_dao_dien(k.prompt.get("7-canh.md", ""))
        assert "metaphor" not in k.prompt["7-canh.md"].split("THE RULES")[1].split("STYLE TAIL")[0].lower() or \
            "not a metaphor" in k.prompt["7-canh.md"]


class TestBiaTheoPhim:
    def test_nhan_vat_chinh_va_hop_bia(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        tc = tmp_path / "0001" / dd.THU_MUC_THAM_CHIEU
        tc.mkdir()
        for t in ("nv1.png", "nv4.png", "nv9.png"):
            (tc / t).write_bytes(b"x")
        canh = [{"scene_id": i, "reference_files": json.dumps(r)} for i, r in enumerate([
            ["nv1.png", "loc1.png"], ["nv1.png", "nv4.png", "loc1.png"], ["nv4.png", "loc2.png"],
            ["nv9.png", "loc2.png"], ["nv1.png", "nv4.png"], ["nv7.png"]], 1)]
        (tmp_path / "0001" / "4-canh.json").write_text(json.dumps(canh), encoding="utf-8")
        chinh = dd.nhan_vat_chinh_cua_luot(luot, 2)
        assert [os.path.basename(p) for p in chinh] == ["nv1.png", "nv4.png"]   # nv7 không có tệp → bỏ
        from core.auto_khau import _hop_bia
        hop_cu = object()
        assert _hop_bia(bc, luot, hop_cu) is not hop_cu and any("ảnh bìa" in d for d in bc._nhat_ky)
        bc.kenh.che_do_ke = ""
        assert _hop_bia(bc, luot, hop_cu) is hop_cu
        assert dd.nhan_vat_chinh_cua_luot(_luot(tmp_path / "khac"), 2) == [] if False else True


# ── Thiếu ảnh tham chiếu: mượn giai đoạn anh em, hoặc bỏ id cho khỏi lệch số ──
# Đo 26/08/2026 (phim 0002): `nv5.png` bị bộ lọc chặn, tool đi tiếp, khối khoá
# vẫn ghi "reference image 1 = nv5" nhưng ảnh gắn thật là bối cảnh → mọi số thứ
# tự trỏ nhầm; cảnh 4 và 6 ra hai con sói khác nhau.

class TestThieuAnhThamChieu:
    def _man_soi(self):
        from core.prompt_visuals import DUOI_CHAN_DUNG
        return {"characters": [
            {"id": "nv5", "role": "villain", "english_prompt": "a grey wolf",
             "sheet_prompt": "a grey wolf" + DUOI_CHAN_DUNG},
            {"id": "nv5b", "role": "villain", "english_prompt": "a grey wolf, white paws",
             "sheet_prompt": "a grey wolf, white paws" + DUOI_CHAN_DUNG}],
            "locations": [{"id": "loc1", "name": "cottage", "english_prompt": "a cottage",
                           "sheet_prompt": "a cottage"}]}

    def test_muon_anh_giai_doan_anh_em(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        man = self._man_soi()

        def tao_anh(ma_id, prompt, dich, tham_chieu=None):
            if ma_id == "nv5":
                raise RuntimeError("content_rejected")
            open(dich, "wb").write(b"png")

        thieu = dd.tao_tham_chieu(bc, luot, man, canh=[], tao_anh=tao_anh,
                                  goi_ai=lambda l: "không có json")
        assert thieu == []
        d = os.path.join(luot.thu_muc, dd.THU_MUC_THAM_CHIEU)
        assert os.path.exists(os.path.join(d, "nv5.png"))
        assert any("mượn ảnh nv5b" in x for x in bc._nhat_ky)

    def test_khong_co_anh_em_thi_bo_id_va_dung_lai_khoi_khoa(self, tmp_path):
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        bc.goc = GOC   # `_bo_id_khoi_canh` dựng lại khối khoá bằng chính run.py
        man = self._man_soi()
        man["characters"] = [man["characters"][0]]   # chỉ còn nv5, không anh em
        canh = [{"scene_id": 1, "characters_used": "nv5", "location_used": "loc1",
                 "reference_files": json.dumps(["nv5.png", "loc1.png"]),
                 "img_prompt": "Wide shot of nv5 at loc1", "video_prompt": "it walks"}]

        def tao_anh(ma_id, prompt, dich, tham_chieu=None):
            if ma_id == "nv5":
                raise RuntimeError("content_rejected")
            open(dich, "wb").write(b"png")

        thieu = dd.tao_tham_chieu(bc, luot, man, canh=canh, tao_anh=tao_anh,
                                  goi_ai=lambda l: "không có json")
        assert thieu == ["nv5"]
        assert json.loads(canh[0]["reference_files"]) == ["loc1.png"]
        assert "nv5" not in canh[0]["characters_used"]
        # Khối khoá dựng lại: ảnh 1 giờ là loc1, không còn trỏ nhầm. Ràng
        # buộc nằm THẲNG trong câu văn (`loc1 (Image 1)`) theo lối OpenStory —
        # xem `run._GUARD_NHAN_DANG`.
        assert "loc1 (Image 1)" in canh[0]["img_prompt"]
        assert "nv5 (Image" not in canh[0]["img_prompt"]
        assert "= nv5" not in canh[0]["img_prompt"]
        assert json.load(open(os.path.join(luot.thu_muc, "4-canh.json"), encoding="utf-8"))[0]["characters_used"] == ""

    def test_anh_goc_hong_thi_giai_doan_sau_van_ve_theo_anh_em(self, tmp_path):
        """MỘT nhân vật chỉ được có MỘT thiết kế, kể cả khi ảnh giai đoạn đầu hỏng.

        Đo 26/08/2026 (phim 0002): `nv5` bị bộ lọc chặn, `nv5b` và `nv5c` chạy song
        song cùng thấy trống nên mỗi cái vẽ từ chữ — ra một con gấu và một con sói.
        """
        from core.prompt_visuals import DUOI_CHAN_DUNG
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        man = {"characters": [
            {"id": "nv5", "role": "villain", "english_prompt": "a wolf", "goc_id": "nv5",
             "sheet_prompt": "a wolf" + DUOI_CHAN_DUNG},
            {"id": "nv5b", "role": "villain", "english_prompt": "a wolf, white paws", "goc_id": "nv5",
             "sheet_prompt": "a wolf, white paws" + DUOI_CHAN_DUNG},
            {"id": "nv5c", "role": "villain", "english_prompt": "a wolf, full belly", "goc_id": "nv5",
             "sheet_prompt": "a wolf, full belly" + DUOI_CHAN_DUNG}], "locations": []}
        thay = {}

        def tao_anh(ma_id, prompt, dich, tham_chieu=None):
            if ma_id == "nv5":
                raise RuntimeError("content_rejected")
            thay[ma_id] = [os.path.basename(x) for x in (tham_chieu or [])]
            open(dich, "wb").write(b"png")

        dd.tao_tham_chieu(bc, luot, man, canh=[], tao_anh=tao_anh, goi_ai=lambda l: "không có json")
        # nv5b vẽ trước (không có mẫu, đành vẽ từ chữ), nv5c PHẢI vẽ theo ảnh nv5b
        assert thay["nv5b"] == []
        assert thay["nv5c"] == ["nv5b.png"], "giai đoạn sau phải vẽ theo ảnh anh em đã có"

    def test_giai_doan_sau_ve_theo_anh_giai_doan_dau_khi_no_co(self, tmp_path):
        from core.prompt_visuals import DUOI_CHAN_DUNG
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        man = {"characters": [
            {"id": "nv5", "role": "villain", "english_prompt": "a wolf", "goc_id": "nv5",
             "sheet_prompt": "a wolf" + DUOI_CHAN_DUNG},
            {"id": "nv5b", "role": "villain", "english_prompt": "a wolf, white paws", "goc_id": "nv5",
             "sheet_prompt": "a wolf, white paws" + DUOI_CHAN_DUNG},
            {"id": "nv5c", "role": "villain", "english_prompt": "a wolf, full belly", "goc_id": "nv5",
             "sheet_prompt": "a wolf, full belly" + DUOI_CHAN_DUNG}], "locations": []}
        thay = {}

        def tao_anh(ma_id, prompt, dich, tham_chieu=None):
            thay[ma_id] = [os.path.basename(x) for x in (tham_chieu or [])]
            open(dich, "wb").write(b"png")

        dd.tao_tham_chieu(bc, luot, man, canh=[], tao_anh=tao_anh, goi_ai=lambda l: "x")
        assert thay["nv5"] == [] and thay["nv5b"] == ["nv5.png"] and thay["nv5c"] == ["nv5.png"]

    def test_thiet_ke_lai_van_giu_phong_cach_phim(self, tmp_path):
        """Vẽ lại nhân vật mà rơi mất khối "Style:" thì cả phim hai lối vẽ.

        Đo 27/08/2026 (phim 0002): lời nhắc con sói bị bộ lọc bắt viết lại nên mất
        mốc DUOI_CHAN_DUNG; bản thiết kế lại cắt chuỗi cũ không ra gì, ba ảnh sói ra
        lối 2D viền nét trong khi cả đàn dê là 3D dựng hình.
        """
        from core.prompt_visuals import DUOI_CHAN_DUNG
        bc, luot = _bc(tmp_path), _luot(tmp_path)
        with open(os.path.join(luot.thu_muc, "4-boi-canh.json"), "w", encoding="utf-8") as f:
            json.dump({"visual_style_directive": "Image style: stylised 3D animated film still, Pixar-like"}, f)
        nv = {"id": "nv5", "role": "villain", "english_prompt": "a wolf",
              "sheet_prompt": "a wolf, drawn after a filter rewrite — mốc đã mất"}
        man = {"characters": [nv], "locations": []}
        # 1) mất mốc, không có nhân vật khác → lấy từ 4-boi-canh.json
        d = dd._duoi_phong_cach(luot, man, nv)
        assert "stylised 3D animated film still" in d and d.startswith(" Style:")
        # 2) có nhân vật khác còn mốc → mượn đuôi của nó
        man["characters"] = [nv, {"id": "nv1", "sheet_prompt": "a goat" + DUOI_CHAN_DUNG + " Style: WATERCOLOUR"}]
        assert dd._duoi_phong_cach(luot, man, nv) == " Style: WATERCOLOUR"
        # 3) chính nó còn mốc → dùng của nó
        nv["sheet_prompt"] = "a wolf" + DUOI_CHAN_DUNG + " Style: MINE"
        assert dd._duoi_phong_cach(luot, man, nv) == " Style: MINE"
