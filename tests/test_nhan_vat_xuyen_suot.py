"""Prompt Visuals kế thừa "nhân vật xuyên suốt + phong cách" của tab Tự động.

Chủ dự án, 22/08/2026: *"ở tab visual khách đẩy voice và cần nhận lại file excel
có prompt, vậy các prompt trong excel cũng cần phù hợp theo dạng 1 nhân vật cố
định hay là xây sheet tham chiếu… rồi phong cách… tức là nó kế thừa ở auto."*

Tab Tự động khoá một `nv1` xuyên suốt vì có sẵn `nv1.png` của kênh. Prompt
Visuals không có kênh, nên nó **tự rút** dàn nhân vật từ chính lời đọc (một lượt
"casting"), rồi mọi cảnh dùng chung — sheet `characters` được điền, mỗi cảnh trỏ
`reference_files` tới `<id>.png`, và một phong cách giữ nguyên cả video.

Không bài nào gọi mạng: lượt casting và lượt chia cảnh đều truyền hàm giả
(`cast_fn`, `chia_fn`).
"""

from __future__ import annotations

import importlib.util
import json
import os
import sys

import pytest

GOC = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if GOC not in sys.path:
    sys.path.insert(0, GOC)


def cue(so: int, dai: float = 3.0):
    dau = (so - 1) * dai
    return {"index": so, "start": dau, "end": dau + dai,
            "text": "cau so {0}".format(so)}


def canh_ai(tu: int, den: int, **them):
    m = {"srt_from": tu, "srt_to": den, "img_prompt": "anh {0}".format(tu),
         "video_prompt": "clip {0}".format(tu)}
    m.update(them)
    return m


def _nap_run_py():
    """Nạp `tool-catalog/prompt.workbook/run.py` — tên thư mục có dấu chấm."""
    duong = os.path.join(GOC, "tool-catalog", "prompt.workbook", "run.py")
    spec = importlib.util.spec_from_file_location("prompt_workbook_run", duong)
    mo_dun = importlib.util.module_from_spec(spec)
    sys.modules["prompt_workbook_run"] = mo_dun
    spec.loader.exec_module(mo_dun)
    return mo_dun


@pytest.fixture(scope="module")
def wb():
    return _nap_run_py()


SRT = "\n\n".join(
    "{0}\n00:00:{1:02d},000 --> 00:00:{2:02d},000\nCau so {0}".format(
        i, (i - 1) * 3, i * 3)
    for i in range(1, 11))


@pytest.fixture
def yeu_cau(tmp_path):
    srt = tmp_path / "loi.srt"
    srt.write_text(SRT, encoding="utf-8")
    lam = tmp_path / "lam-viec"
    lam.mkdir()
    return {"inputs": {"subtitles": {"path": str(srt), "artifact_id": "a1"}},
            "config": {"engine": "veo3", "model": "claude-sonnet-5"},
            "workspace": str(lam), "workflow_id": "kiem-tra"}


def _dan_hai_nguoi(_cues, _ctx):
    return {"style": {"image_style": "cinematic", "palette": "warm",
                      "motion": "slow push-in"},
            "characters": [
                {"id": "nv1", "role": "protagonist", "name": "Anh",
                 "english_prompt": "a young fisherman, weathered face, blue coat"},
                {"id": "nv2", "role": "sister", "name": "Em",
                 "english_prompt": "a girl, long black hair, red scarf"}]}


def _chia_giu_khuc(khuc, _t, _n):
    return [canh_ai(khuc[0]["index"], khuc[-1]["index"])]


# ── 1. Có dàn → manifest + sheet characters + reference_files ────────────────

class TestCoNhanVat:
    def test_manifest_va_sheet_characters_duoc_dien(self, wb, yeu_cau):
        from openpyxl import load_workbook

        ra = wb.handle(yeu_cau, cast_fn=_dan_hai_nguoi, chia_fn=_chia_giu_khuc)
        nhan_vat = ra["scenes"]["json"]["characters"]
        assert [c["id"] for c in nhan_vat] == ["nv1", "nv2"]

        duong = os.path.join(yeu_cau["workspace"], "scene-prompts.xlsx")
        sach = load_workbook(duong)
        try:
            assert "characters" in sach.sheetnames
            sheet = sach["characters"]
            dau = [o.value for o in next(sheet.iter_rows(max_row=1))]
            assert dau == wb.CHARACTER_HEADERS
            dong2 = {dau[i]: o.value
                     for i, o in enumerate(next(sheet.iter_rows(min_row=2, max_row=2)))}
            assert dong2["id"] == "nv1"
            assert dong2["image_file"] == "nv1.png"
            assert dong2["status"] == "pending"
            assert "fisherman" in (dong2["english_prompt"] or "")
        finally:
            sach.close()

    def test_moi_canh_tro_reference_files_toi_anh_nhan_vat(self, wb, yeu_cau):
        ra = wb.handle(yeu_cau, cast_fn=_dan_hai_nguoi, chia_fn=_chia_giu_khuc)
        canh = ra["scenes"]["json"]["scenes"]
        assert canh, "phải có cảnh"
        for c in canh:
            assert c["characters_used"] == "nv1", \
                "có dàn thì cảnh nào chưa gán ai phải mặc định nhân vật chính"
            assert json.loads(c["reference_files"]) == ["nv1.png"], \
                "cảnh dùng nv1 phải trỏ reference_files tới nv1.png"

    def test_style_di_vao_settings(self, wb, yeu_cau):
        ra = wb.handle(yeu_cau, cast_fn=_dan_hai_nguoi, chia_fn=_chia_giu_khuc)
        style = ra["scenes"]["json"]["settings"]["style"]
        assert style["image_style"] == "cinematic"


# ── 2. Khối CAST_STYLE thật sự đi vào lời nhắc chia cảnh ─────────────────────

class TestCastStyleVaoLoiNhac:
    def test_khoi_cast_style_liet_ke_nhan_vat_va_style(self, wb):
        khoi = wb._khoi_cast_style(_dan_hai_nguoi(None, None))
        assert "nv1" in khoi and "nv2" in khoi
        assert "fisherman" in khoi
        assert "cinematic" in khoi and "warm" in khoi
        assert "RECURRING CHARACTERS" in khoi

    def test_dan_rong_thi_khoi_rong(self, wb):
        assert wb._khoi_cast_style({"style": {}, "characters": []}) == ""

    def test_loi_nhac_khuc_co_khoi_cast_khi_co_dan(self, wb):
        """`_bo_chia` phải nhét khối dàn nhân vật vào lời nhắc từng khúc."""
        khoi = wb._khoi_cast_style(_dan_hai_nguoi(None, None))
        chup = {}

        def goi_gia(loi_nhac, _khoa):
            chup["loi_nhac"] = loi_nhac
            return json.dumps({"scenes": [canh_ai(1, 2)]})

        chia = wb._bo_chia(goi_gia, {}, "veo3", khoi)
        chia([cue(1), cue(2)], 0, 1)
        assert "RECURRING CHARACTERS" in chup["loi_nhac"]
        assert "nv1" in chup["loi_nhac"]
        assert "<<" not in chup["loi_nhac"], "placeholder chưa điền phải được dọn"

    def test_khong_dan_thi_loi_nhac_khong_bia_nv1(self, wb):
        chup = {}

        def goi_gia(loi_nhac, _khoa):
            chup["loi_nhac"] = loi_nhac
            return json.dumps({"scenes": [canh_ai(1, 2)]})

        chia = wb._bo_chia(goi_gia, {}, "veo3", "")   # không dàn
        chia([cue(1), cue(2)], 0, 1)
        assert "nv1" not in chup["loi_nhac"]
        assert "<<" not in chup["loi_nhac"]


# ── 3. Dàn rỗng / cờ tắt → về hành vi cũ ────────────────────────────────────

class TestVeHanhViCu:
    def test_cast_rong_thi_characters_rong_van_ra_canh(self, wb, yeu_cau):
        ra = wb.handle(yeu_cau,
                       cast_fn=lambda *_a: {"style": {}, "characters": []},
                       chia_fn=_chia_giu_khuc)
        assert ra["scenes"]["json"]["characters"] == []
        canh = ra["scenes"]["json"]["scenes"]
        assert canh and all(c["characters_used"] == "" for c in canh)
        assert all(c["reference_files"] == "" for c in canh)

    def test_co_tat_thi_khong_casting_du_co_cast_fn(self, wb, yeu_cau):
        yeu_cau["config"]["nhat_quan_nhan_vat"] = False
        goi_cast = []

        def cast_fn(_c, _x):
            goi_cast.append(1)
            return _dan_hai_nguoi(_c, _x)

        ra = wb.handle(yeu_cau, cast_fn=cast_fn, chia_fn=_chia_giu_khuc)
        assert goi_cast == [], "tắt cờ thì không được chạy lượt casting"
        assert ra["scenes"]["json"]["characters"] == []


# ── 4. Casting hỏng không giết cả lượt ──────────────────────────────────────

def test_casting_hong_van_ra_file_va_noi_ra(wb, yeu_cau, capsys):
    def cast_no(_c, _x):
        raise RuntimeError("may chu casting tam gian doan")

    ra = wb.handle(yeu_cau, cast_fn=cast_no, chia_fn=_chia_giu_khuc)
    assert ra["scenes"]["json"]["characters"] == [], \
        "casting hỏng thì dàn rỗng, không giết cả lượt"
    assert ra["scenes"]["json"]["scenes"], "vẫn phải ra cảnh"
    loi_nhan = [json.loads(d).get("message", "")
                for d in capsys.readouterr().out.splitlines() if d.strip()]
    assert any("dan nhan vat" in c for c in loi_nhan), \
        "casting hỏng phải nói ra, không im lặng"


# ── 5. Cờ nhất quán luồn vào config node prompt ─────────────────────────────

class TestCoTrongConfig:
    def test_mac_dinh_bat(self):
        from core.prompt_visuals import dung_workflow

        wf = dung_workflow("art-1")
        node = [n for n in wf["nodes"] if n["id"] == "prompt"][0]
        assert node["config"]["nhat_quan_nhan_vat"] is True

    def test_tat_duoc(self):
        from core.prompt_visuals import dung_workflow

        wf = dung_workflow("art-1", nhat_quan=False)
        node = [n for n in wf["nodes"] if n["id"] == "prompt"][0]
        assert node["config"]["nhat_quan_nhan_vat"] is False
