"""Kịch bản + phong cách hình ảnh + sửa prompt tại chỗ cho tab Prompt Visuals.

Chủ dự án đòi ba thứ ở tab này: (1) ngoài phụ đề còn đẩy được **kịch bản** để
tăng chính xác, (2) chọn được **phong cách hình ảnh** như chọn template visual ở
tab Tự động, (3) prompt ra rồi phải **sửa được** ngay. Bài này khoá phần thuần
tuý của cả ba: gói `context`, luồn vào workflow, và ghi prompt đã sửa ngược lại
file Excel. Không mạng, không cần model.
"""

from __future__ import annotations

import importlib.util
import os

from openpyxl import load_workbook

from core.prompt_visuals import MAU_HINH, dung_boi_canh, dung_workflow


def _pw_run():
    """Nạp `tool-catalog/prompt.workbook/run.py` để mượn `render_workbook`."""
    goc = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    duong = os.path.join(goc, "tool-catalog", "prompt.workbook", "run.py")
    spec = importlib.util.spec_from_file_location("pw_run_ctx", duong)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ── 1. Gói context: kịch bản + phong cách ──────────────────────────────────

def test_boi_canh_rong_thi_khong_gui_gi():
    # Không kịch bản, phong cách "auto" → không có gì để gửi, workflow về dạng cũ.
    assert dung_boi_canh("", "auto") == {}
    assert dung_boi_canh("   ", "auto") == {}


def test_boi_canh_chi_kich_ban():
    ra = dung_boi_canh("Ông lão và con cá", "auto")
    assert ra == {"script": "Ông lão và con cá"}


def test_boi_canh_phong_cach_dien_anh():
    ra = dung_boi_canh("", "dien_anh")
    assert "script" not in ra
    assert "cinematic" in ra["visual_style_directive"].lower()


def test_boi_canh_ca_hai():
    ra = dung_boi_canh("kịch bản", "hoat_hinh_3d")
    assert ra["script"] == "kịch bản"
    assert "3d animation" in ra["visual_style_directive"].lower()


def test_moi_mau_hinh_deu_dung_duoc():
    # Mọi mã trong MAU_HINH phải cho ra một context hợp lệ (auto thì rỗng).
    for ma, _ten, chi_dan in MAU_HINH:
        ra = dung_boi_canh("", ma)
        if chi_dan:
            assert ra["visual_style_directive"].endswith(chi_dan)
        else:
            assert ra == {}


# ── 2. context luồn vào node prompt của workflow ────────────────────────────

def test_workflow_khong_context_thi_input_rong():
    wf = dung_workflow("art-1")
    prompt = [n for n in wf["nodes"] if n["id"] == "prompt"][0]
    assert prompt["inputs"] == {}


def test_workflow_co_context_thi_them_cong():
    wf = dung_workflow("art-1", ma_artifact_context="ctx-9")
    prompt = [n for n in wf["nodes"] if n["id"] == "prompt"][0]
    assert prompt["inputs"] == {"context": "ctx-9"}
    # Cạnh phụ đề vẫn còn nguyên — context là input rời, không phải edge.
    assert any(e["target_port"] == "subtitles" for e in wf["edges"])


# ── 3. Sửa prompt tại chỗ rồi ghi ngược lại Excel ───────────────────────────

def test_ghi_prompt_da_sua_vao_xlsx(tmp_path):
    """Ghi hợp đồng thật: xuất workbook, sửa prompt hai cảnh, đọc lại thấy đúng."""
    from ui_qt.trang_prompt_visuals import TrangPromptVisuals

    pw = _pw_run()
    manifest = {
        "scenes": [
            {"scene_id": 1, "img_prompt": "anh cu 1", "video_prompt": "clip cu 1"},
            {"scene_id": 2, "img_prompt": "anh cu 2", "video_prompt": "clip cu 2"},
        ],
        "characters": [],
    }
    out = tmp_path / "scene-prompts.xlsx"
    pw.render_workbook(out, manifest)

    sua = {1: ("ANH MOI 1", "CLIP MOI 1"), 2: ("ANH MOI 2", "CLIP MOI 2")}
    so = TrangPromptVisuals._ghi_prompt_vao_xlsx(str(out), sua)
    assert so == 2

    wb = load_workbook(str(out), read_only=True, data_only=True)
    try:
        hang = [list(r) for r in wb["scenes"].iter_rows(values_only=True)]
    finally:
        wb.close()
    tieu = list(hang[0])
    c_id, c_anh, c_vid = (tieu.index("scene_id"), tieu.index("img_prompt"),
                          tieu.index("video_prompt"))
    theo_so = {int(d[c_id]): (d[c_anh], d[c_vid]) for d in hang[1:] if d[c_id]}
    assert theo_so[1] == ("ANH MOI 1", "CLIP MOI 1")
    assert theo_so[2] == ("ANH MOI 2", "CLIP MOI 2")


def test_ghi_prompt_bo_qua_canh_khong_sua(tmp_path):
    """Chỉ sửa cảnh có trong `sua`; cảnh khác giữ nguyên."""
    from ui_qt.trang_prompt_visuals import TrangPromptVisuals

    pw = _pw_run()
    manifest = {"scenes": [
        {"scene_id": 1, "img_prompt": "giu nguyen", "video_prompt": "giu nguyen"},
        {"scene_id": 2, "img_prompt": "sua", "video_prompt": "sua"},
    ], "characters": []}
    out = tmp_path / "scene-prompts.xlsx"
    pw.render_workbook(out, manifest)

    so = TrangPromptVisuals._ghi_prompt_vao_xlsx(str(out), {2: ("X", "Y")})
    assert so == 1

    wb = load_workbook(str(out), read_only=True, data_only=True)
    try:
        hang = [list(r) for r in wb["scenes"].iter_rows(values_only=True)]
    finally:
        wb.close()
    tieu = list(hang[0])
    c_id, c_anh = tieu.index("scene_id"), tieu.index("img_prompt")
    theo_so = {int(d[c_id]): d[c_anh] for d in hang[1:] if d[c_id]}
    assert theo_so[1] == "giu nguyen"
    assert theo_so[2] == "X"


# ── 4. Ghi sửa ảnh bìa + nhạc theo khoá dòng (24/08/2026) ───────────────────

def test_ghi_cot_thumbnail_va_music(tmp_path):
    from ui_qt.trang_prompt_visuals import TrangPromptVisuals

    pw = _pw_run()
    manifest = {
        "scenes": [{"scene_id": 1, "img_prompt": "a", "video_prompt": "b"}],
        "characters": [], "title": "T", "thumb_text": "HOOK",
        "thumbnails": [{"thumb_id": 1, "version_desc": "portrait_main",
                        "img_prompt": "cu", "status_img": "pending"}],
        "music": [{"music_id": 1, "start_time": 0, "end_time": 30,
                   "suno_prompt": "cu", "mood": "calm", "status": "pending"}],
    }
    out = tmp_path / "scene-prompts.xlsx"
    pw.render_workbook(out, manifest)
    assert TrangPromptVisuals._ghi_cot_vao_xlsx(
        str(out), "thumbnail", "thumb_id", {"1": {"img_prompt": "MOI"}}) == 1
    assert TrangPromptVisuals._ghi_cot_vao_xlsx(
        str(out), "music", "music_id", {"1": {"suno_prompt": "NHAC MOI"}}) == 1

    from core.prompt_visuals import bia_de_xem, nhac_de_xem
    wb = load_workbook(str(out), read_only=True, data_only=True)
    try:
        bia = bia_de_xem([list(r) for r in wb["thumbnail"].iter_rows(values_only=True)])
        nhac = nhac_de_xem([list(r) for r in wb["music"].iter_rows(values_only=True)])
    finally:
        wb.close()
    assert bia[0]["img_prompt"] == "MOI" and bia[0]["thumb_text"] == "HOOK"
    assert nhac[0]["suno_prompt"] == "NHAC MOI"


# ── 5. Đổi tên tệp tham chiếu thành đường dẫn thật (ảnh tham chiếu tự tạo) ──

def test_ghi_duong_tham_chieu_thay_dung_ten_giu_ten_chua_co(tmp_path):
    from ui_qt.trang_prompt_visuals import TrangPromptVisuals

    pw = _pw_run()
    out = tmp_path / "scene-prompts.xlsx"
    pw.render_workbook(out, {
        "scenes": [{"scene_id": 1, "img_prompt": "a", "video_prompt": "b",
                    "reference_files": '["nv1.png", "loc1.png"]'},
                   {"scene_id": 2, "img_prompt": "c", "video_prompt": "d",
                    "reference_files": ""}],
        "characters": [], "thumbnails": [
            {"thumb_id": 1, "version_desc": "x", "img_prompt": "p",
             "reference_files": '["nv1.png"]'}]})
    anh = str(tmp_path / "loc1.png")
    assert TrangPromptVisuals._ghi_duong_tham_chieu(str(out), {"loc1.png": anh}) == ""
    wb = load_workbook(str(out), read_only=True, data_only=True)
    try:
        hang = [list(r) for r in wb["scenes"].iter_rows(values_only=True)]
        bia = [list(r) for r in wb["thumbnail"].iter_rows(values_only=True)]
    finally:
        wb.close()
    c = hang[0].index("reference_files")
    assert hang[1][c] == "nv1.png, " + anh      # loc1 thành đường thật, nv1 chờ
    assert not hang[2][c]                         # ô trống giữ trống
    # Lượt hai: nv1 về → thay nốt, đường cũ giữ nguyên.
    nv = str(tmp_path / "nv1.png")
    TrangPromptVisuals._ghi_duong_tham_chieu(str(out), {"nv1.png": nv})
    wb = load_workbook(str(out), read_only=True, data_only=True)
    try:
        hang = [list(r) for r in wb["scenes"].iter_rows(values_only=True)]
        bia = [list(r) for r in wb["thumbnail"].iter_rows(values_only=True)]
    finally:
        wb.close()
    assert hang[1][c] == nv + ", " + anh
    cb = bia[0].index("reference_files")
    assert bia[1][cb] == nv
