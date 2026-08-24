"""Cắt cảnh theo NGHĨA, không theo đồng hồ — và cùng một cách cắt cho hai tab.

Chủ dự án, 15/08/2026: *"ở tab prompt visuals thì như auto và tool gốc nó không
làm mặc định 8s mà nó theo nội dung srt"*.

Hai chỗ cần đúng cách cắt này: khâu bảng cảnh của tab Tự động và tool
`tool-catalog/prompt.workbook`. Bài kiểm dưới đây canh hai thứ:

1. **Cách cắt làm đúng việc của nó** — phủ hết phụ đề, không chồng lấn, không
   quá trần engine, và cảnh AI quên lời nhắc thì nhập vào cảnh trước chứ không
   giết cả lượt.
2. **Cả hai tab thật sự dùng chung một bản** — không phải hai bản chép tay.

Và một thứ nữa, đắt hơn cả hai: khi không gọi được AI thì Prompt Visuals vẫn
ra file, nhưng **phải nói thẳng** là cảnh đang bị cắt theo đồng hồ. Lùi im lặng
thì khách mở Excel ra thấy đủ 111 cảnh, không cách nào biết cảnh bị cắt giữa
câu.

Không bài nào gọi mạng: lời gọi AI được truyền vào bằng hàm giả.
"""

from __future__ import annotations

import importlib.util
import json
import os
import sys

import pytest

from core.chia_canh import (
    KHUON_MAC_DINH, bang_phu_de, canh_lai, chia_khuc, chia_theo_nghia,
    loi_nhac_chia,
)
from core.su_co import LoiNoiDung

GOC = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def cue(so: int, dai: float = 2.0, bat_dau: float = None):
    """Một dòng phụ đề, mặc định 2 giây một dòng, nối tiếp nhau."""
    dau = (so - 1) * dai if bat_dau is None else bat_dau
    return {"index": so, "start": dau, "end": dau + dai,
            "text": "cau so {0}".format(so)}


def canh_ai(tu: int, den: int, **them):
    m = {"srt_from": tu, "srt_to": den, "img_prompt": "anh {0}".format(tu),
         "video_prompt": "clip {0}".format(tu)}
    m.update(them)
    return m


class TestCanhLai:
    """AI chia theo nghĩa rất tốt nhưng đếm dở — máy vá lại chỗ đếm."""

    def test_phu_het_moi_dong_phu_de(self):
        cues = [cue(i) for i in range(1, 7)]
        ra = canh_lai([canh_ai(1, 2), canh_ai(3, 6)], cues, 8.0)
        phu = [i for c in ra for i in c["_cue"]]
        assert phu == [1, 2, 3, 4, 5, 6], \
            "dòng phụ đề bị bỏ sót thì đoạn đó không có hình"

    def test_ai_bo_sot_dong_o_giua_thi_duoc_va_lai(self):
        """AI trả 1–2 rồi nhảy sang 5–6: hai dòng ở giữa không được mất tiếng."""
        cues = [cue(i) for i in range(1, 7)]
        ra = canh_lai([canh_ai(1, 2), canh_ai(5, 6)], cues, 8.0)
        phu = [i for c in ra for i in c["_cue"]]
        assert phu == [1, 2, 3, 4, 5, 6]

    def test_ai_tra_khoang_chong_lan_thi_bi_ep_ve_lien_mach(self):
        cues = [cue(i) for i in range(1, 7)]
        ra = canh_lai([canh_ai(1, 3), canh_ai(2, 6)], cues, 8.0)
        phu = [i for c in ra for i in c["_cue"]]
        assert phu == [1, 2, 3, 4, 5, 6], "một dòng không được nằm ở hai cảnh"

    def test_ai_bo_quen_duoi_thi_nhap_vao_canh_cuoi(self):
        cues = [cue(i) for i in range(1, 7)]
        ra = canh_lai([canh_ai(1, 3)], cues, 8.0)
        assert ra[-1]["_cue"][-1] == 6, "đuôi bị bỏ là mấy giây không có hình"

    def test_canh_dai_qua_tran_bi_cat_deu(self):
        """Veo3 ra clip 8 giây. Cảnh 20 giây nghĩa là 12 giây không có hình."""
        cues = [cue(i, dai=5.0) for i in range(1, 5)]   # 20 giây liền
        ra = canh_lai([canh_ai(1, 4)], cues, 8.0)
        assert len(ra) == 3, "20 giây phải cắt thành 3 phần dưới trần 8 giây"
        for c in ra:
            assert c["_ket_thuc"] - c["_bat_dau"] <= 8.0 + 1e-9
        assert {c["_tong_phan"] for c in ra} == {3}

    def test_canh_thieu_loi_nhac_nhap_vao_canh_truoc(self):
        """Sót lời nhắc ở MỘT cảnh không được giết cả lượt — đo thật 15/08."""
        cues = [cue(i) for i in range(1, 7)]
        ra = canh_lai([canh_ai(1, 2), {"srt_from": 3, "srt_to": 4},
                       canh_ai(5, 6)], cues, 8.0)
        assert len(ra) == 2, "cảnh thiếu lời nhắc phải nhập vào cảnh trước"
        assert ra[0]["_cue"] == [1, 2, 3, 4]

    def test_canh_dau_tien_thieu_loi_nhac_thi_doi_hoi_lai(self):
        """Không có cảnh nào trước để nhập vào — đáng hỏi lại chứ không đi tiếp."""
        cues = [cue(i) for i in range(1, 5)]
        with pytest.raises(LoiNoiDung):
            canh_lai([{"srt_from": 1, "srt_to": 4}], cues, 8.0)

    def test_ai_tra_rac_thi_bao_ro_chu_khong_tra_bang_rong(self):
        cues = [cue(i) for i in range(1, 5)]
        with pytest.raises(LoiNoiDung):
            canh_lai(["khong phai canh", 123], cues, 8.0)


class TestChiaTheoNghia:
    def test_chia_khuc_roi_ghep_lai_dung_thu_tu(self):
        """Các khúc chạy song song, `ThreadPoolExecutor` không hứa thứ tự xong."""
        cues = [cue(i) for i in range(1, 71)]     # 70 dòng → 3 khúc

        def hoi(khuc, _thu_tu, _tong):
            return [canh_ai(khuc[0]["index"], khuc[-1]["index"])]

        canh = chia_theo_nghia(cues, hoi, tran=8.0, moi_khuc=30)
        assert [c["scene_id"] for c in canh] == list(range(1, len(canh) + 1))
        phu = [i for c in canh for i in c["srt_indices"]]
        assert phu[0] == 1 and phu[-1] == 70

    def test_moi_khuc_duoc_hoi_dung_mot_lan(self):
        cues = [cue(i) for i in range(1, 71)]
        da_hoi = []

        def hoi(khuc, thu_tu, tong):
            da_hoi.append((thu_tu, tong, len(khuc)))
            return [canh_ai(khuc[0]["index"], khuc[-1]["index"])]

        chia_theo_nghia(cues, hoi, tran=8.0, moi_khuc=30, song_song=3)
        assert sorted(t for t, _n, _d in da_hoi) == [0, 1, 2]
        assert {n for _t, n, _d in da_hoi} == {3}

    def test_khong_khuc_nao_qua_tran_engine(self):
        cues = [cue(i, dai=6.0) for i in range(1, 11)]

        def hoi(khuc, _t, _n):
            return [canh_ai(khuc[0]["index"], khuc[-1]["index"])]

        for c in chia_theo_nghia(cues, hoi, tran=8.0):
            assert c["duration"] <= 8.0 + 1e-9

    def test_mot_khuc_hong_thi_dung_han_chu_khong_ra_bang_cut(self):
        cues = [cue(i) for i in range(1, 71)]

        def hoi(khuc, thu_tu, _n):
            if thu_tu == 1:
                raise LoiNoiDung("JSON dut")
            return [canh_ai(khuc[0]["index"], khuc[-1]["index"])]

        with pytest.raises(LoiNoiDung):
            chia_theo_nghia(cues, hoi, tran=8.0, moi_khuc=30)

    def test_moc_thoi_gian_bam_theo_phu_de_that(self):
        cues = [cue(i) for i in range(1, 5)]     # 0→2, 2→4, 4→6, 6→8

        def hoi(khuc, _t, _n):
            return [canh_ai(1, 2), canh_ai(3, 4)]

        canh = chia_theo_nghia(cues, hoi, tran=8.0)
        assert canh[0]["srt_start"] == "00:00:00,000"
        assert canh[0]["srt_end"] == "00:00:04,000"
        assert canh[1]["srt_start"] == "00:00:04,000"

    def test_nhan_vat_mac_dinh_chi_dien_khi_noi_goi_co_yeu_cau(self):
        """Tab Tự động có `nv1.png` thật; Prompt Visuals thì không có tấm nào.

        Điền bừa `nv1` cho Prompt Visuals là mọi lời nhắc trỏ vào một tấm ảnh
        không tồn tại.
        """
        cues = [cue(i) for i in range(1, 5)]

        def hoi(khuc, _t, _n):
            return [canh_ai(1, 4)]

        assert chia_theo_nghia(cues, hoi, tran=8.0)[0]["characters_used"] == ""
        co = chia_theo_nghia(cues, hoi, tran=8.0, nhan_vat_mac_dinh="nv1")
        assert co[0]["characters_used"] == "nv1"


class TestLoiNhac:
    def test_phu_de_gui_di_co_danh_so(self):
        """Không có số thì không ghép lại được với mốc thời gian thật."""
        dong = bang_phu_de([cue(1), cue(2)])
        assert dong.splitlines()[0].startswith("1 | 0.00")

    def test_khuon_mac_dinh_noi_ro_san_va_tran(self):
        chu = loi_nhac_chia(KHUON_MAC_DINH, [cue(1), cue(2)], 10.0)
        assert "3 and 10 seconds" in chu
        assert "cau so 1" in chu
        assert "<<" not in chu, "chỗ chưa điền phải được dọn, không để AI đọc"

    def test_khuon_mac_dinh_khong_bia_ra_anh_nhan_vat(self):
        """Prompt Visuals không có `nv1.png` — nhắc tới nó là trỏ vào chỗ trống."""
        assert "nv1" not in KHUON_MAC_DINH

    def test_chia_khuc_khong_bo_sot_dong_nao(self):
        cues = [cue(i) for i in range(1, 71)]
        khuc = chia_khuc(cues, 30)
        assert [len(k) for k in khuc] == [30, 30, 10]
        assert [c["index"] for k in khuc for c in k] == list(range(1, 71))


def _nap_run_py():
    """Nạp `tool-catalog/prompt.workbook/run.py` — tên thư mục có dấu chấm."""
    duong = os.path.join(GOC, "tool-catalog", "prompt.workbook", "run.py")
    spec = importlib.util.spec_from_file_location("prompt_workbook_run", duong)
    mo_dun = importlib.util.module_from_spec(spec)
    sys.modules["prompt_workbook_run"] = mo_dun
    spec.loader.exec_module(mo_dun)
    return mo_dun


@pytest.fixture(scope="module")
def workbook_tool():
    return _nap_run_py()


SRT = "\n\n".join(
    "{0}\n00:00:{1:02d},000 --> 00:00:{2:02d},000\nCau so {0}".format(
        i, (i - 1) * 3, i * 3)
    for i in range(1, 11))


@pytest.fixture
def yeu_cau(tmp_path):
    srt = tmp_path / "loi.srt"
    srt.write_text(SRT, encoding="utf-8")
    lam = tmp_path / "lam-viec"
    lam.mkdir()
    return {"inputs": {"subtitles": {"path": str(srt), "artifact_id": "a1"}},
            "config": {"engine": "veo3", "model": "claude-sonnet-5"},
            "workspace": str(lam), "workflow_id": "kiem-tra"}


class TestPromptWorkbookDungChungCachChia:
    """Prompt Visuals phải cắt theo nghĩa y như tab Tự động."""

    def test_chia_theo_noi_dung_thi_khong_hoi_them_luot_nao(self, workbook_tool,
                                                            yeu_cau):
        """Chia cảnh và viết lời nhắc là MỘT lượt gọi, không phải hai."""
        goi_them = []

        def chia(khuc, _t, _n):
            return [canh_ai(khuc[0]["index"], khuc[-1]["index"])]

        def enrich(batch, _ctx):
            goi_them.append(len(batch))
            return []

        ra = workbook_tool.handle(yeu_cau, chia_fn=chia, enrich_fn=enrich)
        assert ra["scenes"]["json"]["settings"]["cach_chia"] == \
            workbook_tool.THEO_NGHIA
        assert not goi_them, "cắt theo nghĩa xong còn hỏi lời nhắc là trả tiền hai lần"
        canh = ra["scenes"]["json"]["scenes"]
        assert canh and all(c["img_prompt"] for c in canh)

    def test_khong_goi_duoc_ai_thi_lui_ve_dong_ho_VA_NOI_RA(self, workbook_tool,
                                                            yeu_cau, capsys):
        def chia(_khuc, _t, _n):
            raise RuntimeError("may chu tam gian doan")

        def enrich(batch, _ctx):
            return [{"scene_id": s["scene_id"], "img_prompt": "anh",
                     "video_prompt": "clip"} for s in batch]

        ra = workbook_tool.handle(yeu_cau, chia_fn=chia, enrich_fn=enrich)
        assert ra["scenes"]["json"]["settings"]["cach_chia"] == \
            workbook_tool.THEO_DONG_HO
        loi_nhan = [json.loads(d).get("message", "")
                    for d in capsys.readouterr().out.splitlines() if d.strip()]
        assert any("theo dong ho" in c for c in loi_nhan), \
            "lùi về cắt theo đồng hồ mà im lặng thì khách không cách nào biết"

    def test_du_lui_ve_dong_ho_van_phu_het_phu_de(self, workbook_tool, yeu_cau):
        def enrich(batch, _ctx):
            return [{"scene_id": s["scene_id"], "img_prompt": "anh",
                     "video_prompt": "clip"} for s in batch]

        ra = workbook_tool.handle(
            yeu_cau, chia_fn=lambda *_a: (_ for _ in ()).throw(RuntimeError("x")),
            enrich_fn=enrich)
        canh = ra["scenes"]["json"]["scenes"]
        assert canh[0]["srt_indices"][0] == 1
        assert canh[-1]["srt_indices"][-1] == 10

    def test_khong_dung_toi_khoa_api_khi_da_dua_san_ham_goi(self, workbook_tool,
                                                            yeu_cau, monkeypatch):
        """Bài kiểm không được chạm mạng — và không được đòi khoá của khách."""
        monkeypatch.delenv("SHOPAPI_API_KEY", raising=False)

        def chia(khuc, _t, _n):
            return [canh_ai(khuc[0]["index"], khuc[-1]["index"])]

        assert workbook_tool.handle(yeu_cau, chia_fn=chia)["scenes"]

    def test_file_excel_that_su_duoc_ghi_ra(self, workbook_tool, yeu_cau):
        from openpyxl import load_workbook

        def chia(khuc, _t, _n):
            return [canh_ai(khuc[0]["index"], khuc[-1]["index"])]

        workbook_tool.handle(yeu_cau, chia_fn=chia)
        duong = os.path.join(yeu_cau["workspace"], "scene-prompts.xlsx")
        assert os.path.isfile(duong)
        sach = load_workbook(duong)
        try:
            assert "scenes" in sach.sheetnames
            dau = [o.value for o in next(sach["scenes"].iter_rows(max_row=1))]
            assert dau == workbook_tool.SCENE_COLUMNS
        finally:
            sach.close()


# ── Bản 24/08/2026: khuôn "đạo diễn storyboard" + đuôi cấm ép bằng mã ────────
#
# Chủ dự án: *"prompt tạo ảnh video phải minh hoạ được nội dung"*, chỉ sang
# `D:\AFFILIATE`. Đo trên TL4-T7/0010: 147/297 cảnh nhân vật chỉ ngồi/đứng,
# 297/297 clip "slowly/gently". Bài này khoá phần đo được bằng mã: đuôi cấm,
# thống kê, và khuôn mặc định có đủ các luật + vị trí khúc.

from core.chia_canh import DUOI_CAM, ep_duoi, thong_ke_canh  # noqa: E402


class TestDuoiCam:
    def test_noi_duoi_khi_thieu(self):
        assert ep_duoi("Wide shot of a door.", DUOI_CAM) == (
            "Wide shot of a door, " + DUOI_CAM)

    def test_khong_noi_hai_lan(self):
        co = "Wide shot, No Text, no letters, no numbers, no watermark"
        assert ep_duoi(co, DUOI_CAM) == co

    def test_rong_thi_giu_rong(self):
        assert ep_duoi("", DUOI_CAM) == ""

    def test_chia_theo_nghia_ep_duoi_moi_canh(self):
        cues = [cue(i) for i in range(1, 5)]

        def hoi(khuc, _t, _n):
            return [canh_ai(1, 2), canh_ai(3, 4)]

        for c in chia_theo_nghia(cues, hoi, tran=8.0, duoi=DUOI_CAM):
            assert c["img_prompt"].endswith(DUOI_CAM)
            assert c["video_prompt"].endswith(DUOI_CAM)


class TestThongKe:
    def test_dem_ngoi_dung_cham_va_lap(self):
        canh = [
            {"img_prompt": "Wide shot of a man sitting by a window",
             "video_prompt": "The light slowly warms"},
            {"img_prompt": "Wide shot of a clock melting",
             "video_prompt": "The clock face cracks and the hands fall off"},
            {"img_prompt": "Extreme close-up of a hand",
             "video_prompt": "Gently, the hand opens"},
        ]
        tk = thong_ke_canh(canh)
        assert tk == {"tinh": 1, "cham": 2, "lap": 1, "tong": 3}

    def test_rong(self):
        assert thong_ke_canh([]) == {"tinh": 0, "cham": 0, "lap": 0, "tong": 0}


class TestKhuonDaoDien:
    def test_co_du_cac_luat_cua_affiliate_va_7_canh(self):
        chu = KHUON_MAC_DINH.lower()
        for dau_hieu in ("metaphor", "rejected", "accent", "different at the end",
                         "never a grid", "carry writing", "style tail",
                         "hard ceiling"):
            assert dau_hieu in chu, dau_hieu

    def test_dien_vi_tri_khuc(self):
        chu = loi_nhac_chia(KHUON_MAC_DINH, [cue(1)], 8.0, {
            "KHUC_THU": 3, "TONG_KHUC": 9, "LA_KHUC_DAU": "no",
            "TY_LE_KHUNG": "16:9 horizontal"})
        assert "piece **3 of 9**" in chu
        assert "FIRST piece? **no**" in chu
        assert "16:9 horizontal composition" in chu
        assert "<<" not in chu

    def test_khong_dien_thi_don_sach(self):
        # Nơi gọi cũ không truyền vị trí khúc: chỗ trống phải được dọn, không
        # để AI đọc thấy `<<KHUC_THU>>`.
        chu = loi_nhac_chia(KHUON_MAC_DINH, [cue(1)], 8.0)
        assert "<<" not in chu


# ── Cảnh ngắn hơn sàn thì gộp — đo thật 24/08/2026: cảnh 0,7 giây ─────────

from core.chia_canh import gop_ngan  # noqa: E402


class TestGopNgan:
    def test_canh_lai_gop_canh_ngan_vao_canh_truoc(self):
        # cue 1,2 dài 3s; cue 3 chỉ 0,5s; cue 4 dài 3s.
        cues = [cue(1), cue(2),
                {"index": 3, "start": 6.0, "end": 6.5, "text": "ngan"},
                {"index": 4, "start": 6.5, "end": 9.5, "text": "cau so 4"}]
        ra = canh_lai([canh_ai(1, 2), canh_ai(3, 3), canh_ai(4, 4)], cues, 8.0)
        # Cảnh 3 (0,5s) nhập vào cảnh trước: còn 2 cảnh, phủ đủ 4 dòng.
        assert [c["_cue"] for c in ra] == [[1, 2, 3], [4]]
        assert ra[0]["img_prompt"] == "anh 1"

    def test_canh_dau_ngan_nhap_vao_canh_sau(self):
        cues = [{"index": 1, "start": 0.0, "end": 0.8, "text": "ơ"},
                {"index": 2, "start": 0.8, "end": 4.0, "text": "cau dai"}]
        ra = canh_lai([canh_ai(1, 1), canh_ai(2, 2)], cues, 8.0)
        assert [c["_cue"] for c in ra] == [[1, 2]]
        assert ra[0]["img_prompt"] == "anh 2"

    def test_gop_ngan_thuan(self):
        theo_so = {1: {"start": 0, "end": 3}, 2: {"start": 3, "end": 3.5},
                   3: {"start": 3.5, "end": 7}}
        ds = [{"srt_from": 1, "srt_to": 1}, {"srt_from": 2, "srt_to": 2},
              {"srt_from": 3, "srt_to": 3}]
        ra = gop_ngan(ds, theo_so, "srt_from", "srt_to", 3.0)
        assert [(m["srt_from"], m["srt_to"]) for m in ra] == [(1, 2), (3, 3)]


class TestTachDaiVaDuoi:
    def test_tach_dai_tai_ranh_gioi_dong(self):
        from core.chia_canh import tach_dai
        theo_so = {i: {"start": (i - 1) * 3.0, "end": i * 3.0} for i in range(1, 7)}
        # Một beat 6 dòng = 18 giây, trần 8 → 3 phần [1-2][3-4][5-6], giữ trường khác.
        ra = tach_dai([{"srt_from": 1, "srt_to": 6, "purpose": "p"}], theo_so,
                      "srt_from", "srt_to", 8.0)
        assert [(m["srt_from"], m["srt_to"], m["purpose"]) for m in ra] == [
            (1, 2, "p"), (3, 4, "p"), (5, 6, "p")]

    def test_tach_dai_mot_dong_dai_thi_de_nguyen(self):
        from core.chia_canh import tach_dai
        theo_so = {1: {"start": 0.0, "end": 12.0}}
        assert tach_dai([{"srt_from": 1, "srt_to": 1}], theo_so, "srt_from", "srt_to", 8.0) == [
            {"srt_from": 1, "srt_to": 1}]

    def test_ep_duoi_khong_noi_khi_da_du_y(self):
        co = "Wide shot, no readable text, no letters, no numbers, no watermark"
        assert ep_duoi(co, DUOI_CAM) == co
        assert ep_duoi("Wide shot, no watermark", DUOI_CAM).endswith(DUOI_CAM)


class TestChiaKhucTheoGiay:
    def test_chan_theo_giay_va_theo_so_dong(self):
        # 10 dòng, mỗi dòng 6 giây (kiểu tiếng Nhật): trần 20 giây → 3 dòng/khúc.
        cues = [cue(i, 6.0) for i in range(1, 11)]
        khuc = chia_khuc(cues, 30, giay_moi_khuc=20.0)
        assert [len(k) for k in khuc] == [3, 3, 3, 1]
        assert [c["index"] for k in khuc for c in k] == list(range(1, 11))

    def test_khong_dat_giay_thi_nhu_cu(self):
        cues = [cue(i, 6.0) for i in range(1, 11)]
        assert [len(k) for k in chia_khuc(cues, 4)] == [4, 4, 2]

    def test_dong_don_le_dai_hon_tran_van_thanh_khuc(self):
        cues = [cue(1, 50.0), cue(2, 50.0)]
        assert [len(k) for k in chia_khuc(cues, 30, giay_moi_khuc=20.0)] == [1, 1]
