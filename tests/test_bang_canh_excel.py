"""Tab Ảnh & Video nhận Excel, và ảnh tham chiếu phải tới được máy chủ.

Chủ dự án, 15/08/2026: *"ở tab tạo ảnh video có thêm mục tải excel lên không
phải txt… tab hàng loạt hay thủ công cần có chỗ điền ô đường dẫn ảnh tham
chiếu, vì tạo ảnh và video đều cần tham chiếu"*.

Rà soát tìm ra một lỗ chưa ai nêu: tab Hàng loạt **gõ cứng
`reference_images = None`** ở hai chỗ, nên mọi ảnh tạo hàng loạt đều không bám
nhân vật nào — mỗi cảnh ra một người khác nhau, đúng thứ `nv1.png` sinh ra để
chặn. Bài `test_reference_images_khong_con_la_None` dưới đây canh đúng chỗ ấy.

Không bài nào gọi mạng.
"""

from __future__ import annotations

import os

import pytest

pytest.importorskip("openpyxl")

from core.bang_canh_excel import (  # noqa: E402
    COT, COT_BAT_BUOC, LoiBangCanh, doc_excel, viet_mau,
)


def _sach(tmp_path, ten, cot, cac_dong):
    from openpyxl import Workbook

    s = Workbook()
    t = s.active
    t.title = "scenes"
    t.append(list(cot))
    for d in cac_dong:
        t.append(list(d))
    duong = str(tmp_path / ten)
    s.save(duong)
    return duong


class TestVongTronKhepKin:
    """Tải mẫu → điền → nạp lên. Hỏng chỗ này là tính năng vô dụng."""

    def test_mau_tai_ve_roi_nap_len_duoc_ngay(self, tmp_path):
        mau = viet_mau(str(tmp_path / "mau.xlsx"))
        dong = doc_excel(mau)
        assert len(dong) == 3, "file mẫu phải có sẵn ba dòng để khách thấy ba kiểu dùng"

    def test_mau_co_trang_huong_dan(self, tmp_path):
        from openpyxl import load_workbook

        s = load_workbook(viet_mau(str(tmp_path / "mau.xlsx")))
        assert "scenes" in s.sheetnames
        assert "huong-dan" in s.sheetnames, \
            "không có trang này thì khách không biết cột nào bắt buộc"

    def test_ba_dong_mau_la_ba_kieu_dung_khac_nhau(self, tmp_path):
        dong = doc_excel(viet_mau(str(tmp_path / "mau.xlsx")))
        assert dong[0]["anh"] and dong[0]["video"], "dòng 1: cả ảnh lẫn clip"
        assert dong[1]["anh"] and not dong[1]["video"], "dòng 2: chỉ ảnh"
        assert not dong[2]["anh"] and dong[2]["video"], "dòng 3: chỉ clip"

    def test_mau_dien_duong_dan_khong_phai_ten_file(self, tmp_path):
        """Chủ dự án 22/08/2026: điền tên file khó, phải là ĐƯỜNG DẪN đầy đủ."""
        dong = doc_excel(viet_mau(str(tmp_path / "mau.xlsx")))
        assert "\\" in dong[0]["tham_chieu"], "ô tham chiếu mẫu phải là đường dẫn"
        assert "," in dong[1]["tham_chieu"], "dòng 2 làm mẫu nhiều ảnh cách dấu phẩy"

    def test_mau_huong_dan_cach_lay_duong_dan(self, tmp_path):
        from openpyxl import load_workbook

        s = load_workbook(viet_mau(str(tmp_path / "mau.xlsx")))
        chu = "\n".join(
            str(o.value or "")
            for hang in s["huong-dan"].iter_rows() for o in hang)
        assert "đường dẫn" in chu.lower()
        assert "Copy as path" in chu, "phải chỉ khách cách lấy đường dẫn"


class TestGiuTenCotTiengAnh:
    """Giữ tên VE3 để file từ Prompt Visuals nạp thẳng sang được."""

    def test_ten_cot_dung_khuon_VE3(self):
        assert COT == ("scene_id", "img_prompt", "video_prompt",
                       "reference_files")

    def test_file_kieu_prompt_visuals_nap_thang_duoc(self, tmp_path):
        """File từ tab Prompt Visuals có thừa nhiều cột — vẫn phải nạp được."""
        thua = ("scene_id", "srt_start", "srt_end", "duration", "srt_text",
                "img_prompt", "video_prompt", "reference_files", "status_img")
        duong = _sach(tmp_path, "tu-pv.xlsx", thua, [
            (1, "00:00:00,000", "00:00:06,000", 6.0, "Câu một",
             "a quiet room", "slow push in", "nv1.png", "pending"),
        ])
        dong = doc_excel(duong)
        assert dong[0]["anh"] == "a quiet room"
        assert dong[0]["tham_chieu"] == "nv1.png"


class TestTuChoiChoRaHon:
    """Khách điền sai mà tool im lặng chạy là hỏng cả mẻ."""

    def test_thieu_cot_thi_GOI_TEN_cot_thieu(self, tmp_path):
        duong = _sach(tmp_path, "thieu.xlsx", ("scene_id", "img_prompt"),
                      [(1, "a room")])
        with pytest.raises(LoiBangCanh, match="video_prompt"):
            doc_excel(duong)

    def test_file_khong_co_dong_nao_dien_thi_bao_ro(self, tmp_path):
        duong = _sach(tmp_path, "trong.xlsx", COT, [(1, "", "", "nv1.png")])
        with pytest.raises(LoiBangCanh, match="img_prompt hoặc video_prompt"):
            doc_excel(duong)

    def test_khong_thay_file_thi_bao_ro(self, tmp_path):
        with pytest.raises(LoiBangCanh, match="Không thấy"):
            doc_excel(str(tmp_path / "khong-co.xlsx"))

    def test_khong_doi_cot_reference_files(self, tmp_path):
        """Bỏ trống được — khách có thể dùng ảnh chung cho cả loạt."""
        duong = _sach(tmp_path, "khong-tc.xlsx", COT_BAT_BUOC,
                      [(1, "a room", "push in")])
        assert doc_excel(duong)[0]["tham_chieu"] == ""


class TestDocNoiDung:
    def test_bo_qua_dong_trong_o_cuoi_file(self, tmp_path):
        duong = _sach(tmp_path, "co-dong-thua.xlsx", COT, [
            (1, "a room", "", "nv1.png"),
            (None, None, None, None),
            (None, "", "", ""),
        ])
        assert len(doc_excel(duong)) == 1

    def test_thieu_scene_id_thi_tu_danh_so(self, tmp_path):
        duong = _sach(tmp_path, "khong-so.xlsx", COT, [
            ("", "a room", "", ""), ("", "another room", "", "")])
        assert [d["so"] for d in doc_excel(duong)] == ["1", "2"]


pytest.importorskip("PyQt5.QtWidgets", reason="máy chạy test không có giao diện")


@pytest.fixture(scope="module")
def qt_app():
    """Một `QApplication` cho cả tệp.

    Dựng nó **trong fixture module-scope**, không dựng lẻ trong từng bài: hai
    `QApplication` trong một tiến trình là Qt chết hẳn, không phải báo lỗi. Đây
    là khuôn `tests/test_auto_luot.py` đã chạy được, dùng lại chứ đừng tự chế.
    """
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    from PyQt5.QtWidgets import QApplication

    return QApplication.instance() or QApplication([])


class _AppGia:
    """Cửa sổ chính giả — chỉ đủ thứ `TabHangLoat` hỏi tới."""

    client = None
    prices = None

    def __init__(self, thu_muc: str):
        self._thu_muc = thu_muc
        self.da_hien = []
        self.da_chay = []

    def default_output_dir(self, _kind):
        return self._thu_muc

    def show_message(self, tieu_de, chu):
        self.da_hien.append((tieu_de, chu))

    def show_error(self, loi):
        self.da_hien.append(("loi", str(loi)))

    def start_batch(self, specs, folder=""):
        self.da_chay.append((list(specs), folder))

    def run_bg(self, viec, on_ok=None, on_err=None):
        """Chạy ngay tại chỗ. Bài kiểm không có vòng sự kiện để mà đợi luồng."""
        try:
            ket = viec()
        except Exception as loi:  # noqa: BLE001
            if on_err:
                on_err(loi)
            return
        if on_ok:
            on_ok(ket)


def _dung_tab(thu_muc: str):
    from ui_qt.trang_anh_video import TabHangLoat

    app = _AppGia(thu_muc)
    return TabHangLoat(app), app


class TestTabHangLoat:
    """Ảnh tham chiếu phải thật sự tới được `JobSpec`."""

    def test_reference_images_khong_con_la_None(self, qt_app, tmp_path):
        """Đây là lỗi thật: mọi ảnh hàng loạt đều không bám nhân vật nào.

        Gõ cứng `None` thì mỗi cảnh ra một người khác — đúng thứ ảnh tham
        chiếu sinh ra để chặn.
        """
        tab, _app = _dung_tab(str(tmp_path))
        anh = tmp_path / "nv1.png"
        anh.write_bytes(b"anh-gia")
        tab.bang.setRowCount(0)
        tab.them_dong("a quiet room", "", str(anh))
        assert tab._anh_cua_dong(0) == [str(anh)], \
            "cột “Ảnh tham chiếu” của dòng phải tới được chỗ dựng việc"

    def test_o_tham_chieu_nhan_nhieu_duong_dan_cach_nhau_dau_phay(
            self, qt_app, tmp_path):
        """Chủ dự án 22/08/2026: điền ĐƯỜNG DẪN, nhiều ảnh cách nhau dấu phẩy."""
        tab, _app = _dung_tab(str(tmp_path))
        a = tmp_path / "nv1.png"
        b = tmp_path / "nv1-nghieng.png"
        for t in (a, b):
            t.write_bytes(b"anh-gia")
        tab.bang.setRowCount(0)
        tab.them_dong("a room", "", "{0}, {1}".format(a, b))
        assert tab._anh_cua_dong(0) == [str(a), str(b)]

    def test_o_tham_chieu_boc_ngoac_kep_kieu_copy_as_path(self, qt_app, tmp_path):
        """“Sao chép dưới dạng đường dẫn” của Windows kẹp ngoặc kép — phải bóc."""
        tab, _app = _dung_tab(str(tmp_path))
        anh = tmp_path / "nv1.png"
        anh.write_bytes(b"anh-gia")
        tab.bang.setRowCount(0)
        tab.them_dong("a room", "", '"{0}"'.format(anh))
        assert tab._anh_cua_dong(0) == [str(anh)]

    def test_o_tham_chieu_ten_tro_troi_bi_bo_qua(self, qt_app, tmp_path):
        """Tên không phải đường dẫn thật → bỏ, không âm thầm gán nhầm."""
        tab, _app = _dung_tab(str(tmp_path))
        tab.bang.setRowCount(0)
        tab.them_dong("a room", "", "nv1.png")
        assert tab._anh_cua_dong(0) == []

    def test_anh_rieng_cua_dong_thang_anh_chung(self, qt_app, tmp_path):
        tab, _app = _dung_tab(str(tmp_path))
        # Tệp phải có thật: `AnhThamChieu` từ chối đường dẫn không tồn tại, và
        # nó từ chối đúng — bắt lỗi ở lúc chọn hơn là lúc gửi đi.
        chung = tmp_path / "chung.png"
        rieng = tmp_path / "rieng.png"
        for t in (chung, rieng):
            t.write_bytes(b"anh-gia")
        tab.anh_vao.dat(str(chung))
        tab.bang.setRowCount(0)
        tab.them_dong("a room", "", str(rieng))
        tab.them_dong("another room", "", "")
        assert tab._anh_cua_dong(0) == [str(rieng)], "dòng điền riêng thì thắng"
        assert tab._anh_cua_dong(1) == [str(chung)], "dòng bỏ trống thì dùng chung"

    def test_dong_chi_co_mo_ta_clip_van_duoc_nhan(self, qt_app, tmp_path):
        """Trước 15/08/2026 `canh()` bỏ qua dòng không có mô tả ảnh."""
        tab, _app = _dung_tab(str(tmp_path))
        tab.bang.setRowCount(0)
        tab.them_dong("", "the camera drifts left", "nv1.png")
        assert len(tab.canh()) == 1
        assert tab.canh()[0][1] == "", "mô tả ảnh trống là đúng ý"
        assert tab.canh()[0][2] == "the camera drifts left"

    def test_dong_trong_hoan_toan_van_bi_bo_qua(self, qt_app, tmp_path):
        tab, _app = _dung_tab(str(tmp_path))
        tab.bang.setRowCount(0)
        tab.them_dong("", "", "nv1.png")
        assert tab.canh() == []

    def test_nap_excel_do_thang_vao_bang(self, qt_app, tmp_path):
        """Vòng tròn khép kín: tải mẫu, nạp lên, bảng có đúng ba dòng."""
        tab, _app = _dung_tab(str(tmp_path))
        mau = viet_mau(str(tmp_path / "mau.xlsx"))
        tab.bang.setRowCount(0)
        for m in doc_excel(mau):
            tab.them_dong(m["anh"], m["video"], m["tham_chieu"])
        assert tab.bang.rowCount() == 3
        assert len(tab.canh()) == 3, "cả ba kiểu dòng đều phải có việc để làm"


class TestBaCheDo:
    """Ba chế độ (Tạo ảnh / Tạo video / Ảnh→Video) — trả lời "bắt đầu từ đâu"."""

    def test_che_do_an_hien_dung_cot(self, qt_app, tmp_path):
        from ui_qt.trang_anh_video import CD_ANH, CD_VIDEO, CD_CHUOI, _CotBang
        tab, _app = _dung_tab(str(tmp_path))

        tab._dat_che_do(CD_ANH)
        assert tab.bang.isColumnHidden(_CotBang.VIDEO)
        assert not tab.bang.isColumnHidden(_CotBang.ANH)
        # Cột Trạng thái / Kết quả / Làm lại luôn hiện ở mọi chế độ.
        assert not tab.bang.isColumnHidden(_CotBang.TRANG_THAI)
        assert not tab.bang.isColumnHidden(_CotBang.LAM_LAI)

        tab._dat_che_do(CD_VIDEO)
        assert tab.bang.isColumnHidden(_CotBang.ANH)
        assert not tab.bang.isColumnHidden(_CotBang.VIDEO)
        assert not tab.bang.isColumnHidden(_CotBang.TRANG_THAI)
        assert not tab.bang.isColumnHidden(_CotBang.LAM_LAI)

        tab._dat_che_do(CD_CHUOI)
        for cot in (_CotBang.ANH, _CotBang.VIDEO,
                    _CotBang.TRANG_THAI, _CotBang.LAM_LAI):
            assert not tab.bang.isColumnHidden(cot)

    def test_che_do_anh_khong_lam_video_du_con_sot_chu(self, qt_app, tmp_path):
        """Gõ cả hai ở chế độ chuỗi rồi đổi sang "Tạo ảnh" thì không lỡ tạo clip."""
        from ui_qt.trang_anh_video import CD_ANH, CD_VIDEO
        tab, _app = _dung_tab(str(tmp_path))
        tab.bang.setRowCount(0)
        tab.them_dong("a room", "the camera drifts", "")

        tab._dat_che_do(CD_ANH)
        assert tab.canh() == [(0, "a room", "")], "chế độ ảnh bỏ mô tả video"

        tab._dat_che_do(CD_VIDEO)
        assert tab.canh() == [(0, "", "the camera drifts")], "chế độ video bỏ mô tả ảnh"

    def test_nut_tham_chieu_luu_duong_dan(self, qt_app, tmp_path):
        """Ảnh tham chiếu riêng lưu trên nút trong ô, không gõ đường dẫn."""
        from ui_qt.trang_anh_video import _CotBang
        tab, _app = _dung_tab(str(tmp_path))
        anh = tmp_path / "nv1.png"
        anh.write_bytes(b"anh-gia")
        tab.bang.setRowCount(0)
        dong = tab.them_dong("a room", "", str(anh))
        nut = tab.bang.cellWidget(dong, _CotBang.THAM_CHIEU)
        assert nut is not None and nut.duong_dan == [str(anh)]
        assert tab._anh_cua_dong(dong) == [str(anh)]

    def test_dong_chi_clip_gui_video_khong_TypeError(self, qt_app, tmp_path):
        """Lỗi cũ: `_gui_video(dong, mo_ta, url)` sai chữ ký → TypeError lúc chạy.

        Không gọi mạng thật: client giả trả URL giả, `start_batch` chỉ ghi lại.
        Chế độ "Tạo video" gửi dòng chỉ-clip qua đúng đường `_gui_video`.
        """
        from core.pricing import DEFAULT_PRICES
        from ui_qt.trang_anh_video import CD_VIDEO, KIND_VIDEO

        class _Uploads:
            def upload_file(self, path):
                return "https://fake/" + os.path.basename(path)

        class _Client:
            uploads = _Uploads()

        tab, app = _dung_tab(str(tmp_path))
        app.client = _Client()
        app.prices = DEFAULT_PRICES
        anh = tmp_path / "khung.png"
        anh.write_bytes(b"anh-gia")
        tab._dat_che_do(CD_VIDEO)
        tab.bang.setRowCount(0)
        tab.them_dong("", "the camera drifts left", str(anh))

        tab.chay()  # không được ném TypeError

        assert len(app.da_chay) == 1, "phải gửi đúng một lô video"
        specs, _folder = app.da_chay[0]
        assert len(specs) == 1 and specs[0].kind == KIND_VIDEO
        assert specs[0].params["image_url"] == "https://fake/khung.png"


def test_hai_tab_con_deu_co_o_anh_tham_chieu():
    """Chủ dự án: *"tạo ảnh và video đều cần tham chiếu"* — cả hai lối làm việc."""
    import inspect

    from ui_qt.trang_anh_video import TabHangLoat, TabThuCong

    for lop in (TabThuCong, TabHangLoat):
        ma = inspect.getsource(lop)
        assert "AnhThamChieu(" in ma,             "{0} chưa có ô ảnh tham chiếu".format(lop.__name__)
