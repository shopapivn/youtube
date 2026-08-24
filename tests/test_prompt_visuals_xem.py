"""Xem trước prompt từng cảnh trong tab Prompt Visuals.

Chủ dự án: tab visual cần thiết kế lại theo lối tab Tự động — chạy xong phải
**thấy prompt từng cảnh, đánh số, đúng thứ tự**, chứ không chỉ ném ra một file
Excel rồi báo "xong". Bài này khoá lại phần ghép/sắp thuần tuý (`canh_de_xem`,
`dan_de_xem`, `tom_tat_dan`) — không mạng, không giao diện.
"""

from __future__ import annotations

from core.prompt_visuals import canh_de_xem, dan_de_xem, tom_tat_dan


def _scenes_hang(so_thu_tu):
    """Dựng các hàng sheet `scenes` (hàng đầu tiêu đề) theo thứ tự cho trước."""
    tieu_de = ["scene_id", "srt_start", "srt_end", "img_prompt", "video_prompt",
               "characters_used"]
    hang = [tieu_de]
    for so in so_thu_tu:
        hang.append([so, "00:00", "00:08", "anh {0}".format(so),
                     "clip {0}".format(so), "nv1"])
    return hang


def test_canh_xep_theo_so_du_thu_tu_trong_file_lon_xon():
    # Các khúc chạy song song nên file có thể ghi lộn xộn 3,1,2.
    canh = canh_de_xem(_scenes_hang([3, 1, 2]))
    assert [c["scene_id"] for c in canh] == [1, 2, 3]
    assert canh[0]["img_prompt"] == "anh 1"
    assert canh[2]["video_prompt"] == "clip 3"


def test_canh_bo_hang_khong_co_so_va_hang_rong():
    hang = _scenes_hang([1, 2])
    hang.append([None, "", "", "", "", ""])          # hàng trống
    hang.append(["", "", "", "rác", "rác", ""])      # thiếu số cảnh
    canh = canh_de_xem(hang)
    assert [c["scene_id"] for c in canh] == [1, 2]


def test_canh_file_chi_co_tieu_de_thi_rong():
    assert canh_de_xem([["scene_id", "img_prompt"]]) == []
    assert canh_de_xem([]) == []


def test_scene_id_dang_so_thuc_van_doc_duoc():
    # openpyxl đọc số về dạng float (1.0); vẫn phải ra cảnh #1.
    canh = canh_de_xem(_scenes_hang([1.0, 2.0]))
    assert [c["scene_id"] for c in canh] == [1, 2]


def test_dan_nhan_vat_bo_hang_thieu_id():
    tieu_de = ["id", "role", "name", "english_prompt"]
    hang = [tieu_de,
            ["nv1", "protagonist", "Lan", "a young woman"],
            ["", "", "", "no id -> bỏ"]]
    dan = dan_de_xem(hang)
    assert len(dan) == 1
    assert dan[0]["id"] == "nv1"
    assert dan[0]["name"] == "Lan"


def test_tom_tat_dan_co_va_khong_co_nhan_vat():
    assert "tự do" in tom_tat_dan([])
    chu = tom_tat_dan([{"id": "nv1", "role": "protagonist", "name": "Lan"},
                       {"id": "nv2", "role": "friend", "name": ""}])
    assert "nv1" in chu and "nv2" in chu
    assert "Lan" in chu


def test_doc_lai_workbook_that_qua_render_workbook(tmp_path):
    """Ghép với `prompt.workbook.render_workbook`: xuất file thật rồi đọc lại.

    Đây là hợp đồng thật giữa hai đầu — cột `scenes`/`characters` mà lệch tên
    thì xem trước ra bảng trống mà không ai biết vì sao.
    """
    import importlib.util
    import os

    from openpyxl import load_workbook

    goc = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    duong_run = os.path.join(goc, "tool-catalog", "prompt.workbook", "run.py")
    spec = importlib.util.spec_from_file_location("pw_run_xem", duong_run)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    manifest = {
        "scenes": [
            {"scene_id": 2, "srt_start": "00:08", "srt_end": "00:16",
             "img_prompt": "anh hai", "video_prompt": "clip hai",
             "characters_used": "nv1"},
            {"scene_id": 1, "srt_start": "00:00", "srt_end": "00:08",
             "img_prompt": "anh mot", "video_prompt": "clip mot",
             "characters_used": "nv1"},
        ],
        "characters": [
            {"id": "nv1", "role": "protagonist", "name": "Lan",
             "english_prompt": "a young woman in a red coat"},
        ],
    }
    out = tmp_path / "scene-prompts.xlsx"
    mod.render_workbook(out, manifest)

    wb = load_workbook(str(out), read_only=True, data_only=True)
    try:
        def hang_cua(ten):
            return [list(r) for r in wb[ten].iter_rows(values_only=True)]
        canh = canh_de_xem(hang_cua("scenes"))
        dan = dan_de_xem(hang_cua("characters"))
    finally:
        wb.close()

    assert [c["scene_id"] for c in canh] == [1, 2]
    assert canh[0]["img_prompt"] == "anh mot"
    assert dan and dan[0]["name"] == "Lan"
    assert "nv1" in tom_tat_dan(dan)
