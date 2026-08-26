"""Ba cách kể chuyện + bối cảnh tham chiếu + ảnh bìa + nhạc Suno trong Excel.

Chủ dự án, 24/08/2026: *"excel vẫn chưa có prompt tạo thumbnail và nhạc suno…
loại 1 là có 1 nhân vật cố định cho kênh; loại 2 có 1 nhân vật cố định nhưng
vẫn cần các nhân vật tham chiếu khác và bối cảnh tham chiếu; loại 3 có thể xây
nhân vật tham chiếu và bối cảnh tuỳ nội dung"* (tham khảo VE3_SUITE).

Không bài nào gọi mạng: casting, chia cảnh, ảnh bìa, nhạc đều bơm hàm giả.
"""

from __future__ import annotations

import json
import os

from openpyxl import load_workbook

from core.prompt_visuals import (
    CHE_DO_KE, bia_de_xem, boi_canh_de_xem, dung_boi_canh, dung_workflow,
    nhac_de_xem,
)
from test_nhan_vat_xuyen_suot import (  # noqa: F401 — dùng lại fixture
    _chia_giu_khuc, canh_ai, wb, yeu_cau,
)


def _cast_phu_va_boi_canh(_cues, _ctx):
    return {"style": {"image_style": "ink", "palette": "grey", "motion": "slow"},
            "characters": [
                {"id": "nv1", "role": "protagonist",
                 "english_prompt": "AI thu dinh nghia lai"},
                {"id": "nv2", "role": "friend",
                 "english_prompt": "a tall friend in a green coat"}],
            "locations": [
                {"id": "loc1", "name": "Kitchen",
                 "english_prompt": "a small tiled kitchen",
                 "location_lock": "blue tiles",
                 "lighting_default": "morning window light"}]}


def _bia_gia(_cues, _ctx, _cast):
    return {"title": "Khi ban im lang", "thumb_text": "IM LANG",
            "thumbnails": [
                {"version_desc": "portrait_main", "img_prompt": "close portrait"},
                {"version_desc": "dramatic_scene", "img_prompt": "wide scene"},
                {"version_desc": "youtube_ctr", "img_prompt": "symbol front"}]}


def _nhac_gia(_cues, _ctx, _cast):
    return {"music": [{"music_id": 1, "start_time": 0, "end_time": 999,
                       "suno_prompt": "Ambient piano. No vocals, instrumental only.",
                       "mood": "calm"}]}


def _sheet(yeu_cau, ten):
    duong = os.path.join(str(yeu_cau["workspace"]), "scene-prompts.xlsx")
    sach = load_workbook(duong, read_only=True, data_only=True)
    try:
        return [list(r) for r in sach[ten].iter_rows(values_only=True)]
    finally:
        sach.close()


# ── Loại 1: một nhân vật cố định, không casting ─────────────────────────────

def test_loai_1_khong_goi_casting_chi_co_nv1(wb, yeu_cau):
    goi = {"n": 0}

    def cast_fn(*_a):
        goi["n"] += 1
        return _cast_phu_va_boi_canh(None, None)

    yeu_cau["config"]["che_do_ke"] = "mot_nhan_vat"
    ra = wb.handle(yeu_cau, cast_fn=cast_fn, chia_fn=_chia_giu_khuc)
    m = ra["scenes"]["json"]
    assert goi["n"] == 0, "loại 1 không được tốn một lượt casting"
    assert [c["id"] for c in m["characters"]] == ["nv1"]
    assert m["characters"][0]["image_file"] == "nv1.png"
    assert m["locations"] == []
    assert m["settings"]["che_do_ke"] == "mot_nhan_vat"
    assert all(s["characters_used"] == "nv1" for s in m["scenes"])
    # Sheet characters: nv1 có ảnh sẵn → status done.
    hang = _sheet(yeu_cau, "characters")
    assert hang[1][0] == "nv1" and hang[1][6] == "done"


# ── Loại 2: nv1 cố định + AI dựng nv2.. và bối cảnh ─────────────────────────

def test_loai_2_nv1_giu_nguyen_ai_them_phu_va_boi_canh(wb, yeu_cau):
    yeu_cau["config"]["che_do_ke"] = "nhan_vat_va_boi_canh"
    ra = wb.handle(yeu_cau, cast_fn=_cast_phu_va_boi_canh, chia_fn=_chia_giu_khuc)
    m = ra["scenes"]["json"]
    assert [c["id"] for c in m["characters"]] == ["nv1", "nv2"]
    # nv1 là của kênh, AI có trả về nv1 cũng không được định nghĩa lại.
    assert "AI thu dinh nghia lai" not in m["characters"][0]["english_prompt"]
    assert m["characters"][0].get("co_dinh") is True
    assert [l["id"] for l in m["locations"]] == ["loc1"]
    hang = _sheet(yeu_cau, "locations")
    assert hang[0][:3] == ["id", "name", "english_prompt"]
    assert hang[1][0] == "loc1" and hang[1][5] == "loc1.png"


def test_khoi_cast_style_co_boi_canh_va_luat_nv1(wb):
    cast = {"style": {}, "characters": [wb._nhan_vat_co_dinh({})],
            "locations": [{"id": "loc1", "name": "Kitchen",
                           "english_prompt": "tiled kitchen",
                           "lighting_default": "morning"}]}
    khoi = wb._khoi_cast_style(cast)
    assert "RECURRING LOCATIONS" in khoi and "loc1" in khoi
    assert "location_used" in khoi
    assert "nv1 (nv1.png)" in khoi and "NEVER describe" in khoi


# ── Loại 3: AI tự xây cả dàn lẫn bối cảnh ───────────────────────────────────

def test_loai_3_ai_xay_ca_dan_va_boi_canh(wb, yeu_cau):
    yeu_cau["config"]["che_do_ke"] = "tu_xay"
    ra = wb.handle(yeu_cau, cast_fn=_cast_phu_va_boi_canh, chia_fn=_chia_giu_khuc)
    m = ra["scenes"]["json"]
    assert [c["id"] for c in m["characters"]] == ["nv1", "nv2"]
    assert m["characters"][0]["english_prompt"] == "AI thu dinh nghia lai"
    assert [l["id"] for l in m["locations"]] == ["loc1"]


# ── Ảnh bìa + nhạc Suno vào Excel ───────────────────────────────────────────

def test_thumbnail_va_music_vao_sheet(wb, yeu_cau):
    ra = wb.handle(yeu_cau, cast_fn=_cast_phu_va_boi_canh, chia_fn=_chia_giu_khuc,
                   bia_fn=_bia_gia, nhac_fn=_nhac_gia)
    m = ra["scenes"]["json"]
    assert m["title"] == "Khi ban im lang" and m["thumb_text"] == "IM LANG"
    assert [t["version_desc"] for t in m["thumbnails"]] == [
        "portrait_main", "dramatic_scene", "youtube_ctr"]
    assert m["thumbnails"][0]["characters_used"] == "nv1"
    # Nhạc: thời gian được kẹp vào độ dài thật của video (10 câu × 3 giây).
    assert m["music"][0]["end_time"] == 30.0
    bia = bia_de_xem(_sheet(yeu_cau, "thumbnail"))
    assert len(bia) == 3 and bia[0]["thumb_text"] == "IM LANG" and bia[0]["title"]
    nhac = nhac_de_xem(_sheet(yeu_cau, "music"))
    assert nhac[0]["suno_prompt"].endswith("instrumental only.")
    assert boi_canh_de_xem(_sheet(yeu_cau, "locations"))[0]["name"] == "Kitchen"


def test_bom_tay_khong_co_ham_bia_nhac_thi_bo_qua(wb, yeu_cau):
    # Có chia_fn mà không có bia_fn/nhac_fn → không tự gọi mạng, sheet trống.
    ra = wb.handle(yeu_cau, cast_fn=_cast_phu_va_boi_canh, chia_fn=_chia_giu_khuc)
    m = ra["scenes"]["json"]
    assert m["thumbnails"] == [] and m["music"] == []


def test_bia_hong_khong_giet_luot(wb, yeu_cau):
    def bia_no(*_a):
        raise RuntimeError("503")

    ra = wb.handle(yeu_cau, cast_fn=_cast_phu_va_boi_canh, chia_fn=_chia_giu_khuc,
                   bia_fn=bia_no, nhac_fn=_nhac_gia)
    assert ra["scenes"]["json"]["thumbnails"] == []
    assert len(ra["scenes"]["json"]["scenes"]) >= 1


def test_nhac_nhieu_track_noi_lien_va_phu_het(wb):
    raw = {"music": [
        {"end_time": 10, "suno_prompt": "a. No vocals, instrumental only."},
        {"end_time": 5, "suno_prompt": "b. No vocals, instrumental only."}]}
    ra = wb._sach_nhac(raw, 30.0)
    assert [(m["start_time"], m["end_time"]) for m in ra] == [
        (0.0, 10.0), (10.0, 30.0)]


# ── Cột bản dịch tiếng Việt ─────────────────────────────────────────────────

def test_cot_srt_text_vi_di_vao_excel(wb, yeu_cau):
    def chia(khuc, _t, _n):
        return [canh_ai(khuc[0]["index"], khuc[-1]["index"],
                        narration_vi="lời dịch")]

    ra = wb.handle(yeu_cau, chia_fn=chia)
    assert ra["scenes"]["json"]["scenes"][0]["srt_text_vi"] == "lời dịch"
    assert "srt_text_vi" in wb.SCENE_COLUMNS


# ── Phía UI: context + workflow ─────────────────────────────────────────────

def test_boi_canh_loai_1_2_mang_nhan_vat_co_dinh():
    ra = dung_boi_canh("", "auto", che_do_ke="mot_nhan_vat",
                       nhan_vat_co_dinh={"english_prompt": "pencil person"})
    assert ra["story_mode"] == "mot_nhan_vat"
    assert ra["fixed_character"]["id"] == "nv1"
    assert ra["fixed_character"]["image_file"] == "nv1.png"
    assert dung_boi_canh("", "auto", che_do_ke="tu_xay") == {}


def test_workflow_mang_che_do_ke_va_co_bia_nhac():
    wf = dung_workflow("art-1", che_do_ke="nhan_vat_va_boi_canh", bia=True,
                       nhac=False)
    cfg = [n for n in wf["nodes"] if n["id"] == "prompt"][0]["config"]
    assert cfg["che_do_ke"] == "nhan_vat_va_boi_canh"
    assert cfg["thumbnail"] is True and cfg["nhac"] is False
    assert {ma for ma, _t, _m in CHE_DO_KE} == {
        "tu_xay", "mot_nhan_vat", "nhan_vat_va_boi_canh"}


# ── Đạo diễn (loại 2, 3): đọc phim → màn → kế hoạch beat → prompt có kế hoạch ─

def _phim_gia(_cues, _ctx):
    return {"genre": "psychology", "arc": "man_in_hole",
            "context_lock": "rainy city, cold blue with one warm lamp",
            "segments": [
                {"segment_id": 1, "name": "Mo dau", "message": "co don",
                 "emotion": "quiet", "motif": "unanswered phone",
                 "srt_from": 1, "srt_to": 4},
                {"segment_id": 2, "name": "Ket", "message": "chap nhan",
                 "emotion": "warm", "motif": "lamp", "srt_from": 5, "srt_to": 99}],
            "characters_mentioned": ["the narrator"], "locations_mentioned": ["kitchen"]}


def _ke_hoach_gia(seg, dong, _cast):
    dau, cuoi = dong[0]["index"], dong[-1]["index"]
    return {"beats": [
        {"srt_from": dau, "srt_to": dau, "purpose": "hook", "characters": "nv1",
         "location": "loc1", "shot_size": "WIDE", "camera": "push_in",
         "element_motion": "phone lights up", "emotion": "tense", "motif": "phone"},
        {"srt_from": dau + 1, "srt_to": cuoi + 50, "purpose": "turn", "characters": "",
         "location": "", "shot_size": "CLOSE", "camera": "static",
         "element_motion": "light dims", "emotion": "quiet", "motif": ""}]}


def test_dao_dien_loai_3_ra_story_va_director_plan(wb, yeu_cau):
    yeu_cau["config"]["che_do_ke"] = "tu_xay"
    ra = wb.handle(yeu_cau, cast_fn=_cast_phu_va_boi_canh, chia_fn=_chia_giu_khuc,
                   phim_fn=_phim_gia, ke_hoach_fn=_ke_hoach_gia)
    m = ra["scenes"]["json"]
    # Màn được canh lại phủ hết 10 dòng, không hở.
    assert [(s["srt_from"], s["srt_to"]) for s in m["story"]["segments"]] == [(1, 4), (5, 10)]
    assert m["story"]["arc"] == "man_in_hole"
    # Beat cuối bị kẹp vào cuối màn; beat dài quá trần 8 giây (mỗi dòng 3 giây)
    # được tách tại ranh giới dòng: 2-4 (9s) → 2-3, 4; 6-10 (15s) → 6-7, 8-9, 10.
    kh = m["director_plan"]
    assert [(b["segment_id"], b["beat"], b["srt_from"], b["srt_to"]) for b in kh] == [
        (1, 1, 1, 1), (1, 2, 2, 3), (1, 3, 4, 4),
        (2, 1, 5, 5), (2, 2, 6, 7), (2, 3, 8, 9), (2, 4, 10, 10)]
    st = _sheet(yeu_cau, "story")
    assert st[0][:2] == ["segment_id", "name"] and st[1][1] == "Mo dau"
    pl = _sheet(yeu_cau, "director_plan")
    assert pl[0][0] == "segment_id" and len(pl) == 8
    # Nhân vật AI dựng và bối cảnh có prompt ảnh tham chiếu; nv cố định thì không.
    nv = {c["id"]: c for c in m["characters"]}
    assert "reference portrait" in nv["nv2"]["sheet_prompt"]
    assert "establishing wide shot" in m["locations"][0]["sheet_prompt"]


def test_dao_dien_loai_2_nv1_khong_co_sheet_prompt(wb, yeu_cau):
    yeu_cau["config"]["che_do_ke"] = "nhan_vat_va_boi_canh"
    ra = wb.handle(yeu_cau, cast_fn=_cast_phu_va_boi_canh, chia_fn=_chia_giu_khuc,
                   phim_fn=_phim_gia, ke_hoach_fn=_ke_hoach_gia)
    nv = {c["id"]: c for c in ra["scenes"]["json"]["characters"]}
    assert "sheet_prompt" not in nv["nv1"] and "sheet_prompt" in nv["nv2"]


def test_loai_1_khong_dao_dien(wb, yeu_cau):
    goi = {"n": 0}

    def phim_fn(*_a):
        goi["n"] += 1
        return _phim_gia(None, None)

    yeu_cau["config"]["che_do_ke"] = "mot_nhan_vat"
    ra = wb.handle(yeu_cau, chia_fn=_chia_giu_khuc, phim_fn=phim_fn)
    assert goi["n"] == 0
    assert ra["scenes"]["json"]["story"] == {} and ra["scenes"]["json"]["director_plan"] == []


def test_bom_tay_khong_phim_fn_thi_bo_qua_dao_dien(wb, yeu_cau):
    yeu_cau["config"]["che_do_ke"] = "tu_xay"
    ra = wb.handle(yeu_cau, cast_fn=_cast_phu_va_boi_canh, chia_fn=_chia_giu_khuc)
    assert ra["scenes"]["json"]["story"] == {}


def test_khoi_ke_hoach_chi_lay_beat_cham_khuc(wb):
    kh = [{"segment_id": 1, "beat": 1, "srt_from": 1, "srt_to": 3, "purpose": "hook",
           "characters": "nv1", "location": "loc1", "shot_size": "WIDE",
           "camera": "push_in", "element_motion": "x", "emotion": "y", "motif": "m"},
          {"segment_id": 1, "beat": 2, "srt_from": 4, "srt_to": 9, "purpose": "turn",
           "characters": "", "location": "", "shot_size": "CLOSE", "camera": "static",
           "element_motion": "z", "emotion": "w", "motif": ""}]
    khuc = [{"index": i} for i in range(4, 7)]
    khoi = wb._khoi_ke_hoach(kh, khuc)
    assert "DIRECTOR'S PLAN" in khoi and "lines 4-9" in khoi and "lines 1-3" not in khoi
    assert wb._khoi_ke_hoach([], khuc) == ""


def test_reference_files_co_ca_boi_canh(wb):
    scenes = [{"characters_used": "nv1 nv2", "location_used": "loc1"},
              {"characters_used": "", "location_used": "loc9"}]
    wb._gan_reference_files(scenes, [{"id": "nv1"}, {"id": "nv2"}], [{"id": "loc1"}])
    import json
    assert json.loads(scenes[0]["reference_files"]) == ["nv1.png", "nv2.png", "loc1.png"]
    assert not scenes[1].get("reference_files")


def test_ke_hoach_gop_beat_ngan(wb):
    seg = {"segment_id": 1, "name": "x", "message": "", "emotion": "", "motif": "",
           "srt_from": 1, "srt_to": 3}
    dong = [{"index": 1, "start": 0.0, "end": 3.0, "text": "a"},
            {"index": 2, "start": 3.0, "end": 3.6, "text": "b"},
            {"index": 3, "start": 3.6, "end": 7.0, "text": "c"}]
    raw = {"beats": [{"srt_from": 1, "srt_to": 1, "purpose": "p1"},
                     {"srt_from": 2, "srt_to": 2, "purpose": "p2"},
                     {"srt_from": 3, "srt_to": 3, "purpose": "p3"}]}
    ra = wb._sach_ke_hoach(raw, seg, dong)
    assert [(b["beat"], b["srt_from"], b["srt_to"]) for b in ra] == [(1, 1, 2), (2, 3, 3)]


# ── Casting phải phủ hết danh sách bước đọc phim đã nhận ra ─────────────────

def test_casting_bo_sot_thi_goi_bo_sung(wb, yeu_cau):
    goi = {"n": 0}

    def cast_fn(_cues, ctx):
        goi["n"] += 1
        if "bo_sung" in ctx:
            # Lượt bổ sung: chỉ xin những mục thiếu.
            assert ctx["bo_sung"]["characters"] == ["the ogre / giant"]
            assert ctx["bo_sung"]["locations"] == ["the king's palace"]
            return {"characters": [{"id": "nv3", "role": "the ogre", "english_prompt": "a huge ogre"}],
                    "locations": [{"id": "loc3", "name": "King's Palace", "english_prompt": "marble hall"}]}
        return {"style": {}, "characters": [
                    {"id": "nv1", "role": "the youngest son", "english_prompt": "a young man"},
                    {"id": "nv2", "role": "the cat", "english_prompt": "a cat in boots"}],
                "locations": [{"id": "loc1", "name": "The Miller's Cottage", "english_prompt": "cottage"},
                              {"id": "loc2", "name": "The River", "english_prompt": "river"}]}

    def phim_fn(*_a):
        return {"genre": "fable", "arc": "hook_payoff", "context_lock": "x",
                "segments": [{"segment_id": 1, "name": "a", "message": "", "emotion": "",
                              "motif": "", "srt_from": 1, "srt_to": 99}],
                "characters_mentioned": ["youngest son", "the cat", "the ogre / giant"],
                "locations_mentioned": ["the miller's cottage", "the river", "the king's palace"]}

    yeu_cau["config"]["che_do_ke"] = "tu_xay"
    ra = wb.handle(yeu_cau, cast_fn=cast_fn, chia_fn=_chia_giu_khuc, phim_fn=phim_fn,
                   ke_hoach_fn=_ke_hoach_gia)
    m = ra["scenes"]["json"]
    assert goi["n"] == 2
    assert [c["id"] for c in m["characters"]] == ["nv1", "nv2", "nv3"]
    assert [l["id"] for l in m["locations"]] == ["loc1", "loc2", "loc3"]


def test_casting_du_thi_khong_goi_them(wb):
    cast = {"characters": [{"id": "nv1", "role": "the cat Puss", "name": "", "english_prompt": ""}],
            "locations": [{"id": "loc1", "name": "The Ogre's Castle", "english_prompt": ""}]}
    ctx = {"film_analysis": {"characters_mentioned": ["the cat (Puss in Boots)"],
                             "locations_mentioned": ["the ogre's castle"]}}
    assert wb._con_thieu(cast, ctx) == ([], [])
    ctx2 = {"film_analysis": {"characters_mentioned": ["the two older brothers"],
                              "locations_mentioned": ["the king's palace"]}}
    assert wb._con_thieu(cast, ctx2) == (["the two older brothers"], ["the king's palace"])


# ── Không ghi cứng theo độ dài kịch bản ─────────────────────────────────────

def test_so_man_theo_do_dai_khong_tran(wb):
    assert wb._so_man_goi_y(0) == 3
    assert wb._so_man_goi_y(180) == 3
    assert wb._so_man_goi_y(480) == 6
    assert wb._so_man_goi_y(1800) == 24


def test_loi_doc_cho_bia_nhac_casting_la_ca_bai(wb):
    cues = [{"index": i, "start": i * 3.0, "end": i * 3.0 + 3.0, "text": "cau {0} ".format(i) + "x" * 300}
            for i in range(1, 80)]
    ca_bai = " ".join(c["text"].strip() for c in cues)
    assert wb._loi_doc_mau(cues) == ca_bai
    assert wb.BIA_KY_TU == 0 and not hasattr(wb, "CAST_KY_TU")
    assert wb.TOKEN_CAST >= wb.TOKEN_CANH


# ── Từ bị bộ lọc an toàn chặn vô cớ ─────────────────────────────────────────

def test_thay_tu_bi_loc_trong_moi_prompt(wb):
    assert wb._lam_lanh_prompt("small anthropomorphic cat with a sly smile") == \
        "small upright humanlike cat with a knowing smile"
    assert wb._lam_lanh_prompt("Slyvia walks") == "Slyvia walks"   # không cắt giữa từ
    cast = {"characters": [{"id": "nv4", "english_prompt": "anthropomorphic cat", "sheet_prompt": "sly cat"}],
            "locations": []}
    scenes = [{"img_prompt": "an anthropomorphic fox", "video_prompt": "sly grin"}]
    bia = {"thumbnails": [{"img_prompt": "sly cat"}]}
    wb._lam_lanh_moi_prompt(cast, scenes, bia)
    assert "anthropomorphic" not in cast["characters"][0]["english_prompt"]
    assert scenes[0]["img_prompt"] == "an upright humanlike fox"
    assert bia["thumbnails"][0]["img_prompt"] == "knowing cat"


def test_excel_khong_con_tu_bi_loc(wb, yeu_cau):
    def cast_fn(_c, _x):
        return {"style": {}, "characters": [{"id": "nv1", "role": "cat", "english_prompt": "anthropomorphic cat, sly"}],
                "locations": []}

    yeu_cau["config"]["che_do_ke"] = "tu_xay"
    ra = wb.handle(yeu_cau, cast_fn=cast_fn, chia_fn=_chia_giu_khuc)
    m = ra["scenes"]["json"]
    assert "anthropomorphic" not in m["characters"][0]["english_prompt"]
    assert all("anthropomorphic" not in s["img_prompt"] and " sly" not in s["img_prompt"] for s in m["scenes"])


# ── Khoá nhận dạng: mô hình vẽ phải biết ảnh tham chiếu nào là ai ───────────

def test_khoa_nhan_dang_noi_ten_voi_anh_tham_chieu(wb):
    dan = [{"id": "nv4", "role": "the cat", "english_prompt": "small cat in a plumed hat and boots"}]
    noi = [{"id": "loc9", "name": "The ogre's castle", "english_prompt": "stone castle"}]
    sc = [{"img_prompt": "Wide shot of nv4 (nv4) facing the ogre inside loc9, pencil sketch",
           "video_prompt": "The cat steps forward.",
           "reference_files": json.dumps(["nv4.png", "loc9.png"])},
          {"img_prompt": "Empty road", "video_prompt": "wind", "reference_files": ""}]
    wb._khoa_nhan_dang(sc, dan, noi)
    p = sc[0]["img_prompt"]
    # ═══ LỐI OPENSTORY: RÀNG BUỘC TRONG CÂU, CHÚ THÍCH MỘT DÒNG ═══
    #
    # Bản cũ dán một đoạn văn ~1.100 ký tự liệt kê mặt, mắt, tỉ lệ, da, lông,
    # trang phục phải chép y hệt. Đo 27/08/2026 trên đúng những cảnh máy vẽ
    # nhầm người (openstory/0002, cảnh 25): đoạn văn dài 3,00 điểm; một dòng
    # của OpenStory 4,00. Liệt kê từng nét hoá ra lại là chữ đi cãi với ảnh.
    assert "nv4 (Image 1)" in p and "loc9 (Image 2)" in p
    # Ảnh nào câu văn đã ràng buộc thì KHÔNG lặp lại ở chú thích dưới.
    assert "= nv4" not in p and "= loc9" not in p
    # Vẫn tuyệt đối không tả ngoại hình.
    assert "plumed hat" not in p and "small cat in a" not in p and "stone castle" not in p
    assert "Reference images define identity" in p
    assert sc[0]["video_prompt"].startswith("IDENTITY LOCK")
    # KHÔNG tả lại nhân vật trong khoá video — chữ và ảnh cãi nhau thì mô hình bỏ ảnh.
    assert "plumed hat" not in sc[0]["video_prompt"] and "nothing is removed" in sc[0]["video_prompt"]
    assert sc[0]["video_prompt"].endswith("The cat steps forward.")
    # Cảnh không có tham chiếu thì để nguyên.
    assert sc[1]["img_prompt"] == "Empty road" and sc[1]["video_prompt"] == "wind"
    # Gọi lần hai không chồng khối.
    wb._khoa_nhan_dang(sc, dan, noi)
    assert sc[0]["img_prompt"].count("REFERENCE IMAGES are attached") == 1


def test_excel_moi_canh_co_khoi_khoa(wb, yeu_cau):
    yeu_cau["config"]["che_do_ke"] = "tu_xay"
    ra = wb.handle(yeu_cau, cast_fn=_cast_phu_va_boi_canh, chia_fn=_chia_giu_khuc)
    m = ra["scenes"]["json"]
    co_ref = [s for s in m["scenes"] if s.get("reference_files")]
    assert co_ref and all("Reference images define identity" in s["img_prompt"]
                          for s in co_ref)
    assert all("(Image 1)" in s["img_prompt"] or "Image 1 =" in s["img_prompt"]
               for s in co_ref)


def test_khoa_cho_phep_doi_trang_phuc_khi_canh_noi_ro(wb):
    from core.chia_canh import KHUON_MAC_DINH
    # Cảnh KHÔNG tả lại nhân vật — chỉ id + tư thế; giai đoạn có id riêng.
    assert "ONLY by id" in KHUON_MAC_DINH and "one id per stage" in KHUON_MAC_DINH
    # 27/08/2026: câu "được đổi trang phục nếu cảnh nói rõ" nằm trong đoạn văn
    # khoá cũ, đã bỏ cùng đoạn ấy (xem `_GUARD_NHAN_DANG`). Đổi trang phục giờ
    # đi bằng ĐÚNG một đường: mỗi giai đoạn một id, một ảnh tham chiếu riêng —
    # đường ấy chắc hơn một câu dặn, và có bài kiểm riêng bên dưới.
    assert not hasattr(wb, "_KHOA_NHAN_VAT"), "đừng dựng lại đoạn văn khoá cũ"


# ── Nhân vật đổi trang phục giữa truyện: mỗi giai đoạn một tham chiếu ───────

def _cast_meo_hai_giai_doan(_c, _x):
    return {"style": {}, "characters": [
        {"id": "nv1", "role": "hero", "english_prompt": "a slim young man"},
        {"id": "nv2", "role": "the cat", "english_prompt": "a chubby orange cat, big green eyes",
         "stages": [{"when": "at the start", "outfit": "no clothes"},
                    {"when": "after the hero gives it a hat and boots", "outfit": "red beret, green vest, small brown boots"}]}],
        "locations": []}


def test_tach_giai_doan_thanh_id_rieng_cung_mat(wb):
    cast = wb._sach_cast(_cast_meo_hai_giai_doan(None, None))
    ids = [c["id"] for c in cast["characters"]]
    assert ids == ["nv1", "nv2", "nv2b"]
    a, b = cast["characters"][1], cast["characters"][2]
    assert a["english_prompt"].startswith("a chubby orange cat, big green eyes") and "no clothes" in a["english_prompt"]
    assert "red beret" in b["english_prompt"] and b["goc_id"] == "nv2" and b["giai_doan"] == 2
    assert "Stage 2/2 of nv2" in b["notes"]
    khoi = wb._khoi_cast_style(cast)
    assert "one id PER STAGE" in khoi and "nv2b (the cat, STAGE 2: after the hero gives it a hat and boots)" in khoi


def test_canh_dung_id_giai_doan_thi_tham_chieu_va_khoa_theo_giai_doan(wb, yeu_cau):
    def chia(khuc, _t, _n):
        c = canh_ai(khuc[0]["index"], khuc[-1]["index"])
        c["characters_used"] = "nv2b"
        c["img_prompt"] = "Wide shot of nv2b (nv2b) walking, pencil"
        return [c]

    yeu_cau["config"]["che_do_ke"] = "tu_xay"
    ra = wb.handle(yeu_cau, cast_fn=_cast_meo_hai_giai_doan, chia_fn=chia)
    m = ra["scenes"]["json"]
    assert [c["id"] for c in m["characters"]] == ["nv1", "nv2", "nv2b"]
    assert all(c.get("sheet_prompt") for c in m["characters"]), "mỗi giai đoạn phải có prompt ảnh tham chiếu riêng"
    s = m["scenes"][0]
    assert json.loads(s["reference_files"]) == ["nv2b.png"]
    # Trang phục của giai đoạn nằm trong ẢNH tham chiếu riêng của nó (nv2b.png),
    # không nằm trong chữ — xem `_mo_ta_tham_chieu`.
    assert "nv2b (Image 1)" in s["img_prompt"] and "red beret" not in s["img_prompt"]
    assert "red beret" in [c for c in m["characters"] if c["id"] == "nv2b"][0]["sheet_prompt"]


# ── Kế hoạch đạo diễn có thẩm quyền: cảnh phải dùng đúng bối cảnh + nhân vật ─

def test_ep_canh_theo_ke_hoach_boi_canh_va_nhan_vat(wb):
    ke_hoach = [{"srt_from": 1, "srt_to": 3, "location": "loc1", "characters": "nv1, nv2"},
                {"srt_from": 4, "srt_to": 6, "location": "loc2", "characters": "nv2"}]
    scenes = [{"srt_indices": [1, 2], "location_used": "loc9", "characters_used": "nv1"},
              {"srt_indices": [3], "location_used": "loc1", "characters_used": "nv2, nv1"},
              {"srt_indices": [4, 5, 6], "location_used": "", "characters_used": ""},
              {"srt_indices": [], "location_used": "loc7", "characters_used": ""}]
    doi = wb._ep_theo_ke_hoach(scenes, ke_hoach)
    assert [s["location_used"] for s in scenes] == ["loc1", "loc1", "loc2", "loc7"]
    assert scenes[0]["characters_used"] == "nv1, nv2" and scenes[1]["characters_used"] == "nv2, nv1"
    assert scenes[2]["characters_used"] == "nv2"
    assert doi == {"boi_canh": 2, "nhan_vat": 2}
    assert wb._so_lan_doi_boi_canh(scenes) == 2


def test_ke_hoach_khong_con_luat_doi_cho_cho_da_dang(wb):
    assert "do not stage two consecutive beats in the same place" not in wb._KHUON_KE_HOACH
    assert "FOLLOWS THE STORY" in wb._KHUON_KE_HOACH
    khoi = wb._khoi_cast_style({"characters": [], "locations": [{"id": "loc1", "name": "a", "english_prompt": "b"}]})
    assert "stay in the same place until the narration says" in khoi


def test_ep_ke_hoach_HOP_nhan_vat_khong_thay(wb):
    # Cảnh 14 (25/08): beat ghi thiếu mèo, bản cũ ghi đè → mất tham chiếu mèo.
    ke_hoach = [{"srt_from": 1, "srt_to": 2, "location": "loc1", "characters": "nv1"}]
    scenes = [{"srt_indices": [1], "location_used": "loc1", "characters_used": "nv1, nv4"},
              {"srt_indices": [2], "location_used": "loc1", "characters_used": "nv4b"}]
    wb._ep_theo_ke_hoach(scenes, ke_hoach)
    assert scenes[0]["characters_used"] == "nv1, nv4"
    assert scenes[1]["characters_used"] == "nv1, nv4b"
    # Beat ghi nv4b (giai đoạn sau) mà cảnh khai nv4 → giữ id của beat, không giữ cả hai.
    ke_hoach = [{"srt_from": 1, "srt_to": 1, "location": "", "characters": "nv4b"}]
    scenes = [{"srt_indices": [1], "location_used": "", "characters_used": "nv4, nv1"}]
    wb._ep_theo_ke_hoach(scenes, ke_hoach)
    assert scenes[0]["characters_used"] == "nv4b, nv1"


def test_khoa_boi_canh_dung_tren_dat_va_anh_thiet_lap_ngang_tam_mat(wb):
    # Luật vật lý tách khỏi khối khoá cũ và giữ nguyên — nó không nói gì về
    # danh tính nên không dính dáng tới `_GUARD_NHAN_DANG`.
    assert "never standing on water" in wb._LUAT_DUNG_CHAN and "correct scale" in wb._LUAT_DUNG_CHAN
    assert "human eye level" in wb._DUOI_BOI_CANH and "lower third" in wb._DUOI_BOI_CANH


# ── Thiết kế nhân vật ở MỘT chỗ: đổi thiết kế là mọi cảnh đổi theo ──────────

def test_doi_thiet_ke_nhan_vat_lan_ra_moi_canh_va_moi_giai_doan(wb):
    dan = [{"id": "nv4", "role": "the cat", "english_prompt": "orange cat; outfit at this stage: no clothes",
            "sheet_prompt": "x"},
           {"id": "nv4b", "role": "the cat", "english_prompt": "orange cat; outfit at this stage: red beret, boots",
            "sheet_prompt": "x"},
           {"id": "nv1", "role": "hero", "english_prompt": "a young man", "sheet_prompt": "y"}]
    sc = [{"img_prompt": "Wide shot of nv4b [reference image 1] waving\nREFERENCE IMAGES are attached, in this order:\n- reference image 1 = nv4b, the the cat: orange cat; outfit at this stage: red beret, boots\n- reference image 2 = nv1, the hero: a young man"},
          {"img_prompt": "nv1 alone\nREFERENCE IMAGES are attached, in this order:\n- reference image 1 = nv1, the hero: a young man"}]
    n = wb.doi_thiet_ke_nhan_vat(sc, dan, "nv4b", "grey cat with a white chest", duoi_style=" Style: 3D")
    assert n == 1
    assert dan[0]["english_prompt"] == "grey cat with a white chest; outfit at this stage: no clothes"
    assert dan[1]["english_prompt"] == "grey cat with a white chest; outfit at this stage: red beret, boots"
    assert dan[1]["sheet_prompt"].startswith("grey cat with a white chest; outfit at this stage: red beret, boots — full-body")
    assert dan[1]["sheet_prompt"].endswith(" Style: 3D")
    assert "reference image 1 = nv4b, the the cat: grey cat with a white chest; outfit at this stage: red beret, boots" in sc[0]["img_prompt"]
    assert dan[2]["english_prompt"] == "a young man" and "a young man" in sc[1]["img_prompt"]
    assert wb.doi_thiet_ke_nhan_vat(sc, dan, "nv9", "x") == 0


def test_loi_nhac_thiet_ke_lai_co_vai_mo_ta_ly_do(wb):
    p = wb.loi_nhac_thiet_ke_lai({"role": "helper", "english_prompt": "a cat in a feathered hat"}, "content_rejected")
    assert "Role in the story: helper" in p and "a cat in a feathered hat" in p and "Rejection reason: content_rejected" in p
    assert '"english_prompt"' in p


def test_khoi_dan_bao_ai_chi_goi_bang_id(wb):
    khoi = wb._khoi_cast_style({"characters": [{"id": "nv4", "role": "cat", "english_prompt": "x"}], "locations": []})
    assert "call it ONLY by its id" in khoi and "NEVER describe" in khoi


def test_casting_yeu_cau_tu_ngu_than_thien_gia_dinh(wb):
    assert "FAMILY-FRIENDLY" in wb._KHUON_CAST and "no weapons" in wb._KHUON_CAST


def test_khuon_chia_cua_kenh_hay_mac_dinh(wb):
    from core.chia_canh import KHUON_MAC_DINH
    run = wb
    du = "KENH " + " ".join(run._CHO_TRONG_KHUON_CHIA)
    assert run._khuon_chia({"storyboard_template": du}) == du
    assert run._khuon_chia({"storyboard_template": "thieu <<SRT>>"}) == KHUON_MAC_DINH
    assert run._khuon_chia({}) == KHUON_MAC_DINH
    assert run._khuon_chia(None) == KHUON_MAC_DINH


def test_pha_lap_canh_doi_khung_va_giu_id(wb):
    run = wb
    tail = ", stylised 3D animated film still, 16:9, no text"
    scenes = [
        {"scene_id": 1, "img_prompt": "Medium shot of nv1 (nv1) on the doorstep of loc2 looking sad" + tail, "video_prompt": "a"},
        {"scene_id": 2, "img_prompt": "Medium shot of nv1 (nv1) on the doorstep of loc2 looking sadder" + tail, "video_prompt": "b"},
        {"scene_id": 3, "img_prompt": "Close-up of nv1 (nv1) on the doorstep of loc2 sighing" + tail, "video_prompt": "c"},
        {"scene_id": 4, "img_prompt": "Wide shot of loc5 the castle gate at dawn" + tail, "video_prompt": "d"},
    ]
    assert run._canh_lap(scenes) == [1]            # 3 đổi cỡ khung → không lặp; 4 khác hẳn
    goi_lan = []

    def goi(loi_nhac, phan_khoa):
        goi_lan.append(loi_nhac)
        assert "previous opening: medium shot" in loi_nhac and "[2]" in loi_nhac
        return '{"2": {"img_prompt": "Over-the-shoulder shot from behind nv1 (nv1) toward the doorstep of loc2, his hand on the step%s", "video_prompt": "hand lifts"}}' % tail

    assert run._pha_lap_canh(scenes, goi) == 1
    assert scenes[1]["img_prompt"].startswith("Over-the-shoulder") and scenes[1]["video_prompt"] == "hand lifts"
    assert run._canh_lap(scenes) == []
    # Bản viết lại đổi id nhân vật → bị bỏ.
    scenes2 = [dict(scenes[0]), {"scene_id": 2, "img_prompt": scenes[0]["img_prompt"], "video_prompt": "x"}]
    assert run._pha_lap_canh(scenes2, lambda l, k: '{"2": {"img_prompt": "Close-up of nv9 (nv9) at loc2 smiling%s"}}' % tail) == 0
    assert scenes2[1]["img_prompt"] == scenes[0]["img_prompt"]


def test_don_mo_ta_bo_lenh_va_phu_dinh(wb):
    run = wb
    chu = ("a small upright kitten with solid plain golden-yellow fur without any stripes or patches, "
           "bare soft paws with no footwear; describe only fur, face and body here; outfit at this stage: "
           "no clothes at all, no hat, bare paws, plain golden-yellow fur only")
    ra = run._don_mo_ta(chu)
    assert "describe only" not in ra and "no footwear" not in ra and "no hat" not in ra and "no clothes" not in ra
    assert "without any stripes" in ra                     # hoa văn lông thì giữ
    assert "bare soft paws" in ra and "outfit at this stage:" in ra and "plain golden-yellow fur only" in ra
    assert run._don_mo_ta("a jolly king with a golden crown and a red robe") == "a jolly king with a golden crown and a red robe"
    dan = [{"id": "nv1", "english_prompt": "a cat, no hat", "stages": [{"when": "x", "outfit": "a vest, no boots"}]}]
    assert run._don_dan(dan) == 1
    assert dan[0]["english_prompt"] == "a cat" and dan[0]["stages"][0]["outfit"] == "a vest"


def test_boi_canh_chu_khong_cat_truyen_dai_va_bo_khuon(wb):
    run = wb
    ctx = {"script": "x" * 50000, "storyboard_template": "KHUON " * 2000, "style": {"a": 1}}
    chu = run._boi_canh_chu(ctx)
    assert "KHUON" not in chu and "x" * 50000 in chu and '"style"' in chu
    assert run._boi_canh_chu({}, "(none)") == "(none)" and run._boi_canh_chu(None, "") == ""
    assert len(run._boi_canh_chu({"s": "y" * 200000})) == run.TRAN_CONTEXT_CHU


def test_ap_same_as_gop_ten_khong_them_id(wb):
    run = wb
    cast = {"characters": [{"id": "nv1", "name": "Cat", "role": "hero", "english_prompt": "a cat"}],
            "locations": [{"id": "loc9", "name": "The mill", "english_prompt": "a watermill"}]}
    n = run._ap_same_as(cast, {"Cối xay bột": "loc9", "Con mèo": "nv1", "không có": "loc99"})
    assert n == 2
    assert cast["locations"][0]["name"] == "The mill (Cối xay bột)" and len(cast["locations"]) == 1
    assert cast["characters"][0]["name"] == "Cat (Con mèo)"
    # Sau khi gộp, kiểm phủ theo từ khoá thấy "Cối xay bột" đã có → không còn thiếu.
    assert run._duoc_phu("Cối xay bột", cast["locations"], "name", "english_prompt")
    assert run._ap_same_as(cast, None) == 0


def test_giai_doan_bien_hinh_la_than_the_moi(wb):
    run = wb
    goc = {"id": "nv15", "name": "Giant", "role": "villain", "english_prompt": "a very big rounded giant with a bushy beard", "notes": ""}
    ra = run._tach_giai_doan(goc, [
        {"when": "start", "outfit": "mossy-green tunic, big brown boots"},
        {"when": "shows off", "outfit": "a large rounded lion, tawny fur, fluffy mane, bare paws"},
        {"when": "tricked", "outfit": "the whole body becomes a tiny grey mouse with a pink tail"},
    ])
    assert [c["id"] for c in ra] == ["nv15", "nv15b", "nv15c"]
    assert ra[0]["english_prompt"].startswith("a very big rounded giant") and "outfit at this stage: mossy-green" in ra[0]["english_prompt"]
    assert ra[1]["english_prompt"].startswith(run.DAU_BIEN_HINH + "a large rounded lion") and "giant with a bushy beard" not in ra[1]["english_prompt"]
    assert ra[2]["english_prompt"].startswith(run.DAU_BIEN_HINH) and "transformed form of Giant" in ra[2]["english_prompt"]
    # Đổi thiết kế thân gốc: giai đoạn biến hình giữ nguyên.
    from core.prompt_visuals import doi_thiet_ke_nhan_vat
    doi_thiet_ke_nhan_vat([], ra, "nv15", "a huge round giant with a red beard")
    assert ra[0]["english_prompt"].startswith("a huge round giant") and "outfit at this stage: mossy-green" in ra[0]["english_prompt"]
    assert ra[1]["english_prompt"].startswith(run.DAU_BIEN_HINH + "a large rounded lion")


# ── Id KHÔNG có ảnh kèm: thay bằng vai, không để trần ───────────────────────

def test_id_khong_duoc_gui_anh_thi_thay_bang_vai(wb):
    """Trần 2 nhân vật/cảnh, nhưng nhân vật thứ ba vẫn bị viết vào lời nhắc.

    `nv4` là thứ chỉ TOOL hiểu — máy vẽ nhận một dãy ảnh theo thứ tự, không
    nhận tên tệp. Để `nv4` trần là đưa cho nó một chuỗi vô nghĩa và nó bịa ra
    một nhân vật mới, mỗi cảnh một kiểu. Đo trên lượt thật openstory/0002
    (26/08/2026): 12/30 cảnh dính.
    """
    dan = [{"id": "nv1", "role": "boy", "english_prompt": "a small boy"},
           {"id": "nv2", "role": "grandmother", "english_prompt": "an old woman"},
           {"id": "nv5", "role": "duck", "english_prompt": "a plump duck"}]
    noi = [{"id": "loc1", "name": "The hut"}, {"id": "loc7", "name": "The pond"}]
    sc = [{"img_prompt": "Wide shot of nv1 and nv2 inside loc1 while nv5 swims past loc7",
           "video_prompt": "nv5 paddles toward nv1.",
           "reference_files": json.dumps(["nv1.png", "nv2.png", "loc1.png"])}]
    wb._khoa_nhan_dang(sc, dan, noi)
    p = sc[0]["img_prompt"]
    # Hai nhân vật + một nơi CÓ ảnh: giữ id, gắn số ảnh.
    assert "nv1 (Image 1)" in p and "nv2 (Image 2)" in p
    assert "loc1 (Image 3)" in p
    # Nhân vật và nơi KHÔNG có ảnh: id biến mất, còn lại vai / tên nơi.
    than = p.split("REFERENCE IMAGES")[0]
    # Nhãn lấy nguyên cụm nhận dạng trong dàn — xem
    # `test_nhan_vat_khong_co_anh_mang_NHAN_CO_DINH` để biết vì sao không phải
    # chỉ cái vai.
    assert "nv5" not in than and "a plump duck" in than
    assert "loc7" not in than and "The pond" in than
    # Lời nhắc clip cũng vậy.
    assert "nv5" not in sc[0]["video_prompt"] and "a plump duck" in sc[0]["video_prompt"]
    assert "nv1" in sc[0]["video_prompt"]  # nhân vật CÓ ảnh giữ nguyên id


def test_nhan_vat_khong_co_anh_mang_NHAN_CO_DINH(wb):
    """Nhãn phải kèm cụm nhận dạng, và giống hệt nhau ở mọi cảnh.

    Bản đầu chỉ cho cái vai ("the duck"). Đo trên lượt thật openstory/0002
    (26/08/2026, 30 cảnh): con vịt hiện ra HAI kiểu trong cùng một phim — vịt
    cổ xanh nâu ở cảnh 11-15, vịt trắng ở cảnh 28-30, trong khi dàn ghi rõ
    "creamy-white". Vai suông nói cho máy biết vẽ CON GÌ, không nói CON NÀO.

    Không mâu thuẫn luật 26/08 ("đừng tả chi tiết, dùng ảnh"): luật ấy nói về
    nhân vật CÓ ảnh — chữ và ảnh cãi nhau thì máy nghe chữ. Ở đây không có ảnh
    nào để cãi; bỏ chữ đi là không còn gì.
    """
    vit = "a friendly plump duck, soft creamy-white feathers and light orange bill"
    dan = [{"id": "nv1", "role": "boy", "english_prompt": "a small boy"},
           {"id": "nv5", "role": "helper animal", "english_prompt": vit}]
    sc = [{"img_prompt": "Close-up of nv5 swimming while nv1 watches",
           "video_prompt": "nv5 paddles.",
           "reference_files": json.dumps(["nv1.png"])},
          {"img_prompt": "Wide shot of nv5 on the bank near nv1",
           "video_prompt": "nv5 waddles.",
           "reference_files": json.dumps(["nv1.png"])}]
    wb._khoa_nhan_dang(sc, dan, [])
    for s in sc:
        than = s["img_prompt"].split("REFERENCE IMAGES")[0]
        assert "nv5" not in than
        assert "creamy-white" in than, "mất màu là mất thứ giữ con vật khỏi đổi"
    # GIỐNG HỆT NHAU giữa hai cảnh — đó là thứ thay cho ảnh tham chiếu.
    lay = lambda t: t.split("nv1")[0]  # noqa: E731
    assert "creamy-white" in sc[0]["video_prompt"]
    assert wb._nhan_co_dinh(dan[1]) == wb._nhan_co_dinh(dan[1])


def test_nhan_co_dinh_khong_cut_giua_tu(wb):
    """Nhãn cụt ("soft dark grey strip") là nhãn vô nghĩa."""
    dai = ("a friendly ordinary house cat, grey-and-brown mackerel tabby fur "
           "with soft dark grey stripes along the back and a pale belly")
    nhan = wb._nhan_co_dinh({"english_prompt": dai})
    assert len(nhan) <= wb.DAI_NHAN_CO_DINH
    assert dai.startswith(nhan)
    assert nhan.split()[-1].lower() not in wb._TU_NOI_CUT
    assert "grey-and-brown" in nhan, "phải giữ được màu"


def test_nhan_co_dinh_thieu_mo_ta_thi_lui_ve_vai(wb):
    assert wb._nhan_co_dinh({"role": "healer"}) == "the healer"


def test_id_la_khong_co_trong_dan_thi_de_nguyen(wb):
    """Id AI bịa ra thì giữ nguyên — che đi là giấu mất một lỗi khác."""
    dan = [{"id": "nv1", "role": "boy", "english_prompt": "a small boy"}]
    sc = [{"img_prompt": "Wide shot of nv1 and nv77 by the river",
           "video_prompt": "they walk.",
           "reference_files": json.dumps(["nv1.png"])}]
    wb._khoa_nhan_dang(sc, dan, [])
    assert "nv77" in sc[0]["img_prompt"]


def test_khoi_khoa_khong_goi_ten_net_ve_nao(wb):
    """Khối khoá dán vào MỌI kênh — gọi tên một nét vẽ là phá kênh khác.

    Đo trên lượt thật openstory/0002 (26/08/2026, hoạt hình 3D): cả 30 lời
    nhắc ảnh đều kết bằng *"the same pencil-sketch line style"* — chép từ kênh
    bút chì giấy trắng. Lời nhắc tả phim 3D kiểu Pixar, câu cuối bảo nhân vật
    có nét vẽ chì; máy phải chọn một, và nó vẽ lại nhân vật cho khớp.
    """
    for net in ("pencil", "sketch", "watercolor", "cel-shaded", "3D", "Pixar"):
        assert net.lower() not in wb._GUARD_NHAN_DANG.lower(), net
    # Một dòng, không phải một đoạn văn: đó là cả điểm của bản 27/08.
    assert len(wb._GUARD_NHAN_DANG) < 250
    assert "consistent with its reference image" in wb._GUARD_NHAN_DANG


def test_chan_dung_bo_qua_tu_the_trong_mo_ta():
    """Ảnh gốc phải là NGƯỜI ĐỨNG, dù truyện đang bảo nhân vật nằm ngủ.

    Đo 27/08/2026 (phim hoathinh-3d/0002): AI ghi tư thế vào ô "trang phục" của
    giai đoạn — "lying down on its side as an ordinary sleeping wolf" — nên ảnh
    tham chiếu của sói giai đoạn 3 là một con sói ĐANG NGỦ. Mọi cảnh vẽ theo ảnh
    đó đều lệch, và nó cũng không còn khớp hai giai đoạn kia.
    """
    from core.prompt_visuals import DUOI_CHAN_DUNG
    d = DUOI_CHAN_DUNG.lower()
    assert "ignore any action, pose" in d
    assert "asleep" in d and "lying down" in d
    assert "awake, standing upright" in d
