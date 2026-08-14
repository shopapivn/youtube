"""Nạp sẵn một khâu bằng file của bạn, thay vì để tool tự làm khâu đó.

═══ VÌ SAO CẦN THỨ NÀY ═══

Chủ dự án, 14/08/2026: *"ở tab auto thì có thêm các yếu tố can thiệp, ví dụ có
thể đẩy content sẵn có viết chỗ khác vào theo dạng txt, hay là chỗ excel có thể
tải file mẫu về điền và up lên"*.

Việc thật đằng sau câu đó: người làm YouTube **đã có sẵn kịch bản** — viết
trong Google Docs, thuê người viết, hoặc lấy từ video cũ. Bắt tool viết lại từ
đầu là vừa mất tiền vừa ra bản kém hơn bản họ đã ưng. Với bảng cảnh cũng vậy:
có người muốn tự tay quyết từng cảnh chứ không giao cho AI.

═══ MỘT CƠ CHẾ, KHÔNG PHẢI TÁM ═══

Mỗi khâu đã tự khai tên sản phẩm của nó trong `core/auto.KHAU`. Nên không cần
viết riêng "nạp kịch bản", "nạp Excel", "nạp giọng đọc" — chỉ cần **chép file
vào đúng chỗ rồi đánh dấu khâu ấy đã xong**. Tám khâu dùng chung một đường.

Nhờ vậy khách nạp được cả những thứ chưa ai nghĩ tới lúc viết dòng này: giọng
đọc thu bằng micro thật, phụ đề sửa tay, ảnh bìa tự thiết kế.

═══ VÌ SAO PHẢI KIỂM TRƯỚC KHI CHÉP ═══

Khâu sau tiêu tiền thật. Nạp nhầm một file Excel thiếu cột `srt_start` thì tool
không chết ngay — nó chạy tiếp, tạo 111 tấm ảnh, rồi mới hỏng ở khâu dựng. Lúc
đó tiền đã đi.

Nên mỗi loại file bị soi trước khi được nhận, và lời từ chối phải nói rõ
**thiếu đúng cái gì** — "file Excel thiếu cột srt_start" thì khách còn sửa
được, "file không hợp lệ" thì họ chỉ biết ngồi nhìn.
"""

from __future__ import annotations

import os
import shutil
from typing import List, Sequence, Tuple

from .auto import KHAU, san_pham_khau, ten_khau

__all__ = ["kieu_file_cua_khau", "kiem_file", "nap_file", "co_mau",
           "viet_mau", "LoiNapSan"]


class LoiNapSan(ValueError):
    """File khách đưa không dùng được — câu chữ trong này hiện thẳng lên màn hình."""


#: Đuôi file mỗi khâu nhận. Khâu đẻ ra thư mục (ảnh, clip) thì nhận cả thư mục.
_DUOI = {
    "kich-ban": (".txt",),
    "giong-doc": (".mp3", ".wav", ".m4a"),
    "phu-de": (".srt",),
    "bang-canh": (".xlsx",),
    "dung": (".mp4",),
}

#: Khâu đẻ ra cả một thư mục nhiều tệp — nạp bằng cách chọn thư mục.
_LA_THU_MUC = {"anh", "clip", "thumbnail"}


def kieu_file_cua_khau(ma: str) -> Tuple[str, Tuple[str, ...]]:
    """Trả `(mô tả cho hộp chọn file, các đuôi nhận)`. Đuôi rỗng = chọn thư mục."""
    if ma in _LA_THU_MUC:
        return ("Thư mục", ())
    duoi = _DUOI.get(ma, ())
    if not duoi:
        return ("Mọi loại file (*)", ())
    ten = {"kich-ban": "Kịch bản", "giong-doc": "Giọng đọc",
           "phu-de": "Phụ đề", "bang-canh": "Bảng cảnh",
           "dung": "Video"}.get(ma, ten_khau(ma))
    return ("{0} ({1})".format(ten, " ".join("*" + d for d in duoi)), duoi)


# ── Soi file trước khi nhận ──────────────────────────────────────────────────

def _kiem_kich_ban(duong: str) -> None:
    try:
        chu = open(duong, encoding="utf-8", errors="replace").read()
    except OSError as loi:
        raise LoiNapSan("Không đọc được file: {0}".format(loi)) from loi
    if len(chu.strip()) < 200:
        raise LoiNapSan(
            "File kịch bản quá ngắn ({0} chữ). Một video mười phút cần khoảng "
            "vài nghìn chữ — bạn kiểm lại xem có chọn nhầm file không."
            .format(len(chu.strip())))


def _kiem_phu_de(duong: str) -> None:
    try:
        chu = open(duong, encoding="utf-8", errors="replace").read()
    except OSError as loi:
        raise LoiNapSan("Không đọc được file: {0}".format(loi)) from loi
    if "-->" not in chu:
        raise LoiNapSan(
            "File này không giống phụ đề: không có dòng thời gian nào dạng "
            "“00:00:01,000 --> 00:00:04,000”.")


def _kiem_bang_canh(duong: str) -> None:
    from .auto_khau import COT_CANH

    try:
        from openpyxl import load_workbook
    except ImportError as loi:  # pragma: no cover — SETUP.bat cài sẵn
        raise LoiNapSan(
            "Máy chưa có bộ đọc Excel. Bạn nhấp đúp SETUP.bat một lần rồi "
            "thử lại.") from loi
    try:
        sach = load_workbook(duong, read_only=True)
    except Exception as loi:  # noqa: BLE001 — file hỏng đủ kiểu
        raise LoiNapSan("Không mở được file Excel: {0}".format(loi)) from loi
    try:
        if "scenes" not in sach.sheetnames:
            raise LoiNapSan(
                "File Excel thiếu trang tính tên “scenes”. Bạn tải file mẫu về "
                "rồi điền vào đó cho chắc.")
        trang = sach["scenes"]
        dau = [str(o.value or "").strip() for o in next(trang.iter_rows(max_row=1))]
        # Chỉ đòi những cột khâu sau THẬT SỰ đọc. Đòi đủ 25 cột là bắt khách
        # điền hai chục ô mà tool không bao giờ nhìn tới.
        can = ("scene_id", "srt_start", "duration", "img_prompt")
        thieu = [c for c in can if c not in dau]
        if thieu:
            raise LoiNapSan(
                "File Excel thiếu cột: {0}.\n\nBạn tải file mẫu về, điền vào "
                "đó rồi tải lên lại.".format(", ".join(thieu)))
        so_dong = sum(1 for _ in trang.iter_rows(min_row=2))
        if so_dong < 1:
            raise LoiNapSan("Trang “scenes” chưa có dòng cảnh nào.")
    finally:
        sach.close()


def _kiem_thu_muc(duong: str, duoi: Sequence[str]) -> None:
    if not os.path.isdir(duong):
        raise LoiNapSan("Đây không phải một thư mục.")
    co = [t for t in os.listdir(duong)
          if os.path.splitext(t)[1].lower() in duoi]
    if not co:
        raise LoiNapSan(
            "Thư mục này không có file {0} nào."
            .format(" hoặc ".join(duoi)))


def kiem_file(ma_khau: str, duong: str) -> None:
    """Soi file. Không hợp thì ném `LoiNapSan` với câu nói rõ thiếu gì."""
    if ma_khau in _LA_THU_MUC:
        _kiem_thu_muc(duong, (".png", ".jpg", ".jpeg")
                      if ma_khau != "clip" else (".mp4",))
        return
    if not os.path.isfile(duong):
        raise LoiNapSan("Không thấy file này.")
    if os.path.getsize(duong) == 0:
        raise LoiNapSan("File rỗng.")
    duoi = os.path.splitext(duong)[1].lower()
    nhan = _DUOI.get(ma_khau, ())
    if nhan and duoi not in nhan:
        raise LoiNapSan(
            "Khâu “{0}” nhận file {1}, còn file bạn chọn là {2}."
            .format(ten_khau(ma_khau), " hoặc ".join(nhan), duoi or "(không đuôi)"))
    {"kich-ban": _kiem_kich_ban,
     "phu-de": _kiem_phu_de,
     "bang-canh": _kiem_bang_canh}.get(ma_khau, lambda _d: None)(duong)


def nap_file(thu_muc_luot: str, ma_khau: str, duong_nguon: str) -> str:
    """Chép file của khách vào lượt chạy. Trả về đường dẫn đã chép tới.

    Soi trước rồi mới chép — không để lại một nửa.
    """
    kiem_file(ma_khau, duong_nguon)
    san_pham = san_pham_khau(ma_khau)
    if not san_pham:
        raise LoiNapSan("Khâu này không nạp sẵn được.")
    dich = os.path.join(thu_muc_luot, san_pham)

    if ma_khau in _LA_THU_MUC:
        os.makedirs(dich, exist_ok=True)
        for ten in sorted(os.listdir(duong_nguon)):
            nguon = os.path.join(duong_nguon, ten)
            if os.path.isfile(nguon):
                shutil.copy2(nguon, os.path.join(dich, ten))
        return dich

    os.makedirs(os.path.dirname(dich) or ".", exist_ok=True)
    # Giữ nguyên đuôi tool mong đợi, kể cả file nguồn tên khác. Khâu sau đi tìm
    # đúng cái tên trong `KHAU`, không đi dò thư mục.
    shutil.copy2(duong_nguon, dich)
    return dich


# ── File mẫu ─────────────────────────────────────────────────────────────────

def co_mau(ma_khau: str) -> bool:
    """Khâu này có file mẫu để tải về không."""
    return ma_khau in ("bang-canh", "kich-ban")


def viet_mau(ma_khau: str, duong_dich: str) -> str:
    """Ghi file mẫu ra `duong_dich`. Trả về đường dẫn đã ghi."""
    if ma_khau == "kich-ban":
        return _mau_kich_ban(duong_dich)
    if ma_khau == "bang-canh":
        return _mau_bang_canh(duong_dich)
    raise LoiNapSan("Khâu này chưa có file mẫu.")


def _mau_kich_ban(duong: str) -> str:
    with open(duong, "w", encoding="utf-8") as ra:
        ra.write(
            "Đây là file mẫu cho kịch bản.\n"
            "\n"
            "Bạn xoá hết chữ trong file này rồi dán kịch bản của bạn vào, lưu "
            "lại, sau đó bấm “Nạp file có sẵn” ở tab Tự động.\n"
            "\n"
            "Vài điều nên biết:\n"
            "\n"
            "• Chỉ viết LỜI ĐỌC. Đừng ghi “Cảnh 1”, “(nhạc nổi lên)”, tên nhân "
            "vật ở đầu dòng — máy đọc sẽ đọc ra hết những thứ đó.\n"
            "\n"
            "• Độ dài quyết định độ dài video. Khoảng 340 chữ Nhật là một phút; "
            "tiếng Việt khoảng 830 chữ một phút; tiếng Anh khoảng 920.\n"
            "\n"
            "• Xuống dòng thoải mái. Tool tự cắt câu khi làm phụ đề.\n"
            "\n"
            "• Lưu ở dạng .txt và mã UTF-8 (Notepad: Lưu thành → Mã hoá UTF-8). "
            "Lưu sai mã thì chữ có dấu thành ô vuông.\n")
    return duong


def _mau_bang_canh(duong: str) -> str:
    from .auto_khau import COT_CANH

    try:
        from openpyxl import Workbook
    except ImportError as loi:  # pragma: no cover
        raise LoiNapSan(
            "Máy chưa có bộ ghi Excel. Bạn nhấp đúp SETUP.bat một lần rồi "
            "thử lại.") from loi

    sach = Workbook()
    trang = sach.active
    trang.title = "scenes"
    trang.append(list(COT_CANH))

    # Hai dòng mẫu điền sẵn, không phải một. Một dòng không cho thấy cảnh sau
    # phải nối tiếp cảnh trước ở cột thời gian — mà đó chính là chỗ hay điền sai.
    trang.append([1, 0.0, 5.5, 5.5, 5.5,
                  "Câu đầu tiên của lời đọc, đúng như trong phụ đề.",
                  "establishing", "solo", "nv1",
                  "ngồi bên cửa sổ, nhìn ra ngoài",
                  "ánh sáng chiều qua rèm mỏng", "chữ, logo",
                  "Warm afternoon light through thin curtains, a person sitting "
                  "by the window seen from behind, quiet room, soft shadows",
                  "", "", "", "", "", "", "nv1", "phòng ngủ", "nv1.png", "", "", 1])
    trang.append([2, 5.5, 11.0, 5.5, 5.5,
                  "Câu thứ hai của lời đọc.",
                  "detail", "solo", "nv1",
                  "cầm tách trà, tay hơi run",
                  "hơi nước bốc lên trong ánh sáng ngược", "chữ, logo",
                  "Close-up of hands holding a warm ceramic cup, steam catching "
                  "the backlight, shallow depth of field",
                  "", "", "", "", "", "", "nv1", "phòng ngủ", "nv1.png", "", "", 1])

    huong = sach.create_sheet("huong-dan")
    for dong in (
        ("CỘT", "PHẢI ĐIỀN?", "GHI CHÚ"),
        ("scene_id", "Có", "Số thứ tự cảnh, đếm từ 1, không nhảy số."),
        ("srt_start", "Có", "Giây cảnh này bắt đầu, tính từ đầu video. Cảnh "
                            "đầu là 0."),
        ("duration", "Có", "Cảnh dài mấy giây. Nên 3–8 giây; clip Veo3 luôn "
                           "dài 8 giây nên cảnh dài hơn 8 sẽ bị cắt."),
        ("img_prompt", "Có", "Mô tả ảnh, VIẾT BẰNG TIẾNG ANH. Đây là thứ quyết "
                             "định ảnh đẹp hay xấu."),
        ("srt_end", "Nên", "srt_start + duration."),
        ("srt_text", "Nên", "Lời đọc của cảnh này. Tool dùng để đối chiếu."),
        ("reference_files", "Nên", "nv1.png nếu cảnh có nhân vật — thiếu nó thì "
                                   "mỗi cảnh ra một người khác nhau."),
        ("must_not_show", "Nên", "Thứ không được xuất hiện. Luôn để “chữ, logo” "
                                 "— chữ do AI vẽ ra luôn sai chính tả."),
        ("", "", ""),
        ("Các cột còn lại", "Không", "Để trống. Tool tự điền trong lúc chạy."),
        ("", "", ""),
        ("Cảnh chia theo Ý,", "", "không theo đồng hồ. Một ý nói xong thì sang "
                                  "cảnh mới, dù mới 3 giây."),
    ):
        huong.append(list(dong))
    huong.column_dimensions["A"].width = 18
    huong.column_dimensions["B"].width = 12
    huong.column_dimensions["C"].width = 70

    sach.save(duong)
    return duong
