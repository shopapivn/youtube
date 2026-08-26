"""Dựng video từ ảnh/clip + lời đọc + phụ đề — chạy bằng FFmpeg **trên máy khách**.

Khâu này không gọi máy chủ, không tốn tiền: ghép file là việc của CPU nhà mình.

═══ MỘT HÀM KIỂM DUY NHẤT ═══

Tool tham chiếu (`D:\\AUTO\\ve3-tool-simple`) có hai chỗ kiểm điều kiện chạy —
giao diện chấp nhận *bất kỳ* file mp3 nào, còn bộ dựng đòi đúng `<mã>.mp3`. Hệ
quả: bảng ghi "Sẵn sàng" rồi bộ dựng lặng lẽ bỏ qua, và khách không có cách nào
biết vì sao. Ở đây :func:`doc_du_an` là chỗ **duy nhất** quyết định một dự án
chạy được hay không; giao diện chỉ hiển thị lại `DuAn.thieu`.

═══ QUÉT KHÔNG ĐƯỢC PHÁ ═══

Tool tham chiếu `shutil.rmtree` thư mục nguồn ngay trong hàm quét khi thấy thiếu
ảnh, và xoá cả nguồn sau khi dựng xong. Ở đây **không có một lời gọi xoá nào**.
Quét là đọc. Thiếu thì báo thiếu.

═══ THƯ MỤC ═══

Khách chọn một thư mục gốc; mỗi thư mục con là một dự án::

    <gốc>/<tên dự án>/
        *.mp3 | *.wav        lời đọc      (bắt buộc)
        *.srt                phụ đề       (tuỳ chọn)
        *.png *.jpg *.mp4 …  hình ảnh     (bắt buộc, ≥1)
        nhac/*.mp3           nhạc nền     (tuỳ chọn)

Không ép đặt tên file theo tên thư mục: người làm YouTube tải ảnh về với đủ thứ
tên, bắt họ đổi tên tay là chặn ngay ở bước đầu.

═══ VÀ NHẬN LUÔN THƯ MỤC DO CHÍNH TOOL GHI RA ═══

Kiểu trên là kiểu khách tự xếp tay. Nhưng **chính tool này không bao giờ ghi ra
kiểu đó**: tab Voice ghi vào `VOICE/`, tab Ảnh & Video ghi vào `VISUAL/`, lượt
Tự động ghi vào `5-anh/` và `6-clip/`. Ai đi hết dây chuyền rồi mở tab Dựng
video và trỏ vào thư mục của mình sẽ nhận đúng một dòng: *"0 dự án, 0 sẵn
sàng"* — báo cáo của khách ngày 26/08/2026, ảnh chụp trỏ vào
`PROJECTS/video-dau-tien/VISUAL`.

Nên :func:`doc_du_an` tìm cả ba kiểu, và :func:`quet_thu_muc` nhận đúng chỗ
khách trỏ tới (một ngăn, một dự án, hay thư mục chứa nhiều dự án). Bảng vẫn chỉ
có **một** chỗ kết luận chạy được hay không, như cũ.
"""

from __future__ import annotations

import os
import re
import shutil
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Sequence, Tuple

from .tron_tieng import loc_tron_nhac

__all__ = [
    "DuAn", "CaiDatDung", "DUOI_ANH", "DUOI_CLIP", "DUOI_TIENG", "DO_PHAN_GIAI",
    "MAU_CHU", "VI_TRI_PHU_DE", "NGAN_DU_AN", "tim_ffmpeg", "doc_du_an",
    "quet_thu_muc", "la_thu_muc_du_an", "du_an_cha", "khoa_tu_nhien",
    "doc_bang_canh", "khop_canh_voi_hinh", "giay_tung_hinh",
    "du_an_chon_tay", "phu_de_tu_txt",
    "lenh_ffmpeg", "loc_srt_style", "thoi_luong_moi_anh", "la_clip", "doc_thoi_luong",
]

DUOI_ANH = (".png", ".jpg", ".jpeg", ".webp", ".bmp")
DUOI_CLIP = (".mp4", ".mov", ".mkv", ".webm")
DUOI_TIENG = (".mp3", ".wav", ".m4a", ".aac")

#: Tên → (rộng, cao). 4K để cuối vì render lâu gấp nhiều lần.
DO_PHAN_GIAI = {
    "1080p": (1920, 1080),
    "1440p": (2560, 1440),
    "4K": (3840, 2160),
}

#: Tên tiếng Việt → màu kiểu ASS (`&HBBGGRR`, ngược thứ tự so với web).
MAU_CHU = {
    "Trắng": "&H00FFFFFF",
    "Vàng": "&H0000D7FF",
    "Xanh lá": "&H0000FF00",
    "Đỏ": "&H000000FF",
}

#: Tên → mã canh lề ASS.
VI_TRI_PHU_DE = {"Dưới": 2, "Giữa": 5, "Trên": 8}

#: Nhạc nền nhỏ hơn lời đọc nhiều — đây là video kể chuyện, không phải MV.
#:
#: Chỉ còn dùng cho **đường lui**, khi bản FFmpeg trong máy thiếu bộ lọc
#: `sidechaincompress`. Bình thường nhạc tự né giọng và độ to lúc không có lời
#: lấy theo `core.tron_tieng.AM_LUONG_NE`.
AM_LUONG_NHAC = 0.18


@dataclass(frozen=True)
class DuAn:
    """Một thư mục dự án đã đọc xong. `thieu` rỗng nghĩa là dựng được."""

    ten: str
    thu_muc: str
    tieng: str = ""
    phu_de: str = ""
    hinh: Tuple[str, ...] = ()
    nhac: Tuple[str, ...] = ()
    thieu: Tuple[str, ...] = ()
    da_xong: str = ""
    #: Bảng cảnh (`.xlsx` hoặc `.json`) — nơi ghi **giây bắt đầu** của từng
    #: cảnh. Có nó thì hình bám đúng lời; không có thì đành chia đều.
    bang_canh: str = ""

    @property
    def chay_duoc(self) -> bool:
        return not self.thieu

    @property
    def trang_thai(self) -> str:
        if self.da_xong:
            return "đã dựng xong"
        if self.thieu:
            return "thiếu " + ", ".join(self.thieu)
        return "sẵn sàng" if self.bang_canh else "sẵn sàng (chia đều)"


@dataclass
class CaiDatDung:
    """Tuỳ chọn dựng. Một chỗ duy nhất, không rải ra ba file như tool tham chiếu."""

    do_phan_giai: str = "1080p"
    fps: int = 30
    chuyen_canh: float = 0.5
    ken_burns: str = "Nhẹ"
    phu_de: bool = True
    font: str = "Arial"
    co_chu: int = 28
    mau_chu: str = "Trắng"
    vi_tri: str = "Dưới"
    nhac_nen: bool = True
    am_luong_nhac: float = AM_LUONG_NHAC
    thu_muc_ra: str = ""
    tang_toc_gpu: bool = False  # bật thì dùng GPU encode (nhanh hơn, chất lượng hơi kém)


def tim_ffmpeg() -> str:
    """Đường dẫn FFmpeg dùng được, hoặc chuỗi rỗng.

    Ưu tiên bản cài trong máy; không có thì lấy bản đi kèm `imageio-ffmpeg` —
    khách không biết code không nên phải tự đi cài FFmpeg rồi sửa biến PATH.
    """
    san_co = shutil.which("ffmpeg")
    if san_co:
        return san_co
    try:
        import imageio_ffmpeg

        duong_dan = imageio_ffmpeg.get_ffmpeg_exe()
    except Exception:  # noqa: BLE001 — thiếu gói là chuyện bình thường
        return ""
    return duong_dan if duong_dan and os.path.isfile(duong_dan) else ""


#: Các ngăn của một dự án do chính tool tạo ra (xem `core/du_an.NGAN`).
NGAN_DU_AN = ("CONTENT", "VOICE", "EXCEL", "VISUAL", "DONE")

#: Nơi tìm **lời đọc**, theo thứ tự ưu tiên. `""` = ngay trong thư mục dự án.
_CHO_TIENG = ("", "voice", "giong", "giong-doc", "audio", "am-thanh")

#: Nơi tìm **ảnh/clip**. `6-clip` đứng trước `5-anh`: lượt Tự động làm ảnh
#: trước rồi biến ảnh thành clip, có clip thì dựng bằng clip.
_CHO_HINH = ("", "visual", "6-clip", "5-anh", "anh", "img", "images", "hinh", "clip")

#: Nơi tìm **phụ đề**.
_CHO_PHU_DE = ("", "excel", "visual")

#: Nơi tìm **nhạc nền**.
_CHO_NHAC = ("nhac", "music")

#: Tên file là **bản dựng xong**, không phải nguồn. Thư mục một lượt Tự động
#: có `8-video.mp4` nằm ngay bên cạnh `6-clip/`; nhận nhầm nó là "hình" thì tab
#: Dựng video đem chính bản đã dựng ra dựng lại.
_TEP_KET_QUA = ("8-video.mp4", "8-video.cu.mp4")


def khoa_tu_nhien(ten: str):
    """Khoá xếp thứ tự **theo số người đọc**, không theo bảng mã.

    ═══ CẢNH 1 → 10 → 11 → 2 ═══

    `sorted()` xếp theo từng ký tự, nên `10.png` đứng trước `2.png`. Ảnh của
    một lượt Tự động tên đúng bằng số cảnh (`5-anh/1.png` … `5-anh/12.png`), mà
    video thật nào cũng hơn 10 cảnh — tức là **mọi** video dựng ở tab này đều
    ra sai thứ tự cảnh, và chỉ lộ ra khi ngồi xem lại. Khâu ghép của tab Tự
    động không dính vì nó đi theo số cảnh trong bảng cảnh, không đi theo tên
    tệp; lỗi nằm đúng một chỗ này.

    >>> sorted(["1.png", "10.png", "2.png"], key=khoa_tu_nhien)
    ['1.png', '2.png', '10.png']
    """
    manh = re.split(r"(\d+)", str(ten).lower())
    # `(0, số)` và `(1, chữ)` để số luôn đứng trước chữ ở cùng vị trí, và để
    # hai kiểu không bao giờ bị đem so với nhau (Python 3 ném lỗi nếu so).
    return [(0, int(m), "") if m.isdigit() else (1, 0, m) for m in manh]


def _liet_ke(thu_muc: str, duoi: Sequence[str],
             bo_ten: Sequence[str] = ()) -> List[str]:
    try:
        ten = sorted(os.listdir(thu_muc), key=khoa_tu_nhien)
    except OSError:
        return []
    bo = {t.lower() for t in bo_ten}
    return [os.path.join(thu_muc, t) for t in ten
            if os.path.splitext(t)[1].lower() in duoi
            and t.lower() not in bo
            and os.path.isfile(os.path.join(thu_muc, t))]


def _thu_muc_con(thu_muc: str) -> Dict[str, str]:
    """`{tên viết thường: đường dẫn}` của các thư mục con. Đọc đĩa một lần."""
    ket: Dict[str, str] = {}
    try:
        ten_con = sorted(os.listdir(thu_muc), key=khoa_tu_nhien)
    except OSError:
        return ket
    for t in ten_con:
        duong = os.path.join(thu_muc, t)
        if os.path.isdir(duong):
            ket.setdefault(t.lower(), duong)
    return ket


def _tim_theo_cho(thu_muc: str, con: Dict[str, str], cho: Sequence[str],
                  duoi: Sequence[str], bo_ten: Sequence[str] = ()) -> List[str]:
    """File đầu tiên tìm thấy theo thứ tự ưu tiên trong `cho`."""
    for ten in cho:
        duong = thu_muc if ten == "" else con.get(ten, "")
        if not duong:
            continue
        co = _liet_ke(duong, duoi, bo_ten if ten == "" else ())
        if co:
            return co
    return []


def la_thu_muc_du_an(thu_muc: str) -> bool:
    """Thư mục này là **một video**, hay là chỗ chứa nhiều video?

    Nhận ra ba kiểu thư mục thật đang có trên máy khách:

    * dự án của tool — có các ngăn `VOICE/`, `VISUAL/`… (`core/du_an.py`);
    * một lượt Tự động — có `1-kich-ban.txt` hoặc `trang-thai.json`;
    * thư mục khách tự xếp — mp3 và ảnh nằm chung một chỗ.
    """
    if not thu_muc or not os.path.isdir(thu_muc):
        return False
    con = _thu_muc_con(thu_muc)
    if sum(1 for n in NGAN_DU_AN if n.lower() in con) >= 2:
        return True
    for dau_hieu in ("trang-thai.json", "1-kich-ban.txt", "2-giong-doc.mp3"):
        if os.path.isfile(os.path.join(thu_muc, dau_hieu)):
            return True
    return bool(_liet_ke(thu_muc, DUOI_TIENG)
                and _liet_ke(thu_muc, DUOI_ANH + DUOI_CLIP, _TEP_KET_QUA))


def du_an_cha(thu_muc: str) -> str:
    """Trỏ vào một ngăn (`…/VISUAL`) thì lùi ra thư mục dự án chứa nó.

    Khách đi hết dây chuyền của tool rồi mở tab Dựng video và trỏ đúng vào chỗ
    họ thấy ảnh — `PROJECTS/<dự án>/VISUAL`. Bắt họ lùi ra một cấp bằng cách in
    "0 dự án" là bắt họ đoán.
    """
    if not thu_muc or not os.path.isdir(thu_muc):
        return thu_muc
    ten = os.path.basename(os.path.normpath(thu_muc))
    if ten.upper() not in NGAN_DU_AN:
        return thu_muc
    cha = os.path.dirname(os.path.normpath(thu_muc))
    return cha if la_thu_muc_du_an(cha) else thu_muc


def doc_du_an(thu_muc: str, *, thu_muc_ra: str = "", can_phu_de: bool = False) -> DuAn:
    """Đọc một thư mục dự án. **Chỉ đọc** — không tạo, không sửa, không xoá.

    Nhận cả thư mục khách tự xếp (mp3 và ảnh nằm chung) lẫn thư mục do **chính
    tool này** ghi ra — tab Voice ghi vào `VOICE/`, tab Ảnh & Video ghi vào
    `VISUAL/`, lượt Tự động ghi vào `5-anh/` và `6-clip/`. Trước 26/08/2026 hàm
    này chỉ biết kiểu đầu, nên đi hết dây chuyền của tool xong thì tab Dựng
    video báo "0 dự án" — đúng lỗi khách gặp.
    """
    ten = os.path.basename(os.path.normpath(thu_muc))
    con = _thu_muc_con(thu_muc)
    con.pop("done", None)  # DONE là bản dựng xong, không phải nguồn
    bo_qua = tuple(_TEP_KET_QUA) + (ten + ".mp4",)
    tieng = _tim_theo_cho(thu_muc, con, _CHO_TIENG, DUOI_TIENG)
    srt = _tim_theo_cho(thu_muc, con, _CHO_PHU_DE, (".srt",))
    hinh = _tim_theo_cho(thu_muc, con, _CHO_HINH, DUOI_ANH + DUOI_CLIP, bo_qua)
    nhac: List[str] = []
    for ten_con in _CHO_NHAC:
        if ten_con in con:
            nhac += _liet_ke(con[ten_con], DUOI_TIENG)
    bang_canh = _tim_bang_canh(thu_muc, con)

    thieu: List[str] = []
    if not tieng:
        thieu.append("lời đọc (.mp3/.wav)")
    if not hinh:
        thieu.append("ảnh hoặc clip")
    if can_phu_de and not srt:
        thieu.append("phụ đề (.srt)")

    da_xong = ""
    if thu_muc_ra:
        ra = os.path.join(thu_muc_ra, ten + ".mp4")
        if os.path.isfile(ra) and os.path.getsize(ra) > 0:
            da_xong = ra

    return DuAn(ten=ten, thu_muc=thu_muc, tieng=tieng[0] if tieng else "",
                phu_de=srt[0] if srt else "", hinh=tuple(hinh), nhac=tuple(nhac),
                thieu=tuple(thieu), da_xong=da_xong, bang_canh=bang_canh)


def _tim_bang_canh(thu_muc: str, con: Dict[str, str]) -> str:
    """Bảng cảnh của dự án — trong thư mục, hay trong ngăn `EXCEL/`.

    Chỉ nhận file **đọc ra được mốc thời gian**: một bảng cảnh chỉ có lời nhắc
    mà không có `srt_start` thì giữ lại làm gì cũng không biết cảnh nào dài bao
    lâu, mà bảng lại ghi "theo bảng cảnh" — hứa suông.
    """
    cho = [thu_muc] + [con[t] for t in ("excel",) if t in con]
    for duong in cho:
        for ten in _TEN_BANG_CANH:
            tep = os.path.join(duong, ten)
            if os.path.isfile(tep) and doc_bang_canh(tep):
                return tep
    for duong in cho:
        for tep in _liet_ke(duong, (".xlsx", ".json")):
            if doc_bang_canh(tep):
                return tep
    return ""


def du_an_chon_tay(ten: str, thu_muc_hinh: str, tieng: str, *,
                   phu_de: str = "", bang_canh: str = "", nhac: str = "",
                   thu_muc_ra: str = "", can_phu_de: bool = False) -> DuAn:
    """Dựng một `DuAn` từ những thứ khách **tự chỉ vào**, không qua quét.

    ═══ VÌ SAO CẦN, DÙ ĐÃ CÓ QUÉT TỰ ĐỘNG ═══

    Chủ dự án, 26/08/2026: *"có thể có 1 option để khách thêm các dữ liệu đó
    vào"*. Quét tự động chỉ đúng khi file nằm đúng chỗ tool quen. Người làm
    YouTube thì có ảnh mua ngoài, giọng thu bằng micro thật, bảng cảnh sửa tay
    — mỗi thứ một ổ đĩa. Bắt họ chép hết vào một thư mục cho vừa ý tool là bắt
    làm lại việc tool sinh ra để đỡ.

    `tieng` và `nhac` nhận **cả file lẫn thư mục**: khách trỏ vào ngăn `VOICE`
    cũng phải chạy, không thì lại thành một cái bẫy nữa.
    """
    hinh = _liet_ke(thu_muc_hinh, DUOI_ANH + DUOI_CLIP) if thu_muc_hinh else []
    tep_tieng = _mot_tep(tieng, DUOI_TIENG)
    tep_nhac = tuple(_liet_ke(nhac, DUOI_TIENG)) if os.path.isdir(nhac or "") \
        else ((nhac,) if nhac and os.path.isfile(nhac) else ())

    thieu: List[str] = []
    if not tep_tieng:
        thieu.append("lời đọc (.mp3/.wav)")
    if not hinh:
        thieu.append("ảnh hoặc clip")
    if can_phu_de and not phu_de:
        thieu.append("phụ đề (.srt)")

    ten = (ten or os.path.basename(os.path.normpath(thu_muc_hinh or "")) or "video")
    da_xong = ""
    if thu_muc_ra:
        ra = os.path.join(thu_muc_ra, ten + ".mp4")
        if os.path.isfile(ra) and os.path.getsize(ra) > 0:
            da_xong = ra
    return DuAn(ten=ten, thu_muc=thu_muc_hinh, tieng=tep_tieng, phu_de=phu_de,
                hinh=tuple(hinh), nhac=tep_nhac, thieu=tuple(thieu),
                da_xong=da_xong,
                bang_canh=bang_canh if doc_bang_canh(bang_canh) else "")


def _mot_tep(duong: str, duoi: Sequence[str]) -> str:
    """File đầu tiên dùng được — nhận cả đường dẫn file lẫn thư mục."""
    if not duong:
        return ""
    if os.path.isfile(duong):
        return duong if os.path.splitext(duong)[1].lower() in duoi else ""
    co = _liet_ke(duong, duoi)
    return co[0] if co else ""


def phu_de_tu_txt(duong_txt: str, duong_tieng: str, dich_srt: str, *,
                  ngon_ngu: str = "", on_log=None, nghe=None) -> str:
    """Kịch bản `.txt` → phụ đề `.srt`, ép khớp vào chính giọng đọc.

    Chạy trên máy, **miễn phí**. Máy yếu không chạy nổi bộ nghe thì
    `core.phu_de` tự rải đều theo số ký tự — chữ vẫn đúng 100%, mốc thời gian
    xê xích vài phần mười giây. Hỏng hẳn thì trả chuỗi rỗng và người gọi dựng
    tiếp không phụ đề, chứ không chặn cả video.
    """
    from .phu_de import tao_phu_de, viet_srt  # noqa: PLC0415

    try:
        with open(duong_txt, "r", encoding="utf-8", errors="replace") as tep:
            kich_ban = tep.read()
    except OSError:
        return ""
    if not kich_ban.strip():
        return ""
    ket = tao_phu_de(duong_tieng, kich_ban, ngon_ngu=ngon_ngu, on_log=on_log,
                     nghe=nghe)
    if not ket.cau:
        return ""
    os.makedirs(os.path.dirname(dich_srt) or ".", exist_ok=True)
    viet_srt(dich_srt, ket.cau)
    return dich_srt


def _cung_cho(mot: str, hai: str) -> bool:
    if not mot or not hai:
        return False
    try:
        return os.path.normcase(os.path.abspath(mot)) == os.path.normcase(
            os.path.abspath(hai))
    except OSError:
        return False


#: Quét sâu tối đa mấy tầng dưới thư mục khách chọn. `PROJECTS/AUTO/<kênh>/
#: <lượt>` là chỗ sâu nhất tool tự ghi ra — ba tầng.
SAU_NHAT = 3

#: Trần số dự án một lượt quét. Khách lỡ trỏ vào `C:\\` thì dừng, đừng treo máy.
TRAN_DU_AN = 300


def quet_thu_muc(goc: str, *, thu_muc_ra: str = "", can_phu_de: bool = False) -> List[DuAn]:
    """Tìm các dự án dựng được dưới `goc`. **Chỉ đọc**, không tạo, không xoá.

    Nhận đúng chỗ khách trỏ tới, thay vì bắt khách đoán đúng một tầng duy nhất:

    * trỏ vào một ngăn (`…/VISUAL`) → lùi ra dự án chứa nó;
    * trỏ thẳng vào **một** dự án → trả về đúng dự án đó;
    * trỏ vào `PROJECTS/` → mỗi thư mục con là một dự án, và thư mục nào chưa
      phải dự án (như `AUTO/<kênh>`) thì đi sâu tiếp, tối đa :data:`SAU_NHAT`
      tầng.

    **Bỏ qua chính thư mục kết quả.** Khách rất hay đặt thư mục lưu ngay bên
    trong thư mục dự án; không loại nó ra thì mỗi lần quét lại mọc thêm một "dự
    án" toàn file mp4 và luôn báo thiếu lời đọc — nhìn như tool đếm sai.
    """
    goc = du_an_cha(goc)
    if not goc or not os.path.isdir(goc):
        return []
    if la_thu_muc_du_an(goc):
        return [doc_du_an(goc, thu_muc_ra=thu_muc_ra, can_phu_de=can_phu_de)]

    ket: List[DuAn] = []

    def di(thu_muc: str, sau: int) -> None:
        try:
            ten_con = sorted(os.listdir(thu_muc), key=khoa_tu_nhien)
        except OSError:
            return
        for ten in ten_con:
            if len(ket) >= TRAN_DU_AN:
                return
            duong = os.path.join(thu_muc, ten)
            if (ten.startswith(".") or ten.upper() == "DONE"
                    or not os.path.isdir(duong)
                    or _cung_cho(duong, thu_muc_ra)):
                continue
            if la_thu_muc_du_an(duong) or sau <= 1:
                ket.append(doc_du_an(duong, thu_muc_ra=thu_muc_ra,
                                     can_phu_de=can_phu_de))
            else:
                # Chưa phải dự án mà bên trong còn thư mục con: đi tiếp. Hiện
                # `AUTO/` ra thành một dòng "thiếu lời đọc, thiếu ảnh" chỉ làm
                # bảng bẩn — thứ khách cần là các lượt nằm sâu bên trong.
                truoc = len(ket)
                di(duong, sau - 1)
                if len(ket) == truoc:
                    ket.append(doc_du_an(duong, thu_muc_ra=thu_muc_ra,
                                         can_phu_de=can_phu_de))

    di(goc, SAU_NHAT)
    return ket


# ═══════════════════════════════════════════════════════════════════════════
# BẢNG CẢNH — GIÂY BẮT ĐẦU CỦA TỪNG CẢNH
# ═══════════════════════════════════════════════════════════════════════════
#
# Chủ dự án, 26/08/2026: *"trong excel có cái thời gian bắt đầu của cảnh đó,
# mày không có nó làm sao biết cảnh đó xuất hiện kết thúc khi nào"*.
#
# Đúng. Tab này trước đó chia đều thời lượng lời đọc cho số ảnh
# (:func:`thoi_luong_moi_anh`) — mà cảnh thì **chia theo nội dung**, không đều:
# đo trên một lượt thật, cảnh ngắn nhất 2,8 giây, dài nhất 8,0. Chia đều là
# hình trôi khỏi lời ngay từ cảnh thứ hai, và càng về cuối càng lệch.
#
# Khâu ghép của tab Tự động đã bỏ cách chia đều từ 14/08/2026 vì đúng lý do
# này (xem `core/auto_khau._khau_dung`). Tab Dựng video thì chưa — nên nó vẫn
# giữ nguyên cái sai ấy cho tới hôm nay.
#
# Luật, giống hệt khâu ghép: cảnh `i` chiếm từ `srt_start[i]` tới
# `srt_start[i+1]`. Khoảng người đọc ngừng lấy hơi tự khắc thuộc về cảnh
# trước — đúng như người dựng tay vẫn làm, hình đứng yên trong lúc ngừng.

#: Tên tệp bảng cảnh tool tự ghi ra, theo thứ tự ưu tiên.
_TEN_BANG_CANH = ("4-canh.xlsx", "4-canh.json", "bang-canh.xlsx")


def _giay(moc) -> float:
    """`00:01:23,450` hoặc `83.45` → số giây. Không đọc được thì `0.0`."""
    if moc is None or moc == "":
        return 0.0
    if isinstance(moc, (int, float)):
        return float(moc)
    chu = str(moc).strip().replace(",", ".")
    if ":" not in chu:
        try:
            return float(chu)
        except ValueError:
            return 0.0
    phan = chu.split(":")
    try:
        so = [float(p) for p in phan]
    except ValueError:
        return 0.0
    tong = 0.0
    for p in so:
        tong = tong * 60 + p
    return tong


def _so_trong_ten(duong_dan: str) -> Optional[int]:
    """Số cảnh nằm trong tên tệp (`5-anh/12.png` → 12). Không có thì `None`."""
    ten = os.path.splitext(os.path.basename(duong_dan))[0]
    so = re.findall(r"\d+", ten)
    return int(so[-1]) if so else None


def doc_bang_canh(duong: str) -> List[Dict[str, float]]:
    """Đọc bảng cảnh → `[{"so", "bat_dau", "ket_thuc"}]`, xếp theo số cảnh.

    Nhận cả `.xlsx` (sheet `scenes`, hoặc sheet đầu) lẫn `.json` — hai thứ tool
    tự ghi ra. Đọc hỏng thì trả danh sách rỗng chứ không ném: thiếu bảng cảnh
    là dựng kém đi, không phải là dừng lại.
    """
    if not duong or not os.path.isfile(duong):
        return []
    duoi = os.path.splitext(duong)[1].lower()
    try:
        hang = _hang_json(duong) if duoi == ".json" else _hang_excel(duong)
    except Exception:  # noqa: BLE001 — file của khách hỏng đủ kiểu
        return []
    ra: List[Dict[str, float]] = []
    for d in hang:
        try:
            so = int(float(str(d.get("scene_id", "")).strip()))
        except (TypeError, ValueError):
            continue
        bat_dau = _giay(d.get("srt_start"))
        ket_thuc = _giay(d.get("srt_end"))
        dai = _giay(d.get("duration"))
        if ket_thuc <= bat_dau and dai > 0:
            ket_thuc = bat_dau + dai
        ra.append({"so": so, "bat_dau": bat_dau, "ket_thuc": ket_thuc})
    ra.sort(key=lambda c: c["so"])
    # Bảng không có mốc thời gian nào (toàn 0) thì coi như không có bảng: thà
    # chia đều và nói thẳng, còn hơn dồn mọi cảnh vào giây 0.
    if not any(c["ket_thuc"] > 0 or c["bat_dau"] > 0 for c in ra):
        return []
    return ra


def _hang_json(duong: str) -> List[Dict]:
    import json  # noqa: PLC0415

    with open(duong, "r", encoding="utf-8") as tep:
        du_lieu = json.load(tep)
    if isinstance(du_lieu, dict):
        for khoa in ("scenes", "canh", "rows"):
            if isinstance(du_lieu.get(khoa), list):
                return [d for d in du_lieu[khoa] if isinstance(d, dict)]
        return []
    return [d for d in du_lieu if isinstance(d, dict)]


def _hang_excel(duong: str) -> List[Dict]:
    from openpyxl import load_workbook  # noqa: PLC0415

    sach = load_workbook(duong, read_only=True, data_only=True)
    try:
        trang = sach["scenes"] if "scenes" in sach.sheetnames else sach.worksheets[0]
        hang = list(trang.iter_rows(values_only=True))
    finally:
        sach.close()
    if len(hang) < 2:
        return []
    dau = [str(o or "").strip() for o in hang[0]]
    return [{ten: d[i] for i, ten in enumerate(dau) if ten and i < len(d)}
            for d in hang[1:]]


def khop_canh_voi_hinh(canh: Sequence[Dict[str, float]],
                       hinh: Sequence[str]) -> List[Dict[str, float]]:
    """Chọn ra các cảnh **có ảnh thật trên đĩa**, theo đúng thứ tự ảnh.

    Ảnh của một lượt Tự động tên đúng bằng số cảnh (`6-clip/12.mp4`), nên ghép
    theo số là chắc nhất — và nó tự lo luôn chuyện **thiếu cảnh**: cảnh nào
    không có ảnh thì bỏ khỏi danh sách, cảnh liền trước giữ hình bù vào, tiếng
    và phụ đề không xê dịch một mi-li-giây.

    Tên ảnh không mang số cảnh (khách tự đặt tên) thì ghép theo thứ tự, và chỉ
    khi hai bên **cùng số lượng** — đoán mò khi lệch số lượng là đẩy sai lệch
    vào cả video mà không ai biết.
    """
    if not canh or not hinh:
        return []
    theo_so = {int(c["so"]): c for c in canh}
    ket: List[Dict[str, float]] = []
    for duong_dan in hinh:
        so = _so_trong_ten(duong_dan)
        if so is not None and so in theo_so:
            ket.append(theo_so[so])
    if len(ket) == len(hinh):
        return ket
    return list(canh[:len(hinh)]) if len(canh) == len(hinh) else []


def giay_tung_hinh(canh: Sequence[Dict[str, float]], hinh: Sequence[str],
                   giay_tieng: float = 0.0) -> List[float]:
    """Số giây mỗi ảnh/clip được chiếm, lấy từ mốc bắt đầu của cảnh kế tiếp.

    Trả danh sách rỗng nếu bảng cảnh không ghép được với đống ảnh đang có —
    người gọi khi ấy quay về chia đều và **nói thẳng** là đang chia đều.
    """
    khop = khop_canh_voi_hinh(canh, hinh)
    if len(khop) != len(hinh) or not khop:
        return []
    giay: List[float] = []
    for i, c in enumerate(khop):
        if i + 1 < len(khop):
            giay.append(max(0.1, khop[i + 1]["bat_dau"] - c["bat_dau"]))
        else:
            giay.append(max(0.1, (c["ket_thuc"] or c["bat_dau"]) - c["bat_dau"]))
    # ═══ HÌNH NGẮN HƠN TIẾNG THÌ CẢNH CUỐI GIỮ HÌNH BÙ ═══
    #
    # Lệnh ghép dùng `-shortest`. Hình dài hơn tiếng thì nó cắt đuôi hình thừa —
    # vô hại. Nhưng hình NGẮN hơn tiếng thì nó cắt **lời đọc**, tức là khách
    # mất mấy câu cuối của video mà không có một dòng báo nào.
    if giay_tieng > 0 and sum(giay) < giay_tieng:
        giay[-1] += giay_tieng - sum(giay)
    return giay


def thoi_luong_moi_anh(giay_tieng: float, so_hinh: int) -> float:
    """Chia đều thời lượng lời đọc cho số ảnh, có sàn 2 giây.

    Ảnh nháy nhanh hơn 2 giây làm người xem chóng mặt; thà video dài hơn tiếng
    một chút — FFmpeg cắt phần thừa bằng `-shortest`.
    """
    if so_hinh <= 0:
        return 0.0
    return max(2.0, giay_tieng / so_hinh) if giay_tieng > 0 else 4.0


def loc_srt_style(cai: CaiDatDung) -> str:
    """Chuỗi `force_style` cho bộ lọc `subtitles` của FFmpeg."""
    return ",".join((
        "FontName={0}".format(cai.font),
        "FontSize={0}".format(int(cai.co_chu)),
        "PrimaryColour={0}".format(MAU_CHU.get(cai.mau_chu, MAU_CHU["Trắng"])),
        "OutlineColour=&H00000000",
        "BorderStyle=1", "Outline=2", "Shadow=0",
        "Alignment={0}".format(VI_TRI_PHU_DE.get(cai.vi_tri, 2)),
        "MarginV=48",
    ))


def _thoat_loc(duong_dan: str) -> str:
    """Escape đường dẫn nằm TRONG chuỗi bộ lọc FFmpeg.

    Bộ lọc `subtitles=` phân tích chuỗi hai lần, nên `D:\\a\\b.srt` phải thành
    `D\\\\:/a/b.srt`. Đây là chỗ hỏng kinh điển trên Windows: đường dẫn có dấu
    hai chấm ổ đĩa làm FFmpeg tưởng đang sang tuỳ chọn mới.
    """
    return duong_dan.replace("\\", "/").replace(":", "\\:").replace("'", "\\'")


def la_clip(duong_dan: str) -> bool:
    return os.path.splitext(duong_dan)[1].lower() in DUOI_CLIP


def doc_thoi_luong(ffmpeg: str, duong_dan: str) -> float:
    """Số giây của một file, đọc từ dòng `Duration:` FFmpeg in ra.

    Dùng chính FFmpeg thay vì `ffprobe`: bản FFmpeg đi kèm `imageio-ffmpeg` chỉ
    có mỗi `ffmpeg`, đòi thêm `ffprobe` là đòi khách tự đi cài.
    """
    import re
    import subprocess

    if not ffmpeg or not duong_dan:
        return 0.0
    try:
        xong = subprocess.run(
            [ffmpeg, "-hide_banner", "-i", duong_dan],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True,
            encoding="utf-8", errors="replace", check=False, timeout=60)
    except (OSError, subprocess.SubprocessError):
        return 0.0
    khop = re.search(r"Duration:\s*(\d+):(\d\d):(\d\d(?:\.\d+)?)", xong.stdout or "")
    if not khop:
        return 0.0
    gio, phut, giay = khop.groups()
    return int(gio) * 3600 + int(phut) * 60 + float(giay)


def lenh_ffmpeg(du_an: DuAn, cai: CaiDatDung, ffmpeg: str, dich: str, *,
                giay_moi_anh: float = 4.0, ne_giong: bool = True,
                giay: Optional[Sequence[float]] = None) -> List[str]:
    """Dựng danh sách tham số FFmpeg. Thuần tính toán — không chạy gì.

    Tách rời như vậy để test kiểm được nội dung lệnh mà không cần cài FFmpeg
    trong máy chạy test.

    `ne_giong` là **câu trả lời**, không phải câu hỏi: người gọi tự hỏi FFmpeg
    bằng `core.tron_tieng.co_ne_giong` rồi đưa kết quả xuống đây. Hỏi ngay tại
    đây thì hàm này hết thuần tính toán, và bài kiểm phải cài FFmpeg mới chạy
    được.

    Dùng **bộ lọc `concat`** chứ không dùng concat demuxer: demuxer đòi mọi đầu
    vào cùng codec, cùng khổ, cùng fps — ảnh PNG chụp màn hình lẫn với clip mp4
    tải về là hỏng. Bộ lọc thì nắn từng đầu vào về cùng khuôn rồi mới nối, nên
    trộn ảnh với clip vẫn chạy.

    `giay[i]` là số giây cảnh thứ `i` được chiếm, lấy từ bảng cảnh
    (:func:`giay_tung_hinh`). Có nó thì **clip cũng bị cắt về đúng khoảng của
    cảnh** — engine bán clip dài cố định, còn cảnh chia theo nội dung, nên
    không cắt là hình trôi khỏi lời. Để trống thì mọi ảnh dùng chung
    `giay_moi_anh` và clip chạy hết độ dài vốn có — cách cũ, chỉ đúng khi
    không có bảng cảnh.
    """
    if not du_an.chay_duoc:
        raise ValueError("Dự án {0} còn thiếu: {1}".format(
            du_an.ten, ", ".join(du_an.thieu)))
    if giay is not None and len(giay) != len(du_an.hinh):
        raise ValueError(
            "Bảng cảnh có {0} mốc thời gian nhưng có {1} ảnh/clip".format(
                len(giay), len(du_an.hinh)))
    rong, cao = DO_PHAN_GIAI.get(cai.do_phan_giai, DO_PHAN_GIAI["1080p"])
    lenh = [ffmpeg, "-hide_banner", "-loglevel", "error", "-y"]

    for i, duong_dan in enumerate(du_an.hinh):
        can = float(giay[i]) if giay is not None else max(0.5, giay_moi_anh)
        if la_clip(duong_dan):
            lenh += ["-i", duong_dan]
        else:
            # Ảnh tĩnh không có thời lượng: phải nói rõ giữ bao lâu, không thì
            # nó chỉ ra đúng một khung hình.
            lenh += ["-loop", "1", "-t", "{0:.3f}".format(max(0.1, can)),
                     "-i", duong_dan]
    so_hinh = len(du_an.hinh)
    chi_so_tieng = so_hinh
    lenh += ["-i", du_an.tieng]
    co_nhac = bool(cai.nhac_nen and du_an.nhac)
    if co_nhac:
        lenh += ["-stream_loop", "-1", "-i", du_an.nhac[0]]

    # Nắn mỗi đầu vào về đúng khổ. Dùng `pad` chứ không `crop`: cắt cho vừa
    # khung là cắt mất đầu nhân vật ở ảnh dọc.
    #
    # `flags=lanczos`: không ghi gì thì FFmpeg dùng `bicubic` — mềm. Ảnh của
    # kênh thường nhỏ hơn khung đích (nhà cung cấp trả 1408 chiều ngang, khung
    # 1080p là 1920) nên gần như tấm nào cũng bị phóng lên, và phóng bằng
    # lanczos nét hơn thấy được. Không tốn thêm thời gian đáng kể.
    khuon = ("scale={0}:{1}:force_original_aspect_ratio=decrease:flags=lanczos,"
             "pad={0}:{1}:(ow-iw)/2:(oh-ih)/2,setsar=1,fps={2},format=yuv420p")

    def _cat_clip(i: int) -> str:
        """Cắt clip thứ `i` về đúng khoảng cảnh của nó. Rỗng = để nguyên.

        `tpad=clone` trước, `trim` sau: cảnh dài hơn clip thì khung cuối đứng
        yên cho tới hết khoảng (đúng như người dựng tay để hình đứng lúc người
        đọc ngừng lấy hơi), cảnh ngắn hơn clip thì cắt bớt. Thiếu bước này,
        FFmpeg chèn đen — một nháy đen giữa video là lỗi ai cũng thấy.
        """
        if giay is None or not la_clip(du_an.hinh[i]):
            return ""
        can = max(0.1, float(giay[i]))
        return ("tpad=stop_mode=clone:stop_duration={0:.3f},"
                "trim=duration={0:.3f},setpts=PTS-STARTPTS,".format(can))

    phan = ["[{0}:v]{1}{2}[v{0}]".format(
        i, _cat_clip(i), khuon.format(rong, cao, cai.fps))
        for i in range(so_hinh)]
    phan.append("{0}concat=n={1}:v=1:a=0[vcat]".format(
        "".join("[v{0}]".format(i) for i in range(so_hinh)), so_hinh))
    nhan_v = "[vcat]"
    if cai.phu_de and du_an.phu_de:
        phan.append("[vcat]subtitles='{0}':force_style='{1}'[vsub]".format(
            _thoat_loc(du_an.phu_de), loc_srt_style(cai)))
        nhan_v = "[vsub]"
    if co_nhac:
        phan.append(loc_tron_nhac(
            "{0}:a".format(chi_so_tieng), "{0}:a".format(chi_so_tieng + 1),
            "aout", am_luong_deu=cai.am_luong_nhac, ne_giong=ne_giong))
        nhan_a = "[aout]"
    else:
        nhan_a = "{0}:a:0".format(chi_so_tieng)

    # Chọn encoder: mặc định CPU (libx264 medium crf 20). Khách bật "Tăng tốc
    # GPU" ở tab Dựng video, và máy có NVIDIA thật, thì dùng h264_nvenc — nhanh
    # hơn nhiều, chất lượng hơi kém một chút ở cùng dung lượng.
    lenh += ["-filter_complex", ";".join(phan), "-map", nhan_v, "-map", nhan_a]
    lenh += _tham_so_video(cai)
    lenh += ["-c:a", "aac", "-b:a", "192k",
             "-shortest", "-movflags", "+faststart", dich]
    return lenh


def _tham_so_video(cai: CaiDatDung) -> List[str]:
    """Tham số `-c:v` + preset cho tab Dựng video, theo lựa chọn tăng tốc GPU."""
    if getattr(cai, "tang_toc_gpu", False):
        try:
            from core.phan_cung import doc_ket_qua
            pc = doc_ket_qua(".")
            if pc and pc.gpu_nvidia and "h264_nvenc" in pc.ffmpeg_encoders:
                return ["-c:v", "h264_nvenc", "-preset", "p4", "-cq", "20",
                        "-pix_fmt", "yuv420p"]
        except Exception:  # noqa: BLE001 — dò hỏng thì lui về CPU
            pass
    # Mặc định an toàn: CPU
    return ["-c:v", "libx264", "-preset", "medium", "-crf", "20",
            "-pix_fmt", "yuv420p"]
