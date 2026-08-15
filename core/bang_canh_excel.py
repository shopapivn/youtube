"""Đọc và ghi bảng cảnh Excel cho tab **Ảnh & Video**.

═══ VÌ SAO CẦN THÊM MỘT BỘ ĐỌC NỮA ═══

`core/nap_san.py` đã có một bộ đọc bảng cảnh, nhưng nó phục vụ tab **Tự động**
nên đòi cả `srt_start` và `duration` — hai cột chỉ có nghĩa khi bảng cảnh sinh
ra từ phụ đề. Tab Ảnh & Video thì khác: khách gõ tay hoặc dán từ đâu đó, không
có phụ đề, không có mốc thời gian. Bắt họ điền hai cột ấy là bắt điền thứ tool
không bao giờ đọc tới.

Nên chỗ này soi **nhẹ hơn**: chỉ đòi có `scene_id`, và mỗi dòng có ít nhất một
trong `img_prompt` / `video_prompt`.

═══ VÌ SAO GIỮ TÊN CỘT TIẾNG ANH ═══

`scene_id`, `img_prompt`, `video_prompt`, `reference_files` — đúng tên VE3_SUITE
và tab Prompt Visuals dùng. Chủ dự án chọn giữ nguyên (15/08/2026) để **file
xuất từ Prompt Visuals nạp thẳng sang tab Ảnh & Video được**, không phải qua
bước chuyển đổi nào.

Đổi sang tên tiếng Việt thì khách đọc file dễ hơn một chút, nhưng mất hẳn cái
vòng tròn khép kín ấy. Phần dễ đọc đã có chỗ khác lo: trang `huong-dan` trong
chính file mẫu giải nghĩa từng cột bằng tiếng Việt.
"""

from __future__ import annotations

import os
from typing import Any, Dict, List, Optional

__all__ = ["COT", "COT_BAT_BUOC", "LoiBangCanh", "doc_excel", "viet_mau"]


class LoiBangCanh(ValueError):
    """File không dùng được — câu chữ trong này hiện thẳng lên màn hình."""


#: Cột của file mẫu, theo đúng thứ tự.
COT = ("scene_id", "img_prompt", "video_prompt", "reference_files")

#: Cột bắt buộc phải có mặt trong file. `reference_files` để trống được — khách
#: có thể dùng ô ảnh tham chiếu chung cho cả loạt thay vì điền từng dòng.
COT_BAT_BUOC = ("scene_id", "img_prompt", "video_prompt")


def _mo_sach(duong: str):
    try:
        from openpyxl import load_workbook
    except ImportError as loi:  # pragma: no cover — SETUP.bat cài sẵn
        raise LoiBangCanh(
            "Máy chưa có bộ đọc Excel. Bạn nhấp đúp SETUP.bat một lần rồi thử "
            "lại.") from loi
    try:
        return load_workbook(duong, read_only=True, data_only=True)
    except Exception as loi:  # noqa: BLE001 — file hỏng đủ kiểu
        raise LoiBangCanh("Không mở được file Excel: {0}".format(loi)) from loi


def doc_excel(duong: str) -> List[Dict[str, Any]]:
    """Đọc bảng cảnh. Trả về danh sách dòng, hoặc ném `LoiBangCanh`.

    Mỗi dòng là `{"so", "anh", "video", "tham_chieu"}`. Dòng trống hoàn toàn thì
    bỏ qua — người ta hay để vài dòng thừa ở cuối file.
    """
    if not os.path.isfile(duong):
        raise LoiBangCanh("Không thấy file này.")
    sach = _mo_sach(duong)
    try:
        trang = sach["scenes"] if "scenes" in sach.sheetnames else sach.worksheets[0]
        try:
            dau_tien = next(trang.iter_rows(max_row=1))
        except StopIteration:
            raise LoiBangCanh("File Excel rỗng.") from None
        dau = [str(o.value or "").strip() for o in dau_tien]

        thieu = [c for c in COT_BAT_BUOC if c not in dau]
        if thieu:
            # Gọi tên cột thiếu chứ không nói "file không hợp lệ": khách sửa
            # được cái thứ nhất, còn cái thứ hai thì chỉ biết ngồi nhìn.
            raise LoiBangCanh(
                "File Excel thiếu cột: {0}.\n\nBạn bấm “Tải file mẫu”, điền "
                "vào đó rồi tải lên lại.".format(", ".join(thieu)))

        vi = {c: dau.index(c) for c in COT if c in dau}
        ra: List[Dict[str, Any]] = []
        for hang in trang.iter_rows(min_row=2, values_only=True):
            lay = lambda c: (  # noqa: E731 — gọn hơn một hàm con ở đây
                str(hang[vi[c]] or "").strip()
                if c in vi and vi[c] < len(hang) else "")
            anh, video = lay("img_prompt"), lay("video_prompt")
            if not anh and not video:
                continue
            ra.append({
                "so": lay("scene_id") or str(len(ra) + 1),
                "anh": anh,
                "video": video,
                "tham_chieu": lay("reference_files"),
            })
        if not ra:
            raise LoiBangCanh(
                "File Excel không có dòng nào điền mô tả. Mỗi dòng cần ít nhất "
                "một trong hai cột img_prompt hoặc video_prompt.")
        return ra
    finally:
        sach.close()


def viet_mau(duong_dich: str) -> str:
    """Ghi file mẫu ra `duong_dich`. Trả về đường dẫn đã ghi."""
    try:
        from openpyxl import Workbook
    except ImportError as loi:  # pragma: no cover
        raise LoiBangCanh(
            "Máy chưa có bộ ghi Excel. Bạn nhấp đúp SETUP.bat một lần rồi thử "
            "lại.") from loi

    sach = Workbook()
    trang = sach.active
    trang.title = "scenes"
    trang.append(list(COT))

    # Ba dòng mẫu, mỗi dòng một kiểu dùng — vì ba kiểu ấy cho ra ba kết quả
    # khác hẳn nhau, và không nhìn thấy cả ba thì khách không đoán được.
    trang.append([1,
                  "Warm afternoon light through thin curtains, a person seen "
                  "from behind by the window, quiet room",
                  "slow push in, dust drifting in the light",
                  "nv1.png"])
    trang.append([2,
                  "Close-up of hands holding a warm ceramic cup, steam catching "
                  "the backlight",
                  "", "nv1.png"])
    trang.append([3, "", "the camera drifts left across an empty room",
                  "anh-cua-toi.png"])

    huong = sach.create_sheet("huong-dan")
    for dong in (
        ("CỘT", "PHẢI ĐIỀN?", "GHI CHÚ"),
        ("scene_id", "Nên", "Số thứ tự, đếm từ 1. Bỏ trống thì tôi tự đánh số."),
        ("img_prompt", "Xem ghi chú",
         "Mô tả ảnh, VIẾT BẰNG TIẾNG ANH. Đây là thứ quyết định ảnh đẹp hay xấu."),
        ("video_prompt", "Xem ghi chú",
         "Mô tả clip. Để trống thì dòng này chỉ tạo ảnh, không làm clip."),
        ("reference_files", "Nên",
         "Ảnh tham chiếu cho riêng dòng này. Bỏ trống thì dùng ảnh bạn chọn "
         "chung cho cả loạt."),
        ("", "", ""),
        ("Mỗi dòng cần ít nhất MỘT", "", "trong hai cột img_prompt và video_prompt."),
        ("", "", ""),
        ("Ba dòng mẫu là ba kiểu dùng:", "", ""),
        ("  dòng 1", "", "có cả hai → tạo ảnh trước, rồi cho ảnh ấy động đậy."),
        ("  dòng 2", "", "chỉ có mô tả ảnh → chỉ tạo ảnh."),
        ("  dòng 3", "", "chỉ có mô tả clip + ảnh tham chiếu → làm clip thẳng "
                         "từ ảnh bạn đưa, không tạo ảnh mới."),
    ):
        huong.append(list(dong))
    huong.column_dimensions["A"].width = 26
    huong.column_dimensions["B"].width = 14
    huong.column_dimensions["C"].width = 72
    trang.column_dimensions["B"].width = 52
    trang.column_dimensions["C"].width = 40
    trang.column_dimensions["D"].width = 20

    sach.save(duong_dich)
    return duong_dich
