"""Việc thật của tám khâu — chỗ nối `core/auto.py` vào năng lực sẵn có của tool.

`core/auto.py` chỉ biết thứ tự và trạng thái. Tệp này biết **làm**, và làm bằng
cách gọi lại thứ MyTool đã có chứ không viết mới:

| khâu | dựa vào |
|---|---|
| kịch bản | `core/script_video.py` (lấy tư liệu) + chuỗi 7 lời nhắc của kênh |
| giọng đọc | `client.tts.create` — cùng cửa với tab Voice |
| phụ đề | `core/phu_de.py` — ép khớp, chạy trên máy, miễn phí |
| bảng cảnh | `core/chia_canh.py` — AI chia theo nghĩa, chung với Prompt Visuals |
| ảnh | `client.images.create`, có `nv1.png` làm tham chiếu |
| clip | `client.videos.create`, lấy chính ảnh cảnh đó làm khung đầu |
| ảnh bìa | `client.images.create`, ba kiểu khác nhau |
| dựng | `core/dung_video.py` + FFmpeg — chạy trên máy, miễn phí |

═══ HAI LUẬT CHUNG CHO MỌI KHÂU ═══

**Một, đã có tệp thì không làm lại.** Mỗi khâu nhìn đĩa trước: ảnh cảnh 47 đã
nằm đó thì bỏ qua cảnh 47. Nhờ vậy chạy tiếp không trả tiền hai lần cho cùng
một tấm ảnh, kể cả khi lượt trước chết ở cảnh 118/120.

**Hai, mọi lời gọi tốn tiền đều có `idempotency_key` cố định theo *việc*, không
theo *lần gọi*.** Cảnh 47 của lượt 0001 luôn mang đúng một khoá. Mạng rớt giữa
lúc máy chủ đã nhận thì lần gọi lại rơi vào đúng job cũ chứ không đẻ job mới —
đây là tầng chống trừ tiền hai lần nằm dưới cả luật một.

Việc gọi mạng được **truyền vào** (`goi_chat`, `client`), nên cả tệp này chạy
thử được bằng đồ giả, không tốn đồng nào và không cần mạng.
"""

from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import collections
import threading
import time
from dataclasses import dataclass, field, replace
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

from .auto import LuotChay, TrangThaiKhau
from .chia_canh import (DUOI_CAM, MIN_GIAY_CANH, bang_phu_de, chia_theo_nghia,
                        dien_khuon, khoi_ke_hoach, loi_nhac_chia, nhip_tu_khuon,
                        sach_ke_hoach)
# `loc_json` chuyển sang ở cạnh `goi_van_ban` — nơi nào đòi AI trả JSON cũng
# phải bóc kiểu ấy, kể cả tool `prompt.workbook` ngoài `core/`. Vẫn nhập lại
# vào đây vì `__all__` của tệp này đã hứa có nó.
from .goi_van_ban import loc_json
from .kenh import Kenh, ten_khung, ten_tieng
from .nang_anh import KHUNG
from .ghi_dia import (duong_tam, ghi_chu as _ghi_chu_dia, ghi_json,
                      thay_the)
from .the_cam_xuc import TEP_CO_THE, chen_the, kiem_the
from .tron_tieng import co_ne_giong, loc_tron_nhac
from .su_co import (SUAT_TAI_TEP, LoiNoiDung, LoiTaiVe, goi_kien_nhan,
                    phan_loai, xin_nhip)

__all__ = [
    "BoiCanh", "dung_bo_viec", "chia_doan_doc", "chia_doan_va_nghi",
    "tach_phan", "DAU_NGAT_PHAN", "GIAY_NGHI_PHAN", "GIAY_NGHI_GIUA_KHUC",
    "dem_tien_do",
    "CHU_MOI_LUOT_DOC", "loc_json", "sua_loi_nhac_canh",
]

#: Số ký tự tối đa gửi cho một lượt đọc — **trần cứng của cổng, đã đo**.
#:
#: Cùng con số với tool nội bộ đời trước (`ANON_MAX_CHARS = 1000`, ghi rõ
#: "giới hạn cứng của endpoint, đã đo"). Đằng sau cả hai là một cổng giọng nói,
#: nên trần giống nhau là phải.
#:
#: Bản trước để 2.500 — con số đoán, không đo. Nó không làm hỏng lượt nào chỉ
#: vì `SO_DOAN_DOC` cắt vụn xuống dưới trần trước khi chạm tới; bỏ cái cắt vụn
#: ấy đi mà giữ 2.500 là cổng từ chối ngay.
CHU_MOI_LUOT_DOC = 1000

#: Mấy đoạn đọc cùng lúc **khi chưa hỏi được máy chủ**. Con số thật lấy từ
#: `GET /v1/me` (`concurrent_jobs.tts`), đo được là 3.
SONG_SONG_DOC = 3

#: Số cảnh gửi cho AI trong một lượt viết lời nhắc.
#:
#: `tool-catalog/prompt.workbook` để 20, và tôi chép theo. Lượt chạy thật đầu
#: tiên cho thấy 20 là quá nhiều: mỗi cảnh cần `img_prompt` + `video_prompt`
#: tiếng Anh chi tiết, hai mươi cảnh vượt trần chữ trả về và JSON **đứt giữa
#: câu** — `Unterminated string ... (char 20748)`.
#:
#: Cách chữa **bây giờ nằm ở chỗ khác**, và hằng số cũ ở đây đã chết:
#: `CANH_MOI_LUOT = 8` không còn nơi nào dùng, và ghi chú của nó trỏ tới một
#: hàm `_viet_loi_nhac` **không còn tồn tại**. Đã bỏ cả hai 17/08/2026.
#:
#: Chỗ chia lô thật: `core/chia_canh.CUE_MOI_KHUC` (30 dòng phụ đề một khúc),
#: chạy `KHUC_SONG_SONG = 3` khúc cùng lúc. Giữ đoạn ghi chú phía trên vì bài
#: học đo được của nó vẫn đúng — chỉ con số và tên hàm là cũ.

#: Trần chữ cho lượt viết lời nhắc cảnh. Cao hơn hẳn mặc định vì đây là lời gọi
#: đẻ ra nhiều chữ nhất trong cả dây chuyền.
TOKEN_CANH = 16384

#: Kịch bản được coi là đúng độ dài khi lệch dưới ngần này.
#:
#: Vì sao phải canh: độ dài kịch bản quyết định **mọi thứ phía sau** — thời
#: lượng giọng đọc, số cảnh, số ảnh, số clip. Lệch 14% ở đây là video 8,6 phút
#: thay vì 10 phút, và người xem không nhận ra nhưng thuật toán YouTube thì có.
#:
#: 8% vì đó là ngưỡng còn nắn được bằng một vòng viết lại mà không phải bịa
#: thêm ý; quá đó thì phải viết lại cả đoạn, và bịa thêm ý là chỗ chất lượng
#: tụt.
#:
#: ═══ 0,25 QUÁ RỘNG ĐỂ CÒN GỌI LÀ MỤC TIÊU (sửa 29/08/2026) ═══
#:
#: Bản 2.13.1 nới 0,08 → 0,25 để "thôi gọi API nhiều vô ích" — đúng ý định,
#: nhưng ghi chú ngay trên đây vẫn nói 8%, nên chỗ lệch không ai thấy.
#:
#: Hệ quả đo được trên TL4-T7: kênh nhắm 13 phút, bản viết ra 10,0 phút, tool
#: in "độ dài đạt (lệch 23%)" rồi bỏ qua luôn vòng nắn. 0,25 biến "13 phút"
#: thành "10 tới 16 phút" — rộng tới mức con số mục tiêu không còn nghĩa gì.
#:
#: 0,15 là chỗ đứng giữa: vẫn rộng gần gấp đôi ý định gốc nên không kéo lại
#: cảnh nắn ba vòng, nhưng đủ chặt để một bản hụt 23% bị nắn. Và lời nhắc giờ
#: đo bằng SỐ CÂU (xem prompt/2-viet.md) nên bản viết ra đã sát đích hơn hẳn
#: — phần lớn lượt chỉ tốn thêm đúng một vòng.
#:
#: ═══ ĐÂY LÀ MẶC ĐỊNH, KÊNH ĐƯỢC NỚI RIÊNG (04/09/2026) ═══
#:
#: Mỗi kênh chịu được một mức lệch khác nhau, và đó là quyết định của người
#: làm kênh chứ không phải một hằng số chung. TL4-T7 nới lên 0,30 — chủ dự án:
#: *"về độ dài tao không quá quan trọng trong khoảng từ 10-15 phút"*, và đo
#: bốn lượt thật thì bước viết đã tự về đích không cần ai nắn.
#:
#: Đặt `chenh_cho_phep` trong `kenh.yaml` để nới; bỏ trống thì lấy số này.
CHENH_CHO_PHEP = 0.15


def _chenh_cho_phep(k: Kenh) -> float:
    """Mức lệch độ dài kênh này chịu được — của riêng kênh, hoặc mặc định."""
    rieng = getattr(k, "chenh_cho_phep", 0.0) or 0.0
    return float(rieng) if rieng > 0 else CHENH_CHO_PHEP


#: Nắn nhiều nhất mấy vòng. Ba là đủ: đo thật thì vòng đầu đã kéo được phần
#: lớn khoảng cách, vòng bốn trở đi chỉ đổi chỗ chữ chứ không đổi độ dài.
VONG_NAN_TOI_DA = 3

#: Bao nhiêu cảnh làm cùng lúc ở khâu ảnh và khâu clip.
#:
#: ═══ VÌ SAO PHẢI SONG SONG, VÀ VÌ SAO 12 ═══
#:
#: Bản đầu làm tuần tự: tạo job → **đứng đợi nó xong** → tải về → sang cảnh
#: sau. Đo thật: 50 giây một tấm ảnh, mà gần hết chừng ấy là đứng đợi. 99 cảnh
#: thành 80 phút cho một khâu.
#:
#: Chủ dự án, 14/08/2026: *"1 phút bên shopapi ra được vài nghìn ảnh vài trăm
#: video"*. Đúng — trần 60 lượt/phút là trần **số yêu cầu**, không phải trần
#: số job chạy cùng lúc. Đứng đợi từng cái là tự trói mình.
#:
#: Số luồng **khi chưa hỏi được máy chủ**. Con số thật lấy từ `GET /v1/me`.
#:
#: ⚠ Ghi chú cũ ở đây giải thích rất kỹ vì sao 6 là đúng — và giải thích ấy
#: dựng trên một tiền đề sai: rằng cổng chỉ cho 60 lượt gọi mỗi phút. Hỏi thẳng
#: ngày 15/08/2026 thì cổng cho **600.000 lượt/phút**, **979 job ảnh** và
#: **316 job clip** chạy cùng lúc, hàng chờ 100.000 job, và nói rõ *"gửi nhiều
#: hơn KHÔNG bị từ chối — phần vượt nằm ở hàng chờ và tự chạy khi có chỗ"*.
#:
#: Cái giá: một video mười phút mất 38 phút tạo ảnh và 56 phút tạo clip, trong
#: khi nhà máy làm được hàng chục nghìn ảnh mỗi giờ. Chủ dự án, 15/08/2026:
#: *"hiệu suất nhà máy video và ảnh rất lớn… đôi khi tạo ảnh và video 5 phút là
#: xong"* — đúng, và chỗ chậm là tool chứ không phải cổng.
SONG_SONG_CANH = 6

#: Nhiều nhất bao nhiêu job cùng lúc cho một khâu.
#:
#: ═══ ĐO ĐƯỢC 15/08/2026 — NHIỀU HƠN KHÔNG PHẢI LÚC NÀO CŨNG NHANH HƠN ═══
#:
#: Cổng khai cho **979 job ảnh** cùng lúc, nên bản đầu bắn thẳng cả 117 việc
#: một lượt. Đo trên cùng một lượt, cùng bảng cảnh 114 cảnh:
#:
#:     48 job cùng lúc   ->  khâu ảnh  5,9 phút
#:    117 job cùng lúc   ->  khâu ảnh 22,3 phút   (chậm gấp bốn)
#:
#: Vài cảnh nằm "đang xử lý" tới 12 phút rồi phải đặt lại bằng khoá mới. Cổng
#: khai 979 chỗ nhưng `workers_online` là **1** — con số ấy là chỗ *nhận việc*,
#: không phải chỗ *làm việc*. Nhồi quá thì hàng chờ dài ra chứ sản lượng không
#: tăng, và độ trễ từng cái thì tăng thật.
#:
#: 48 là mức đo được là tốt. Đừng nâng vì thấy cổng khai số lớn — hãy đo lại.
TRAN_LUONG_MAY = 48

#: Bao lâu hỏi cả sổ job một lượt.
#:
#: ═══ 2,0 → 30,0 NGÀY 16/08/2026 ═══
#:
#: Lý lẽ cũ ghi ở đây — "một lượt hỏi là **một** lời gọi cho cả trăm job" — sai
#: ở hai chỗ, và cả hai đều đo được:
#:
#: 1. Không phải một lời gọi. `_mot_luot` lật tới `TRANG_MOI_LUOT` trang cho
#:    MỖI trạng thái trong ("succeeded", "failed"), tức tối đa 5 × 2 = 10 lời
#:    gọi mỗi vòng. Ở nhịp 2 giây là **5 lời gọi/giây**.
#: 2. `GET /v1/jobs?limit=100` không rẻ: nó trả về cả trăm job KÈM toàn bộ
#:    output, tốn gấp ~200 lần `jobs.get(id)` ở phía máy chủ.
#:
#: Đo trên nginx của máy chủ ngày 16/08/2026: 1.212 lượt `GET /v1/jobs` trong 5
#: phút từ đúng máy này. Cùng lúc đó worker chạy trên chính máy ấy tải ảnh tham
#: chiếu từ CDN về chỉ đạt 23 KB/s — 516 lượt tải hết giờ giữa chừng, kéo theo
#: 15–25% job của khách hỏng với câu báo lỗi đổ tại "địa chỉ ảnh của bạn".
#:
#: Nhịp hỏi không làm job xong sớm hơn một giây nào. Nó chỉ giành đường truyền
#: và CPU của đúng cái máy chủ đang kết sổ tiền cho job ta đang chờ.
#:
#: 30 giây là mốc chủ dự án chốt, vì job nhanh nhất của cả ba nhà máy — ảnh —
#: cũng đã 30 giây. Giá phải trả cao nhất là một tấm ảnh xong ở giây 31 nằm chờ
#: thêm một nhịp; đổi lại là 15 lời gọi/phút thay vì 300.
NHIP_HOI_CHUNG = 30.0

#: Mỗi trang lấy nhiều nhất bao nhiêu job, và lật nhiều nhất mấy trang.
#:
#: 114 tấm ảnh xong gần như cùng lúc, mà một trang chỉ chứa 100 — nên phải lật
#: tiếp, nếu không job xong nằm ở trang hai phải đợi lượt hỏi sau. Năm trang là
#: 500 job, rộng hơn hẳn mẻ lớn nhất tool từng chạy.
SO_MOI_TRANG = 100
TRANG_MOI_LUOT = 5

#: Bao lâu thì hỏi riêng một job mà sổ chung chưa thấy tăm hơi.
#:
#: Đây là lưới an toàn cho ba chỗ hụt của sổ chung: job hỏng theo kiểu không
#: nằm trong danh sách nào ta hỏi, job xong từ lâu nên đã trôi khỏi năm trang
#: đầu, và cổng cũ chưa có `GET /v1/jobs`. Thưa (45 giây) vì bình thường nó
#: không bao giờ phải chạy.
NHIP_HOI_RIENG = 45.0

#: Nhớ lại câu trả lời của `GET /v1/me` để không hỏi lại mỗi khâu.
_HAN_MUC: Dict[str, Any] = {}
_KHOA_HAN_MUC = threading.Lock()


def han_muc_may_chu(bc: "BoiCanh") -> Dict[str, Any]:
    """Hỏi cổng xem nó cho chạy bao nhiêu. Hỏi hỏng thì trả về `{}`.

    Hỏi **một lần cho cả lượt chạy**: con số này đổi theo tải nhà máy, nhưng
    không đổi từng phút, và hỏi lại ở mỗi khâu chỉ tốn thêm lượt gọi.
    """
    with _KHOA_HAN_MUC:
        if _HAN_MUC:
            return _HAN_MUC
        try:
            tra = bc.client.request("GET", "/v1/me")
            goi = tra.to_dict() if hasattr(tra, "to_dict") else dict(tra)
            _HAN_MUC.update(goi.get("limits") or {})
        except Exception:  # noqa: BLE001 — hỏi không được thì dùng số an toàn
            pass
        return _HAN_MUC


def _so_luong(bc: "BoiCanh", loai: str, mac_dinh: int = SONG_SONG_CANH,
              can: int = 0) -> int:
    """Bao nhiêu luồng cho khâu này — theo con số máy chủ tự khai.

    `can` là số việc thật sự phải làm: mở 128 luồng cho 3 tấm ảnh bìa chỉ tổ
    tốn chỗ. `mac_dinh` là đường lui khi không hỏi được máy chủ.
    """
    from .su_co import dat_tran_moi_phut  # noqa: PLC0415

    han = han_muc_may_chu(bc)
    if not han:
        ra = mac_dinh
    else:
        dat_tran_moi_phut(han.get("requests_per_minute") or 0)
        try:
            cua_loai = int((han.get("concurrent_jobs") or {}).get(loai) or 0)
        except (TypeError, ValueError):
            cua_loai = 0
        ra = min(cua_loai, TRAN_LUONG_MAY) if cua_loai > 0 else mac_dinh
    if can > 0:
        ra = min(ra, can)
    return max(1, ra)


@dataclass
class BoiCanh:
    """Mọi thứ một khâu cần để làm việc."""

    goc: str
    kenh: Kenh
    #: Gọi AI viết chữ. `(lời nhắc, mô hình=…, khoa=…) -> chữ trả về`.
    #:
    #: `khoa` là Idempotency-Key **cố định theo bước**, không phải theo lần gọi.
    #: Xem ghi chú ở `_goi` trong tệp này — đây là chỗ đã làm hỏng lượt chạy
    #: thật đầu tiên (14/08/2026).
    goi_chat: Callable[..., str]
    #: Gọi AI viết chữ **riêng cho khâu kịch bản** — để trống thì dùng chung
    #: `goi_chat`. Có nó khi chủ máy bật "Kịch bản viết bằng Claude Code" trong
    #: Cài đặt: khâu kịch bản đi qua thuê bao Claude của máy, còn lời nhắc
    #: ảnh/clip và mọi khâu khác vẫn đi ví ShopAPI. Xem `core/viet_max.py`.
    goi_chat_kich_ban: Optional[Callable[..., str]] = None
    #: Client ShopAPI cho giọng đọc / ảnh / clip. Có thể là đồ giả khi chạy thử.
    client: Any = None
    on_log: Optional[Callable[[str], None]] = None
    cancel: Optional[threading.Event] = None
    #: Lấy lời thoại video đối thủ. Tách ra để test không cần mạng.
    lay_tu_lieu: Optional[Callable[..., Any]] = None
    #: Bộ nghe cho khâu phụ đề. Để trống thì dùng `faster-whisper` trên máy.
    #: Tách ra vì hai lý do, không chỉ để test: bộ nghe là thứ dễ đổi nhất
    #: trong cả dây chuyền (bản model mới, hay máy khách không cài được), và
    #: chạy thử cả luồng bằng tiếng giả thì whisper thật sẽ không nghe ra chữ
    #: nào — đúng như nó nên thế.
    nghe: Optional[Callable[..., Any]] = None
    ffmpeg: str = ""
    #: Hàm ngủ. Tách ra để bài kiểm chạy được các nhịp chờ dài mà không phải
    #: ngồi đợi thật mười sáu phút.
    ngu: Callable[[float], None] = time.sleep
    #: Mấy giây một lượt hỏi cả sổ job (`SoTheoDoi`). Tách ra cùng lý do với
    #: `ngu`: bài kiểm không phải ngồi đợi hai giây cho mỗi nhịp giả.
    nhip_hoi: float = NHIP_HOI_CHUNG
    #: Chốt "cổng ngừng nhận loại việc này". Một luồng gạt, mọi luồng dừng.
    nha_may_tat: Optional[threading.Event] = None
    #: "Trạng thái vừa đổi **giữa lúc một khâu còn đang chạy** — ghi ra đĩa và
    #: vẽ lại đi."
    #:
    #: `core/auto.chay` chỉ báo đổi ở đầu và cuối mỗi khâu. Với khâu 99 cảnh thì
    #: giữa hai mốc ấy là bốn mươi phút chỉ hiện đúng hai chữ "ĐANG CHẠY", và
    #: không cách nào phân biệt "đang làm cảnh 87" với "tool treo".
    on_nhip: Optional[Callable[[LuotChay], None]] = None
    #: Tải một ảnh về theo URL, trả bytes. Tách ra để bài kiểm đọc ảnh bìa đối
    #: thủ không phải chạm mạng. Để trống thì dùng `_tai_anh_thumb` (urllib).
    tai_anh: Optional[Callable[[str], bytes]] = None

    def ghi(self, dong: str) -> None:
        if self.on_log is not None:
            self.on_log(dong)

    def nhip(self, luot: LuotChay) -> None:
        if self.on_nhip is not None:
            self.on_nhip(luot)

    def cho_kich_ban(self) -> "BoiCanh":
        """Bối cảnh dành cho khâu kịch bản: đổi đường viết chữ nếu có đường riêng.

        Trả về **bản sao** chứ không sửa tại chỗ: cùng một `BoiCanh` còn được
        các khâu khác dùng song song, mà lời nhắc ảnh/clip thì phải tiếp tục đi
        ví ShopAPI.
        """
        if self.goi_chat_kich_ban is None:
            return self
        return replace(self, goi_chat=self.goi_chat_kich_ban)

    def kiem_dung(self) -> None:
        """Có phải dừng không — vì người bấm Dừng, HOẶC vì cổng đã ngừng nhận.

        Gộp hai lý do vào một chỗ là cố ý. Chốt nhà máy mà chỉ kiểm ở lúc **bắt
        đầu** một mục thì luồng nào đang ngủ giữa nhịp chờ 300 giây vẫn ngủ đủ
        300 giây — mười hai luồng như vậy là cả mẻ treo mười sáu phút cho một
        thứ đã biết chắc không xong. Đặt ở đây thì mọi vòng chờ đều tỉnh dậy
        trong nửa giây.
        """
        from .auto import Cancelled  # noqa: PLC0415 — tránh vòng nhập

        if self.cancel is not None and self.cancel.is_set():
            raise Cancelled()
        if self.nha_may_tat is not None and self.nha_may_tat.is_set():
            raise Cancelled()


# ── Tiện ích chung ───────────────────────────────────────────────────────────


def _goi(bc: "BoiCanh", loi_nhac: str, khoa: str,
         toi_da_token: int = 8192, anh: str = "") -> str:
    """Gọi AI, kèm Idempotency-Key **cố định theo bước**.

    ═══ VÌ SAO CẦN KHOÁ CỐ ĐỊNH, VÀ VÌ SAO KHÔNG ĐƯỢC GỬI LẠI ═══

    Lượt chạy thật đầu tiên chết ở đây, và đáng ghi lại vì nó ngược với trực
    giác. Viết kịch bản là lời gọi **dài**: 3.410 ký tự tiếng Nhật dựng từ bản
    gỡ băng 1.924 chữ mất vài phút. Client thì chỉ đợi 60 giây rồi bỏ cuộc.

    Bỏ cuộc **không** có nghĩa máy chủ đã dừng — nó vẫn đang viết. Nên lần gọi
    sau nhận đúng câu:

        Yêu cầu với Idempotency-Key này đang được xử lý. Vui lòng đợi vài
        giây rồi kiểm tra lại kết quả, đừng gửi lại.

    Câu đó là **lời chỉ dẫn**, không phải lỗi. Đường đúng là đợi rồi hỏi lại
    **cùng một khoá** — máy chủ trả bản đã viết xong, và chỉ trừ tiền một lần.
    Đường sai là sinh khoá mới gửi lại: vừa bỏ mất bài đang viết, vừa trả tiền
    lần hai cho cùng một việc.

    Khoá ở đây gắn với `(mã lượt, tên bước)`, nên chạy lại lượt cũ cũng rơi vào
    đúng kết quả cũ chứ không đẻ ra lượt tính tiền mới.
    """
    # ═══ LỜI GỌI AI CŨNG PHẢI QUA VAN NHỊP ═══
    #
    # Tôi để sót đúng chỗ này: van 48 lượt/phút áp cho tạo job, tải tệp và hỏi
    # job — nhưng **không** áp cho lời gọi viết chữ. Bình thường không sao vì
    # các bước viết chạy tuần tự.
    #
    # Khâu chia cảnh thì khác: ba khúc bắn cùng lúc. Đo được: cùng lời nhắc ấy
    # gửi **một mình** xong trong 46 giây và trả về 7 cảnh đúng; gửi ba cái một
    # lượt là ăn "gửi quá nhanh", rồi mỗi lần hỏi lại nhận "đang được xử lý" —
    # kẹt hơn tám phút mà nhìn vào tưởng máy chủ chậm.
    #
    # Một lời gọi viết chữ nặng hơn một lời gọi tạo job nhiều, nên tính hai
    # suất: nó chiếm cổng lâu hơn hẳn.
    # ═══ "CHƯA NHẬN ĐƯỢC YÊU CẦU" THÌ PHẢI ĐỔI KHOÁ, KHÔNG PHẢI ĐỢI ═══
    #
    # Đây là cái bẫy tốn nhiều thời gian nhất của cả ngày, và nó ngược hẳn với
    # trực giác. Hai câu của máy chủ nghe giống nhau nhưng nghĩa trái ngược:
    #
    #   "tạm gián đoạn, CHƯA NHẬN ĐƯỢC yêu cầu, KHÔNG bị trừ tiền"
    #        -> chưa có gì tồn tại. Đợi là đợi một thứ không có.
    #   "Idempotency-Key này ĐANG ĐƯỢC XỬ LÝ"
    #        -> đang có thật. Đợi là đúng.
    #
    # Chỗ chết người: sau câu thứ nhất, hỏi lại **cùng khoá** thì máy chủ ghi
    # nhận cái khoá ấy, và từ đó mọi lần hỏi tiếp đều trả câu thứ hai — kẹt
    # vĩnh viễn. Đúng chuỗi mà máy khách gặp ở khâu "đọc lại lần cuối":
    # trục trặc → trục trặc → đang xử lý → đang xử lý → …
    #
    # Nên: gặp "chưa nhận được" thì **đổi khoá**. An toàn tuyệt đối, vì chính
    # máy chủ đã nói chưa trừ tiền và chưa tạo việc gì.
    # ═══ ĐỔI KHOÁ RỒI THÌ PHẢI ĐỢI, KHÔNG ĐƯỢC HỎI NGAY ═══
    #
    # Bản trước gọi lại **tức thì** sau mỗi lần hỏng. `xin_nhip` ở đầu vòng lặp
    # trông như một nhịp nghỉ nhưng không phải: nó là bộ giữ **hạn mức gọi mỗi
    # phút**, chưa chạm trần thì trả về ngay.
    #
    # Đo được trên lượt chạy thật 17/08/2026, khâu cắt cảnh, đúng lúc máy chủ
    # báo *"Hệ thống đang quá tải… vui lòng thử lại sau ít phút"*:
    #
    #     [5313s] thử lại lần 1, 2, 3      ← cùng một giây
    #     [5314s] thử lại lần 1, 2, 3      ← cùng một giây
    #     [5315s] thử lại lần 1, 2
    #
    # Khoảng **mười lăm lời gọi trong hai giây**, ném vào một máy chủ vừa nói
    # nó đang quá tải. Đó đúng là thứ `CLAUDE.md` cấm: hỏi dày không làm việc
    # xong sớm hơn một giây nào, nó chỉ lấy thêm CPU của chính máy chủ đang
    # nghẹt — và làm cả bốn lần thử đều hỏng vì cùng một lý do.
    #
    # `core/su_co.py` vốn đã có bảng nhịp lùi tính sẵn cho từng loại sự cố.
    # Chỗ này chỉ việc dùng nó.
    from .su_co import (MAT_MANG, TAM_NGHI, dau_vet, nhip_cho,  # noqa: PLC0415
                        phan_loai as _phan)

    for lan in range(4):
        xin_nhip(bc.on_log, ngu=bc.ngu, so_suat=2)
        khoa_lan = khoa if lan == 0 else "{0}:r{1}".format(khoa, lan)
        try:
            # `anh` (đọc chữ trên ảnh bìa đối thủ) chỉ truyền khi thật có ảnh —
            # `goi_chat` cũ không nhận kwarg này, nên không đưa vào lúc viết chữ
            # thường để khỏi phá các nơi gọi khác.
            them = {"anh": anh} if anh else {}
            ket = bc.goi_chat(loi_nhac, mo_hinh=bc.kenh.mo_hinh,
                              khoa=khoa_lan, toi_da_token=toi_da_token, **them)
            # ═══ BÓC LỚP VỎ "GỌI CÔNG CỤ" GIẢ ═══
            #
            # Có lượt mô hình không trả thẳng lời đọc, mà "diễn" một pha ghi tệp
            # (```bash … / name write_file / {"content": "…"} / </function…>).
            # Lời thật nằm trong trường JSON "content"; nếu để nguyên thì bộ đọc
            # giọng đọc cả vỏ ("んてんてん…"), và bước đo độ dài đếm cả rác. Bóc
            # ngay tại cửa AI để MỌI bước hưởng, và chỉ bóc khi chắc là vỏ giả.
            from .lam_sach import go_boc_tool_gia  # noqa: PLC0415
            return go_boc_tool_gia(ket) if isinstance(ket, str) else ket
        except Exception as loi:  # noqa: BLE001
            if lan == 3:
                raise
            loai = _phan(loi)
            # Ghi kèm `request_id` — không có nó thì bên vận hành không tra
            # được, và cả lượt chạy hỏng thành một câu than vô ích.
            bc.ghi("  {0} — làm lại bằng khoá mới (lần {1}).{2}".format(
                "máy chủ chưa nhận được yêu cầu"
                if loai == TAM_NGHI else str(loi)[:60], lan + 1,
                dau_vet(loi)))
            cho = nhip_cho(loai, lan)
            if cho > 0:
                bc.ghi("    đợi {0:.0f} giây cho máy chủ thở.".format(cho))
                bc.ngu(cho)
                bc.kiem_dung()
    raise RuntimeError("không gọi được AI sau 4 lần đổi khoá")


#: Điền `<<TÊN>>` trong lời nhắc.
#:
#: Việc này nằm ở `core/chia_canh.py` vì lời nhắc chia cảnh cũng cần nó, và
#: `chia_canh` không được nhập ngược lại tệp này (vòng nhập). Một bản duy nhất,
#: hai nơi gọi — tên cũ giữ nguyên để tám khâu bên dưới không phải sửa.
_thay = dien_khuon


#: Chỗ cắt tốt dần từ trên xuống. Mỗi mục là (mẫu tìm, số ký tự ăn thêm).
#:
#: Thứ tự này là thứ tự **chỗ nghỉ tự nhiên của người đọc**: hết đoạn văn nghỉ
#: dài nhất, hết câu nghỉ vừa, giữa câu nghỉ ngắn. Cắt đúng chỗ người ta vốn đã
#: nghỉ thì chỗ nối không nghe ra; cắt giữa câu thì nghe rõ một nhịp hụt.
_CHO_CAT = (
    ("\n\n", 2),   # hết đoạn văn — tự nhiên nhất
    ("\n", 1),     # hết dòng
)
#: Hết câu: dấu chấm/hỏi/than rồi tới khoảng trắng.
_HET_CAU = re.compile(r"[.!?。．！？…][\"'”’)\]]?\s")
#: Giữa câu — dùng khi cả một khúc dài không có lấy một dấu chấm.
_GIUA_CAU = ("; ", ", ", ": ", "；", "，", "、")


def _cho_cat_tot_nhat(chu: str, tran: int) -> int:
    """Vị trí cắt tốt nhất trong `chu[:tran]`, tính theo thứ tự `_CHO_CAT`.

    Luôn lấy chỗ **gần `tran` nhất** trong hạng tốt nhất tìm được: cắt sớm là
    thừa ra một đoạn nữa, mà mỗi đoạn thừa là một chỗ đổi tông.

    Sàn 30%: chỗ cắt quá gần đầu thì thà xuống hạng kém hơn mà lấy được đoạn
    dài, còn hơn sinh ra một mẩu vài chục chữ.
    """
    cua_so = chu[:tran]
    san = tran * 0.3

    for dau, an_them in _CHO_CAT:
        vi_tri = cua_so.rfind(dau)
        if vi_tri > san:
            return vi_tri + an_them

    khop = list(_HET_CAU.finditer(cua_so))
    if khop and khop[-1].end() > san:
        return khop[-1].end()

    for dau in _GIUA_CAU:
        vi_tri = cua_so.rfind(dau)
        if vi_tri > san:
            return vi_tri + len(dau)

    vi_tri = cua_so.rfind(" ")
    if vi_tri > san:
        return _ne_giua_the(chu, vi_tri + 1)

    # Không có lấy một khoảng trắng — cắt cứng. Gần như không xảy ra với chữ
    # tiếng Việt, nhưng thiếu nhánh này thì hàm quay vòng vô tận.
    return _ne_giua_the(chu, tran)


def _ne_giua_the(chu: str, cat: int) -> int:
    """Kéo chỗ cắt lùi lại nếu nó rơi vào giữa một thẻ cảm xúc.

    ═══ VÌ SAO CẦN ═══

    Thẻ cảm xúc v3 có loại **chứa khoảng trắng**: `[laughs harder]`,
    `[short pause]`, `[clears throat]`. Mà hàm cắt ở trên được phép cắt tại
    khoảng trắng — nên nó có thể cắt đúng giữa thẻ, thành `…[laughs` ở đoạn này
    và `harder]…` ở đoạn sau.

    Hai mảnh ấy không còn là thẻ nữa. Model không hiểu, và cái nó làm với chữ
    không hiểu thì tài liệu **không nói** — có thể bỏ qua, mà cũng có thể đọc
    to lên giữa video.

    Kéo lùi về trước dấu `[` là xong: thẻ đi trọn vẹn sang đoạn sau, còn đoạn
    này kết thúc sớm hơn vài chữ.
    """
    mo = chu.rfind("[", 0, cat)
    if mo < 0:
        return cat
    dong = chu.find("]", mo)
    if dong < 0 or dong < cat:
        return cat              # thẻ đã đóng trước chỗ cắt — không sao
    # Lùi về trước dấu `[`. Nếu lùi tới 0 thì cả đoạn chỉ có mỗi cái thẻ dở
    # dang — giữ nguyên chỗ cắt cũ, không thì hàm gọi quay vòng vô tận.
    return mo if mo > 0 else cat


#: Dòng đánh dấu **hết một phần** trong kịch bản đọc.
#:
#: Ba gạch đứng một mình trên một dòng — thứ người không biết lập trình gõ
#: được, và là dấu ngắt quen thuộc trong văn bản. Bước rà soát
#: (`prompt/3-sua.md`) được dặn đặt dấu này ở ranh giới các phần.
DAU_NGAT_PHAN = "---"

#: Nghỉ mấy giây ở chỗ cắt VÌ QUÁ DÀI (giữa một phần), để che chỗ đổi tông.
#:
#: ═══ VÌ SAO CHỖ NÀY CŨNG PHẢI CÓ MỘT NHỊP ═══
#:
#: Chủ dự án, 27/08/2026: *"voice mỗi lần là 1 tông giọng nên việc nghỉ vậy ở
#: 1 phần đã xong sẽ giúp khán giả không nhận ra việc thay đổi tông"*. Đúng —
#: và nó cũng đúng cho chỗ cắt vì trần 1.000 chữ: đó cũng là một lượt gọi mới,
#: cũng đổi tông, chỉ khác là nó rơi giữa một mạch đang kể.
#:
#: Nên chỗ ấy vẫn cần một nhịp, nhưng phải NGẮN hơn hẳn nhịp giữa hai phần:
#: 0,35 giây nghe như một hơi lấy đà tự nhiên và che được chỗ ghép, trong khi
#: 1,2 giây ở giữa câu chuyện thì nghe như đứt băng.
#:
#: Cách chắc nhất vẫn là **đừng để phải cắt**: kịch bản 15 phút khoảng 4.700
#: chữ chia 5–9 phần thì mỗi phần ~600 chữ, dưới trần 1.000 — mọi chỗ đổi tông
#: rơi đúng vào nhịp nghỉ giữa hai phần. Đó là lý do `prompt/3-sua.md` dặn đặt
#: 5–9 dấu `---`.
GIAY_NGHI_GIUA_KHUC = 0.35

#: Nghỉ mấy giây giữa hai phần khi kênh không khai `giay_nghi_phan`.
#:
#: ═══ VÌ SAO LÀ KHOẢNG LẶNG THẬT, KHÔNG PHẢI THẺ `[long pause]` ═══
#:
#: Chủ dự án, 27/08/2026: *"mỗi phần đó có nhịp nghỉ… để khán giả được chuyển
#: mình giữa các phần, với lại đoạn nghỉ đó khi edit nhìn vào thấy nó không có
#: âm thanh, dễ edit các phần"*.
#:
#: Thẻ thì nhà máy giọng nói lúc nghe lúc không — đo trên lượt 0053: mọi thẻ
#: đều bị nuốt, quãng nghỉ ra 1,2–1,8 giây tuỳ chỗ, không điều khiển được.
#: Khoảng lặng chèn lúc ghép thì đúng từng phần mười giây, nhìn thấy trên sóng
#: âm khi dựng, và không tốn một lượt gọi nào.
#:
#: 1,2 giây: đủ để người xem chuyển mình, đủ để người dựng thấy quãng trắng rõ
#: ràng, mà mười phần cũng chỉ cộng thêm 12 giây vào cả video.
GIAY_NGHI_PHAN = 1.2


def _bo_dau_ngat(chu: str) -> str:
    """Bỏ các dòng đánh dấu ngắt phần — chúng không bao giờ được đọc lên."""
    ra = []
    for dong in (chu or "").splitlines():
        g = dong.strip()
        if g and len(g) >= 3 and set(g) <= set("-_*"):
            continue
        ra.append(dong)
    return "\n".join(ra)


def tach_phan(kich_ban: str) -> List[str]:
    """Cắt kịch bản tại các dòng `---`. Không có dấu nào thì trả về một phần."""
    phan: List[str] = []
    dang: List[str] = []
    for dong in (kich_ban or "").splitlines():
        g = dong.strip()
        if g and len(g) >= 3 and set(g) <= set("-_*"):
            phan.append("\n".join(dang))
            dang = []
            continue
        dang.append(dong)
    phan.append("\n".join(dang))
    return [m for m in (x.strip() for x in phan) if m]


def chia_doan_va_nghi(kich_ban: str, tran: int = CHU_MOI_LUOT_DOC,
                      giay_nghi: float = GIAY_NGHI_PHAN,
                      ) -> Tuple[List[str], List[float]]:
    """Chia kịch bản thành đoạn đọc, kèm **số giây nghỉ sau mỗi đoạn**.

    Nghỉ chỉ rơi vào chỗ hết một phần (dòng `---`), không rơi vào chỗ cắt vì
    quá dài — chỗ ấy nằm giữa một ý, chèn khoảng lặng vào nghe như đứt băng.

    `giay_nghi <= 0` thì mọi ô nghỉ đều 0: kênh không muốn nhịp nghỉ vẫn chạy
    y như trước.
    """
    phan = tach_phan(kich_ban)
    doan: List[str] = []
    nghi: List[float] = []
    for i, mot in enumerate(phan):
        cac = chia_doan_doc(mot, tran)
        for j, x in enumerate(cac):
            doan.append(x)
            cuoi_bai = (i == len(phan) - 1) and (j == len(cac) - 1)
            het_phan = (j == len(cac) - 1)
            if cuoi_bai or giay_nghi <= 0:
                nghi.append(0.0)
            elif het_phan:
                nghi.append(float(giay_nghi))
            else:
                # Cắt vì quá trần: vẫn là một lượt đọc mới, vẫn đổi tông — che
                # bằng một nhịp ngắn. Xem `GIAY_NGHI_GIUA_KHUC`.
                nghi.append(min(float(giay_nghi), GIAY_NGHI_GIUA_KHUC))
    return doan, nghi


def chia_doan_doc(kich_ban: str, tran: int = CHU_MOI_LUOT_DOC) -> List[str]:
    """Cắt kịch bản thành các đoạn vừa một lượt đọc.

    ═══ ÍT ĐOẠN NHẤT CÓ THỂ, VÀ CÁC ĐOẠN ĐỀU NHAU ═══

    Mỗi đoạn là **một lượt gọi riêng tới nhà máy giọng nói**, và nhà máy không
    nhớ nó vừa đọc gì ở lượt trước. Nên mỗi chỗ nối giữa hai đoạn là một chỗ
    **tông giọng đổi** — nghe ra được, và càng nhiều chỗ nối thì càng lộ.

    Vì vậy luật ở đây là: **ít đoạn nhất có thể**. Số đoạn ít nhất là
    `⌈độ dài ÷ trần⌉`, không cách nào ít hơn — trần là giới hạn cứng của cổng.

    Bản trước làm ngược lại: nó **cố tình cắt vụn** ra mười đoạn để ba suất
    chạy song song của cổng có việc mà làm. Nhanh hơn thật, nhưng đổi bằng đúng
    thứ người xem nghe thấy. Chủ dự án chỉ ra 16/08/2026 khi thấy kịch bản
    2.726 chữ bị chia **tám** đoạn ~390 chữ, trong khi ba đoạn ~909 chữ là đủ.

    Mà thực ra cũng chẳng mất mấy phần nhanh: kịch bản mười phút dài khoảng
    15.000 chữ, chia theo trần 1.000 vẫn ra mười lăm đoạn — thừa việc cho cả ba
    suất. Chỉ những kịch bản ngắn mới ra ít hơn ba đoạn, và với chúng thì chia
    vụn cũng chẳng nhanh thêm được bao nhiêu.

    Chia **đều** chứ không nhồi đầy từng đoạn: 2.726 chữ nhồi đầy ra
    1.000+1.000+726, chia đều ra 909+909+908. Cùng ba đoạn, nhưng đoạn đều nhau
    thì tông giọng giữa chúng cũng gần nhau hơn.
    """
    tho = _bo_dau_ngat(kich_ban or "").strip()
    if not tho:
        return []
    tran = max(1, int(tran))

    ra: List[str] = []
    con_lai = tho
    while con_lai:
        con_lai = con_lai.strip()
        if not con_lai:
            break
        if len(con_lai) <= tran:
            ra.append(con_lai)
            break
        # Tính lại mỗi vòng: đoạn vừa cắt ra ngắn hơn dự tính thì phần còn lại
        # tự san lại cho đều, khỏi dồn hết chỗ hụt vào đoạn cuối.
        so_doan_con = (len(con_lai) + tran - 1) // tran
        deu = (len(con_lai) + so_doan_con - 1) // so_doan_con
        cat = _cho_cat_tot_nhat(con_lai, min(tran, deu))
        khuc = con_lai[:cat].strip()
        if khuc:
            ra.append(khuc)
        con_lai = con_lai[cat:]
    return [m for m in ra if m]


def _ghi_chu(duong: str, chu: str) -> None:
    _ghi_chu_dia(duong, chu)


def _doc_chu(duong: str) -> str:
    try:
        with open(duong, "r", encoding="utf-8") as tep:
            return tep.read()
    except OSError:
        return ""


# ═══ NHỚ TIÊU ĐỀ + MÃ VIDEO ĐỐI THỦ ĐỂ CHẠY LẠI VẪN CÒN ═══
#
# Chạy lại một lượt thì bước lấy lời thoại bị bỏ qua (đã có `0-tu-lieu.txt`),
# nên `ket.title`/`ket.video_id` không còn. Kênh `nguyen_goc` cần cả hai để
# lấy nguyên tiêu đề và dựng địa chỉ ảnh bìa. Ghi ra một tệp bên cạnh lời thoại.
def _ghi_doi_thu(d: str, tieu_de: str, video_id: str, duration_s: int = 0) -> None:
    _ghi_chu(os.path.join(d, "0-doi-thu.txt"),
             "TITLE: {0}\nVIDEO_ID: {1}\nDURATION: {2}\n".format(
                 tieu_de or "", video_id or "", int(duration_s or 0)))


def _doc_doi_thu(d: str) -> Dict[str, str]:
    ra = {"title": "", "video_id": "", "duration": "0"}
    for dong in _doc_chu(os.path.join(d, "0-doi-thu.txt")).splitlines():
        if dong.startswith("TITLE:"):
            ra["title"] = dong[len("TITLE:"):].strip()
        elif dong.startswith("VIDEO_ID:"):
            ra["video_id"] = dong[len("VIDEO_ID:"):].strip()
        elif dong.startswith("DURATION:"):
            ra["duration"] = dong[len("DURATION:"):].strip() or "0"
    return ra


#: Ngân sách token đầu ra cho bước viết: sàn cũ 8.192 (≈ 20 phút đọc tiếng
#: Việt) và trần 32.768. Truyện dài theo nguồn (chủ dự án 26/08/2026: "đừng
#: giới hạn, nó do nguồn đầu vào") mà giữ 8.192 là bài bị cụt ở phút 20.
TOKEN_VIET_SAN = 8192
TOKEN_VIET_TRAN = 32768


def _token_viet(so_ky_tu_goc: int, muc_tieu: int = 0) -> int:
    """Token đầu ra cho một lượt viết/sửa kịch bản, theo độ dài nguồn hoặc mục tiêu.

    Tiếng Việt ≈ 2 ký tự/token (đo thô, có dấu); nhân 1,3 cho dư. Lấy cái lớn
    hơn giữa nguồn và mục tiêu, kẹp trong [sàn, trần].
    """
    can = max(int(so_ky_tu_goc or 0), int(muc_tieu or 0))
    uoc = int(can * 1.3 / 2)
    return max(TOKEN_VIET_SAN, min(TOKEN_VIET_TRAN, uoc))


def _muc_tieu_do_dai(k: "Kenh", tu_lieu: str, giay_goc: int) -> int:
    """Số ký tự nhắm tới cho bước viết.

    Kênh thường: mốc phút cố định (`ky_tu_muc_tieu`). Kênh remake bám gốc
    (`do_dai_theo_goc`): đo theo SỐ GIÂY video đối thủ × `ky_tu_moi_phut` — thước
    không phụ thuộc ngôn ngữ, tránh chuyện bản gỡ băng là bản dịch dài gấp đôi.
    Thiếu số giây thì lui về số ký tự tư liệu; thiếu cả tư liệu (quên đưa link)
    thì về mốc phút cho an toàn — lấy 0 làm mục tiêu sẽ tắt luôn sàn chống "kịch
    bản quá ngắn" trong `_kiem_kich_ban_dung_duoc`.
    """
    if getattr(k, "do_dai_tu_do", False):
        return 0  # độ dài tự do: không mục tiêu, không nắn, không chấm lệch
    if not (k.do_dai_theo_goc and tu_lieu):
        return k.ky_tu_muc_tieu
    if giay_goc > 0 and k.ky_tu_moi_phut > 0:
        return int(round(giay_goc / 60 * k.ky_tu_moi_phut))
    return len(tu_lieu)


def _tai_anh_thumb(url: str) -> bytes:
    """Tải một ảnh về theo URL, trả bytes. Mặc định của `BoiCanh.tai_anh`."""
    from urllib.request import Request, urlopen  # noqa: PLC0415

    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=30) as ph:  # noqa: S310 — URL cố định của YouTube
        return ph.read()


def _anh_thanh_data_url(byte: bytes, kieu: str = "image/jpeg") -> str:
    import base64  # noqa: PLC0415

    return "data:{0};base64,{1}".format(
        kieu, base64.b64encode(byte).decode("ascii"))


#: Lời nhắc đọc chữ trên ảnh bìa — ngắn, chỉ xin đúng chữ nhìn thấy.
_LOI_NHAC_DOC_BIA = (
    "Đây là ảnh bìa (thumbnail) một video. Hãy đọc và trả về ĐÚNG dòng chữ lớn "
    "in trên ảnh, y nguyên từng chữ, không thêm giải thích, không dịch, không "
    "thêm dấu ngoặc. Nếu ảnh không có chữ thì trả về một dòng trống."
)

#: Chữ bìa dài hơn ngần này thì gần như chắc là AI đã tả ảnh / giải thích thay
#: vì đọc dòng chữ — chữ bìa thật chỉ vài từ. Dài quá thì bỏ, lấy tiêu đề thay.
_TOI_DA_CHU_BIA = 120

#: Dấu hiệu mô hình **không thấy ảnh** (câu từ chối), không phải chữ trên bìa.
#:
#: Đo lượt chạy thật 22/08/2026: khi ảnh gửi sai định dạng, cổng lặng lẽ bỏ phần
#: ảnh và mô hình trả "I don't see any image attached…" — đúng 111 chữ, LỌT qua
#: rào 120 chữ ở trên. Nên bắt thêm bằng câu từ chối: thấy là bỏ, lấy tiêu đề.
#: Đã sửa định dạng ảnh (xem `goi_van_ban.khoi_anh`) nên đường này hiếm khi chạm
#: tới, nhưng giữ làm chốt chặn — chữ bìa sai còn hại hơn không có.
_DAU_TU_CHOI_BIA = (
    "i don't see", "i do not see", "don't see any image", "no image",
    "cannot see", "can't see", "unable to see", "please share",
    "please provide", "chưa thấy ảnh", "không thấy ảnh", "không có ảnh",
)


def _giong_tu_choi(chu: str) -> bool:
    """Câu này có phải mô hình đang nói 'không thấy ảnh' không?"""
    thap = chu.lower()
    return any(dau in thap for dau in _DAU_TU_CHOI_BIA)


def _doc_chu_bia_doi_thu(bc: "BoiCanh", luot: LuotChay, video_id: str) -> str:
    """Đọc chữ trên ảnh bìa đối thủ. Hỏng thì trả "" để nơi gọi lấy đường lui.

    Không bao giờ ném lỗi ra ngoài: cổng có thể không nhận ảnh, ảnh có thể không
    tải được — chuyện ấy không đáng làm vỡ cả lượt chạy. Nơi gọi thấy "" thì
    lấy tiêu đề đối thủ làm chữ bìa.
    """
    vid = str(video_id or "").strip()
    if not vid:
        return ""
    tai = bc.tai_anh or _tai_anh_thumb
    byte = b""
    for ten in ("maxresdefault.jpg", "hqdefault.jpg"):
        url = "https://i.ytimg.com/vi/{0}/{1}".format(vid, ten)
        try:
            byte = tai(url)
        except Exception as loi:  # noqa: BLE001
            bc.ghi("  (không tải được ảnh bìa {0}: {1})".format(ten, loi))
            byte = b""
        if byte:
            break
    if not byte:
        return ""
    try:
        tra = _goi(bc, _LOI_NHAC_DOC_BIA, _khoa_chat(luot, "thumb-ocr"),
                   anh=_anh_thanh_data_url(byte))
    except Exception as loi:  # noqa: BLE001
        bc.ghi("  (không đọc được chữ trên ảnh bìa đối thủ: {0})".format(loi))
        return ""
    chu = " ".join((tra or "").split())
    # Cổng nhận ảnh nhưng mô hình "kể" thay vì "đọc" thì trả về cả đoạn — dài
    # bất thường so với một dòng chữ bìa. Không tin, để nơi gọi lấy tiêu đề.
    if len(chu) > _TOI_DA_CHU_BIA:
        bc.ghi("  (chữ đọc từ ảnh bìa dài bất thường — bỏ, dùng tiêu đề đối thủ)")
        return ""
    # Câu từ chối "không thấy ảnh" ngắn hơn rào dài ở trên nên lọt qua — bắt
    # riêng, kẻo ghi nguyên câu tiếng Anh ấy vào chữ bìa.
    if _giong_tu_choi(chu):
        bc.ghi("  (mô hình báo không thấy ảnh bìa — bỏ, dùng tiêu đề đối thủ)")
        return ""
    return chu


def _giay_srt(moc: Any) -> float:
    """Đổi `00:01:23,450` thành số giây. Không đọc được thì trả 0."""
    try:
        gio, phut, con = str(moc).replace(".", ",").split(":")
        giay, mili = con.split(",")
        return int(gio) * 3600 + int(phut) * 60 + int(giay) + int(mili) / 1000.0
    except Exception:  # noqa: BLE001
        return 0.0


def _dai_clip(ffmpeg: str, duong: str) -> float:
    """Độ dài một clip, tính bằng giây. Không đo được thì trả `0.0`.

    Dùng để biết còn bao nhiêu chỗ dư mà bỏ đoạn lấy đà đầu clip.
    """
    try:
        tho = subprocess.run(
            [ffmpeg, "-hide_banner", "-i", duong], capture_output=True, text=True, timeout=60,
            creationflags=_co_tao_ffmpeg()).stderr or ""
        chu = tho.split("Duration:", 1)[1].split(",", 1)[0].strip()
        gio, phut, giay = chu.split(":")
        return int(gio) * 3600 + int(phut) * 60 + float(giay)
    except (OSError, ValueError, IndexError, subprocess.SubprocessError):
        return 0.0


def _kiem_media(bc: BoiCanh, duong: str) -> None:
    """Mở thử tệp bằng FFmpeg. Hỏng thì **xoá** để lần sau tải lại.

    Kiểm bằng chính công cụ sẽ dùng nó ở khâu dựng: `Content-Length` khớp mà
    tệp vẫn hỏng là chuyện có thật (máy chủ cắt giữa chừng nhưng vẫn khai đủ),
    và cách duy nhất chắc chắn là thử mở.
    """
    ffmpeg = bc.ffmpeg or _tim_ffmpeg()
    if not ffmpeg or not os.path.exists(duong):
        return
    # ═══ GIẢI MÃ CẢ TỆP, KHÔNG CHỈ 0,1 GIÂY ĐẦU ═══
    #
    # 25/08/2026 (story-3d/0001): clip 99 mở được, khung đầu đẹp, nhưng từ
    # giữa tệp toàn "Invalid NAL unit size" — máy chủ trả tệp cụt mà khai đủ.
    # Kiểm 0,1 giây đầu cho qua; khâu dựng cắt tới nó mới đổ, ba lần thử đều
    # đổ ở đúng đoạn ấy. Giải mã hết 8 giây 720p mất chưa tới một giây CPU,
    # rẻ hơn mọi phút ngồi tìm xem cảnh nào làm hỏng video.
    loi = _loi_giai_ma(ffmpeg, duong)
    if loi:
        try:
            os.remove(duong)
        except OSError:
            pass
        raise LoiNoiDung("tệp tải về không mở được: {0}".format(loi[:120]))


def _loi_giai_ma(ffmpeg: str, duong: str) -> str:
    """Giải mã cả tệp bằng FFmpeg (`-xerror`: dừng ở lỗi đầu). Rỗng = lành."""
    ket = subprocess.run(
        [ffmpeg, "-v", "error", "-xerror", "-i", duong, "-f", "null", "-"],
        capture_output=True, text=True,
        creationflags=_co_tao_ffmpeg())
    if ket.returncode == 0:
        return ""
    return _loi_ffmpeg(ket.stderr) or "FFmpeg trả mã {0}".format(ket.returncode)


def _loai_clip_hong(bc: BoiCanh, ffmpeg: str, thu_muc_clip: str,
                    canh: Sequence[Dict[str, Any]]) -> List[int]:
    """Trước khi dựng: soi từng clip, clip hỏng thì cất thành `<n>.mp4.hong`.

    Clip hỏng được coi như THIẾU — video vẫn dựng (cảnh trước giữ hình bù
    vào), và "Làm lại khâu clip" chỉ tạo lại đúng clip ấy. Trả về danh sách
    cảnh bị cất.
    """
    from concurrent.futures import ThreadPoolExecutor  # noqa: PLC0415

    viec = []
    for c in canh:
        so_canh = int(c["scene_id"])
        tep = os.path.join(thu_muc_clip, "{0}.mp4".format(so_canh))
        if os.path.exists(tep):
            viec.append((so_canh, tep))

    def mot(v):
        so_canh, tep = v
        return so_canh, tep, _loi_giai_ma(ffmpeg, tep)

    hong: List[int] = []
    with ThreadPoolExecutor(max_workers=4) as pool:
        for so_canh, tep, loi in pool.map(mot, viec):
            if not loi:
                continue
            try:
                os.replace(tep, tep + ".hong")
            except OSError:
                continue
            hong.append(so_canh)
            bc.ghi("  clip cảnh {0} hỏng ({1}) — cất sang .hong, dựng như cảnh thiếu; "
                   "“Làm lại khâu clip” sẽ tạo lại đúng clip này.".format(so_canh, loi[:90]))
    return hong



def _tao_job(bc: BoiCanh, ham: Callable, **kw):
    """Tạo một job, kiên nhẫn qua lúc máy chủ trục trặc.

    Việc phân loại sự cố nằm ở `core/su_co.py` — một chỗ duy nhất cho cả tool,
    xem ghi chú đầu tệp ấy. Ở đây chỉ cần bảo đảm một điều: `kw` mang
    **cùng một Idempotency-Key** qua mọi lần thử, nên thử lại không bao giờ
    trừ tiền hai lần.

    ═══ ÉP KHOÁ VỀ ASCII NGAY TẠI ĐÂY ═══

    Idempotency-Key đi trong **header HTTP**, mà header chỉ nhận ASCII. Lọt một
    chữ có dấu là mọi lời gọi chết với câu `'ascii' codec can't encode
    character` — câu ấy không nhắc gì tới khoá, tới header hay tới tiếng Việt,
    nên nhìn vào không đoán ra.

    Mọi nơi gọi hôm nay đều đã đi qua `khoa_viec`/`_khoa_chat` (đã ép ASCII).
    Chặn thêm lần nữa ở đây vì đây là **cửa duy nhất** mọi việc tốn tiền phải
    qua: nơi gọi thứ mười lăm quên ép thì cũng không thủng. Đường viết chữ
    (`core/goi_van_ban.py`) đã theo đúng nếp này từ trước.
    """
    khoa = kw.get("idempotency_key")
    if isinstance(khoa, str) and not khoa.isascii():
        kw = dict(kw, idempotency_key=(
            khoa.encode("ascii", "replace").decode("ascii").replace("?", "-")))

    def mot_lan():
        xin_nhip(bc.on_log, ngu=bc.ngu)
        return ham(**kw)

    return goi_kien_nhan(mot_lan, on_log=bc.on_log, kiem_dung=bc.kiem_dung,
                         ngu=bc.ngu)


def _ngu_ngat(bc: BoiCanh, giay: float) -> None:
    """Ngủ mà vẫn bấm Dừng được — chia nhỏ ra, mỗi nhịp ngó lại cờ dừng."""
    con = max(0.0, float(giay))
    while con > 0:
        bc.kiem_dung()
        buoc = min(0.5, con)
        time.sleep(buoc)
        con -= buoc


class LoiKetJob(RuntimeError):
    """Job đã nhận nhưng đợi mãi không xong — gần như chắc chắn khoá bị kẹt.

    Tách riêng khỏi lỗi thường vì nơi gọi xử khác hẳn: gọi lại y nguyên là rơi
    vào đúng job kẹt ấy, phải **đặt job mới bằng khoá mới** mới thoát ra được.
    """


def khoa_thoat_ket(lan: int) -> str:
    """Đuôi khoá idempotency để THOÁT một khoá đã bị kẹt ở máy chủ.

    ═══ MỘT KHOÁ HỎNG WEDGE CẢ LƯỢT CHẠY — ĐO 28/08/2026 ═══

    Phim `openstory/0010`, đoạn giọng đọc số 5 hỏng **mười hai lần liền** với
    `engine_unavailable`, trong khi bốn đoạn trước làm được. Loại dần:

    * gửi **đúng 857 ký tự ấy** qua đường trần trụi, khoá mới → **xong trong 62
      giây**. Không phải nội dung, không phải độ dài, không phải nhà máy.
    * cái khác duy nhất là **khoá**: tool gọi lại bằng đúng khoá cũ.

    Job hỏng ở máy chủ thì bản ghi của khoá ấy giữ luôn cái hỏng. Gọi lại bằng
    khoá cũ là nhận lại đúng cái xác ấy, mãi mãi.

    Tool vốn có một nấc thoát — đuôi `":k2"` — nhưng chỉ **một** nấc. Khâu ngoài
    thử lại cả khâu ba lần, mỗi lần lại dựng đúng hai khoá `""` và `":k2"` cũ,
    nên sau lần đầu là cả hai đều đã hỏng. Lượt chạy kẹt ba tiếng vì thế.

    Đuôi ở đây bám theo THỜI GIAN nên **đổi cả khi tool khởi động lại** — chỗ
    mà một biến đếm trong bộ nhớ không cứu được.

    ⚠ Chỉ gọi hàm này SAU một `LoiKetJob` (máy chủ nhận rồi bỏ). Job hỏng thì
    không bị trừ tiền, nên đặt lại bằng khoá mới là an toàn. Đừng dùng nó cho
    lần gọi ĐẦU: khoá đổi mỗi lần gọi thì mất luôn cái chống trả tiền hai lần.
    """
    import time as _t  # noqa: PLC0415

    return ":k{0}-{1}".format(int(lan), int(_t.time()) // 60)


#: Bao lâu thì nhắc một câu trong lúc đợi job.
#:
#: 90 giây: đủ thưa để không làm rác màn hình, đủ dày để người ngồi trước máy
#: biết tool còn sống. Clip veo3 mất 1–3 phút nên phần lớn job xong trước khi
#: có dòng nhắc nào.
KHOANG_KE_CHO = 90.0

#: Đợi một job tối đa bao lâu rồi coi như kẹt.
#:
#: Từng để 3600 giây. Số đó sai theo cả hai hướng: quá dài để hữu ích (clip
#: veo3 mất 1–3 phút; đợi một tiếng là đợi một thứ không bao giờ tới) và quá
#: ngắn để an toàn (nó vẫn hết hạn, chỉ là sau khi đã giữ chỗ một luồng suốt
#: 60 phút). 12 phút là gấp bốn lần thời gian thật của job chậm nhất — đủ rộng
#: cho lúc nhà máy đông, đủ hẹp để cái kẹt lộ ra khi khách còn ngồi đó.
TRAN_CHO_JOB = 12 * 60.0

#: Riêng khâu đọc thành giọng đợi lâu hơn.
#:
#: Một đoạn kịch bản dài ba nghìn ký tự đọc ra vài phút tiếng, và máy chủ phải
#: dựng xong cả đoạn mới trả. Đo 15/08/2026: để trần chung 12 phút thì hai
#: lượt liền chết ở đúng khâu này — và chết ở đây là mất luôn kịch bản đã trả
#: tiền của khâu trước.
TRAN_CHO_TTS = 25 * 60.0

#: Dấu hiệu máy chủ tự khai "job này hỏng nhưng bạn KHÔNG bị trừ tiền".
#:
#: Đọc chính câu máy chủ nói chứ không đoán theo mã lỗi: câu tiền nong là thứ
#: cổng nói rất rõ ràng và rất ổn định, còn mã lỗi thì thêm bớt theo từng bản.
_KHONG_TRU_TIEN = ("không bị trừ tiền", "khong bi tru tien",
                   "engine_unavailable", "not charged",
                   # Câu cổng dùng từ 02/09/2026 cho job hỏng sau khi thử lại:
                   # "Toàn bộ tiền tạm giữ đã được hoàn về ví bạn".
                   "tiền tạm giữ", "tien tam giu", "hoàn về ví", "refunded")


def _khong_bi_tru_tien(loi_goi) -> bool:
    """Máy chủ có tự nói là chưa trừ tiền cho job này không."""
    chu = str(loi_goi).lower()
    return any(d in chu for d in _KHONG_TRU_TIEN)


#: Mã lỗi / câu chữ nói rằng hỏng là do CHÍNH NỘI DUNG yêu cầu — đặt lại y
#: nguyên bằng khoá mới thì hỏng y nguyên, chỉ tốn thời gian (tiền thì vẫn
#: được hoàn, nhưng vòng lặp vô ích là thứ phải tránh).
_LOI_DO_NOI_DUNG_MA = ("content_rejected", "invalid_prompt", "invalid_request",
                       "validation_error", "rejected")
_LOI_DO_NOI_DUNG_CHU = ("vi phạm", "vi pham", "quy định nội dung",
                        "không hợp lệ", "khong hop le", "prohibited")


def _hong_do_noi_dung(trang_thai: str, loi_goi) -> bool:
    """Job hỏng vì nội dung yêu cầu (từ chối, sai định dạng) — không phải vì máy chủ."""
    if str(trang_thai or "") == "rejected":
        return True
    ma = ""
    if isinstance(loi_goi, dict):
        ma = str(loi_goi.get("code") or "").strip().lower()
    if ma in _LOI_DO_NOI_DUNG_MA:
        return True
    chu = str(loi_goi).lower()
    return any(d in chu for d in _LOI_DO_NOI_DUNG_CHU)


#: Những trạng thái nghĩa là "job này chấm hết rồi, đừng hỏi nữa".
_XONG_HAN = ("succeeded", "completed", "failed", "cancelled", "canceled",
             "rejected")


def _goi_dict(x) -> Dict[str, Any]:
    """Đổi thứ SDK trả về (`Model`) thành `dict` thường."""
    if hasattr(x, "to_dict"):
        return x.to_dict()
    try:
        return dict(x or {})
    except (TypeError, ValueError):
        return {}


def _xong_han(trang_thai: str) -> bool:
    return str(trang_thai or "") in _XONG_HAN


def _ket_job(goi: Dict[str, Any]) -> Dict[str, Any]:
    """Job đã chấm hết: trả về gói nếu xong, ném lỗi đúng loại nếu hỏng.

    Tách riêng vì có **hai** đường đợi job (hỏi từng cái, và hỏi cả lượt bằng
    `SoTheoDoi`), mà cách phân xử lúc job hỏng thì phải giống hệt nhau — nhất
    là chỗ `LoiKetJob`, thứ quyết định có được đặt lại bằng khoá mới hay không.
    """
    trang_thai = str(goi.get("status") or "")
    if trang_thai in ("succeeded", "completed"):
        return goi
    loi_goi = goi.get("error") or trang_thai

    # ═══ JOB HỎNG THÌ MẶC ĐỊNH LÀ ĐẶT LẠI ĐƯỢC — ĐO 03/09/2026 ═══
    #
    # Bản trước chỉ ném `LoiKetJob` (= cho phép khoá mới) khi câu lỗi có chữ
    # "không bị trừ tiền". Ngày 03/09 cổng đổi câu thành *"Yêu cầu này không
    # thể hoàn thành dù thử lại. Toàn bộ tiền tạm giữ đã được hoàn về ví"* —
    # không khớp bảng chữ → rơi xuống `RuntimeError` thường → không đổi khoá →
    # khâu ngoài thử lại ba lần bằng ĐÚNG khoá cũ → cổng phát lại ĐÚNG xác job
    # cũ → ba lần "hỏng" y hệt trong tích tắc, kênh openstory lượt 0015 kẹt ở
    # đoạn giọng đọc số 3 dù 5/6 đoạn kia đã xong. Một lỗi thoáng qua của máy
    # chủ (job chết đúng lúc máy chủ đổi ca) hoá thành vĩnh viễn chỉ vì một
    # câu chữ đổi.
    #
    # Luật đúng không nằm ở câu chữ: hợp đồng cổng ShopAPI (CONTRACT §2.2) là
    # **job hỏng được hoàn 100% tiền** — mọi job `failed`/`cancelled` đều đã
    # được hoàn, nên đặt lại bằng khoá mới KHÔNG BAO GIỜ tốn thêm đồng nào.
    # Ngoại lệ duy nhất là hỏng vì chính NỘI DUNG yêu cầu (từ chối, sai định
    # dạng): đặt lại y nguyên thì hỏng y nguyên — ném lỗi thường để khâu ngoài
    # dừng và người ta sửa nội dung, đừng quay vòng vô ích.
    if _hong_do_noi_dung(trang_thai, loi_goi):
        raise RuntimeError("máy chủ báo job hỏng vì nội dung: {0}".format(loi_goi))
    # Vẫn giữ chữ "job hỏng" trong câu để nhật ký/bài kiểm cũ đọc được;
    # `LoiKetJob` là `RuntimeError` con, nơi bắt lỗi thường vẫn bắt được.
    raise LoiKetJob(
        "máy chủ báo job hỏng (đã hoàn tiền) — đặt lại bằng khoá mới: {0}".format(loi_goi))


class SoTheoDoi:
    """Một chỗ hỏi chung cho cả trăm job — thay cho mỗi job một luồng ngồi canh.

    ═══ VÌ SAO PHẢI GOM VIỆC HỎI LẠI ═══

    Bản cũ để **mỗi job một luồng** tự gọi `jobs.retrieve` theo nhịp giãn dần
    2 → 10 giây. Hai cái giá, và cái thứ hai mới đắt:

    1. Muốn theo dõi 114 job thì phải mở 114 luồng. Không mở nổi thì phải chạy
       theo đợt — mà chờ theo đợt là tự dựng lại đúng hàng rào ta đang gỡ.
    2. Một tấm ảnh nhà máy làm xong ở giây 31 thì tới tận ~giây 40 tool mới
       biết, vì nhịp hỏi lúc ấy đã giãn ra 9–10 giây. Nhân với trăm tấm là hàng
       chục phút chờ suông, mà việc thì đã xong từ lâu.

    SDK có sẵn thứ cần: `client.jobs.list(status="succeeded")` — **một** lời gọi
    trả về cả trăm job. Nên: một luồng duy nhất quét danh sách mỗi 2 giây, đối
    chiếu với sổ những mã đang chờ, thấy cái nào xong thì gạt `Event` của cái
    ấy. Luồng đang chờ chỉ nằm im trên `Event`: không tốn CPU, không tốn một
    lượt gọi nào, và tỉnh dậy trong nửa giây kể từ lúc việc xong.

    Ba lưới an toàn giữ nguyên tính chất cũ:

    * hỏi cả `succeeded` lẫn `failed`, nên job hỏng vẫn báo về đúng chỗ;
    * job nào sổ chung chưa thấy sau 45 giây thì **hỏi riêng** một cái;
    * `GET /v1/jobs` mà hỏng hai lượt liền thì tự quay hẳn về lối hỏi từng
      cái — tool vẫn chạy, chỉ chậm như bản cũ.
    """

    def __init__(self, bc: BoiCanh, nhip: float = NHIP_HOI_CHUNG) -> None:
        self._bc = bc
        self._nhip = max(0.1, float(nhip))
        self._khoa = threading.Lock()
        self._so: Dict[str, Dict[str, Any]] = {}
        self._thoat = threading.Event()
        self._luong: Optional[threading.Thread] = None
        #: Còn tin `GET /v1/jobs` không. Hỏng hai lượt liền thì thôi.
        self.hoi_ca_luot = True
        self._hong_lien = 0

    # ── Ghi tên vào sổ ───────────────────────────────────────────────────────

    def dat(self, ma: str) -> None:
        """Ghi một mã job vào sổ chờ, và mở luồng quét nếu chưa có."""
        if not ma:
            return
        mo = None
        with self._khoa:
            if ma not in self._so:
                self._so[ma] = {"co": threading.Event(), "goi": None}
            if self._luong is None and not self._thoat.is_set():
                self._luong = threading.Thread(target=self._vong, daemon=True,
                                               name="so-theo-doi-job")
                mo = self._luong
        if mo is not None:
            mo.start()

    def bo(self, ma: str) -> None:
        with self._khoa:
            self._so.pop(ma, None)

    def dong(self) -> None:
        """Đóng sổ. Gọi trong `finally` để luồng quét không sống dai hơn khâu."""
        self._thoat.set()
        luong = self._luong
        if luong is not None and luong.is_alive():
            luong.join(timeout=3.0)

    # ── Đợi một job ──────────────────────────────────────────────────────────

    def cho(self, ma: str, tran: float = TRAN_CHO_JOB,
            ten_viec: str = "") -> Dict[str, Any]:
        """Đợi job `ma` xong. Cùng giao kèo với `_cho_job`, kể cả `LoiKetJob`."""
        self.dat(ma)
        bat_dau = time.time()
        het_han = bat_dau + max(60.0, float(tran))
        lan_ke = bat_dau + KHOANG_KE_CHO
        cho_rieng = 2.0
        moc_rieng = bat_dau + (NHIP_HOI_RIENG if self.hoi_ca_luot else cho_rieng)
        try:
            while True:
                self._bc.kiem_dung()
                goi = self._lay(ma)
                if goi is not None:
                    return _ket_job(goi)
                bay_gio = time.time()
                if bay_gio >= het_han:
                    break
                # Nằm trên `Event` chứ không hỏi máy chủ: nửa giây một nhịp là
                # đủ để nút Dừng ăn ngay, mà không tốn lượt gọi nào.
                self._nam_cho(ma, min(0.5, het_han - bay_gio))
                bay_gio = time.time()
                # Sổ chung vừa hỏng giữa lúc đang chờ: kéo mốc hỏi riêng về
                # ngay, đừng ngồi đợi hết 45 giây của lối cũ.
                if not self.hoi_ca_luot and moc_rieng > bay_gio + cho_rieng:
                    moc_rieng = bay_gio + cho_rieng
                if bay_gio >= moc_rieng:
                    if self.hoi_ca_luot:
                        moc_rieng = bay_gio + NHIP_HOI_RIENG
                    else:
                        # Sổ chung hỏng thì đây là đường duy nhất — nhịp giãn
                        # dần đúng như bản cũ.
                        moc_rieng = bay_gio + cho_rieng
                        cho_rieng = min(10.0, cho_rieng * 1.4)
                    goi = self._hoi_rieng(ma)
                    if goi is not None:
                        return _ket_job(goi)
                if bay_gio >= lan_ke:
                    lan_ke = bay_gio + KHOANG_KE_CHO
                    self._bc.ghi(
                        "    {0}: máy chủ vẫn đang làm, đã đợi {1:.0f} phút…"
                        .format(ten_viec or ma[:16], (bay_gio - bat_dau) / 60.0))
        finally:
            self.bo(ma)
        raise LoiKetJob(
            "đợi {0:.0f} phút mà máy chủ vẫn chưa trả kết quả".format(
                (time.time() - bat_dau) / 60.0))

    # ── Bên trong ────────────────────────────────────────────────────────────

    def _lay(self, ma: str) -> Optional[Dict[str, Any]]:
        with self._khoa:
            muc = self._so.get(ma)
            return dict(muc["goi"]) if muc and muc.get("goi") else None

    def _nam_cho(self, ma: str, giay: float) -> None:
        with self._khoa:
            muc = self._so.get(ma)
        co = muc.get("co") if muc else None
        if co is None:
            time.sleep(max(0.0, giay))
        else:
            co.wait(max(0.0, giay))

    def _nhan(self, ma: str, goi: Dict[str, Any]) -> None:
        with self._khoa:
            muc = self._so.get(ma)
            if muc is None:
                return
            muc["goi"] = goi
            muc["co"].set()

    def _con_cho(self) -> List[str]:
        with self._khoa:
            return [m for m, v in self._so.items() if not v.get("goi")]

    def _hoi_rieng(self, ma: str) -> Optional[Dict[str, Any]]:
        """Hỏi riêng đúng một job. Hỏng thì im lặng — trần chờ sẽ lo phần còn lại.

        Nuốt lỗi ở đây là cố ý: một cú rớt mạng lúc *hỏi thăm* không được giết
        một job **đã trả tiền** và vẫn đang chạy ngon lành trên máy chủ.
        """
        try:
            xin_nhip(self._bc.on_log, ngu=self._bc.ngu)
            goi = _goi_dict(self._bc.client.jobs.retrieve(ma))
        except Exception:  # noqa: BLE001
            return None
        if _xong_han(goi.get("status")):
            self._nhan(ma, goi)
            return goi
        return None

    def _vong(self) -> None:
        """Luồng quét: mỗi `nhip` giây hỏi một lượt cho cả sổ."""
        while not self._thoat.wait(self._nhip):
            try:
                self._mot_luot()
            except Exception:  # noqa: BLE001 — sổ hỏng không được giết cả mẻ
                pass

    def _mot_luot(self) -> None:
        if not self.hoi_ca_luot:
            return
        con = set(self._con_cho())
        if not con:
            return
        try:
            for trang_thai in ("succeeded", "failed"):
                con_tro = None
                for _ in range(TRANG_MOI_LUOT):
                    if not con:
                        return
                    xin_nhip(self._bc.on_log, ngu=self._bc.ngu)
                    trang = _goi_dict(self._bc.client.jobs.list(
                        status=trang_thai, limit=SO_MOI_TRANG, cursor=con_tro))
                    for m in (trang.get("data") or []):
                        goi = _goi_dict(m)
                        ma = _ma_job(goi)
                        if ma in con:
                            con.discard(ma)
                            self._nhan(ma, goi)
                    con_tro = trang.get("next_cursor")
                    if not trang.get("has_more") or not con_tro:
                        break
        except Exception as loi:  # noqa: BLE001
            self._hong_lien += 1
            if self._hong_lien >= 2:
                self.hoi_ca_luot = False
                self._bc.ghi("  (không hỏi được cả lượt job — quay về hỏi từng "
                             "cái, chậm hơn nhưng vẫn chạy: {0})".format(
                                 str(loi)[:70]))
            return
        self._hong_lien = 0


def _cho_job(bc: BoiCanh, job, tran: float = TRAN_CHO_JOB,
             ten_viec: str = "", so: Optional[SoTheoDoi] = None) -> Dict[str, Any]:
    """Đợi một job của cổng ShopAPI xong, vẫn bấm Dừng được.

    `so` là **sổ theo dõi chung**: có nó thì việc hỏi thăm gom về một luồng duy
    nhất (xem `SoTheoDoi`) và hàm này chỉ còn nằm chờ. Không có thì rơi về lối
    cũ ngay dưới đây — mỗi job tự hỏi lấy, dùng cho những chỗ chỉ có một job.

    Không dùng `client.jobs.wait()`: nó ngủ trong luồng và không nhả ra, nên nút
    Dừng mất tác dụng — đúng lý do `core/jobs.py` cũng tự chờ lấy.

    Quá `tran` giây thì ném `LoiKetJob` — xem giải thích ở hằng `TRAN_CHO_JOB`.

    `ten_viec` là tên hiện trong dòng nhắc lúc đợi ("cảnh 47"), để người nhìn
    biết cái nào đang chậm chứ không phải chỉ biết "có cái gì đó đang chậm".
    """
    goi = job.to_dict() if hasattr(job, "to_dict") else dict(job or {})
    ma = str(goi.get("id") or goi.get("job_id") or "")
    trang_thai = str(goi.get("status") or "")
    if not ma:
        return goi
    # ═══ JOB ĐÃ CHẤM HẾT NGAY LÚC ĐƯA VÀO — PHÂN XỬ NHƯ MỌI JOB CHẤM HẾT ═══
    #
    # Bản trước liệt kê `"failed"` chung với `"succeeded"` ở đây và **trả gói
    # về như xong**. Đo 03/09/2026: khoá idempotency trùng → cổng phát lại gói
    # 202 của một job đã `failed` → hàm này trả nó về nguyên xi → nơi gọi đem
    # đi tải kết quả → `/download` trả 400 → `LoiTaiVe(400)` → phân loại `CHET`
    # → khâu dừng với câu "tải kết quả hỏng (400)", che mất sự thật là job đã
    # hỏng và đáng được đặt lại bằng khoá mới. Job chấm hết thì đi qua đúng một
    # cửa `_ket_job`, dù nó chấm hết TRƯỚC hay SAU khi ta bắt đầu đợi.
    if _xong_han(trang_thai):
        return _ket_job(goi)
    if so is not None:
        return so.cho(ma, tran=tran, ten_viec=ten_viec)
    # ═══ HỎI THƯA DẦN, KHÔNG HỎI ĐỀU MỖI 2 GIÂY ═══
    #
    # Hỏi mỗi 2 giây là **30 lượt/phút cho một job**. Chạy hai lượt song song là
    # chạm trần 60 lượt/phút của cổng chỉ bằng việc hỏi thăm, chưa làm gì cả —
    # và khi ấy chính lời gọi tạo job bị chặn. Mẻ chạy thật gãy đúng vậy.
    #
    # Job ảnh mất chục giây, job clip mất vài phút. Hỏi dày ở đầu (bắt kịp job
    # nhanh) rồi thưa dần tới 15 giây (đỡ tốn nhịp cho job chậm) là hợp cả hai.
    # Hỏi thưa: clip mất vài phút mới xong, hỏi mỗi 3 giây là 20 lượt/phút cho
    # MỘT job — sáu job song song đã gấp đôi trần của cả tài khoản. Bắt đầu ở 8
    # giây, giãn tới 25 giây. Ảnh xong nhanh thì chậm nhất cũng chỉ trễ 8 giây.
    # ═══ HỎI THĂM DÀY HƠN, VÌ NGÂN SÁCH GỌI RẤT RỘNG ═══
    #
    # Nhịp cũ bắt đầu ở 8 giây và giãn tới 25, chọn theo tiền đề "cổng chỉ cho
    # 60 lượt/phút". Cổng cho **600.000**. Với ngân sách ấy thì hỏi thưa chỉ có
    # một tác dụng: một tấm ảnh xong ở giây thứ 9 phải nằm chờ tới giây 20 mới
    # được nhận ra — nhân với trăm cảnh là hàng chục phút chờ suông.
    #
    # Van chung ở `core/su_co.py` vẫn đứng đó phòng khi trần thật hẹp hơn lời
    # khai; chỗ này chỉ thôi tự bóp thêm một lần nữa.
    cho = 2.0
    bat_dau = time.time()
    het_han = bat_dau + max(60.0, float(tran))
    lan_ke = bat_dau + KHOANG_KE_CHO
    while time.time() < het_han:
        bc.kiem_dung()
        _ngu_ngat(bc, cho)
        cho = min(10.0, cho * 1.4)
        xin_nhip(bc.on_log, ngu=bc.ngu)
        moi = bc.client.jobs.retrieve(ma)
        goi = moi.to_dict() if hasattr(moi, "to_dict") else dict(moi or {})
        trang_thai = str(goi.get("status") or "")
        # Job chấm hết (xong hay hỏng) → MỘT cửa phân xử duy nhất là `_ket_job`,
        # chung với đường hỏi cả lượt của `SoTheoDoi`. Lịch sử vì sao job hỏng
        # phải được đặt lại bằng khoá mới (cảnh 112/112 ngày 15/08/2026, đoạn
        # giọng đọc 3 ngày 03/09/2026) ghi ở đó.
        if _xong_han(trang_thai):
            return _ket_job(goi)
        # ═══ NÓI RA TRONG LÚC ĐỢI ═══
        #
        # Vòng này từng đợi trong im lặng tuyệt đối. Đã xảy ra thật
        # (14/08/2026): chín clip cuối kẹt, sáu luồng cùng ngồi đợi, màn hình
        # không nhích một dòng nào suốt mười hai phút — không cách gì phân biệt
        # "máy chủ đang làm" với "tool treo", kể cả người dựng tool cũng chịu.
        #
        # Im lặng là câu trả lời tệ nhất cho câu hỏi "nó còn chạy không".
        if time.time() >= lan_ke:
            lan_ke = time.time() + KHOANG_KE_CHO
            bc.ghi("    {0}: máy chủ vẫn đang làm, đã đợi {1:.0f} phút…".format(
                ten_viec or ma[:16], (time.time() - bat_dau) / 60.0))
    raise LoiKetJob(
        "đợi {0:.0f} phút mà máy chủ vẫn chưa trả kết quả".format(
            (time.time() - bat_dau) / 60.0))


#: Nhớ URL ảnh tham chiếu ở cấp KÊNH, không phải cấp lượt chạy.
TEP_THAM_CHIEU = "anh-tham-chieu.json"

#: Dùng lại URL trong bao lâu, khi không đọc được hạn thật từ chính URL.
#:
#: ⚠ Đừng tin câu "URL sống 24 giờ" trong tài liệu SDK. URL thật là **link ký
#: hạn** kiểu S3, và tham số `X-Amz-Expires` trong chính nó nói **7198 giây —
#: đúng 2 giờ**. Tôi từng để 20 giờ và mẻ chạy thật gãy đúng kiểu khó đoán
#: nhất: ảnh 1, 2 ra bình thường rồi cảnh 3 trở đi báo *"Ảnh tham chiếu tải
#: không được"* — vì chữ ký hết hiệu lực giữa chừng.
#:
#: Nên: đọc hạn **từ chính URL**, và chỉ dùng 80% hạn ấy. Con số dưới chỉ là
#: đường lui khi URL không mang tham số hạn.
GIO_DUNG_LAI = 45 * 60


def _han_cua_url(url: str) -> float:
    """Đọc `X-Amz-Expires` ngay trong URL. Trả về số giây nên dùng lại.

    Lấy 80% hạn thật để chừa biên: một lượt 100 cảnh chạy cả tiếng, và hết hạn
    giữa chừng thì hỏng ở cảnh thứ n chứ không hỏng ngay — loại lỗi tốn nhất
    để tìm ra.
    """
    try:
        from urllib.parse import parse_qs, urlparse  # noqa: PLC0415

        gia_tri = parse_qs(urlparse(url).query).get("X-Amz-Expires")
        if gia_tri:
            return max(60.0, float(gia_tri[0]) * 0.8)
    except Exception:  # noqa: BLE001 — URL kiểu khác thì dùng đường lui
        pass
    return float(GIO_DUNG_LAI)


#: Nơi để lại bản sao ảnh vừa đẩy lên, cho worker trên CÙNG máy dùng thẳng.
#:
#: Phải khớp biến `SHOPAPI_ANH_CUC_BO` mà phía nhận ảnh cũng đọc. Đặt biến môi
#: trường thì theo biến; không thì theo mặc định dưới đây, và hai bên tự khớp
#: nhau mà không cần ai cấu hình gì.
KHO_ANH_CUC_BO = os.environ.get("SHOPAPI_ANH_CUC_BO") or os.path.join(
    os.environ.get("ProgramData", os.path.expanduser("~")), "ShopAPI", "anh-cuc-bo"
)

#: Giữ bản sao bao lâu rồi dọn (giây). 6 giờ — rộng hơn hẳn đời một mẻ chạy.
HAN_ANH_CUC_BO = 6 * 3600.0

#: Bao lâu mới quét dọn một lượt (giây). Dọn mỗi lần đẩy ảnh là phí; kho này chỉ
#: cần không phình vô hạn chứ không cần sạch từng phút.
NHIP_DON_ANH_CUC_BO = 600.0

_LAN_DON_CUOI = 0.0
_KHOA_DON = threading.Lock()


def _luu_ban_cuc_bo(duong: str, url: str) -> None:
    """Để lại một bản ảnh vừa đẩy lên, ngay trên đĩa máy này.

    ═══ VÌ SAO ═══

    Tool này và worker chạy trên CÙNG một máy. Trước bản này, tool đẩy ảnh lên
    kho ở Singapore rồi worker trên đúng máy ấy tải chính tấm ảnh đó ngược về —
    một tấm ảnh đi Việt Nam → Singapore → Việt Nam.

    Đo ngày 16/08/2026: tool đẩy 463 ảnh (178 MB) mỗi 5 phút, kín đường lên của
    mạng nhà. Đường lên kín thì tín hiệu báo nhận của đường xuống cũng nghẹt,
    nên chặng về chỉ còn 23 KB/s — 516 lượt tải hết hạn 30 giây giữa chừng, kéo
    15–25% job hỏng.

    Chặng đẩy lên vẫn phải giữ (máy chủ cần URL để nhận job), nhưng chặng về thì
    bỏ được, và nó đúng là chặng đang chết.

    ═══ HỎNG THÌ IM LẶNG ═══

    Mọi lỗi ở đây đều nuốt. Đây thuần tuý là một lối tắt: không có bản sao thì
    worker tải mạng như cũ, job vẫn chạy. Ném lỗi ở đây là làm hỏng một khâu
    đang chạy được để đổi lấy một thứ chỉ có tác dụng tăng tốc.
    """
    try:
        ma = _ma_upload(url)
        if not ma:
            return
        os.makedirs(KHO_ANH_CUC_BO, exist_ok=True)
        dich = os.path.join(KHO_ANH_CUC_BO, ma)
        # Ghi ra tên tạm rồi mới đổi tên: worker có thể đọc bất cứ lúc nào, và
        # một file đang ghi dở là một tấm ảnh cụt — thứ hỏng âm thầm tận nhà máy.
        tam = dich + ".dang-ghi"
        shutil.copyfile(duong, tam)
        os.replace(tam, dich)
    except Exception:  # noqa: BLE001 — xem khối "HỎNG THÌ IM LẶNG" ở trên
        return
    _don_kho_cuc_bo()


def _ma_upload(url: str) -> str:
    """Rút mã `upl_...` khỏi URL trả về của máy chủ."""
    try:
        from urllib.parse import urlsplit  # noqa: PLC0415

        ten = urlsplit(url).path.rsplit("/", 1)[-1]
    except Exception:  # noqa: BLE001
        return ""
    ma = ten.split(".", 1)[0]
    return ma if re.fullmatch(r"upl_[a-z0-9]{1,64}", ma) else ""


def _don_kho_cuc_bo() -> None:
    """Xoá bản sao quá hạn. Không dọn thì kho phình ~1,6 GB mỗi giờ."""
    global _LAN_DON_CUOI
    bay_gio = time.time()
    with _KHOA_DON:
        if bay_gio - _LAN_DON_CUOI < NHIP_DON_ANH_CUC_BO:
            return
        _LAN_DON_CUOI = bay_gio
    try:
        for ten in os.listdir(KHO_ANH_CUC_BO):
            duong = os.path.join(KHO_ANH_CUC_BO, ten)
            try:
                if bay_gio - os.path.getmtime(duong) > HAN_ANH_CUC_BO:
                    os.remove(duong)
            except OSError:
                continue
    except OSError:
        return


def _url_tham_chieu(bc: BoiCanh, bo_qua_nho: bool = False) -> List[str]:
    """URL ảnh nhân vật tham chiếu. **Tải lên một lần, dùng lại mãi.**

    ═══ VÌ SAO PHẢI NHỚ, KHÔNG TẢI LẠI MỖI CẢNH ═══

    `reference_images` của cổng ShopAPI nhận **URL**, không nhận đường dẫn trên
    máy. Nên mỗi tấm ảnh cần một lần tải `nv1.png` lên trước — và nếu tải lại
    cho từng cảnh thì một video 97 cảnh là 97 lần tải cùng một tệp 0,5 MB.

    Kho tạm của tài khoản chỉ có 500 MB, và tệp chỉ tự hết hạn sau 2 giờ không
    dùng. Chạy 16 video kiểu ấy là **1.600 lần tải ≈ 880 MB** — vượt trần giữa
    chừng, và lỗi hiện ra ở khâu tạo ảnh chứ không ở chỗ gây ra nó. Chủ dự án
    gặp đúng màn hình ấy, 14/08/2026.

    Nhớ ở cấp **kênh** chứ không cấp lượt: mười sáu video của cùng một kênh
    dùng chung đúng một nhân vật, nên chỉ cần đúng một lần tải cho cả mẻ.

    Khoá theo `(đường dẫn, cỡ tệp, lần sửa cuối)` — thay `nv1.png` là URL cũ tự
    hết giá trị, không phải nhớ đi xoá.
    """
    anh = bc.kenh.anh_nv[:1]
    if not anh:
        return []
    duong = anh[0]
    try:
        dau_vet = "{0}|{1}|{2}".format(
            os.path.basename(duong), os.path.getsize(duong),
            int(os.path.getmtime(duong)))
    except OSError:
        return []

    kho = os.path.join(bc.goc, "PROJECTS", "AUTO", bc.kenh.ma, TEP_THAM_CHIEU)
    goi = {}
    try:
        with open(kho, "r", encoding="utf-8") as tep:
            goi = json.load(tep)
    except (OSError, ValueError):
        goi = {}
    cu = goi.get(dau_vet) if isinstance(goi, dict) else None
    if isinstance(cu, dict) and cu.get("url") and not bo_qua_nho:
        # Hạn đọc TỪ CHÍNH URL, không phải một hằng số đoán bừa — xem
        # `_han_cua_url`. `bo_qua_nho` là đường cho `_tao_anh` ép tải lại khi
        # máy chủ báo nó không tải nổi ảnh tham chiếu.
        if (time.time() - float(cu.get("luc") or 0)) < _han_cua_url(str(cu["url"])):
            return [str(cu["url"])]

    bc.ghi("  tải ảnh nhân vật lên (một lần cho cả kênh)…")

    def tai():
        xin_nhip(bc.on_log, ngu=bc.ngu, so_suat=SUAT_TAI_TEP)
        return bc.client.uploads.upload_file(duong)

    # Trước đây gọi thẳng. Việc tải ảnh cũng tính vào trần 60 lượt/phút, và
    # gặp chặn ở đây thì cả khâu ảnh gãy — mẻ chạy thật dính đúng vậy.
    url = goi_kien_nhan(tai, on_log=bc.on_log, kiem_dung=bc.kiem_dung,
                        ngu=bc.ngu)
    _luu_ban_cuc_bo(duong, str(url))
    goi = goi if isinstance(goi, dict) else {}
    goi[dau_vet] = {"url": str(url), "luc": time.time()}
    _ghi_chu(kho, json.dumps(goi, ensure_ascii=False, indent=1))
    return [str(url)]


def _khoa_ascii(tho: str) -> str:
    """Ép khoá về thuần ASCII. Một bản duy nhất cho mọi loại khoá.

    Mã kênh do người dùng đặt, mà `kiem_ma_kenh` chỉ chặn mấy ký tự Windows
    cấm — chữ có dấu vẫn lọt. Từ khi khoá mang cả mã kênh thì đây không còn là
    rủi ro xa: một kênh đặt tên "Kênh Việt" là đủ làm chết mọi lời gọi.
    """
    return tho.encode("ascii", "replace").decode("ascii").replace("?", "-")


def _khoa_chat(luot: LuotChay, buoc: str) -> str:
    """Idempotency-Key cho một lượt gọi viết chữ.

    Mang cả **mã kênh** chứ không chỉ mã lượt — xem ghi chú ở `khoa_viec`.
    """
    return _khoa_ascii("{0}:{1}:chat:{2}".format(
        luot.ma_kenh or "?", luot.ma_luot, buoc))


def khoa_viec(luot: LuotChay, viec: str, so: Any, *dau_vao: Any) -> str:
    """Idempotency-Key cho một việc. **Phủ cả đầu vào của việc ấy.**

    ═══ VÌ SAO KHOÁ PHẢI ĐỔI KHI ĐẦU VÀO ĐỔI ═══

    Khoá cố định theo *việc* là thứ giữ cho ta không trả tiền hai lần. Nhưng
    "cùng một việc" phải nghĩa là **cùng một việc với cùng đầu vào** — không
    phải "cùng vị trí trong danh sách".

    Mẻ chạy thật cho thấy chỗ hụt: cảnh 2 của lượt L01 từng được gửi đi kèm một
    URL ảnh tham chiếu **đã hết hạn**, job ấy hỏng. Chạy tiếp thì tool tải ảnh
    mới, nhưng vẫn gửi kèm khoá cũ `L01:img:2` — và máy chủ trả lại đúng cái
    job hỏng ấy. Tool ngồi đợi mãi một thứ không bao giờ tốt lên.

    Cho đầu vào vào khoá là hết: URL mới → khoá mới → job mới. Còn chạy lại y
    nguyên đầu vào cũ thì khoá vẫn trùng, vẫn không trả tiền lần hai.

    Băm cho ngắn vì lời nhắc ảnh dài cả nghìn ký tự, mà khoá thì nên gọn.
    """
    import hashlib  # noqa: PLC0415

    van = "|".join(str(m) for m in dau_vao)
    dau = hashlib.sha256(van.encode("utf-8")).hexdigest()[:10]
    # ═══ KHOÁ PHẢI MANG CẢ MÃ KÊNH ═══
    #
    # Mọi kênh đều đánh số lượt từ `0001`, nên khoá chỉ có mã lượt là kênh thứ
    # hai đâm thẳng vào kênh thứ nhất. Đo thật 19/08/2026: lượt `0001` của
    # TL5-T7 va vào lượt `0001` của TL4-T7, cổng trả `409 idempotency_conflict`
    # và lượt chạy kẹt 25 phút. Với khách vừa tạo ba kênh mới thì đây không
    # phải ca hiếm — nó dính ngay từ lượt đầu tiên của kênh thứ hai.
    tho = "{0}:{1}:{2}:{3}:{4}".format(luot.ma_kenh or "?", luot.ma_luot,
                                       viec, so, dau)
    # ═══ KHOÁ PHẢI THUẦN ASCII ═══
    #
    # Idempotency-Key đi trong **header HTTP**, mà header chỉ nhận ASCII. Lọt
    # một chữ có dấu vào đây là **mọi lời gọi đều chết** với câu
    # `'ascii' codec can't encode character '\\u1ea3'` — câu đó không nhắc gì
    # tới khoá, tới header, hay tới tiếng Việt, nên nhìn vào không đoán ra.
    #
    # Hôm nay mã lượt là `L01`, `0001` — toàn ASCII. Nhưng mã lượt do người
    # dùng đặt (tab Tự động sinh theo số thứ tự, còn chạy tay thì tuỳ), và một
    # cái tên kênh có dấu là đủ làm hỏng cả mẻ. Chặn ngay tại đây rẻ hơn nhiều.
    #
    # Từ 19/08/2026 phép này dùng chung với khoá viết chữ — xem `_khoa_ascii`.
    # Khoá viết chữ từng dựng riêng ở nơi khác và bỏ sót đúng phép này, nên nó
    # là chỗ duy nhất trong tool còn có thể lọt chữ có dấu vào header.
    return _khoa_ascii(tho)


#: Nhớ URL của ảnh từng cảnh, ở cấp LƯỢT (mỗi lượt một bộ ảnh riêng).
TEP_URL_ANH = "anh-canh-url.json"


def _url_anh_canh(bc: BoiCanh, luot: LuotChay, so: int, duong: str,
                  bo_qua_nho: bool = False) -> str:
    """URL của ảnh cảnh `so`, để làm khung đầu cho clip. Tải lên một lần rồi nhớ.

    ═══ VÌ SAO KHÔNG ĐƯA THẲNG ĐƯỜNG DẪN ═══

    `image_url` của cổng nhận **URL http/https**, không nhận đường dẫn trên
    máy — y như `reference_images`. Bản đầu tôi đưa thẳng
    `…/5-anh/3.png` và cả khâu clip chết ngay tấm đầu với câu *"Chỉ chấp nhận
    URL http/https"*.

    Khác ảnh nhân vật ở một điểm quan trọng: ảnh nhân vật **một tấm cho cả
    kênh**, còn ảnh cảnh thì **mỗi cảnh một tấm** — 99 cảnh là 99 lần tải, không
    có cách nào gộp. Nên phải nhớ kỹ: tải rồi thì đừng tải lại, kẻo chạy tiếp
    một lượt dở là tải lại từ đầu và đụng trần kho tạm 500 MB.

    Nhớ theo `(tên tệp, cỡ, lần sửa cuối)` như ảnh nhân vật: tạo lại ảnh cảnh
    ấy thì URL cũ tự hết giá trị.
    """
    try:
        dau_vet = "{0}|{1}|{2}".format(
            os.path.basename(duong), os.path.getsize(duong),
            int(os.path.getmtime(duong)))
    except OSError:
        return ""
    kho = os.path.join(luot.thu_muc, TEP_URL_ANH)
    with _KHOA_URL_ANH:
        goi = {}
        try:
            with open(kho, "r", encoding="utf-8") as tep:
                goi = json.load(tep)
        except (OSError, ValueError):
            goi = {}
        cu = goi.get(dau_vet) if isinstance(goi, dict) else None
        if isinstance(cu, dict) and cu.get("url") and not bo_qua_nho:
            if (time.time() - float(cu.get("luc") or 0)) < _han_cua_url(str(cu["url"])):
                return str(cu["url"])

    def tai():
        xin_nhip(bc.on_log, ngu=bc.ngu, so_suat=SUAT_TAI_TEP)
        return bc.client.uploads.upload_file(duong)

    url = goi_kien_nhan(tai, on_log=bc.on_log, kiem_dung=bc.kiem_dung,
                        ngu=bc.ngu)
    _luu_ban_cuc_bo(duong, str(url))
    with _KHOA_URL_ANH:
        goi = {}
        try:
            with open(kho, "r", encoding="utf-8") as tep:
                goi = json.load(tep)
        except (OSError, ValueError):
            goi = {}
        goi = goi if isinstance(goi, dict) else {}
        goi[dau_vet] = {"url": str(url), "luc": time.time()}
        _ghi_chu(kho, json.dumps(goi, ensure_ascii=False, indent=1))
    return str(url)


#: Mười hai luồng cùng ghi một tệp nhớ thì bản ghi của luồng này đè luồng kia.
_KHOA_URL_ANH = threading.Lock()


#: Máy chủ báo nó không tải nổi ảnh tham chiếu ta đưa.
_ANH_THAM_CHIEU_HONG = ("ảnh tham chiếu tải không được",
                        "không tải được ảnh từ địa chỉ",
                        "reference image",
                        # Cổng từ chối ngay từ khâu kiểm đầu vào khi đường dẫn
                        # không phải URL — gặp câu này thì tải lên rồi gửi URL.
                        "chỉ chấp nhận url",
                        "phải là url", "must be a url")


def _tao_anh(bc: BoiCanh, luot: LuotChay, loi_nhac: str,
             hop: "ThamChieu", khoa: str, ten_hien: str = "",
             so: Optional[SoTheoDoi] = None):
    """Tạo một tấm ảnh. Chữ ký ảnh tham chiếu hết hạn thì **tự tải lại**.

    ═══ VÌ SAO CẦN TỰ CHỮA, KHÔNG CHỈ CẦN NHỚ ĐÚNG HẠN ═══

    URL kho tạm là link ký hạn: hết hiệu lực là máy chủ không tải nổi ảnh nữa,
    và nó trả `invalid_request` — loại lỗi mà bộ phân loại xếp vào "hỏng thật,
    đừng thử lại". Xếp vậy là đúng với phần lớn `invalid_request`, nhưng sai
    với đúng cái này: chữ ký hết hạn thì **tải lại là xong**.

    Đo được ở mẻ chạy thật: ảnh 1 và 2 ra bình thường, cảnh 3 trở đi báo *"Ảnh
    tham chiếu tải không được"*. Chữ ký sống 2 giờ, mà lượt chạy dài hơn thế.

    Chỉ tính đúng hạn thôi thì chưa đủ — hạn có thể đổi, đồng hồ máy có thể
    lệch, và lượt chạy có thể bị treo giữa chừng rồi mới tiếp. Nên có thêm
    tầng này: gặp đúng câu ấy thì bỏ bản nhớ, tải lại, thử lại một lần.

    Trả về danh sách URL ảnh, hoặc `(danh sách, tham chiếu mới)` khi vừa phải
    tải lại — để nơi gọi cập nhật cho những cảnh sau khỏi hỏng tiếp.
    """
    def mot_lan(anh_tc, hau_to=""):
        job = _tao_job(bc, bc.client.images.create,
                       prompt=loi_nhac, n=1, aspect_ratio="16:9",
                       reference_images=anh_tc or None,
                       idempotency_key=khoa + hau_to)
        return _cho_job(bc, job, ten_viec=ten_hien, so=so)

    dang_dung = hop.lay()
    try:
        return mot_lan(dang_dung)
    except LoiKetJob:
        # Job đã nhận nhưng bỏ đó. Khoá mới là đường duy nhất — xem ghi chú
        # dài ở chỗ tạo clip.
        # Hai nấc, đuôi bám thời gian — xem `khoa_thoat_ket`. Một nấc `":k2"`
        # cố định thì chạy lại lượt là gặp đúng khoá đã hỏng.
        for _lan in range(1, 3):
            bc.ghi("    {0}: máy chủ nhận việc rồi bỏ đó — đặt lại bằng khoá "
                   "mới ({1}/2).".format(ten_hien or "ảnh", _lan))
            try:
                return mot_lan(dang_dung, khoa_thoat_ket(_lan))
            except LoiKetJob:
                if _lan == 2:
                    raise
    except Exception as loi:  # noqa: BLE001
        chu = str(loi).lower()
        if not any(d in chu for d in _ANH_THAM_CHIEU_HONG) or not dang_dung:
            raise
        # Nhờ CÁI HỘP làm mới, không tự tải: luồng khác vừa làm rồi thì mình
        # dùng bản của họ. Đây là chỗ chặn 12 lần tải xuống còn 1.
        moi = hop.lam_moi(dang_dung)
        # Khoá phải đổi: khoá cũ đã gắn với một job hỏng, hỏi lại là nhận lại
        # đúng cái hỏng ấy — xem `core/su_co.py`.
        return mot_lan(moi, ":tc2")


def _ma_job(goi: Dict[str, Any]) -> str:
    return str((goi or {}).get("id") or (goi or {}).get("job_id") or "")


def _tai_ket_qua(bc: BoiCanh, goi: Dict[str, Any], chi_so: int, dich: str,
                 ngu: Callable[[float], None] = time.sleep) -> str:
    """Tải file kết quả của một job về `dich`, **kiên nhẫn qua trục trặc tạm**.

    ═══ VÌ SAO PHẢI KIÊN NHẪN NGAY Ở ĐÂY ═══

    Trước 18/08/2026 hàm này gọi thẳng một lần rồi ném lỗi lên. Lỗi ấy đi lên
    tận `core/auto.chay`, nơi chỉ có **ba lượt thử cho CẢ KHÂU** — nên một tệp
    tải trượt là bỏ luôn cả mẻ 133 ảnh, và ba lượt thử của khâu chỉ tổ tạo lại
    y hệt tình huống cũ.

    Đo trên hai lượt chạy thật cùng ngày: chết ở ảnh 71/133 và 87/115. Mỗi lượt
    là hàng chục ảnh **đã trả tiền** mà không dùng được.

    Tải về là việc **không tốn tiền và không đổi trạng thái** — một `GET` thuần.
    Nên chỗ này là chỗ rẻ nhất trong cả dây chuyền để kiên nhẫn: đợi vài chục
    giây ở đây rẻ hơn bỏ cả mẻ đúng một trăm lần.

    `goi_kien_nhan` chỉ đợi những loại đáng đợi; `CHET` và `NOI_DUNG` vẫn ném
    lên ngay như cũ.
    """
    return goi_kien_nhan(
        lambda: _tai_ket_qua_mot_lan(bc, goi, chi_so, dich),
        on_log=bc.ghi, kiem_dung=bc.kiem_dung, ngu=ngu)


def _tai_ket_qua_mot_lan(bc: BoiCanh, goi: Dict[str, Any], chi_so: int,
                         dich: str) -> str:
    """Một lượt tải, không thử lại. Chỗ gọi là `_tai_ket_qua`.

    ═══ VÌ SAO ĐI QUA `/v1/jobs/{id}/download` CHỨ KHÔNG DÙNG `output.url` ═══

    Từ 14/08/2026 cổng giao ảnh và clip bằng **link Google**, và link ấy chỉ
    sống **khoảng 6 giờ**. `output.url` vẫn dùng được, nhưng chỉ nếu tải ngay.

    Dây chuyền này thì không "ngay" được: một lượt 99 clip chạy hàng giờ, và
    người dùng có thể bấm Chạy tiếp vào hôm sau. Link 6 giờ là loại hạn chắc
    chắn sẽ hết giữa chừng — đúng kiểu hỏng đã cắn tôi hai lần với chữ ký ảnh
    tham chiếu.

    `/download` **không hết hạn chừng nào job còn**: mỗi lần gọi nó tự lái sang
    một đường tải còn tươi. Nên chỗ cần nhớ là **mã job**, không phải đường
    tải — và mã job thì tôi đã nhớ sẵn trong `idempotency_key` rồi.

    `bc.client._http` đã bật `follow_redirects=True`, nên cái bẫy "quên `-L` rồi
    nhận 302 rỗng" không dính ở đây.
    """
    ma = _ma_job(goi)
    if not ma:
        raise LoiNoiDung("máy chủ không trả về mã job để tải kết quả")
    goc = str(getattr(bc.client, "base_url", "") or "https://api.shopapi.vn")
    dia_chi = "{0}/v1/jobs/{1}/download".format(goc.rstrip("/"), ma)
    if chi_so:
        dia_chi += "?index={0}".format(int(chi_so))

    os.makedirs(os.path.dirname(dich) or ".", exist_ok=True)
    tam = duong_tam(dich)
    xin_nhip(bc.on_log, ngu=bc.ngu)
    dau = bc.client._build_headers(accept="*/*")  # noqa: SLF001
    da_nhan = 0
    with bc.client._http.stream("GET", dia_chi, headers=dau) as ph:  # noqa: SLF001
        if ph.status_code >= 400:
            ph.read()
            raise LoiTaiVe("tải kết quả hỏng ({0}) cho job {1}".format(
                ph.status_code, ma), ph.status_code)
        khai = ph.headers.get("Content-Length")
        with open(tam, "wb") as tep:
            for khuc in ph.iter_bytes(1 << 16):
                tep.write(khuc)
                da_nhan += len(khuc)
    if da_nhan == 0:
        _bo_tep(tam)
        raise LoiNoiDung("tải về tệp rỗng")
    if khai:
        try:
            can = int(khai)
        except (TypeError, ValueError):
            can = 0
        if can and da_nhan < can:
            _bo_tep(tam)
            raise LoiNoiDung(
                "tải thiếu: nhận {0} byte trong khi máy chủ báo {1}".format(
                    da_nhan, can))
    thay_the(tam, dich)
    return dich


def _bo_tep(duong: str) -> None:
    try:
        os.remove(duong)
    except OSError:
        pass


def _khau_kich_ban(bc_goc: BoiCanh):
    def lam(luot: LuotChay, tt: TrangThaiKhau):
        # Khâu này — và CHỈ khâu này — được đổi đường viết chữ sang thuê bao
        # Claude của máy khi chủ máy bật nút ấy. `cho_kich_ban` trả về bản sao,
        # nên các khâu sau (bảng cảnh, ảnh bìa…) vẫn nhận `bc_goc` nguyên vẹn.
        bc = bc_goc.cho_kich_ban()
        k = bc.kenh
        d = luot.thu_muc
        duong_kb = os.path.join(d, "1-kich-ban.txt")

        # ═══ NÓI RÕ ĐI ĐƯỜNG NÀO, NGAY DÒNG ĐẦU ═══
        #
        # Chủ dự án, 25/08/2026, đọc nhật ký lượt 0049: *"sao bước viết content
        # ở máy này lại là dùng api… có thể mày đang viết content bằng
        # shopapi"*. Một dòng này để ai dán nhật ký lên cũng thấy ngay.
        if bc_goc.goi_chat_kich_ban is not None:
            from .viet_max import (NHIP_THU_LAI, THANG_MO_HINH,  # noqa: PLC0415
                                   mo_hinh_dang_dung)

            # Nói model ĐANG dùng, không nói model đứng đầu thang: hôm nay
            # fable cạn hạn mức thì dòng này phải hiện đúng bậc thay thế, chứ
            # không phải một cái tên đã bị bỏ qua từ nãy.
            dang = mo_hinh_dang_dung(bc.goc)
            bc.ghi("  đường viết chữ: Claude Code (thuê bao Claude Max của máy), "
                   "model {0}{1}, KHÔNG tiêu ví; hết hạn mức thì tụt bậc ngay, "
                   "hỏng vì lý do khác thì thử lại {2} lần rồi mới báo lỗi.".format(
                       dang,
                       "" if dang == THANG_MO_HINH[0] else
                       " (đã tụt từ {0} vì hết hạn mức)".format(THANG_MO_HINH[0]),
                       1 + len(NHIP_THU_LAI)))
        else:
            bc.ghi("  đường viết chữ: ví ShopAPI, model {0} (mỗi lượt gọi trừ "
                   "tiền).".format(k.mo_hinh))

        # Tư liệu: lời thoại video đối thủ.
        tu_lieu = _doc_chu(os.path.join(d, "0-tu-lieu.txt"))
        link = str(luot.dau_vao.get("link") or "").strip()
        if not tu_lieu and link:
            bc.ghi("  đang lấy lời thoại video tư liệu…")
            lay = bc.lay_tu_lieu
            if lay is None:
                from .script_video import lay_script  # noqa: PLC0415

                lay = lay_script
            ket = lay(link, cancel=bc.cancel, cho_phep_nghe=True,
                      ngon_ngu_uu_tien=k.ngon_ngu, on_log=bc.on_log)
            tu_lieu = getattr(ket, "text", "") or ""
            if not tu_lieu:
                raise RuntimeError(
                    "không lấy được lời thoại của video tư liệu: {0}".format(
                        getattr(ket, "loi", "") or "không rõ"))
            # ═══ BỎ KHOẢNG TRẮNG THỪA CHO TIẾNG VIẾT LIỀN ═══
            #
            # Phụ đề `json3` của YouTube tách từng "từ" bằng khoảng trắng. Với
            # tiếng Nhật/Trung/Thái (viết liền, không cách) thì bản gỡ băng phình
            # ~60% số ký tự so với chữ thật. Dọn ngay để `CHARS_GOC` và mục tiêu
            # độ dài đo trên chữ THẬT, không phải trên khoảng trắng máy chèn.
            from .lam_sach import go_cach_cjk  # noqa: PLC0415
            tu_lieu = go_cach_cjk(tu_lieu, k.ngon_ngu)
            _ghi_chu(os.path.join(d, "0-tu-lieu.txt"), tu_lieu)
            # Giữ lại tiêu đề + mã video + ĐỘ DÀI video đối thủ: kênh `nguyen_goc`
            # cần tiêu đề/mã, còn bước đo độ dài cần số giây. Chạy lại thì bước
            # lấy này bị bỏ qua nên `ket` không còn — đọc lại từ tệp bên cạnh.
            _ghi_doi_thu(d, getattr(ket, "title", "") or "",
                         getattr(ket, "video_id", "") or "",
                         getattr(ket, "duration_s", 0) or 0)
            # Đếm KÝ TỰ, không đếm "chữ" theo dấu cách: tiếng Nhật/Trung không có
            # dấu cách nên 14.000 ký tự in ra "1 chữ" — chủ dự án tưởng tư liệu rỗng.
            bc.ghi("  tư liệu: {0} ký tự ≈ {1} phút đọc.".format(
                len(tu_lieu), _phut(len(tu_lieu), k.ky_tu_moi_phut)))

        # ═══ ĐỘ DÀI NHẮM TỚI: THEO PHÚT, HAY THEO VIDEO GỐC ═══
        #
        # Kênh thường nhắm một số phút cố định (`ky_tu_muc_tieu`). Kênh remake
        # "gần như giống đối thủ nhất" (`do_dai_theo_goc`) thì nhắm đúng độ dài
        # bản gốc: video dài bằng video đối thủ, không bị kéo/nén về mốc cố định.
        # Con số này dẫn dắt bước viết, bước nắn (nếu có) và chốt chặn quá ngắn.
        #
        # ═══ ĐO THEO SỐ GIÂY VIDEO, KHÔNG PHẢI SỐ KÝ TỰ BẢN GỠ BĂNG ═══
        #
        # Trước đây mục tiêu = số ký tự tư liệu. Nhưng tư liệu có thể là bản DỊCH
        # (YouTube trả phụ đề tiếng Việt cho video Nhật) — mà một ý tiếng Việt
        # dài gấp đôi tiếng Nhật. Đo thật link GJjYlTjNV8g (16/08/2026): bản dịch
        # Việt 16.187 ký tự, bản Nhật gốc chỉ ~4.847. Lấy 16.187 làm mục tiêu thì
        # bài Nhật viết ra dài gấp rưỡi — video 16 phút thành 26 phút.
        #
        # Số giây video là thước KHÔNG phụ thuộc ngôn ngữ: nhân với `ky_tu_moi_
        # phut` (đo từ bảy video thật của kênh) ra đúng số ký tự tiếng ấy cần cho
        # ngần ấy phút. 975 giây × 298 ÷ 60 = 4.842 ≈ 4.847 chữ Nhật gốc — khớp.
        # Thiếu số giây (dán tay bản gỡ băng, không link) thì lui về số ký tự.
        giay_goc = 0
        try:
            giay_goc = int(_doc_doi_thu(d).get("duration") or 0)
        except (TypeError, ValueError):
            giay_goc = 0
        muc_tieu_kt = _muc_tieu_do_dai(k, tu_lieu, giay_goc)

        chung = {
            "LANGUAGE": k.giong_van or k.ngon_ngu,
            # Tên tiếng bằng tiếng Việt ("tiếng Nhật") cho lời nhắc đọc rõ —
            # `LANGUAGE` ở trên là mô tả giọng văn (tiếng Anh) của kênh.
            "NGON_NGU": ten_tieng(k.ngon_ngu) or (k.ngon_ngu or ""),
            "CHANNEL": _mo_ta_kenh(k),
            "CHARS": (muc_tieu_kt if muc_tieu_kt > 0
                      else "không giới hạn — dài ngắn theo câu chuyện"),
            # ═══ VÌ SAO PHẢI NÓI CẢ SÀN LẪN TRẦN ═══
            #
            # Lời nhắc chỉ nói một con số đích thì AI viết theo độ dài BẢN GỐC:
            # lượt TL4-T7/0001 nhắm 13 phút, cả năm bản ra 23–25 phút vì gốc dài
            # 23 phút. Sửa lời nhắc thành "tối đa X, vượt là loại" thì nó lộn
            # sang đầu kia — bản thử ra 1.466 ký tự, đúng 4,9 phút trên đích 13.
            #
            # Nên đưa cả hai đầu, và đưa bằng SỐ chứ không bằng chữ "khoảng":
            # AI không đếm ký tự được, nhưng nó bám được vào một khoảng có biên.
            "CHARS_MIN": (int(muc_tieu_kt * 0.85) if muc_tieu_kt > 0 else ""),
            "CHARS_MAX": (int(muc_tieu_kt * 1.15) if muc_tieu_kt > 0 else ""),
            # ═══ MỘT THƯỚC: PHÚT ĐỌC ═══
            #
            # Chủ dự án, 25/08/2026: *"lúc thì đo bằng độ dài ký tự lúc thì
            # phút"*. Mục tiêu là PHÚT; ký tự chỉ là quy đổi theo giọng đọc
            # của kênh (`ky_tu_moi_phut`). Lời nhắc và bộ chấm nói cả hai,
            # phút trước, ký tự sau trong ngoặc.
            "PHUT": (_phut(muc_tieu_kt, k.ky_tu_moi_phut) if muc_tieu_kt > 0
                     else "tự do (không giới hạn)"),
            "PHUT_GOC": _phut(len(tu_lieu), k.ky_tu_moi_phut),
            # ═══ MỘT CÁI THƯỚC ĐỂ SO, KHÔNG PHẢI MỘT SỐ ĐỂ ĐẾM ═══
            #
            # AI không đếm được ký tự. Bảo nó "khoảng 4.470 ký tự" thì nó viết
            # 2.563 — đo được ngày 19/08/2026. Nhưng nó ước lượng rất tốt khi
            # có một văn bản nằm ngay trước mắt để so.
            #
            # Đây không phải chuyện thẩm mỹ. Bản nháp hụt độ dài chính là thứ
            # mở đường cho bước SỬA đi lấp phần thiếu bằng chữ của bản gốc. Đo
            # từng mốc trong một lượt chạy thật:
            #
            #     bản gốc                4.463 ký tự
            #     sau bước viết   30,5% trùng, 2.563 ký tự
            #     sau bước sửa    77,6% trùng, 4.460 ký tự
            #
            # Bài sau khi sửa dài 4.460 trong khi bản gốc 4.463 — lệch ba ký
            # tự. Bước sửa không sửa; nó lấp chỗ hụt bằng bản gốc.
            "CHARS_GOC": len(tu_lieu),
            # ═══ ĐỘ DÀI HOOK: MỘT CON SỐ, KHÔNG PHẢI MỘT CÂU MÔ TẢ ═══
            #
            # `2d-hook.md` từng nói *"độ dài xấp xỉ đoạn mở của bản gốc"*. Nghe
            # thì rõ, nhưng "đoạn mở của bản gốc" không có ranh giới: lượt 0007
            # model đọc là vài câu đầu (viết ra 128–243 ký tự), lượt 0008 đọc là
            # toàn bộ phần trước 一つ目 (viết ra 1.012–1.074). Cả hai đều là
            # cách đọc hợp lý — và rào chắn 60–600 ký tự đánh rớt CẢ BA hook của
            # 0008, nên lượt ấy chạy không có hook viết riêng.
            #
            # Một phút đọc là mốc có nghĩa: chỗ rớt nặng nhất nằm trong 60 giây
            # đầu, và đoạn mở của ba video đối thủ đã thắng đo được ~240 ký tự,
            # sát đúng một phút giọng kênh này.
            #
            # KẸP VÀO TRONG RÀO CHẮN, nếu không thì tự tay dựng lại ca 0008:
            # tiếng châu Âu đọc ~900 ký tự/phút, khai 900 là vượt trần 600 của
            # `hook_dung_duoc` và mọi hook bị bỏ. Trần 480 chừa chỗ cho nết
            # viết dôi của model.
            "CHARS_HOOK": min(480, max(120, int(
                getattr(k, "ky_tu_moi_phut", 0) or 270))),
            "COMPETITOR_TRANSCRIPT": tu_lieu,
            "TRANSCRIPT_SAMPLE": tu_lieu[:1500],
            # Bám bản gốc hay đặt lại theo chất kênh — do kênh chọn qua
            # `che_do_tieu_de` (mặc định "faithful", nết cũ). Lời nhắc
            # `1-tieu-de.md` đã có sẵn hai nhánh này; đây chỉ chọn nhánh nào.
            "MODE": k.che_do_tieu_de,
            "CASING": ("Viết hoa toàn bộ." if k.chu_bia_hoa
                       else "Giữ nguyên chữ như bạn viết."),
        }

        # ═══ TIÊU ĐỀ: NGƯỜI ĐƯA THÌ DÙNG CỦA NGƯỜI ═══
        #
        # Chủ dự án, 14/08/2026: *"thì tao cung cấp thì dùng của tao nếu không
        # cung cấp thì mày tự làm"*. Đè lên thứ người ta đã nghĩ ra là vừa tốn
        # một lượt gọi vừa làm hỏng ý họ.
        tieu_de = str(luot.dau_vao.get("tieu_de") or "").strip()
        chu_bia = str(luot.dau_vao.get("chu_bia") or "").strip()
        khuon_tieu_de = k.prompt.get("1-tieu-de.md", "")
        # Tiêu đề + mã video đối thủ đã nhớ lúc lấy lời thoại (còn cả khi chạy
        # lại). Vừa để nhánh `nguyen_goc` lấy nguyên, vừa cho nhánh gọi AI biết
        # tiêu đề thật của đối thủ thay vì một ô rỗng như trước.
        doi_thu = _doc_doi_thu(d)
        tieu_de_doi_thu = (str(luot.dau_vao.get("tieu_de_doi_thu") or "").strip()
                           or doi_thu.get("title", ""))
        if tieu_de and chu_bia:
            bc.ghi("  dùng tiêu đề và chữ bìa bạn đã đưa — bỏ qua bước đặt tên.")
        elif k.che_do_tieu_de == "nguyen_goc" and tieu_de_doi_thu:
            # ═══ LẤY NGUYÊN TIÊU ĐỀ + ĐỌC CHỮ TRÊN ẢNH BÌA ĐỐI THỦ ═══
            #
            # Chủ dự án, 22/08/2026: kênh remake "gần như giống đối thủ nhất" thì
            # lấy nguyên tiêu đề đối thủ, và chữ bìa đọc thẳng từ ảnh bìa đối thủ
            # — bỏ hẳn lượt gọi AI viết lại. Đọc ảnh bìa hỏng (cổng không nhận
            # ảnh, ảnh không tải được) thì lấy tiêu đề làm chữ bìa: đường lui an
            # toàn, không bao giờ làm vỡ lượt chạy.
            tieu_de = tieu_de or tieu_de_doi_thu
            if not chu_bia:
                bc.kiem_dung()
                bc.ghi("  đang đọc chữ trên ảnh bìa đối thủ…")
                doc = _doc_chu_bia_doi_thu(bc, luot, doi_thu.get("video_id", ""))
                chu_bia = doc or tieu_de
                if not doc:
                    bc.ghi("  (không lấy được chữ bìa — tạm dùng tiêu đề đối thủ)")
        elif not khuon_tieu_de.strip():
            # ═══ THIẾU LỜI NHẮC THÌ BỎ QUA, ĐỪNG GỬI LỜI NHẮC RỖNG ═══
            #
            # Mọi bước không bắt buộc khác đều có cửa này (`3-sua`, `4-do-dai`,
            # `5-hoan-thien`, `6-seo`) — riêng bước đặt tên thì quên. Kênh nào
            # dùng bộ lời nhắc gọn, bỏ tệp này đi, là tool gửi một lời nhắc
            # RỖNG lên cổng: trả tiền cho một lượt gọi vô nghĩa rồi nhận về
            # một tiêu đề bịa không liên quan gì tới video.
            #
            # Lấy tạm câu mở đầu tư liệu làm tiêu đề. Xấu, nhưng thật — và
            # khách sửa được ở ô nhập trước khi chạy.
            tieu_de = tieu_de or (tu_lieu.strip().splitlines() or [""])[0][:80]
            chu_bia = chu_bia or tieu_de[:20]
            bc.ghi("  (kênh không có lời nhắc đặt tên — tạm lấy câu đầu tư "
                   "liệu, bạn sửa lại ở ô Tiêu đề khi chạy)")
        else:
            bc.kiem_dung()
            bc.ghi("  đang đặt tiêu đề…")
            tra = _goi(bc, _thay(khuon_tieu_de, dict(
                chung, COMPETITOR_TITLE=tieu_de_doi_thu)),
                _khoa_chat(luot, "tieu-de"))
            t, b = _doc_tieu_de(tra)
            tieu_de = tieu_de or t
            chu_bia = chu_bia or b
        _ghi_chu(os.path.join(d, "1-tieu-de.txt"),
                 "TITLE: {0}\nTHUMB: {1}\n".format(tieu_de, chu_bia))
        chung["TITLE"] = tieu_de
        chung["THUMB"] = chu_bia

        # ═══ MỘT BƯỚC ĐỌC BẢN GỐC TRƯỚC KHI VIẾT ═══
        #
        # Chỉ chạy khi kênh có `2a-phan-tich.md`; kênh không có thì bỏ qua, y
        # như mọi bước không bắt buộc khác.
        #
        # Chủ dự án, 19/08/2026: *"có thể có 1 bước trước khi viết đó là việc
        # phân tích kịch bản đối thủ, để biết nó hay chỗ nào và chưa hay chỗ
        # nào để khi viết có thêm yêu cầu về các vấn đề đó"*.
        #
        # Vì sao đáng một lượt gọi: bảo AI *"hãy viết hay hơn"* là giao một
        # mục tiêu không có chỗ bám. Đưa nó một danh sách **chỗ cụ thể để
        # vượt** thì việc trở nên làm được. Bản phân tích đi vào lời nhắc viết
        # qua `<<PHAN_TICH>>`.
        #
        # Ghi ra đĩa để đứt là chạy tiếp — cùng nết với các bước nháp bên dưới.
        khuon_pt = k.prompt.get("2a-phan-tich.md", "")
        if khuon_pt.strip() and not _doc_chu(duong_kb):
            nhap_pt = os.path.join(d, "1-nhap-phan-tich.txt")
            da_co = _doc_chu(nhap_pt).strip()
            if da_co:
                bc.ghi("  phân tích bản gốc — đã có từ lần trước, dùng lại.")
                chung["PHAN_TICH"] = da_co
            else:
                bc.kiem_dung()
                bc.ghi("  đọc và phân tích bản gốc…")
                pt = _goi(bc, _thay(khuon_pt, chung),
                          _khoa_chat(luot, "2a-phan-tich.md")).strip()
                chung["PHAN_TICH"] = pt
                _ghi_chu(nhap_pt, pt + "\n")

        # Bốn bước viết, chạy lần lượt, mỗi bước ăn kết quả bước trước.
        ban_nhap = _doc_chu(duong_kb)
        if not ban_nhap:
            # ═══ NHỚ LẠI TỪNG BƯỚC, ĐỪNG VIẾT LẠI CẢ BÀI ═══
            #
            # Cả khâu này được `core/auto.chay` chạy lại tới ba lần khi hỏng.
            # Trước đây không bước nào để lại dấu, nên một cú `500` ở bước
            # "đối chiếu và sửa" là vứt luôn bản nháp vừa viết xong rồi viết
            # lại từ đầu — ba lần.
            #
            # Đo thật 15/08/2026: một lượt tiêu 32.505₫ qua ba vòng và không ra
            # bài nào, trong đó phần lớn là viết đi viết lại đúng một bản nháp.
            # Chủ dự án: *"việc viết content tao thấy nó cũng đang api hơi
            # nhiều"* — đây là chỗ nhiều nhất, và nó nhiều một cách vô ích.
            #
            # Giờ mỗi bước ghi kết quả ra một tệp nháp. Chạy lại thì nhặt đúng
            # chỗ đứt, y như cách các khâu lớn nhìn đĩa trước khi làm.
            # Nhãn nhật ký nói đúng luồng hiện tại của kênh: viết N bản → chấm
            # & chọn → hoàn thiện → rà soát. Chủ dự án, 25/08/2026: *"sửa log ở
            # tự động để nó đúng với luồng logic hiện tại"*.
            so_ban = max(1, int(getattr(k, "so_ban_nhap", 1) or 1))
            nhan_viet = ("viết {0} bản kịch bản, rồi chấm & chọn một bản{1}"
                         .format(so_ban, " và hoàn thiện bản đó"
                                 if getattr(k, "hoan_thien", False) else "")
                         if so_ban > 1 else "viết kịch bản")
            # ═══ NẮN ĐỘ DÀI XEN GIỮA VIẾT VÀ RÀ SOÁT (đảo lại 04/09/2026) ═══
            #
            # Trước: viết → rà soát → nắn. Bước rà soát (`3-sua.md`) tách mỗi
            # câu một dòng và chèn thẻ cảm xúc, rồi bước nắn viết lại cả bài —
            # thẻ đặt trên bản chữ cũ không còn khớp, nên mã phải VỨT bản có
            # thẻ đi và bắt khâu giọng đọc chèn lại từ đầu. Tức mỗi lần bước
            # nắn chạy là công của bước rà soát đổ đi một nửa.
            #
            # Nay: viết → NẮN → rà soát. Bước nắn nhận bản chữ liền mạch (thứ
            # nó dễ nén nhất), và thẻ chỉ được chèn một lần, lên bản cuối cùng.
            for ten, nhan in (("2-viet.md", nhan_viet),):
                khuon = k.prompt.get(ten, "")
                if not khuon.strip():
                    continue
                nhap = os.path.join(d, "1-nhap-{0}.txt".format(ten[0]))
                da_co = _doc_chu(nhap).strip()
                if da_co:
                    bc.ghi("  {0} — đã có từ lần trước, dùng lại.".format(nhan))
                    ban_nhap = da_co
                    continue
                bc.kiem_dung()
                bc.ghi("  {0}…".format(nhan))
                if ten == "2-viet.md" and k.so_ban_nhap > 1:
                    # Viết nhiều bản rồi chấm — xem `_viet_nhieu_ban`.
                    ban_nhap = _viet_nhieu_ban(bc, luot, k, chung, khuon,
                                               tu_lieu, muc_tieu_kt, d)
                else:
                    ban_nhap = _goi(
                        bc, _thay(khuon, dict(chung, DRAFT=ban_nhap)),
                        _khoa_chat(luot, ten),
                        toi_da_token=_token_viet(len(tu_lieu), muc_tieu_kt)).strip()
                truoc_go = ban_nhap
                ban_nhap = _don_ban(ban_nhap, k.ngon_ngu)
                if ban_nhap != truoc_go:
                    bc.ghi("  (bỏ {0} ký tự lời dẫn / ghi chú kỹ thuật AI in "
                           "kèm bài)".format(len(truoc_go) - len(ban_nhap)))
                if not ban_nhap:
                    raise RuntimeError("bước “{0}” trả về rỗng".format(nhan))
                _ghi_chu(nhap, ban_nhap + "\n")

            # ═══ NẮN ĐỘ DÀI: ĐO RỒI NẮN, KHÔNG NẮN MÙ ═══
            #
            # Lượt chạy thật đầu tiên ra 2.933/3.410 ký tự — hụt 14%, tức video
            # 8,6 phút thay vì 10. Bước nắn có nhận con số mục tiêu, nhưng lời
            # nhắc chỉ nói "khoảng ngần này ký tự", mà "khoảng" với tiếng Nhật
            # thì AI hiểu rất rộng.
            #
            # Bảo AI *"khoảng 3.410"* là giao một mục tiêu mơ hồ. Bảo nó
            # *"đang có 2.933, thêm khoảng 480 nữa"* là giao một việc đo được.
            # Nên: đo, nói chênh lệch cụ thể, nắn, đo lại — tối đa ba vòng.
            truoc_nan = ban_nhap
            ban_nhap = _nan_do_dai(bc, luot, k, chung, ban_nhap, muc_tieu_kt)

            # ═══ ĐỌC LẠI CHỈ KHI ĐÃ NẮN ═══
            #
            # Bước này là lượt gọi thứ tư trên cùng một bài, và ba lượt trước
            # đã viết rồi soát rồi. Tool gốc (`D:/CONTENT`) gọi `review` **chỉ
            # trong nhánh nắn độ dài** — đúng chỗ nó có việc để làm: cắt hay
            # thêm vài trăm chữ bao giờ cũng để lại chỗ gợn.
            #
            # Bản chưa phải nắn thì nó chỉ là một lượt gọi nữa trên một bài đã
            # ổn. Chủ dự án, 15/08/2026: *"việc viết content tao thấy nó cũng
            # đang api hơi nhiều"*.
            khuon_cuoi = k.prompt.get("5-hoan-thien.md", "")
            if khuon_cuoi.strip() and ban_nhap != truoc_nan:
                bc.kiem_dung()
                bc.ghi("  đọc lại lần cuối…")
                # Bước này CÓ thể làm ngắn lại, nên đo lần nữa sau nó và giữ
                # bản tốt hơn — đọc mượt mà hụt 15% thì vẫn là hỏng.
                cuoi = _goi(bc, _thay(khuon_cuoi,
                                      dict(chung, DRAFT=ban_nhap)),
                            "{0}:chat:5-hoan-thien.md".format(
                                luot.ma_luot)).strip()
                cuoi = _don_ban(cuoi, k.ngon_ngu)
                if cuoi and _lech(cuoi, muc_tieu_kt) <= max(
                        _lech(ban_nhap, muc_tieu_kt), _chenh_cho_phep(k)):
                    ban_nhap = cuoi
                elif cuoi:
                    bc.ghi("  (bỏ bản đọc lại: nó làm lệch {0:.0%}, bản trước "
                           "lệch {1:.0%})".format(_lech(cuoi, muc_tieu_kt),
                                                  _lech(ban_nhap, muc_tieu_kt)))

            # ═══ RÀ SOÁT CHO GIỌNG ĐỌC — BƯỚC CUỐI, SAU KHI ĐỘ DÀI ĐÃ CHỐT ═══
            #
            # `3-sua.md` tách mỗi câu một dòng, chèn thẻ cảm xúc và đặt dấu
            # `---` ngăn phần. Cả ba thứ đều gắn với BẢN CHỮ CỤ THỂ, nên nó
            # phải chạy sau khi không còn ai viết lại bài nữa.
            #
            # Trước 04/09/2026 nó chạy TRƯỚC bước nắn, và mỗi lần nắn phải vứt
            # bản có thẻ đi vì thẻ không còn khớp — công của cả một lượt gọi AI
            # đổ đi. Đặt sau thì thẻ chỉ chèn một lần, lên đúng bản cuối cùng.
            khuon_sua = k.prompt.get("3-sua.md", "")
            if khuon_sua.strip():
                nhap3 = os.path.join(d, "1-nhap-3.txt")
                da_co3 = _doc_chu(nhap3).strip()
                if da_co3:
                    bc.ghi("  rà soát bản cuối — đã có từ lần trước, dùng lại.")
                    ban_nhap = da_co3
                else:
                    bc.kiem_dung()
                    bc.ghi("  rà soát bản cuối: lệch tiếng, tách câu, chèn thẻ "
                           "cảm xúc…")
                    sua = _goi(
                        bc, _thay(khuon_sua, dict(chung, DRAFT=ban_nhap)),
                        _khoa_chat(luot, "3-sua.md"),
                        toi_da_token=_token_viet(len(tu_lieu),
                                                 muc_tieu_kt)).strip()
                    sua = _don_ban(sua, k.ngon_ngu)
                    if not sua:
                        raise RuntimeError("bước “rà soát bản cuối” trả về rỗng")
                    ban_nhap = sua
                    _ghi_chu(nhap3, ban_nhap + "\n")

            # ═══ KỊCH BẢN CÓ SẴN THẺ CẢM XÚC THÌ TÁCH LÀM HAI ═══
            #
            # Chủ dự án, 24/08/2026: *"kết hợp cái review và cài chèn thẻ cảm
            # xúc đi… đơn giản hiệu quả để kịch bản ok nhất, đưa vào voice
            # được luôn"*. Nên bước rà soát trả về bài ĐÃ CÓ THẺ.
            #
            # Nhưng `1-kich-ban.txt` còn được khâu phụ đề (ép chữ lên giọng
            # đọc), khâu ảnh bìa và phép đo độ dài dùng — thẻ lọt vào đó là
            # `[sighs]` hiện lên màn hình. Nên tách ngay tại đây: bản có thẻ
            # để riêng cho giọng đọc (`TEP_CO_THE`, khâu giọng đọc tự nhặt,
            # không gọi AI chèn nữa), bản gỡ thẻ đi tiếp như mọi khi.
            ban_nhap = _tach_the_cam_xuc(bc, d, ban_nhap)

            # Gỡ dấu markdown TRƯỚC khi ghi: tệp này đi thẳng vào bộ đọc giọng
            # nói, và AI hay in đậm mấy chữ nó cho là quan trọng dù lời nhắc đã
            # dặn xuất dạng txt. Xem `go_dinh_dang`.
            from .lam_sach import go_dinh_dang  # noqa: PLC0415
            # Dọn thêm một lượt ngay trước khi ghi: đây là cửa CUỐI trước khi
            # chữ thành tệp, và `ban_nhap` có thể tới từ một tệp nháp do bản
            # tool cũ ghi ra (chạy tiếp một lượt đứt giữa chừng).
            ban_nhap = go_dinh_dang(_don_ban(ban_nhap, k.ngon_ngu))
            _ghi_chu(duong_kb, ban_nhap + "\n")
            # Bài đã xong thì mấy tệp nháp không còn việc gì. Để lại chỉ làm
            # thư mục kết quả rối, và người mở ra không biết tệp nào là bài thật.
            for ten in ("2", "3", "phan-tich"):
                try:
                    os.remove(os.path.join(d, "1-nhap-{0}.txt".format(ten)))
                except OSError:
                    pass

        # Đo bằng chữ ĐỌC LÊN, không phải `len()`: bản này đã qua `3-sua.md`
        # nên nó mang 210–406 ký tự xuống dòng — xem `_do_doc`.
        doc_duoc = _do_doc(ban_nhap)
        lech = _lech(ban_nhap, muc_tieu_kt)
        if muc_tieu_kt > 0:
            bc.ghi("  kịch bản: {3} ký tự ≈ {4} phút đọc (nhắm {5} phút ≈ {1} ký "
                   "tự, lệch {2:.0%}).".format(
                       doc_duoc, muc_tieu_kt, lech, doc_duoc,
                       _phut(doc_duoc, k.ky_tu_moi_phut),
                       _phut(muc_tieu_kt, k.ky_tu_moi_phut)))
        else:
            bc.ghi("  kịch bản: {0} ký tự ≈ {1} phút đọc (độ dài tự do, không nhắm mốc).".format(
                doc_duoc, _phut(doc_duoc, k.ky_tu_moi_phut)))
        _kiem_kich_ban_dung_duoc(doc_duoc, muc_tieu_kt, duong_kb,
                                 tu_do=bool(getattr(k, "do_dai_tu_do", False)))
        _kiem_ban_sach(bc, ban_nhap, duong_kb)

        # SEO — thiếu cũng vẫn ra được video, nên hỏng thì chỉ ghi nhật ký.
        duong_seo = os.path.join(d, "1-seo.txt")
        if k.prompt.get("6-seo.md") and not os.path.exists(duong_seo):
            try:
                bc.kiem_dung()
                # CHANNEL_KEYWORDS từng lấy `style_name` — một KEY TÀI SẢN
                # (vd `blank_white_figure_warm_peach`), và mô hình chép nguyên
                # nó vào KEYWORDS của 1-seo.txt hai lượt liền (0004 + 0005,
                # TL4-T7, 03/09/2026). Tên kênh mới là thứ mô tả kênh.
                # CHAPTERS để trống: mốc thời gian THẬT chỉ có sau khâu phụ
                # đề — `_chen_muc_luc_seo` sẽ chèn 目次 vào lúc đó.
                seo = _goi(bc, _thay(k.prompt["6-seo.md"], dict(
                    chung, SCRIPT_OPENING=ban_nhap[:1500],
                    CHANNEL_KEYWORDS=k.ten, CHAPTERS="")),
                    _khoa_chat(luot, "seo"))
                _ghi_chu(duong_seo, seo)
            except Exception as loi:  # noqa: BLE001
                bc.ghi("  (bỏ qua SEO: {0})".format(str(loi)[:100]))

        # Thẻ cảm xúc **không** làm ở đây. Chủ dự án, 16/08/2026: *"sẽ tách ra
        # khỏi khâu content mà thay vào đó ở khâu voice thì sẽ hợp lý hơn"* —
        # và đúng vậy: thẻ là chỉ đạo cho người đọc, không phải một phần của
        # nội dung. Xem `_chen_the_cam_xuc`, gọi từ khâu giọng đọc.
        _lam_sach_ket_qua(bc, duong_kb, duong_seo,
                          os.path.join(d, "1-tieu-de.txt"))
        return {"so_ky_tu": doc_duoc, "lech": round(lech, 3),
                "tieu_de": tieu_de, "chu_bia": chu_bia}

    # Nhãn cho nhật ký của `core/auto.chay`: khâu này có đi ví hay không.
    lam.khong_tieu_vi = bc_goc.goi_chat_kich_ban is not None
    return lam


def _trung_nguyen_van(moi: str, goc: str, n: int = 10) -> float:
    """Xem `core.viet_nhieu_ban.trung_nguyen_van` — giữ tên cũ cho bài kiểm."""
    from .viet_nhieu_ban import trung_nguyen_van  # noqa: PLC0415

    return trung_nguyen_van(moi, goc, n)


#: Tên tệp lưu từng bản viết và bản chấm — để lại trong thư mục lượt cho chủ
#: kênh soi, không dọn như tệp nháp.
TEP_BAN_VIET = "1-ban-{0}.txt"
TEP_CHAM_DIEM = "1-cham-diem.txt"
TEP_HOOK = "1-hook-{0}.txt"


def _viet_nhieu_ban(bc: BoiCanh, luot: LuotChay, k: Kenh, chung: Dict[str, Any],
                    khuon: str, tu_lieu: str, muc_tieu: int, d: str) -> str:
    """Viết `k.so_ban_nhap` bản rồi chấm, trả về bản được chọn.

    ═══ VÌ SAO ═══

    Chủ dự án, 25/08/2026, sau khi thấy cùng một prompt lượt bám gốc rất tốt
    (0014–0016) lượt lại thay hết nghiên cứu (0013, 0024): *"cho nó viết nhiều
    lần, và chấm điểm các lần tức là chọn bản tốt nhất ok nhất khi viết ví dụ
    3 lần chẳng hạn"*. Làm tay thì chính người viết là người chọn; tool thì
    lấy bản đầu tiên — nên phải có người chấm thay.

    Chấm bằng AI theo `prompt/2b-cham.md`, kèm **số đo tính sẵn** (độ dài so
    mục tiêu, mức trùng nguyên văn với bản gốc) để nó không phải đoán. Chấm
    hỏng — thiếu tệp, JSON lỗi — thì chọn theo số đo: gần mục tiêu độ dài nhất
    trong các bản không chép quá nửa. Từng bản và bản chấm ghi ra đĩa để chủ
    kênh soi lại được.

    Mỗi bản dùng một Idempotency-Key riêng (`2-viet.md:banN`), nên chạy tiếp
    một lượt đứt giữa chừng thì nhặt đúng bản đã viết, không viết lại.
    """
    n = max(1, int(getattr(k, "so_ban_nhap", 1) or 1))
    ban: List[str] = []
    for i in range(n):
        nhan = chr(65 + i)
        tep = os.path.join(d, TEP_BAN_VIET.format(nhan))
        da_co = _doc_chu(tep).strip()
        if da_co:
            bc.ghi("  bản {0} — đã có từ lần trước, dùng lại.".format(nhan))
            ban.append(da_co)
            continue
        bc.kiem_dung()
        bc.ghi("  viết bản {0}/{1}…".format(nhan, n))
        bat_dau = time.time()
        chu = _goi(bc, _thay(khuon, dict(chung, DRAFT="")),
                   _khoa_chat(luot, "2-viet.md:ban{0}".format(i + 1)),
                   toi_da_token=_token_viet(len(tu_lieu), muc_tieu)).strip()
        chu = _don_ban(chu, k.ngon_ngu)
        if chu:
            bc.ghi("  bản {0}: {1} ký tự ≈ {2} phút, mất {3:.0f} giây.".format(
                nhan, len(chu), _phut(len(chu), int(getattr(k, "ky_tu_moi_phut", 0)
                                                     or 0)), time.time() - bat_dau))
        if not chu:
            bc.ghi("  (bản {0} trả về rỗng — bỏ)".format(nhan))
            continue
        _ghi_chu(tep, chu + "\n")
        ban.append(chu)
    if not ban:
        raise RuntimeError("không bản nào viết được")
    if len(ban) == 1:
        return ban[0]

    from .viet_nhieu_ban import cham_va_chon  # noqa: PLC0415

    khuon_cham = k.prompt.get("2b-cham.md", "")

    def goi_cham(loi_nhac: str) -> str:
        bc.kiem_dung()
        return _goi(bc, loi_nhac, _khoa_chat(luot, "2b-cham.md"))

    nhip = int(getattr(k, "ky_tu_moi_phut", 0) or 0)
    if khuon_cham.strip():
        chon, ly_do, diem, bang = cham_va_chon(
            goi_cham, ban, tu_lieu, khuon_cham=khuon_cham, chung=chung,
            muc_tieu=muc_tieu, ghi=bc.ghi, ky_tu_moi_phut=nhip)
    else:
        # Không có prompt chấm → không gọi AI, chọn theo số đo.
        chon, ly_do, diem, bang = cham_va_chon(
            None, ban, tu_lieu, muc_tieu=muc_tieu, ghi=bc.ghi,
            ky_tu_moi_phut=nhip)
    ban_chon = ban[chon]
    ghi_ht = ""
    if getattr(k, "hoan_thien", False):
        # ═══ HOÀN THIỆN BẢN ĐÃ CHỌN, RỒI CHO BỘ CHẤM SO LẠI ═══
        #
        # Chủ dự án, 25/08/2026: "remake với prompt đơn giản vài lần nó sẽ ra
        # bài ok nhất, và chỉnh lại bài đó để hoàn thiện các điểm yếu và nổi
        # bật phát huy điểm tốt, làm mượt lại". Bản hoàn thiện chỉ được dùng
        # khi (a) không viết lại từ đầu — giữ ≥60% câu, độ dài ±25% (mã chốt)
        # và (b) chính bộ chấm, so hai bản cạnh nhau, thích bản mới hơn. Hai
        # cửa để bước này không bao giờ làm bài xấu đi.
        from .viet_nhieu_ban import (dien_o_giu_lai, hoan_thien_ban,  # noqa: PLC0415
                                     tach_diem)

        diem_manh, diem_yeu = tach_diem(ly_do)
        if not (diem_manh or diem_yeu):
            bc.ghi("  (bộ chấm không nêu điểm mạnh/yếu — bỏ qua hoàn thiện)")
        else:
            if diem_yeu:
                bc.ghi("  bộ chấm chê bản {0}: {1}".format(
                    chr(65 + chon), diem_yeu[:220]))

            def goi_ht(loi_nhac: str) -> str:
                bc.kiem_dung()
                return _goi(bc, loi_nhac, _khoa_chat(luot, "2c-hoan-thien.md"),
                            toi_da_token=16384)

            ban_ht, da_ht, ghi_ht = hoan_thien_ban(
                goi_ht, ban_chon, tu_lieu, diem_manh=diem_manh, diem_yeu=diem_yeu,
                ngon_ngu=ten_tieng(k.ngon_ngu), phut=str(chung.get("PHUT", "")),
                chars=int(muc_tieu or 0),
                # KHÔNG dùng `_thay` (xoá ô còn sót) — `<<DRAFT>>`, `<<DIEM_*>>`
                # phải còn nguyên cho `hoan_thien_ban` điền.
                khuon=dien_o_giu_lai(k.prompt.get("2c-hoan-thien.md", ""),
                                     dict(chung, SO_BAN=len(ban))),
                ghi=bc.ghi)
            if da_ht:
                _ghi_chu(os.path.join(d, "1-ban-hoan-thien.txt"), ban_ht + "\n")
                ten_goc = "bản {0} chưa hoàn thiện".format(chr(65 + chon))
                i_hon, ly_do_so, _d, _b = cham_va_chon(
                    goi_cham if khuon_cham.strip() else None,
                    [ban_chon, ban_ht], tu_lieu, khuon_cham=khuon_cham,
                    chung=chung, muc_tieu=muc_tieu, ghi=bc.ghi,
                    ky_tu_moi_phut=nhip,
                    ten_ban=(ten_goc, "bản {0} đã hoàn thiện".format(
                        chr(65 + chon))))
                if i_hon == 1:
                    ban_chon = ban_ht
                    ghi_ht += " — bộ chấm chọn bản hoàn thiện: " + ly_do_so[:200]
                else:
                    ghi_ht += (" — bộ chấm vẫn thích bản chưa hoàn thiện: "
                               + ly_do_so[:200])
                    bc.ghi("  (bản hoàn thiện không hơn — dùng {0})".format(
                        ten_goc))
    # ═══ THAY ĐOẠN MỞ ĐẦU BẰNG HOOK VIẾT RIÊNG ═══
    #
    # Chạy CUỐI CÙNG, sau cả bước hoàn thiện — vì hoàn thiện viết lại cả bài
    # (chỉ buộc giữ ≥60% câu), mà hook chỉ chiếm ~4% bài nên nó thừa sức viết
    # đè lên hook vừa chọn. Đặt sau thì hook thắng cuộc là hook cuối cùng.
    #
    # Hook mới được nhận `<<THAN_BAI>>` — đoạn thân nối ngay sau — nên câu cuối
    # hook dẫn được vào thân, không cần bước làm mượt nào nữa.
    ghi_hook = ""
    so_hook = max(0, int(getattr(k, "so_ban_hook", 0) or 0))
    khuon_hook = k.prompt.get("2d-hook.md", "")
    if so_hook > 1 and khuon_hook.strip():
        from .viet_nhieu_ban import tach_hook, thay_hook  # noqa: PLC0415

        # Mỗi bản một khoá riêng — chạy tiếp lượt đứt giữa chừng thì nhặt đúng
        # bản đã viết, không trả tiền hai lần. Đếm bằng biến đóng vì
        # `thay_hook` gọi hàm viết mà không truyền chỉ số.
        dem_hook = [0]

        def goi_viet_hook(loi_nhac: str) -> str:
            bc.kiem_dung()
            dem_hook[0] += 1
            return _don_ban(_goi(bc, loi_nhac,
                                 _khoa_chat(luot, "2d-hook.md:ban{0}".format(
                                     dem_hook[0])),
                                 toi_da_token=2048), k.ngon_ngu)

        def goi_cham_hook(loi_nhac: str) -> str:
            bc.kiem_dung()
            return _goi(bc, loi_nhac, _khoa_chat(luot, "2e-cham-hook.md"))

        def goi_va_hook(loi_nhac: str) -> str:
            bc.kiem_dung()
            return _don_ban(_goi(bc, loi_nhac,
                                 _khoa_chat(luot, "2f-va-hook.md"),
                                 toi_da_token=2048), k.ngon_ngu)

        def luu_hook(i: int, chu: str) -> None:
            _ghi_chu(os.path.join(d, TEP_HOOK.format(chr(65 + i))), chu + "\n")

        def hook_da_co(i: int) -> str:
            return _doc_chu(os.path.join(d, TEP_HOOK.format(chr(65 + i))))

        bc.ghi("  viết riêng {0} đoạn mở đầu…".format(so_hook))
        try:
            ban_moi, da_thay, ghi_hook = thay_hook(
                goi_viet_hook, goi_cham_hook, ban_chon,
                # ═══ ĐƯA NGUYÊN KỊCH BẢN GỐC, KHÔNG CẮT SẴN ═══
                #
                # Từng cắt lấy "đoạn mở của bản gốc" bằng bộ tách từ khoá rồi
                # mới đưa cho AI. Cắt bằng từ khoá là cách thô: bản gốc có thể
                # không dùng dấu nào trong danh sách, và cắt sai thì chuẩn đối
                # chiếu sai theo — hook viết ra bắt chước một đoạn cụt.
                #
                # Model đủ thông minh để tự nhìn ra bản gốc mở đầu thế nào.
                # Đưa nguyên bài, để nó tự đọc.
                tu_lieu or "",
                so_ban=so_hook, khuon_viet=khuon_hook,
                khuon_cham=k.prompt.get("2e-cham-hook.md", ""),
                # Vá hook chỉ chạy khi kênh bật `hoan_thien` — cùng cờ với
                # bước hoàn thiện thân bài, vì cùng một quyết định: có trả
                # thêm một lượt chữ để sửa theo lời chê hay không.
                khuon_va=(k.prompt.get("2f-va-hook.md", "")
                          if getattr(k, "hoan_thien", False) else ""),
                goi_va=goi_va_hook,
                chung=chung, ghi=bc.ghi,
                luu_ban=luu_hook, da_co=hook_da_co)
            if da_thay:
                ban_chon = ban_moi
                _ghi_chu(os.path.join(d, "1-ban-hook-moi.txt"), ban_chon + "\n")
            else:
                bc.ghi("  (giữ đoạn mở cũ: {0})".format(ghi_hook))
        except Exception as loi:  # noqa: BLE001 — hook hỏng không được vỡ bài
            bc.ghi("  (bỏ qua bước hook: {0})".format(str(loi)[:120]))
            ghi_hook = "lỗi: " + str(loi)[:120]
    _ghi_chu(os.path.join(d, TEP_CHAM_DIEM),
             "{0}\n\nChọn: bản {1}\nĐiểm: {2}\nLý do: {3}\n{4}{5}".format(
                 bang, chr(65 + chon), json.dumps(diem, ensure_ascii=False),
                 ly_do, ("Hoàn thiện: " + ghi_ht + "\n") if ghi_ht else "",
                 ("\n── HOOK ──\n" + ghi_hook + "\n") if ghi_hook else ""))
    bc.ghi("  → dùng bản {0}{1}; các bản và bản chấm nằm trong thư mục lượt "
           "(1-ban-*.txt, 1-cham-diem.txt).".format(
               chr(65 + chon),
               " đã hoàn thiện" if "chọn bản hoàn thiện" in ghi_ht else ""))
    return ban_chon


#: Mã tiếng viết bằng chữ KHÔNG Latinh — với những tiếng này, vài dòng đầu
#: thuần tiếng Anh chắc chắn không phải bài, mà là AI "kể sắp làm gì".
_TIENG_KHONG_LATINH = frozenset(("ja", "zh", "ko", "th", "ar", "ru", "hi",
                                 "he", "el", "ka", "bn", "uk", "fa", "ur"))


def _go_loi_dan_dau(ban: str, ngon_ngu: str) -> str:
    """Bỏ đoạn dẫn tiếng Anh AI in TRƯỚC bài, với kênh viết chữ không Latinh.

    Đo 24/08/2026, lượt 0019: bước sửa trả về *"I'll read the situation,
    verify the two scripts against each other, then produce the fixed txt
    with tags. Let me first…"* rồi mới tới tiếng Nhật. Chốt độ dài không bắt
    được (tổng vẫn đủ dài), và 374 ký tự tiếng Anh ấy đi thẳng vào giọng đọc.

    Chỉ làm với tiếng viết chữ không Latinh: ở đó, dòng không có lấy một chữ
    cái ngoài ASCII thì không thể là lời đọc. Kênh tiếng Việt/Anh không phân
    biệt được nên không đụng. Cắt tới dòng đầu tiên có chữ bản ngữ; đầu bài
    phải trông như văn tiếng Anh (từ ba từ Latinh) mới cắt — một dòng thẻ
    `[short pause]` đứng đầu thì để yên.
    """
    if (ngon_ngu or "").strip().lower()[:2] not in _TIENG_KHONG_LATINH:
        return ban
    dong = (ban or "").splitlines()

    def co_chu_ban_ngu(d: str) -> bool:
        return any((not c.isascii()) and c.isalpha() for c in d)

    i = 0
    while i < len(dong) and not co_chu_ban_ngu(dong[i]):
        i += 1
    if i == 0 or i >= len(dong):
        return ban          # không có gì để cắt, hoặc cả bài không có bản ngữ
    dau = "\n".join(dong[:i])
    if len(re.findall(r"[A-Za-z]{2,}", dau)) < 3:
        return ban
    return "\n".join(dong[i:]).lstrip()


def _don_ban(ban: str, ngon_ngu: str = "") -> str:
    """Dọn một bản kịch bản AI vừa trả về: bỏ lời dẫn và ghi chú kỹ thuật.

    ═══ GỌI Ở MỌI CHỖ AI TRẢ VỀ CHỮ KỊCH BẢN, KHÔNG CHỈ CHỖ ĐẦU ═══

    Khách báo 28/08/2026: kịch bản đem đi đọc có lẫn lời AI tả việc nó vừa
    làm. Đi dò lại thì `_go_loi_dan_dau` chỉ đứng ở **hai** trong **năm** cửa
    AI trả chữ về, và chỉ chạy với tiếng viết chữ không Latinh:

        2-viet.md         có  (nhưng chỉ tiếng không Latinh)
        3-sua.md          có  (nt)
        2c-hoan-thien.md  KHÔNG
        4-do-dai.md       KHÔNG  ← lượt gọi CUỐI của phần lớn kênh
        5-hoan-thien.md   KHÔNG

    Tức bước cuối cùng — bước quyết định chữ nào đi vào giọng đọc — là bước
    không ai gác. Và cả năm cửa đều bỏ ngỏ với kênh tiếng Việt / tiếng Anh.

    Dọn ở đây còn có một tác dụng nữa, quan trọng không kém: **phép đo độ dài
    ăn theo**. Một khối ghi chú 400 ký tự làm bước nắn độ dài tưởng bài đã đủ
    dài, rồi nó cắt bớt 400 ký tự lời đọc thật để bù vào.
    """
    from .lam_sach import go_ghi_chu_ky_thuat  # noqa: PLC0415

    return go_ghi_chu_ky_thuat(_go_loi_dan_dau(ban or "", ngon_ngu)).strip()


def _phut(so_ky_tu: int, ky_tu_moi_phut: int) -> str:
    """Quy ký tự ra phút đọc theo giọng của kênh: 3926 ký tự, 302/phút → "13,0"."""
    return "{0:.1f}".format(so_ky_tu / max(1, int(ky_tu_moi_phut or 1))).replace(
        ".", ",")


def _tach_the_cam_xuc(bc: BoiCanh, thu_muc: str, ban: str) -> str:
    """Bài có thẻ `[…]` thì ghi bản có thẻ ra `TEP_CO_THE`, trả về bản sạch.

    Bài không có thẻ thì trả nguyên văn, không ghi gì — kênh nào không chèn
    thẻ ở bước sửa vẫn chạy y như cũ. Thẻ lạ (AI bịa) bị gỡ trước khi ghi,
    cùng cửa lọc với đường chèn riêng (`loc_the_la`).
    """
    from .the_cam_xuc import (TEP_CO_THE, bo_the, loc_the_la,  # noqa: PLC0415
                              thua_the)

    co_the, da_bo = loc_the_la(ban)
    # AI chèn dày gấp đôi lời dặn (đo 4 lượt thật) — chốt bằng mã.
    co_the = thua_the(co_the)
    sach = bo_the(co_the)
    if sach == co_the:
        return ban
    # Gỡ thẻ để lại khoảng trắng thừa ở đầu câu; dọn cho phụ đề khỏi lệch.
    sach = "\n".join(dong.strip() for dong in sach.splitlines())
    sach = re.sub(r"[ \t]{2,}", " ", sach).strip()
    if da_bo:
        bc.ghi("  (bỏ thẻ không dùng được: {0})".format(
            ", ".join(sorted(set(da_bo))[:5])))
    so_the = len(re.findall(r"\[[a-z][a-z \-]*\]", co_the))
    _ghi_chu(os.path.join(thu_muc, TEP_CO_THE), co_the.strip() + "\n")
    bc.ghi("  kịch bản có sẵn thẻ cảm xúc ({0} thẻ sau khi thưa bớt) — bản có "
           "thẻ để riêng cho giọng đọc, bản sạch cho phụ đề.".format(so_the))
    return sach


#: Dấu ngăn phần trong `prompt/3-sua.md` — tool đổi nó thành một quãng lặng
#: thật, nên nó KHÔNG được đọc lên.
_DAU_NGAN = re.compile(r"(?m)^\s*-{3,}\s*$")


def _do_doc(chu: str) -> int:
    """Số ký tự THẬT SỰ được đọc lên — thước duy nhất đúng cho độ dài video.

    ═══ VÌ SAO KHÔNG DÙNG `len()` (đo 04/09/2026) ═══

    Bước `3-sua.md` tách mỗi câu một dòng cho giọng đọc. Bản cuối vì thế mang
    **210–406 ký tự xuống dòng**, cộng 6–10 dấu `---`. Trên đích 3.926 của
    TL4-T7 thì đó là **5–10%** con số đem đi so — mà không ai đọc chúng.

    Đo bốn lượt thật (`1-kich-ban.txt` so với chính `2-giong-doc.mp3`):

        lượt   thô    sạch   xuống dòng   giọng đọc thật
        0002  3.834   3.404       406       11,97 phút
        0004  4.076   3.848       210       14,90 phút
        0005  4.051   3.820       213       14,84 phút
        0006  4.529   4.143       356       15,05 phút

    Lượt 0006 vượt trần dải ±15% đúng **14 ký tự** trong khi nó mang 356 ký tự
    xuống dòng: bước nắn bị gọi dậy bởi thứ không có trong video. Đo bằng chữ
    sạch thì nó nằm giữa dải, và bước nắn nằm im như nó phải thế.
    """
    return len(re.sub(r"\s+", "", _DAU_NGAN.sub("", _bo_the_cam_xuc(chu))))


def _bo_the_cam_xuc(chu: str) -> str:
    """Gỡ thẻ `[sighs]`, `[long pause]`… — chúng điều khiển giọng, không đọc lên."""
    return re.sub(r"\[[a-z][a-z \-]*\]", "", chu)


def _lech(chu: str, muc_tieu: int) -> float:
    """Kịch bản này lệch bao nhiêu phần trăm so với độ dài nhắm tới.

    `muc_tieu` là số ký tự nhắm tới — hoặc `ky_tu_muc_tieu` của kênh (theo
    phút), hoặc độ dài bản gốc khi kênh bật `do_dai_theo_goc`.

    Đo bằng `_do_doc`, không phải `len()` — xem ghi chú ở đó.
    """
    if muc_tieu <= 0:
        return 0.0  # độ dài tự do: không có gì để lệch
    return abs(_do_doc(chu) - muc_tieu) / max(1, muc_tieu)


def _nan_do_dai(bc: BoiCanh, luot: LuotChay, k: Kenh, chung: Dict[str, Any],
                ban_nhap: str, muc_tieu: Optional[int] = None) -> str:
    """Nắn kịch bản về đúng độ dài, đo lại sau mỗi vòng.

    Trả về bản tốt nhất đo được — kể cả khi hết vòng mà vẫn chưa đạt, vì một
    bản hụt 9% vẫn dùng được, còn ném lỗi ở đây là vứt cả kịch bản đã trả tiền.

    ═══ TAM SAO THẤT BẢN — VÌ SAO LUÔN NẮN TỪ BẢN GỐC ═══

    Chủ dự án, 17/08/2026: *"kịch bản đang không hay"*. Nguyên nhân nằm ở đây.

    Bản trước đưa **bản vừa viết lại** vào viết lại tiếp:

        moi = _goi(... DRAFT=tot_nhat ...)
        tot_nhat = moi          # vòng sau lại lấy chính nó

    Ba vòng như vậy là bản sao của bản sao của bản sao. Mỗi lần viết lại, AI
    làm mượt đi một chút, mất một chi tiết cụ thể, thay một câu sắc bằng một
    câu tròn. Kịch bản không hỏng ở vòng nào cả — nó **nhạt dần**, và không có
    dòng nhật ký nào báo điều đó.

    Tool gốc `D:\\CONTENT` ghi thẳng lý do trong mã, và đó là bài học mua bằng
    kinh nghiệm: *"Lượt 1-3: nén từ BẢN GỐC (đủ chất liệu, tránh tam sao thất
    bản)."* Bản gốc là bản có nhiều chất liệu nhất; mọi bản sau đều nghèo hơn.

    Nên vòng phản hồi ở đây tác động vào **con số mình khai với AI**, không vào
    bản chữ. Khai 3.400 mà ra 2.700 thì lần sau khai cao hơn — bản gốc vẫn
    nguyên vẹn để nắn lại từ đầu.
    """
    khuon = k.prompt.get("4-do-dai.md", "")
    # Không truyền mục tiêu thì lấy theo phút của kênh — nết cũ, và là thứ các
    # bài kiểm tra gọi thẳng `_nan_do_dai` trông đợi.
    if muc_tieu is None:
        muc_tieu = k.ky_tu_muc_tieu
    if muc_tieu <= 0:
        bc.ghi("  độ dài tự do — không nắn độ dài.")
        return ban_nhap
    if not khuon.strip():
        # ═══ THIẾU TỆP LỜI NHẮC THÌ PHẢI NÓI, ĐỪNG TẮT TRONG IM LẶNG ═══
        #
        # Bản 2.31.0 xoá `4-do-dai.md` khỏi bộ lời nhắc gọn, và dòng `return`
        # này lặng lẽ vô hiệu hoá cả bước nắn độ dài. Không ai biết, cho tới khi
        # hai lượt chạy thật ra kịch bản **dài 38% và 84%** so với mục tiêu —
        # tức video 14 và 18 phút thay vì 10.
        #
        # Kiểm tra ngay nếu đọc dòng này khi đang đi tìm lỗi: bước nắn CHỈ chạy
        # khi kênh có tệp `prompt/4-do-dai.md`.
        if abs(_do_doc(ban_nhap) - muc_tieu) > muc_tieu * _chenh_cho_phep(k):
            nhip = int(getattr(k, "ky_tu_moi_phut", 0) or 0)
            bc.ghi("  (bài {0} ký tự ≈ {2} phút, lệch nhắm {1} phút quá 20% — "
                   "kênh không có bước nắn độ dài prompt/4-do-dai.md, giữ "
                   "nguyên)".format(_do_doc(ban_nhap), _phut(muc_tieu, nhip),
                                    _phut(_do_doc(ban_nhap), nhip)))
        return ban_nhap

    dich = muc_tieu
    cho_phep = _chenh_cho_phep(k)
    duoi, tren = dich * (1 - cho_phep), dich * (1 + cho_phep)
    khai = dich                       # lượt đầu khai đúng mục tiêu
    tot_nhat, cach_nhat = ban_nhap, abs(_do_doc(ban_nhap) - dich)

    for vong in range(1, VONG_NAN_TOI_DA + 1):
        if duoi <= _do_doc(tot_nhat) <= tren:
            bc.ghi("  độ dài đạt: {0} ký tự (lệch {1:.0%}).".format(
                _do_doc(tot_nhat), _lech(tot_nhat, muc_tieu)))
            return tot_nhat

        # ═══ LUÔN NẮN TỪ BẢN GỐC ═══
        #
        # Trừ đúng một ca ở dưới. Đây là chỗ quan trọng nhất của cả hàm.
        nguon = ban_nhap
        ghi_chu = ""
        if vong >= VONG_NAN_TOI_DA and tot_nhat is not ban_nhap \
                and _do_doc(tot_nhat) > tren:
            # Vòng cuối, và bản tốt nhất đang DÀI hơn trần: rút gọn từ chính
            # nó. Cắt 5.000 xuống 3.400 dễ hơn nhiều so với nén 18.000 xuống
            # 3.400, và tới đây thì đã hết lượt để thử lại từ đầu.
            nguon = tot_nhat
            khai = dich
            ghi_chu = " (rút từ bản gần nhất)"

        thieu = dich - _do_doc(nguon)
        viec = ("THÊM khoảng {0} ký tự nữa".format(thieu) if thieu > 0
                else "CẮT bớt khoảng {0} ký tự".format(-thieu))
        bc.ghi("  nắn độ dài vòng {0}: khai {1} ký tự{2}…".format(
            vong, khai, ghi_chu))
        try:
            moi = _goi(bc, _thay(khuon, dict(
                chung, DRAFT=nguon, CHARS=khai,
                CHARS_NOW=_do_doc(nguon), CHARS_DELTA=thieu,
                LENGTH_TASK=viec)),
                _khoa_chat(luot, "4-do-dai.md:v{0}".format(vong))).strip()
            # Dọn TRƯỚC khi đo: ghi chú kỹ thuật cũng là ký tự, và bước này
            # quyết định độ dài bằng chính con số đếm được — xem `_don_ban`.
            moi = _don_ban(moi, getattr(k, "ngon_ngu", ""))
        except Exception as loi:  # noqa: BLE001 — nắn hỏng thì giữ bản đang có
            bc.ghi("  (vòng nắn {0} hỏng: {1}) — giữ bản hiện tại.".format(
                vong, str(loi)[:100]))
            return tot_nhat
        if not moi:
            return tot_nhat

        n = _do_doc(moi)
        bc.ghi("    → {0} ký tự (nhắm {1}).".format(n, dich))
        if abs(n - dich) < cach_nhat:
            tot_nhat, cach_nhat = moi, abs(n - dich)
        if duoi <= n <= tren:
            return moi

        # ═══ CHỈNH CON SỐ KHAI, KHÔNG CHỈNH BẢN CHỮ ═══
        #
        # AI không đếm được ký tự lúc viết, và nó vượt/hụt số khai theo một tỉ
        # lệ dao động. Nên vòng phản hồi tác động vào **con số mình khai**, chứ
        # không vào bản chữ: khai 3.400 mà ra 2.700 thì lần sau khai cao hơn.
        #
        # Mũ 0.6 là giảm chấn. Chỉnh thẳng theo tỉ lệ thì số khai nhảy vọt qua
        # lại (tool gốc đo được một cú nhảy 17k → 5k) vì lượng chữ ra tăng
        # nhanh hơn tuyến tính theo số khai.
        he_so = (dich / max(1, n)) ** 0.6
        khai = int(max(dich * 0.3, min(dich * 1.5, khai * he_so)))

    bc.ghi("  độ dài cuối: {0} ký tự (lệch {1:.0%}).".format(
        _do_doc(tot_nhat), _lech(tot_nhat, muc_tieu)))
    return tot_nhat


def _mo_ta_kenh(k: Kenh) -> str:
    """Vài dòng tả kênh, đưa vào mọi lời nhắc để giọng văn không trôi."""
    phan = [k.giong_van or "", str(k.style.get("audience_culture_note") or "")]
    return "\n".join(p for p in phan if p).strip()


def _doc_tieu_de(chu: str):
    tieu_de = chu_bia = ""
    for dong in (chu or "").splitlines():
        thap = dong.strip()
        if thap.upper().startswith("TITLE:"):
            tieu_de = thap.split(":", 1)[1].strip()
        elif thap.upper().startswith("THUMB:"):
            chu_bia = thap.split(":", 1)[1].strip()
    return tieu_de, chu_bia


# ── Khâu 2: giọng đọc ────────────────────────────────────────────────────────


def _khau_giong_doc(bc: BoiCanh):
    def lam(luot: LuotChay, tt: TrangThaiKhau):
        d = luot.thu_muc
        dich = os.path.join(d, "2-giong-doc.mp3")
        if os.path.exists(dich):
            return {"da_co": True}
        kich_ban = _doc_chu(os.path.join(d, "1-kich-ban.txt")).strip()
        if not kich_ban:
            raise RuntimeError("chưa có kịch bản để đọc")
        # ═══ CHÈN THẺ CẢM XÚC — ĐÚNG CHỖ NÀY, KHÔNG SỚM HƠN KHÔNG MUỘN HƠN ═══
        #
        # Trước khâu cắt đoạn: thẻ tính vào trần 1.000 ký tự của cổng, nên nó
        # phải có mặt **trước** lúc chia. Cắt trước rồi chèn sau là đoạn phình
        # quá trần và cổng từ chối — cùng bài học ghi trong tool nội bộ đời
        # trước: *"thêm break tags TRƯỚC khi split"*.
        #
        # Và chỉ khâu này thấy bản có thẻ. Khâu phụ đề vẫn ép **bản sạch** lên
        # giọng đọc, nên `[whispers]` không bao giờ hiện lên màn hình.
        kich_ban = _chen_the_cam_xuc(bc, luot, kich_ban)
        if not bc.kenh.voice_id:
            raise RuntimeError("kênh chưa chọn giọng — điền voice_id vào kenh.yaml")

        doan, nghi_doan = chia_doan_va_nghi(
            kich_ban, giay_nghi=float(getattr(bc.kenh, "giay_nghi_phan",
                                              GIAY_NGHI_PHAN) or 0.0))
        if not doan:
            raise RuntimeError("kịch bản rỗng, không có gì để đọc")
        thu_muc_doan = os.path.join(d, "2-doan")
        os.makedirs(thu_muc_doan, exist_ok=True)
        manh = [os.path.join(thu_muc_doan, "{0:03d}.mp3".format(so))
                for so in range(1, len(doan) + 1)]
        # Ba đoạn chạy cùng lúc thì cũng nên hỏi chung một lượt: mỗi đoạn tự
        # hỏi lấy là ba lần chờ hết nhịp 2 giây đầu tiên cho mỗi đợt.
        theo_doi = SoTheoDoi(bc, nhip=bc.nhip_hoi)
        # Lấy một lần cho cả mẻ: cửa soi âm đầu chạy trong luồng con, đừng để
        # mỗi luồng đi dò FFmpeg lại từ đầu.
        try:
            ffmpeg_doc = bc.ffmpeg or _tim_ffmpeg()
        except Exception:  # noqa: BLE001
            ffmpeg_doc = ""

        def mot_doan(muc):
            so, chu = muc
            tep = manh[so - 1]
            if os.path.exists(tep):
                return so, True
            bc.ghi("  đọc đoạn {0}/{1} ({2} ký tự)…".format(
                so, len(doan), len(chu)))

            def doc(hau_to=""):
                job = _tao_job(
                    bc, bc.client.tts.create,
                    text=chu, voice_id=bc.kenh.voice_id, format="mp3",
                    idempotency_key=khoa_viec(luot, "tts", so, chu,
                                              bc.kenh.voice_id) + hau_to)
                # Đọc một đoạn lâu hơn hẳn tạo một tấm ảnh, nên trần chờ ở đây
                # phải rộng hơn trần chung. Đo 15/08/2026: để trần chung 12
                # phút thì M02 và M03 cùng chết ở khâu này, mỗi lượt mất luôn
                # cả kịch bản đã trả tiền.
                return _cho_job(bc, job, tran=TRAN_CHO_TTS,
                                ten_viec="đoạn {0}".format(so), so=theo_doi)

            # Máy chủ nhận việc rồi bỏ đó — đặt lại bằng khoá MỚI. Hai nấc,
            # và đuôi bám thời gian nên chạy lại lượt cũng ra khoá khác; xem
            # `khoa_thoat_ket` để biết vì sao một nấc là không đủ.
            goi_tts = None
            for _lan in range(3):
                try:
                    goi_tts = doc("" if _lan == 0 else khoa_thoat_ket(_lan))
                    break
                except LoiKetJob:
                    if _lan == 2:
                        raise
                    bc.ghi("    đoạn {0}: máy chủ nhận việc rồi bỏ đó — đặt "
                           "lại bằng khoá mới ({1}/2).".format(so, _lan + 1))
            _tai_ket_qua(bc, goi_tts, 0, tep)
            # ═══ SOI ÂM ĐẦU: NHÀ MÁY HAY XÉN MẤT CHỮ ĐẦU ═══
            #
            # Xem `NGUONG_AM_DAU`. Đọc lại đúng MỘT lần bằng khoá mới — bản
            # thứ hai thường lành, và một lượt đọc lại rẻ hơn nhiều so với
            # một video mở màn bằng chữ cụt.
            if _bi_xen_am_dau(ffmpeg_doc, tep):
                bc.ghi("    đoạn {0}: nhà máy trả về bản bị cắt mất âm đầu — "
                       "đọc lại.".format(so))
                try:
                    _tai_ket_qua(bc, doc(":am-dau"), 0, tep)
                except Exception as loi:  # noqa: BLE001
                    bc.ghi("    đoạn {0}: đọc lại không được ({1}) — giữ bản "
                           "cũ.".format(so, str(loi)[:60]))
                else:
                    if _bi_xen_am_dau(ffmpeg_doc, tep):
                        bc.ghi("    đoạn {0}: bản đọc lại vẫn cụt âm đầu — "
                               "giữ, nhưng chữ đầu đoạn có thể nghe hụt."
                               .format(so))
            return so, False

        # ═══ ĐỌC SONG SONG, VÀ THIẾU MỘT ĐOẠN LÀ HỎNG CẢ KHÂU ═══
        #
        # Cổng cho 3 job đọc cùng lúc. Bản trước chạy tuần tự nên chỉ dùng một
        # suất trong ba — đo được 11,3 phút. Giờ bắn ba đoạn một lượt.
        #
        # Số đoạn thì **không** nắn theo chỗ này: `chia_doan_doc` cắt ít đoạn
        # nhất có thể vì mỗi chỗ nối là một chỗ đổi tông giọng. Muốn nhanh thì
        # tăng suất song song, đừng cắt vụn kịch bản ra — xem ghi chú ở đó.
        #
        # `chiu_thieu=False` vì giọng đọc khác ảnh: thiếu một cảnh thì khâu
        # dựng giữ hình cảnh trước bù vào, còn thiếu một đoạn đọc là **mất hẳn
        # một khúc lời** giữa bài, mà mọi mốc thời gian phía sau (phụ đề, cảnh)
        # đều bám vào file này.
        try:
            _chay_song_song(bc, list(enumerate(doan, start=1)), mot_doan,
                            "đoạn đọc", loai_job="tts", mac_dinh=SONG_SONG_DOC,
                            chiu_thieu=False)
        finally:
            theo_doi.dong()
        _noi_mp3(bc, manh, dich, nghi=nghi_doan)
        doi = _doi_cao_do_giong(bc, dich)
        _lam_sach_ket_qua(bc, dich)
        return {"so_doan": len(manh), "doi_cao_do": doi}

    return lam


#: Đỉnh âm lượng cho phép trong 50 mili giây ĐẦU của một đoạn giọng đọc.
#:
#: Đo 27/08/2026 trên các tệp thật: đoạn lành mở đầu bằng im lặng
#: (−58 … −84 dB), đoạn bị xén mở đầu bằng tiếng (−1,8 … −11 dB). Không có
#: đoạn nào rơi vào giữa, nên mốc −30 dB tách sạch hai nhóm mà không bắt oan.
NGUONG_AM_DAU = -30.0

#: Nghe bao nhiêu mili giây đầu để biết đoạn có bị xén không.
GIAY_NGHE_AM_DAU = 0.05


def _dinh_am_dau(ffmpeg: str, tep: str, giay: float = GIAY_NGHE_AM_DAU) -> Optional[float]:
    """Đỉnh âm lượng (dB) trong `giay` giây đầu của tệp tiếng, hoặc `None`.

    `None` = không đo được (thiếu FFmpeg, tệp lạ). Nơi gọi coi như "không có
    ý kiến" và đi tiếp — cửa soi hỏng không được làm hỏng khâu đọc.
    """
    if not ffmpeg or not os.path.exists(tep):
        return None
    try:
        ket = subprocess.run(
            [ffmpeg, "-hide_banner", "-nostats", "-t", "{0:.3f}".format(giay),
             "-i", tep, "-af", "volumedetect", "-f", "null", "-"],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            creationflags=_co_tao_ffmpeg())
    except OSError:
        return None
    dau = re.search(r"max_volume:\s*(-?[0-9.]+)", ket.stderr or "")
    return float(dau.group(1)) if dau else None


def _bi_xen_am_dau(ffmpeg: str, tep: str) -> bool:
    """Đoạn tiếng này có bị cắt mất âm đầu không. Xem `NGUONG_AM_DAU`."""
    dinh = _dinh_am_dau(ffmpeg, tep)
    return dinh is not None and dinh > NGUONG_AM_DAU


def _tep_im_lang(ffmpeg: str, thu_muc: str, giay: float, mau: str) -> str:
    """Dựng một tệp mp3 im lặng dài `giay`, cùng khuôn tiếng với `mau`.

    Cùng tần số lấy mẫu / số kênh / bitrate thì nối bằng `-c copy` được — khỏi
    mã hoá lại mười lăm phút tiếng chỉ để thêm mấy quãng lặng.
    """
    hz, kenh, bit = 44100, 1, "128k"
    try:
        ket = subprocess.run(
            [ffmpeg, "-hide_banner", "-i", mau, "-f", "null", "-"],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            creationflags=_co_tao_ffmpeg())
        mo = re.search(r"(\d+) Hz, (mono|stereo), [^,]+, (\d+) kb/s", ket.stderr or "")
        if mo:
            hz = int(mo.group(1))
            kenh = 1 if mo.group(2) == "mono" else 2
            bit = "{0}k".format(int(mo.group(3)))
    except OSError:
        pass
    dich = os.path.join(thu_muc, "_im-lang-{0:.0f}ms.mp3".format(giay * 1000))
    if not os.path.exists(dich):
        subprocess.run(
            [ffmpeg, "-y", "-hide_banner", "-loglevel", "error", "-f", "lavfi",
             "-i", "anullsrc=r={0}:cl={1}".format(hz, "mono" if kenh == 1 else "stereo"),
             "-t", "{0:.3f}".format(giay), "-c:a", "libmp3lame", "-b:a", bit, dich],
            check=True, creationflags=_co_tao_ffmpeg())
    return dich


def _noi_mp3(bc: BoiCanh, manh: Sequence[str], dich: str,
             nghi: Optional[Sequence[float]] = None) -> None:
    """Nối các đoạn mp3 thành một file. Dùng FFmpeg, chạy trên máy.

    `nghi[i]` = mấy giây im lặng chèn **sau** đoạn thứ i — nhịp nghỉ giữa các
    phần: khán giả có chỗ chuyển mình, người dựng nhìn sóng âm là thấy ranh
    giới phần để cắt. Xem `chia_doan_va_nghi`.
    """
    nghi = list(nghi or [])
    if len(manh) == 1 and not any(x > 0 for x in nghi[:-1] or [0]):
        import shutil  # noqa: PLC0415

        shutil.copyfile(manh[0], dich)
        return
    ffmpeg = bc.ffmpeg or _tim_ffmpeg()
    if not ffmpeg:
        raise RuntimeError("máy chưa có FFmpeg để nối các đoạn giọng đọc")
    danh_sach = dich + ".txt"
    with open(danh_sach, "w", encoding="utf-8") as tep:
        for i, m in enumerate(manh):
            tep.write("file '{0}'\n".format(os.path.abspath(m).replace("'", "'\\''")))
            giay = float(nghi[i]) if i < len(nghi) else 0.0
            if giay <= 0 or i == len(manh) - 1:
                continue
            try:
                im = _tep_im_lang(ffmpeg, os.path.dirname(os.path.abspath(m)),
                                  giay, m)
            except (OSError, subprocess.SubprocessError):
                # Không dựng được quãng lặng thì thôi — đừng làm hỏng cả khâu
                # đọc vì một nhịp nghỉ.
                continue
            tep.write("file '{0}'\n".format(im.replace("'", "'\\''")))
    lenh = [ffmpeg, "-y", "-hide_banner", "-nostats", "-f", "concat",
            "-safe", "0", "-i", danh_sach, "-c", "copy", dich]
    ket = subprocess.run(lenh, capture_output=True, text=True,
                         encoding="utf-8", errors="replace",
                         creationflags=_co_tao_ffmpeg())
    try:
        os.remove(danh_sach)
    except OSError:
        pass
    if ket.returncode != 0 or not os.path.exists(dich):
        raise RuntimeError("nối giọng đọc hỏng: {0}".format(
            (ket.stderr or "")[-300:]))


def _tim_ffmpeg() -> str:
    from .dung_video import tim_ffmpeg  # noqa: PLC0415

    return tim_ffmpeg()


def _bao_dam_ffmpeg(bc: BoiCanh) -> str:
    """FFmpeg đủ dựng được video — thiếu thì TẢI VỀ THƯ MỤC TOOL rồi dùng.

    ═══ ĐỪNG BẮT KHÁCH TỰ ĐI CÀI Ở KHÂU CUỐI ═══

    Tới đây khách đã trả tiền cho kịch bản, giọng đọc, 99 ảnh và 99 clip. Báo
    *"máy chưa có FFmpeg"* rồi dừng là để họ ôm đủ nguyên liệu của một video
    mười phút mà không có video nào — đúng cái cảnh khâu ảnh bìa từng gây ra
    (xem `core.auto.KHAU_KHONG_CHAN`).

    Tải 40 MB mất chưa tới một phút, và chỉ mất đúng một lần cho cả đời máy đó.
    Nó rẻ hơn mọi cách khác, kể cả cách bảo khách mở trình duyệt đi tìm.

    `bc.ffmpeg` có sẵn thì tin ngay, không soi lại: đó là đường khai đè của bài
    kiểm và của người gọi biết rõ mình đang làm gì.
    """
    from .ffmpeg_goi_san import cai_ffmpeg, thieu_gi  # noqa: PLC0415

    if bc.ffmpeg:
        return bc.ffmpeg
    ffmpeg = _tim_ffmpeg()
    if ffmpeg:
        thieu = thieu_gi(ffmpeg)
        if not thieu:
            return ffmpeg
        bc.ghi("  bản FFmpeg trên máy này thiếu {0} — không dựng đủ được, tải "
               "bản đầy đủ về thư mục tool.".format(", ".join(thieu)))
    else:
        bc.ghi("  máy chưa có FFmpeg — tải một bản về thư mục tool.")
    return cai_ffmpeg(bc.goc, bao=bc.ghi)


# ── Khâu 3: phụ đề ───────────────────────────────────────────────────────────


# ── Mục lục (目次) cho mô tả video ───────────────────────────────────────────
#
# Khâu SEO viết mô tả từ lúc kịch bản mới xong — khi CHƯA có giọng đọc, nên
# không thể biết chương nào rơi vào phút nào. Kịch bản lại đặt sẵn một dòng
# `---` ở ranh giới mỗi phần (xem `giay_nghi_phan`), và dòng ấy đi nguyên vào
# SRT — nên SRT là nơi DUY NHẤT vừa có nhãn phần vừa có mốc thời gian THẬT.
# Vậy: chèn 目次 vào 1-seo.txt ngay sau khi SRT ra đời, bằng ghép chuỗi thuần,
# không tốn một lượt gọi AI nào. Chủ dự án yêu cầu 03/09/2026: "fix tool để
# về sau nó làm đúng loại mô tả có [timestamps]".


def _muc_luc_tu_srt(srt: str) -> "list[str]":
    """Rút các dòng chương `MM:SS nhãn` từ dấu `---` trong phụ đề.

    Chương đầu luôn là `00:00` lấy câu mở màn làm nhãn. Trả `[]` khi không đủ
    3 chương — YouTube chỉ nhận mục lục từ 3 mốc trở lên, mốc đầu tại 00:00.
    """
    cau = []  # (giây bắt đầu, chữ)
    for khoi in srt.replace("\r\n", "\n").split("\n\n"):
        dong = [x for x in khoi.strip().split("\n") if x.strip()]
        if len(dong) < 3 or "-->" not in dong[1]:
            continue
        giay = _giay_srt(dong[1].split("-->")[0].strip())
        cau.append((giay, " ".join(dong[2:]).strip()))
    if not cau:
        return []

    def nhan(chu: str) -> str:
        return chu.lstrip("-—– ").strip().rstrip("。.、,").strip()[:60]

    muc = [(0.0, nhan(cau[0][1]))]
    for i, (giay, chu) in enumerate(cau):
        if chu.startswith("---") and giay >= 10:
            chu = nhan(chu)
            # Dòng `---` đôi khi chỉ là câu chuyển ("では、") — cụt quá thì
            # không làm nhãn chương được, ghép thêm câu ngay sau nó.
            if len(chu) < 6 and i + 1 < len(cau):
                chu = (chu + "、" + nhan(cau[i + 1][1])).lstrip("、")
            muc.append((giay, chu))
    if len(muc) < 3:
        return []
    ra = []
    for giay, chu in muc:
        phut, s = divmod(int(giay), 60)
        gio, phut = divmod(phut, 60)
        moc = ("{0}:{1:02d}:{2:02d}".format(gio, phut, s) if gio
               else "{0:02d}:{1:02d}".format(phut, s))
        ra.append("{0} {1}".format(moc, chu))
    return ra


def _chen_muc_luc_seo(bc: BoiCanh, d: str, duong_srt: str) -> None:
    """Chèn khối 目次 (mốc thật từ SRT) vào DESCRIPTION trong `1-seo.txt`.

    Không bao giờ ném lỗi ra ngoài, và chạy lại không chèn đúp: mô tả thiếu
    mục lục vẫn đăng được video, không đáng làm vỡ khâu phụ đề.
    """
    try:
        duong_seo = os.path.join(d, "1-seo.txt")
        seo = _doc_chu(duong_seo)
        if not seo or "目次" in seo or "Chapters\n" in seo:
            return
        muc = _muc_luc_tu_srt(_doc_chu(duong_srt))
        if not muc:
            return
        dau = "📌 目次" if bc.kenh.ngon_ngu == "ja" else "📌 Chapters"
        khoi = ("━━━━━━━━━━━━━━\n{0}\n".format(dau)
                + "\n".join(muc) + "\n━━━━━━━━━━━━━━")
        dong = seo.split("\n")
        # Chèn cuối DESCRIPTION: trước nhãn HASHTAGS:, và trước cả dòng
        # hashtag chốt mô tả nếu có — mục lục nằm trong mô tả, hashtag vẫn cuối.
        vi_tri = None
        for i, dg in enumerate(dong):
            if dg.strip().upper().startswith("HASHTAGS:"):
                vi_tri = i
                break
        if vi_tri is None:
            dong += ["", khoi]
        else:
            j = vi_tri - 1
            while j >= 0 and not dong[j].strip():
                j -= 1
            if j >= 0 and dong[j].lstrip().startswith("#"):
                vi_tri = j
            dong[vi_tri:vi_tri] = [khoi, ""]
        _ghi_chu(duong_seo, "\n".join(dong))
        bc.ghi("  đã chèn 目次 {0} chương (mốc thật từ SRT) vào 1-seo.txt."
               .format(len(muc)))
    except Exception as loi:  # noqa: BLE001
        bc.ghi("  (không chèn được 目次 vào 1-seo.txt: {0})".format(
            str(loi)[:100]))


def _khau_phu_de(bc: BoiCanh):
    def lam(luot: LuotChay, tt: TrangThaiKhau):
        from .phu_de import (  # noqa: PLC0415
            do_khop_voi_kich_ban, tao_phu_de, viet_srt,
        )

        d = luot.thu_muc
        dich = os.path.join(d, "3-phu-de.srt")
        if os.path.exists(dich):
            # SRT có sẵn từ lượt trước nhưng 1-seo.txt có thể chưa có mục lục
            # (tệp do bản tool cũ ghi) — chèn bù, hàm tự bỏ qua nếu đã có.
            _chen_muc_luc_seo(bc, d, dich)
            return {"da_co": True}
        mp3 = os.path.join(d, "2-giong-doc.mp3")
        if not os.path.exists(mp3):
            raise RuntimeError("chưa có file giọng đọc")
        kich_ban = _doc_chu(os.path.join(d, "1-kich-ban.txt"))
        bc.ghi("  đang ép kịch bản khớp vào giọng đọc (chạy trên máy)…")
        ket = tao_phu_de(mp3, kich_ban, ngon_ngu=bc.kenh.ngon_ngu,
                         nghe=bc.nghe, cancel=bc.cancel, on_log=bc.on_log)
        if not ket.cau:
            raise RuntimeError(ket.loi or "không tạo được phụ đề")
        viet_srt(dich, ket.cau)
        # ═══ SOI LẠI THỨ MÌNH VỪA GHI RA ═══
        #
        # Khách báo 28/08/2026: *"srt bị sai nội dung"*. Một tệp `.srt` sai chữ
        # trông y hệt một tệp đúng — không có dòng nào đỏ lên, khâu sau vẫn
        # chạy ngon, và chỗ phát hiện ra là lúc video đã lên sóng.
        #
        # Chỗ duy nhất soi được là ngay đây, khi kịch bản và phụ đề còn nằm
        # cạnh nhau. So một lần tốn vài mili giây, và nó biến một lỗi câm
        # thành một dòng nhật ký có con số.
        khop_chu = do_khop_voi_kich_ban(ket.cau, kich_ban)
        if khop_chu < 0.99:
            bc.ghi("  ⚠ chữ trong phụ đề chỉ trùng kịch bản {0:.0%} — báo lại "
                   "chỗ này, đây là lỗi của tool chứ không phải của bạn."
                   .format(khop_chu))
        if ket.moc_uoc_luong:
            bc.ghi("  LƯU Ý: chữ đúng nguyên kịch bản, nhưng mốc thời gian chỉ "
                   "là ước lượng (máy chưa nghe được file giọng đọc) — câu có "
                   "thể hiện sớm hoặc muộn vài phần mười giây.")
        _lam_sach_ket_qua(bc, dich)
        _chen_muc_luc_seo(bc, d, dich)
        return {"so_cau": len(ket.cau), "khop": round(ket.ty_le_khop, 3),
                "khop_chu": round(khop_chu, 3), "dang_tin": ket.dang_tin}

    return lam


# ── Khâu 4: bảng cảnh ────────────────────────────────────────────────────────


def _canh_dung_duoc(canh, bc: BoiCanh):
    """Soi bảng cảnh lấy từ tệp cũ. Dùng được thì trả về, hỏng thì trả `None`.

    ═══ VÌ SAO PHẢI SOI LẠI THỨ MÌNH TỰ GHI RA ═══

    Mọi khâu đều nhìn đĩa trước và bỏ qua nếu đã có tệp — đó là thứ giữ cho
    "Chạy tiếp" không trả tiền hai lần. Nhưng nó cũng nghĩa là **tệp do một bản
    tool cũ ghi ra vẫn được tin tuyệt đối**, kể cả khi bản ấy có lỗi.

    Xảy ra đúng thế ngày 15/08/2026. Một bản trước đó ghi ra bảng cảnh mà phần
    lớn dòng không có lời nhắc — đo lại: 83/91, 50/90, 52/108 cảnh trống. Chốt
    chặn được thêm vào sau đó nằm trong khâu cắt cảnh, mà khâu ấy thấy tệp có
    sẵn nên bỏ qua, nên chốt không bao giờ được hỏi tới. Khâu ảnh cứ thế gửi
    lời nhắc rỗng đi cho tới khi cổng từ chối — một lượt tiêu 48.000₫ ra 56 tấm
    ảnh rồi chết.

    Nên: tệp cũ vẫn dùng, nhưng phải qua cửa này trước. Hỏng thì làm lại khâu
    cắt cảnh — một lượt gọi AI, rẻ hơn nhiều lần so với trả tiền cho ảnh hỏng.
    """
    if not isinstance(canh, list) or not canh:
        return None
    thieu = [c for c in canh
             if not str((c or {}).get("img_prompt") or "").strip()]
    if not thieu:
        return canh
    bc.ghi("  bảng cảnh cũ có {0}/{1} cảnh thiếu lời nhắc — bỏ, cắt lại từ "
           "đầu.".format(len(thieu), len(canh)))
    return None


def _khau_bang_canh(bc: BoiCanh):
    def lam(luot: LuotChay, tt: TrangThaiKhau):
        from .srt_scenes import parse_srt  # noqa: PLC0415

        d = luot.thu_muc
        dich = os.path.join(d, "4-canh.xlsx")
        goi_json = os.path.join(d, "4-canh.json")
        srt = _doc_chu(os.path.join(d, "3-phu-de.srt"))
        if not srt.strip():
            raise RuntimeError("chưa có phụ đề để cắt cảnh")

        # ═══ NHÁNH ĐẠO DIỄN — chỉ khi kenh.yaml khai `che_do_ke` ═══
        #
        # Kênh truyện nhiều nhân vật (story-3d) đi dây chuyền của Prompt Visuals:
        # dàn nhân vật có giai đoạn + bối cảnh + kế hoạch + ảnh tham chiếu từng
        # nhân vật. Kênh không khai thì đường cũ y nguyên (xem core/dao_dien_auto).
        from .dao_dien_auto import (  # noqa: PLC0415
            TEP_DAN, chay_dao_dien, che_do_dao_dien, tao_tham_chieu,
        )

        if che_do_dao_dien(bc.kenh):
            canh = None
            man: Dict[str, Any] = {}
            if os.path.exists(goi_json) and os.path.exists(os.path.join(d, TEP_DAN)):
                canh = _canh_dung_duoc(json.loads(_doc_chu(goi_json)), bc)
                try:
                    man = json.loads(_doc_chu(os.path.join(d, TEP_DAN)))
                except ValueError:
                    man = {}
            if not canh:
                canh, man = chay_dao_dien(bc, luot)
                _ghi_chu(goi_json, json.dumps(canh, ensure_ascii=False, indent=1))
            thieu = tao_tham_chieu(bc, luot, man, canh=canh)
            # Thiết kế lại nhân vật có thể đã sửa khối khoá trong `canh`.
            _ghi_chu(goi_json, json.dumps(canh, ensure_ascii=False, indent=1))
            if thieu:
                bc.ghi("  [CHÚ Ý] thiếu ảnh tham chiếu: {0} — cảnh có chúng sẽ "
                       "mỗi cảnh một kiểu. Sửa mô tả trong 4-canh-dan.json rồi "
                       "“Làm lại khâu này”.".format(", ".join(thieu)))
            _viet_xlsx(dich, canh, bc.kenh, ke_hoach=list(man.get("director_plan") or []),
                       dan=list(man.get("characters") or []),
                       boi_canh=list(man.get("locations") or []))
            return {"so_canh": len(canh), "so_nhan_vat": len(man.get("characters") or []),
                    "so_boi_canh": len(man.get("locations") or []),
                    "thieu_tham_chieu": thieu}

        if os.path.exists(goi_json):
            canh = _canh_dung_duoc(json.loads(_doc_chu(goi_json)), bc)
        if not os.path.exists(goi_json) or not canh:
            cue = parse_srt(srt)
            if not cue:
                raise RuntimeError("phụ đề không đọc được dòng nào")
            khuon = bc.kenh.prompt.get("7-canh.md", "")
            if not khuon.strip():
                raise RuntimeError("kênh thiếu lời nhắc `7-canh.md`")
            canh = _chia_canh_theo_nghia(bc, luot, cue, khuon)
            _ghi_chu(goi_json, json.dumps(canh, ensure_ascii=False, indent=1))

        ke_hoach: List[Dict[str, Any]] = []
        tep_kh = os.path.join(d, TEP_KE_HOACH)
        if os.path.exists(tep_kh):
            try:
                ke_hoach = json.loads(_doc_chu(tep_kh)) or []
            except (ValueError, TypeError):
                ke_hoach = []
        _viet_xlsx(dich, canh, bc.kenh, ke_hoach=ke_hoach)
        return {"so_canh": len(canh), "so_chuong": len(ke_hoach)}

    def soi_lai(luot: LuotChay) -> bool:
        """Bảng cảnh đã ghi ra còn dùng được không — hỏi TRƯỚC khi bỏ qua khâu.

        Cửa `_canh_dung_duoc` ở trên chỉ chạy khi khâu này chạy. Một bản tool cũ
        ghi ra bảng cảnh thiếu lời nhắc rồi đánh dấu xong thì khâu không chạy
        nữa, và cửa ấy nằm im trong khi khâu ảnh phía sau cứ thế tiêu tiền cho
        tới lúc gặp dòng trống đầu tiên. Xem `core.auto._con_dung_duoc`.
        """
        goi_json = os.path.join(luot.thu_muc, "4-canh.json")
        if not os.path.exists(goi_json):
            # Chưa có tệp thì không phải việc của cửa này: khâu ảnh sẽ báo
            # thiếu bảng cảnh bằng câu của nó, rõ hơn câu ở đây.
            return True
        return _canh_dung_duoc(json.loads(_doc_chu(goi_json)), bc) is not None

    lam.soi_lai = soi_lai
    return lam


def _chia_canh_theo_nghia(bc: BoiCanh, luot: LuotChay, cue: List[Dict[str, Any]],
                          khuon: str) -> List[Dict[str, Any]]:
    """Đưa phụ đề cho AI **tự chia cảnh theo nghĩa**, rồi dựng bảng cảnh.

    Cách chia nằm ở `core/chia_canh.py` — chung với tab Prompt Visuals. Ở đây
    chỉ còn phần riêng của tab Tự động: lời nhắc `7-canh.md` của kênh, và lời
    gọi AI đi qua `BoiCanh.goi_chat` (có khoá idempotency theo lượt chạy, nên
    chạy tiếp không trả tiền hai lần).
    """
    from .srt_scenes import max_seconds_for  # noqa: PLC0415

    tran = float(max_seconds_for(bc.kenh.engine))
    # Mạch chia (bao nhiêu giây một cảnh) là hai con số trong khối PACING của
    # chính `7-canh.md` — chủ kênh sửa số trong prompt là đổi mạch. Trần một
    # ý được lớn hơn clip engine; `chia_theo_nghia` vẫn cắt theo `tran` engine.
    cap = nhip_tu_khuon(khuon)
    san, tran_y = cap if cap else (float(MIN_GIAY_CANH), tran)
    if cap:
        bc.ghi("  mạch chia trong prompt: {0:.0f}–{1:.0f} giây một cảnh.".format(
            san, tran_y))
    # Bản đồ hình cho CẢ bài, lập TRƯỚC khi chia khúc — để 9 khúc chạy song
    # song cùng một thế giới. Không có (kênh thiếu `7-ke-hoach.md`, hay AI trả
    # rác) thì `[]`, và các khúc chia y như trước.
    ke_hoach = _ke_hoach_hinh(bc, luot, cue)

    def hoi(khuc, thu_tu, tong_khuc):
        return _hoi_chia_canh(bc, luot, khuon, list(khuc), thu_tu, tong_khuc,
                              tran_y, ke_hoach=ke_hoach, san=san, clip=tran)

    # `duoi`: ep "no text, no letters…" vao cuoi moi prompt bang ma — luoi an
    # toan hoc tu D:\AFFILIATE, phong khi AI quen duoi o mot cang.
    return chia_theo_nghia(cue, hoi, tran=tran, san=min(san, tran - 1.0),
                           nhan_vat_mac_dinh="nv1",
                           ghi=bc.ghi, kiem_dung=bc.kiem_dung, duoi=DUOI_CAM)


#: Trần chữ cho lượt lập bản đồ hình: mươi chương, mỗi chương vài dòng.
TOKEN_KE_HOACH = 8192

#: Tệp bản đồ hình trong thư mục lượt. Giữ lại để "chạy tiếp" không hỏi AI
#: lần hai, và để người dùng mở ra xem video được chia chương ra sao.
TEP_KE_HOACH = "4-ke-hoach.json"


def _doc_tieu_de_luot(luot: LuotChay) -> str:
    """Tiêu đề đã chốt của lượt (dòng `TITLE:` trong `1-tieu-de.txt`), hoặc rỗng."""
    chu = _doc_chu(os.path.join(luot.thu_muc, "1-tieu-de.txt"))
    for dong in chu.splitlines():
        if dong.strip().upper().startswith("TITLE:"):
            return dong.split(":", 1)[1].strip()
    return ""


def _ke_hoach_hinh(bc: BoiCanh, luot: LuotChay,
                   cue: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Lập bản đồ hình cho cả video: chương → bối cảnh, vật ẩn dụ, câu bản lề.

    ═══ VÌ SAO KHÔNG BAO GIỜ ĐƯỢC LÀM HỎNG LƯỢT CHẠY ═══

    Bản đồ là thứ làm video HAY HƠN, không phải thứ làm video CÓ ĐƯỢC. Kênh
    thiếu `7-ke-hoach.md`, AI trả JSON hỏng, cổng nghẽn — mọi đường hỏng đều
    trả `[]` và ghi một dòng nhật ký; khâu chia cảnh vẫn chạy, chỉ là không có
    bản đồ (đúng như mọi lượt trước 25/08/2026). Ném lỗi ở đây là đổi một video
    kém liền mạch lấy một lượt chạy chết — mà kịch bản và giọng đọc đã trả tiền.

    Lý do có bước này nằm ở `core/chia_canh.sach_ke_hoach`.
    """
    khuon = bc.kenh.prompt.get("7-ke-hoach.md", "")
    if not khuon.strip() or not cue:
        return []
    tep = os.path.join(luot.thu_muc, TEP_KE_HOACH)
    if os.path.exists(tep):
        try:
            cu = sach_ke_hoach(json.loads(_doc_chu(tep)), cue)
        except (ValueError, TypeError):
            cu = []
        if cu:
            bc.ghi("  bản đồ hình có sẵn: {0} chương.".format(len(cu)))
            return cu
    st = bc.kenh.style
    dong = bang_phu_de(cue)
    tong_giay = 0.0
    try:
        tong_giay = float(cue[-1]["end"])
    except (TypeError, ValueError, KeyError):
        pass
    from .chia_canh import CUE_MOI_KHUC  # noqa: PLC0415
    loi_nhac = dien_khuon(khuon, {
        "SRT": dong,
        "DONG_CUOI": cue[-1]["index"],
        "TONG_GIAY": "{0:.0f}".format(tong_giay),
        # ~1 chương / 90 giây, ít nhất 3 — cùng lý lẽ với
        # `prompt.workbook._so_man_goi_y`: đừng ghi chết một con số cho mọi
        # độ dài.
        "SO_CHUONG": max(3, int(round(tong_giay / 90.0))),
        "SO_KHUC": max(1, -(-len(cue) // CUE_MOI_KHUC)),
        "TITLE": _doc_tieu_de_luot(luot),
        "AUDIENCE_LANGUAGE": st.get("audience_language", bc.kenh.ngon_ngu),
        "AUDIENCE_CULTURE_NOTE": st.get("audience_culture_note", ""),
        "CULTURAL_PROPS": st.get("cultural_props", ""),
        "CULTURAL_METAPHORS": st.get("cultural_metaphors", ""),
    })
    bc.ghi("  lập bản đồ hình cho cả bài ({0} dòng, {1:.0f} giây)…".format(
        len(cue), tong_giay))
    for lan in range(2):
        try:
            # `loi_nhac` nằm trong khoá vì bản đồ này dựng từ bản gỡ băng VÀ
            # từ khuôn kịch bản của kênh — khách sửa khuôn ở tab Nâng cao là
            # lời nhắc khác hẳn, trong khi `dong`/`lan` không đổi. Thiếu nó
            # thì cổng báo "khoá này đã dùng cho nội dung khác" và lượt chạy
            # kẹt ở đây mãi. Xem `khoa_viec` và ca lượt 0016 ghi ở
            # `core/goi_van_ban.py`.
            tra = _goi(bc, loi_nhac,
                       khoa_viec(luot, "ke-hoach", 0, dong, lan, loi_nhac),
                       toi_da_token=TOKEN_KE_HOACH)
            ke_hoach = sach_ke_hoach(loc_json(tra), cue)
        except Exception as loi:  # noqa: BLE001
            # Cùng lý lẽ với vòng hỏi lại của `_hoi_chia_canh`: khoá kẹt thì
            # đổi khoá, không đợi. Xem ghi chú dài ở đó.
            bc.ghi("  bản đồ hình chưa ra ({0}){1}".format(
                str(loi)[:70], " — hỏi lại bằng khoá mới." if lan == 0 else "."))
            continue
        if not ke_hoach:
            bc.ghi("  AI trả bản đồ hình không dùng được{0}".format(
                " — hỏi lại." if lan == 0 else "."))
            continue
        _ghi_chu(tep, json.dumps(ke_hoach, ensure_ascii=False, indent=1))
        bc.ghi("  bản đồ hình: {0} chương — {1}".format(
            len(ke_hoach), "; ".join(
                (c.get("place") or c.get("title") or "?")[:40]
                for c in ke_hoach)))
        return ke_hoach
    bc.ghi("  không lập được bản đồ hình — chia cảnh không có bản đồ.")
    return []


def _hoi_chia_canh(bc: BoiCanh, luot: LuotChay, khuon: str,
                   cue: List[Dict[str, Any]], thu_tu: int, tong_khuc: int,
                   tran: float,
                   ke_hoach: Optional[List[Dict[str, Any]]] = None,
                   san: float = MIN_GIAY_CANH, clip: float = 0.0,
                   ) -> List[Dict[str, Any]]:
    """Hỏi AI chia một khúc phụ đề. Trả về nguyên thứ nó đưa ra.

    `tran` là trần MỘT Ý ghi trong prompt; `clip` là clip engine (`<<CLIP_SEC>>`).
    Canh lại là việc của `core.chia_canh.canh_lai`, không phải của chỗ này.
    """
    st = bc.kenh.style
    dong = bang_phu_de(cue)
    loi_nhac = loi_nhac_chia(khuon, cue, tran, san=san, clip=clip, them={
        # Phần bản đồ hình chạm vào khúc này (rỗng nếu không có bản đồ —
        # `dien_khuon` dọn chỗ trống, lời nhắc y như cũ).
        "KE_HOACH": khoi_ke_hoach(ke_hoach or [], cue),
        "IMAGE_STYLE": st.get("image_style", ""),
        "VIDEO_STYLE": st.get("video_style", ""),
        "PALETTE": st.get("palette", ""),
        "REFERENCE_LOCK": st.get("reference_lock", ""),
        "NEGATIVE_PROMPT": st.get("negative_prompt", ""),
        "AUDIENCE_LANGUAGE": st.get("audience_language", bc.kenh.ngon_ngu),
        "AUDIENCE_CULTURE_NOTE": st.get("audience_culture_note", ""),
        "CULTURAL_PROPS": st.get("cultural_props", ""),
        "CULTURAL_METAPHORS": st.get("cultural_metaphors", ""),
        # ═══ LỜI NHẮC PHẢI BIẾT NÓ ĐANG Ở KHÚC MẤY ═══
        #
        # Video ở đây **dài và ngang**, chia thành nhiều khúc, mỗi khúc một lượt
        # gọi riêng — khác hẳn tool `D:\AFFILIATE` vốn làm video dọc ngắn xong
        # trong một lượt.
        #
        # Khác biệt ấy đổi hẳn vài luật. "Cảnh đầu là cú hook" chỉ đúng với
        # khúc 1; bê nguyên sang khúc 5 là mỗi khúc lại dựng một cú mở màn, và
        # video thành năm lần mở bài. Nên đưa vị trí khúc vào để lời nhắc tự
        # biết đường xử.
        "KHUC_THU": thu_tu + 1,
        "TONG_KHUC": tong_khuc,
        "LA_KHUC_DAU": "yes" if thu_tu == 0 else "no",
        "TY_LE_KHUNG": st.get("aspect_ratio", "16:9 horizontal"),
    })
    bc.ghi("  khúc {0}/{1}: {2} dòng ({3}–{4})…".format(
        thu_tu + 1, tong_khuc, len(cue), cue[0]["index"], cue[-1]["index"]))

    def mot_lan(lan: int):
        # Bản đồ nằm trong khoá: đổi bản đồ (làm lại khâu) là lời nhắc khác,
        # phải là một lượt gọi khác — cùng luật với ảnh tham chiếu ở khâu ảnh.
        #
        # Trước đây chỗ này chỉ cho lời nhắc vào khoá KHI CÓ bản đồ, cho rằng
        # không bản đồ thì lời nhắc đứng yên. Không đúng: lời nhắc còn mang
        # khuôn chia cảnh của kênh, mà khuôn ấy khách sửa được. Bỏ điều kiện
        # đi thì luật thành một câu: lời nhắc nào, khoá nấy. Chạy lại y nguyên
        # vẫn ra đúng khoá cũ nên không trả tiền lần hai.
        tra = _goi(bc, loi_nhac,
                   khoa_viec(luot, "canh", cue[0]["index"], dong, lan, loi_nhac),
                   toi_da_token=TOKEN_CANH)
        try:
            goi = loc_json(tra)
        except (ValueError, TypeError) as loi:
            # `json.loads` ném `JSONDecodeError` — con của `ValueError`, KHÔNG
            # phải `LoiNoiDung`. Không gói lại thì vòng hỏi lại ở dưới không
            # bắt được, và câu trả lời đứt cụt làm hỏng luôn cả khúc dù chỉ cần
            # hỏi lại một lần là xong. Đo được bằng bài kiểm: chỉ gọi 1 lần rồi
            # bỏ cuộc.
            raise LoiNoiDung(str(loi)) from loi
        ds = goi.get("scenes") if isinstance(goi, dict) else goi
        if not isinstance(ds, list) or not ds:
            raise LoiNoiDung("AI không trả về danh sách `scenes`")
        # ═══ KIỂM LỜI NHẮC NGAY Ở ĐÂY, ĐỪNG ĐỂ `canh_lai` KIỂM HỘ ═══
        #
        # `canh_lai` cũng kiểm, nhưng nó chạy **sau** vòng hỏi lại này — nên
        # một khúc trả về đủ `scenes` mà rỗng lời nhắc sẽ lọt qua đây, rồi làm
        # gãy cả khâu ở tầng trên.
        #
        # Lượt chạy thật S03 ngày 18/08/2026: `11/11 cảnh thiếu lời nhắc`, và
        # vì hỏng ở tầng trên nên `core/auto.chay` làm lại **cả 18 khúc** —
        # ba lần, mất 11 phút, mỗi lần lại hỏng đúng ở đó.
        #
        # Kiểm ở đây thì đúng một khúc được hỏi lại, bằng một khoá khác, và
        # 17 khúc kia giữ nguyên.
        rong = [c for c in ds if isinstance(c, dict)
                and not (str(c.get("img_prompt") or "").strip()
                         and str(c.get("video_prompt") or "").strip())]
        if rong:
            raise LoiNoiDung(
                "khúc {0}/{1}: {2}/{3} cảnh thiếu lời nhắc".format(
                    thu_tu + 1, tong_khuc, len(rong), len(ds)))
        return ds

    ds = None
    for lan in range(3):
        try:
            ds = mot_lan(lan)
            break
        except Exception as loi:  # noqa: BLE001
            # ═══ KHOÁ BỊ KẸT THÌ PHẢI ĐỔI KHOÁ, KHÔNG PHẢI ĐỢI TIẾP ═══
            #
            # Bắt cả `Exception` chứ không riêng `LoiNoiDung`, vì có một kiểu
            # kẹt rất khó chịu: lời gọi đầu bị chặn nhịp, nhưng máy chủ đã kịp
            # **ghi nhận cái khoá**. Từ đó trở đi mọi lần hỏi lại cùng khoá ấy
            # đều nhận "đang được xử lý" — và cứ đợi mãi.
            #
            # Đo được: cùng lời nhắc ấy gửi một mình xong trong 46 giây; còn
            # trong mẻ song song thì kẹt hơn tám phút rồi vẫn "đang xử lý".
            #
            # `lan` nằm trong khoá, nên vòng sau là một khoá hoàn toàn mới —
            # thoát hẳn cái khoá kẹt thay vì kiên nhẫn với nó. Đây là chỗ
            # **kiên nhẫn đúng cách là bỏ đi làm lại**, khác với mọi chỗ khác
            # trong tool nơi kiên nhẫn nghĩa là đợi.
            if lan == 2:
                raise
            bc.ghi("  khúc {0} chưa xong ({1}) — hỏi lại bằng khoá mới.".format(
                thu_tu + 1, str(loi)[:70]))
    return ds or []


#: Cột của sheet `scenes` — **giữ đúng tên và thứ tự này**, đây là khuôn mà
#: các tool dựng video theo bảng cảnh đọc. Đổi tên cột là mở bên đó không ra gì.
COT_CANH = ("scene_id", "srt_start", "srt_end", "duration", "planned_duration",
            "srt_text", "scene_kind", "subject_mode", "primary_subject",
            "primary_action", "visual_anchor", "must_not_show", "img_prompt",
            "prompt_json", "video_prompt", "img_path", "video_path",
            "status_img", "status_vid", "characters_used", "location_used",
            "reference_files", "media_id", "video_note", "segment_id")

COT_NHAN_VAT = ("id", "role", "name", "english_prompt", "vietnamese_prompt",
                "character_lock", "image_file", "status", "is_child", "media_id")

COT_THUMB = ("thumb_id", "version_desc", "img_prompt", "characters_used",
             "location_used", "reference_files", "img_path", "status_img")

#: Sheet `story_map` — bản đồ hình, mỗi chương một dòng.
COT_KE_HOACH = ("chuong", "srt_from", "srt_to", "title", "place", "time_light",
                "people", "motif", "emotion", "key_line")


def _viet_xlsx(duong: str, canh: List[Dict[str, Any]], k: Kenh,
               ke_hoach: Optional[List[Dict[str, Any]]] = None,
               dan: Optional[List[Dict[str, Any]]] = None,
               boi_canh: Optional[List[Dict[str, Any]]] = None) -> None:
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.styles import Font, PatternFill  # noqa: PLC0415

    sach = Workbook()
    ws = sach.active
    ws.title = "scenes"
    for cot, ten in enumerate(COT_CANH, 1):
        o = ws.cell(1, cot, ten)
        o.font = Font(bold=True, color="FFFFFF")
        o.fill = PatternFill("solid", fgColor="70AD47")
    for hang, c in enumerate(canh, 2):
        for cot, ten in enumerate(COT_CANH, 1):
            gia_tri = c.get(ten, "")
            # Đường cũ: mọi cảnh mặc định trỏ `nv1.png`. Nhánh đạo diễn (có
            # `dan`) thì mỗi cảnh tự khai — cảnh không có tham chiếu để trống.
            if ten == "reference_files" and not gia_tri and dan is None:
                gia_tri = json.dumps(["nv1.png"])
            if isinstance(gia_tri, (list, dict)):
                gia_tri = json.dumps(gia_tri, ensure_ascii=False)
            ws.cell(hang, cot, gia_tri)

    nv = sach.create_sheet("characters")
    for cot, ten in enumerate(COT_NHAN_VAT, 1):
        nv.cell(1, cot, ten)
    if dan:
        # Dàn do đạo diễn dựng: mỗi nhân vật (kể cả giai đoạn `nv4b`) một dòng,
        # ảnh tham chiếu `<id>.png` trong `<lượt>/tham-chieu/`. Cột thêm:
        # `sheet_prompt` (lời nhắc vẽ chân dung tham chiếu).
        nv.cell(1, len(COT_NHAN_VAT) + 1, "sheet_prompt")
        for hang, c in enumerate(dan, 2):
            for cot, ten in enumerate(COT_NHAN_VAT, 1):
                gia = c.get(ten, "")
                if ten == "image_file" and not gia:
                    gia = "{0}.png".format(c.get("id", ""))
                if ten == "status" and not gia:
                    gia = "done"
                nv.cell(hang, cot, gia if not isinstance(gia, (list, dict)) else json.dumps(gia))
            nv.cell(hang, len(COT_NHAN_VAT) + 1, str(c.get("sheet_prompt") or ""))
    else:
        nv.cell(2, 1, "nv1")
        nv.cell(2, 2, "protagonist")
        nv.cell(2, 3, "Reference")
        nv.cell(2, 4, str(k.style.get("default_character_prompt", "")))
        nv.cell(2, 6, str(k.style.get("reference_lock", "")))
        nv.cell(2, 7, "nv1.png")
        nv.cell(2, 8, "done")
    if boi_canh:
        lo = sach.create_sheet("locations")
        cot_loc = ("id", "name", "english_prompt", "location_lock", "lighting_default",
                   "image_file", "sheet_prompt")
        for cot, ten in enumerate(cot_loc, 1):
            lo.cell(1, cot, ten)
        for hang, l in enumerate(boi_canh, 2):
            for cot, ten in enumerate(cot_loc, 1):
                gia = l.get(ten, "")
                if ten == "image_file" and not gia:
                    gia = "{0}.png".format(l.get("id", ""))
                lo.cell(hang, cot, gia if not isinstance(gia, (list, dict)) else json.dumps(gia))

    tb = sach.create_sheet("thumbnail")
    for cot, ten in enumerate(COT_THUMB, 1):
        tb.cell(1, cot, ten)

    # Bản đồ hình — để người dựng mở Excel là thấy video chia chương ra sao,
    # chương nào ở đâu, câu nào là bản lề. Không có bản đồ thì không có sheet.
    if ke_hoach:
        kh = sach.create_sheet("story_map")
        for cot, ten in enumerate(COT_KE_HOACH, 1):
            o = kh.cell(1, cot, ten)
            o.font = Font(bold=True, color="FFFFFF")
            o.fill = PatternFill("solid", fgColor="4472C4")
        for hang, c in enumerate(ke_hoach, 2):
            for cot, ten in enumerate(COT_KE_HOACH, 1):
                kh.cell(hang, cot, (c or {}).get(ten, ""))

    os.makedirs(os.path.dirname(duong) or ".", exist_ok=True)
    tam = duong_tam(duong)
    sach.save(tam)
    thay_the(tam, duong)


# ── Khâu 5 & 6: ảnh và clip ──────────────────────────────────────────────────


def _doc_canh(luot: LuotChay) -> List[Dict[str, Any]]:
    tho = _doc_chu(os.path.join(luot.thu_muc, "4-canh.json"))
    if not tho.strip():
        raise RuntimeError("chưa có bảng cảnh")
    return json.loads(tho)


def sua_loi_nhac_canh(luot: LuotChay, so_canh: int, *,
                      img_prompt: Optional[str] = None,
                      video_prompt: Optional[str] = None) -> Dict[str, Any]:
    """Sửa lời nhắc ảnh/clip của **một cảnh** trong `4-canh.json`.

    Đây là đường để người dùng xem lại một cảnh chưa ưng, sửa lời nhắc, rồi bảo
    tool tạo lại đúng cảnh ấy — không đụng tới 118 cảnh kia.

    Chỉ ghi vào `4-canh.json` (bản chính mà khâu ảnh và khâu clip đọc). Không
    đụng `PROJECTS/` kết quả và không đụng file cảnh nguồn của kênh — đây là file
    làm việc của chính lượt đang chạy.

    `None` nghĩa là **giữ nguyên** lời nhắc cũ; chuỗi rỗng bị chặn, vì khâu ảnh
    từ chối lời nhắc rỗng và cửa soi bảng cảnh (`_canh_dung_duoc`) coi cảnh thiếu
    lời nhắc ảnh là bảng hỏng — để lọt là cả bảng bị cắt lại từ đầu bằng AI.

    Trả về đúng cảnh vừa sửa. Không tìm thấy cảnh thì ném lỗi.
    """
    goi_json = os.path.join(luot.thu_muc, "4-canh.json")
    tho = _doc_chu(goi_json)
    if not tho.strip():
        raise RuntimeError("chưa có bảng cảnh để sửa")
    canh = json.loads(tho)
    for c in canh:
        if int(c.get("scene_id") or 0) != int(so_canh):
            continue
        if img_prompt is not None:
            moi = str(img_prompt).strip()
            if not moi:
                raise ValueError("lời nhắc ảnh không được để trống")
            c["img_prompt"] = moi
        if video_prompt is not None:
            c["video_prompt"] = str(video_prompt).strip()
        _ghi_chu(goi_json, json.dumps(canh, ensure_ascii=False, indent=1))
        return c
    raise RuntimeError("không thấy cảnh {0} trong bảng".format(so_canh))


def don_canh_de_lam_lai(luot: LuotChay, cac_canh: Sequence[int], *,
                        ca_anh: bool) -> Tuple[int, int]:
    """Xoá tệp ảnh/clip của **đúng những cảnh này**, để hai khâu ấy làm lại chúng.

    Khâu ảnh và khâu clip nhìn đĩa trước: cảnh nào còn tệp thì bỏ qua. Nên "tạo
    lại cảnh 7 và 19" không phải là chạy lại cả khâu — chỉ cần xoá tệp của hai
    cảnh ấy, 117 cảnh kia vẫn nằm nguyên và **không trả tiền lần thứ hai**.

    `ca_anh=True`: xoá cả ảnh lẫn clip. Phải xoá clip theo, vì clip lấy ảnh làm
    khung đầu — giữ clip cũ là giữ một đoạn chuyển động của tấm ảnh không còn
    nữa. `ca_anh=False`: giữ ảnh, chỉ dựng lại chuyển động.

    Cảnh chưa từng có ảnh vẫn đưa vào danh sách được: không có tệp thì bỏ qua,
    khâu ảnh sẽ tự làm. Xoá không được (tệp đang mở trong trình xem video) thì
    ném `OSError` để người gọi nói ra, chứ không im lặng bỏ qua — im lặng ở đây
    nghĩa là khâu ảnh thấy tệp cũ, bỏ qua cảnh, và người dùng ngồi đợi một tấm
    ảnh không bao giờ đổi.
    """
    thu_muc_anh = os.path.join(luot.thu_muc, "5-anh")
    thu_muc_clip = os.path.join(luot.thu_muc, "6-clip")
    xoa_anh = xoa_clip = 0
    for so in sorted({int(s) for s in cac_canh}):
        can = [(os.path.join(thu_muc_clip, "{0}.mp4".format(so)), "clip")]
        if ca_anh:
            can.append((os.path.join(thu_muc_anh, "{0}.png".format(so)), "anh"))
        for duong, loai in can:
            if not os.path.isfile(duong):
                continue
            os.remove(duong)
            if loai == "anh":
                xoa_anh += 1
            else:
                xoa_clip += 1
    return xoa_anh, xoa_clip


class ThamChieu:
    """Giữ URL ảnh nhân vật cho **cả mẻ song song**, và chỉ làm mới **một lần**.

    ═══ VÌ SAO CẦN MỘT CÁI HỘP CHUNG ═══

    Bản đầu để mỗi luồng giữ URL riêng. Chữ ký hết hạn giữa mẻ thì cả mười hai
    luồng cùng phát hiện, cùng tải lại — và tệ hơn: luồng nào làm mới xong cũng
    **giữ cho riêng nó**, nên cảnh sau vẫn cầm URL cũ rồi lại hỏng, lại tải.

    Đo trên mẻ thật: **67 lần tải cho 61 tấm ảnh**. Gần một lần tải mỗi tấm,
    trong khi đúng ra cả lượt chỉ cần một. Với mười sáu video là ~600 MB, vượt
    trần kho tạm 500 MB — đúng cái lỗi tôi tưởng đã sửa xong.

    Cái hộp này sửa cả hai vế: một chỗ chung cho mọi luồng, và làm mới theo
    kiểu *so-rồi-đổi* — luồng nào tới sau thấy URL đã khác cái mình cầm thì
    dùng luôn bản mới chứ không tải nữa.
    """

    def __init__(self, bc: BoiCanh) -> None:
        self._bc = bc
        self._khoa = threading.Lock()
        self._url: List[str] = _url_tham_chieu(bc)

    def lay(self) -> List[str]:
        with self._khoa:
            return list(self._url)

    def lam_moi(self, cu: List[str]) -> List[str]:
        """Làm mới URL, trừ khi luồng khác vừa làm rồi."""
        with self._khoa:
            if list(self._url) != list(cu):
                # Luồng khác đã làm mới trong lúc mình đang hỏng. Dùng bản của
                # họ — đây là chỗ chặn 12 lần tải xuống còn 1.
                return list(self._url)
            self._bc.ghi("  chữ ký ảnh nhân vật hết hạn — làm mới cho cả mẻ.")
            self._url = _url_tham_chieu(self._bc, bo_qua_nho=True)
            return list(self._url)


#: Thiếu bao nhiêu phần trăm cảnh thì vẫn đi tiếp.
#:
#: 3% của 112 cảnh là 3 cảnh. Thiếu tới đó thì video vẫn xem được bình thường —
#: khâu dựng giữ khung cuối của cảnh trước lâu thêm vài giây, đúng như người
#: dựng tay để hình đứng yên lúc người đọc ngừng lấy hơi.
#:
#: Quá mức đó thì dừng: hỏng không còn là chuyện lẻ tẻ mà là hỏng có hệ thống
#: (lời nhắc sai khuôn, ảnh tham chiếu chết, nhà máy chập chờn), và đi tiếp chỉ
#: tốn thêm tiền cho một video vá chằng chịt.
TY_LE_THIEU_CHO_PHEP = 0.03


#: Giữa hai lần ghi tiến độ ra đĩa, ít nhất bấy nhiêu giây.
#:
#: Cần từ lúc cả mẻ bắn một lượt: 114 tấm ảnh giờ xong **trong cùng một hai
#: giây**, và mỗi tấm xong là một lần ghi `trang-thai.json` cộng một lần vẽ lại
#: bảng trên luồng giao diện. Trăm lần như thế dồn vào hai giây là cửa sổ khựng
#: — đúng thứ mà việc chạy nền sinh ra để tránh.
#:
#: Con số trong bộ nhớ vẫn cập nhật từng cái một; chỗ này chỉ thưa bớt **lần
#: ghi**. Mốc cuối (`xong == tổng`) luôn được ghi, và `core/auto.chay` còn ghi
#: lại lần nữa khi khâu kết thúc.
NHIP_GHI_TIEN_DO = 0.4


def dem_tien_do(bc: BoiCanh, luot: LuotChay, tt: TrangThaiKhau, viec: str,
                giu_nhip: float = 0.0):
    """Trả về hàm `(xong, tổng)` ghi tiến độ **trong** một khâu.

    Ghi vào `tt.ghi_chu` chứ không vào một chỗ riêng: `ghi_chu` đã đi thẳng ra
    `trang-thai.json`, nên đóng tool giữa chừng rồi mở lại vẫn thấy lần trước
    dừng ở cảnh thứ mấy.

    `giu_nhip` là số giây tối thiểu giữa hai lần **ghi ra đĩa** — xem
    `NHIP_GHI_TIEN_DO`.

    ═══ GHI SỔ HỎNG KHÔNG ĐƯỢC LÀM HỎNG KHÂU ═══

    Đây là **sổ sách**, không phải sản phẩm. Ảnh đã tải về vẫn nằm trên đĩa dù
    có ghi được tiến độ hay không, và khâu chạy lại sẽ nhìn đĩa mà nhặt tiếp.
    Nên một lỗi ghi tệp ở đây chỉ được phép làm mất *con số hiển thị*, tuyệt
    đối không được ném lên trên để giết cả mẻ.

    Có luật này vì đã mất thật: 27/08/2026, máy khách chết ở đúng dòng đổi tên
    `trang-thai.json` (`WinError 5`), khâu ảnh hỏng 12 lần liền, 97 cảnh không
    ra nổi một tấm — trong khi kịch bản, giọng đọc, phụ đề và bảng cảnh đã
    xong hết từ ba tiếng trước. `core/ghi_dia.py` lo phần thử lại; chỗ này là
    lưới cuối, cho trường hợp thử lại vẫn không xong (đĩa đầy, thư mục bị khoá
    hẳn). Lỗi vẫn được nói ra một lần trong nhật ký chứ không nuốt im.
    """
    moc = [0.0]
    da_than = [False]

    def bao(xong: int, tong: int) -> None:
        tt.ghi_chu["xong"] = int(xong)
        tt.ghi_chu["tong"] = int(tong)
        tt.ghi_chu["viec"] = viec
        bay_gio = time.time()
        if giu_nhip and xong < tong and bay_gio - moc[0] < float(giu_nhip):
            return
        moc[0] = bay_gio
        try:
            bc.nhip(luot)
        except OSError as loi:
            if not da_than[0]:
                da_than[0] = True
                bc.ghi("    (không ghi được tiến độ ra đĩa: {0} — vẫn chạy "
                       "tiếp, kết quả không mất)".format(str(loi)[:120]))

    return bao


def _chay_song_song(bc: BoiCanh, muc: List[Dict[str, Any]], lam, ten: str,
                    nhip=None, loai_job: str = "", mac_dinh: int = 0,
                    chiu_thieu: bool = True) -> int:
    """Chạy `lam(mục)` cho cả danh sách, **bắn hết một lượt**. Trả về số đã xong.

    ═══ HAI LUẬT ═══

    **Giữ nguyên phần đã làm.** Một cảnh hỏng không được vứt 98 cảnh kia — chờ
    cả mẻ chạy xong rồi mới báo. Tệp đã tải về vẫn nằm trên đĩa, nên chạy tiếp
    chỉ làm phần còn thiếu.

    **Bấm Dừng là dừng.** Đặt cờ rồi thì các luồng đang chạy tự thoát ở lần
    `kiem_dung()` kế tiếp, không đợi hết mẻ.

    `chiu_thieu=False` cho những khâu mà thiếu một mảnh là hỏng cả: giọng đọc
    thiếu một đoạn giữa bài thì video mất hẳn một khúc lời, không như ảnh thiếu
    một cảnh (khâu dựng giữ hình cảnh trước bù vào).

    ═══ VÌ SAO KHÔNG CÒN "THĂM DÒ MỘT CÁI TRƯỚC" ═══

    Bản trước làm **cái đầu tiên một mình**, xong xuôi mới bung luồng, để bắt
    sớm cảnh nhà máy đang tắt. Ý tốt, giá đắt: một tấm ảnh mất 30,8 giây ở nhà
    máy, nên mỗi khâu phải đứng yên **nửa phút** trước khi mở luồng nào — và
    khâu ảnh với khâu clip đều trả cái giá ấy.

    Mà nó cũng không mua được gì thật: gặp nhà máy tắt thì chính lượt thăm dò
    cũng ngồi đợi qua cả thang thử lại của `goi_kien_nhan` (~16 phút) rồi mới
    kêu. Cái chốt `nha_may_tat` dưới đây mới là thứ chặn được sớm — luồng nào
    thấy nhà máy tắt thì gạt chốt, các luồng còn lại thôi nhận việc mới.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed  # noqa: PLC0415

    con_lai = list(muc)
    if not con_lai:
        return 0
    xong = 0
    da_co = 0
    loi_dau: List[BaseException] = []
    tong = len(con_lai)

    def bao_nhip() -> None:
        if nhip is not None:
            nhip(xong, tong)

    bao_nhip()

    nha_may_tat = threading.Event()
    bc.nha_may_tat = nha_may_tat
    # Số luồng theo con số MÁY CHỦ TỰ KHAI, không phải số gõ sẵn trong mã.
    # Cổng cho 979 job ảnh cùng lúc; tool từng chạy 6 vì tin một trần 60
    # lượt/phút vốn không có thật. Xem `_so_luong` và `SONG_SONG_CANH`.
    so_luong = (_so_luong(bc, loai_job, mac_dinh or SONG_SONG_CANH, can=tong)
                if loai_job else min(mac_dinh or SONG_SONG_CANH, tong))
    if so_luong > SONG_SONG_CANH:
        bc.ghi("  bắn {0} việc cùng lúc (cổng cho phép).".format(so_luong))
    try:
        with ThreadPoolExecutor(max_workers=so_luong) as bo:
            cho = {bo.submit(_boc(bc, lam, c, loi_dau, nha_may_tat)): c
                   for c in con_lai}
            for xong_roi in as_completed(cho):
                ket = xong_roi.result()
                if ket is None:
                    continue
                _so, san_co = ket
                xong += 1
                da_co += 1 if san_co else 0
                bao_nhip()
                if xong % 10 == 0 or xong == tong:
                    bc.ghi("  {0}: {1}/{2} xong.".format(ten, xong, tong))
    finally:
        bc.nha_may_tat = None
    # ═══ BẤM DỪNG THÌ PHẢI BÁO ĐÚNG LÀ ĐÃ DỪNG ═══
    #
    # `_boc` nuốt mọi lỗi vào `loi_dau` để một cảnh hỏng không chôn cả mẻ — kể
    # cả `Cancelled`. Nhưng `core/auto.chay` đối xử với `Cancelled` khác hẳn
    # lỗi thường: nó giữ nguyên mọi thứ và không đốt lượt thử lại nào. Nên phải
    # lôi nó ra trước, đừng để một lỗi cảnh nào đó nói thay.
    from .auto import Cancelled  # noqa: PLC0415

    for _l in loi_dau:
        if isinstance(_l, Cancelled):
            raise _l
    if nha_may_tat.is_set():
        from .su_co import HET_TIEN, phan_loai as _phan  # noqa: PLC0415

        # Ví hết tiền cũng gạt chốt này, nhưng câu báo của nó đã đúng sẵn rồi
        # ("ví hết tiền, nạp thêm") — đừng đắp lên trên một câu nói về nhà máy.
        if any(_phan(l) == HET_TIEN for l in loi_dau):
            raise loi_dau[0]
        raise RuntimeError(
            "cổng ShopAPI ngừng nhận việc {0} giữa chừng. Đã giữ {1}/{2} — "
            "đây là phía máy chủ, không phải tool. Bật lại thì bấm Chạy tiếp, "
            "tool chỉ làm phần còn thiếu.".format(ten, xong, tong))
    if loi_dau:
        # ═══ MỘT CẢNH HỎNG KHÔNG ĐƯỢC CHÔN CẢ LƯỢT ═══
        #
        # Xảy ra thật hai lần liền (15/08/2026): cảnh 112/112 hỏng, và cả khâu
        # dừng. 111 clip đã trả tiền nằm im, ảnh bìa không làm, video không
        # dựng. Khách trả tiền cho 111 cảnh và nhận về **không có gì xem được**.
        #
        # Mà một cảnh thiếu thì khâu dựng chỉ giữ khung cuối của cảnh trước lâu
        # thêm vài giây — hầu như không ai nhận ra. So với việc không có video
        # thì đó là đổi chác quá hời.
        #
        # Nên: thiếu ít thì đi tiếp và NÓI RÕ thiếu cảnh nào, thiếu nhiều thì
        # dừng — vì lúc đó hỏng không còn là chuyện lẻ tẻ mà là hỏng có hệ
        # thống, và đi tiếp chỉ tốn thêm tiền cho một video vá chằng chịt.
        thieu = tong - xong
        if not chiu_thieu or thieu > max(1, int(tong * TY_LE_THIEU_CHO_PHEP)):
            raise loi_dau[0]
        bc.ghi("  {0}: thiếu {1}/{2} — đi tiếp, phần thiếu sẽ giữ hình cảnh "
               "trước lâu hơn một chút.".format(ten, thieu, tong))
        bc.ghi("    lý do cảnh hỏng: {0}".format(str(loi_dau[0])[:160]))
        bc.ghi("    muốn làm nốt thì bấm “Làm lại khâu này” — tool chỉ làm "
               "phần còn thiếu.")
    if da_co:
        bc.ghi("  ({0}/{1} {2} đã có sẵn, không làm lại)".format(
            da_co, tong, ten))
    return xong


def _boc(bc: BoiCanh, lam, c, loi_dau: List[BaseException],
         nha_may_tat: Optional[threading.Event] = None):
    """Bọc một mục: nuốt lỗi vào `loi_dau` để mẻ chạy tiếp, không gãy giữa chừng.

    ═══ NHÀ MÁY TẮT GIỮA MẺ THÌ CẢ MẺ DỪNG, KHÔNG AI ĐỢI RIÊNG ═══

    Nhà máy bên cổng có lúc **tắt giữa chừng** — 15 phút trước còn chạy, đang
    làm dở thì dừng. Khi ấy mười hai luồng mỗi luồng tự đợi 60 → 120 → 180
    giây, và mỗi lần thử lại vẫn ăn một suất trong trần 60 lượt/phút. Kết quả
    đo được: màn hình đầy chữ giống hệt nhau, suất gọi cạn sạch, mà không tấm
    ảnh nào ra.

    Một cái chốt chung sửa việc đó: luồng nào thấy nhà máy tắt thì gạt chốt,
    các luồng còn lại **thôi không nhận việc mới**. Mẻ kết thúc gọn, giữ nguyên
    phần đã làm, và người dùng bấm "Chạy tiếp" khi nhà máy bật lại.

    Dừng sớm ở đây **không mất gì**: ảnh đã tải về vẫn nằm trên đĩa, và khâu
    nào cũng nhìn đĩa trước.
    """
    from .su_co import HET_TIEN, NHA_MAY_NGHI, phan_loai as _phan  # noqa: PLC0415

    def chay_mot():
        if nha_may_tat is not None and nha_may_tat.is_set():
            return None
        try:
            bc.kiem_dung()
            return lam(c)
        except Exception as loi:  # noqa: BLE001
            # ═══ VÍ HẾT TIỀN CŨNG LÀ LÝ DO DỪNG CẢ MẺ ═══
            #
            # Trước đây chỗ này chỉ chặn khi nhà máy tắt, còn "ví hết tiền" thì
            # bắt được nhờ lượt thăm dò chạy một mình. Lượt thăm dò đã bỏ (nó
            # tốn nửa phút mỗi khâu), nên lý do ấy phải chuyển vào đây: hết tiền
            # thì 113 lời gọi còn lại chắc chắn cũng hết tiền, gửi tiếp chỉ làm
            # màn hình đầy chữ giống hệt nhau.
            loai = _phan(loi)
            if loai in (NHA_MAY_NGHI, HET_TIEN) and nha_may_tat is not None:
                nha_may_tat.set()
            loi_dau.append(loi)
            return None

    return chay_mot


def _giay_clip(bc: BoiCanh) -> int:
    """Một clip dài bao nhiêu giây, theo engine của kênh.

    ═══ THỜI LƯỢNG LÀ CỐ ĐỊNH THEO ENGINE, KHÔNG PHẢI THEO CẢNH ═══

    Veo3 bán **đúng** clip 8 giây, Seedance **đúng** 10 — không có số nào khác.
    Bản đầu tôi gửi độ dài của chính cảnh, làm tròn: cảnh 7,3 giây thành 7, và
    cổng trả *"Engine veo3 chỉ nhận video 8 giây, bạn đang chọn 7 giây"*.

    Lỗi này lẩn rất kỹ: cảnh nào tình cờ tròn 8 giây thì qua, nên mẻ thật ra
    được 18/99 clip rồi mới chết — nhìn vào tưởng chập chờn.

    `core/srt_scenes.py` đã ép trần cảnh theo đúng engine từ khâu cắt cảnh, nên
    lấy thẳng số cố định là an toàn: cảnh không bao giờ dài hơn trần, và phần
    hình thừa ra ở cảnh ngắn thì khâu dựng cắt.

    Lấy từ chính bảng của SDK, không tự đoán: thêm engine mới mà quên sửa chỗ
    này là cả khâu clip hỏng, và hỏng kiểu chập chờn.
    """
    try:
        from shopapi._constants import (  # noqa: PLC0415
            DEFAULT_VIDEO_DURATION_BY_ENGINE as _DL)
        return int(_DL.get(bc.kenh.engine, 8))
    except Exception:  # noqa: BLE001
        return 8 if bc.kenh.engine == "veo3" else 10


def _ghim_hai_dau(bc: BoiCanh) -> bool:
    """Kênh này có vẽ thêm ảnh KHUNG CUỐI và ghim clip cả hai đầu không?"""
    return bool(getattr(getattr(bc, "kenh", None), "ghim_hai_dau", False))


#: Câu ghim vào MỌI lời nhắc clip của kênh giữ tiếng cảnh.
#:
#: ═══ VÌ SAO TOOL GHIM CHỨ KHÔNG TRÔNG VÀO AI ═══
#:
#: Không tách được nhạc và lời ra khỏi tiếng động sau khi engine đã trộn chúng
#: vào một đường tiếng. Nên chỗ duy nhất chặn được là lúc đặt hàng.
#:
#: Lời nhắc `7-canh.md` đã dặn "no music, no speech" từ đầu. Đo trên phim
#: `openstory/0008` (28/08/2026) xem AI có chép lại không: **25/30** cảnh có
#: `ambient:`/`sfx:`, nhưng **0/30** cảnh nhắc "no music" hay "no speech". Dặn
#: trong lời nhắc là điều kiện cần; ghim bằng mã mới là điều kiện đủ.
LUAT_TIENG_CANH = (
    " AUDIO — ambient and sound effects ONLY: footsteps, water, wind, birds, "
    "cloth, wood, animal calls. NO music of any kind, no song, no score, no "
    "melody. NO speech, no dialogue, no voice-over, no narration, no singing, "
    "no humming, no whispering — not one spoken word from anyone."
)


def _giu_tieng_canh(bc: BoiCanh) -> bool:
    """Kênh này có giữ tiếng của clip (bước chân, chim hót) khi dựng không?

    `getattr` vì khâu clip chạy trong luồng phụ, nơi mọi lỗi bị nuốt thành
    "cảnh này hỏng" — xem `_co_khung_dau`.
    """
    return bool(getattr(getattr(bc, "kenh", None), "giu_tieng_canh", False))


def _anh_khung_cuoi(bc: BoiCanh, luot: LuotChay, c: Dict[str, Any],
                    anh_dau: str, hop: "ThamChieu",
                    so: Optional[SoTheoDoi] = None) -> str:
    """Vẽ tấm KHUNG CUỐI của một cảnh; trả đường dẫn, hoặc "" nếu không vẽ được.

    ═══ VÌ SAO CẦN TẤM NÀY ═══

    Ghim khung ĐẦU xong thì đầu clip trùng khít ảnh (đo 27/08/2026: lệch 3,6 /
    255, 30/30 clip). Nhưng Veo vẫn có 8 giây tự do sau đó, và đuôi trôi: AI
    chấm khung cuối 3,70, hai clip rơi xuống 2 điểm.

    Ghim CẢ HAI đầu trên đúng ba clip trôi nặng nhất: cảnh 11 đi 2 → 4, cảnh 2
    đi 3 → 4, và cả ba clip kết thúc gần trùng khít tấm khung cuối (lệch
    1,4–5,4). Cảnh 8 vẫn 2 điểm — vì chính TẤM KHUNG CUỐI của nó vẽ sai (lệch
    khung đầu 52,6, tức một bố cục khác hẳn).

    Bài học nằm ở cảnh 8: ghim hai đầu không tự chữa nhân vật, nó **dời chỗ
    hỏng từ bên trong video ra thành một tấm ảnh**. Mà ảnh thì đo được, chấm
    được, vẽ lại được — nên tấm này đi qua ĐÚNG cửa chấm như ảnh khung đầu.

    Ảnh khung đầu được đính làm tham chiếu CUỐI CÙNG: `prompt_khung_cuoi` dặn
    máy "cùng cú máy với tấm tham chiếu cuối, máy không di chuyển, chỉ nhân
    vật diễn tiếp".
    """
    from .noi_canh import prompt_khung_cuoi  # noqa: PLC0415

    so_canh = int(c["scene_id"])
    dich = os.path.join(os.path.dirname(anh_dau), "{0}-cuoi.png".format(so_canh))
    if os.path.exists(dich):
        return dich
    loi_nhac = prompt_khung_cuoi(str(c.get("img_prompt") or ""))
    if not loi_nhac.strip():
        return ""
    from .dao_dien_auto import ThamChieuCanh  # noqa: PLC0415

    duong = list(getattr(hop, "_duong", []) or [])
    # Ảnh khung ĐẦU đính làm tham chiếu CUỐI CÙNG: `prompt_khung_cuoi` dặn máy
    # "cùng cú máy với tấm tham chiếu cuối". Hộp không có đường cục bộ (kênh
    # một nhân vật cố định) thì dùng nguyên hộp cũ, mất câu neo ấy nhưng vẫn
    # vẽ được.
    hop_cuoi = ThamChieuCanh(bc, duong + [anh_dau]) if duong else hop
    try:
        goi = _tao_anh(bc, luot, loi_nhac, hop_cuoi,
                       khoa_viec(luot, "img", so_canh, loi_nhac,
                                 "|".join(hop_cuoi.lay()), "kc"),
                       ten_hien="khung cuối cảnh {0}".format(so_canh), so=so)
        _tai_ket_qua(bc, goi, 0, dich)
        _xoa_dau(bc, dich)
    except Exception as loi:  # noqa: BLE001 — không có khung cuối thì ghim một đầu
        bc.ghi("    cảnh {0}: không vẽ được khung cuối ({1}) — ghim một đầu."
               .format(so_canh, str(loi)[:80]))
        return ""
    # Cùng cửa chấm với ảnh khung đầu: cảnh 8 hỏng vì chính tấm này vẽ sai.
    _cham_va_ve_lai(bc, luot, dict(c, img_prompt=loi_nhac), dich, hop_cuoi, so=so)
    return dich


def _co_khung_dau(bc: BoiCanh) -> bool:
    """Kênh này có ghim khung đầu clip không.

    `getattr` chứ không `bc.kenh.khung_dau`: khâu clip chạy trong luồng phụ và
    mọi lỗi ở đó bị nuốt thành "cảnh này hỏng" — một `AttributeError` vì đồ
    giả trong bài kiểm thiếu trường sẽ hiện ra thành *"không ra clip nào"* chứ
    không hiện ra thành lỗi thuộc tính. Mất nửa giờ mới lần ra (27/08/2026).
    """
    return bool(getattr(getattr(bc, "kenh", None), "khung_dau", False))


def _lam_clip(bc: BoiCanh, luot: LuotChay, c: Dict[str, Any], anh: str,
              dich: str, giay: int, so: Optional[SoTheoDoi] = None,
              khung_dau: bool = False, anh_cuoi: Optional[str] = None) -> None:
    """Tạo clip cho một cảnh, tải về, mở thử bằng FFmpeg.

    `khung_dau=True` gửi thêm `frame_mode: start_frame`: khung hình đầu của clip
    CHÍNH LÀ ảnh (Flow "Frames"). Khoá idempotency đổi theo, vì clip cũ cùng
    prompt nhưng làm ở chế độ nguyên liệu là một sản phẩm khác.

    ═══ BA NƠI GỌI, TRƯỚC 27/08/2026 CHỈ MỘT NƠI TRUYỀN CỜ ═══

    Kênh khai `khung_dau: true` mà đường thường (không nối cảnh) vẫn gọi hàm
    này với mặc định `False` — cờ của kênh bị bỏ qua **lặng lẽ**, không một
    dòng nhật ký nào.

    Đo được vì so khung đầu clip với chính tấm ảnh gửi vào (ảnh xám 160×90,
    thang 0–255): phim `openstory/0002` lệch trung bình **37,6** dù kênh đã
    bật cờ; làm lại đúng một clip có cờ thì lệch **3,5**. Một phép đo rẻ mà
    nói ngay được "cờ có tới nơi không" — đắt hơn nhiều là ngồi đoán vì sao
    nhân vật trôi.

    `anh_cuoi` gửi thêm `image_url_end`: clip bị ghim CẢ HAI đầu — nó bắt đầu
    đúng bằng `anh` và kết thúc đúng bằng `anh_cuoi`, engine chỉ làm phần chuyển
    động ở giữa. Đo 26/08/2026: chỉ ghim một đầu thì sau 8 giây nhân vật đã đổi.

    Tách ra khỏi khâu clip vì giờ có **hai** nơi gọi: dây chuyền ở khâu ảnh
    (ảnh nào xong là bắn clip của nó ngay) và khâu clip (làm nốt phần còn
    thiếu). Cùng một mã, nên hai đường không bao giờ lệch nhau về khoá
    idempotency — thứ giữ cho chạy tiếp không trả tiền hai lần.
    """
    so_canh = int(c["scene_id"])
    # ═══ KHÔNG CÓ ẢNH THÌ KHÔNG LÀM CLIP ═══
    #
    # Bản trước để `url_anh = ""` rồi vẫn bắn — tức sinh clip **không có ảnh
    # tham chiếu nào**, không nhân vật, không phong cách kênh.
    #
    # Đo trên lượt chạy thật U01 ngày 18/08/2026: ảnh cảnh 18 bị cổng từ chối
    # (`content_rejected`), nhưng clip 18 vẫn ra. Mở ba khung hình liền nhau
    # xem thì cảnh 17 và 19 đúng phong cách kênh, còn cảnh 18 có một nhân vật
    # **không có mặt** — đầu trống trơn, không mắt không miệng — trên nền xám
    # nhợt thay vì kem ấm. Một clip lạc hẳn, ghép giữa hai cảnh đúng.
    #
    # Thiếu clip thì khâu dựng giữ khung của cảnh trước lâu thêm vài giây, gần
    # như không ai nhận ra. Một clip sai phong cách thì ai cũng thấy. Nên thà
    # thiếu còn hơn sai.
    #
    # Chặn ở ĐÂY chứ không ở nơi gọi: có hai nơi gọi, và `bat_clip` ở khâu ảnh
    # vốn đã tự kiểm — chỉ khâu clip chạy riêng là quên. Đặt phép kiểm vào chỗ
    # chung thì không nơi nào quên được nữa.
    if not os.path.exists(anh):
        raise LoiNoiDung(
            "cảnh {0} chưa có ảnh nên chưa làm clip được — làm lại khâu ảnh "
            "trước đã".format(so_canh))
    # Ảnh của chính cảnh này làm khung đầu — đây là thứ giữ cho nhân vật không
    # đổi mặt giữa các cảnh. Cổng nhận URL, không nhận đường dẫn máy — xem
    # `_url_anh_canh`.
    url_anh = _url_anh_canh(bc, luot, so_canh, anh)
    url_cuoi = _url_anh_canh(bc, luot, so_canh, anh_cuoi) if (khung_dau and anh_cuoi and os.path.exists(anh_cuoi)) else ""

    # ═══ GHIM HAI ĐẦU THÌ CẤM "HIỆN RA RỒI BIẾN MẤT" ═══
    #
    # Ghim cả hai đầu là ra lệnh cho engine: bắt đầu ở đây, kết thúc ở kia. Nếu
    # lời nhắc lại tả một biến cố KHÔNG còn trong khung cuối, engine buộc phải
    # diễn nó ra rồi **dọn sạch** kịp hạ đúng khung cuối — và lúc dọn là lúc
    # hình khựng.
    #
    # Đo bởi phiên `kho-github-77` ngày 27/08/2026 trên clip 1 của phim
    # `timelapse/0001` (cùng cơ chế ghim hai đầu): lệch tiền cảnh so với ảnh
    # mốc đầu là 11,4 ở giây 0,1 — vọt lên **42,1** ở giây 4,0 — rồi tụt về
    # 3,9 ở giây 7,9. Xem tận mắt giây 4: một đám đông tràn kín bãi cát rồi
    # biến sạch. `freezedetect`/`blackdetect` không báo gì, nên không phải lỗi
    # ghép: chính engine dựng lên rồi nuốt đi.
    # ═══ VÀ PHẢI CHO NÓ MỘT LỐI RA ═══
    #
    # Cấm suông thì engine vẫn phải lấp đủ mấy giây giữa hai khung ghim: bí
    # đường thì nó rơi về thứ nó thuộc nhất về nơi ấy. Phiên `kho-github-77`
    # đo 28/08/2026 trên phim Paris: giữa hai khung ghim thời trung cổ, clip
    # trôi hẳn sang Paris **hôm nay** — cầu thép, ô tô, tàu du lịch — rồi mới
    # quay về đúng khung cuối. Không luật cấm nào bắt được, vì nó có quay về.
    #
    # Nên câu thứ hai chỉ đường: không thấy cách đi thì ĐỨNG GẦN KHUNG ĐẦU và
    # đổi ít thôi. Một clip nhạt vẫn ghép được vào phim; một clip lạc thì không.
    luat_mot_chieu = (
        " ONE DIRECTION ONLY: every change in this clip must still be there in "
        "the final frame. Nothing may appear and then vanish, and nothing may "
        "be undone before the clip ends."
        " If you cannot find a way from the first frame to the last, stay close "
        "to the first frame and change very little — a quiet clip is fine. Never "
        "drift to anything outside the world of these two frames to fill the time."
        if url_cuoi else "")

    # Kênh giữ tiếng cảnh thì phải đặt hàng đúng thứ tiếng ấy — xem
    # `LUAT_TIENG_CANH`. Khoá idempotency băm `c["video_prompt"]` chứ không băm
    # câu ghim thêm, nên phải đánh dấu `:tc` như `luat_mot_chieu` đánh `:kc`;
    # thiếu dấu ấy thì chạy tiếp lấy lại đúng clip cũ có nhạc.
    luat_tieng = LUAT_TIENG_CANH if _giu_tieng_canh(bc) else ""

    def goi_clip(dia_chi, hau_to=""):
        job = _tao_job(
            bc, bc.client.videos.create,
            prompt=c["video_prompt"] + luat_mot_chieu + luat_tieng,
            engine=bc.kenh.engine,
            duration=giay, aspect_ratio="16:9",
            image_url=dia_chi or None,
            extra_body=(dict({"frame_mode": "start_frame"},
                             **({"image_url_end": url_cuoi} if url_cuoi else {}))
                        if khung_dau else None),
            idempotency_key=khoa_viec(luot, "vid", so_canh,
                                      c["video_prompt"], dia_chi,
                                      giay) + (":kd" if khung_dau else "")
            + (":kc" + url_cuoi[-12:] if url_cuoi else "")
            + (":tc" if luat_tieng else "") + hau_to)
        return _cho_job(bc, job, ten_viec="cảnh {0}".format(so_canh), so=so)

    try:
        goi = goi_clip(url_anh)
    except LoiKetJob:
        # ═══ JOB KẸT: ĐẶT LẠI BẰNG KHOÁ MỚI ═══
        #
        # Máy chủ đã nhận việc nhưng mười hai phút vẫn "đang xử lý", với một
        # clip lẽ ra mất 1–3 phút. Gọi lại y nguyên là rơi vào đúng job kẹt đó,
        # mãi mãi. Chỉ khoá mới mới thoát ra.
        #
        # Có tốn thêm tiền không? Có thể — nếu job cũ rốt cuộc vẫn chạy xong.
        # Đổi lại là cả lượt chạy không đứng hình. Đã đo thật 14/08/2026: chín
        # clip cuối kẹt, sáu luồng cùng ngồi đợi, cả mẻ 112 cảnh không nhích
        # suốt mười hai phút. Đường còn lại — bỏ cả lượt — đắt hơn nhiều lần.
        # MỘT nấc là không đủ: khâu ngoài thử lại cả khâu, mỗi lần lại dựng
        # đúng hai khoá `""` và `":k2"` cũ — cả hai đã hỏng từ lần trước. Đo
        # 28/08/2026 trên phim `openstory/0011` cảnh 40: `:k2` đặt lúc 17:44,
        # tới 17:55 vẫn "đang làm". Xem `khoa_thoat_ket`.
        goi = None
        for _lan in range(1, 3):
            bc.ghi("    cảnh {0}: máy chủ nhận việc rồi bỏ đó — đặt lại bằng "
                   "khoá mới ({1}/2).".format(so_canh, _lan))
            try:
                goi = goi_clip(url_anh, khoa_thoat_ket(_lan))
                break
            except LoiKetJob:
                if _lan == 2:
                    raise
    except Exception as loi:  # noqa: BLE001
        chu = str(loi).lower()
        if not url_anh or not any(d in chu for d in _ANH_THAM_CHIEU_HONG):
            raise
        # Chữ ký ảnh khung đầu hết hạn giữa mẻ — tải lại đúng tấm ấy.
        url_anh = _url_anh_canh(bc, luot, so_canh, anh, bo_qua_nho=True)
        goi = goi_clip(url_anh, ":tc2")
    _tai_ket_qua(bc, goi, 0, dich)
    try:
        _kiem_media(bc, dich)
    except LoiNoiDung as loi:
        # ═══ TỆP HỎNG TỪ NGUỒN → LÀM LẠI BẰNG KHOÁ MỚI, MỘT LẦN ═══
        #
        # Tải lại cùng job là nhận lại đúng tệp hỏng ấy (máy chủ giữ bản đã
        # hỏng). Đo 25/08/2026: clip 99, 106, 110 của story-3d/0001 đều cụt
        # giữa tệp dù tải đủ byte. Chỉ một job mới mới cho tệp lành.
        bc.ghi("    cảnh {0}: clip tải về hỏng ({1}) — tạo lại bằng khoá mới."
               .format(so_canh, str(loi)[:80]))
        goi = goi_clip(url_anh, ":hong2")
        _tai_ket_qua(bc, goi, 0, dich)
        _kiem_media(bc, dich)


def _hop_cho_canh(bc: BoiCanh, luot: LuotChay, c: Dict[str, Any], hop: "ThamChieu"):
    """Nhánh đạo diễn: mỗi cảnh dùng đúng ảnh tham chiếu nó khai (`reference_files`
    → `<lượt>/tham-chieu/<id>.png`). Đường cũ: hộp `nv1.png` chung như trước."""
    from .dao_dien_auto import ThamChieuCanh, che_do_dao_dien, duong_tham_chieu_canh  # noqa: PLC0415

    if not che_do_dao_dien(bc.kenh):
        return hop
    duong = duong_tham_chieu_canh(luot, c)
    return ThamChieuCanh(bc, duong) if duong else _HopTrong()


def _hop_bia(bc: BoiCanh, luot: LuotChay, hop: "ThamChieu"):
    """Tham chiếu cho ẢNH BÌA: kênh đường đạo diễn dùng nhân vật chính của chính
    bộ phim (hai id xuất hiện nhiều nhất); đường cũ giữ hộp nv1.png của kênh."""
    from .dao_dien_auto import ThamChieuCanh, che_do_dao_dien, nhan_vat_chinh_cua_luot  # noqa: PLC0415

    if not che_do_dao_dien(bc.kenh):
        return hop
    duong = nhan_vat_chinh_cua_luot(luot, 2)
    if not duong:
        return hop
    bc.ghi("  ảnh bìa: dùng nhân vật chính của phim làm tham chiếu ({0}).".format(
        ", ".join(os.path.basename(d)[:-4] for d in duong)))
    return ThamChieuCanh(bc, duong)


class _HopTrong:
    def lay(self) -> List[str]:
        return []

    def lam_moi(self, _cu: List[str]) -> List[str]:
        return []


def _lam_anh_canh(bc: BoiCanh, luot: LuotChay, c: Dict[str, Any], tep: str,
                  hop: "ThamChieu", so: Optional[SoTheoDoi] = None) -> None:
    """Tạo ảnh cho một cảnh rồi tải về."""
    so_canh = int(c["scene_id"])
    # ═══ KHÔNG GỬI LỜI NHẮC RỖNG ═══
    #
    # Cổng từ chối bằng *"Bạn cần mô tả thứ muốn tạo — `prompt` đang rỗng"*,
    # nhưng chỉ sau khi mẻ đã chạy được một đoạn. Đo thật 15/08/2026: một lượt
    # trả tiền cho 56 tấm ảnh rồi mới chết, vì bảng cảnh có 52/108 dòng thiếu
    # lời nhắc.
    #
    # Chặn ở đây là chặn trước khi gọi, tức trước khi mất gì.
    if not str(c.get("img_prompt") or "").strip():
        raise LoiNoiDung(
            "cảnh {0} không có lời nhắc ảnh — bảng cảnh hỏng, hãy chọn dòng "
            "“Cắt cảnh và viết lời nhắc” rồi bấm “Làm lại từ khâu này”".format(
                so_canh))
    # ═══ KHOÁ PHẢI PHỦ CẢ ẢNH THAM CHIẾU ═══
    #
    # Thân yêu cầu gồm lời nhắc **và** URL ảnh tham chiếu. URL ấy là link ký
    # hạn: hết hạn thì tool tải lại và nhận một URL khác. Khoá chỉ đúc từ lời
    # nhắc nên nó không đổi theo — và cổng từ chối thẳng:
    #
    #     "Idempotency-Key này đã được dùng cho một yêu cầu có nội dung khác."
    #
    # Đo được 15/08/2026: chạy lại khâu ảnh của một lượt cũ là hỏng 100% ngay
    # lượt gọi đầu. Khâu clip đã đưa `dia_chi` vào khoá từ trước; khâu ảnh và
    # khâu ảnh bìa thì quên — cùng một bài học, sót hai chỗ.
    try:
        goi = _tao_anh(bc, luot, c["img_prompt"], hop,
                       khoa_viec(luot, "img", so_canh, c["img_prompt"],
                                 "|".join(hop.lay())),
                       ten_hien="ảnh cảnh {0}".format(so_canh), so=so)
    except Exception as loi:  # noqa: BLE001
        # ═══ BỘ LỌC TỪ CHỐI → VIẾT LẠI LỜI NHẮC MỘT LẦN, KHÔNG BỎ CẢNH ═══
        #
        # Đo 25/08/2026 (story-3d/0001, 123 cảnh): ba cảnh bị chặn chỉ vì chữ
        # "cheeks flushing", "swing violently", "coy teasing" — không có gì
        # để chặn, nhưng bộ lọc là máy. Trước đây cảnh ấy bị bỏ ("giữ hình
        # cảnh trước lâu hơn") dù tab Hàng loạt đã biết viết lại từ 3899466.
        # Chủ dự án: "prompt bị từ chối thì phải có logic làm lại prompt".
        moi = _viet_lai_khi_bi_tu_choi(bc, luot, c, loi)
        if not moi:
            raise
        try:
            goi = _tao_anh(bc, luot, moi, hop,
                           khoa_viec(luot, "img", so_canh, moi, "|".join(hop.lay()), "vl"),
                           ten_hien="ảnh cảnh {0}".format(so_canh), so=so)
        except Exception as loi2:  # noqa: BLE001
            # Lần 3: thay từ thô (mồm, liếm, nuốt, vũ khí…) — không tốn lượt chữ.
            # Đo 26/08/2026: bản AI viết lại vẫn giữ "toward his open mouth" và
            # bị chặn lần nữa; thay thẳng từ mới qua.
            tho = _lam_lanh_tho_neu_bi_tu_choi(bc, luot, c, moi, loi2)
            if not tho:
                raise
            goi = _tao_anh(bc, luot, tho, hop,
                           khoa_viec(luot, "img", so_canh, tho, "|".join(hop.lay()), "vl2"),
                           ten_hien="ảnh cảnh {0}".format(so_canh), so=so)
    _tai_ket_qua(bc, goi, 0, tep)
    _xoa_dau(bc, tep)
    _cham_va_ve_lai(bc, luot, c, tep, hop, so=so)


#: Mỗi cảnh yếu được vẽ thêm tối đa bấy nhiêu ứng viên.
#:
#: Chủ dự án, 28/08/2026: *"tao sợ tốn tiền đâu, tao cần là mọi thứ ok tốt
#: nhất"*. Một tấm ảnh 50 ₫; một cảnh nhân vật sai thì cả phim hỏng ở chỗ ấy
#: và người xem thấy ngay. Nên cửa chấm được phép kiên nhẫn: vẽ tới bốn ứng
#: viên, dừng ngay khi có tấm đạt (xem vòng lặp — `tot > NGUONG_LAM_LAI` thì
#: thoát), nên cảnh nào ra tốt lần đầu vẫn chỉ tốn đúng một tấm.
SO_UNG_VIEN_CHAM = 4


def _cham_va_ve_lai(bc: BoiCanh, luot: LuotChay, c: Dict[str, Any], tep: str,
                    hop: "ThamChieu", so: Optional[SoTheoDoi] = None) -> None:
    """Chấm tấm vừa vẽ với ảnh tham chiếu; lệch quá thì vẽ thêm, giữ tấm hơn.

    ═══ VÌ SAO KHÂU NÀY CẦN CỬA CHẤM RIÊNG ═══

    Tool đã có bộ chấm `core/cham_anh.py` từ 25/08/2026, nhưng nó chỉ gắn vào
    **hàng đợi của giao diện** (`ui_qt/app.py` → `core/jobs.py`). Luồng Tự động
    gọi thẳng `client.images.create`, nên nó đi vòng qua cửa ấy: vẽ xong là đi
    tiếp, tấm nào lệch thì lệch luôn vào phim.

    Đo 27/08/2026 trên phim `openstory/0002` (30 cảnh): 4 cảnh ra 2–3 điểm —
    nhân vật bị vẽ lại thành người khác. Vẽ thêm ứng viên rồi giữ tấm cao điểm
    hơn cứu được **cả bốn** lên 4 điểm (cảnh 2 đi 2 → 3 → 4).

    ═══ MẶC ĐỊNH TẮT ═══

    Mỗi lượt chấm là một lời gọi chữ có kèm ảnh, và mỗi lần vẽ lại là một tấm
    ảnh nữa — tiền thật. Kênh nào muốn thì khai `cham_anh: true` trong
    `kenh.yaml`; kênh của khách không tự dưng đắt lên.
    """
    if not bool(getattr(bc.kenh, "cham_anh", False)):
        return
    from .cham_anh import NGUONG_LAM_LAI, cham_anh  # noqa: PLC0415

    refs = _duong_tham_chieu(bc, luot, c, hop)
    if not refs or not os.path.isfile(tep):
        return
    so_canh = int(c["scene_id"])

    def goi_cham(noi_dung):
        from .goi_van_ban import goi_van_ban  # noqa: PLC0415

        return goi_van_ban(bc.client, [{"role": "user", "content": noi_dung}],
                           mo_hinh=bc.kenh.mo_hinh, toi_da_token=200)

    diem = cham_anh(goi_cham, tep, refs, mo_ta=str(c.get("img_prompt") or "")[:600])
    # `0` KHÔNG phải điểm kém: giám khảo được bảo trả 0 khi nhân vật không có
    # trong khung (máy quay đang ở chỗ khác, cảnh chỉ có đồ vật) — không có gì
    # để so thì không có gì để sửa. `core/jobs.py` chặn vế này từ 25/08/2026;
    # bản đầu của hàm này quên, và cảnh 27 của phim 0002 bị vẽ lại oan
    # (27/08/2026).
    if diem is None or diem == 0 or diem > NGUONG_LAM_LAI:
        return
    bc.ghi("    ảnh cảnh {0}: {1}/5 điểm giống — vẽ thêm để chọn tấm hơn…".format(
        so_canh, diem))
    tot = diem
    giu = tep + ".giu"
    try:
        shutil.copyfile(tep, giu)
    except OSError:
        return
    try:
        for lan in range(SO_UNG_VIEN_CHAM):
            bc.kiem_dung()
            try:
                goi = _tao_anh(bc, luot, c["img_prompt"], hop,
                               khoa_viec(luot, "img", so_canh, c["img_prompt"],
                                         "|".join(hop.lay()), "uv{0}".format(lan + 1)),
                               ten_hien="ảnh cảnh {0} (ứng viên {1})".format(
                                   so_canh, lan + 1), so=so)
                _tai_ket_qua(bc, goi, 0, tep)
                _xoa_dau(bc, tep)
            except Exception as loi:  # noqa: BLE001 — ứng viên hỏng thì giữ tấm cũ
                bc.ghi("    ảnh cảnh {0}: ứng viên {1} không vẽ được ({2})".format(
                    so_canh, lan + 1, str(loi)[:80]))
                break
            moi = cham_anh(goi_cham, tep, refs,
                           mo_ta=str(c.get("img_prompt") or "")[:600])
            bc.ghi("    ảnh cảnh {0}: ứng viên {1} được {2}/5 (tấm đang giữ {3})".format(
                so_canh, lan + 1, moi, tot))
            if moi is not None and moi > tot:
                tot = moi
                shutil.copyfile(tep, giu)
            if tot > NGUONG_LAM_LAI:
                break
        # Tấm cuối cùng vẽ ra chưa chắc là tấm hơn — chép lại bản đang giữ.
        shutil.copyfile(giu, tep)
        # ═══ VẼ HẾT ỨNG VIÊN MÀ VẪN KHÔNG ĐẠT = LỖI LỜI NHẮC ═══
        #
        # Vẽ lại là để cứu những lượt xui. Nếu CẢ BỐN ứng viên cùng trượt thì
        # không còn là xui: lời nhắc ấy sai từ gốc, vẽ thêm bao nhiêu cũng thế.
        #
        # Đo 28/08/2026 (phim 0008 cảnh 7): lời nhắc mở bằng "Over-the-shoulder
        # shot from just behind nv1's shoulder" — chỗ to nhất khung hình là một
        # cái lưng không có mặt để khớp ảnh, nên máy bịa ra người lạ, lần nào
        # cũng bịa. Bốn ứng viên, cả bốn 2/5, rồi tấm hỏng vẫn lặng lẽ vào phim.
        #
        # Ghi một dòng RÕ để lượt chạy còn nói ra được chỗ hỏng, thay vì im.
        if tot <= NGUONG_LAM_LAI:
            bc.ghi("    ⚠ ảnh cảnh {0}: vẽ {1} lượt vẫn {2}/5 — lỗi nằm ở LỜI NHẮC "
                   "(khung hình hoặc tham chiếu), không phải ở lượt vẽ.".format(
                       so_canh, SO_UNG_VIEN_CHAM + 1, tot))
    finally:
        try:
            os.remove(giu)
        except OSError:
            pass


def _duong_tham_chieu(bc: BoiCanh, luot: LuotChay, c: Dict[str, Any],
                      hop: "ThamChieu") -> List[str]:
    """Đường dẫn ảnh tham chiếu THẬT của một cảnh, cho bộ chấm mở ra xem.

    Chỉ đường đạo diễn (`ThamChieuCanh`) mới giữ đường dẫn cục bộ. Kênh một
    nhân vật cố định đi hộp khác, chưa có đường ấy — trả rỗng thì `cham_anh`
    lặng lẽ bỏ qua, đúng nết "không có gì để so thì đừng chấm".
    """
    duong = list(getattr(hop, "_duong", []) or [])
    return [p for p in duong if isinstance(p, str) and os.path.isfile(p)]


_KHOA_SUA_CANH = threading.Lock()


def _lam_lanh_tho_neu_bi_tu_choi(bc: BoiCanh, luot: LuotChay, c: Dict[str, Any],
                                 prompt: str, loi: Exception) -> str:
    """Bị chặn lần hai: thay từ thô (`core.viet_lai_prompt.lam_lanh_tho`). Khác bản
    cũ thì ghi vào 4-canh.json và trả về; không đổi được gì thì ""."""
    from .viet_lai_prompt import la_bi_tu_choi, lam_lanh_tho  # noqa: PLC0415

    if not la_bi_tu_choi("", str(loi)):
        return ""
    tho = lam_lanh_tho(prompt)
    if not tho or tho.strip() == prompt.strip():
        return ""
    so_canh = int(c["scene_id"])
    bc.ghi("    ảnh cảnh {0}: vẫn bị chặn — thay từ thô rồi thử lần cuối…".format(so_canh))
    c["img_prompt"] = tho
    _ghi_loi_nhac_da_sua(luot, so_canh, tho)
    return tho


def _viet_lai_khi_bi_tu_choi(bc: BoiCanh, luot: LuotChay, c: Dict[str, Any],
                             loi: Exception) -> str:
    """Bị bộ lọc chặn thì nhờ AI viết lại lời nhắc ảnh (giữ chủ thể, khối khoá,
    đuôi phong cách — xem `core/viet_lai_prompt`). Trả về lời nhắc mới, hoặc ""
    nếu không phải lỗi bộ lọc / viết lại không ra. Lời nhắc mới được ghi lại
    vào `4-canh.json` để khâu clip và lần "Làm lại" dùng đúng bản đã qua."""
    from .viet_lai_prompt import la_bi_tu_choi, viet_lai_prompt  # noqa: PLC0415

    if not la_bi_tu_choi("", str(loi)):
        return ""
    so_canh = int(c["scene_id"])
    cu = str(c.get("img_prompt") or "")
    bc.ghi("    ảnh cảnh {0}: bị bộ lọc từ chối — viết lại lời nhắc rồi thử lại…".format(so_canh))

    def goi_ai(loi_nhac: str) -> str:
        return bc.goi_chat(loi_nhac, mo_hinh=str(bc.kenh.mo_hinh or "claude-sonnet-5"),
                           khoa=khoa_viec(luot, "vl-img", so_canh, cu), toi_da_token=2048)

    try:
        moi = viet_lai_prompt(goi_ai, cu, str(loi))
    except Exception as loi2:  # noqa: BLE001
        bc.ghi("    ảnh cảnh {0}: không viết lại được ({1}).".format(so_canh, str(loi2)[:100]))
        return ""
    if not moi or moi.strip() == cu.strip():
        return ""
    c["img_prompt"] = moi
    _ghi_loi_nhac_da_sua(luot, so_canh, moi)
    return moi


def _ghi_loi_nhac_da_sua(luot: LuotChay, so_canh: int, moi: str) -> None:
    """Lưu lời nhắc đã sửa vào 4-canh.json — KHÔNG kèm đuôi nối cảnh (đuôi được
    nối lại lúc tạo; lưu kèm là lần sau nối đôi, prompt vượt 5.000 ký tự)."""
    from .noi_canh import bo_duoi_noi_canh  # noqa: PLC0415

    with _KHOA_SUA_CANH:
        try:
            sua_loi_nhac_canh(luot, so_canh, img_prompt=bo_duoi_noi_canh(moi))
        except Exception:  # noqa: BLE001 — không ghi được cũng vẫn tạo ảnh bằng bản mới
            pass


def _xoa_dau(bc: BoiCanh, tep: str) -> None:
    """Xoá dấu nhà cung cấp ngay khi ảnh vừa tải về.

    ═══ VÌ SAO PHẢI Ở ĐÂY, KHÔNG PHẢI SAU ═══

    Ảnh của cảnh nào là **khung đầu của clip cảnh ấy**. Dấu còn trên ảnh thì nó
    nằm luôn trong clip, và tám giây clip nào cũng đeo nó. Phát hiện muộn thì
    phải làm lại từ khâu clip — trả tiền lại cho cả trăm clip vì một cái dấu ở
    góc.

    Chạy trên máy khách, 27 mili giây một ảnh, không gọi mạng. Hỏng thì im
    lặng giữ ảnh nguyên: ảnh còn dấu vẫn dùng được, còn làm hỏng cả khâu ảnh vì
    một việc làm đẹp thì không.
    """
    try:
        from .xoa_dau_anh import xoa_dau_tep  # noqa: PLC0415

        xoa_dau_tep(tep)
    except Exception:  # noqa: BLE001
        pass
    _lam_sach_anh(bc, tep)


def _lam_sach_anh(bc: BoiCanh, tep: str) -> None:
    """Bỏ thẻ nguồn gốc AI khỏi ảnh vừa tải về, nếu khách bật ở Cài đặt.

    ═══ CHỖ THẬT SỰ ĂN THUA LÀ ẢNH BÌA ═══

    Đo 16/08/2026 trên kết quả thật: ảnh nhà cung cấp trả về mang đủ thẻ nguồn
    gốc AI (`c2pa`, nhãn "ảnh do AI tạo", và lời khai đã nhúng dấu chìm
    SynthID). Nhưng **video cuối thì vốn đã sạch** — khâu dựng mã hoá lại nên
    thẻ mất hết.

    Còn **ảnh bìa cũng lên YouTube**, mà nó là tệp nhà cung cấp trả về gần như
    nguyên vẹn. Đó là chỗ duy nhất bước này đổi được điều gì.

    Vẫn chạy cho cả ảnh cảnh, dù chúng không được tải lên đâu: cùng một chỗ
    móc, mất chừng vài phần trăm giây một tấm, và đỡ phải nhớ tấm nào cần tấm
    nào không. Ảnh **không bị nén lại** nên không mất một chút nét nào — xem
    `core/lam_sach.lam_sach_anh`.

    Hỏng thì im lặng để nguyên: đây là việc vệ sinh, không đáng làm hỏng một
    tấm ảnh đã trả tiền để tạo ra.
    """
    try:
        if not _bat_lam_sach(bc):
            return
        from .lam_sach import lam_sach_anh  # noqa: PLC0415

        lam_sach_anh(tep)
    except Exception:  # noqa: BLE001
        pass


def _chen_the_cam_xuc(bc: BoiCanh, luot: LuotChay, kich_ban: str) -> str:
    """Trả về chữ sẽ đem đi đọc — có thẻ cảm xúc nếu khách bật, không thì y cũ.

    ═══ GHI RA TỆP RIÊNG, KHÔNG ĐÈ LÊN KỊCH BẢN ═══

    `1-kich-ban.txt` còn được khâu phụ đề và khâu ảnh bìa đọc. Khâu phụ đề **ép
    chính chữ ấy lên giọng đọc**, nên thẻ lọt vào đó là `[whispers]` hiện lên
    màn hình cho người xem đọc. Bản có thẻ phải nằm riêng.

    Không có tệp riêng ấy thì mọi thứ chạy y như trước — đó là đường lui của cả
    tính năng này, và nó là đường lui **không cần viết thêm dòng nào**.

    Đã có tệp thì dùng lại: chèn thẻ tốn mấy lượt gọi AI, và "Chạy tiếp" một
    lượt cũ không nên trả tiền lại cho việc đã làm.
    """
    tep = os.path.join(luot.thu_muc, TEP_CO_THE)
    da_co = _doc_chu(tep).strip()
    # ═══ BẢN CÓ THẺ SẴN TRÊN ĐĨA THÌ DÙNG, KHÔNG CẦN BẬT NÚT ═══
    #
    # Tệp này giờ có hai nguồn: nút "Chèn thẻ cảm xúc" ở Cài đặt (đường chèn
    # riêng bên dưới), HOẶC chính bước sửa của kênh trả về bài đã có thẻ
    # (`_tach_the_cam_xuc`). Nguồn thứ hai là ý của người soạn lời nhắc kênh
    # — họ đã quyết kênh này đọc có thẻ — nên không bắt họ bật thêm nút nào.
    # Vẫn kiểm khớp chữ: tệp có thể do bản tool cũ ghi, hoặc khách sửa tay.
    if da_co:
        if kiem_the(kich_ban, da_co):
            bc.ghi("  dùng bản đã có thẻ cảm xúc.")
            return da_co
        bc.ghi("  (bản chèn thẻ cũ không khớp kịch bản — bỏ)")
    try:
        from . import cai_dat  # noqa: PLC0415

        bat = bool(cai_dat.doc(bc.goc).get("the_cam_xuc", False))
    except Exception:  # noqa: BLE001
        bat = False
    if not bat:
        return kich_ban
    try:
        bc.ghi("  chèn thẻ cảm xúc cho giọng đọc…")

        # ═══ MỘT LƯỢT GỌI, KHÔNG LEO THANG THỬ LẠI ═══
        #
        # Cố ý **không** dùng `_goi`. `_goi` là thang bốn lần thử kèm nhịp lùi
        # 30–120 giây, và bên trong `goi_van_ban` còn ba lần đổi khoá nữa —
        # đúng như những việc phải làm cho bằng được cần.
        #
        # Chèn thẻ thì không phải loại việc ấy. Bỏ qua nó chỉ mất mấy cái thẻ,
        # còn kiên nhẫn với nó thì mất cả lượt chạy: đo 17/08/2026, khâu giọng
        # đọc đứng im hơn hai mươi lăm phút ở đúng đây, trong khi hàng đợi của
        # cổng trống rỗng.
        #
        # Một lượt, hỏng thì thôi. `chen_the` tự lo phần quay về bản sạch.
        # Thẻ là BƯỚC CUỐI của phần chữ, nên đi cùng đường với kịch bản: máy
        # nào bật "Kịch bản viết bằng Claude Code" thì chèn thẻ cũng qua đó
        # (`cho_kich_ban` trả về chính `bc` khi không bật — không đổi gì).
        bc = bc.cho_kich_ban()

        def goi(loi_nhac: str) -> str:
            # Đi qua `_khoa_chat` chứ KHÔNG tự ghép chuỗi: nó lo hai việc mà
            # bản ghép tay ở đây thiếu cả hai — ép khoá về thuần ASCII (mã kênh
            # có dấu là chết cả mẻ, xem `_khoa_ascii`) và nhét MÃ KÊNH vào khoá
            # (mọi kênh đều đánh số lượt từ `0001`, thiếu mã kênh là kênh này
            # đâm vào kênh kia, cổng trả 409 idempotency_conflict — đã mất một
            # lượt 25 phút vì đúng chuyện đó, 19/08/2026).
            return bc.goi_chat(loi_nhac, mo_hinh=bc.kenh.mo_hinh,
                               khoa=_khoa_chat(luot, "the-cam-xuc:{0}".format(
                                   len(loi_nhac))))

        co_the = chen_the(kich_ban, goi, giong_van=bc.kenh.giong_van,
                          ngon_ngu=bc.kenh.ngon_ngu, ghi=bc.ghi)
    except Exception as loi:  # noqa: BLE001 — thẻ là việc làm đẹp, không bắt buộc
        bc.ghi("  (bỏ qua thẻ cảm xúc: {0})".format(str(loi)[:100]))
        return kich_ban
    if not co_the:
        return kich_ban
    _ghi_chu(tep, co_the)
    return co_the


def _doi_cao_do_giong(bc: BoiCanh, mp3: str) -> bool:
    """Dịch nhẹ cao độ giọng đọc, nếu khách bật ở Cài đặt.

    ═══ PHẢI LÀM Ở ĐÂY, TRƯỚC KHÂU PHỤ ĐỀ ═══

    Khâu phụ đề nghe chính tệp này để ra mốc thời gian, và bảng cảnh bám theo
    những mốc ấy. Dịch cao độ **sau** khi đã có phụ đề thì phụ đề mô tả một tệp
    không còn tồn tại nữa.

    Phép dịch giữ nguyên độ dài (xem `loc_doi_cao_do`), nên kể cả làm đúng thứ
    tự thì mốc thời gian cũng không xê dịch. Nhưng đặt đúng chỗ vẫn hơn dựa vào
    một tính chất có thể đổi.

    Hỏng thì giữ nguyên giọng cũ và **nói ra** — khác với mấy bước vệ sinh im
    lặng ở trên. Khách bật nút này là họ đang chờ một thứ cụ thể; im lặng bỏ
    qua là để họ tin nhầm là đã làm.
    """
    try:
        from . import cai_dat  # noqa: PLC0415

        if not cai_dat.doc(bc.goc).get("doi_cao_do_giong", False):
            return False
        from .lam_sach import CENT_DOI, doi_cao_do  # noqa: PLC0415

        ffmpeg = bc.ffmpeg or _tim_ffmpeg()
        if not ffmpeg:
            bc.ghi("  (không đổi được cao độ: máy chưa có FFmpeg)")
            return False
        bc.ghi("  đổi nhẹ cao độ giọng ({0} cent)…".format(CENT_DOI))
        if doi_cao_do(ffmpeg, mp3):
            return True
        bc.ghi("  (đổi cao độ không thành — giữ nguyên giọng gốc)")
    except Exception as loi:  # noqa: BLE001
        try:
            bc.ghi("  (đổi cao độ không thành: {0})".format(str(loi)[:80]))
        except Exception:  # noqa: BLE001
            pass
    return False


#: Kịch bản ngắn hơn bấy nhiêu phần của mục tiêu thì coi như không dùng được.
#:
#: `_nan_do_dai` đã cố nắn cho vừa; con số này là **sàn cuối**, chỗ nói "cố
#: không nổi thì dừng" thay vì đi tiếp.
#:
#: 0,45 chứ không phải 0,8: một kịch bản hụt 30-40% vẫn là kịch bản thật, chỉ
#: là video ngắn hơn ý muốn — dừng ở đó là cướp của khách một bài dùng được.
#: Còn hụt quá nửa thì nó không còn là "hơi ngắn", nó là một thứ khác hẳn.
SAN_DO_DAI_KICH_BAN = 0.45
#: Sàn tuyệt đối (ký tự) khi kênh để độ dài tự do — dưới mức này không phải kịch bản.
SAN_KICH_BAN_TU_DO = 1500


def _kiem_kich_ban_dung_duoc(so_ky_tu: int, muc_tieu: int,
                             duong_kb: str = "", tu_do: bool = False) -> None:
    """Kịch bản ngắn tới mức vô lý thì dừng NGAY, đừng đem đi đọc.

    ═══ MỘT LƯỢT CHẠY THẬT, 18/08/2026 ═══

    Tệp `1-kich-ban.txt` của lượt R01 chứa nguyên văn:

        "Bạn gửi tôi một kịch bản bằng tiếng Nhật, nhưng yêu cầu đánh giá so với
         kịch bản tiếng Việt đã viral. Tôi cần **kịch bản tiếng Việt** mà bạn
         vừa viết để đánh giá và sửa. Bạn có thể gửi lại không?"

    Đó là AI **hỏi lại**, không phải kịch bản. Tool ghi câu ấy vào tệp kịch bản,
    in ra `lệch 94%`, rồi báo khâu **XONG** và đem 218 ký tự ấy đi tạo giọng
    nói. Nếu không ai để ý, nó sẽ chạy tiếp qua phụ đề, cắt cảnh, và hàng trăm
    lượt tạo ảnh — tất cả dựng từ một câu hỏi.

    ═══ VÌ SAO ĐO ĐỘ DÀI, KHÔNG ĐI DÒ CÂU HỎI ═══

    Bắt "AI đang hỏi lại" bằng cách dò chữ là việc bạc: kịch bản tiếng Việt gọi
    khán giả là "bạn" ở gần như mọi câu, và câu hỏi tu từ là một lối viết hay
    dùng. Dò kiểu ấy sẽ giết những bài hoàn toàn tốt.

    Nhưng **mọi** kiểu hỏng ở khâu này đều để lại một dấu vết chung: bản ra
    ngắn. AI hỏi lại thì ngắn. AI từ chối thì ngắn. Trả về rỗng thì ngắn. Bị cắt
    giữa chừng thì ngắn. Một phép đo bắt được cả bốn, và không nhầm với văn hay.
    """
    # Độ dài tự do (mục tiêu 0) vẫn cần sàn: bản rỗng, AI hỏi lại, bị cắt
    # giữa chừng đều ngắn — sàn tuyệt đối thay cho sàn theo mục tiêu.
    if tu_do:
        san = SAN_KICH_BAN_TU_DO
    elif muc_tieu > 0:
        san = int(muc_tieu * SAN_DO_DAI_KICH_BAN)
    else:
        return  # không biết mục tiêu (dán tay, không link) thì không chặn gì — như cũ
    if so_ky_tu >= san:
        return
    # ═══ DỜI BẢN HỎNG SANG MỘT BÊN, ĐỪNG ĐỂ NÓ CHẶN LƯỢT SAU ═══
    #
    # Khâu này mở đầu bằng `if not ban_nhap:` — có tệp kịch bản rồi thì nó bỏ
    # qua cả phần viết. Nên nếu để nguyên bản hỏng, ba lượt thử lại của
    # `core/auto.chay` đều đọc lại đúng câu ấy và hỏng y hệt, còn người dùng
    # bấm “Chạy tiếp” bao nhiêu lần cũng vậy.
    #
    # ĐỔI TÊN chứ không xoá: đây vẫn là thứ khách trả tiền để tạo ra, và nó là
    # bằng chứng duy nhất cho biết máy đã trả về cái gì.
    _doi_ten_ban_hong(duong_kb)
    raise LoiNoiDung(
        "kịch bản chỉ có {0} ký tự, dưới sàn {1} ký tự{2} — ngắn tới mức này "
        "thì thường không phải bài viết, mà là câu AI hỏi lại hoặc trả về dở. "
        "Dừng ở đây thay vì đem nó đi tạo giọng nói và hàng trăm tấm ảnh. "
        "Bản hỏng đã dời sang 1-kich-ban-KHONG-DUNG-DUOC.txt để bạn xem máy đã "
        "trả về gì; bấm Chạy tiếp là tool viết lại từ đầu."
        # Nói SÀN trước, mục tiêu sau. Kênh để độ dài tự do thì `muc_tieu` là
        # 0, mà câu cũ in thẳng `muc_tieu` nên màn hình hiện "cần khoảng 0" —
        # vô nghĩa với người đọc, lại giấu mất con số thật đang chặn (đo
        # 26/08/2026, lượt thử kênh openstory: bài 1.452 ký tự bị sàn 1.500
        # chặn, màn hình nói 0).
        .format(so_ky_tu, san,
                " (mục tiêu {0})".format(muc_tieu) if muc_tieu > 0 else ""))


def _doi_ten_ban_hong(duong_kb: str) -> None:
    """Dời bản hỏng sang `…-KHONG-DUNG-DUOC.txt`, đừng để nó chặn lượt sau.

    Khâu kịch bản mở đầu bằng `if not ban_nhap:` — có tệp rồi thì nó bỏ qua cả
    phần viết. Để nguyên bản hỏng thì ba lượt thử lại của `core/auto.chay` đều
    đọc lại đúng bản ấy và hỏng y hệt. ĐỔI TÊN chứ không xoá: đây vẫn là thứ
    khách trả tiền để tạo ra, và là bằng chứng duy nhất máy đã trả về cái gì.
    """
    if not (duong_kb and os.path.exists(duong_kb)):
        return
    goc, duoi = os.path.splitext(duong_kb)
    try:
        os.replace(duong_kb, goc + "-KHONG-DUNG-DUOC" + duoi)
    except OSError:
        pass


def _kiem_ban_sach(bc: BoiCanh, ban: str, duong_kb: str = "") -> None:
    """Kịch bản còn lẫn ghi chú kỹ thuật thì DỪNG, đừng đem đi đọc.

    ═══ KHÁCH BÁO, 28/08/2026 ═══

    *"kịch bản trước khi voice nó bị lẫn cả các ghi chú kỹ thuật — tức nó là
    cái AI miêu tả kết quả lại đi kèm vào — như vậy thì ở logic hiện tại nó
    làm voice cả phần đó"*.

    `_don_ban` đã cắt được khối ghi chú ở đầu và ở cuối. Cái lọt tới đây là
    thứ nằm **giữa bài**, lẫn vào lời đọc — và giữa bài thì không có cách nào
    cắt mà chắc tay: cắt nhầm là mất một câu của bài, mà không dòng lỗi nào
    báo. Nên chốt này không sửa, nó **dừng**.

    ═══ VÌ SAO DỪNG CHỨ KHÔNG CHỈ GHI NHẬT KÝ ═══

    Kịch bản là chữ đầu dây chuyền. Đi tiếp từ đây là giọng đọc (vài phút),
    phụ đề ép theo chính chữ ấy, rồi hàng trăm tấm ảnh và clip cắt theo cảnh —
    tất cả dựng trên một bài có câu *"Đã chèn 32 thẻ cảm xúc v3"* nằm giữa.
    Dừng ở đây tốn một lượt viết lại; đi tiếp tốn cả lượt chạy.

    Chỉ dừng vì những dấu **không thể** là lời đọc (`DAU_KY_THUAT`: rào mã, tên
    cổng giọng nói, `write_file`, ô lời nhắc chưa điền…). Mấy dòng nhãn kiểu
    `"Tóm tắt:"` thì chỉ ghi nhật ký — "gần như chắc" không đủ để vứt một
    kịch bản đã trả tiền viết.
    """
    from .lam_sach import (ghi_chu_ky_thuat_con_lai,  # noqa: PLC0415
                           nhan_ghi_chu_con_lai)

    ngo = nhan_ghi_chu_con_lai(ban)
    if ngo:
        bc.ghi("  (ngờ còn ghi chú lẫn trong kịch bản: {0} — mở "
               "1-kich-ban.txt xem lại trước khi đăng)".format("; ".join(ngo[:3])))
    dau = ghi_chu_ky_thuat_con_lai(ban)
    if not dau:
        return
    _doi_ten_ban_hong(duong_kb)
    raise LoiNoiDung(
        "kịch bản còn lẫn ghi chú kỹ thuật của AI ({0}) — mấy chữ này không "
        "phải lời đọc, mà máy đọc giọng nói không phân biệt được nên nó sẽ "
        "đọc luôn vào video. Dừng ở đây thay vì đem đi tạo giọng nói và hàng "
        "trăm tấm ảnh. Bản hỏng đã dời sang 1-kich-ban-KHONG-DUNG-DUOC.txt để "
        "bạn xem máy đã trả về gì; bấm Chạy tiếp là tool viết lại từ đầu."
        .format(", ".join('"{0}"'.format(d) for d in dau[:3])))


def _lam_sach_ket_qua(bc: BoiCanh, *tep: str) -> None:
    """Bỏ dấu nguồn gốc AI khỏi kết quả của một khâu, nếu khách đã bật.

    Gọi ở **cuối mỗi khâu có tệp giao cho khách**: chữ (kịch bản, SEO, tiêu
    đề), giọng đọc, phụ đề, ảnh, video. Chủ dự án, 16/08/2026: *"cần bỏ dấu vân
    tay AI cho tất cả content, voice, ảnh, video"*.

    Nói thẳng cho người đọc mã sau này khỏi kỳ vọng nhầm: **phần lớn mấy tệp
    này vốn đã sạch**. Đo trên kết quả thật cùng ngày — chữ không có ký tự ẩn
    nào, giọng đọc và video cuối chỉ mang thẻ của FFmpeg. Chỗ hở thật chỉ có
    ảnh bìa. Gọi cho cả bốn loại là để **không phải nhớ** loại nào cần, và để
    còn đúng khi nhà cung cấp đổi cách gắn thẻ mà không báo ai.

    Không ném lỗi ra ngoài: đây là việc vệ sinh, không đáng làm hỏng một khâu
    khách đã trả tiền.
    """
    if not tep:
        return
    try:
        if not _bat_lam_sach(bc):
            return
        from .lam_sach import lam_sach_tep  # noqa: PLC0415

        ffmpeg = bc.ffmpeg or _tim_ffmpeg()
        for t in tep:
            try:
                lam_sach_tep(t, ffmpeg)
            except Exception:  # noqa: BLE001 — một tệp hỏng không dừng cả khâu
                pass
    except Exception:  # noqa: BLE001
        pass


def _bat_lam_sach(bc: BoiCanh) -> bool:
    """Khách có bật xoá dấu nguồn gốc không. **Hỏi một lần cho cả lượt chạy.**

    Hàm gọi nó chạy cho từng tấm ảnh, mà một lượt có hơn trăm tấm — đọc lại
    `workspace/cai-dat.json` cả trăm lần cho một câu trả lời không đổi là việc
    thừa, và nó nằm ngay trong luồng chạy song song của khâu ảnh, tức đúng chỗ
    không nên thêm việc.

    Đọc một lần cũng **đúng hơn**: khách gạt nút giữa chừng thì cả lượt vẫn xử
    như nhau, chứ không nửa số ảnh một kiểu.
    """
    da = getattr(bc, "_nho_lam_sach", None)
    if da is not None:
        return bool(da)
    from . import cai_dat  # noqa: PLC0415

    bat = bool(cai_dat.doc(bc.goc).get("lam_sach_dau_ai", True))
    try:
        bc._nho_lam_sach = bat  # noqa: SLF001 — nhớ trên chính bối cảnh lượt này
    except Exception:  # noqa: BLE001 — không gán được thì đọc lại, vẫn đúng
        pass
    return bat


class VanTay:
    """Nhớ mỗi tệp đã được tạo từ **lời nhắc nào**, để sửa lời nhắc là làm lại.

    ═══ VÌ SAO KHÔNG CHỈ HỎI "TỆP CÓ CHƯA" ═══

    Mọi khâu đều nhìn đĩa trước rồi bỏ qua nếu thấy tệp — đó là thứ giữ cho
    "Chạy tiếp" không trả tiền hai lần. Nhưng nó chỉ đúng khi **đầu vào không
    đổi**. Đo trên lượt thật TL4-T7/0051 (26/08/2026): khách sửa lời nhắc rồi
    "Làm lại từ khâu cắt cảnh" lúc 18:54, chạy tiếp — 169/173 ảnh và 172/172
    clip vẫn là bản cũ, video dựng ra y hệt lần trước. Cả buổi sửa lời nhắc
    không có tác dụng gì, mà nhìn bảng trạng thái vẫn thấy xanh hết.

    Nên cạnh thư mục kết quả có một sổ nhỏ: `_van-tay.json` = {số cảnh: vân
    tay của lời nhắc đã dùng}. Tệp có sẵn mà vân tay khác thì lời nhắc đã đổi
    → làm lại. Vân tay giống thì bỏ qua như trước.

    ═══ KHÔNG CÓ SỔ THÌ KHÔNG ĐOÁN ═══

    Lượt chạy trước bản này không có sổ. Khi ấy `khac()` trả `False` — giữ
    đúng nết cũ, không tự ý tiêu tiền vẽ lại cả trăm tấm ảnh của khách chỉ vì
    tool vừa lên đời. Muốn làm lại thì có "Làm lại khâu này" và
    `don_canh_de_lam_lai`.
    """

    #: Sổ nằm ở thư mục LƯỢT, không nằm trong `5-anh/` hay `6-clip/`: thư mục
    #: kết quả là thứ khách mở ra xem và chép đi, đừng rắc tệp kỹ thuật vào đó.
    TEN_ANH = "_van-tay-anh.json"
    TEN_CLIP = "_van-tay-clip.json"

    def __init__(self, duong: str) -> None:
        self._duong = duong
        self._khoa = threading.Lock()
        try:
            with open(self._duong, "r", encoding="utf-8") as tep:
                self._so = {str(k): str(v) for k, v in json.load(tep).items()}
        except (OSError, ValueError, AttributeError):
            self._so = {}

    @staticmethod
    def _dau(*phan: Any) -> str:
        import hashlib  # noqa: PLC0415

        van = "|".join(str(m) for m in phan)
        return hashlib.sha1(van.encode("utf-8")).hexdigest()[:16]

    def khac(self, so: Any, *phan: Any) -> bool:
        """Đã ghi vân tay cho `so` và vân tay ấy KHÁC đầu vào lần này chưa?"""
        with self._khoa:
            cu = self._so.get(str(so))
        return bool(cu) and cu != self._dau(*phan)

    def dat(self, so: Any, *phan: Any) -> None:
        """Ghi vân tay cho `so` rồi lưu sổ ra đĩa (an toàn với đa luồng)."""
        with self._khoa:
            self._so[str(so)] = self._dau(*phan)
            ban = dict(self._so)
        try:
            ghi_json(self._duong, ban, indent=None)
        except OSError:
            # Sổ hỏng không được làm hỏng lượt chạy: mất sổ thì chỉ mất khả
            # năng phát hiện lời nhắc đổi, ảnh vẫn có và tiền vẫn không mất.
            pass


def _cat_tep_cu(tep: str) -> bool:
    """Đổi tên `tep` thành `tep.cu` (giữ lại chứ không xoá). Trả True nếu có làm."""
    if not os.path.exists(tep):
        return False
    try:
        os.replace(tep, tep + ".cu")
    except OSError:
        return False
    return True


def _bo_clip_cu(bc: BoiCanh, tep_clip: str) -> bool:
    """Ảnh của cảnh vừa được tạo lại → clip cũ (nếu còn) cất đi (`<n>.mp4.cu`), trả True.

    Gọi đúng lúc tool tạo lại ảnh — không so mtime: ảnh còn bị ghi lại sau khi
    clip đã tạo (xoá dấu, làm sạch thẻ), so mtime là bắt nhầm (25/08/2026: sáu
    clip bị làm lại vô cớ, 3.000 ₫).
    """
    if not os.path.exists(tep_clip):
        return False
    try:
        os.replace(tep_clip, tep_clip + ".cu")
    except OSError:
        return False
    bc.ghi("    {0}: ảnh vừa làm lại — clip cũ lỗi thời, sẽ tạo lại (bản cũ giữ ở .cu)."
           .format(os.path.basename(tep_clip)))
    return True


def _bo_clip_cu_hon_anh(bc: BoiCanh, tep_clip: str, tep_anh: str) -> bool:
    """Clip CŨ HƠN ảnh của chính nó thì cất đi (`<n>.mp4.cu`) và trả True.

    ═══ VÌ SAO ═══

    Clip của cảnh nào là ảnh cảnh ấy cử động. Khách "Làm lại khâu ảnh" để sửa
    một tấm xấu, tool tạo ảnh mới — nhưng clip cũ vẫn nằm đó, khâu clip thấy
    "đã có" nên bỏ qua, và video cuối vẫn là con mèo cũ. Đo 25/08/2026
    (story-3d/0001): 18 ảnh mèo làm lại lúc 20:05, 18 clip vẫn là bản 19:3x.
    Ảnh mới hơn clip nghĩa là clip đã lỗi thời — làm lại, không bỏ qua.
    """
    try:
        if not (os.path.exists(tep_clip) and os.path.exists(tep_anh)):
            return False
        if os.path.getmtime(tep_clip) >= os.path.getmtime(tep_anh):
            return False
        os.replace(tep_clip, tep_clip + ".cu")
    except OSError:
        return False
    bc.ghi("    {0}: ảnh mới hơn clip — làm lại clip (bản cũ giữ ở .cu).".format(
        os.path.basename(tep_clip)))
    return True


def _nguon_moi_hon_video(tep_video: str, thu_muc_clip: str, mp3: str) -> str:
    """Tên nguồn (clip hoặc giọng đọc) MỚI HƠN video cuối, hoặc "" nếu video còn đúng."""
    try:
        moc = os.path.getmtime(tep_video)
    except OSError:
        return ""
    try:
        if os.path.exists(mp3) and os.path.getmtime(mp3) > moc:
            return "giọng đọc"
        for ten in sorted(os.listdir(thu_muc_clip)) if os.path.isdir(thu_muc_clip) else []:
            if ten.endswith(".mp4") and os.path.getmtime(os.path.join(thu_muc_clip, ten)) > moc:
                return "clip " + ten[:-4]
    except OSError:
        return ""
    return ""


def _khau_anh(bc: BoiCanh):
    """Khâu ảnh — và thật ra là **cả dây chuyền ảnh → clip**.

    ═══ VÌ SAO BA KHÂU GỘP LÀM MỘT MẺ ═══

    Đo trên máy chủ thật, 15/08/2026: một tấm ảnh mất 30,8 giây ở nhà máy,
    nhưng **ba mươi tấm bắn cùng lúc thì xong hết trong 38,2 giây** — nhà máy
    chạy song song thật, 30 tấm gần như không đắt hơn 1 tấm. Cổng cho 979 job
    ảnh và 172–316 job clip chạy cùng lúc, hàng chờ 100.000, và nói rõ *"gửi
    nhiều hơn KHÔNG bị từ chối"*.

    Vậy mà tool mất 5,9 phút cho 114 ảnh, 14,8 phút cho 114 clip, và **3,3 phút
    cho đúng 3 tấm ảnh bìa** — vì ba việc ấy là ba hàng rào nối đuôi nhau: ảnh
    bìa đợi clip, clip đợi **đủ** 114 ảnh.

    Không có hàng rào nào trong đó là thật:

    * ảnh bìa chẳng cần clip nào — nó chỉ cần tiêu đề, có từ khâu 1;
    * clip của cảnh 7 chỉ cần **ảnh của cảnh 7**, không cần 113 ảnh kia.

    Nên mẻ này bắn **hết một lượt**: 114 ảnh cảnh cộng 3 ảnh bìa. Ảnh nào về
    thì tải xuống rồi bắn clip của chính nó ngay, không đợi ai.

    Hai khâu sau (`clip`, `thumbnail`) vẫn còn nguyên và vẫn chạy — chúng nhìn
    đĩa trước, thấy đủ tệp thì đi qua trong một nháy, thiếu cái nào thì làm nốt
    cái ấy. Nhờ vậy mọi thứ giữ nguyên: bảng trạng thái tám khâu, "Làm lại khâu
    này", và **chạy tiếp nhặt đúng chỗ đứt**.
    """

    def lam(luot: LuotChay, tt: TrangThaiKhau):
        from .auto import Cancelled  # noqa: PLC0415
        from .su_co import NHA_MAY_NGHI, phan_loai as _phan  # noqa: PLC0415

        canh = _doc_canh(luot)
        thu_muc = os.path.join(luot.thu_muc, "5-anh")
        thu_muc_clip = os.path.join(luot.thu_muc, "6-clip")
        os.makedirs(thu_muc, exist_ok=True)
        os.makedirs(thu_muc_clip, exist_ok=True)
        # Lấy URL tham chiếu TRƯỚC khi bung luồng: để trong luồng thì cả trăm
        # luồng cùng thấy chưa có rồi cùng tải một tệp lên.
        hop = ThamChieu(bc)
        # Sổ vân tay lời nhắc: sửa lời nhắc rồi chạy tiếp thì phải vẽ lại,
        # không được lấy ảnh cũ ra dùng (xem `VanTay`).
        van_tay = VanTay(os.path.join(luot.thu_muc, VanTay.TEN_ANH))
        van_tay_clip = VanTay(os.path.join(luot.thu_muc, VanTay.TEN_CLIP))
        thu_muc_bia, muc_bia, _thieu_bia, ta_bia, tieu_de, chu_bia = \
            _chuan_bi_bia(bc, luot)
        giay = _giay_clip(bc)
        so = SoTheoDoi(bc, nhip=bc.nhip_hoi)

        # ═══ MỘT KHOÁ CHO CẢ BA BỘ ĐẾM ═══
        #
        # `dem_tien_do` ghi thẳng ra `trang-thai.json`. Trước đây chỉ luồng
        # chính gọi nó; giờ mỗi luồng tự báo khi việc của mình xong, mà hai
        # luồng cùng ghi một tệp qua đường tệp tạm thì bản này đè bản kia —
        # trên Windows còn ném hẳn lỗi giữa mẻ.
        khoa_dem = threading.Lock()
        dem = {"anh": 0, "clip": 0, "bia": 0}
        tong = {"anh": len(canh), "clip": len(canh), "bia": len(muc_bia)}
        bao = {
            "anh": dem_tien_do(bc, luot, tt, "ảnh", NHIP_GHI_TIEN_DO),
            "clip": dem_tien_do(bc, luot, luot.tt("clip"), "clip",
                                NHIP_GHI_TIEN_DO),
            "bia": dem_tien_do(bc, luot, luot.tt("thumbnail"), "ảnh bìa",
                               NHIP_GHI_TIEN_DO),
        }
        for loai in ("anh", "clip", "bia"):
            bao[loai](0, tong[loai])

        def them(loai: str) -> None:
            with khoa_dem:
                dem[loai] += 1
                bao[loai](dem[loai], tong[loai])

        #: Nhà máy clip tắt thì thôi bắn clip, nhưng **vẫn làm nốt ảnh**: ảnh đã
        #: có trên đĩa là tiền đã tiêu không mất, và khi nhà máy clip bật lại
        #: thì khâu clip chỉ còn việc bắn.
        clip_tat = threading.Event()

        def bat_clip(c, tep_anh: str) -> None:
            """Ảnh vừa về thì bắn clip của chính nó — không đợi các cảnh khác."""
            so_canh = int(c["scene_id"])
            dich = os.path.join(thu_muc_clip, "{0}.mp4".format(so_canh))
            if os.path.exists(dich):
                if van_tay_clip.khac(so_canh, c.get("video_prompt") or ""):
                    bc.ghi("    cảnh {0}: lời nhắc clip đã đổi — làm lại "
                           "clip.".format(so_canh))
                    _cat_tep_cu(dich)
                else:
                    them("clip")
                    return
            if clip_tat.is_set() or not os.path.exists(tep_anh):
                return
            if not str(c.get("video_prompt") or "").strip():
                return
            try:
                anh_cuoi = (_anh_khung_cuoi(bc, luot, c, tep_anh,
                                            _hop_cho_canh(bc, luot, c, hop), so=so)
                            if _ghim_hai_dau(bc) else "")
                _lam_clip(bc, luot, c, tep_anh, dich, giay, so=so,
                          khung_dau=_co_khung_dau(bc) or bool(anh_cuoi),
                          anh_cuoi=anh_cuoi or None)
                van_tay_clip.dat(so_canh, c.get("video_prompt") or "")
            except Cancelled:
                raise
            except Exception as loi:  # noqa: BLE001
                # ═══ CLIP HỎNG KHÔNG ĐƯỢC LÀM HỎNG KHÂU ẢNH ═══
                #
                # Ở đây clip là việc **làm thêm cho sớm**. Nó hỏng thì phần ảnh
                # vẫn đúng và vẫn phải được ghi nhận là xong — khâu clip ngay
                # sau đó sẽ làm nốt, và nếu vẫn hỏng thì nó mới là chỗ báo lỗi.
                # Ném lên từ đây là đổ tội cho khâu ảnh một việc không phải của
                # nó, rồi bảng trạng thái chỉ vào sai chỗ.
                if _phan(loi) == NHA_MAY_NGHI:
                    clip_tat.set()
                    bc.ghi("  nhà máy clip đang tắt — làm nốt ảnh đã, phần "
                           "clip để khâu sau.")
                else:
                    bc.ghi("    cảnh {0}: chưa ra clip ({1}) — khâu clip sẽ "
                           "làm nốt.".format(so_canh, str(loi)[:80]))
                return
            them("clip")

        def mot_muc(m):
            loai, x = m
            if loai == "bia":
                ket = _lam_bia(bc, luot, _hop_bia(bc, luot, hop), thu_muc_bia, x, ta_bia,
                               tieu_de, chu_bia, so=so)
                them("bia")
                return ket
            so_canh = int(x["scene_id"])
            tep = os.path.join(thu_muc, "{0}.png".format(so_canh))
            san_co = os.path.exists(tep)
            if san_co and van_tay.khac(so_canh, x.get("img_prompt") or ""):
                # Lời nhắc của cảnh này đã đổi kể từ lúc vẽ tấm đang nằm đây.
                # Giữ tấm cũ lại (`.cu`) rồi vẽ tấm mới — đây chính là chỗ mà
                # cả buổi sửa lời nhắc của khách từng không có tác dụng gì.
                bc.ghi("    cảnh {0}: lời nhắc đã đổi — vẽ lại ảnh.".format(
                    so_canh))
                _cat_tep_cu(tep)
                san_co = False
            if not san_co:
                _lam_anh_canh(bc, luot, x, tep, _hop_cho_canh(bc, luot, x, hop), so=so)
                van_tay.dat(so_canh, x.get("img_prompt") or "")
                # Ảnh VỪA tạo lại → clip cũ của cảnh này (nếu còn) đã lỗi thời.
                _bo_clip_cu(bc, os.path.join(thu_muc_clip, "{0}.mp4".format(so_canh)))
            else:
                # Ảnh có sẵn trên đĩa từ lượt chạy trước — có thể là lượt chạy
                # bằng bản tool chưa biết xoá dấu. Xoá lại ở đây thì lượt cũ
                # chạy tiếp cũng ra clip sạch. Ảnh đã sạch rồi thì `xoa_dau`
                # tự nhận ra và không đụng vào, nên gọi thừa không hại gì.
                _xoa_dau(bc, tep)
            them("anh")
            bat_clip(x, tep)
            return so_canh, san_co

        # Ảnh bìa đứng đầu hàng: chỉ có ba tấm, và trước đây chúng là cái đuôi
        # 3,3 phút chạy sau cùng khi mọi thứ khác đã xong.
        muc = [("bia", m) for m in muc_bia] + [("canh", c) for c in canh]
        try:
            _chay_song_song(bc, muc, mot_muc, "ảnh", loai_job="image")
        finally:
            so.dong()
        if dem["clip"] < tong["clip"]:
            bc.ghi("  còn {0}/{1} clip chưa xong — khâu clip làm nốt.".format(
                tong["clip"] - dem["clip"], tong["clip"]))
        return {"so_anh": dem["anh"], "so_clip": dem["clip"],
                "so_thumbnail": dem["bia"]}

    return lam


def _la_noi_canh(kenh: Any) -> bool:
    from .noi_canh import la_noi_canh  # noqa: PLC0415

    return la_noi_canh(kenh)


def _khau_anh_noi_canh(bc: BoiCanh):
    """Khâu ảnh ở chế độ NỐI CẢNH: ảnh N → clip N → cắt → khung cuối → ảnh N+1.

    Xem `core/noi_canh.py`. Chuỗi = dãy cảnh cùng bối cảnh; các chuỗi chạy song
    song, trong chuỗi tuần tự. Ảnh bìa để khâu thumbnail làm; clip thiếu để khâu
    clip làm nốt; video cuối vẫn qua khâu dựng (bản cắt ở đây khớp bản khâu dựng
    cắt lại).
    """
    def lam(luot: LuotChay, tt: TrangThaiKhau):
        from .dao_dien_auto import THU_MUC_THAM_CHIEU, ThamChieuCanh  # noqa: PLC0415
        from .noi_canh import (  # noqa: PLC0415
            ChuoiNoiCanh, CuMayDai, SONG_SONG_CHUOI, cat_clip_theo_canh, chay_cac_chuoi,
            chuoi_theo_boi_canh, giu_khung_dau, khung_cuoi, noi_cac_clip,
        )
        from .phan_cung import chon_encoder, doc_ket_qua  # noqa: PLC0415

        canh = _doc_canh(luot)
        thu_muc = os.path.join(luot.thu_muc, "5-anh")
        thu_muc_clip = os.path.join(luot.thu_muc, "6-clip")
        thu_muc_tc = os.path.join(luot.thu_muc, THU_MUC_THAM_CHIEU)
        os.makedirs(thu_muc, exist_ok=True)
        os.makedirs(thu_muc_clip, exist_ok=True)
        ffmpeg = bc.ffmpeg or _tim_ffmpeg()
        if not ffmpeg:
            raise RuntimeError("máy chưa có FFmpeg — chế độ nối cảnh cần FFmpeg để cắt clip và lấy khung cuối")
        codec, opts = chon_encoder(doc_ket_qua(bc.goc), intermediate=True)
        giay = _giay_clip(bc)
        so = SoTheoDoi(bc, nhip=bc.nhip_hoi)
        chuoi = chuoi_theo_boi_canh(canh)
        bc.ghi("  nối cảnh: {0} cảnh thành {1} chuỗi theo bối cảnh (chuỗi dài nhất {2} cảnh); "
               "tối đa {3} chuỗi chạy cùng lúc.".format(
                   len(canh), len(chuoi), max(len(x) for x in chuoi), SONG_SONG_CHUOI))
        khoa_dem = threading.Lock()
        dem = {"anh": 0, "clip": 0}
        bao = {"anh": dem_tien_do(bc, luot, tt, "ảnh", NHIP_GHI_TIEN_DO),
               "clip": dem_tien_do(bc, luot, luot.tt("clip"), "clip", NHIP_GHI_TIEN_DO)}
        for loai in ("anh", "clip"):
            bao[loai](0, len(canh))

        def them(loai: str):
            def _t() -> None:
                with khoa_dem:
                    dem[loai] += 1
                    bao[loai](dem[loai], len(canh))
            return _t

        def lam_anh(c, tep, refs, prompt):
            c2 = dict(c, img_prompt=prompt)
            _lam_anh_canh(bc, luot, c2, tep, ThamChieuCanh(bc, refs), so=so)
            _bo_clip_cu(bc, os.path.join(thu_muc_clip, "{0}.mp4".format(int(c["scene_id"]))))

        def lam_clip(c, anh, tho, anh_cuoi=None):
            _lam_clip(bc, luot, c, anh, tho, giay, so=so, khung_dau=_co_khung_dau(bc),
                      anh_cuoi=anh_cuoi)

        def cat(tho, clip, giay_canh):
            from .noi_canh import bat_dau_cat  # noqa: PLC0415

            cat_clip_theo_canh(ffmpeg, tho, clip, giay_canh, codec, opts,
                               bat_dau=bat_dau_cat(giay_canh, giay))

        def trich(clip, khung):
            return khung_cuoi(ffmpeg, clip, khung)

        def cat_tu(nguon, dich, bat_dau, giay_canh):
            cat_clip_theo_canh(ffmpeg, nguon, dich, giay_canh, codec, opts, bat_dau=bat_dau)

        def noi(nguon, dich):
            noi_cac_clip(ffmpeg, nguon, dich)

        loi_chung: List[str] = []
        # Khung đầu thật (Veo 3 Frames / Seedance): cả chuỗi là MỘT cú máy dài ghép từ
        # các đoạn 8 s. Không thì đường cũ: ảnh mới mỗi cảnh + khung cuối làm tham chiếu.
        cu_may_dai = giu_khung_dau(bc.kenh)
        if cu_may_dai:
            bc.ghi("  khung đầu thật: mỗi chuỗi là một cú máy dài, cắt từng cảnh ra từ đó.")

        def lam_chuoi(ch):
            chung = dict(thu_muc_anh=thu_muc, thu_muc_clip=thu_muc_clip, thu_muc_tham_chieu=thu_muc_tc,
                         lam_anh=lam_anh, lam_clip=lam_clip, cat=cat, trich_khung=trich,
                         ghi=bc.ghi, kiem_dung=bc.kiem_dung, bao_anh=them("anh"), bao_clip=them("clip"))
            if cu_may_dai:
                ct = CuMayDai(cat_tu=cat_tu, noi_clip=noi, lien_mach=True, **chung)
            else:
                ct = ChuoiNoiCanh(lien_mach=False, **chung)
            n = ct.chay(ch)
            with khoa_dem:
                loi_chung.extend(ct.loi)
            return n

        try:
            chay_cac_chuoi(chuoi, lam_chuoi, SONG_SONG_CHUOI)
        finally:
            so.dong()
        thieu_anh = [int(c["scene_id"]) for c in canh
                     if not os.path.exists(os.path.join(thu_muc, "{0}.png".format(int(c["scene_id"]))))]
        if thieu_anh and not cu_may_dai:
            bc.ghi("  ảnh: thiếu {0}/{1} ({2}) — “Làm lại khâu này” làm nốt.".format(
                len(thieu_anh), len(canh), ", ".join(str(x) for x in thieu_anh[:12])))
        if dem["clip"] < len(canh):
            bc.ghi("  còn {0}/{1} clip chưa xong — khâu clip làm nốt.".format(len(canh) - dem["clip"], len(canh)))
        return {"so_anh": dem["anh"], "so_clip": dem["clip"], "so_chuoi": len(chuoi),
                "loi": loi_chung[:20]}

    return lam


def _khau_clip(bc: BoiCanh):
    """Làm nốt những clip dây chuyền ở khâu ảnh chưa kịp ra.

    Bình thường khâu này chỉ còn việc nhìn đĩa rồi đi qua — clip đã được bắn từ
    lúc ảnh của nó vừa về. Nó vẫn phải đứng đây vì hai lý do: người dùng có
    quyền bấm "Làm lại khâu này" cho riêng clip, và nếu nhà máy clip tắt giữa
    chừng thì đây là chỗ làm lại phần còn thiếu.
    """

    def lam(luot: LuotChay, tt: TrangThaiKhau):
        canh = _doc_canh(luot)
        thu_muc = os.path.join(luot.thu_muc, "6-clip")
        thu_muc_anh = os.path.join(luot.thu_muc, "5-anh")
        os.makedirs(thu_muc, exist_ok=True)
        giay = _giay_clip(bc)
        so = SoTheoDoi(bc, nhip=bc.nhip_hoi)
        van_tay_clip = VanTay(os.path.join(luot.thu_muc, VanTay.TEN_CLIP))

        def mot_canh(c):
            so_canh = int(c["scene_id"])
            tep = os.path.join(thu_muc, "{0}.mp4".format(so_canh))
            anh = os.path.join(thu_muc_anh, "{0}.png".format(so_canh))
            if os.path.exists(tep):
                # ═══ "ĐÃ CÓ TỆP" CHƯA ĐỦ ĐỂ BỎ QUA ═══
                #
                # Hai cách một clip trở nên lỗi thời mà tệp vẫn nằm đó:
                # ảnh của chính nó vừa được vẽ lại (`_bo_clip_cu_hon_anh` —
                # viết ra từ 25/08 nhưng chưa nơi nào gọi), và lời nhắc clip
                # đã đổi (`VanTay`). Cả hai đều từng làm khách xem lại đúng
                # video cũ sau khi đã sửa và trả tiền cho lượt sửa.
                cu = _bo_clip_cu_hon_anh(bc, tep, anh)
                if not cu and van_tay_clip.khac(so_canh, c.get("video_prompt") or ""):
                    bc.ghi("    cảnh {0}: lời nhắc clip đã đổi — làm lại "
                           "clip.".format(so_canh))
                    _cat_tep_cu(tep)
                    cu = True
                if not cu:
                    return so_canh, True
            # Cùng luật với dây chuyền ở khâu ảnh — hai nhánh lệch nhau đúng
            # một tham số là cờ chết lặng, không ai thấy (bài học 27/08/2026
            # với `khung_dau`, xem `_lam_clip`).
            anh_cuoi = (_anh_khung_cuoi(bc, luot, c, anh,
                                        _hop_cho_canh(bc, luot, c, _HopTrong()), so=so)
                        if _ghim_hai_dau(bc) else "")
            _lam_clip(bc, luot, c, anh, tep, giay, so=so,
                      khung_dau=_co_khung_dau(bc) or bool(anh_cuoi),
                      anh_cuoi=anh_cuoi or None)
            van_tay_clip.dat(so_canh, c.get("video_prompt") or "")
            return so_canh, False

        try:
            xong = _chay_song_song(
                bc, canh, mot_canh, "clip", loai_job="video",
                nhip=dem_tien_do(bc, luot, tt, "clip", NHIP_GHI_TIEN_DO))
        finally:
            so.dong()
        return {"so_clip": xong}

    return lam


# ── Khâu 7: ảnh bìa ──────────────────────────────────────────────────────────

#: Ba kiểu ảnh bìa khác nhau, không phải ba bản ngẫu nhiên của cùng một kiểu —
#: chép đúng nết `version_desc` của tool cũ (`portrait_main`, `dramatic_scene`…).
#:
#: ⚠ TÊN Ở ĐÂY PHẢI KHỚP TÊN TRONG `prompt/8-thumbnail.md` CỦA KÊNH.
#:
#: Đây là chỗ làm chữ bìa mất hẳn trên tấm thứ ba, và triệu chứng không hề chỉ
#: về đây. Đo trên lượt chạy thật TL4-T7 lượt 0009 (22/08/2026): cả mười một
#: tệp `8-thumbnail.md` trên đĩa đều xin AI trả về `youtube_ctr`, còn bảng này
#: gọi tấm thứ ba là `symbolic_object`. Tra `ta_bia["symbolic_object"]` là
#: **trượt**, nên tấm thứ ba rơi về `_bia_du_phong` — bản ghép cứng vốn không
#: có một dòng nào yêu cầu chữ. Kết quả: hai tấm có chữ, tấm thứ ba trắng chữ,
#: và không có lấy một dòng nhật ký nào nói vì sao.
#:
#: Nên tên tấm thứ ba đổi về `youtube_ctr` cho khớp lời nhắc, và việc tra thì đi
#: qua `_lay_ta_bia` — có tên khác thì nhận, AI đặt tên lạ thì lấy theo thứ tự.
KIEU_THUMB = (
    ("portrait_main", "close-up portrait of the reference character, direct "
                      "emotional gaze, single clear feeling"),
    ("dramatic_scene", "the most emotionally charged moment of the story, "
                       "character small in a meaningful environment"),
    ("youtube_ctr", "one strong symbolic object in the foreground with the "
                    "character reacting behind it"),
)

#: Tên cũ / tên khác cho cùng một kiểu bìa. Lời nhắc là tệp người dùng sửa được,
#: nên tên trong đó sẽ lệch — nhận cả họ tên thay vì bắt gõ đúng một chữ.
_TEN_BIA_KHAC: Dict[str, Tuple[str, ...]] = {
    "youtube_ctr": ("symbolic_object", "high_ctr", "symbolic"),
    "portrait_main": ("portrait", "main_portrait"),
    "dramatic_scene": ("dramatic", "scene"),
}


def _lay_ta_bia(ta_bia: Dict[str, str], ten_kieu: str, so_bia: int) -> str:
    """Lấy lời nhắc AI đã viết cho tấm bìa này. Không có thì trả "".

    Ba tầng, nới dần: đúng tên → tên khác đã biết → **theo thứ tự**. Tầng cuối
    là tầng quan trọng nhất: lời nhắc nằm trong tệp người dùng sửa được, nên AI
    có thể trả về ba cái tên chẳng giống bảng nào. Ba lời nhắc đúng thứ tự vẫn
    tốt hơn hẳn một bản ghép cứng — miễn là còn đủ ba cái.
    """
    if not ta_bia:
        return ""
    for ten in (ten_kieu,) + _TEN_BIA_KHAC.get(ten_kieu, ()):
        if ta_bia.get(ten):
            return ta_bia[ten]
    ds = [v for v in ta_bia.values() if v]
    if len(ds) >= so_bia >= 1:
        return ds[so_bia - 1]
    return ""


#: Bắt lời nhắc ảnh bìa dùng ĐÚNG chữ bìa, không tự nghĩ chữ khác.
#:
#: Kênh remake "gần như giống đối thủ nhất" thì chữ trên bìa là chữ đã đọc được
#: từ bìa đối thủ — cả công đoạn đọc ảnh chỉ để có đúng dòng chữ ấy. Mà
#: `8-thumbnail.md` lại bảo AI *"text: <hook in the channel's language>"*, tức
#: mời nó tự nghĩ một câu hook mới. Đo lượt 0009: tấm thứ hai đội chữ
#: 「温度が違う理由」 — câu AI tự bịa, không có trên bìa đối thủ.
#: ═══ CHỮ BÌA NGUYÊN VĂN THÌ NGÂN SÁCH 14 KÝ TỰ KHÔNG ÁP DỤNG ═══
#:
#: `8-thumbnail.md` ép *"tối đa 2 khối chữ và 14 ký tự TỔNG"*, và gọi con số
#: ấy là NON-NEGOTIABLE. Luật đó đúng cho kênh tự nghĩ câu hook ngắn.
#:
#: Nhưng kênh `nguyen_goc` lấy NGUYÊN chữ bìa đối thủ, và chữ ấy dài bao nhiêu
#: là chuyện của đối thủ. Lượt 0009: 『これ』を一人でしているなら あなたのIQは
#: 非常に高いかも — **27 ký tự**, nhét vào ngân sách 14 là bất khả.
#:
#: Model xử mâu thuẫn bằng cách đẻ ra bốn khối rồi rải chữ, và 「かも」 —
#: một trợ từ KẾT CÂU — rơi vào giữa: người Nhật đọc ra 「あなたのIQは / かも /
#: 非常に高い」, câu vỡ. Cả ba tấm của lượt 0009 đều hỏng kiểu này.
#:
#: Nên phải nói thẳng hai điều: ngân sách ký tự KHÔNG áp dụng, và thứ tự đọc
#: là bất khả xâm phạm.
_LUAT_CHU_BIA_NGUYEN = (
    "\n\n## MANDATORY — EXACT THUMBNAIL TEXT\n"
    "The hook text is FIXED ({1} characters). It must appear on the image "
    "EXACTLY as written below, character for character:\n\n    {0}\n\n"
    "This overrides the character budget in the `TEXT STYLE` block: ignore the "
    "\"maximum 2 text blocks and 14 characters TOTAL\" limit and use as many "
    "blocks as this text needs. Shrink the type instead — the whole string "
    "must fit.\n"
    "READING ORDER IS FIXED. Reading the blocks top to bottom, left to right "
    "must give back exactly the sentence above, word for word. Break it only "
    "between clauses; never lift a word, particle or suffix out of the middle "
    "of a phrase to use as a separate block.\n"
    "Do not translate it, rewrite it, shorten it or extend it. Add no other "
    "text to the image.\n"
)


def _loi_nhac_bia(bc: BoiCanh, luot: LuotChay, khuon: str, tieu_de: str,
                  chu_bia: str, kieu) -> Dict[str, str]:
    """Nhờ AI viết lời nhắc cho ba ảnh bìa. Hỏng thì trả rỗng, không giết khâu.

    Ảnh bìa thiếu thì video vẫn dùng được — nên chỗ này không được phép làm
    hỏng cả lượt. Hỏng thì rơi về bản ghép cứng ở `_bia_du_phong`.
    """
    if not khuon.strip():
        return {}
    st = bc.kenh.style
    mo_dau = _doc_chu(os.path.join(luot.thu_muc, "1-kich-ban.txt"))[:1200]
    loi_nhac = _thay(khuon, {
        "TITLE": tieu_de, "THUMB": chu_bia,
        "SCRIPT_OPENING": mo_dau,
        "THUMBNAIL_STYLE": st.get("thumbnail_style", st.get("image_style", "")),
        "PALETTE": st.get("palette", ""),
        "REFERENCE_LOCK": st.get("reference_lock", ""),
        "NEGATIVE_PROMPT": st.get("negative_prompt", ""),
        "THUMB_TEXT_STYLE": st.get("thumb_text_style", ""),
        "THUMB_TEXT_FONT": st.get("thumb_text_font", ""),
        "THUMB_TEXT_SHADOW": st.get("thumb_text_shadow", ""),
    })
    # Kênh lấy nguyên chữ bìa đối thủ thì chữ ấy là **cố định** — chốt lại, kẻo
    # `8-thumbnail.md` mời AI tự nghĩ một câu hook mới (xem `_LUAT_CHU_BIA_NGUYEN`).
    if bc.kenh.che_do_tieu_de == "nguyen_goc" and chu_bia.strip():
        loi_nhac += _LUAT_CHU_BIA_NGUYEN.format(chu_bia.strip(),
                                                len(chu_bia.strip()))
    try:
        goi = loc_json(_goi(bc, loi_nhac,
                            khoa_viec(luot, "chat", "thumb", tieu_de, chu_bia)))
    except Exception as loi:  # noqa: BLE001
        bc.ghi("  (AI chưa viết được lời nhắc ảnh bìa: {0}) — dùng bản mặc "
               "định.".format(str(loi)[:90]))
        return {}
    ds = goi.get("thumbnails") if isinstance(goi, dict) else goi
    ra: Dict[str, str] = {}
    for m in (ds or []):
        if isinstance(m, dict) and m.get("img_prompt"):
            ra[str(m.get("version_desc") or "")] = str(m["img_prompt"])
    if ra:
        bc.ghi("  AI viết {0} lời nhắc ảnh bìa.".format(len(ra)))
    return ra


def _bia_du_phong(st: Dict[str, Any], tieu_de: str, chu_bia: str,
                  ta: str) -> str:
    """Bản ghép cứng, chỉ dùng khi AI không viết được.

    ═══ BẢN NÀY CŨNG PHẢI XIN CHỮ ═══

    Bản trước chỉ đưa chữ bìa vào làm *"Emotional message"* — một câu tả tâm
    trạng, không phải một yêu cầu in chữ. Nên tấm nào rơi vào đây là tấm ấy ra
    lò **không có chữ nào**, trong khi hai tấm kia (do AI viết lời nhắc) đội chữ
    to đùng. Đo trên lượt 0009: đúng tấm thứ ba như thế.

    Ảnh bìa không chữ thì gần như vô dụng với kênh này, nên bản đường-cùng vẫn
    phải xin chữ — kèm cả kiểu chữ của kênh nếu có.
    """
    kieu_chu = " ".join(m for m in (st.get("thumb_text_font", ""),
                                    st.get("thumb_text_style", ""),
                                    st.get("thumb_text_shadow", "")) if m)
    return ("{0}\nYouTube thumbnail, {1}. Video topic: {2}.\n"
            "TEXT STYLE (HIGH CTR YOUTUBE):\n"
            'text: "{3}"\n'
            "render this text large and dominant, integrated into the "
            "composition, negative space reserved for it. {4}\n"
            "no other text, no watermark, no logo. {5} Avoid: {6}".format(
                st.get("thumbnail_style", st.get("image_style", "")),
                ta, tieu_de, chu_bia, kieu_chu,
                st.get("reference_lock", ""), st.get("negative_prompt", "")))


def _tep_bia(thu_muc: str, so: int) -> str:
    return os.path.join(thu_muc, "thumb_{0:03d}.png".format(so))


def _chuan_bi_bia(bc: BoiCanh, luot: LuotChay):
    """Dọn sẵn mọi thứ để tạo ảnh bìa. Trả về `(thư mục, mục, thiếu, lời nhắc,
    tiêu đề, chữ bìa)`.

    ═══ LỜI NHẮC ẢNH BÌA DO AI VIẾT, KHÔNG PHẢI CHUỖI GHÉP CỨNG ═══

    Bản trước ghép chuỗi ngay trong code: style + một câu tả kiểu + tiêu đề. Ra
    ảnh đúng phong cách nhưng **không có chữ hook** và không bám nội dung —
    trong khi ảnh bìa là thứ duy nhất quyết định người ta có bấm vào hay không.

    Tool gốc để AI viết ba lời nhắc, ba ý cảm xúc khác nhau, và ảnh bìa thì
    **CÓ chữ** (khác hẳn các cảnh trong video vốn không có chữ nào). Nay cũng
    vậy, và lời nhắc nằm ở `prompt/8-thumbnail.md` của kênh nên người dùng sửa
    được.

    **Chỉ hỏi AI khi còn thiếu tấm nào.** Bản trước hỏi mỗi lần chạy, kể cả khi
    ba tấm đã nằm sẵn trên đĩa — chạy tiếp một lượt dở là trả tiền cho một lời
    nhắc không ai dùng.
    """
    thu_muc = os.path.join(luot.thu_muc, "7-thumbnail")
    os.makedirs(thu_muc, exist_ok=True)
    kieu = list(KIEU_THUMB[:max(1, bc.kenh.so_thumbnail)])
    muc = list(enumerate(kieu, start=1))
    thieu = [m for m in muc if not os.path.exists(_tep_bia(thu_muc, m[0]))]
    tieu_de, chu_bia = _doc_tieu_de(
        _doc_chu(os.path.join(luot.thu_muc, "1-tieu-de.txt")))
    ta_bia: Dict[str, str] = {}
    if thieu:
        ta_bia = _loi_nhac_bia(bc, luot, bc.kenh.prompt.get("8-thumbnail.md", ""),
                               tieu_de, chu_bia, kieu)
    return thu_muc, muc, thieu, ta_bia, tieu_de, chu_bia


def _lam_bia(bc: BoiCanh, luot: LuotChay, hop: "ThamChieu", thu_muc: str,
             muc, ta_bia: Dict[str, str], tieu_de: str, chu_bia: str,
             so: Optional[SoTheoDoi] = None):
    """Tạo một tấm ảnh bìa. Đã có trên đĩa thì bỏ qua."""
    so_bia, (ten_kieu, mac_dinh) = muc
    tep = _tep_bia(thu_muc, so_bia)
    if os.path.exists(tep):
        return so_bia, True
    loi_nhac = _lay_ta_bia(ta_bia, ten_kieu, so_bia)
    if not loi_nhac:
        # Nói ra chứ đừng lặng lẽ: bản ghép cứng ra ảnh khác hẳn hai tấm kia,
        # và bản trước không ghi gì nên không ai biết vì sao tấm này lệch.
        bc.ghi("  (không có lời nhắc AI cho ảnh bìa {0} — dùng bản mặc "
               "định)".format(so_bia))
        loi_nhac = _bia_du_phong(bc.kenh.style, tieu_de, chu_bia, mac_dinh)
    # Khoá phủ cả ảnh tham chiếu — xem ghi chú ở `_lam_anh_canh`.
    goi = _tao_anh(bc, luot, loi_nhac, hop,
                   khoa_viec(luot, "thumb", so_bia, loi_nhac,
                             "|".join(hop.lay())),
                   ten_hien="ảnh bìa {0}".format(so_bia), so=so)
    _tai_ket_qua(bc, goi, 0, tep)
    _xoa_dau(bc, tep)
    return so_bia, False


def _khau_thumbnail(bc: BoiCanh):
    """Làm nốt ảnh bìa dây chuyền ở khâu ảnh chưa kịp ra.

    Ba tấm này đã được bắn cùng mẻ với 114 ảnh cảnh, nên bình thường khâu này
    chỉ nhìn đĩa rồi đi qua. Vẫn đứng đây cho "Làm lại khâu này" và cho lúc mẻ
    trước hụt mất một tấm.
    """

    def lam(luot: LuotChay, tt: TrangThaiKhau):
        thu_muc, muc, thieu, ta_bia, tieu_de, chu_bia = _chuan_bi_bia(bc, luot)
        # Đủ ba tấm rồi thì đừng dựng `ThamChieu`: nó có thể phải tải lại ảnh
        # nhân vật lên, tốn chỗ trong kho tạm 500 MB cho một việc không có.
        hop = _hop_bia(bc, luot, ThamChieu(bc)) if thieu else None
        so = SoTheoDoi(bc, nhip=bc.nhip_hoi) if thieu else None

        def mot_bia(m):
            return _lam_bia(bc, luot, hop, thu_muc, m, ta_bia, tieu_de,
                            chu_bia, so=so)

        try:
            xong = _chay_song_song(
                bc, muc, mot_bia, "ảnh bìa",
                nhip=dem_tien_do(bc, luot, tt, "ảnh bìa", NHIP_GHI_TIEN_DO),
                loai_job="image")
        finally:
            if so is not None:
                so.dong()
        return {"so_thumbnail": xong}

    return lam


# ── Khâu 8: dựng ─────────────────────────────────────────────────────────────


def _khau_dung(bc: BoiCanh):
    def lam(luot: LuotChay, tt: TrangThaiKhau):
        d = luot.thu_muc
        dich = os.path.join(d, "8-video.mp4")
        if os.path.exists(dich):
            moi = _nguon_moi_hon_video(dich, os.path.join(d, "6-clip"),
                                       os.path.join(d, "2-giong-doc.mp3"))
            if not moi:
                # Video còn tốt — nhưng phần đuôi CapCut (nếu kênh bật) có thể
                # chưa làm hoặc hỏng lần trước. Chạy nốt rồi mới nhận "đã có".
                tep_cc = _xuat_capcut_neu_bat(bc, d, dich)
                return {"da_co": True, **({"capcut": tep_cc} if tep_cc else {})}
            bc.ghi("  video đã có nhưng {0} mới hơn nó — dựng lại (bản cũ giữ ở "
                   "8-video.cu.mp4).".format(moi))
            try:
                os.replace(dich, os.path.join(d, "8-video.cu.mp4"))
            except OSError:
                pass
        ffmpeg = _bao_dam_ffmpeg(bc)
        canh = _doc_canh(luot)
        thu_muc_clip = os.path.join(d, "6-clip")
        _loai_clip_hong(bc, ffmpeg, thu_muc_clip, canh)
        # ═══ THIẾU VÀI CLIP THÌ VẪN DỰNG, CẢNH TRƯỚC GIỮ HÌNH BÙ VÀO ═══
        #
        # Trước đây thiếu một clip là không dựng, chấm hết. Đã xảy ra thật hai
        # lần liền (15/08/2026) với **đúng một** cảnh trong 112: khách trả tiền
        # cho 111 cảnh rồi nhận về không có gì xem được.
        #
        # Bỏ cảnh thiếu ra khỏi danh sách thì cảnh liền trước tự động chiếm chỗ
        # của nó — `giay[i]` tính bằng `srt_start` của cảnh kế **còn lại**, nên
        # hình đứng yên thêm vài giây rồi đi tiếp. Tiếng và phụ đề không xê
        # dịch một mi-li-giây nào, vì cả hai bám mốc thời gian tuyệt đối chứ
        # không bám thứ tự clip.
        con = [c for c in canh
               if os.path.exists(os.path.join(
                   thu_muc_clip, "{0}.mp4".format(int(c["scene_id"]))))]
        thieu = len(canh) - len(con)
        if not con:
            raise RuntimeError("chưa có clip nào, không dựng được")
        if thieu:
            bc.ghi("  thiếu {0}/{1} clip — dựng bằng {2} clip đang có, cảnh "
                   "trước sẽ giữ hình bù vào chỗ trống.".format(
                       thieu, len(canh), len(con)))
        canh = con
        manh = [os.path.join(thu_muc_clip, "{0}.mp4".format(int(c["scene_id"])))
                for c in canh]
        mp3 = os.path.join(d, "2-giong-doc.mp3")
        srt = os.path.join(d, "3-phu-de.srt")
        # ═══ CẮT MỖI CLIP VỀ ĐÚNG ĐỘ DÀI CẢNH CỦA NÓ ═══
        #
        # Engine bán clip **cố định 8 giây**, nhưng cảnh thì **chia theo nội
        # dung** — đo trên lượt thật: từ 2,8 tới 8,0 giây, phần lớn 5–7.
        #
        # Ghép thẳng 99 clip 8 giây là 792 giây hình cho 645 giây tiếng: hình
        # trôi khỏi tiếng **147 giây**. Tới cuối video thì lời đang nói chuyện
        # này mà hình chiếu chuyện khác — hỏng đúng thứ cả dây chuyền sinh ra
        # để làm: hình phải nói cùng điều lời đang nói.
        #
        # Chủ dự án nhắc đúng chỗ này, 14/08/2026: *"scenes ở excel nó chia
        # không cố định đâu, nó chia theo nội dung, để về sau khi edit nội dung
        # nói tới phần nào thì clip sẽ thể hiện điều đó"*.
        # ═══ MỖI CLIP PHẢI NẰM ĐÚNG MỐC `srt_start` CỦA NÓ ═══
        #
        # Bản trước lấy `duration` rồi nối tiếp nhau. Sai, vì giữa hai cảnh có
        # **khoảng hở** — chỗ người đọc ngừng lấy hơi. Đo trên lượt thật: 99
        # cảnh, tổng hở **61,9 giây**. Nối tiếp nhau là bỏ hết chỗ hở ấy, nên
        # càng về cuối hình càng chạy trước tiếng, tới cảnh cuối lệch hơn một
        # phút — lời đang nói chuyện này, hình đã sang chuyện khác.
        #
        # Chủ dự án nhìn ra ngay khi xem video đầu tiên, 14/08/2026: *"edit
        # đang không theo log của excel để khống chế việc clip chỉ xuất hiện
        # theo excel là có thời gian bắt đầu"*.
        #
        # Đúng phải là: cảnh `i` chiếm đúng khoảng từ `srt_start[i]` tới
        # `srt_start[i+1]`. Khoảng hở tự khắc được cảnh trước giữ hình — đúng
        # như người dựng tay vẫn làm: người đọc ngừng thì hình vẫn ở đó.
        moc = [_giay_srt(c.get("srt_start")) for c in canh]
        het = [_giay_srt(c.get("srt_end")) for c in canh]
        giay = []
        for i in range(len(canh)):
            if i + 1 < len(canh):
                giay.append(max(0.1, moc[i + 1] - moc[i]))
            else:
                giay.append(max(0.1, het[i] - moc[i]))
        _keo_canh_cuoi_cho_du_tieng(giay, mp3, bc.ghi)
        # Cách dựng lấy từ kênh, không hỏi lại từng lượt: mọi video của một
        # kênh dựng giống hệt nhau. Xem `core/kenh.Kenh.dot_phu_de`.
        dot = bool(getattr(bc.kenh, "dot_phu_de", True)) and os.path.exists(srt)
        giu_tieng = _giu_tieng_canh(bc)
        nhac = _duong_nhac(bc.kenh)
        ten_dpg = chon_do_phan_giai(bc.goc, bc.kenh)
        khung = KHUNG.get(ten_dpg)
        bc.ghi("  ghép {0} clip (cắt theo độ dài từng cảnh: {1:.0f} giây hình "
               "cho {2:.0f} giây tiếng){3}{4}{5}{6}…".format(
                   len(manh), sum(giay), sum(giay),
                   " + phụ đề" if dot else "",
                   " + tiếng cảnh" if giu_tieng else "",
                   " + nhạc nền" if nhac else "",
                   " + phóng lên {0}".format(ten_dpg) if khung else ""))
        if khung:
            # Nói thật ngay lúc chạy: phóng lên thì lâu hơn hẳn, và khách đang
            # ngồi nhìn dòng nhật ký này chứ không đọc tài liệu.
            bc.ghi("    (phóng {0}×{1} — khâu này lâu hơn giữ nguyên khoảng "
                   "bốn lần; phần nét thêm ra là máy đoán, không phải chi "
                   "tiết có thật. Đổi ở Cài đặt.)".format(khung[0], khung[1]))
        # Nói ngay từ đầu khâu này chạy bằng gì và sẽ báo tiến độ ra sao. Đây là
        # khâu duy nhất chạy trên máy khách, và cũng là khâu lâu nhất: khách báo
        # 28/08/2026 rằng tool "Not responding" chính vì ngồi nhìn một dòng nhật
        # ký đứng im mà không biết nó còn sống hay không.
        bc.ghi("    (dựng trên máy bạn, không tốn tiền: {0} luồng CPU, mức nén "
               "“{1}” chọn theo cấu hình máy này. Cứ khoảng {2:.0f} giây tôi "
               "báo một dòng phần trăm; cửa sổ vẫn bấm được, và bấm Dừng là "
               "dừng ngay.)".format(
                   so_van_ffmpeg(), _muc_nen(bc.goc), GIAY_BAO_TIEN_DO))
        _ghep_video(ffmpeg, manh, mp3, srt if dot else "", dich,
                    giay=giay, ghi=bc.ghi, nhac=nhac,
                    am_luong=float(getattr(bc.kenh, "am_luong_nhac", 0.12)),
                    khung=khung, base_dir=bc.goc,
                    loc_them=_so_nam_len_hinh(bc, canh, giay, khung),
                    dung=bc.kiem_dung, giu_tieng=giu_tieng,
                    am_luong_tieng=float(getattr(
                        bc.kenh, "am_luong_tieng_canh", AM_LUONG_TIENG_CANH)),
                    nguong_tieng_nguoi=float(getattr(
                        bc.kenh, "nguong_tieng_nguoi", 0.0) or 0.0))
        # Video dựng xong vốn đã sạch thẻ — FFmpeg mã hoá lại là thẻ của tệp
        # nguồn mất hết. Vẫn chạy một lượt cho chắc: nó chỉ chép luồng sang tệp
        # mới, mất vài giây cho cả video mười phút, và nó bảo hiểm cho ngày nào
        # đó bản FFmpeg mới giữ lại thẻ mà không ai để ý.
        try:
            if _bat_lam_sach(bc):
                from .lam_sach import lam_sach_video  # noqa: PLC0415

                lam_sach_video(ffmpeg, dich)
        except Exception:  # noqa: BLE001 — vệ sinh hỏng không được hỏng video
            pass
        tep_cc = _xuat_capcut_neu_bat(bc, d, dich)
        return {"so_clip": len(manh), "giay_hinh": round(sum(giay)),
                "phu_de_dot": dot, "tieng_canh": giu_tieng,
                "nhac": os.path.basename(nhac) if nhac else "",
                "do_phan_giai": ten_dpg,
                **({"capcut": tep_cc} if tep_cc else {})}

    def soi_lai(luot: LuotChay) -> bool:
        """Video cũ còn đúng không — hỏi TRƯỚC khi bỏ qua khâu dựng.

        ═══ VÌ SAO CẦN, ĐO NGÀY 26/08/2026 ═══

        Thân khâu này ĐÃ biết so ngày: clip nào mới hơn `8-video.mp4` thì dựng
        lại. Nhưng khâu đã đánh dấu "xong" thì `core.auto.chay` bỏ qua thẳng,
        nên đoạn so ngày ấy nằm im — y hệt chuyện đã xảy ra với bảng cảnh
        (xem `core.auto._con_dung_duoc`).

        Hậu quả đúng chỗ khách vừa trả tiền: sửa lời nhắc vài cảnh, tạo lại
        clip, bấm "Chạy tiếp" — khâu dựng bị bỏ qua, và khách mở `8-video.mp4`
        ra xem thì vẫn là **bản cũ**, đúng những cảnh vừa sửa vẫn nguyên si.
        Cùng lỗi ấy dính cả nút "Làm lại khâu này" trên dòng clip.

        Cửa này chỉ đọc ngày tệp, không gọi mạng, không tiêu một đồng nào.
        """
        d = luot.thu_muc
        dich = os.path.join(d, "8-video.mp4")
        if not os.path.exists(dich):
            return True     # chưa có video thì không phải việc của cửa này
        if _nguon_moi_hon_video(dich, os.path.join(d, "6-clip"),
                                os.path.join(d, "2-giong-doc.mp3")):
            return False
        # Kênh bật xuất qua CapCut thì bản CapCut cũng phải có và không cũ
        # hơn video — thiếu nó là khâu phải chạy lại (chỉ chạy phần đuôi).
        if getattr(bc.kenh, "xuat_capcut", False):
            cc = os.path.join(d, "9-video-capcut.mp4")
            if (not os.path.exists(cc)
                    or os.path.getmtime(cc) < os.path.getmtime(dich)):
                return False
        return True

    lam.soi_lai = soi_lai
    return lam


def _xuat_capcut_neu_bat(bc: BoiCanh, thu_muc: str, video: str) -> str:
    """Đuôi của khâu dựng: kênh bật `xuat_capcut` thì đưa video vào CapCut
    xuất lại thành `9-video-capcut.mp4`. Trả tên tệp, hoặc rỗng khi kênh tắt.

    Chạy trên máy, không tốn tiền. Hỏng thì ném — khâu dựng mang dấu HỎNG với
    câu chữ thật của `core.capcut`, còn `8-video.mp4` vẫn nguyên; bấm "Chạy
    tiếp" là chỉ làm lại đúng phần đuôi này (video có rồi thì khâu đi thẳng
    vào đây, xem nhánh `da_co`).
    """
    if not getattr(bc.kenh, "xuat_capcut", False):
        return ""
    dich = os.path.join(thu_muc, "9-video-capcut.mp4")
    if (os.path.exists(dich)
            and os.path.getmtime(dich) >= os.path.getmtime(video)):
        return "9-video-capcut.mp4"
    from .capcut import xuat_qua_capcut  # noqa: PLC0415

    bc.ghi("  đưa video vào CapCut để xuất lại (chạy trên máy, miễn phí — "
           "CapCut sẽ tự mở, tự bấm rồi tự đóng; đừng dùng chuột phím trong "
           "lúc nó bấm).")
    xuat_qua_capcut(video, dich, ghi=bc.ghi, dung=bc.kiem_dung)
    return "9-video-capcut.mp4"


def _muc_nen(goc: str) -> str:
    """Mức nén x264 máy này sẽ dùng cho bản cuối — chỉ để ghi vào nhật ký."""
    from .phan_cung import chon_encoder, doc_ket_qua  # noqa: PLC0415

    try:
        return chon_encoder(doc_ket_qua(goc), intermediate=False)[1].get(
            "-preset", "medium")
    except Exception:  # noqa: BLE001 — một dòng nhật ký không được làm hỏng khâu
        return "medium"


def chon_do_phan_giai(goc: str, kenh) -> str:
    """Độ phân giải video ra, gộp hai tầng cài đặt lại thành một câu trả lời.

    Thứ tự: **kênh khai gì thì theo kênh**, không khai thì theo cài đặt chung
    của tool, hỏng cả hai thì `"4K"`.

    Hai tầng chứ không một, vì hai câu hỏi khác nhau: *"nhà tôi làm video kiểu
    gì"* hỏi một lần ở Cài đặt, còn *"riêng kênh này khác"* mới hỏi ở kênh. Bắt
    khai lại cho từng kênh là bắt trả lời cùng một câu mười lần.

    Kênh khai một chữ tool không hiểu thì rơi về cài đặt chung chứ không lặng
    lẽ tắt tính năng — xem `core.kenh.ten_khung`.
    """
    from . import cai_dat  # noqa: PLC0415 — tránh vòng nhập lúc khởi động

    rieng = ten_khung(getattr(kenh, "do_phan_giai", ""))
    if rieng:
        return rieng
    return ten_khung(cai_dat.doc(goc).get("do_phan_giai")) or "4K"


#: Hình ngắn hơn tiếng bao nhiêu giây thì mới đáng kéo dài cảnh cuối.
#:
#: Dưới mức này là sai số làm tròn của mốc phụ đề (đo trên hai video thật:
#: lệch 0,027 và 0,026 giây) — kéo thêm chỉ tổ đẻ ra một khung hình thừa.
KE_HO_TIENG_BO_QUA = 0.35


def _keo_canh_cuoi_cho_du_tieng(giay: List[float], mp3: str,
                                ghi: Callable[[str], None]) -> None:
    """Hình ngắn hơn tiếng thì cho cảnh cuối giữ hình bù, đừng cắt lời đọc.

    ═══ `-shortest` CHỈ AN TOÀN MỘT CHIỀU ═══

    Lệnh ghép dùng `-shortest`, kèm ghi chú *"tổng clip thường dài hơn tiếng
    vài giây vì mỗi cảnh làm tròn lên"*. Giả định ấy đúng gần như mọi lúc — và
    `-shortest` khi ấy cắt phần ĐUÔI HÌNH thừa, hoàn toàn vô hại.

    Nhưng khi nó sai thì `-shortest` cắt phần kia: **lời đọc**. Đo trên lượt
    chạy thật R01 ngày 18/08/2026:

        tiếng 11,76 giây  ·  video 9,13 giây  →  mất 2,6 giây cuối

    Người xem nghe câu cuối đứt ngang giữa chừng. Trong khi hai video khác cùng
    ngày lệch có 0,03 giây, nên chuyện này không lộ ra nếu chỉ nhìn lướt.

    ═══ VÌ SAO KÉO CẢNH CUỐI, KHÔNG PHẢI BỎ `-shortest` ═══

    Bỏ `-shortest` thì gặp ca ngược lại — hình dài hơn tiếng — video sẽ có một
    đoạn đuôi câm. Giữ `-shortest` và kéo cảnh cuối cho đủ thì cả hai chiều đều
    đúng, và cách xử này giống hệt điều dây chuyền vẫn làm cho cảnh thiếu: giữ
    khung cuối lâu thêm vài giây, đúng như người dựng tay vẫn làm khi người đọc
    còn nói mà hình đã hết.

    Mất mấy giây cuối của lời đọc là hỏng nội dung. Giữ thêm một khung hình vài
    giây thì gần như không ai nhận ra.
    """
    from .phu_de import do_dai_tieng  # noqa: PLC0415 — cùng gói, dùng lại

    if not giay:
        return
    dai_tieng = do_dai_tieng(mp3)
    thieu = dai_tieng - sum(giay)
    if dai_tieng <= 0 or thieu <= KE_HO_TIENG_BO_QUA:
        return
    giay[-1] += thieu
    ghi("    (hình ngắn hơn tiếng {0:.1f} giây — giữ hình cảnh cuối thêm bấy "
        "nhiêu để không cắt mất câu cuối)".format(thieu))


def _so_nam_len_hinh(bc: BoiCanh, canh: Sequence[Dict[str, Any]],
                     giay: Sequence[float], khung: Optional[Sequence[int]]) -> str:
    """Số năm chạy ở góc phải dưới — chỉ kênh timelapse, và chỉ khi máy có phông.

    Phim không lời đọc thì đây là thứ DUY NHẤT nói cho người xem biết họ đang ở
    thế kỷ nào. Đo trên video đối thủ ngày 27/08/2026: số năm nhảy liên tục suốt
    22 phút, không lúc nào đứng im.
    """
    from .timelapse import la_timelapse, loc_so_nam  # noqa: PLC0415

    if not la_timelapse(bc.kenh) or not canh:
        return ""
    ra = loc_so_nam(canh, cao=int(khung[1]) if khung else 1080, giay=giay)
    if ra:
        bc.ghi("    + số năm chạy ở góc trái dưới ({0} → {1}).".format(
            canh[0].get("nam_tu"), canh[-1].get("nam_den")))
    else:
        bc.ghi("    (máy này không có phông chữ đậm nào — bỏ số năm trên hình.)")
    return ra


#: Độ to tiếng cảnh lúc KHÔNG có giọng đọc — mặc định, kênh đè được bằng
#: `am_luong_tieng_canh` trong `kenh.yaml`.
#:
#: ═══ VÌ SAO PHẢI NHÌN ĐỈNH, KHÔNG NHÌN TRUNG BÌNH ═══
#:
#: Bản đầu để 0,7 vì tiếng cảnh nghe "khá nhỏ": trung bình -31,3 dB so với
#: giọng đọc -14,7 dB. Chủ dự án nghe thì bảo nó lấn lời.
#:
#: Đo lại bằng ĐỈNH thì rõ ngay (phim `openstory/0008`, 28/08/2026):
#:
#:     giọng đọc    trung bình -14,7 dB   đỉnh -1,4 dB
#:     tiếng cảnh   trung bình -31,3 dB   đỉnh -1,6 dB   ← ngang giọng đọc
#:
#: Tiếng nền thì nhỏ thật, nhưng một tiếng nước bắn, một tiếng gỗ va là vọt
#: lên bằng lời kể. Trung bình không nói ra điều đó vì nó bị 8 giây im lặng
#: kéo xuống.
#:
#: 0,35 (nửa mức cũ, -9 dB) đưa đỉnh tiếng cảnh xuống -10,7 dB — thấp hơn đỉnh
#: giọng đọc 9,3 dB — và chừa chỗ cho nhạc nền khách tự chèn ở CapCut.
AM_LUONG_TIENG_CANH = 0.35


def _clip_co_tieng(ffmpeg: str, duong: str) -> bool:
    """Clip này có luồng tiếng không.

    Phải hỏi từng tệp: `concat` đòi mọi mảnh cùng bộ luồng, nên một clip câm
    lọt vào giữa là hỏng cả video. Clip câm thì khâu cắt lắp một đường im lặng
    vào cho đủ bộ (`anullsrc`).

    Chạy trên đĩa, không tốn tiền. Ghi chú cũ trong hàm này từng nói "Veo 3 qua
    đường này trả clip KHÔNG có tiếng (đo 27/08/2026)" — đo lại 28/08/2026 trên
    phim `openstory/0008` thì **có**: aac 48 kHz stereo, 140 kb/s. Nên đừng tin
    câu ấy nữa, hỏi tệp.
    """
    try:
        ra = subprocess.run([ffmpeg, "-hide_banner", "-i", duong],
                            capture_output=True, text=True)
    except OSError:
        return False
    return "Audio:" in ((ra.stderr or "") + (ra.stdout or ""))


def _ghep_video(ffmpeg: str, clip: Sequence[str], mp3: str, srt: str,
                dich: str, giay: Optional[Sequence[float]] = None,
                ghi: Optional[Callable[[str], None]] = None,
                nhac: str = "", am_luong: float = 0.12,
                khung: Optional[Sequence[int]] = None,
                base_dir: str = ".", loc_them: str = "",
                dung: Optional[Callable[[], None]] = None,
                giu_tieng: bool = False,
                am_luong_tieng: float = AM_LUONG_TIENG_CANH,
                nguong_tieng_nguoi: float = 0.0) -> None:
    """Cắt từng clip về đúng độ dài cảnh, nối lại, gắn tiếng, đốt phụ đề.

    `giay[i]` là độ dài **cảnh thứ i** lấy từ bảng cảnh — không phải độ dài
    clip. Engine bán clip cố định (Veo3 8 giây), còn cảnh chia theo nội dung,
    nên phải cắt thì hình mới bám đúng lời. Xem ghi chú ở `_khau_dung`.

    `srt` rỗng = không đốt phụ đề vào hình (kênh chỉ đăng YouTube thường muốn
    thế: tải tệp `.srt` lên riêng thì người xem bật/tắt được).

    `nhac` là tệp nhạc nền; rỗng = không có. `am_luong` là phần độ to còn lại
    của nhạc so với giọng đọc.

    `khung` là (rộng, cao) muốn xuất ra; `None` = giữ đúng cỡ nhà cung cấp trả
    về. Xem `Kenh.do_phan_giai` để biết vì sao cần và cái được thật là gì.

    `base_dir` là thư mục gốc của tool — cần để đọc kết quả khảo sát phần cứng.

    `loc_them` là chuỗi lọc FFmpeg cài thêm vào hình, chạy SAU khi phóng cỡ và
    TRƯỚC khi đốt phụ đề — kênh timelapse dùng nó để in số năm lên góc phải dưới
    (xem `core.timelapse.loc_so_nam`).

    `nguong_tieng_nguoi` là ngưỡng nhận ra tiếng người của kênh; 0 = dùng ngưỡng
    chung `tieng_canh.NGUONG_TIENG_NGUOI`.

    `am_luong_tieng` là độ to tiếng cảnh trong bản đã trộn — xem
    `AM_LUONG_TIENG_CANH`. Tệp `8-tieng-canh.m4a` xuất riêng KHÔNG bị hạ theo
    số này: khách chỉnh to nhỏ ở CapCut.

    `giu_tieng=True` giữ TIẾNG CẢNH của từng clip (bước chân, chim hót, nước,
    gió) nằm dưới giọng đọc, và xuất thêm `8-tieng-canh.m4a` cạnh `dich` — đường
    tiếng rời để mang sang CapCut. Xem `Kenh.giu_tieng_canh`.
    """
    from core.phan_cung import doc_ket_qua, chon_encoder

    thu_muc = os.path.dirname(dich) or "."
    tam = os.path.join(thu_muc, "_cat")
    os.makedirs(tam, exist_ok=True)

    # Lần cuối có phải mã lại hình không. Đốt phụ đề phải mã lại; đổi độ phân
    # giải cũng vậy — `-c:v copy` chỉ sao chép nguyên si, không phóng được.
    #
    # `loc_them` cũng phải nằm ở đây. Bỏ sót nó một lần rồi: kênh timelapse
    # không có phụ đề, không phóng cỡ, nên `ma_lai` thành False, FFmpeg chép
    # thẳng luồng và bộ lọc số năm bị vứt im lặng — nhật ký vẫn báo "+ số năm"
    # mà phim ra không có số nào (đo 27/08/2026, phim timelapse/0001).
    ma_lai = bool(srt) or bool(khung) or bool(loc_them)

    # ═══ NÉN HAI LẦN THÌ LẦN ĐẦU PHẢI GẦN NHƯ KHÔNG MẤT GÌ ═══
    #
    # Có đốt phụ đề thì hình đi qua **hai** vòng H.264: cắt từng clip ở đây,
    # rồi mã lại lần nữa lúc đốt chữ. H.264 là nén mất dữ liệu, nên lần hai nén
    # lên cái đã hỏng của lần một — hỏng chồng hỏng.
    #
    # Bản trước để `veryfast -crf 20` cho lần một. `veryfast` là mức nhanh thứ
    # nhì của x264: cùng một CRF nó cho ảnh xấu hơn hẳn `medium`. Mà đây lại là
    # **bản gốc cho lần mã hoá thứ hai** — hỏng từ đầu thì lần sau không cứu
    # được.
    #
    # KHÔNG mã lại thì ngược hẳn: bước sau dùng `-c:v copy`, nên bản cắt ở đây
    # **chính là video giao cho khách**. Để `crf 14` cho nó là giao một tệp to
    # gấp mấy lần cần thiết, tải lên YouTube lâu mà YouTube vẫn nén lại hết.
    # Nên hai đường phải khác nhau, không dùng chung một con số.
    #
    # Từ ngày 20/08/2026: chọn encoder theo khảo sát phần cứng. Bản trung gian
    # (intermediate=True) dùng GPU nếu có — nhanh gấp 4–8 lần. Bản master cuối
    # (intermediate=False) luôn dùng CPU để đảm bảo chất lượng.
    pc = doc_ket_qua(base_dir)
    codec_cuoi, opts_cuoi = chon_encoder(pc, intermediate=False)

    # Chọn encoder cho BƯỚC CẮT clip. Chỗ này quyết định theo `ma_lai`:
    #   - ma_lai=True: bản cắt là TRUNG GIAN (sẽ mã lại lần nữa) → GPU nếu có.
    #   - ma_lai=False: bản cắt CHÍNH LÀ video giao khách (bước sau `-c copy`)
    #     → phải dùng encoder bản cuối (CPU an toàn), không được dùng GPU.
    codec_cat, opts_cat = chon_encoder(pc, intermediate=ma_lai)

    # ═══ CLIP NÀO CÓ NGƯỜI NÓI THÌ TẮT TIẾNG CLIP ẤY ═══
    #
    # Chủ dự án 28/08/2026: *"tiếng nói chuyện — nó bị khác ngôn ngữ nên tao
    # muốn tận dụng âm thanh ngoài, còn chỗ nào có âm thanh nói chuyện thì
    # thôi"*. Lời nhắc đã cấm thoại (`LUAT_TIENG_CANH`) nhưng engine vẫn thoại,
    # nên phải soi lại bản đã về. Xem `core/tieng_canh.py`.
    cam_tieng = set()
    if giu_tieng:
        from .tieng_canh import clip_co_nguoi_noi  # noqa: PLC0415

        from .tieng_canh import NGUONG_TIENG_NGUOI  # noqa: PLC0415

        cam_tieng = clip_co_nguoi_noi(
            ffmpeg, clip, ghi=ghi,
            nguong=float(nguong_tieng_nguoi) or NGUONG_TIENG_NGUOI)

    da_cat = []
    for i, m in enumerate(clip):
        if giay is None:
            da_cat.append(m)
            continue
        ra = os.path.join(tam, "{0:04d}.mp4".format(i))
        if not os.path.exists(ra):
            can = float(giay[i])
            # ═══ CẢNH DÀI HƠN CLIP THÌ GIỮ KHUNG CUỐI ═══
            #
            # Engine bán clip cố định 8 giây, nhưng khoảng một cảnh phải chiếm
            # (tính cả chỗ người đọc ngừng lấy hơi) có khi hơn 8 giây. Thiếu
            # hình thì FFmpeg chèn đen — một nháy đen giữa video là lỗi ai cũng
            # thấy. `tpad=clone` giữ nguyên khung cuối cho tới hết khoảng, đúng
            # như người dựng tay để hình đứng yên trong lúc người đọc ngừng.
            loc = "tpad=stop_mode=clone:stop_duration={0:.3f}".format(
                max(0.0, can))
            # `-t` sau `-i` = cắt theo thời gian PHÁT. Phải mã lại chứ không
            # `-c copy` được: copy chỉ cắt được ở khung khoá, lệch tới cả giây,
            # và 99 lần lệch cộng dồn là hình lại trôi khỏi tiếng.
            lenh_cat = ["-y", "-hide_banner", "-nostats", "-i", m]
            # ═══ GIỮ TIẾNG THÌ MẢNH NÀO CŨNG PHẢI CÓ ĐƯỜNG TIẾNG ═══
            #
            # `concat` chép thẳng luồng, nên nó đòi mọi mảnh cùng một bộ luồng.
            # Một clip câm lọt vào giữa 30 mảnh có tiếng là hỏng cả video. Clip
            # câm thì lắp một đường im lặng vào cho đủ bộ.
            #
            # `apad` cho tiếng: cảnh dài hơn clip thì `tpad` đã giữ khung cuối
            # cho hình, còn tiếng mà không đệm thì FFmpeg cắt ngắn đường tiếng
            # và `concat` lệch dần — hình một đằng tiếng một nẻo.
            if giu_tieng and i not in cam_tieng and _clip_co_tieng(ffmpeg, m):
                lenh_cat.extend(["-vf", loc, "-af", "apad"])
            elif giu_tieng:
                lenh_cat.extend(["-f", "lavfi", "-i",
                                 "anullsrc=channel_layout=stereo:sample_rate=48000",
                                 "-vf", loc, "-map", "0:v:0", "-map", "1:a:0"])
            else:
                lenh_cat.extend(["-vf", loc])
            lenh_cat.extend(["-t", "{0:.3f}".format(can), "-c:v", codec_cat])
            for k, v in opts_cat.items():
                lenh_cat.extend([k, str(v)])
            lenh_cat.extend(["-threads", str(so_van_ffmpeg()),
                             "-pix_fmt", "yuv420p"])
            lenh_cat.extend(["-c:a", "aac", "-b:a", "160k", "-ar", "48000",
                             "-ac", "2"] if giu_tieng else ["-an"])
            lenh_cat.append(ra)
            _chay(ffmpeg, lenh_cat)
        da_cat.append(ra)
        if dung is not None:
            dung()      # bấm Dừng giữa 99 lần cắt thì dừng ngay tại đây
        if ghi is not None and (i + 1) % 20 == 0:
            ghi("    cắt {0}/{1} clip…".format(i + 1, len(clip)))

    danh_sach = os.path.join(thu_muc, "_clip.txt")
    with open(danh_sach, "w", encoding="utf-8") as tep:
        for m in da_cat:
            tep.write("file '{0}'\n".format(os.path.abspath(m).replace("'", "'\\''")))
    tam_noi = os.path.join(thu_muc, "_noi.mp4")
    _chay(ffmpeg, ["-y", "-hide_banner", "-nostats", "-f", "concat", "-safe",
                   "0", "-i", danh_sach, "-c", "copy", tam_noi])

    # ═══ ĐƯỜNG TIẾNG RỜI CHO CAPCUT ═══
    #
    # Chủ dự án 28/08/2026: *"video sau khi xong tao còn cho vào capcut để edit
    # lại sẽ chèn nhạc, làm phụ đề"*. Ở đó cần tiếng cảnh thành MỘT TỆP riêng —
    # bản đã trộn sẵn với giọng đọc thì không tách ra được nữa.
    #
    # Chép luồng (`-c:a copy`), không mã lại: vài giây cho cả phim, và không
    # nén chồng lên cái đã nén.
    if giu_tieng:
        rieng = os.path.join(thu_muc, "8-tieng-canh.m4a")
        try:
            _chay(ffmpeg, ["-y", "-hide_banner", "-nostats", "-i", tam_noi,
                           "-vn", "-c:a", "copy", rieng])
            if ghi is not None:
                ghi("    tiếng cảnh để riêng ở 8-tieng-canh.m4a (mang sang "
                    "CapCut trộn tay).")
        except Exception:  # noqa: BLE001 — tách hỏng không được hỏng video
            if ghi is not None:
                ghi("    (không tách được đường tiếng riêng — video vẫn dựng "
                    "bình thường.)")

    co_tieng = os.path.exists(mp3)
    # Tiếng cảnh chỉ tính là CÓ khi vừa bật cờ vừa thật sự có luồng tiếng trong
    # bản nối — cờ bật mà mọi clip đều câm thì đừng dựng chuỗi lọc rỗng.
    co_tieng_canh = bool(giu_tieng) and _clip_co_tieng(ffmpeg, tam_noi)
    # Nhạc chạy được KHÔNG CẦN giọng đọc: kênh timelapse không có lời đọc, cả
    # đường tiếng của nó chỉ là nhạc nền (xem `core/timelapse.py`).
    co_nhac = bool(nhac) and os.path.exists(nhac)

    lenh = ["-y", "-hide_banner", "-nostats", "-i", tam_noi]
    if co_tieng:
        lenh += ["-i", mp3]
    if co_nhac:
        # `-stream_loop -1`: bài nhạc thường ngắn hơn video nhiều, cho nó lặp
        # tới hết. `-shortest` ở dưới chốt điểm dừng nên lặp vô hạn không sao.
        lenh += ["-stream_loop", "-1", "-i", nhac]

    if ma_lai:
        # ═══ PHÓNG TRƯỚC, ĐỐT CHỮ SAU — THỨ TỰ NÀY KHÔNG ĐƯỢC ĐẢO ═══
        #
        # Đốt phụ đề ở 720p rồi mới phóng lên 4K là phóng luôn cả chữ: nét chữ
        # bị kéo giãn, viền răng cưa, nhìn ra ngay. Phóng hình trước rồi mới vẽ
        # chữ thì chữ được vẽ thẳng ở cỡ 4K — sắc nét đúng bằng cỡ xuất ra.
        #
        # `flags=lanczos`: mặc định của FFmpeg là `bicubic`, mềm. Đây là chỗ
        # phóng gấp ba (1280 → 3840) nên chọn phép nào thấy rõ nhất.
        #
        # Không thêm `unsharp` ở đây, dù làm nét sau khi phóng thì đẹp hơn trên
        # máy. Lý do: làm nét sinh viền sáng quanh mép, và bộ mã hoá của
        # YouTube khuếch đại đúng loại viền đó thành vệt bẩn. Nét vừa phải
        # trước khi tải lên cho kết quả đẹp hơn nét gắt.
        buoc = []
        if khung:
            buoc.append(
                "scale={0}:{1}:force_original_aspect_ratio=decrease:"
                "flags=lanczos,pad={0}:{1}:(ow-iw)/2:(oh-ih)/2,setsar=1".format(
                    int(khung[0]), int(khung[1])))
        if loc_them:
            buoc.append(loc_them)
        if srt:
            buoc.append("subtitles='{0}'".format(
                os.path.abspath(srt).replace("\\", "/").replace(":", "\\:")))
        # Lần nén cuối: `slow -crf 18` thay cho `medium -crf 20`. Đây là bản
        # giao cho YouTube, mà việc của mình là đưa cho nó **bản gốc sạch** —
        # YouTube mã hoá lại hết, nên nén tiếc ở đây chỉ tổ mất nét hai lần.
        # Từ ngày 20/08/2026: luôn dùng CPU cho bản cuối (an toàn), nhưng vẫn
        # đọc từ kết quả khảo sát để nhất quán.
        lenh += ["-vf", ",".join(buoc), "-c:v", codec_cuoi]
        for k, v in opts_cuoi.items():
            if k not in ("-preset", "-crf"):
                lenh.extend([k, str(v)])
        lenh.extend(["-preset", opts_cuoi.get("-preset", "slow")])
        lenh.extend(["-crf", opts_cuoi.get("-crf", "18")])
        lenh.extend(["-threads", str(so_van_ffmpeg())])
        lenh.extend(["-pix_fmt", "yuv420p"])
    else:
        lenh += ["-c:v", "copy"]

    if co_tieng_canh and co_tieng:
        # ═══ TIẾNG CẢNH NÉ GIỌNG ĐỌC, ĐÚNG CÁCH NHẠC NỀN VẪN NÉ ═══
        #
        # Dùng lại `loc_tron_nhac`: tiếng cảnh vào chỗ "nhạc", nên nó tự hạ khi
        # người đọc nói và tự lên lại lúc ngừng. Khác nhạc ở đúng một con số —
        # `AM_LUONG_TIENG_CANH` cao hơn, vì tiếng bước chân là thứ đang xảy ra
        # trong hình chứ không phải nền để lấp khoảng lặng.
        #
        # Có cả nhạc nền thì trộn nhạc vào tiếng cảnh TRƯỚC, rồi cho cả cụm ấy
        # né giọng một lần — né hai lần chồng nhau là tiếng nhấp nhô.
        muc_tieng = max(0.0, min(1.0, float(am_luong_tieng)))
        if co_nhac:
            nen = ("[0:a]volume=1.0[sc];[2:a]volume={0:.3f}[ms];"
                   "[sc][ms]amix=inputs=2:duration=first:"
                   "dropout_transition=0[nen];".format(
                       max(0.0, min(1.0, float(am_luong) / max(1e-6, muc_tieng)))))
            nhan_nen = "nen"
        else:
            nen = ""
            nhan_nen = "0:a"
        tron = nen + loc_tron_nhac("1:a", nhan_nen, "ra",
                                   am_luong_ne=muc_tieng,
                                   am_luong_deu=muc_tieng,
                                   ne_giong=co_ne_giong(ffmpeg))
        lenh += ["-filter_complex", tron, "-map", "0:v:0", "-map", "[ra]",
                 "-c:a", "aac", "-b:a", "192k", "-shortest"]
    elif co_tieng_canh and co_nhac:
        # ═══ TIẾNG CẢNH + NHẠC, KHÔNG CÓ GIỌNG ĐỌC ═══
        #
        # Kênh timelapse rơi đúng vào đây: không lời đọc, mà chủ kênh vẫn thả
        # được một tệp nhạc vào `nhac_nen`. Không có giọng thì không có gì để
        # né — trộn thẳng, nhạc hạ theo `am_luong_nhac` của kênh.
        #
        # Nhánh này viết ra vì thiếu nó là **nuốt mất nhạc** một cách lặng lẽ:
        # nhánh "chỉ tiếng cảnh" ở dưới map đúng `0:a` và bỏ hẳn đầu vào nhạc.
        # Hôm nay `nhac_nen` của kênh ấy còn rỗng nên chưa ai thấy; ngày chủ
        # kênh thả tệp nhạc vào thì mới vỡ, và vỡ không một dòng báo.
        tron = ("[0:a]volume=1.0[sc];[1:a]volume={0:.3f}[ms];"
                "[sc][ms]amix=inputs=2:duration=first:"
                "dropout_transition=0[ra]".format(
                    max(0.0, min(1.0, float(am_luong)))))
        lenh += ["-filter_complex", tron, "-map", "0:v:0", "-map", "[ra]",
                 "-c:a", "aac", "-b:a", "192k", "-shortest"]
    elif co_tieng_canh:
        # Không có giọng đọc, không nhạc: tiếng cảnh là toàn bộ đường tiếng,
        # để nguyên độ to.
        lenh += ["-map", "0:v:0", "-map", "0:a:0", "-c:a", "aac", "-b:a", "192k"]
    elif co_nhac and co_tieng:
        # Nhạc tự lùi khi có giọng, tự lên lại khi giọng ngừng. Cả lời giải
        # thích lẫn số đo nằm ở `core/tron_tieng.py` — cùng một chuỗi lọc với
        # tab Dựng video thủ công, để hai tab ra tiếng giống nhau.
        tron = loc_tron_nhac("1:a", "2:a", "ra", am_luong_deu=am_luong,
                             ne_giong=co_ne_giong(ffmpeg))
        lenh += ["-filter_complex", tron, "-map", "0:v:0", "-map", "[ra]",
                 "-c:a", "aac", "-b:a", "192k", "-shortest"]
    elif co_nhac:
        # KHÔNG có giọng đọc: nhạc là toàn bộ đường tiếng, và nó nằm ở đầu vào
        # SỐ 1 chứ không phải số 2 (kênh timelapse — xem `core/timelapse.py`).
        # Nhạc đã `-stream_loop -1` nên `-shortest` chốt đúng độ dài hình.
        lenh += ["-map", "0:v:0", "-map", "1:a:0", "-c:a", "aac", "-b:a",
                 "192k", "-shortest"]
    elif co_tieng:
        # `-shortest` để video kết thúc cùng lúc với giọng đọc: tổng clip
        # thường dài hơn tiếng vài giây vì mỗi cảnh làm tròn lên.
        lenh += ["-map", "0:v:0", "-map", "1:a:0", "-c:a", "aac", "-b:a",
                 "192k", "-shortest"]
    else:
        # Không giọng, không nhạc: nói rõ ra là chỉ lấy hình. Veo 3 qua đường
        # này trả clip KHÔNG có tiếng (đo 27/08/2026), nên để FFmpeg tự chọn
        # luồng thì cũng ra tệp câm — nhưng tự chọn là chỗ dễ đổi ngầm sau này.
        lenh += ["-map", "0:v:0", "-an"]
    # `+faststart` đẩy bảng mục lục của tệp lên đầu. Không có nó thì trình phát
    # phải tải hết tệp mới bắt đầu phát được — xem lại bản dựng trên máy là
    # phải chờ. Tab Dựng video thủ công vốn đã có, đường Tự động thì chưa.
    lenh += ["-movflags", "+faststart", dich]
    # Vòng nén cuối là chỗ ngốn hầu hết thời gian của cả khâu — có video mười
    # phút mất hàng giờ. Đưa `ghi` xuống để nó báo phần trăm, và `dung` để nút
    # Dừng cắt được ngang chừng; không có hai thứ ấy thì nhật ký im suốt và
    # khách chỉ còn cách nhìn một cửa sổ trắng mà đoán.
    _chay(ffmpeg, lenh, ghi=ghi, viec="ghép video",
          tong_giay=float(sum(giay)) if giay else 0.0, dung=dung)
    import shutil as _sh  # noqa: PLC0415
    for tep in (danh_sach, tam_noi):
        try:
            os.remove(tep)
        except OSError:
            pass
    _sh.rmtree(tam, ignore_errors=True)


_DAU_LOI_FFMPEG = ("error", "invalid", "failed", "no such", "permission", "denied",
                   "cannot", "could not", "not found", "unable", "conversion failed")


def _loi_ffmpeg(stderr: str) -> str:
    """Rút những dòng LỖI thật trong stderr của FFmpeg.

    Trước 25/08/2026 tool báo 400 ký tự cuối — với libx264 đó là bảng thống kê
    ("i8c dc,h,v,p: 35% 23%…"), còn dòng lỗi thật nằm phía trên và bị cắt mất.
    Hai lần ghép hỏng tối 25/08 không đọc được lý do vì thế.
    """
    dong = [d.strip() for d in (stderr or "").splitlines() if d.strip()]
    loi = [d for d in dong if any(k in d.lower() for k in _DAU_LOI_FFMPEG)
           and not d.startswith("[libx264")]
    chon = loi[-4:] if loi else dong[-3:]
    return " | ".join(chon)[:400]


#: Chừa lại bao nhiêu lõi CPU cho cửa sổ tool khi FFmpeg đang chạy.
LOI_DE_DANH = 1


def so_van_ffmpeg() -> int:
    """Cho FFmpeg mấy luồng — luôn chừa ít nhất một lõi cho giao diện."""
    return max(1, (os.cpu_count() or 2) - LOI_DE_DANH)


def _co_tao_ffmpeg() -> int:
    """Cờ tạo tiến trình FFmpeg: ẩn cửa sổ đen **và** hạ mức ưu tiên.

    ═══ VÌ SAO PHẢI HẠ ƯU TIÊN — KHÁCH BÁO 28/08/2026 ═══

    Khâu dựng là khâu DUY NHẤT chạy hẳn trên máy khách, và nó ăn sạch CPU: mã
    hoá x264 `-preset slow` ở 4K chiếm mọi lõi trong hàng giờ. Luồng vẽ của Qt
    khi ấy tranh không nổi một lát CPU nào; quá 5 giây không trả lời tin nhắn
    cửa sổ là Windows dán ngay chữ **"Not responding"** lên thanh tiêu đề và
    phủ trắng cả trang — trông y hệt tool đã treo, trong khi nó đang chạy đúng.

    Hai việc chữa đúng chỗ ấy, cả hai đều không đổi một khung hình nào của
    video ra:

    * `BELOW_NORMAL_PRIORITY_CLASS` — Windows luôn nhường lát CPU cho luồng vẽ
      trước, FFmpeg nhận phần còn lại. Cửa sổ vẫn bấm được suốt lúc dựng.
    * `so_van_ffmpeg()` (`-threads`) — chừa hẳn một lõi, phòng khi máy chỉ có
      2–4 lõi thì riêng mức ưu tiên chưa đủ.

    Mất thêm khoảng một phần mười thời gian dựng. Đổi lại khách không còn nhìn
    một cửa sổ chết trong hai tiếng rồi tắt máy giữa chừng.
    """
    if os.name != "nt":
        return 0
    return (getattr(subprocess, "CREATE_NO_WINDOW", 0)
            | getattr(subprocess, "BELOW_NORMAL_PRIORITY_CLASS", 0x00004000))


def _giay_dong_ho(chu: str) -> float:
    """`00:01:23.456` → số giây. Đọc không được thì `0.0`."""
    try:
        gio, phut, giay = str(chu).strip().split(":")
        return int(gio) * 3600 + int(phut) * 60 + float(giay)
    except (ValueError, AttributeError):
        return 0.0


#: Mấy giây mới báo tiến độ ghép một lần.
GIAY_BAO_TIEN_DO = 20.0


def _chay(ffmpeg: str, tham_so: Sequence[str], *,
          ghi: Optional[Callable[[str], None]] = None,
          tong_giay: float = 0.0, viec: str = "ghép",
          dung: Optional[Callable[[], None]] = None) -> None:
    """Chạy một lệnh FFmpeg, có báo tiến độ và bấm Dừng được.

    ═══ MỘT DÒNG NHẬT KÝ IM SUỐT HAI TIẾNG CŨNG LÀ MỘT LỖI ═══

    Bản trước gọi `subprocess.run` rồi ngồi đợi. Với video mười phút phóng lên
    4K, giữa dòng "ghép 99 clip…" và dòng kế tiếp là **hàng giờ không có chữ
    nào** — không cách nào phân biệt "đang mã hoá cảnh 40" với "tool treo".
    Cùng lúc ấy nút Dừng bấm cũng không nhả, vì `run` chỉ quay lại khi FFmpeg
    xong.

    Nay đọc `-progress` của chính FFmpeg: cứ `GIAY_BAO_TIEN_DO` giây báo một
    dòng phần trăm, và mỗi nhịp ấy cũng là một lần hỏi "có ai bấm Dừng không" —
    bấm Dừng thì giết FFmpeg ngay, không đợi hết.

    `tong_giay` là độ dài video sẽ ra, để đổi mốc thời gian thành phần trăm;
    bằng 0 thì chỉ báo số giây đã ghép được.
    """
    lenh = [ffmpeg]
    if ghi is not None:
        lenh += ["-progress", "pipe:1"]
    lenh += list(tham_so)
    ra = str(tham_so[-1]) if tham_so else ""

    def bo_tep_do() -> None:
        # Tệp ra viết dở phải bỏ: lần thử lại thấy "đã có" là dùng luôn tệp
        # cụt — video cuối lặng lẽ thiếu một đoạn.
        if ra and not ra.startswith("-") and os.path.isfile(ra):
            try:
                os.remove(ra)
            except OSError:
                pass

    tien = subprocess.Popen(  # noqa: S603
        lenh, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
        encoding="utf-8", errors="replace", creationflags=_co_tao_ffmpeg())

    # stderr phải có người đọc liên tục: FFmpeg in cảnh báo suốt lúc chạy, ống
    # dẫn đầy là nó đứng im giữa chừng và cả hai bên cùng đợi nhau mãi mãi.
    kho_loi: List[str] = []

    def hut_loi() -> None:
        try:
            kho_loi.append(tien.stderr.read() or "")
        except Exception:  # noqa: BLE001
            kho_loi.append("")

    luong = threading.Thread(target=hut_loi, daemon=True)
    luong.start()

    lan_bao = time.monotonic()
    try:
        for dong in tien.stdout or []:
            if dung is not None:
                dung()          # ném Cancelled nếu người dùng bấm Dừng
            if ghi is None or not dong.startswith("out_time="):
                continue
            xong = _giay_dong_ho(dong.split("=", 1)[1])
            bay_gio = time.monotonic()
            if xong <= 0 or bay_gio - lan_bao < GIAY_BAO_TIEN_DO:
                continue
            lan_bao = bay_gio
            if tong_giay > 0:
                ghi("    {0} {1:.0f}% ({2:.0f}/{3:.0f} giây)…".format(
                    viec, min(100.0, 100.0 * xong / tong_giay), xong,
                    tong_giay))
            else:
                ghi("    {0} được {1:.0f} giây…".format(viec, xong))
    except BaseException:
        tien.kill()
        tien.wait()
        bo_tep_do()
        raise
    finally:
        try:
            tien.stdout.close()
        except Exception:  # noqa: BLE001
            pass

    ma = tien.wait()
    luong.join(timeout=10)
    if ma != 0:
        bo_tep_do()
        raise RuntimeError("FFmpeg hỏng ({0}): {1}".format(
            os.path.basename(ra) if ra else "?",
            _loi_ffmpeg("".join(kho_loi))))


# ── Gom lại ──────────────────────────────────────────────────────────────────


def _tra_anh_that(bc: BoiCanh, d: str, noi: str, noi_vi: str, moc_dinh: str,
                  nam_dau: int, nam_cuoi: int, bai: Sequence[tuple],
                  mo_hinh: str) -> Optional[Dict[str, Any]]:
    """Tra ẢNH THẬT của chỗ ấy về đĩa, và chọn lấy một tấm làm ẢNH NHẬN DẠNG.

    Chạy ở khâu KỊCH BẢN, tức TRƯỚC khi dựng bảng mốc — thứ tự này quan trọng:
    có ảnh rồi thì mới bắt AI tả góc máy cho đúng tấm ảnh ấy được. Xem
    `timelapse.loi_nhac_bang_moc`.

    Kết quả ghi vào `4-anh-that.json`, tấm được chọn mang cờ `nhan_dang`. Khâu
    bảng cảnh và khâu vẽ ảnh đọc lại tệp ấy, không tra lần nữa.

    Hỏng ở bất cứ đâu cũng chỉ trả `None`: thiếu ảnh thật thì phim vẫn dựng
    được, chỉ là hình học chỉ dựa vào chữ và sẽ trôi.
    """
    from .timelapse import (  # noqa: PLC0415
        LOI_NHAC_TIM_ANH, SO_UNG_VIEN_NHIN, _tu_khoa,
        chon_anh_nhan_dang_bang_mat, chon_anh_that, gom_anh_that,
        tai_anh_that, ung_vien_nhan_dang,
    )

    tep = os.path.join(d, "4-anh-that.json")
    if os.path.exists(tep):
        try:
            with open(tep, encoding="utf-8") as f:
                kho = json.load(f)
            return next((x for x in kho if x.get("nhan_dang")), None)
        except (OSError, ValueError):
            pass

    bc.ghi("  tra ẢNH THẬT: hỏi thể loại trên Commons…")
    try:
        hoi = loc_json(bc.goi_chat(
            LOI_NHAC_TIM_ANH.format(noi=noi, moc_dinh=moc_dinh,
                                    nam_dau=nam_dau, nam_cuoi=nam_cuoi),
            mo_hinh=mo_hinh, toi_da_token=1500)) or {}
    except Exception as loi:  # noqa: BLE001
        bc.ghi("  (không hỏi được thể loại ảnh: {0})".format(str(loi)[:70]))
        hoi = {}
    tk = _tu_khoa(noi, moc_dinh, noi_vi)
    tho = gom_anh_that(hoi.get("the_loai"), bai, tk, ghi=bc.ghi)
    thu = os.path.join(d, THU_MUC_THAM_CHIEU_TL, "that")
    # ── Một lần tải, để tên tệp không trùng nhau ────────────────────────────
    #
    # `tai_anh_that` đánh số tệp theo thứ tự TRONG MỘT lời gọi. Bản trước gọi nó
    # hai lần (một cho kho đối chiếu, một cho tấm nhận dạng) nên cả hai cùng ghi
    # ra `that-01.jpg` — mở tệp ấy lên thì thấy tấm của lần gọi sau. Đo
    # 28/08/2026: `4-anh-that.json` có hai mục cùng trỏ tới `that-01.jpg`.
    #
    # Hai việc ngược nhau nên vẫn hai phép CHỌN, nhưng chỉ một lần TẢI: ứng viên
    # nhận dạng đứng đầu danh sách, kho đối chiếu theo sau.
    ung = ung_vien_nhan_dang(tho)
    doi = chon_anh_that(tho, None, 12)
    ten_ung = {str(x.get("ten") or "") for x in ung[:SO_UNG_VIEN_NHIN]}
    ds = ung[:SO_UNG_VIEN_NHIN] + [x for x in doi
                                   if str(x.get("ten") or "") not in ten_ung]
    kho = tai_anh_that(ds, thu, ghi=bc.ghi)
    # ── Chọn bằng MẮT, không bằng tên tệp ───────────────────────────────────
    #
    # Tên tệp nói dối. Đo 28/08/2026: phép chọn theo tên lấy ra "Incendie de
    # Notre-Dame-de-Paris 15 avril 2019" — đúng chỗ, đúng tên riêng, giấy phép
    # đẹp, và là ảnh nhà thờ ĐANG CHÁY.
    def nhin(noi_dung):
        from .goi_van_ban import goi_van_ban  # noqa: PLC0415

        return goi_van_ban(bc.client, [{"role": "user", "content": noi_dung}],
                           mo_hinh=mo_hinh, toi_da_token=300)

    nd = chon_anh_nhan_dang_bang_mat(
        nhin, [x for x in kho if str(x.get("ten") or "") in ten_ung],
        noi, moc_dinh, ghi=bc.ghi)
    for x in kho:
        x["nhan_dang"] = bool(nd is not None and x.get("ten") == nd.get("ten"))
    with open(tep, "w", encoding="utf-8") as f:
        json.dump(kho, f, ensure_ascii=False, indent=1)
    bc.ghi("  {0} ảnh thật; ảnh nhận dạng: {1}".format(
        len(kho), (nd.get("ten") or "")[:60] if nd else "KHÔNG CÓ"))
    return nd


#: Thư mục tham chiếu của một lượt — trùng `dao_dien_auto.THU_MUC_THAM_CHIEU`,
#: để ở đây cho `_tra_anh_that` khỏi phải nạp vòng.
THU_MUC_THAM_CHIEU_TL = "tham-chieu"


def _viet_goi_dang(bc: BoiCanh, d: str, bang: Dict[str, Any], phut: float,
                   mo_hinh: str) -> None:
    """Gói đăng YouTube cho kênh timelapse: `1-seo.txt` + `1-tieu-de.txt`.

    ═══ VÌ SAO KÊNH NÀY PHẢI CÓ RIÊNG ═══

    Khâu SEO chung (`_khau_kich_ban`, tệp `6-seo.md`) viết mô tả **từ kịch bản
    lời đọc** — mà phim này không có lời đọc, chỉ có bảng mốc thời gian. Kênh
    timelapse lại đi khâu kịch bản riêng, nên nó rơi qua cả hai cửa và trước bản
    này **không sinh ra tiêu đề, mô tả, thẻ hay mốc chương nào**.

    `1-tieu-de.txt` cũng là tệp khâu ẢNH BÌA đọc (`_doc_tieu_de`, hai dòng
    `TITLE:` / `THUMB:`). Thiếu nó thì bìa vẽ ra không có chữ — mà chữ số năm to
    đùng chính là thứ khiến người ta bấm vào ở thể loại này.

    Chủ dự án 28/08/2026: *"làm all mọi thứ để ra sp có thể đăng youtube"*.

    Hỏng ở bất cứ đâu cũng chỉ ghi nhật ký: bảy khâu trước đã tiêu tiền, một bộ
    phim thiếu sẵn phần mô tả vẫn hơn hẳn một lượt chết ở khâu cuối.
    """
    from .timelapse import (  # noqa: PLC0415
        GIAY_MOT_MOC, TEP_SEO, goi_seo, loi_nhac_seo,
    )

    tep_seo = os.path.join(d, TEP_SEO)
    tep_ten = os.path.join(d, "1-tieu-de.txt")
    if os.path.exists(tep_seo) and os.path.exists(tep_ten):
        return
    if not (bang.get("moc") or []):
        return
    try:
        bc.kiem_dung()
        goi = loc_json(bc.goi_chat(loi_nhac_seo(bang, phut, GIAY_MOT_MOC),
                                   mo_hinh=mo_hinh, toi_da_token=4096))
        _ghi_chu(tep_seo, goi_seo(goi))
        d_ten = goi if isinstance(goi, dict) else {}
        tieu_de = str(d_ten.get("tieu_de_vi") or d_ten.get("tieu_de_en") or "").strip()
        chu_bia = str(d_ten.get("chu_bia") or "").strip()
        if tieu_de or chu_bia:
            _ghi_chu(tep_ten, "TITLE: {0}\nTHUMB: {1}\n".format(tieu_de, chu_bia))
        bc.ghi("  gói đăng YouTube: {0} + 1-tieu-de.txt".format(TEP_SEO))
    except Exception as loi:  # noqa: BLE001
        bc.ghi("  (bỏ qua gói đăng YouTube: {0})".format(str(loi)[:90]))


def _khau_kich_ban_timelapse(bc: BoiCanh):
    """TRA CỨU SỬ THẬT rồi mới dựng bảng mốc — phim này không có lời đọc.

    Ba bước, và bước đầu là bước quan trọng nhất:

      1. **Tra cứu.** Hỏi AI nên đọc những trang bách khoa nào, rồi TẢI đúng
         những trang ấy về (`0-tu-lieu.txt`).
      2. **Dựng bảng mốc** — chỉ được dùng năm có trong tư liệu vừa tải.
      3. **Soát lại** từng mốc so với tư liệu, mốc nào không có thì bỏ.

    Vì sao phải khổ thế. Bản đầu để AI dựng bảng mốc bằng trí nhớ; soi lại 60
    mốc ngày 27/08/2026 thì chỉ ~23 mốc là sự kiện có thật đúng năm, ~37 mốc còn
    lại là chuyện dựng mặc áo năm tháng ("1085 triều đại yên bình", "1137 lụt
    tới chân thành"). Người xem thể loại này đến vì tò mò lịch sử THẬT — nhận ra
    một mốc bịa là thôi tin cả phim.
    """
    def lam(luot: LuotChay, tt: TrangThaiKhau):
        from .timelapse import (  # noqa: PLC0415
            GIAY_MOT_MOC, LOI_NHAC_BU_NGUON, LOI_NHAC_SOAT_MOC, TEP_SEO,
            goi_seo, loi_nhac_seo,
            LOI_NHAC_TIM_NGUON, TEP_MOC, bai_da_doc, doc_bang_moc,
            TEP_DAN_Y, loi_nhac_bang_moc, nguon_bu, so_moc_cho_phut,
            soat_bang_moc,
            tai_tu_lieu_su,
        )

        d = luot.thu_muc
        tep = os.path.join(d, TEP_MOC)
        phut_kenh = float(getattr(bc.kenh, "phut_muc_tieu", 0) or 8)
        if os.path.exists(tep):
            with open(tep, encoding="utf-8") as f:
                bang = json.load(f)
            # Bảng mốc đã có nhưng gói đăng thì có thể chưa: lượt chạy dở trước
            # bản này, hoặc lần trước gói đăng hỏng. Bù ở đây để bấm "Chạy tiếp"
            # là xong, thay vì phải dựng lại cả bảng mốc.
            _viet_goi_dang(bc, d, bang, phut_kenh,
                           str(bc.kenh.mo_hinh or "claude-sonnet-5"))
            return {"da_co": True, "so_moc": len(bang.get("moc") or [])}
        vao = luot.dau_vao or {}
        chu_de = str(vao.get("tieu_de") or "").strip() or str(vao.get("link") or "").strip()
        if not chu_de:
            raise LoiNoiDung("chưa có chủ đề — nhập tiêu đề (ví dụ “Thăng Long "
                             "1000 năm nhìn từ một bến sông”) rồi chạy lại")
        mo_hinh = str(bc.kenh.mo_hinh or "claude-sonnet-5")

        # ── 1. TRA CỨU ──────────────────────────────────────────────────────
        #
        # HAI việc, KHÔNG gộp làm một điều kiện. Bản trước gộp: `0-tu-lieu.txt`
        # đã có thì bỏ qua cả khối — kể cả lời gọi sinh ra `0-nguon.json`.
        #
        # Đo 28/08/2026: tôi sửa `LOI_NHAC_TIM_NGUON` (bắt chọn đúng một chỗ
        # đứng được, thay vì "Paris, France"), rồi chạy lại lượt bằng
        # `LAM_LAI=kich-ban`. Lời nhắc mới **không hề được gọi**, vì tư liệu đã
        # tải xong từ lượt trước. Nhật ký nói thẳng ra hậu quả — *"ảnh nhận
        # dạng: KHÔNG CÓ — hình học sẽ trôi"* — và nó đã bắt đầu vẽ 56 tấm ảnh
        # mốc không có neo hình học nào.
        #
        # Tải tư liệu thì vẫn bỏ qua khi đã có (tải lại 800.000 chữ là vô ích),
        # nhưng `0-nguon.json` thiếu thì PHẢI hỏi lại: nó rẻ (một lời gọi chữ)
        # và nó quyết định máy đứng ở đâu cho cả bộ phim.
        tep_tl = os.path.join(d, "0-tu-lieu.txt")
        tep_nguon = os.path.join(d, "0-nguon.json")
        tu_lieu = _doc_chu(tep_tl)
        nguon = {}
        if not tu_lieu.strip() or not os.path.exists(tep_nguon):
            bc.ghi("  tra cứu sử: hỏi nên đọc những trang nào…")
            nguon = loc_json(bc.goi_chat(
                LOI_NHAC_TIM_NGUON.format(chu_de=chu_de), mo_hinh=mo_hinh,
                toi_da_token=2048)) or {}
            _ghi_chu(os.path.join(d, "0-ngon-ngu.txt"),
                     str(nguon.get("ngon_ngu") or "vi"))
            with open(tep_nguon, "w", encoding="utf-8") as f:
                json.dump(nguon, f, ensure_ascii=False, indent=1)
        if not tu_lieu.strip():
            bc.ghi("  tải tư liệu về:")
            tu_lieu = tai_tu_lieu_su(nguon, ghi=bc.ghi)
            if len(tu_lieu) < 4000:
                raise LoiNoiDung(
                    "không tra cứu được sử của nơi này (tải về {0} chữ). Kênh này "
                    "chỉ dựng phim từ sự kiện CÓ THẬT, nên tôi dừng ở đây thay vì "
                    "bịa ra mốc. Thử ghi tiêu đề rõ tên nơi chốn hơn, ví dụ "
                    "“Hoàng thành Thăng Long, 1010 đến nay”.".format(len(tu_lieu)))
            _ghi_chu(tep_tl, tu_lieu)
        bc.ghi("  tư liệu: {0} chữ.".format(len(tu_lieu)))
        # Mã ngôn ngữ của CHÍNH NƯỚC ấy — vòng tra bù phải đọc đúng bản ấy, không
        # mặc định tiếng Việt. Xem `timelapse.nguon_bu`.
        ngon_ngu_nguon = (_doc_chu(os.path.join(d, "0-ngon-ngu.txt")).strip()
                          or "vi")[:12]

        # ── 1b. TRA ẢNH THẬT — TRƯỚC KHI DỰNG BẢNG MỐC ──────────────────────
        #
        # Thứ tự này là chỗ tôi làm sai một vòng, và nó là lỗi thiết kế chứ
        # không phải lỗi lọc ảnh. Bản trước để AI tự nghĩ ra góc máy rồi mới đi
        # tìm ảnh, và hai thứ nói về hai chỗ khác nhau: AI viết "đứng trên đảo,
        # trong lòng phố, nhìn dọc con đường về phía nhà thờ", còn tấm ảnh tìm
        # được lại chụp Notre-Dame **từ bên kia sông Seine**. Tấm ảnh vì thế
        # chẳng neo được gì.
        #
        # Nay: tìm ảnh trước, rồi bắt AI tả lại đúng góc máy CỦA TẤM ẢNH ẤY.
        nguon_da_luu = {}
        try:
            with open(os.path.join(d, "0-nguon.json"), encoding="utf-8") as f:
                nguon_da_luu = json.load(f)
        except (OSError, ValueError):
            pass
        noi = str(nguon_da_luu.get("noi") or chu_de)
        noi_vi = str(nguon_da_luu.get("noi_vi") or chu_de)
        # Tra ảnh bằng TÊN NGẮN, không bằng câu tả dài: kho ảnh tra theo
        # tên riêng. Thiếu tên ngắn thì lùi về `noi` như cũ.
        ten_ngan = str(nguon_da_luu.get("ten_ngan") or "").strip() or noi
        bc.ghi("  chỗ máy đứng: {0}".format(ten_ngan))
        nhan_dang = _tra_anh_that(bc, d, ten_ngan, noi_vi, "",
                                  _so_nguyen(nguon_da_luu.get("nam_dau")) or 0,
                                  _so_nguyen(nguon_da_luu.get("nam_cuoi")) or 2025,
                                  bai_da_doc(tu_lieu), mo_hinh)

        # ── 2. DỰNG BẢNG MỐC TỪ TƯ LIỆU ─────────────────────────────────────
        phut = float(getattr(bc.kenh, "phut_muc_tieu", 0) or 8)
        so_moc = so_moc_cho_phut(phut, GIAY_MOT_MOC)
        bc.ghi("  dựng bảng mốc: xin khoảng {0} mốc cho ~{1:.0f} phút phim "
               "(sử thật có bao nhiêu thì lấy bấy nhiêu).".format(so_moc, phut))
        # DÀN Ý VIẾT TAY của chủ kênh, nếu có — xương sống của bảng mốc.
        dan_y = _doc_chu(os.path.join(d, TEP_DAN_Y)).strip()
        if dan_y:
            bc.ghi("  có dàn ý của bạn ({0} chữ) — dùng làm xương sống, tra cứu "
                   "để kiểm và bù cho đủ mốc.".format(len(dan_y)))
        bang = doc_bang_moc(loc_json(bc.goi_chat(
            loi_nhac_bang_moc(chu_de, so_moc, tu_lieu, nhan_dang, dan_y),
            mo_hinh=mo_hinh, toi_da_token=32000)))
        if len(bang.get("moc") or []) < 2:
            raise LoiNoiDung("AI không dựng được bảng mốc thời gian — thử lại hoặc "
                             "đổi tiêu đề cho rõ nơi chốn")
        if not bang.get("goc_may"):
            raise LoiNoiDung("bảng mốc thiếu mô tả GÓC MÁY — cả phim dựa vào nó")

        # ── 2b. TRA BÙ CHỖ TRỐNG ────────────────────────────────────────────
        #
        # Vòng tra đầu bám danh sách trang do AI tự nghĩ, và nó hay bỏ sót cả
        # một thời kỳ. Đo 27/08/2026 (lượt 0003): tám trang tải về không có
        # trang nào về chiến tranh Nguyên–Mông hay thời Trịnh–Nguyễn, nên bảng
        # mốc trống gần 200 năm. Nên: nhìn lại bảng vừa dựng, tìm quãng nào hở
        # trên năm mươi năm, rồi tra bù đúng quãng ấy.
        nam = [int(m["nam"]) for m in bang["moc"]]
        ho = [(nam[i], nam[i + 1]) for i in range(len(nam) - 1)
              if nam[i + 1] - nam[i] >= 50]
        if ho:
            ho.sort(key=lambda x: x[1] - x[0], reverse=True)   # quãng dài nhất trước
            bc.ghi("  bảng còn hở {0} quãng dài ({1}) — tra bù…".format(
                len(ho), ", ".join("{0}–{1} ({2} năm)".format(a, b, b - a)
                                   for a, b in ho[:4])))
            da_doc = sorted(set(l.split(":", 1)[-1].strip(" ═")
                                for l in tu_lieu.splitlines() if l.startswith("═══ ")))
            try:
                them = tai_tu_lieu_su(nguon_bu(loc_json(bc.goi_chat(
                    LOI_NHAC_BU_NGUON.format(
                        ngon_ngu=ngon_ngu_nguon,
                        nam_dau=nam[0], nam_cuoi=nam[-1],
                        da_doc="\n".join("  - " + x for x in da_doc),
                        da_co=", ".join(str(x) for x in nam),
                        lo_hong="\n".join(
                            "  {0} → {1}   ({2} năm không có mốc nào)".format(a, b, b - a)
                            for a, b in ho)),
                    mo_hinh=mo_hinh, toi_da_token=2048)), ngon_ngu_nguon),
                    ghi=bc.ghi, da_co=da_doc)
            except Exception as loi:  # noqa: BLE001 — tra bù hỏng thì đi tiếp
                bc.ghi("  (tra bù hỏng: {0}) — dùng bảng đang có.".format(str(loi)[:80]))
                them = ""
            if len(them) > 2000:
                tu_lieu = tu_lieu + "\n\n" + them
                _ghi_chu(tep_tl, tu_lieu)
                bc.ghi("  dựng lại bảng mốc với {0} chữ tư liệu…".format(len(tu_lieu)))
                bang2 = doc_bang_moc(loc_json(bc.goi_chat(
                    loi_nhac_bang_moc(chu_de, so_moc, tu_lieu, nhan_dang,
                                      dan_y),
                    mo_hinh=mo_hinh, toi_da_token=32000)))
                # Chỉ nhận bản mới nếu nó thật sự lấp được chỗ trống.
                if len(bang2.get("moc") or []) > len(bang["moc"]) and bang2.get("goc_may"):
                    bc.ghi("  {0} mốc → {1} mốc.".format(
                        len(bang["moc"]), len(bang2["moc"])))
                    bang = bang2

        # ── 3. SOÁT LẠI: mốc nào không có trong tư liệu thì bỏ ───────────────
        truoc = len(bang["moc"])
        bc.ghi("  soát {0} mốc lại với tư liệu…".format(truoc))
        gon = [{"nam": m["nam"], "su_that": m.get("su_that") or m.get("nhan")}
               for m in bang["moc"]]
        try:
            bang = soat_bang_moc(bang, loc_json(bc.goi_chat(
                LOI_NHAC_SOAT_MOC.format(
                    tu_lieu=tu_lieu[:200000],
                    bang=json.dumps(gon, ensure_ascii=False)),
                mo_hinh=mo_hinh, toi_da_token=16384)), ghi=bc.ghi)
        except Exception as loi:  # noqa: BLE001 — soát hỏng thì giữ bảng, nói thật
            bc.ghi("  (không soát lại được: {0}) — giữ nguyên bảng, "
                   "hãy tự đọc `1-kich-ban.txt` trước khi đăng.".format(str(loi)[:80]))
        if len(bang.get("moc") or []) < 2:
            raise LoiNoiDung("soát xong không còn mốc nào đứng vững — tư liệu quá "
                             "mỏng cho nơi này")
        bc.ghi("  giữ {0}/{1} mốc.".format(len(bang["moc"]), truoc))

        with open(tep, "w", encoding="utf-8") as f:
            json.dump(bang, f, ensure_ascii=False, indent=1)
        dong = ["{0}{1} — {2}".format(
            m.get("nhan") or m.get("nam"),
            "  ⟵ DỪNG LẠI" if int(m.get("tam") or 1) >= 2 else "",
            m.get("su_that") or m.get("canh")) for m in bang["moc"]]
        with open(os.path.join(d, "1-kich-ban.txt"), "w", encoding="utf-8") as f:
            f.write("NƠI: {0}\n\nMỐC NEO: {1}\n\nGÓC MÁY: {2}\n\n".format(
                bang.get("noi"), bang.get("moc_dinh"), bang.get("goc_may")))
            f.write("\n".join(dong) + "\n")
        bc.ghi("  {0} mốc, từ {1} tới {2}.".format(
            len(bang["moc"]), bang["moc"][0].get("nam"), bang["moc"][-1].get("nam")))
        _viet_goi_dang(bc, d, bang, phut, mo_hinh)

        return {"so_moc": len(bang["moc"]), "noi": bang.get("noi", "")}

    return lam


def _so_nguyen(x: Any) -> Optional[int]:
    """Đổi sang số nguyên, hỏng thì `None` — dùng cho năm tháng do AI khai."""
    try:
        return int(x)
    except (TypeError, ValueError):
        return None


def _khau_bang_canh_timelapse(bc: BoiCanh):
    """Bảng mốc → bảng cảnh, + TRA ẢNH THẬT, + một ảnh bối cảnh làm góc máy.

    Khâu tra ảnh thật là thứ chủ dự án gọi là "việc đầu tiên cần làm cho sản phẩm
    này thực sự có giá trị" (28/08/2026):

        *"cuối cùng thì là giai đoạn cuối nó phải giống thật… có thể những gì từ
        lâu không có ảnh nhưng sẽ có các tài liệu mô tả và có thể dựa vào dữ liệu
        để xây dựng phán đoán."*

    Người xem biết chỗ ấy hôm nay trông thế nào. Đoạn cuối phim không giống cái
    họ đã thấy tận mắt thì họ không tin cả nghìn năm phía trước, dù nghìn năm ấy
    dựng đúng sử. Nên ẢNH GÓC MÁY — tấm quyết định hình học của cả phim — được vẽ
    KÈM ảnh chụp thật của chỗ ấy ngày nay, chứ không vẽ từ chữ nghĩa của AI.
    """
    def lam(luot: LuotChay, tt: TrangThaiKhau):
        from .dao_dien_auto import THU_MUC_THAM_CHIEU, ThamChieuCanh  # noqa: PLC0415
        from .timelapse import (  # noqa: PLC0415
            GIAY_MOT_MOC, LOI_NHAC_TIM_ANH, TEP_MOC, _tu_khoa, anh_gan_nam,
            bai_da_doc, canh_tu_bang_moc, chon_anh_nhan_dang, chon_anh_that,
            gom_anh_that, prompt_anh_moc, tai_anh_that,
        )

        d = luot.thu_muc
        with open(os.path.join(d, TEP_MOC), encoding="utf-8") as f:
            bang = json.load(f)
        canh = canh_tu_bang_moc(bang, GIAY_MOT_MOC)
        if not canh:
            raise LoiNoiDung("bảng mốc chưa đủ hai mốc để dựng cảnh")
        tc = os.path.join(d, THU_MUC_THAM_CHIEU)
        os.makedirs(tc, exist_ok=True)

        # ── 1. ẢNH THẬT — đã tra ở khâu KỊCH BẢN, chỉ đọc lại ───────────────
        #
        # Tra ảnh phải chạy TRƯỚC khi dựng bảng mốc, vì bản mô tả góc máy phải
        # tả đúng góc máy của tấm ảnh tìm được. Xem `_tra_anh_that`.
        _tra_anh_that(bc, d, bang.get("ten_ngan") or bang.get("noi") or "",
                      bang.get("noi_vi") or "",
                      bang.get("ten_moc_dinh") or bang.get("moc_dinh") or "",
                      _so_nguyen(bang["moc"][0].get("nam")) or 0,
                      _so_nguyen(bang["moc"][-1].get("nam")) or 2025,
                      bai_da_doc(_doc_chu(os.path.join(d, "0-tu-lieu.txt")))[:8],
                      str(bc.kenh.mo_hinh or "claude-sonnet-5"))
        with open(os.path.join(d, "4-anh-that.json"), encoding="utf-8") as f:
            kho_anh = json.load(f)
        bc.ghi("  có {0} ảnh thật để đối chiếu.".format(len(kho_anh)))

        # ── 2. Ảnh GÓC MÁY: vẽ kèm ảnh thật ngày nay ────────────────────────
        dich = os.path.join(tc, "loc1.png")
        if not os.path.exists(dich):
            nay = next((x for x in kho_anh if x.get("nhan_dang")), None)
            bc.ghi("  vẽ khung hình mốc đầu — góc máy của cả phim{0}.".format(
                " (kèm ảnh thật {0})".format(nay.get("nam")) if nay else
                " (KHÔNG có ảnh thật — hình học chỉ dựa vào chữ)"))
            # `anh_nhan_dang`, KHÔNG phải `anh_that`: tấm này chụp ngày nay, còn
            # mốc đầu phim có khi cách nó hai nghìn năm. Đưa vào ô `anh_that` là
            # bảo máy "vẽ đúng như ảnh" — tức bê cả mái ngói hôm nay vào năm 866.
            prompt = prompt_anh_moc(bang, bang["moc"][0], dau_phim=True,
                                    anh_nhan_dang=nay)
            refs = [nay["tep"]] if nay and os.path.isfile(nay.get("tep") or "") else []
            _lam_anh_canh(bc, luot, dict(canh[0], img_prompt=prompt, scene_id=0),
                          dich, ThamChieuCanh(bc, refs))
        with open(os.path.join(d, "4-canh.json"), "w", encoding="utf-8") as f:
            json.dump(canh, f, ensure_ascii=False, indent=1)
        _viet_xlsx(os.path.join(d, "4-canh.xlsx"), canh, bc.kenh,
                   boi_canh=[{"id": "loc1", "name": bang.get("noi") or "",
                              "english_prompt": bang.get("goc_may") or ""}])
        bc.ghi("  {0} cảnh × {1:.0f} giây = {2:.0f} giây phim.".format(
            len(canh), GIAY_MOT_MOC, len(canh) * GIAY_MOT_MOC))
        return {"so_canh": len(canh), "giay": len(canh) * GIAY_MOT_MOC,
                "goc_may": (bang.get("goc_may") or "")[:80]}

    return lam


def _khung_cuoi_clip(bc: BoiCanh, clip: str, ra: str) -> str:
    """Tấm hình CUỐI CÙNG của một clip — để clip sau nối vào đúng chỗ ấy.

    `-sseof -0.4` tua từ ĐUÔI tệp lùi lại 0,4 giây (tua từ đầu thì phải giải mã
    cả clip), rồi `-update 1` ghi đè liên tiếp mọi khung còn lại lên cùng một
    tệp — nên thứ đọng lại là khung CUỐI CÙNG.

    Đừng đổi thành `-frames:v 1`: nó lấy khung ĐẦU TIÊN của quãng ấy, tức lùi
    trước khung cuối một quãng. Clip sau bắt đầu từ đó thì mỗi chỗ nối phim tua
    lại chừng ấy — 0,4 giây × 44 chỗ nối, và mắt đọc ra đúng là "khựng".
    """
    if os.path.exists(ra):
        return ra
    _chay(bc.ffmpeg or _tim_ffmpeg(),
          ["-y", "-hide_banner", "-nostats", "-sseof", "-0.4", "-i", clip,
           "-update", "1", "-q:v", "2", ra])
    return ra


def _soat_thoi_dai_anh_moc(bc: BoiCanh, luot: LuotChay, c: Dict[str, Any],
                           anh: str, nam: Any, noi: str, hop, so=None) -> int:
    """Soi tấm ảnh mốc vừa vẽ: có vật nào lạc thế kỷ không? Có thì vẽ lại.

    ═══ VÌ SAO KHÂU NÀY CÓ MẶT ═══

    Ngày 28/08/2026 chủ dự án mở phim 0005 ra và thấy **ô tô ở năm 500** — sau
    khi tôi đã soi 24 khung ngẫu nhiên trên 824 giây và báo với họ là phim sạch.
    Một chiếc xe con nằm mép khung, kéo dài 5% thời lượng, thì mẫu thưa nào cũng
    trượt. Mắt người không phải là cửa chặn cho một kênh mà tính đúng là toàn bộ
    giá trị: *"đây là sản phẩm lịch sử, những gì nó vẽ là phải giống, phải như
    sự thật"*.

    Nên soi giao cho máy, và soi TRƯỚC khi vẽ clip: một tấm ảnh mốc bẩn sẽ đẻ ra
    hai clip bẩn (clip GIỮ mở từ nó, clip TUA hạ vào nó), nên chặn ở đây rẻ gấp
    đôi chặn ở sau.

    Vẽ lại thì **nói tên đúng vật máy vừa vẽ nhầm** — cấm chung chung một lần đã
    không ăn thì cấm chung chung lần hai cũng thế. Vẫn còn thì ghi vào nhật ký
    chứ không im lặng cho qua: phim vẫn chạy tiếp, nhưng người xem lại được biết
    tấm nào đáng ngờ.
    """
    from .cham_anh import data_url  # noqa: PLC0415, F401
    from .goi_van_ban import goi_van_ban  # noqa: PLC0415
    from .timelapse import (  # noqa: PLC0415
        NGUONG_LAC_THOI, loi_nhac_ve_lai, soat_thoi_dai,
    )

    def goi(noi_dung):
        return goi_van_ban(bc.client, [{"role": "user", "content": noi_dung}],
                           mo_hinh=bc.kenh.mo_hinh, toi_da_token=300)

    # Bo cham nem loi thi coi nhu tam sach. Cua nay de BAT loi, khong phai
    # de chan ca day chuyen khi chinh no hong -- va con `khoa_the_ky` o loi
    # nhac do lung. Ban dau cua ham nay khong bao boc cho nay, va mot loi
    # cua bo cham la mat ca khau anh (bai kiem
    # `test_bo_cham_hong_thi_day_chuyen_van_chay`).
    try:
        lac, thay_bang = soat_thoi_dai(goi, anh, nam, noi)
    except Exception as loi:  # noqa: BLE001
        bc.ghi("    mốc {0}: cửa soát thời đại hỏng ({1}) — bỏ qua tấm "
               "này.".format(nam, str(loi)[:60]))
        return 0
    if len(lac) < NGUONG_LAC_THOI:
        return 0
    bc.ghi("    mốc {0}: ảnh có {1} vật lạc thế kỷ ({2}) — vẽ lại…".format(
        nam, len(lac), ", ".join(lac[:4])))
    cu = anh + ".lac"
    try:
        shutil.copyfile(anh, cu)
        os.remove(anh)
        c2 = dict(c)
        c2["img_prompt"] = loi_nhac_ve_lai(str(c.get("img_prompt") or ""),
                                           thay_bang, lac, nam)
        _lam_anh_canh(bc, luot, c2, anh, hop, so=so)
        try:
            lai, _ = soat_thoi_dai(goi, anh, nam, noi)
        except Exception:  # noqa: BLE001
            lai = []
        # Hoà thì giữ tấm VẼ LẠI, không giữ tấm đầu: tấm vẽ lại ít nhất đã
        # được vẽ với câu sửa, còn tấm đầu thì chưa ai nói gì với nó. Chỉ giữ
        # tấm đầu khi vẽ lại thật sự TỆ HƠN.
        if len(lai) >= NGUONG_LAC_THOI and len(lai) > len(lac):
            # Ve lai khong hon: giu tam dau, va NOI RA. Doi tam moi lay tam cu
            # chi de khoi mat cong -- ca hai deu ban nhu nhau.
            shutil.copyfile(cu, anh)
            bc.ghi("    mốc {0}: vẽ lại vẫn còn ({1}) — GIỮ tấm đầu, "
                   "đánh dấu để soi tay.".format(nam, ", ".join(lai[:3])))
            return 2
        bc.ghi("    mốc {0}: vẽ lại sạch.".format(nam))
        return 1
    except Exception as loi:  # noqa: BLE001
        # Ve lai hong thi tra tam cu ve cho: mot tam dang ngo van hon khong tam
        # nao -- thieu anh moc la ca khoi hai clip mat trang.
        if os.path.exists(cu) and not os.path.exists(anh):
            shutil.copyfile(cu, anh)
        bc.ghi("    mốc {0}: vẽ lại hỏng ({1}) — giữ tấm đầu.".format(
            nam, str(loi)[:60]))
        return 2


def _khau_anh_timelapse(bc: BoiCanh):
    """Ảnh + clip cho timelapse: ảnh mốc k là khung ĐẦU clip k và khung CUỐI clip k−1.

    Khác `_khau_anh_noi_canh`: ở đây ẢNH CHÍNH LÀ MỐC, không có khung phụ nào.

    ẢNH vẽ TUẦN TỰ, vì ảnh mốc k phải nhìn ảnh mốc k−1 mới giữ được góc máy. Đo
    27/08/2026 trên phim timelapse/0001, cùng lời nhắc, chỉ đổi ảnh tham chiếu —
    tiền cảnh lệch bao nhiêu so với ảnh gốc (thấp = máy đứng yên):

        mốc | nhìn ảnh mốc trước | chỉ nhìn ảnh gốc
          2 |        4,20        |      40,49        ← bố cục tự kéo gần lại
          3 |       14,41        |      17,00

    Nên không bỏ chuỗi được, dù bỏ được thì 60 ảnh chạy song song.

    CLIP thì chạy theo KHỐI, và mỗi khối là một mạch LAI. Đo 27/08/2026, cùng
    thang "đổi thay đi được nửa đường ở giây thứ mấy" (đều thì phải ~4,0 / 8 giây):

        ghim hai đầu MỌI clip : 7,5 và 7,5     — đứng phẳng 7 giây rồi giật một phát
        trôi tự do, nối chuỗi : 1,0 / 3,0 / 6,5 — đổi thay chảy thật

    Nhưng trôi tự do thì hình học đi mất: sau 3 clip, con voi đá bên phải hoá
    thành trống đồng rồi biến hẳn. Nên: `CHUOI_TROI` clip trôi cho đổi thay chảy,
    rồi MỘT clip ghim hạ đúng vào ảnh mốc vẽ sẵn để kéo khung hình về chỗ cũ.

    Nhờ vậy chỉ 1/4 số mốc phải vẽ ảnh — mà vẽ ảnh mới là chỗ tuần tự, mỗi tấm
    ~45 giây. Các KHỐI thì độc lập với nhau (mỗi khối bắt đầu từ ảnh mốc riêng)
    nên chạy song song được; trong một khối thì clip phải nối tiếp nhau.
    """
    def lam(luot: LuotChay, tt: TrangThaiKhau):
        from concurrent.futures import ThreadPoolExecutor  # noqa: PLC0415

        from .dao_dien_auto import THU_MUC_THAM_CHIEU, ThamChieuCanh  # noqa: PLC0415
        from .timelapse import (  # noqa: PLC0415
            SONG_SONG_KHOI, TEP_MOC, anh_gan_nam, prompt_anh_moc,
        )

        d = luot.thu_muc
        canh = _doc_canh(luot)
        thu_muc = os.path.join(d, "5-anh")
        thu_muc_clip = os.path.join(d, "6-clip")
        os.makedirs(thu_muc, exist_ok=True)
        os.makedirs(thu_muc_clip, exist_ok=True)
        goc = os.path.join(d, THU_MUC_THAM_CHIEU, "loc1.png")
        so = SoTheoDoi(bc, nhip=bc.nhip_hoi)
        giay = _giay_clip(bc)

        # ── 1. Vẽ ảnh cho các mốc GHIM, tuần tự, mỗi tấm nhìn tấm trước ──────
        khoi, dang = [], []
        for c in canh:
            dang.append(c)
            if c.get("ghim"):
                khoi.append(dang)
                dang = []
        if dang:                      # đuôi thừa: coi như một khối, clip cuối vẫn ghim
            khoi.append(dang)
        bc.ghi("  {0} cảnh → {1} khối; vẽ {1} ảnh mốc rồi chạy các khối song "
               "song (tối đa {2} khối cùng lúc).".format(
                   len(canh), len(khoi), SONG_SONG_KHOI))
        xong_anh = 0
        truoc = goc
        # Kho ẢNH THẬT tra được ở khâu bảng cảnh. Mốc nào rơi vào thời đã có máy
        # ảnh thì được gắn kèm tấm gần năm nhất — người xem nhận ra ngay chỗ nào
        # phim vẽ sai so với ảnh họ từng thấy. Xem `_khau_bang_canh_timelapse`.
        kho_anh, bang = [], {}
        try:
            with open(os.path.join(d, "4-anh-that.json"), encoding="utf-8") as f:
                kho_anh = json.load(f)
            with open(os.path.join(d, TEP_MOC), encoding="utf-8") as f:
                bang = json.load(f)
        except (OSError, ValueError):
            pass
        moc_theo_nam = {int(m["nam"]): m for m in (bang.get("moc") or [])
                        if str(m.get("nam", "")).lstrip("-").isdigit()}
        # ẢNH NHẬN DẠNG: tấm chụp chỗ ấy NGÀY NAY. Gắn vào mọi tấm vẽ — nó không
        # nói năm nào, nó nói ĐÂY LÀ CHỖ NÀO và máy đứng ở đâu.
        nhan_dang = next((x for x in kho_anh if x.get("nhan_dang")), None)
        if nhan_dang and not os.path.isfile(nhan_dang.get("tep") or ""):
            nhan_dang = None
        bc.ghi("  ảnh nhận dạng: {0}".format(
            "{0} ({1})".format(nhan_dang.get("ten"), nhan_dang.get("nam"))
            if nhan_dang else "KHÔNG CÓ — hình học sẽ trôi, xem prompt_anh_moc"))
        dem_that = 0
        sach_lai = con_ban = 0

        for k in khoi:
            bc.kiem_dung()
            c = k[-1]
            n = int(c["scene_id"])
            anh = os.path.join(thu_muc, "{0}.png".format(n))
            if not os.path.exists(anh):
                nam_moc = _so_nguyen(c.get("nam_den")) or 0
                that = anh_gan_nam(kho_anh, nam_moc)
                if that is nhan_dang:
                    that = None      # đã gắn rồi, đừng gắn hai lần cùng một tấm
                # ẢNH NHẬN DẠNG đứng ĐẦU danh sách — lời nhắc gọi nó là "tấm thứ
                # nhất". Nó phải có mặt ở MỌI tấm: bản trước chỉ neo tấm đầu rồi
                # để 14 tấm sau vẽ chuyền tay nhau, và bố cục trôi tới 99,8/255.
                refs = [nhan_dang["tep"]] if nhan_dang else []
                if os.path.isfile(truoc):
                    refs.append(truoc)
                if that and os.path.isfile(that.get("tep") or ""):
                    refs.append(that["tep"])
                    dem_that += 1
                m = moc_theo_nam.get(nam_moc)
                if m and bang:
                    c = dict(c, img_prompt=prompt_anh_moc(
                        bang, m, anh_that=that, anh_nhan_dang=nhan_dang))
                bc.ghi("    mốc {0}: {1} ảnh tham chiếu{2}.".format(
                    c.get("nam_den"), len(refs),
                    " (có ảnh thật {0})".format(that.get("nam")) if that else ""))
                hop = ThamChieuCanh(bc, refs)
                _lam_anh_canh(bc, luot, c, anh, hop, so=so)
                # Cua soat thoi dai: mot tam ban de ra HAI clip ban.
                v = _soat_thoi_dai_anh_moc(bc, luot, c, anh, nam_moc,
                                           bang.get("noi") or "", hop, so=so)
                if v == 1:
                    sach_lai += 1
                elif v == 2:
                    con_ban += 1
            xong_anh += 1
            truoc = anh
        if kho_anh:
            bc.ghi("  {0}/{1} ảnh mốc có ảnh thật đối chiếu.".format(dem_that, len(khoi)))
        if sach_lai or con_ban:
            bc.ghi("  cửa soát thời đại: {0} tấm vẽ lại thành sạch, {1} tấm "
                   "vẫn đáng ngờ.".format(sach_lai, con_ban))

        # ── 2. Mỗi khối: clip trôi nối chuỗi, clip cuối ghim vào ảnh mốc ─────
        def chay_khoi(i, k):
            dau = goc if i == 0 else os.path.join(
                thu_muc, "{0}.png".format(int(khoi[i - 1][-1]["scene_id"])))
            tu = dau
            xong = 0
            for c in k:
                bc.kiem_dung()
                n = int(c["scene_id"])
                clip = os.path.join(thu_muc_clip, "{0}.mp4".format(n))
                if not os.path.exists(clip):
                    cuoi = (os.path.join(thu_muc, "{0}.png".format(n))
                            if c.get("ghim") else None)
                    try:
                        _lam_clip(bc, luot, c, tu, clip, giay, so=so,
                                  khung_dau=True, anh_cuoi=cuoi)
                    except Exception as loi:  # noqa: BLE001
                        # MỘT clip hỏng chỉ được mất MỘT cảnh. Trước đây lỗi này
                        # ném ra ngoài và cả khối bốn cảnh mất trắng — mà máy chủ
                        # lúc đông thì hỏng lẻ là chuyện thường (đo 27/08/2026:
                        # "máy chủ nhận việc rồi bỏ đó"). Nay: bỏ đúng cảnh ấy,
                        # mở lại chuỗi từ ảnh mốc của khối, đi tiếp.
                        bc.ghi("    cảnh {0}: clip hỏng ({1}) — bỏ cảnh này, mở "
                               "lại chuỗi từ ảnh mốc.".format(n, str(loi)[:80]))
                        tu = dau
                        continue
                xong += 1
                # Clip sau nối vào ĐÚNG khung cuối clip trước — đó là chỗ đổi
                # thay chảy liên tục thay vì giật từng nấc.
                tu = (os.path.join(thu_muc, "{0}.png".format(n)) if c.get("ghim")
                      else _khung_cuoi_clip(bc, clip, os.path.join(
                          thu_muc_clip, "_cuoi-{0}.png".format(n))))
            return xong

        xong_clip = 0
        try:
            with ThreadPoolExecutor(max_workers=SONG_SONG_KHOI) as bo:
                viec = [(i, bo.submit(chay_khoi, i, k)) for i, k in enumerate(khoi)]
                for i, v in viec:
                    try:
                        xong_clip += v.result()
                    except Exception as loi:  # noqa: BLE001
                        # Một khối hỏng chỉ mất mấy bước chuyển của nó; khâu dựng
                        # bỏ qua clip thiếu, phần còn lại của phim vẫn liền.
                        bc.ghi("    khối {0} (mốc {1}): hỏng ({2}) — đi tiếp.".format(
                            i + 1, khoi[i][0].get("nam_tu"), str(loi)[:80]))
        finally:
            so.dong()
        return {"so_anh": xong_anh, "so_clip": xong_clip, "tong": len(canh)}

    return lam


def dung_bo_viec(bc: BoiCanh) -> Dict[str, Callable]:
    """Bảng `mã khâu → hàm làm việc`, đưa thẳng cho `core.auto.chay`.

    Khâu nào KHÔNG có trong bảng thì `core.auto.chay` đánh dấu bỏ qua — đó là
    cách kênh timelapse tắt hai khâu tiếng: nó không có lời đọc, nhịp phim lấy
    từ bảng mốc thời gian (xem `core/timelapse.py`).
    """
    from .timelapse import la_timelapse  # noqa: PLC0415

    if la_timelapse(bc.kenh):
        return {
            "kich-ban": _khau_kich_ban_timelapse(bc),
            "bang-canh": _khau_bang_canh_timelapse(bc),
            "anh": _khau_anh_timelapse(bc),
            # Khâu clip ở đây là CHÍNH mạch khối, không phải `_khau_clip` dùng
            # chung. Lý do: `_khau_clip` đòi MỌI cảnh phải có ảnh, mà mạch lai
            # chỉ vẽ ảnh cho mốc ghim (15/64 cảnh) — nên nó chết ngay ở cảnh đầu
            # tiên không có ảnh ("cảnh 5 chưa có ảnh nên chưa làm clip được").
            # Gọi lại chính mạch khối thì nó bỏ qua clip đã có và thử lại đúng
            # những cảnh còn thiếu, tức bấm "Chạy tiếp" là vá được mẻ hỏng.
            "clip": _khau_anh_timelapse(bc),
            "thumbnail": _khau_thumbnail(bc),
            "dung": _khau_dung(bc),
        }
    return {
        "kich-ban": _khau_kich_ban(bc),
        "giong-doc": _khau_giong_doc(bc),
        "phu-de": _khau_phu_de(bc),
        "bang-canh": _khau_bang_canh(bc),
        "anh": (_khau_anh_noi_canh(bc) if _la_noi_canh(bc.kenh) else _khau_anh(bc)),
        "clip": _khau_clip(bc),
        "thumbnail": _khau_thumbnail(bc),
        "dung": _khau_dung(bc),
    }


def _duong_nhac(kenh) -> str:
    """Đường dẫn đầy đủ tới nhạc nền của kênh, hoặc chuỗi rỗng.

    Tệp thiếu thì trả rỗng chứ **không ném lỗi**: dựng xong một video không
    nhạc vẫn hơn hẳn hỏng cả khâu dựng vì một tệp nhạc đặt sai tên — nhất là
    khi bảy khâu trước đã tiêu tiền rồi.
    """
    ten = str(getattr(kenh, "nhac_nen", "") or "").strip()
    if not ten:
        return ""
    duong = ten if os.path.isabs(ten) else os.path.join(
        str(getattr(kenh, "duong", "") or ""), ten)
    return duong if os.path.isfile(duong) else ""


