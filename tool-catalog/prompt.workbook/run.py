"""SRT -> scene-manifest.v1 + XLSX tuong thich VE3.

Canh duoc cat **theo NGHIA cua loi doc**, khong theo dong ho. Chu du an,
15/08/2026: *"o tab prompt visuals thi nhu auto va tool goc no khong lam mac
dinh 8s ma no theo noi dung srt"*.

Cach chia nam o `core/chia_canh.py` — dung chung voi tab Tu dong, khong chep
tay sang day. Duong lui khi khong goi duoc AI: `core/srt_scenes.group_cues`
cat theo dong ho, va luc do tool **noi thang tren man hinh** rang canh dang bi
cat theo dong ho chu khong theo y.
"""

from __future__ import annotations

import json
import os
import re
import sys
import threading
from pathlib import Path
from typing import Any, Callable, Dict, List, Mapping, Sequence

_STUDIO = Path(__file__).resolve().parents[2]
if str(_STUDIO) not in sys.path:
    sys.path.insert(0, str(_STUDIO))

from core.prompt_visuals import (  # noqa: E402,F401 — doi_thiet_ke/loi_nhac dung qua module
    DUOI_BOI_CANH, DUOI_CHAN_DUNG, DUOI_GIAI_DOAN, bo_tu_the, doi_thiet_ke_nhan_vat,
    goc_cua_id, loi_nhac_thiet_ke_lai,
)
from core.chia_canh import (  # noqa: E402
    DUOI_CAM, KHUON_MAC_DINH, bang_phu_de, chia_theo_nghia, loi_nhac_chia,
)
from core.goi_van_ban import goi_van_ban, loc_json  # noqa: E402
from core.prompt_visuals import CHO_TRONG_KHUON_CHIA  # noqa: E402
from core.chia_canh import MIN_GIAY_CANH, nhip_tu_khuon  # noqa: E402
from core.srt_scenes import (  # noqa: E402
    clock, enforce_max_duration, group_cues, max_seconds_for, parse_srt, target_seconds_for,
)

#: Cot cua sheet `scenes`. Giu DUNG thu tu nay: khau dung video doc file Excel
#: nay theo ten cot, va `video_note = "SKIP"` la cach bao no bo qua mot canh.
SCENE_COLUMNS = [
    "scene_id", "srt_start", "srt_end", "duration", "planned_duration", "srt_text",
    "scene_kind", "subject_mode", "primary_subject", "primary_action", "visual_anchor",
    "must_not_show", "img_prompt", "prompt_json", "video_prompt", "img_path", "video_path",
    "status_img", "status_vid", "characters_used", "location_used", "reference_files", "media_id",
    "video_note", "segment_id",
    # Cot them SAU cac cot VE3 (VE3 doc theo ten nen khong anh huong): ban dich
    # tieng Viet cua loi doc — chu du an 24/08/2026 dua prompt tham khao co cot
    # "Ban dich tieng Viet" de khach hieu canh dang noi gi ma toi uu prompt.
    "srt_text_vi",
]

#: Sheet `locations` — dung cot cua VE3_SUITE (`excel_manager.LOCATIONS_SHEET`).
LOCATION_HEADERS = ["id", "name", "english_prompt", "location_lock",
                    "lighting_default", "image_file", "status", "sheet_prompt"]

#: Sheet `thumbnail` — cot cua tab Tu dong (`auto_khau.COT_THUMB`) + hai cot
#: rieng: chu hook va tieu de de xuat (tab nay khong co tieu de san).
THUMB_HEADERS = ["thumb_id", "version_desc", "img_prompt", "characters_used",
                 "location_used", "reference_files", "img_path", "status_img",
                 "thumb_text", "title"]

#: Sheet `music` — theo `9-nhac.md` cua kenh: moi track ~105 giay, prompt Suno.
MUSIC_HEADERS = ["music_id", "start_time", "end_time", "suno_prompt", "mood",
                 "status"]

#: Ba cach ke chuyen — chu du an 24/08/2026, tham khao VE3_SUITE:
#:   mot_nhan_vat          loai 1: MOT nhan vat co dinh cua kenh (anh nv1.png),
#:                         khong casting, khong boi canh tham chieu;
#:   nhan_vat_va_boi_canh  loai 2: nv1 co dinh + AI dung them nhan vat phu
#:                         (nv2..) va boi canh tham chieu (loc1..) theo noi dung;
#:   tu_xay                loai 3: AI tu xay ca dan nhan vat lan boi canh.
CHE_DO_MOT_NV = "mot_nhan_vat"
CHE_DO_NV_BOI_CANH = "nhan_vat_va_boi_canh"
CHE_DO_TU_XAY = "tu_xay"

#: Mot track Suno dai chung nay (giay) — xem `9-nhac.md`.
GIAY_MOI_TRACK = 105

#: Sheet `story`: phan tich phim + cac MAN (segment). Sheet `director_plan`:
#: ke hoach dao dien tung beat trong man. Ca hai chi co o loai 2, 3.
STORY_HEADERS = ["segment_id", "name", "message", "emotion", "motif",
                 "srt_from", "srt_to", "genre", "arc", "context_lock"]
PLAN_HEADERS = ["segment_id", "beat", "srt_from", "srt_to", "purpose",
                "characters", "location", "shot_size", "camera",
                "element_motion", "emotion", "motif", "status"]

#: ═══ ĐẠO DIỄN CHO LOẠI 2, 3 — HỌC TỪ VE3, ViMax, vox-director ═══
#:
#: Chu du an 24/08/2026: *"o 2 option kia can dong bo nhieu, no nhu 1 bo phim
#: co su lien ket… san pham dau ra nhu 1 bo phim giup khan gia khong the roi
#: mat"*. Truoc day loai 2/3 chi co casting roi chia canh thang — tram canh la
#: tram an du roi, khong man, khong mach. Nay lam TU TREN XUONG nhu VE3:
#:
#:   A. doc phim   → the loai, cung truyen, the gioi hinh, CAC MAN (segment)
#:   B. casting    → nhan vat (dac diem co dinh/thay duoc, chan dung tham chieu)
#:                   + boi canh (mo ta co dinh + anh thiet lap)
#:   C. ke hoach   → moi man mot ke hoach beat: ai, o dau, co canh, chuyen dong
#:   D. viet prompt → moi khuc nhan ke hoach cua dung nhung dong SRT ay
_KHUON_PHIM = """You are the showrunner of a narrated video. Read the WHOLE
narration and plan it like a film before any shot is written.

## Narration
{transcript}

## Context
{context}

## Decide
1. `genre` (e.g. psychology explainer, personal story, history) and `arc` —
   pick one narrative arc that fits: hook_payoff, man_in_hole, timeline,
   three_act, story_spine, myth_buster, listicle, how_it_works.
2. `context_lock`: ONE paragraph fixing the visual world for every scene —
   era, place, time of day range, weather/mood, recurring colour accents.
3. `segments`: split the narration into ACTS by meaning (using the SRT line
   numbers). This narration is {phut} minutes long — expect about {so_man}
   acts (roughly one act per 60–90 seconds; a 3-minute video has ~3, a
   30-minute one 20+). Follow the meaning, not the number: never merge two
   different ideas into one act to hit a count, never split one idea in two.
   Each act has a `name`, the `message` it lands, the dominant
   `emotion`, and ONE visual `motif` that recurs inside the act (an object,
   a gesture, a weather) — the motif is what makes an act feel like one piece.
   Acts are contiguous and cover every line; the FIRST act must open with a
   hook within its first scene.
4. `characters_mentioned` and `locations_mentioned`: EVERY person, creature or
   group that acts in the story, and EVERY distinct place it visits — for the
   casting step. Do not shorten these lists; a long story has a long cast.

## Return JSON only, no commentary
```json
{{"genre": "...", "arc": "...", "context_lock": "...",
 "segments": [{{"segment_id": 1, "name": "...", "message": "...",
                "emotion": "...", "motif": "...", "srt_from": 1, "srt_to": 12}}],
 "characters_mentioned": ["..."], "locations_mentioned": ["..."]}}
```
"""

_KHUON_KE_HOACH = """You are the DIRECTOR planning act {so}/{tong} of a narrated
video: "{ten}". Message of this act: {message}. Emotion: {emotion}. Visual
motif to recur: {motif}.

## Visual world (hold it)
{context_lock}

{cast}

## The narration of this act (each line: `index | start -> end | text`)
{srt}

## Plan the beats — one beat per IDEA, each {min_sec}–{max_sec} seconds
A beat is NOT a line: work out each beat's length from the timestamps and
merge short lines that belong to one thought until the beat is at least
{min_sec} seconds; split a long line where the thought turns. Fewer, fuller
beats beat many thin ones.
For every beat give:
- `srt_from`/`srt_to` — contiguous, covering every line of this act, no gaps;
- `purpose` — what this beat does for the story (establish / show the pain /
  turn / reveal / land the point);
- `characters` — ids from the cast that are IN FRAME ("" if none) and
  `location` — a location id ("" only if the story is truly nowhere). THE PLACE
  IS WHERE THE CHARACTERS ARE IN THAT MOMENT, like a film: a beat stays in the
  SAME place as the previous beat until the characters move. Never cut to
  another place for variety — vary shot size, angle and what moves in frame
  instead. A conversation happens in one place from start to end.
  MOVING COUNTS EVEN WHEN NO SENTENCE ANNOUNCES IT. A journey has stages, and
  the cast list usually has a location for each: pushing off is at the shore,
  paddling and meeting someone out on the water is in the MIDDLE of the water,
  arriving is at the far side. Ask of every beat "where are their feet right
  now?" — not "where did this stretch of the story begin?". Leaving characters
  at the departure point through a whole crossing is the commonest failure
  here: measured 27/08/2026 on a film that had a "middle of the pond" location
  built and unused while the mid-water meeting was drawn on the bank.
  EVERY LOCATION IN THE CAST LIST MUST BE USED by at least one beat. The cast
  built exactly the places this story visits; one that no beat uses means a
  beat is in the wrong place — find it and move it.
  ONCE THEY HAVE MOVED ON, THEY DO NOT SLIDE BACK. A place they have left may
  return only where the narration takes them back (they walk home, they climb
  out). A plan that reads A, B, A, B for one continuous journey is wrong every
  time: measured 27/08/2026, a boy already adrift in the middle of the water
  had his whole conversation with the duck planned back on the bank, then the
  next beat was out in the middle again. Read the beats in a row like a
  storyboard and ask whether a viewer could walk that path;
- `shot_size` from EST_WIDE, WIDE, MEDIUM, CLOSE, DETAIL — build intensity
  toward the act's turning point (WIDE→MEDIUM→CLOSE), open on a wide when a
  place is new;
- `camera` from static, push_in, pull_out, pan, tilt, orbit — never the same
  as the previous beat; reserve `static` for the beat that lands the point;
- `element_motion` — what physically changes inside the frame by the end of
  the beat (something enters, breaks, grows, tips, lights up, empties);
- `emotion` and `motif` (how the act's motif shows up in this beat, or "").
The first beat of act 1 is the hook: the most arresting image of the video.
The last beat of the act closes it on the motif.

## Return JSON only, no commentary
```json
{{"beats": [{{"srt_from": 1, "srt_to": 3, "purpose": "...", "characters": "nv1",
             "location": "loc1", "shot_size": "WIDE", "camera": "push_in",
             "element_motion": "...", "emotion": "...", "motif": "..."}}]}}
```
"""

#: Prompt chan dung tham chieu (ViMax): full-body, nhin thang, nen trang, canh
#: 16:9 — khung hinh nay lam anh tham chieu cho MOI canh co nhan vat ay.
#: "ONE figure only": do 25/08/2026 (Meo di hia 3D) — mo ta cau ut co "later
#: transform into fine marquis attire" nen anh tham chieu ve HAI ban cau ut
#: dung canh nhau, va moi canh sau chep luon bo cuc hai nguoi.
_DUOI_CHAN_DUNG = DUOI_CHAN_DUNG
_DUOI_BOI_CANH = DUOI_BOI_CANH
_DUOI_GIAI_DOAN = DUOI_GIAI_DOAN

#: Bao nhieu ky tu loi doc dua cho luot viet anh bia / nhac. 0 = ca bai:
#: anh bia phai biet cao trao o cuoi, nhac phai theo mach ca video — cat bot
#: la bia va nhac chi biet phan dau.
BIA_KY_TU = 0

#: Toi da bao nhieu GIAY tieng trong mot khuc chia canh. Chu du an 24/08/2026:
#: *"neu ma no dai thi phai co ke hoach chia de api song song nhieu luong cho
#: nhanh"*. 100 giay ≈ 12–20 canh moi luot goi, tra ve trong duoi mot phut;
#: video 10 phut thanh ~6 khuc chay song song thay vi 2–3 khuc nang.
GIAY_MOI_KHUC = 100.0
CREATIVE = ("scene_kind", "subject_mode", "primary_subject", "primary_action", "visual_anchor",
            "must_not_show", "img_prompt", "video_prompt", "characters_used", "location_used")

#: Cot cua sheet `characters` — trung ten voi VE3_SUITE va voi tab Tu dong
#: (`core/auto_khau.COT_NHAN_VAT`), de file mo thang bang VE3 duoc.
CHARACTER_HEADERS = ["id", "role", "name", "body_mode", "english_prompt", "vietnamese_prompt",
                     "image_file", "status", "gender", "age", "notes",
                     # Cot them sau cot VE3: prompt tao anh chan dung tham chieu.
                     "sheet_prompt"]

#: Tran chu cho mot luot chia canh. Cao vi moi canh keo theo hai loi nhac tieng
#: Anh chi tiet; hut tran thi JSON dut giua cau va ca khuc phai hoi lai.
TOKEN_CANH = 16384

#: Tran chu cho luot "casting" — mot luot doc ca loi doc roi dung dan nhan vat.
#: Bang TOKEN_CANH: truyen dai co the 12+ nhan vat, 8+ noi chon, moi muc kem
#: sheet_prompt — 4096 la dut JSON giua chung.
TOKEN_CAST = TOKEN_CANH

#: Lời nhắc dựng dàn nhân vật + phong cách từ chính lời đọc. Prompt Visuals
#: không có kênh nên không có `nv1.png` hay style sẵn — bản này bảo AI **tự rút
#: ra** dàn nhân vật lặp lại và một phong cách thống nhất, để mọi cảnh sau dùng
#: chung. Không có nhân vật lặp lại (ví dụ video toàn cảnh vật) thì trả mảng
#: rỗng — lúc đó tool về đúng hành vi cũ, không bịa người.
_KHUON_CAST = """You are the casting director AND production designer for a
narrated video. Read the whole narration transcript below and decide:

1. The CHARACTERS — every person, creature or group that acts in the story or
   appears in more than one scene: the hero, the helper, the rival, the
   parent, the ruler, a crowd ("the villagers" is one id). For each, write ONE
   fixed English appearance description (`english_prompt`) that every scene
   will reuse verbatim so the character never drifts, and make the characters
   clearly DISTINCT from each other (build, age, clothing colour, one signature
   prop). The description MUST include the FULL costume and the props that
   SAY THE ROLE at first glance — a king: a golden crown AND a royal robe with
   fur trim; a queen: gown and circlet; a princess: gown and tiara; a soldier:
   uniform; a miller: apron. A face and a beard alone are not a king: the
   image generator dressed a crownless "jolly old man" as a beggar (25/08/2026).
   Every human character needs: face, hair, build, complete outfit with
   colours, footwear, one signature prop. If a character's LOOK CHANGES during the story (a cat is given a hat
   and boots, a peasant puts on royal clothes, a king loses his crown), give
   `stages` — a list in story order, each `{{"when": "<the story moment>",
   "outfit": "<clothes and props at that stage>"}}`; the first stage is the
   look at the START. Then `english_prompt` describes only face, body and
   fur/skin (no clothes) — every stage becomes its own reference image with the
   same face. Characters that never change have no `stages`. A character that
   TRANSFORMS (an ogre becomes a lion, then a mouse; a frog becomes a prince)
   also gets `stages` — one per form, `outfit` describing the whole new body —
   because each form needs its own reference picture or the scenes invent it.
   EVERY character must declare `body_mode` — the shape its body has in THIS
   story, as data, not buried in prose:
     "human"          a person;
     "animal"         an ordinary animal that walks on all fours (or swims,
                      or perches) and never stands, talks or wears clothes;
     "upright-animal" an animal that walks on two legs, gestures, talks or
                      wears clothes like a person.
   Judge it from the story, not from how cute it would look. A cat that only
   miaows, jumps and carries things in its mouth is "animal"; a cat that puts
   on boots and speaks to a king is "upright-animal". Getting this wrong
   poisons every scene: measured 27/08/2026 on a film whose cat was an ordinary
   house cat — the reference sheet drew it standing on two legs, and all 19
   scenes with the cat drew a DIFFERENT cat, because no scene could use a
   two-legged cat and each one re-invented an ordinary one.
   The SAME rule applies when only the BODY POSTURE changes: an animal that is an
   ordinary animal at first and only later stands, speaks or acts like a person
   (or the reverse) needs one stage per posture — `outfit` says the whole body
   mode ("on all four paws as an ordinary animal", "standing upright on two legs
   like a person, wearing …"). Do not skip this: with one upright portrait, every
   scene that shows the animal on all fours invents a different animal (measured
   26/08/2026: identity score 3/5, a different cat in each scene).
   Describe what image generators get wrong by default, in POSITIVE words:
   for an animal, the exact fur pattern ("solid plain golden-yellow fur without
   any stripes", not just "yellow cat"), "bare paws" when it wears no shoes,
   and the exact colour of each garment ("plain brown vest") — otherwise the
   scene's accent colour recolours the vest and the model adds stripes, boots
   or a bag from its habits.
   ONE id = ONE individual, because each id becomes one reference portrait of
   a single figure: two brothers are two ids (each with his own look); a crowd
   (guards, villagers, courtiers) is ONE representative member, and the scene
   prompt says how many of them appear. Never an id that means "the two…".
   Image generators REJECT any animal wearing footwear (boots, shoes,
   slippers — measured 25/08/2026 on a cat, every wording) and a cat under a
   wide-brimmed hat of any kind. Never give an animal footwear even if the
   story says boots: the narration may say "boots", the picture gives it a
   beret, a small cap, a vest, a scarf or a satchel instead — never a
   wide-brimmed or feathered hat on a cat.
   Write descriptions in POSITIVE terms only — what IS there. Never put
   "no boots", "without a hat", "no weapons" inside a description: safety
   filters and image models read the noun, not the negation (a description
   saying "bare paws, no boots" was rejected; "bare paws" alone passed).
   Never copy the signature look of a famous copyrighted character (a cat in a
   feathered musketeer hat and tall boots, a mouse in red shorts, a blue
   hedgehog…): image generators reject it. Invent a fresh, simple look instead
   (a beret, a scarf, a small vest, plain boots) and keep it for the whole
   video. When the story itself hands a character a famous look, keep ONLY
   the item the plot needs (the cat gets boots because it asks for boots) and
   drop the copyrighted flourish (no feathered hat of any colour or shape on
   that cat — a plain cap or nothing on the head; and no boots on it, see above). Use FAMILY-FRIENDLY wording in every description — this is for
   children and for image generators with strict safety filters: no weapons
   (guards carry ceremonial staffs, not halberds or swords), no "brutal",
   "hulking", "menacing", "ashen", "iron-studded"; a villain is "very big,
   grumpy, with a wide mouth", never frightening. {fixed_rule}
   Only a pure landscape/abstract video has an empty `characters` list — do NOT
   invent a person there, and do NOT drop a character the story uses.
2. The LOCATIONS — every distinct place the story visits: a cottage, a river
   bank, a palace hall, a castle, a road, a forest edge. For each, write ONE
   fixed English description of the PLACE (`english_prompt`, no people in it),
   a short `location_lock` (what must never change) and `lighting_default`.
   Give ids `loc1`, `loc2`… A palace and a castle are two places; do not
   merge distinct settings into one.
   ONE id = ONE camera-able place: the OUTSIDE of a castle (gate, walls,
   towers) and its INSIDE (hall, throne room) are TWO locations with two ids;
   a cottage interior and its doorstep are two. A location reference is a
   single view, and a scene set "in the castle hall" must not be locked to a
   picture of the castle gate (measured 25/08/2026: 10 of 32 castle scenes
   mismatched for exactly this).
3. ONE consistent visual `style` for the whole video: `image_style`, `palette`,
   `motion`.

## MUST COVER — the story analysis already found these; give each one an id
Characters mentioned: {phai_co_nv}
Places mentioned: {phai_co_loc}
Merge only true duplicates (the same person under two names). Anything on
these lists that is missing from your answer will be treated as an error.

## Context
{context}

## Transcript
{transcript}

## Return JSON only, no commentary
```json
{{"style": {{"image_style": "...", "palette": "...", "motion": "..."}},
 "characters": [
   {{"id": "{nv_dau}", "role": "...", "name": "...",
    "body_mode": "human | animal | upright-animal",
    "english_prompt": "<fixed appearance, no scene-specific pose>",
    "reference_lock": "<short identity anchor>",
    "gender": "", "age": "", "notes": "",
    "stages": [{{"when": "at the start", "outfit": "..."}},
               {{"when": "after ...", "outfit": "..."}}]}}
 ],
 "locations": [
   {{"id": "loc1", "name": "...", "english_prompt": "<fixed look of the place>",
    "location_lock": "<what never changes>", "lighting_default": "..."}}
 ]}}
```
"""

#: Luot BO SUNG khi casting bo sot: chi xin dung nhung muc con thieu, id tiep
#: theo dan da co. Do 25/08/2026 (Puss in Boots, 8 phut): buoc doc phim nhan ra
#: 9 nhan vat + 6 noi, casting tra ve 4 + 4 — mat ca yeu tinh lan cung dien.
_KHUON_BO_SUNG = """You are completing the cast and location list of a narrated
video. The list below ALREADY EXISTS — do not repeat or redefine any of it:
{da_co}

## Still missing — write an entry for EACH of these
Characters: {thieu_nv}
Places: {thieu_loc}

Rules: same JSON shape as before; character ids continue from `{nv_tiep}`,
location ids from `{loc_tiep}`; each `english_prompt` is one fixed English
appearance/place description reused verbatim by every scene; characters must
be clearly DISTINCT from the existing ones; keep the same visual style.
FIRST check whether a "missing" name is simply another name (often Vietnamese)
for something that ALREADY exists in the list above — "Cối xay bột" is the
mill `loc9`, "Lâu đài của nhà vua" is the palace `loc10`. Such names go into
`same_as` (name → existing id) and get NO new entry: two ids for one place
give the film two different castles (measured 26/08/2026).

## Transcript (for reference)
{transcript}

## Return JSON only, no commentary
```json
{{"same_as": {{"<missing name that already exists>": "<existing id>"}},
 "characters": [{{"id": "{nv_tiep}", "role": "...", "name": "...",
                 "body_mode": "human | animal | upright-animal",
                 "english_prompt": "...", "reference_lock": "...",
                 "gender": "", "age": "", "notes": ""}}],
 "locations": [{{"id": "{loc_tiep}", "name": "...", "english_prompt": "...",
                "location_lock": "...", "lighting_default": "..."}}]}}
```
"""

#: Tu vo hai nhung bo loc an toan cua nha cung cap anh hay chan vo co. Do
#: 25/08/2026: anh tham chieu "small upright anthropomorphic cat ... sly small
#: smile" bi content_rejected HAI lan; bo hai tu la qua ngay. Ca 85 canh cua
#: video do co 12 canh "anthropomorphic", 23 canh "sly" — thay truoc khi ghi
#: Excel, khoi de khach gap loi "noi dung bi tu choi" tren mot con meo di hia.
_TU_BI_LOC = [(re.compile(r"\banthropomorphic\b", re.I), "upright humanlike"),
              (re.compile(r"\bsly\b", re.I), "knowing")]


def _lam_lanh_prompt(chu: str) -> str:
    for mau, thay in _TU_BI_LOC:
        chu = mau.sub(thay, str(chu or ""))
    return chu


# ── Dọn mô tả nhân vật: bỏ câu-lệnh lọt vào và mệnh đề phủ định ─────────────
#
# Đo 25/08/2026 22:15 (chạy thử story-3d): AI chép nguyên hướng dẫn vào mô tả
# ("describe only face, body and hair here;") và viết phủ định ("no hat, no
# clothes at all, bare paws with no footwear"). Máy vẽ và bộ lọc đọc DANH TỪ,
# không đọc chữ "no" — "no boots" vẫn là "boots". Dọn bằng mã, không tin lời hứa.
_MENH_DE_LENH = re.compile(r"\s*[;,.]?\s*(?:describe|write|use|include|keep|mention)\b[^;.]*?\bhere\b[^;,.]*", re.I)
_MENH_DE_PHU_DINH = re.compile(r"\s*(?:,|;|\band\b)?\s*(?:with\s+)?(?:no|without|never)\s+(?!(?:any\s+|the\s+)?(?:stripes|spots|patches|markings|pattern))[a-z][a-z\- ]{0,40}?(?=(?:,|;|\.|$|\band\b))", re.I)


def _don_mo_ta(chu: str) -> str:
    """Bỏ câu-lệnh lọt vào và mệnh đề "no X / without X" khỏi một mô tả.

    Giữ "no stripes / no spots" (đó là hoa văn lông, cần nói) — còn "no hat",
    "no boots", "no clothes at all", "with no footwear" thì bỏ.
    """
    t = str(chu or "")
    t = _MENH_DE_LENH.sub("", t)
    t = _MENH_DE_PHU_DINH.sub("", t)
    t = re.sub(r"\s*,\s*,", ",", t)
    t = re.sub(r":\s*,", ":", t)
    t = re.sub(r";\s*;", ";", t)
    t = re.sub(r"\s+", " ", t).strip(" ,;")
    return t


def _don_dan(characters) -> int:
    """Dọn `english_prompt` (và `outfit` trong `stages`) của cả dàn. Trả về số nhân vật đã đổi."""
    n = 0
    for c in characters or []:
        cu = str(c.get("english_prompt") or "")
        moi = _don_mo_ta(cu)
        if moi and moi != cu:
            c["english_prompt"] = moi
            n += 1
        for st in c.get("stages") or []:
            if isinstance(st, dict) and st.get("outfit"):
                st["outfit"] = _don_mo_ta(str(st["outfit"]))
    return n


#: Dang than -> mot cau dat NGAY SAU mo ta, truoc duoi chan dung.
#:
#: ═══ VI SAO PHAI LA DU LIEU, KHONG PHAI VAN XUOI ═══
#:
#: Truoc 27/08/2026 khong cho nao trong day chuyen KHAI ra "con meo nay di bon
#: chan". Dieu ay chi nam chim trong mot cau van ("four bare soft paws"), con
#: `DUOI_CHAN_DUNG` thi dan cung cho MOI nhan vat cau *"standing, arms relaxed
#: at sides"* — cau viet cho NGUOI. Ket qua: anh tham chieu ra con meo dung hai
#: chan, va 19 canh co meo ve ra 19 con meo khac nhau (phim openstory/0002,
#: chu du an: *"luc thi meo 1 loai luc thi ve ra meo khac"*).
#:
#: Chua bang cach dan them mot cau nua vao duoi la chua phan ngon. Goc re la:
#: dang than phai do NGUOI DUNG DAN khai ra, va lenh ve phai doc dung o ay.
_CAU_DANG_THAN = {
    "animal": (" — this is an ORDINARY ANIMAL: draw it in its natural animal "
               "bearing (a four-legged animal stands squarely on all four legs, "
               "a bird perches, a fish swims), never on two legs, never with "
               "human arms or hands, never wearing clothes"),
    "upright-animal": (" — this animal WALKS AND ACTS LIKE A PERSON in this "
                       "story: draw it standing upright on two legs, arms "
                       "relaxed at its sides"),
    "human": "",
}


def _dang_than(c) -> str:
    """`body_mode` da chuan hoa. Thieu thi doan tu mo ta, khong doan tu vai."""
    tho = str(c.get("body_mode") or "").strip().lower().replace("_", "-")
    if tho in _CAU_DANG_THAN:
        return tho
    if tho.startswith("upright"):
        return "upright-animal"
    if tho.startswith("animal"):
        return "animal"
    return ""


def _cau_dang_than(c) -> str:
    """Cau ta dang than cho loi nhac ve chan dung. Khong khai thi tra "" —
    im lang tot hon doan bay, va `DUOI_CHAN_DUNG` van co nhanh du phong."""
    return _CAU_DANG_THAN.get(_dang_than(c), "")


def _lam_lanh_moi_prompt(cast, scenes, bia) -> None:
    """Thay tu bi loc trong MOI prompt se di ra Excel (dan, boi canh, canh, bia)."""
    _don_dan(cast.get("characters") or [])
    for c in cast.get("characters") or []:
        if c.get("sheet_prompt"):
            c["sheet_prompt"] = _don_mo_ta(str(c["sheet_prompt"]))
    for c in list(cast.get("characters") or []) + list(cast.get("locations") or []):
        for k in ("english_prompt", "sheet_prompt", "reference_lock", "location_lock"):
            if c.get(k):
                c[k] = _lam_lanh_prompt(c[k])
    for s in scenes:
        for k in ("img_prompt", "video_prompt", "primary_subject", "primary_action"):
            if s.get(k):
                s[k] = _lam_lanh_prompt(s[k])
    for t in (bia or {}).get("thumbnails") or []:
        if t.get("img_prompt"):
            t["img_prompt"] = _lam_lanh_prompt(t["img_prompt"])


#: Tu khong mang nghia khi so khop ten nhan vat/noi chon voi dan da co.
_TU_RONG = {"the", "a", "an", "of", "and", "or", "his", "her", "their", "its",
            "two", "three", "old", "older", "young", "younger", "who", "with",
            "edge", "side"}

#: Luat cho luot casting khi nv1 DA CO DINH (loai 2): AI chi dung nhan vat
#: PHU tu nv2, khong duoc dinh nghia lai nv1.
_LUAT_NV1_CO_DINH = (
    "The MAIN character `nv1` is ALREADY FIXED and must NOT be redefined — it "
    "is: {mo_ta}. List only the OTHER recurring characters, starting at id "
    "`nv2`, in order of importance.")

#: Luot viet TIEU DE + CHU HOOK + 3 prompt anh bia. Rut tu `8-thumbnail.md` cua
#: kenh va prompt cua D:\\AFFILIATE: anh bia KHONG CO CHU (chu do phan mem chen
#: sau bang font that), chua mot xung dot xa hoi hoac mot bat ngo thi giac.
_KHUON_BIA = """You are an elite YouTube thumbnail prompt writer for faceless
narrated channels. The goal is STOPPING THE SCROLL: curiosity, emotional
discomfort, recognition ("that's me"), social judgment.

## The video (narration, opening and sample)
{transcript}

## Style
{style}
{cast}

## Write
1. `title` — a video title in the SAME language as the narration, at most 70
   characters, honest, curiosity-driven, no clickbait lies.
2. `thumb_text` — the hook the viewer reads on the cover: 2 to 5 SHORT words,
   UPPERCASE, same language as the narration, the emotionally strongest words.
3. Three thumbnail image prompts (16:9), each a DIFFERENT reason to click:
   - `portrait_main` — character close, one clear feeling. Safest.
   - `dramatic_scene` — the most charged moment, character smaller, the
     situation around them carrying the tension.
   - `youtube_ctr` — boldest: one strong symbolic object in front, the
     character reacting behind it. Highest contrast.
   Rules for every prompt: the main character (if any) large, 35–45% of the
   frame, asymmetric composition, background simplified; ONE social conflict
   or visual surprise (a shadow, a reflection, people staring); readable at
   phone size; leave clean NEGATIVE SPACE on one side for the text that will
   be overlaid later. ABSOLUTELY NO TEXT IN THE IMAGE — no letters, no
   numbers, no signs. {fixed_rule} End every prompt with the style tail and
   "no text, no letters, no numbers, no watermark".

## Return JSON only, no commentary
```json
{{"title": "...", "thumb_text": "...",
 "thumbnails": [
   {{"version_desc": "portrait_main", "img_prompt": "..."}},
   {{"version_desc": "dramatic_scene", "img_prompt": "..."}},
   {{"version_desc": "youtube_ctr", "img_prompt": "..."}}
 ]}}
```
"""

#: Luot viet prompt nhac nen Suno — rut tu `9-nhac.md` cua kenh.
_KHUON_NHAC = """Write background music for a narrated video as SUNO prompts.

Suno makes tracks of about 1 minute 45 seconds, so the music can follow the
story instead of sitting flat underneath it: one track per ~{giay_track}
seconds of video. This video is {giay:.0f} seconds long → write exactly
{so_track} track(s) that cover it end to end with no gaps.

## The narration (opening and sample) — read it for the emotional arc
{transcript}

## Visual mood
{style}

## The Suno format — one line, five parts
    [Style]. [Instruments]. [Mood/Emotion]. [Atmosphere]. No vocals, instrumental only.

## Rules
1. Instrumental only. ALWAYS end with "No vocals, instrumental only." — a voice
   in the music fights the narrator.
2. Concrete beats adjectives: name instruments, tempo feel, key or mood.
3. The music sits UNDER the narration: sparse, soft dynamics, no sudden hits.
4. Each track loopable and even — the arc comes from tracks differing from
   each other, not from swells inside one.
5. Consecutive tracks share a family (same instrument palette, same register)
   so the video sounds like one piece, not a playlist.
6. Never name a real artist, band or existing song.

## Return JSON only, no commentary
`start_time`/`end_time` in seconds from the start of the video.
```json
{{"music": [
  {{"music_id": 1, "start_time": 0, "end_time": {giay_track},
   "suno_prompt": "<the one-line Suno prompt>", "mood": "<two or three words>"}}
]}}
```
"""

_KHOA_IN = threading.Lock()


def emit(value: Mapping[str, Any]) -> None:
    # Cac khuc chia canh chay song song nen hai luong co the cung bao tien do.
    # Mot dong JSON bi xen giua chung la mot dong cha khong doc duoc.
    with _KHOA_IN:
        print(json.dumps(dict(value), ensure_ascii=False), flush=True)


def handle(request: Mapping[str, Any], *, enrich_fn: Callable = None,
           chia_fn: Callable = None, cast_fn: Callable = None,
           bia_fn: Callable = None, nhac_fn: Callable = None,
           phim_fn: Callable = None, ke_hoach_fn: Callable = None) -> Mapping[str, Any]:
    inputs = request.get("inputs") if isinstance(request.get("inputs"), dict) else {}
    srt_path = _path(inputs.get("subtitles"), "subtitles")
    context = _optional_json(inputs.get("context"))
    cues = parse_srt(srt_path.read_text(encoding="utf-8-sig"))
    if not cues:
        raise ValueError("SRT khong co dong phu de hop le")
    config = request.get("config") if isinstance(request.get("config"), dict) else {}
    engine = str(config.get("engine") or "veo3")
    model = str(config.get("model") or "claude-sonnet-5")
    nhat_quan = bool(config.get("nhat_quan_nhan_vat", True))
    che_do = str(config.get("che_do_ke") or context.get("story_mode") or CHE_DO_TU_XAY)
    lam_bia = bool(config.get("thumbnail", True))
    lam_nhac = bool(config.get("nhac", True))
    # Che do bom tay (bai kiem dua `chia_fn` ma khong dua ham cua luot nao) thi
    # luot ay KHONG tu goi mang — cung luat voi casting.
    bom_tay = chia_fn is not None

    goi = _hop_goi(request, model)
    try:
        # ═══ DAO DIEN (loai 2, 3): doc phim TRUOC casting, ke hoach TRUOC chia canh ═══
        dao_dien = (nhat_quan and che_do in (CHE_DO_NV_BOI_CANH, CHE_DO_TU_XAY)
                    and not (bom_tay and phim_fn is None))
        phim = _doc_phim(goi, cues, context, phim_fn=phim_fn) if dao_dien else {}
        boi_canh_cast = dict(context)
        if phim:
            boi_canh_cast["film_analysis"] = {
                k: phim.get(k) for k in ("genre", "arc", "context_lock",
                                          "characters_mentioned", "locations_mentioned")}

        # Dung dan nhan vat + boi canh + phong cach TRUOC khi cat canh, de moi
        # canh sau deu nhac dung mot dan va giu mot tong — y het tab Tu dong.
        cast = _dan_nhan_vat(goi, cues, boi_canh_cast, nhat_quan=nhat_quan,
                             che_do=che_do, cast_fn=cast_fn, chia_fn=chia_fn)
        _them_prompt_tham_chieu(cast, context)
        cast_style = _khoi_cast_style(cast)
        nhan_vat_chinh = cast["characters"][0]["id"] if cast["characters"] else ""

        ke_hoach = (_ke_hoach_dao_dien(goi, cues, phim, cast, engine=engine,
                                       ke_hoach_fn=ke_hoach_fn,
                                       nhip=_nhip_canh(context, engine))
                    if phim.get("segments") else [])

        # ═══ ANH BIA + NHAC CHAY SONG SONG VOI CHIA CANH ═══
        #
        # Hai luot nay chi can loi doc + dan nhan vat, khong can bang canh —
        # de sau chia canh la doi them ~25 giay khong vi ly do gi (do 24/08/2026:
        # bia 16s + nhac 9s noi duoi). Chu du an: *"chia de api song song
        # nhieu luong cho nhanh"*.
        from concurrent.futures import ThreadPoolExecutor  # noqa: PLC0415

        trong_bia = {"title": "", "thumb_text": "", "thumbnails": []}
        with ThreadPoolExecutor(max_workers=2) as phu:
            viec_bia = (None if (not lam_bia or (bom_tay and bia_fn is None))
                        else phu.submit(_anh_bia, goi, cues, context, cast, bia_fn=bia_fn))
            viec_nhac = (None if (not lam_nhac or (bom_tay and nhac_fn is None))
                         else phu.submit(_nhac_nen, goi, cues, context, cast, nhac_fn=nhac_fn))
            scenes, cach = _canh(cues, engine=engine, context=context, goi=goi,
                                 chia_fn=chia_fn, enrich_fn=enrich_fn,
                                 cast_style=cast_style, nhan_vat_chinh=nhan_vat_chinh,
                                 ke_hoach=ke_hoach)
            bia = viec_bia.result() if viec_bia is not None else trong_bia
            nhac = viec_nhac.result() if viec_nhac is not None else []
        if ke_hoach:
            doi = _ep_theo_ke_hoach(scenes, ke_hoach)
            emit({"type": "event", "event": "progress", "progress": 0.0,
                  "message": "Ep canh theo ke hoach dao dien: sua boi canh {0} canh, nhan vat {1} "
                             "canh; ca video doi boi canh {2} lan / {3} canh.".format(
                                 doi["boi_canh"], doi["nhan_vat"], _so_lan_doi_boi_canh(scenes),
                                 len(scenes))})
        if chia_fn is None:
            # ═══ PHÁ CẢNH LẶP ═══ (chỉ khi chạy thật; bài kiểm bơm chia_fn không gọi AI)
            so_lap = _pha_lap_canh(scenes, goi)
            if so_lap:
                emit({"type": "event", "event": "progress", "progress": 0.0,
                      "message": "Doi khung hinh {0} canh lap voi canh truoc.".format(so_lap)})
        _gan_reference_files(scenes, cast["characters"], cast.get("locations") or [])
        _validate_coverage(cues, scenes)
        _lam_lanh_moi_prompt(cast, scenes, bia)
        _khoa_nhan_dang(scenes, cast["characters"], cast.get("locations") or [])
        _ep_phong_cach(scenes, cast.get("style") or {}, emit)

        workspace = Path(str(request.get("workspace") or "")).resolve(); workspace.mkdir(parents=True, exist_ok=True)
        manifest = {"schema_version": 1, "project_id": str(request.get("workflow_id") or "project"),
                    "source": {"subtitle_artifact_id": _artifact_id(inputs.get("subtitles")),
                               "content_artifact_id": _artifact_id(inputs.get("context"))},
                    "settings": {"model": model, "cach_chia": cach, "style": cast["style"],
                                 "che_do_ke": che_do},
                    "characters": cast["characters"], "locations": cast.get("locations") or [],
                    "story": phim, "director_plan": ke_hoach,
                    "title": bia["title"], "thumb_text": bia["thumb_text"],
                    "thumbnails": bia["thumbnails"], "music": nhac,
                    "scenes": scenes, "coverage": {"cue_count": len(cues), "covered": len(cues), "percent": 100}}
        workbook = workspace / "scene-prompts.xlsx"
        render_workbook(workbook, manifest)
        emit({"type": "event", "event": "progress", "progress": 1.0, "message": "Da tao workbook"})
        return {"scenes": {"json": manifest, "filename": "scene-manifest.json",
                           "metadata": {"scene_count": len(scenes), "coverage": 100,
                                        "cach_chia": cach}},
                "workbook": {"path": workbook.name,
                             "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             "metadata": {"scene_count": len(scenes)}}}
    finally:
        close = getattr(goi, "close", None)
        if callable(close): close()


def _doc_phim(goi, cues, context, *, phim_fn=None) -> Dict[str, Any]:
    """Buoc A: mot luot AI doc ca loi doc → the loai, cung truyen, the gioi hinh, cac man.

    Best-effort: hong thi tra `{}` va noi ra — khong co man thi ve duong cu
    (casting + chia canh), khong giet luot.
    """
    try:
        if phim_fn is not None:
            raw = phim_fn(cues, context)
        else:
            # Ca loi doc, khong cat: cat la mat nhan vat/noi chon o phan sau.
            loi_doc = bang_phu_de(cues)
            boi_canh = _boi_canh_chu(context, "(none)")
            giay = _giay_video(cues)
            raw = loc_json(goi(_KHUON_PHIM.format(transcript=loi_doc, context=boi_canh,
                                                  phut="{0:.1f}".format(giay / 60.0),
                                                  so_man=_so_man_goi_y(giay)),
                               "phim"))
        phim = _sach_phim(raw, cues)
    except Exception as loi:  # noqa: BLE001 — dao dien la phu, hong thi ve duong cu
        emit({"type": "event", "event": "progress", "progress": 0.0,
              "message": "Chua doc duoc phim de chia man ({0}). Chia canh theo "
                         "duong cu.".format(str(loi)[:120])})
        return {}
    emit({"type": "event", "event": "progress", "progress": 0.0,
          "message": "Da doc phim: {0} man, cung truyen {1}.".format(
              len(phim.get("segments") or []), phim.get("arc") or "?")})
    return phim


def _sach_phim(raw, cues) -> Dict[str, Any]:
    """Chuan hoa: man noi lien nhau, phu het dong SRT (AI khong dem gioi)."""
    if not isinstance(raw, Mapping):
        return {}
    dau, cuoi = int(cues[0]["index"]), int(cues[-1]["index"])
    ds: List[Dict[str, Any]] = []
    ke_tiep = dau
    for item in raw.get("segments") or []:
        if not isinstance(item, Mapping):
            continue
        try:
            den = int(item.get("srt_to"))
        except (TypeError, ValueError):
            continue
        den = max(ke_tiep, min(den, cuoi))
        if ke_tiep > cuoi:
            break
        ds.append({"segment_id": len(ds) + 1, "name": str(item.get("name") or ""),
                   "message": str(item.get("message") or ""),
                   "emotion": str(item.get("emotion") or ""),
                   "motif": str(item.get("motif") or ""),
                   "srt_from": ke_tiep, "srt_to": den})
        ke_tiep = den + 1
    if ds and ke_tiep <= cuoi:
        ds[-1]["srt_to"] = cuoi
    if not ds:
        return {}
    return {"genre": str(raw.get("genre") or ""), "arc": str(raw.get("arc") or ""),
            "context_lock": str(raw.get("context_lock") or ""),
            "characters_mentioned": [str(x) for x in raw.get("characters_mentioned") or []],
            "locations_mentioned": [str(x) for x in raw.get("locations_mentioned") or []],
            "segments": ds}


def _them_prompt_tham_chieu(cast, context) -> None:
    """Moi nhan vat (khong co dinh) va boi canh co `sheet_prompt` — prompt tao
    ANH THAM CHIEU (chan dung full-body nen trang / anh thiet lap boi canh).

    Anh nay tao ngay sau khi co Excel (chu du an: *"tu tao luon khi bam Tao
    prompt"*) va duoc gan vao `reference_files` cua tung canh — nhan vat va noi
    chon giu nguyen qua ca tram canh, thu lam video "nhu mot bo phim".
    """
    style = str(context.get("visual_style_directive") or "").strip()
    duoi_style = (" Style: " + " ".join(style.split())[:400]) if style else ""
    for c in cast.get("characters") or []:
        if c.get("co_dinh"):
            continue
        # Giai đoạn sau vẽ KÈM ảnh giai đoạn đầu -> nói rõ "cùng một cá thể".
        them = _DUOI_GIAI_DOAN if str(c.get("goc_id") or "") not in ("", c.get("id")) else ""
        c["sheet_prompt"] = (c["english_prompt"] + _cau_dang_than(c)
                             + _DUOI_CHAN_DUNG + them + duoi_style)
    for l in cast.get("locations") or []:
        l["sheet_prompt"] = l["english_prompt"] + _DUOI_BOI_CANH + duoi_style


def _ke_hoach_dao_dien(goi, cues, phim, cast, *, engine, ke_hoach_fn=None,
                       nhip=None) -> List[Dict[str, Any]]:
    """Buoc C: moi man mot luot AI, chay song song, tra ve danh sach beat da canh lai.

    Man nao hong thi bo qua man do (noi ra) — cac man con lai van co ke hoach.
    """
    from concurrent.futures import ThreadPoolExecutor  # noqa: PLC0415

    theo_so = {int(c["index"]): c for c in cues}
    segs = phim.get("segments") or []
    cast_khoi = _khoi_cast_style(cast)
    # `nhip` = (sàn, trần một ý) đọc từ khuôn; không có thì mặc định. Beat dài
    # hơn clip engine vẫn được `_sach_ke_hoach` tách theo trần engine.
    tran_engine = float(max_seconds_for(engine))
    san, tran = nhip if nhip else (float(MIN_GIAY_CANH), tran_engine)

    def mot_man(seg):
        dong = [theo_so[i] for i in range(int(seg["srt_from"]), int(seg["srt_to"]) + 1)
                if i in theo_so]
        if not dong:
            return []
        try:
            if ke_hoach_fn is not None:
                raw = ke_hoach_fn(seg, dong, cast)
            else:
                loi_nhac = _KHUON_KE_HOACH.format(
                    so=seg["segment_id"], tong=len(segs), ten=seg["name"],
                    message=seg["message"], emotion=seg["emotion"],
                    motif=seg["motif"], context_lock=phim.get("context_lock") or "",
                    cast=cast_khoi or "(no recurring cast)", srt=bang_phu_de(dong),
                    min_sec=int(san), max_sec=int(tran))
                raw = loc_json(goi(loi_nhac, "man-{0}".format(seg["segment_id"])))
            return _sach_ke_hoach(raw, seg, dong, tran=tran_engine, san=float(san))
        except Exception as loi:  # noqa: BLE001 — mot man hong khong giet ca phim
            emit({"type": "event", "event": "progress", "progress": 0.0,
                  "message": "Man {0} chua co ke hoach dao dien ({1}).".format(
                      seg["segment_id"], str(loi)[:100])})
            return []

    with ThreadPoolExecutor(max_workers=min(8, max(1, len(segs)))) as bo:
        ket = list(bo.map(mot_man, segs))
    ra = [b for man in ket for b in man]
    emit({"type": "event", "event": "progress", "progress": 0.0,
          "message": "Ke hoach dao dien: {0} beat cho {1} man.".format(len(ra), len(segs))})
    return ra


def _sach_ke_hoach(raw, seg, dong, tran: float = 8.0,
                   san: float = MIN_GIAY_CANH) -> List[Dict[str, Any]]:
    """Beat noi lien nhau trong man, phu het dong cua man; danh so beat.

    `tran`: tran giay moi canh cua engine — beat dai hon thi tach tai ranh
    gioi dong (xem `core.chia_canh.tach_dai`).
    """
    ds = raw.get("beats") if isinstance(raw, Mapping) else raw
    if not isinstance(ds, list):
        return []
    dau, cuoi = int(dong[0]["index"]), int(dong[-1]["index"])
    ra: List[Dict[str, Any]] = []
    ke_tiep = dau
    for item in ds:
        if not isinstance(item, Mapping):
            continue
        try:
            den = int(item.get("srt_to"))
        except (TypeError, ValueError):
            continue
        if ke_tiep > cuoi:
            break
        den = max(ke_tiep, min(den, cuoi))
        ra.append({"segment_id": seg["segment_id"], "beat": len(ra) + 1,
                   "srt_from": ke_tiep, "srt_to": den,
                   "purpose": str(item.get("purpose") or ""),
                   "characters": str(item.get("characters") or ""),
                   "location": str(item.get("location") or ""),
                   "shot_size": str(item.get("shot_size") or ""),
                   "camera": str(item.get("camera") or ""),
                   "element_motion": str(item.get("element_motion") or ""),
                   "emotion": str(item.get("emotion") or ""),
                   "motif": str(item.get("motif") or ""), "status": "planned"})
        ke_tiep = den + 1
    if ra and ke_tiep <= cuoi:
        ra[-1]["srt_to"] = cuoi
    # Beat ngan hon san thi gop — AI hay de "moi dong mot beat" (do 24/08/2026:
    # 23 beat / 24 dong, co beat 0,7 giay). Cung ham voi khau chia canh.
    from core.chia_canh import gop_ngan, tach_dai  # noqa: PLC0415

    theo_so = {int(c["index"]): c for c in dong}
    ra = gop_ngan(ra, theo_so, "srt_from", "srt_to", float(san))
    # Beat dai qua tran thi tach tai ranh gioi dong — moi phan mot hinh rieng
    # thay vi khau chia canh cat doi va dung chung mot tam anh.
    ra = tach_dai(ra, theo_so, "srt_from", "srt_to", float(tran))
    for so, b in enumerate(ra, 1):
        b["beat"] = so
    return ra


def _khoi_ke_hoach(ke_hoach, khuc) -> str:
    """Khoi `<<DIRECTOR_PLAN>>` cho MOT khuc: cac beat cham vao dong cua khuc ay."""
    if not ke_hoach or not khuc:
        return ""
    dau, cuoi = int(khuc[0]["index"]), int(khuc[-1]["index"])
    dong: List[str] = []
    for b in ke_hoach:
        if int(b["srt_to"]) < dau or int(b["srt_from"]) > cuoi:
            continue
        dong.append("- lines {0}-{1}: {2} | who: {3} | where: {4} | {5}, {6} | "
                    "change: {7} | emotion: {8}{9}".format(
                        b["srt_from"], b["srt_to"], b["purpose"] or "beat",
                        b["characters"] or "nobody", b["location"] or "free",
                        b["shot_size"] or "?", b["camera"] or "?",
                        b["element_motion"] or "?", b["emotion"] or "?",
                        " | motif: " + b["motif"] if b["motif"] else ""))
    if not dong:
        return ""
    return ("## DIRECTOR'S PLAN for these lines — follow it\n"
            "Each beat below is one scene (or a group of scenes). Keep the shot "
            "size and camera it names, put the named ids in `characters_used` / "
            "`location_used`, describe a named location EXACTLY as in RECURRING "
            "LOCATIONS, and make the named change happen in the video prompt.\n"
            + "\n".join(dong) + "\n")


def _giay_video(cues) -> float:
    try:
        return max(float(c.get("end") or 0) for c in cues)
    except (TypeError, ValueError):
        return 0.0


def _so_man_goi_y(giay: float) -> int:
    """So man goi y cho buoc doc phim: ~1 man / 75 giay, it nhat 3, khong tran.

    Chu du an 25/08/2026: *"dung co gi cung o day… kich ban ngan dai khac
    nhau"* — ban cu ghi chet "3–8 man" nen video 30 phut cung chi 8 man.
    """
    return max(3, int(round(float(giay or 0) / 75.0)))


def _loi_doc_mau(cues, toi_da: int = BIA_KY_TU) -> str:
    """Loi doc dua cho luot anh bia / nhac: ca bai (toi_da <= 0), hoac lay mau."""
    chu = " ".join(str(c.get("text") or "").strip() for c in cues)
    if toi_da <= 0 or len(chu) <= toi_da:
        return chu
    dau = int(toi_da * 0.6)
    con = chu[dau:]
    buoc = max(1, len(con) // max(1, toi_da - dau))
    return chu[:dau] + " … " + " ".join(con[i:i + 200] for i in range(0, len(con), buoc * 200))[:toi_da - dau]


def _nhan_vat_co_dinh(context) -> Dict[str, Any]:
    """`fixed_character` trong context (loai 1, 2) — luon co id nv1 va anh nv1.png."""
    tho = context.get("fixed_character") if isinstance(context, Mapping) else None
    tho = dict(tho) if isinstance(tho, Mapping) else {}
    mo_ta = str(tho.get("english_prompt") or "").strip() or (
        "the reference character exactly as in nv1.png — same face, hair, "
        "clothes and colours in every scene")
    return {"id": "nv1", "role": str(tho.get("role") or "protagonist"),
            "name": str(tho.get("name") or "Nhan vat chinh"),
            "english_prompt": mo_ta,
            "reference_lock": str(tho.get("reference_lock") or
                                  "use nv1.png as the exact identity anchor"),
            "image_file": "nv1.png", "co_dinh": True,
            "gender": str(tho.get("gender") or ""), "age": str(tho.get("age") or ""),
            "notes": str(tho.get("notes") or "")}


def _luat_nv1(cast) -> str:
    """Cau nhac 'chi goi la nv1 (nv1.png), khong ta ngoai hinh' khi co nv1 co dinh."""
    for c in cast.get("characters") or []:
        if c.get("co_dinh"):
            return ("The main character has a reference image attached at "
                    "generation time: refer to it ONLY as `nv1 (nv1.png)`; NEVER "
                    "describe its face, hair, skin, clothes or colours — only "
                    "pose, gesture, expression.")
    return ""


def _anh_bia(goi, cues, context, cast, *, bia_fn=None) -> Dict[str, Any]:
    """Mot luot AI: tieu de + chu hook + 3 prompt anh bia. Best-effort."""
    trong = {"title": "", "thumb_text": "", "thumbnails": []}
    try:
        if bia_fn is not None:
            raw = bia_fn(cues, context, cast)
        else:
            style = str(context.get("visual_style_directive") or "") or json.dumps(
                cast.get("style") or {}, ensure_ascii=False)
            loi_nhac = _KHUON_BIA.format(
                transcript=_loi_doc_mau(cues), style=style or "(pick one look and hold it)",
                cast=_khoi_cast_style(cast), fixed_rule=_luat_nv1(cast))
            raw = loc_json(goi(loi_nhac, "bia"))
        ra = _sach_bia(raw, cast)
    except Exception as loi:  # noqa: BLE001 — anh bia la phu, hong thi bo trong
        emit({"type": "event", "event": "progress", "progress": 0.0,
              "message": "Chua viet duoc prompt anh bia ({0}). Sheet thumbnail "
                         "de trong.".format(str(loi)[:120])})
        return trong
    emit({"type": "event", "event": "progress", "progress": 0.0,
          "message": "Da viet {0} prompt anh bia.".format(len(ra["thumbnails"]))})
    return ra


def _sach_bia(raw, cast) -> Dict[str, Any]:
    if not isinstance(raw, Mapping):
        return {"title": "", "thumb_text": "", "thumbnails": []}
    nv1 = "nv1" if any(c.get("id") == "nv1" for c in cast.get("characters") or []) else ""
    ds = []
    for so, item in enumerate(raw.get("thumbnails") or [], 1):
        if not isinstance(item, Mapping):
            continue
        prompt = str(item.get("img_prompt") or "").strip()
        if not prompt:
            continue
        ds.append({"thumb_id": so,
                   "version_desc": str(item.get("version_desc") or "version_{0}".format(so)),
                   "img_prompt": prompt, "characters_used": nv1,
                   "location_used": "",
                   "reference_files": json.dumps(["nv1.png"]) if nv1 else "",
                   "img_path": "", "status_img": "pending"})
    return {"title": str(raw.get("title") or "").strip(),
            "thumb_text": str(raw.get("thumb_text") or "").strip(),
            "thumbnails": ds}


def _nhac_nen(goi, cues, context, cast, *, nhac_fn=None) -> List[Dict[str, Any]]:
    """Mot luot AI: prompt Suno cho tung track ~105 giay. Best-effort."""
    giay = _giay_video(cues)
    so_track = max(1, int(-(-giay // GIAY_MOI_TRACK))) if giay > 0 else 1
    try:
        if nhac_fn is not None:
            raw = nhac_fn(cues, context, cast)
        else:
            style = str(context.get("visual_style_directive") or "") or json.dumps(
                cast.get("style") or {}, ensure_ascii=False)
            loi_nhac = _KHUON_NHAC.format(
                giay_track=GIAY_MOI_TRACK, giay=giay, so_track=so_track,
                transcript=_loi_doc_mau(cues), style=style or "(calm, warm)")
            raw = loc_json(goi(loi_nhac, "nhac"))
        ra = _sach_nhac(raw, giay)
    except Exception as loi:  # noqa: BLE001 — nhac la phu, hong thi bo trong
        emit({"type": "event", "event": "progress", "progress": 0.0,
              "message": "Chua viet duoc prompt nhac ({0}). Sheet music de "
                         "trong.".format(str(loi)[:120])})
        return []
    emit({"type": "event", "event": "progress", "progress": 0.0,
          "message": "Da viet {0} prompt nhac Suno.".format(len(ra))})
    return ra


def _sach_nhac(raw, giay: float) -> List[Dict[str, Any]]:
    """Chuan hoa track: danh so lai, thoi gian noi lien nhau va phu het video."""
    ds = raw.get("music") if isinstance(raw, Mapping) else raw
    if not isinstance(ds, list):
        return []
    ra: List[Dict[str, Any]] = []
    truoc = 0.0
    for item in ds:
        if not isinstance(item, Mapping):
            continue
        prompt = str(item.get("suno_prompt") or "").strip()
        if not prompt:
            continue
        try:
            ket = float(item.get("end_time"))
        except (TypeError, ValueError):
            ket = truoc + GIAY_MOI_TRACK
        ket = max(truoc + 1.0, min(ket, giay) if giay > 0 else ket)
        ra.append({"music_id": len(ra) + 1, "start_time": round(truoc, 1),
                   "end_time": round(ket, 1), "suno_prompt": prompt,
                   "mood": str(item.get("mood") or ""), "status": "pending"})
        truoc = ket
    if ra and giay > 0:
        ra[-1]["end_time"] = round(giay, 1)
    return ra


#: Hai cach cat canh. Ten nay di vao manifest, nen no la mot phan cua ket qua.
THEO_NGHIA = "theo-noi-dung"
THEO_DONG_HO = "theo-dong-ho"


#: Khuôn chia cảnh của KÊNH (hay khách sửa ở Nâng cao của Prompt Visuals) phải
#: có đủ các chỗ trống này mới dùng được ở đường đạo diễn (dàn nhân vật, kế
#: hoạch, bối cảnh, phụ đề). Danh sách sống ở `core.prompt_visuals` — ô Nâng
#: cao kiểm trước khi gửi bằng đúng danh sách ấy.
_CHO_TRONG_KHUON_CHIA = CHO_TRONG_KHUON_CHIA


def _nhip_canh(context, engine) -> "tuple[float, float]":
    """(sàn, trần) giây MỘT CẢNH, đọc từ khối PACING của chính khuôn chia cảnh
    (`_khuon_chia(context)`) — khách sửa số trong prompt là đổi mạch. Không có
    khối thì mặc định 3 giây và trần engine. Trần cảnh được phép lớn hơn clip
    engine (30 giây một ý): `_canh_theo_nghia` vẫn cắt theo trần engine, mỗi
    phần một hình từ góc máy khác."""
    tran_engine = float(max_seconds_for(engine))
    cap = nhip_tu_khuon(_khuon_chia(context))
    if cap is None:
        return float(MIN_GIAY_CANH), tran_engine
    san, tran = cap
    return min(san, tran_engine - 1.0) if tran_engine > 1.0 else san, tran


#: Trần chữ của khối context đưa vào lời nhắc. Trước 26/08/2026 là 30.000 ký tự
#: — kịch bản dài (truyện 25 phút ≈ 20.000 ký tự) cộng phong cách là bị cắt mất
#: đoạn cuối TRƯỚC khi tuyển vai, nhân vật xuất hiện muộn không có mặt trong dàn.
#: Chủ dự án: "đừng giới hạn, nó do nguồn đầu vào". 160.000 ký tự ≈ 40 phút đọc,
#: vẫn trong cửa sổ của mô hình.
TRAN_CONTEXT_CHU = 160000
#: Khoá trong context KHÔNG phải nội dung — không đưa vào lời nhắc.
_KHOA_CONTEXT_BO = ("storyboard_template",)


def _boi_canh_chu(context, mac_dinh: str = "") -> str:
    """Context → chuỗi JSON cho lời nhắc: bỏ khuôn chia cảnh, trần `TRAN_CONTEXT_CHU`."""
    if not context:
        return mac_dinh
    if isinstance(context, Mapping):
        context = {k: v for k, v in context.items() if k not in _KHOA_CONTEXT_BO}
    return json.dumps(context, ensure_ascii=False)[:TRAN_CONTEXT_CHU]


def _khuon_chia(context) -> str:
    """Khuôn chia cảnh: của kênh (`context["storyboard_template"]`) nếu đủ chỗ trống, không thì mặc định.

    Khuôn mặc định viết cho video người lớn kiểu "ẩn dụ hình" (đồng hồ xoáy,
    kính vỡ) — sai thể loại cho truyện cổ tích trẻ em, nơi mỗi cảnh phải minh
    hoạ ĐÚNG câu kể (chủ dự án 25/08/2026: "cảnh lặp, không minh hoạ nội
    dung"). Kênh nào có `prompt/7-canh.md` đủ chỗ trống thì dùng khuôn của
    kênh: thể loại nằm ở kênh, không cứng trong mã.
    """
    khuon = ""
    if isinstance(context, Mapping):
        khuon = str(context.get("storyboard_template") or "")
    if khuon.strip() and all(ct in khuon for ct in _CHO_TRONG_KHUON_CHIA):
        return khuon
    return KHUON_MAC_DINH


def _canh(cues, *, engine, context, goi, chia_fn=None, enrich_fn=None,
          cast_style="", nhan_vat_chinh="", ke_hoach=None):
    """Cat canh theo nghia; khong duoc thi lui ve dong ho va **noi ra**.

    Duong lui phai con, vi khong co no thi mot cu 503 la khach khong nhan duoc
    gi ca. Nhung lui im lang thi con te hon: khach mo file Excel ra, thay du
    111 canh va du loi nhac, khong the nao biet rang canh da bi cat giua cau.
    """
    chia = chia_fn or _bo_chia(goi, context, engine, cast_style, ke_hoach or [])
    try:
        return (_canh_theo_nghia(cues, chia, engine, nhan_vat_chinh,
                                 nhip=_nhip_canh(context, engine)), THEO_NGHIA)
    except Exception as loi:  # noqa: BLE001 — het duong nay thi con duong lui
        emit({"type": "event", "event": "progress", "progress": 0.0,
              "message": "Chua chia duoc canh theo noi dung ({0}). Toi tam cat "
                         "theo dong ho — canh se deu nhau ve do dai nhung co "
                         "the cat giua mot y.".format(str(loi)[:120])})
    scenes = group_scenes(cues, engine=engine)
    enrich = enrich_fn or _bo_enrich(goi)
    for offset in range(0, len(scenes), 20):
        batch = scenes[offset:offset + 20]
        emit({"type": "event", "event": "progress", "progress": offset / max(1, len(scenes)),
              "message": "Dang tao prompt canh {0}-{1}".format(offset + 1, offset + len(batch))})
        _apply_creative(batch, enrich(batch, context))
    return scenes, THEO_DONG_HO


def _canh_theo_nghia(cues, chia, engine, nhan_vat_chinh="", nhip=None) -> List[Dict[str, Any]]:
    """AI tu chia canh theo nghia va viet luon loi nhac cho tung canh.

    Mot luot goi lam ca hai viec, khong phai hai luot: canh duoc cat o dau la
    thu chi AI biet, nen bat no chia xong roi hoi lai "canh nay ta cai gi" la
    tra tien hai lan cho cung mot doan chu.
    """
    # Trần CẮT là của engine; trần trong lời nhắc (một ý dài bao nhiêu) nằm ở
    # `nhip` và có thể lớn hơn — cảnh dài được quay thành nhiều góc máy.
    san = nhip[0] if nhip else float(MIN_GIAY_CANH)
    ra: List[Dict[str, Any]] = []
    for canh in chia_theo_nghia(cues, chia, tran=max_seconds_for(engine), san=san,
                                nhan_vat_mac_dinh=nhan_vat_chinh,
                                duoi=DUOI_CAM, giay_moi_khuc=GIAY_MOI_KHUC):
        scene = {key: "" for key in SCENE_COLUMNS}
        for ten, gia_tri in canh.items():
            if ten in scene:
                scene[ten] = gia_tri
        scene["prompt_json"] = json.dumps({k: scene[k] for k in CREATIVE}, ensure_ascii=False)
        scene["srt_indices"] = list(canh.get("srt_indices") or ())
        ra.append(scene)
    return ra


def group_scenes(cues: Sequence[Mapping[str, Any]], *, engine: str = "veo3"):
    """DUONG LUI: cat theo dong ho khi khong goi duoc AI.

    Hai buoc, khong gop lam mot duoc:

    1. Gop cac dong phu de lien nhau — chi chot duoc o ranh gioi GIUA hai dong.
    2. Ep tran — mot dong phu de don le dai hon tran thi khong co ranh gioi nao
       ben trong de chot, phai cat khoang thoi gian cua chinh no.
    """
    maximum = max_seconds_for(engine)
    groups = group_cues(cues, target=target_seconds_for(engine), maximum=maximum)
    spans = [{"start": float(group[0]["start"]), "end": float(group[-1]["end"]),
              "text": " ".join(str(cue["text"]) for cue in group),
              "srt_indices": [cue["index"] for cue in group]} for group in groups]
    scenes = []
    for index, span in enumerate(enforce_max_duration(spans, maximum), 1):
        scene = {key: "" for key in SCENE_COLUMNS}
        scene.update({
            "scene_id": index,
            "srt_start": clock(span["start"]), "srt_end": clock(span["end"]),
            "duration": span["duration"], "planned_duration": span["duration"],
            "srt_text": span["text"],
            "status_img": "pending", "status_vid": "pending",
            "srt_indices": list(span["srt_indices"]),
        })
        if int(span.get("split_total", 1)) > 1:
            # Cac phan cat ra tu cung mot khoang deu mang mot `segment_id`, de
            # khau dung video biet chung la mot cau dang duoc doc lien mach chu
            # khong phai hai y roi nhau — no se khong chen chuyen canh giua chung.
            scene["segment_id"] = "seg{0}".format(index - int(span["split_part"]) + 1)
        scenes.append(scene)
    return scenes


def _hop_goi(request: Mapping[str, Any], model: str):
    """Mot cua duy nhat de goi AI, dung cho ca chia canh lan viet loi nhac.

    Di qua `core.goi_van_ban` chu khong tu goi `/v1/chat/completions`: do la
    cho biet doi khi may chu tra 503 *"ban KHONG bi tru tien, thu lai sau 15
    giay"* — ban chep tay khong biet, va mot cu 503 la mat ca luot chay.

    Client dung tren **luot goi dau tien**, khong phai luc dung ham. Nho vay
    bai kiem dua san ham goi thi khong doi khoa API va khong cham mang.
    """
    hop: Dict[str, Any] = {"client": None}

    def khach():
        if hop["client"] is None:
            api_key = os.environ.get("SHOPAPI_API_KEY", "").strip()
            if not api_key:
                raise ValueError("Thieu SHOPAPI_API_KEY")
            sys.path.insert(0, str(_STUDIO / "_sdk"))
            from shopapi import ShopAPI
            # ═══ MOT LOI GOI TREO KHONG DUOC NGON CA LUOT ═══
            #
            # Do 24/08/2026 (180 giay tieng Nhat, loai 3): khuc 1 xong sau 44
            # giay, khuc 2 TREO — tool cho theo thoi gian mac dinh cua SDK, roi
            # `goi_kien_nhan` doi tiep, tong hon 17 phut, va tool `prompt.workbook`
            # bi giet o tran 1200 giay: mat ca luot du moi luot goi truoc do da
            # tra tien. Mot khuc binh thuong het duoi mot phut; cho toi 4 phut la
            # du rong, qua do la treo — cat, doi khoa moi, goi lai.
            hop["client"] = ShopAPI(
                api_key=api_key,
                base_url=os.environ.get("SHOPAPI_BASE_URL", "https://api.shopapi.vn"),
                timeout=240.0,
                default_headers={"X-ShopAPI-Client": "shopapi-tool-builder"})
        return hop["client"]

    def ghi(dong: str) -> None:
        # Cau cho/thu lai cua `goi_van_ban` len man hinh — khach thay "may chu
        # truc trac tam, thu lai sau 30s" thay vi mot thanh tien do dung im.
        emit({"type": "event", "event": "progress", "progress": 0.0,
              "message": str(dong)[:160]})

    def goi(loi_nhac: str, phan_khoa: str) -> str:
        # Khoa co dinh theo (lan chay, nut, viec): mat phan hoi giua chung thi
        # hoi lai dung khoa ay se nhan lai bai da tra tien, khong tra lan hai.
        return goi_van_ban(khach(), [{"role": "user", "content": loi_nhac}],
                           mo_hinh=model, toi_da_token=TOKEN_CANH, on_log=ghi,
                           khoa="{0}:{1}:{2}".format(request.get("run_id", "run"),
                                                     request.get("node_id", "workbook"),
                                                     phan_khoa))

    def dong():
        if hop["client"] is not None:
            hop["client"].close()

    goi.close = dong
    return goi


def _bo_chia(goi, context, engine, cast_style="", ke_hoach=None) -> Callable:
    """Ham hoi AI chia mot khuc phu de. `core.chia_canh` lo phan con lai.

    `cast_style` la khoi chu "dan nhan vat + phong cach" da dung san o
    `_dan_nhan_vat`. Rong thi `dien_khuon` xoa sach `<<CAST_STYLE>>`, template
    ve dung nhu cu (khong bia nhan vat) — dung hanh vi PV truoc day.
    `ke_hoach` (loai 2, 3): beat cua dao dien; moi khuc chi nhan beat cham
    vao dong cua no (xem `_khoi_ke_hoach`).
    """
    san, tran = _nhip_canh(context, engine)
    clip = float(max_seconds_for(engine))
    boi_canh = _boi_canh_chu(context, "")
    xong = {"value": 0}

    def chia(khuc, thu_tu, tong_khuc):
        # Vi tri khuc di vao loi nhac: "canh dau la cu hook" chi dung o khuc 1,
        # be nguyen sang khuc 5 la video mo bai nam lan (xem core/auto_khau).
        loi_nhac = loi_nhac_chia(_khuon_chia(context), khuc, tran, {
            "CONTEXT": boi_canh, "CAST_STYLE": cast_style,
            "DIRECTOR_PLAN": _khoi_ke_hoach(ke_hoach, khuc),
            "KHUC_THU": thu_tu + 1, "TONG_KHUC": tong_khuc,
            "LA_KHUC_DAU": "yes" if thu_tu == 0 else "no",
            "TY_LE_KHUNG": "16:9 horizontal",
        }, san=san, clip=clip)
        tra = goi(loi_nhac, "chia-{0}".format(khuc[0]["index"]))
        goi_ve = loc_json(tra)
        ds = goi_ve.get("scenes") if isinstance(goi_ve, dict) else goi_ve
        if not isinstance(ds, list) or not ds:
            raise ValueError("AI khong tra ve danh sach `scenes`")
        xong["value"] += 1
        emit({"type": "event", "event": "progress",
              "progress": xong["value"] / max(1, tong_khuc),
              "message": "Da chia canh theo noi dung: khuc {0}/{1}".format(
                  xong["value"], tong_khuc)})
        return ds

    return chia


def _dan_nhan_vat(goi, cues, context, *, nhat_quan=True, che_do=CHE_DO_TU_XAY,
                  cast_fn=None, chia_fn=None):
    """Mot luot AI doc ca loi doc roi tra ve dan nhan vat + boi canh + phong cach.

    Ba che do (xem `CHE_DO_*`): loai 1 khong goi AI — dan chi co nv1 co dinh;
    loai 2 goi AI dung nhan vat PHU (nv2..) + boi canh, roi ghep nv1 co dinh
    len dau; loai 3 goi AI dung ca dan lan boi canh.

    Best-effort: casting hong (503, JSON rac) thi ve dan toi thieu va **noi
    ra** — khong duoc giet ca luot, vi con canh van chia duoc theo cach cu.

    Chi goi mang o luong that (khong co `chia_fn` bom vao, va cong bat). Bai
    kiem bom `cast_fn` de tra dan gia; hoac bom `chia_fn` (khong `cast_fn`) thi
    coi nhu nguoi goi tu lai AI -> khong casting, khong cham mang.
    """
    co_dinh = che_do in (CHE_DO_MOT_NV, CHE_DO_NV_BOI_CANH)
    nen = {"style": {}, "characters": [_nhan_vat_co_dinh(context)] if co_dinh else [],
           "locations": []}
    if not nhat_quan or che_do == CHE_DO_MOT_NV:
        return nen
    if cast_fn is None and chia_fn is not None:
        return nen  # che do bom tay: khong tu goi mang de casting
    try:
        raw = (cast_fn(cues, context) if cast_fn is not None
               else _dung_dan_cast(goi, cues, context, nen["characters"]))
        cast = _sach_cast(raw, nv_dau=2 if co_dinh else 1)
    except Exception as loi:  # noqa: BLE001 — casting la phu, hong thi ve hanh vi cu
        emit({"type": "event", "event": "progress", "progress": 0.0,
              "message": "Chua dung duoc dan nhan vat ({0}). {1}".format(
                  str(loi)[:120],
                  "Chi giu nv1 co dinh." if co_dinh
                  else "Moi canh se tu do, khong khoa mot nguoi xuyen suot.")})
        return nen
    if co_dinh:
        # nv1 co dinh luon dung dau; AI lo tra them mot `nv1` thi bo — no khong
        # duoc dinh nghia lai nhan vat cua kenh.
        cast["characters"] = nen["characters"] + [
            c for c in cast["characters"] if c["id"] != "nv1"]
    # ═══ PHU HET DANH SACH BUOC DOC PHIM DA NHAN RA — KIEM BANG MA ═══
    #
    # AI hay "rut gon" dan: do 25/08/2026 tra 4/9 nhan vat, 4/6 noi. Con thieu
    # thi goi DUNG MOT luot bo sung xin nhung muc thieu, roi ghep vao.
    thieu_nv, thieu_loc = _con_thieu(cast, context)
    if thieu_nv or thieu_loc:
        try:
            ctx_bs = dict(context, bo_sung={"characters": thieu_nv, "locations": thieu_loc})
            raw2 = (cast_fn(cues, ctx_bs) if cast_fn is not None
                    else _dung_dan_bo_sung(goi, cues, cast, thieu_nv, thieu_loc))
            them = _sach_cast(raw2, nv_dau=len(cast["characters"]) + 1)
            co_nv = {c["id"] for c in cast["characters"]}
            co_loc = {l["id"] for l in cast["locations"]}
            so_trung = _ap_same_as(cast, raw2.get("same_as") if isinstance(raw2, Mapping) else None)
            cast["characters"] += [c for c in them["characters"] if c["id"] not in co_nv]
            cast["locations"] += [l for l in them["locations"] if l["id"] not in co_loc]
            if so_trung:
                emit({"type": "event", "event": "progress", "progress": 0.0,
                      "message": "{0} ten 'thieu' thuc ra la ten khac cua muc da co — gop, khong them id moi.".format(so_trung)})
            emit({"type": "event", "event": "progress", "progress": 0.0,
                  "message": "Casting bo sot {0} nhan vat, {1} noi chon — da bo sung "
                             "them {2} nhan vat, {3} noi.".format(
                                 len(thieu_nv), len(thieu_loc),
                                 len(them["characters"]), len(them["locations"]))})
        except Exception as loi:  # noqa: BLE001 — bo sung hong thi dung dan hien co
            emit({"type": "event", "event": "progress", "progress": 0.0,
                  "message": "Chua bo sung duoc dan ({0}); thieu: {1}.".format(
                      str(loi)[:80], ", ".join(thieu_nv + thieu_loc)[:200])})
    if cast["characters"] or cast["locations"]:
        emit({"type": "event", "event": "progress", "progress": 0.0,
              "message": "Da dung dan {0} nhan vat va {1} boi canh co dinh cho ca "
                         "video.".format(len(cast["characters"]), len(cast["locations"]))})
    return cast


def _da_nhan_ra(context) -> tuple:
    """Danh sach nhan vat / noi chon ma buoc doc phim da nhan ra (neu co)."""
    phim = context.get("film_analysis") if isinstance(context, Mapping) else None
    if not isinstance(phim, Mapping):
        return [], []
    nv = [str(x).strip() for x in (phim.get("characters_mentioned") or []) if str(x).strip()]
    loc = [str(x).strip() for x in (phim.get("locations_mentioned") or []) if str(x).strip()]
    return nv, loc


def _tu_khoa(chu: str) -> List[str]:
    return [t for t in re_tach(str(chu or "").lower()) if len(t) > 2 and t not in _TU_RONG]


def re_tach(chu: str) -> List[str]:
    return re.findall(r"[a-z0-9']+", chu.replace("'s", ""))


def _duoc_phu(muc: str, dan: List[Mapping[str, Any]], *cot) -> bool:
    """Mot muc da nhan ra co mat trong dan chua — so khop theo tu khoa."""
    tk = _tu_khoa(muc)
    if not tk:
        return True
    for c in dan:
        chu = " ".join(str(c.get(k) or "") for k in cot).lower()
        if any(t in chu for t in tk):
            return True
    return False


def _ap_same_as(cast, same_as) -> int:
    """`same_as` {tên còn thiếu → id đã có}: nối tên ấy vào `name` của mục đã có
    (để `_duoc_phu` khớp từ khoá), KHÔNG tạo id mới. Trả về số tên đã gộp."""
    if not isinstance(same_as, Mapping):
        return 0
    theo_id = {c["id"]: c for c in cast["characters"]}
    theo_id.update({l["id"]: l for l in cast["locations"]})
    n = 0
    for ten, ma in same_as.items():
        muc = theo_id.get(str(ma or "").strip())
        ten = str(ten or "").strip()
        if muc is None or not ten:
            continue
        hien = str(muc.get("name") or "")
        if ten.lower() not in hien.lower():
            muc["name"] = "{0} ({1})".format(hien, ten) if hien else ten
        n += 1
    return n


def _con_thieu(cast, context) -> tuple:
    nv, loc = _da_nhan_ra(context)
    thieu_nv = [m for m in nv if not _duoc_phu(m, cast["characters"], "role", "name",
                                                "english_prompt")]
    thieu_loc = [m for m in loc if not _duoc_phu(m, cast["locations"], "name",
                                                  "english_prompt")]
    return thieu_nv, thieu_loc


def _dung_dan_bo_sung(goi, cues, cast, thieu_nv, thieu_loc) -> Mapping[str, Any]:
    """Luot bo sung: chi xin nhung muc casting bo sot (khoa idempotent `cast-bs`)."""
    loi_doc = " ".join(str(cue.get("text") or "").strip() for cue in cues)
    da_co = "\n".join(
        ["- {0} ({1}): {2}".format(c["id"], c.get("role") or c.get("name"), c["english_prompt"][:160])
         for c in cast["characters"]]
        + ["- {0} ({1}): {2}".format(l["id"], l.get("name"), l["english_prompt"][:160])
           for l in cast["locations"]]) or "(nothing yet)"
    loi_nhac = _KHUON_BO_SUNG.format(
        da_co=da_co, thieu_nv=", ".join(thieu_nv) or "(none)",
        thieu_loc=", ".join(thieu_loc) or "(none)",
        nv_tiep="nv{0}".format(len(cast["characters"]) + 1),
        loc_tiep="loc{0}".format(len(cast["locations"]) + 1), transcript=loi_doc)
    return loc_json(goi(loi_nhac, "cast-bs"))


def _dung_dan_cast(goi, cues, context, co_san) -> Mapping[str, Any]:
    """Goi AI mot lan (`goi(..., "cast")`, khoa idempotent) de rut dan + boi canh + style."""
    # Ca loi doc, khong cat (ban cu cat 12.000 ky tu: video dai mat nhan vat cuoi).
    loi_doc = " ".join(str(cue.get("text") or "").strip() for cue in cues)
    boi_canh = _boi_canh_chu(context, "(khong co)")
    fixed = _LUAT_NV1_CO_DINH.format(mo_ta=co_san[0]["english_prompt"]) if co_san else ""
    nv, loc = _da_nhan_ra(context)
    loi_nhac = _KHUON_CAST.format(context=boi_canh, transcript=loi_doc,
                                  fixed_rule=fixed, nv_dau="nv2" if co_san else "nv1",
                                  phai_co_nv=", ".join(nv) or "(not analysed — decide from the transcript)",
                                  phai_co_loc=", ".join(loc) or "(not analysed — decide from the transcript)")
    return loc_json(goi(loi_nhac, "cast"))


#: Nhan vat DOI TRANG PHUC giua truyen → moi giai doan mot anh tham chieu rieng.
#:
#: Chu du an 25/08/2026 xem video: *"luc dau con meo chua he co mu, giay, mai
#: sau doan no noi chuyen voi chu no moi duoc cap… doan cuoi cau ut va cong
#: chua lai khac… voi mot cau chuyen thi nen co DU tham chieu va cac prompt
#: phai dung DUNG"*. Mot anh tham chieu cho "meo di hia" khong the vua la meo
#: tran vua la meo mang ung; ep khoa theo mot anh thi nua dau phim sai, khong
#: khoa thi nua sau phim troi. Tach thanh nv4 (truoc) va nv4b (sau): cung mat,
#: khac do; canh nao dung id cua giai doan do.
#: Đầu mô tả của một giai đoạn BIẾN HÌNH — cả tool nhận ra nhờ dấu này
#: (`core.prompt_visuals.doi_thiet_ke_nhan_vat` không ghép lại thân cũ vào).
DAU_BIEN_HINH = "full form: "
_TU_BIEN_HINH = re.compile(r"\b(whole body|becomes?|turn(?:s|ed)? into|transform(?:s|ed)?|in the form of|shape-?shift)\b", re.I)
_LOAI_VAT = ("lion", "mouse", "rat", "cat", "dog", "wolf", "fox", "bear", "frog", "toad", "bird", "eagle",
             "dragon", "snake", "fish", "rabbit", "hare", "goat", "sheep", "horse", "donkey", "pig", "cow",
             "tiger", "elephant", "monkey", "swan", "duck", "goose", "beetle", "spider", "ogre", "giant",
             "prince", "princess", "old woman", "old man", "beggar", "statue", "tree", "stone", "cloud")


def _la_bien_hinh(than_goc: str, do: str) -> bool:
    """Giai đoạn này có phải một HÌNH DẠNG khác (biến hình) chứ không phải bộ đồ?"""
    if _TU_BIEN_HINH.search(do or ""):
        return True
    g, d = (than_goc or "").lower(), (do or "").lower()
    for loai in _LOAI_VAT:
        if re.search(r"\b%s\b" % re.escape(loai), d) and not re.search(r"\b%s\b" % re.escape(loai), g):
            return True
    return False


def _tach_giai_doan(goc: Dict[str, Any], stages) -> List[Dict[str, Any]]:
    ds = [s for s in (stages or []) if isinstance(s, Mapping) and str(s.get("outfit") or "").strip()]
    if len(ds) < 2:
        return [goc]
    ra = []
    for k, st in enumerate(ds):
        c = dict(goc)
        c["id"] = goc["id"] if k == 0 else "{0}{1}".format(goc["id"], chr(ord("a") + k))
        do = str(st["outfit"]).strip()
        if k > 0 and _la_bien_hinh(goc["english_prompt"], do):
            # BIẾN HÌNH (yêu tinh → sư tử → chuột): giai đoạn là một THÂN THỂ khác,
            # không phải bộ đồ. Ghép thân cũ + "đồ" mới là máy vẽ ra ông khổng lồ
            # cầm con chuột (đo 26/08/2026, nv15c). Mô tả giai đoạn = hình dạng mới.
            c["english_prompt"] = "{0}{1} — the transformed form of {2}".format(
                DAU_BIEN_HINH, do.rstrip(" ;."), goc.get("name") or goc["id"])
        else:
            # Bo TU THE khoi o "outfit": no la do mac, khong phai nhan vat dang lam
            # gi. Do 27/08/2026 (phim 0002): "lying down ... as an ordinary sleeping
            # wolf" lot vao day -> anh tham chieu ve ra con soi DANG NGU.
            c["english_prompt"] = "{0}; outfit at this stage: {1}".format(
                goc["english_prompt"].rstrip(" ;."), bo_tu_the(do) or do)
        c["giai_doan"] = k + 1
        c["giai_doan_khi"] = str(st.get("when") or "").strip()
        c["goc_id"] = goc["id"]
        c["notes"] = "{0}Stage {1}/{2} of {3} — {4}.".format(
            (goc["notes"] + " ") if goc["notes"] else "", k + 1, len(ds), goc["id"],
            c["giai_doan_khi"] or "unspecified moment")
        ra.append(c)
    return ra


def _sach_cast(raw, nv_dau: int = 1) -> Dict[str, Any]:
    """Chuan hoa dan cast + boi canh: bo dong thieu mo ta, danh id neu AI bo trong."""
    if not isinstance(raw, Mapping):
        return {"style": {}, "characters": [], "locations": []}
    style = raw.get("style") if isinstance(raw.get("style"), Mapping) else {}
    ra: List[Dict[str, Any]] = []
    for thu_tu, item in enumerate(raw.get("characters") or [], nv_dau):
        if not isinstance(item, Mapping):
            continue
        english = str(item.get("english_prompt") or "").strip()
        if not english:
            continue  # khong co mo ta ngoai hinh thi khong khoa duoc, bo qua
        cid = str(item.get("id") or "").strip() or "nv{0}".format(thu_tu)
        # `body_mode` phai co ten o day. Ham nay dung mot dict MOI theo danh
        # sach truong co dinh, nen truong nao khong duoc goi ten la bi vut
        # lang le — them o khuon JSON va o CHARACTER_HEADERS thoi thi chua du,
        # va cai hong khong keu mot tieng nao (27/08/2026).
        goc = {"id": cid, "role": str(item.get("role") or ""),
               "name": str(item.get("name") or ""), "english_prompt": english,
               "body_mode": str(item.get("body_mode") or ""),
               "reference_lock": str(item.get("reference_lock") or ""),
               "gender": str(item.get("gender") or ""),
               "age": str(item.get("age") or ""),
               "notes": str(item.get("notes") or "")}
        ra.extend(_tach_giai_doan(goc, item.get("stages")))
    boi_canh: List[Dict[str, Any]] = []
    for thu_tu, item in enumerate(raw.get("locations") or [], 1):
        if not isinstance(item, Mapping):
            continue
        english = str(item.get("english_prompt") or "").strip()
        if not english:
            continue
        lid = str(item.get("id") or "").strip() or "loc{0}".format(thu_tu)
        boi_canh.append({"id": lid, "name": str(item.get("name") or ""),
                         "english_prompt": english,
                         "location_lock": str(item.get("location_lock") or ""),
                         "lighting_default": str(item.get("lighting_default") or "")})
    return {"style": dict(style), "characters": ra, "locations": boi_canh}


def _khoi_cast_style(cast: Mapping[str, Any]) -> str:
    """Dung chu cho `<<CAST_STYLE>>`: dan nhan vat + boi canh + luat tai dung + style.

    Rong (khong dan, khong style) thi tra "" — `dien_khuon` xoa placeholder,
    template ve dung nhu cu.
    """
    chars = cast.get("characters") or []
    locs = cast.get("locations") or []
    style = cast.get("style") or {}
    dong: List[str] = []
    if chars:
        dong.append("## RECURRING CHARACTERS — reuse, never redesign")
        dong.append("These are the ONLY recurring characters in this video. When "
                    "a scene shows one, put its id in `characters_used` and, in the "
                    "prompt text, call it ONLY by its id (e.g. `nv4`) with pose, "
                    "gesture and expression — NEVER describe its face, hair, fur, "
                    "clothes, props or colours in a scene: its reference image and "
                    "the description below are attached automatically, and words "
                    "you add can only contradict them. The main character is the "
                    "centre of every scene it appears in; supporting figures, props "
                    "and settings stay simple and in the SAME style. The video "
                    "prompt must never change a character — only its expression "
                    "and action:")
        co_giai_doan = any(c.get("giai_doan") for c in chars)
        if co_giai_doan:
            dong.append("Some characters CHANGE LOOK during the story and therefore have "
                        "one id PER STAGE (nv4 = before, nv4b = after…). In each scene "
                        "use the id of the stage that matches that moment of the story — "
                        "never the other one, never both.")
        for c in chars:
            nhan = c.get("role") or c.get("name") or "character"
            if c.get("co_dinh"):
                dong.append("- {0} ({1}): reference image `nv1.png` is attached at "
                            "generation time. Refer to it ONLY as `nv1 (nv1.png)`; "
                            "NEVER describe its face, hair, skin, clothes or colours "
                            "— only pose, gesture, expression.".format(c["id"], nhan))
            elif c.get("giai_doan"):
                dong.append("- {0} ({1}, STAGE {2}: {3}): {4}".format(
                    c["id"], nhan, c["giai_doan"], c.get("giai_doan_khi") or "?",
                    c["english_prompt"]))
            else:
                dong.append("- {0} ({1}): {2}".format(c["id"], nhan, c["english_prompt"]))
    if locs:
        if dong:
            dong.append("")
        dong.append("## RECURRING LOCATIONS — reuse, do not redesign")
        dong.append("When a scene is set in one of these places, put its id in "
                    "`location_used` and describe the place EXACTLY as written. "
                    "The place follows the STORY like a film: consecutive scenes "
                    "stay in the same place until the narration says the "
                    "characters moved; never change place for variety — change "
                    "shot size and angle instead:")
        for l in locs:
            dong.append("- {0} ({1}): {2}{3}".format(
                l["id"], l.get("name") or "place", l["english_prompt"],
                " Lighting: " + l["lighting_default"] if l.get("lighting_default") else ""))
    duoi = [("Image style", style.get("image_style")), ("Palette", style.get("palette")),
            ("Motion", style.get("motion"))]
    duoi = [(ten, str(gt).strip()) for ten, gt in duoi if str(gt or "").strip()]
    if duoi:
        if dong:
            dong.append("")
        dong.append("## STYLE — hold this exact look across every scene")
        for ten, gt in duoi:
            dong.append("{0}: {1}".format(ten, gt))
    return "\n".join(dong)


# ── Cảnh lặp: cùng cỡ khung, cùng người, cùng chỗ, cảnh này nối cảnh kia ───────
#
# Đo 25/08/2026 (story-3d/0001, 123 cảnh): 105/122 cặp cảnh liền nhau trùng
# hơn 50% từ; 45 cảnh mở bằng "Medium shot", 21 "Wide shot"; tám cảnh 5–12 là
# cùng cậu bé + mèo ở cùng bậc cửa, cùng cỡ khung. Người xem thấy "cảnh lặp
# lại". Luật 3 trong lời nhắc chia cảnh đã bảo đổi khung hình mà AI vẫn không
# đổi — nên mã phải đo và bắt sửa, không tin lời hứa.
_NGUONG_LAP = 0.6
_TIEN_TRINH_KHUNG = ("extreme close-up (eyes, paws, hands, one object)", "close-up",
                     "medium shot", "over-the-shoulder shot", "low-angle shot",
                     "high-angle / top-down view", "wide shot", "insert of a prop or detail",
                     "POV shot (what the character sees)", "reaction shot of the listener")


def _mo_dau_khung(prompt: str) -> str:
    """Cụm mở đầu nói cỡ khung: 'medium shot', 'close-up', 'wide shot'…"""
    dau = re.split(r"\b(of|on|at|:)\b", str(prompt or "").strip().lower(), maxsplit=1)[0]
    return re.sub(r"[^a-z\- ]", "", dau).strip()[:40]


def _tu_than(prompt: str) -> set:
    than = str(prompt or "").split("REFERENCE IMAGES", 1)[0].lower()
    return set(re.findall(r"[a-z]{4,}", than))


def _canh_lap(scenes) -> List[int]:
    """Chỉ số các cảnh LẶP cảnh liền trước — người xem thấy "cùng một bức".

    Lặp = cùng cỡ khung mở đầu VÀ trùng > 50% từ (cùng người, cùng chỗ, cùng
    cỡ), hoặc trùng > 80% từ bất kể khung. Hai cỡ khung khác nhau của cùng một
    câu (medium rồi close-up) là cách phim che một câu dài — không phải lặp.
    """
    ra = []
    for i in range(1, len(scenes)):
        a, b = scenes[i - 1], scenes[i]
        pa, pb = str(a.get("img_prompt") or ""), str(b.get("img_prompt") or "")
        cung_khung = bool(_mo_dau_khung(pa)) and _mo_dau_khung(pa) == _mo_dau_khung(pb)
        A, B = _tu_than(pa), _tu_than(pb)
        j = len(A & B) / max(1, len(A | B))
        if (cung_khung and j > 0.5) or j > 0.8:
            ra.append(i)
    return ra


_KHUON_PHA_LAP = """These image prompts are CONSECUTIVE scenes of one animated film, and each one
repeats the framing of the scene before it (same shot size, same people, same place) — the
viewer sees the same picture again and again. Rewrite ONLY the scenes listed under "REWRITE"
so that no two consecutive scenes share a shot size/angle. Use a real film progression and
vary the staging: {tien_trinh}. Keep for every scene: the same characters (their ids in
parentheses, e.g. (nv1), stay), the same place id, the same story beat/action, the same style
words at the end. Change the camera, the distance, the angle, what is in the foreground, and
which part of the action we see. Family-friendly wording. Return ONLY JSON:
{{"<scene_id>": {{"img_prompt": "...", "video_prompt": "..."}}}} — video_prompt is the same
shot in motion (one small camera move, one small action), under 60 words.

CONTEXT (previous scene of each rewritten one, do not change these; each rewritten scene must
open with a DIFFERENT shot than the one named after "previous opening"):
{ngu_canh}

REWRITE:
{can_sua}"""


def _pha_lap_canh(scenes, goi, *, moi_lan: int = 20) -> int:
    """Tìm cảnh lặp khung với cảnh trước và nhờ AI đổi khung (một lượt gọi mỗi ≤20 cảnh).

    Bản viết lại chỉ được nhận khi vẫn giữ đúng các id nhân vật/bối cảnh của
    cảnh gốc — lệch id là bỏ, giữ nguyên bản cũ. Trả về số cảnh đã đổi.
    """
    if goi is None:
        return 0
    doi = 0
    for vong in range(2):  # vòng 2 chỉ còn những cảnh AI viết lại mà vẫn lặp
        lap = _canh_lap(scenes)
        if not lap:
            break
        doi += _viet_lai_canh_lap(scenes, goi, lap, moi_lan, vong)
    return doi


def _viet_lai_canh_lap(scenes, goi, lap, moi_lan: int, vong: int) -> int:
    doi = 0
    for dau in range(0, len(lap), moi_lan):
        nhom = lap[dau:dau + moi_lan]
        ngu_canh = "\n".join("[{0}] (previous opening: {2}) {1}".format(
            scenes[i - 1].get("scene_id"), str(scenes[i - 1].get("img_prompt") or "")[:400],
            _mo_dau_khung(str(scenes[i - 1].get("img_prompt") or "")) or "?") for i in nhom)
        can_sua = "\n".join("[{0}] {1}\n    video: {2}".format(
            scenes[i].get("scene_id"), str(scenes[i].get("img_prompt") or "")[:600],
            str(scenes[i].get("video_prompt") or "")[:200]) for i in nhom)
        loi_nhac = _KHUON_PHA_LAP.format(tien_trinh="; ".join(_TIEN_TRINH_KHUNG),
                                         ngu_canh=ngu_canh, can_sua=can_sua)
        try:
            tra = loc_json(goi(loi_nhac, "pha-lap-{0}-{1}".format(vong, dau)))
        except Exception:  # noqa: BLE001 — không đổi được thì giữ nguyên
            continue
        if not isinstance(tra, dict):
            continue
        for i in nhom:
            sc = scenes[i]
            moi = tra.get(str(sc.get("scene_id")))
            if not isinstance(moi, dict):
                continue
            img = str(moi.get("img_prompt") or "").strip()
            vid = str(moi.get("video_prompt") or "").strip()
            if len(img) < 30:
                continue
            cu_id = set(re.findall(r"\b(?:nv|loc)\d+[a-z]?\b", str(sc.get("img_prompt") or "")))
            moi_id = set(re.findall(r"\b(?:nv|loc)\d+[a-z]?\b", img))
            if cu_id and moi_id != cu_id:
                continue
            sc["img_prompt"] = img
            if len(vid) >= 10:
                sc["video_prompt"] = vid
            doi += 1
    return doi


def _gan_reference_files(scenes, characters, locations=()) -> None:
    """Moi canh dung nhan vat / boi canh nao thi tro toi `<id>.png` — quy uoc VE3.

    Khau dung anh o VE3/tab Hang loat bam vao `reference_files` de dinh anh
    tham chieu, khoa mat mot nguoi va mot noi chon xuyen suot. Chi gan id co
    that trong dan. Giao dien doi ten tep thanh duong dan that sau khi anh
    tham chieu duoc tao ra canh Excel.
    """
    hop_nv = {c["id"] for c in characters}
    hop_loc = {l["id"] for l in locations}
    for scene in scenes:
        if scene.get("reference_files"):
            continue
        tho = str(scene.get("characters_used") or "").replace(",", " ").split()
        # MOI nhan vat co mat trong canh deu phai co anh cua no. Truoc 26/08/2026
        # tran nay la 2, theo phep do 25/08 khi loi nhac CON ta nhan vat bang chu
        # (4–6 anh → 3,75; 2 anh + 1 noi → 3,79). Bo ta chu roi thi phep do do
        # khong con dung: nhan vat khong co anh la nhan vat KHONG CO GI — do
        # 26/08 tren phim 0002, canh 12–13 co 4 nhan vat ma chi gan 2, may ve ra
        # may con cao va mot con gau thay cho dan de con.
        #
        # ═══ VA MOI ANH GAN VAO DEU PHAI CO CHO DUNG TRONG CAU ═══
        #
        # `characters_used` do AI khai, cau van cung do AI viet — hai thu ay
        # LECH NHAU. Gan anh theo o khai thi co tam anh khong ai goi ten, va
        # may lam dung cai viec ta so nhat: no ve TAM ANH THUA ay, roi bo
        # nhung nhan vat cau van co goi.
        #
        # Do 27/08/2026 tren phim openstory/0003: 4/29 canh gan anh cho nhan
        # vat cau van khong nhac. Canh 14 gan `nv5` (con vit) ma cau van chi
        # noi `nv1` va `nv3` — anh ra CHI CO CON VIT, cau be va con meo bien
        # mat. Chu du an: *"van la tinh trang thieu tham chieu"*.
        #
        # Nen: chi gan anh cho nhan vat CAU VAN THAT SU GOI TEN, va xep theo
        # dung thu tu chung xuat hien trong cau. Cau van khong goi ai (hiem)
        # thi moi lui ve o khai.
        than = str(scene.get("img_prompt") or "").split("\nREFERENCE IMAGES")[0]
        trong_cau = []
        for m in re.finditer(r"\b(nv\d+[a-z]?)\b", than):
            if m.group(1) in hop_nv and m.group(1) not in trong_cau:
                trong_cau.append(m.group(1))
        # Cau van khong goi ai thi lui ve o khai — day la hanh vi CO Y cua san
        # pham ("canh nao chua gan ai thi mac dinh nhan vat chinh", xem
        # `tests/test_nhan_vat_xuyen_suot.py`), va no dung cho canh AI quen kê
        # tên.
        #
        # Cho nay con MOT CA CHUA XU: canh co y trong nguoi (phim 0004 canh 23
        # "empty gathering spot with no people", canh 24 can canh mot bong sen)
        # van bi gan anh nhan vat chinh. Chua sua vi chua co cach doc chac chan
        # "canh nay co nguoi khong" tu cau van — do bang tu khoa thi sai nhieu
        # hon dung. Guard co san cau "never render a reference image itself as
        # a separate panel or subject" dang giu cho nay; neu do duoc no thung
        # thi quay lai day.
        ids = (trong_cau or [i for i in tho if i in hop_nv])[:TOI_DA_NV_THAM_CHIEU]
        # Giu o khai khop voi anh that su gan, de khau sau khong doc mot danh
        # sach da khong con dung.
        if ids:
            scene["characters_used"] = ",".join(ids)
        loc = str(scene.get("location_used") or "").strip()
        if loc in hop_loc:
            ids.append(loc)
        if ids:
            scene["reference_files"] = json.dumps(["{0}.png".format(i) for i in ids])


#: Cau khoa nhan dang — chu du an 25/08/2026 xem video: *"nhan vat tham chieu
#: luc co luc khong, luc thi con meo A luc thi con meo B… lon xon"*. Do tren
#: Excel that: 63/63 canh co meo deu gan nv4.png, nhung loi nhac chi ghi "(nv4)"
#: — vo nghia voi mo hinh ve — va moi canh AI ta con meo mot kieu. Mo hinh anh
#: khong biet anh tham chieu nao la ai, nen no ve lai theo chu. Khoa bang ma:
#: noi ro "reference image 1 = nv4, the cat: <mo ta nguyen van>" va bat ve y het.
#: Bao nhieu NHAN VAT duoc gan anh tham chieu cho mot canh (them 1 boi canh).
#:
#: ═══ CONG NHAN 10 ANH, NHUNG MO HINH CHI DUNG NOI 3 ═══
#:
#: Do 27/08/2026 tren phim `openstory/0003`, cham tung canh bang `cham_anh`
#: roi gom theo SO ANH GAN VAO:
#:
#:     2 anh : 4,00
#:     3 anh : 4,33   <- tot nhat
#:     4 anh : 2,80   <- sup
#:     5 anh : 2,00
#:
#: ═══ VA PHEP DO AY BI NHIEU — DUNG DUNG NO DE HA TRAN ═══
#:
#: Bang tren gom theo so anh, nhung canh nhieu anh CUNG LA canh nhieu nhan
#: vat, tuc bo cuc kho hon. No do "canh dong thi diem thap", khong do "nhieu
#: anh thi hong". Phep do CO KIEM SOAT (cung mot loi nhac, doi so anh 1/2/3/4,
#: sau luot) cho ket qua NHIEU: 4 anh co luot dung ca ba nhan vat, 3 anh co
#: luot hong.
#:
#: 27/08 toi ha tran ve 2 dua tren bang bi nhieu ay. Hau qua thay ngay o phim
#: 0007: canh "thay lang toi xem benh" bi ep con hai nhan vat, va NGUOI BENH —
#: nhan vat ca canh noi ve — bi day ra thanh "a small blanketed figure", roi
#: duoc ve thanh mot nguoi la. Chu du an: *"khong nen ha tran, veo 3 cho tham
#: chieu nhieu ma"*.
#:
#: Nen: giu tran 4 (cong nhan toi 10 anh). Cho chua khong nam o so anh ma o
#: KHUNG HINH — canh chon hai nguoi cua cau chuyen lam chu the, con lai la bo
#: vai, ban tay o mep khung, hoac ra ngoai khung. Va KHONG BAO GIO viet mot
#: nhan vat co ten thanh mot cai xac vo danh. Xem `CHANNEL/*/prompt/7-canh.md`.
TOI_DA_NV_THAM_CHIEU = 4

#: ═══ HAI HANG SO CU DA BO — DUNG DUNG LAI ═══
#:
#: `_KHOA_NHAN_VAT` (doan van ~1.100 ky tu liet ke mat, mat, ti le, da, long,
#: trang phuc, dao cu phai chep y het) va `_KHOA_BOI_CANH` bi go 27/08/2026.
#:
#: Do tren canh 25 cua openstory/0002 — dung canh may ve cau be thanh dua khac:
#:
#:     doan van dai (loi cu)      3, 3 diem
#:     mot dong guard (OpenStory) 4, 4 diem
#:
#: Liet ke tung net de "chep cho giong" hoa ra la chu di cai voi anh, va may
#: nghe chu. Xem `_GUARD_NHAN_DANG`. Luat vat ly cua `_KHOA_BOI_CANH` duoc giu
#: lai rieng o `_LUAT_DUNG_CHAN` vi no khong noi gi ve danh tinh.

#: Khoa cho loi nhac VIDEO — dat o DAU, khong phai cuoi. Do 25/08/2026 (12 clip
#: Veo 3, AI cham khung cuoi duoc dung): cau khoa o cuoi → 3,00; khoa o dau kem
#: mo ta tung nhan vat + "khong them, khong bot" → 3,50. Veo nang phan dau.
#: KHONG ta lai nhan vat trong khoa video. Chu du an 25/08/2026: *"khong mo ta chi
#: tiet nhan vat ma bao la anh tham chieu di kem"* — do cung 12 clip: khoa kem mo
#: ta 3,50; khoa chi noi "y het khung dau" 3,67 (9/12 dat >=4). Chu va anh cai nhau
#: thi mo hinh chon chu; bo chu di la no bam anh.
_KHOA_VIDEO_DAU = ("IDENTITY LOCK, highest priority for the entire clip: every character stays "
                   "exactly as drawn in the first frame — same face, eyes, body proportions, "
                   "outfit, props and line style; nothing is added (no extra clothes, cape, "
                   "belt, strap, collar, weapon, fur texture or stripes) and nothing is "
                   "removed. Only pose, gesture, expression and camera move. ")
_KHOA_VIDEO_CU = (" Keep every character exactly as drawn in the first frame — same face, "
                  "outfit, proportions and line style — for the whole clip.")


def _mo_ta_ngan(c: Mapping[str, Any]) -> str:
    """Mo ta ngoai hinh, bo duoi phong cach ('black pencil…') cho gon."""
    chu = str(c.get("english_prompt") or "")
    for moc in (", black pencil", ", pure black", ", Simple hand-drawn", ", simple hand-drawn"):
        if moc in chu:
            chu = chu.split(moc)[0]
    return chu.strip()


def _khoa_video(vid: str, ids, dan: Mapping[str, Any]) -> str:
    """Dat khoi khoa nhan dang vao DAU loi nhac video (bo cau khoa cu o cuoi neu co)."""
    goc = str(vid or "").replace(_KHOA_VIDEO_CU, "").strip()
    if not goc or goc.startswith("IDENTITY LOCK"):
        return goc
    if not any(i in dan for i in ids):
        return goc
    return _KHOA_VIDEO_DAU + goc


def _mo_ta_tham_chieu(i: str, dan: Mapping[str, Any], noi: Mapping[str, Any]) -> str:
    """Mot dong trong khoi khoa: CHI ten + vai, KHONG ta ngoai hinh.

    Chu du an 26/08/2026: *"khong mo ta chi tiet ma chi co trong prompt dung media
    id cua anh nhan vat do, vi neu mo ta chi tiet thi no se bi sai"*. Do cung ngay
    (mau canh 5-12 hoathinh-3d): moi lan ve lai, may doc cau ta roi tuong tuong ra
    mot con meo khac - luc tron map, luc cao gay. Chu va anh danh nhau thi may
    nghe chu. Bo chu di thi no bam anh. Loi nhac CLIP da bo ta tu 25/08 (3,50 ->
    3,67); day la nua con lai.
    """
    if i in dan:
        c = dan[i]
        return "{0}, the {1}".format(i, c.get("role") or c.get("name") or "character")
    l = noi[i]
    return "{0}, {1}".format(i, l.get("name") or "place")


#: Id nhan vat / boi canh trong loi nhac: `nv4`, `nv7b`, `loc12`.
_MAU_ID = re.compile(r"\b((?:nv|loc)\d+[a-z]?)\b(?!\s*\[reference)")


#: Nhan nhan dang dai toi da bao nhieu ky tu. Du de noi ro con vat/nguoi ay
#: la ai (loai, mau), ngan de khong thanh mot doan ta canh tranh voi bo cuc.
DAI_NHAN_CO_DINH = 100

#: Tu noi khong duoc dung cuoi nhan — cat o do la nhan cut nghia.
_TU_NOI_CUT = frozenset(("with", "and", "the", "of", "in", "on", "a",
                         "along", "over", "under", "at", "to", "for"))


def _nhan_co_dinh(e, mac_dinh: str = "") -> str:
    """Nhan cho nhan vat/noi KHONG duoc gui anh tham chieu — LUON GIONG NHAU.

    ═══ VI SAO KHONG PHAI CHI CAI VAI ═══

    Ban dau cho ra "the duck". Do tren luot that openstory/0002 (26/08/2026,
    30 canh): con vit hien ra **hai kieu khac nhau trong cung mot phim** —
    canh 11-15 la vit co xanh nau, canh 28-30 la vit trang. Dan ghi ro nv5 la
    *"creamy-white feathers"*, tuc may bia ra con vit nau. Vai suong ("the
    duck") noi cho may biet VE CON GI, khong noi VE CON NAO.

    Nen nhan phai kem mot cum nhan dang **chep nguyen tu dan va giong het nhau
    o moi canh** — giong het la phan quan trong nhat: no thay cho vai tro ma
    anh tham chieu dang lam cho hai nhan vat kia.

    ═══ KHONG MAU THUAN VOI LUAT 26/08 ═══

    Chu du an hom ay: *"khong mo ta chi tiet ma chi co trong prompt dung media
    id cua anh nhan vat do, vi neu mo ta chi tiet thi no se bi sai"*. Luat ay
    noi ve nhan vat CO anh: chu va anh danh nhau thi may nghe chu, nen phai bo
    chu. O day khong co anh nao de danh nhau — bo chu di la khong con gi.
    """
    tho = str(e.get("english_prompt") or "").strip()
    if not tho:
        return mac_dinh or "the {0}".format(
            e.get("role") or e.get("name") or "character")
    if len(tho) <= DAI_NHAN_CO_DINH:
        return tho
    # Cat o KHOANG TRANG gan tran nhat, KHONG cat o dau phay dau tien: cau ta
    # trong dan de loai o menh de dau ("a friendly ordinary house cat") va MAU
    # o menh de sau ("grey-and-brown mackerel tabby") — cat som la vut di dung
    # thu giu cho con vat khong doi mau giua phim.
    cat = tho.rfind(" ", 0, DAI_NHAN_CO_DINH)
    ra = tho[:cat if cat > 20 else DAI_NHAN_CO_DINH].rstrip(" ,")
    # Bo tu noi bi cut o duoi ("with", "along the") — nhan phai doc duoc.
    while True:
        tu = ra.rsplit(" ", 1)[-1].lower()
        if tu in _TU_NOI_CUT and " " in ra:
            ra = ra.rsplit(" ", 1)[0].rstrip(" ,")
            continue
        return ra


def _goi_lai(i: str, dan, noi) -> str:
    """Cach goi mot nhan vat / mot noi o nhung lan nhac SAU lan dau.

    Dan ghi ten kieu `"Ti (the poor boy)"`, `"Meo muop (the tabby cat)"` — cum
    trong ngoac la cum tieng Anh dung duoc thang trong loi nhac. Khong co thi
    lui ve ten, roi toi vai. Khong bao gio tra rong: tra rong la cau van thung
    mot lo.
    """
    e = dan.get(i) or noi.get(i) or {}
    ten = str(e.get("name") or "").strip()
    m = re.search(r"\(([^)]+)\)", ten)
    if m:
        trong = m.group(1).strip()
        if re.match(r"^(the|a|an) ", trong, re.I):
            return trong
    dau = re.split(r"\s*\(", ten, 1)[0].strip()
    if dau:
        return dau if i in noi else ("the " + dau if not dau.lower().startswith("the ") else dau)
    vai = re.split(r"\s*\(", str(e.get("role") or ""), 1)[0].strip()
    return ("the " + vai) if vai else i


def _thay_id_khong_co_anh(chu: str, gan, dan, noi) -> str:
    """Id CO trong dan nhung KHONG duoc gui anh kem -> thay bang VAI cua no.

    ═══ VI SAO ═══

    Tran `TOI_DA_NV_THAM_CHIEU` (4 nhan vat + 1 noi moi canh) cat bot id du.
    Nhung nhan vat thu nam van bi AI viet vao loi nhac bang chinh cai id — va id la thu chi
    TOOL hieu. May ve khong nhan duoc ten tep, no nhan mot day anh theo thu tu.
    Thay `nv4` cho no la mot chuoi vo nghia: no be ra mot nhan vat moi, moi
    canh mot kieu.

    Do tren luot that (openstory/0002, 26/08/2026): **12/30 canh** co id tran —
    nv4 thay lang, nv5 con vit, nv6 dan ca. Ba nhan vat phu, khong canh nao ve
    giong canh nao.

    Thay bang VAI ("the village healer"), KHONG ta ngoai hinh: nhan vat co anh
    thi chu va anh danh nhau, may nghe chu (chu du an 26/08 — xem
    `_mo_ta_tham_chieu`); nhan vat khong co anh thi khong co gi de danh nhau,
    nhung ta ky van lam may ve moi canh mot kieu. Vai la muc vua du: may biet
    ve ai, ma khong co chi tiet nao de troi.
    """
    def doi(m):
        i = m.group(1)
        if i in gan:
            return i
        if i in dan:
            return _nhan_co_dinh(dan[i])
        if i in noi:
            l = noi[i]
            return _nhan_co_dinh(l, mac_dinh=str(l.get("name") or "the place"))
        return i  # id la, khong co trong dan: de nguyen cho nguoi sua tool thay

    # `(nv4)` -> `(the village healer)` roi moi toi dang tran, de khong sinh
    # ra `((the …))`.
    return _MAU_ID.sub(doi, str(chu or ""))


#: ═══ MOT DONG, KHONG PHAI MOT DOAN VAN ═══
#:
#: Nguyen van cua OpenStory (`src/lib/prompts/reference-image-prompt.ts`,
#: hang `IDENTITY_GUARD`). Ho co y BO doan van dai liet ke mat/mat/ti le/da/
#: trang phuc ma tool nay dung truoc do; ghi chu cua ho: *"One-line identity
#: guard replacing the old IMPORTANT paragraph… phrased as cross-panel
#: CONSISTENCY, not likeness replication"*.
#:
#: Do tren may chu cua ta (26-27/08/2026, canh 25/26/29 cua openstory/0002 —
#: dung nhung canh may ve nham nguoi): doan van dai 3,00 diem; mot dong nay
#: 4,00. Liet ke tung net de "chep cho giong" hoa ra lai la chu di cai voi
#: anh, va may nghe chu.
_GUARD_NHAN_DANG = (
    "Reference images define identity — keep every referenced character, "
    "object and location consistent with its reference image; never render a "
    "reference image itself as a separate panel or subject.")

#: Luat vat ly giu lai tu `_KHOA_BOI_CANH` cu: no khong noi ve danh tinh nen
#: khong dinh dang toi guard, ma bo di thi nhan vat dung tren mat nuoc.
_LUAT_DUNG_CHAN = (
    "Every character stands, sits or walks on the SOLID GROUND of the place "
    "(floor, path, grass, bank, steps) with correct scale and physics — never "
    "standing on water, floating in the air or sunk into walls, unless the "
    "text explicitly says so.")


def _ep_phong_cach(scenes, style, emit=None) -> int:
    """Ep MOI canh mang dung phong cach cua kenh, khong tin AI chep lai.

    ═══ VI SAO PHAI EP BANG MA ═══

    Phong cach kenh duoc dua vao loi nhac qua `<<CAST_STYLE>>` va AI duoc dan
    la chep lai vao cuoi moi `img_prompt`. Do tren phim `openstory/0011`
    (Thach Sanh, 28/08/2026), 64 canh:

        37 canh chep dung "stylised 3D animated film still, Pixar-like..."
        23 canh KHONG co cau phong cach nao
         4 canh tu viet "hand-painted 2D animated feature style"  ← khac han

    Tuc mot phan tu bo phim co the ra mot net ve khac. Nguoi xem thay ngay,
    va do la loi khong the "chua sau" duoc: anh da ve roi.

    Cung mot hinh dang voi `LUAT_TIENG_CANH`: dan trong loi nhac la dieu kien
    CAN, ma khong du. Cai gi phai dung y het o moi canh thi de MA ghim, dung
    de AI chep.

    Cat cau phong cach AI tu viet truoc khi noi cau that vao: hai cau phong
    cach danh nhau thi may ve nghe cau nao khong ai doan duoc.
    """
    import re as _re

    duoi = str((style or {}).get("image_style") or "").strip().rstrip(".,")
    if not duoi:
        return 0
    # Cac cum "phong cach" AI hay tu che ra. Chi cat dung cum, khong cat cau
    # chua no: phan con lai cua cau van la noi dung canh.
    _LAC = _re.compile(
        r",?\s*(?:hand[- ]painted\s+)?(?:2D|3D)?\s*"
        r"(?:animated\s+(?:feature|film|movie|cartoon)\s*(?:style|still)?|"
        r"anime\s*style|watercolou?r\s*(?:style|illustration)|"
        r"oil\s*painting|storybook\s*illustration|comic\s*book\s*style)",
        _re.I)
    # Cau phong cach phai nam cuoi phan TA CANH, khong phai cuoi ca chuoi:
    # phia sau con khoi "REFERENCE IMAGES are attached..." (luat nhan dang, luat
    # dung chan). Noi ra dang sau la dan phong cach vao cuoi mot cau luat —
    # do 28/08/2026 tren phim 0011, canh 42 ra "...sunk into walls, unless the
    # text explicitly says so, stylised 3D animated film still, ...".
    NEO = "REFERENCE IMAGES"
    doi = 0
    for s in scenes:
        chu = str(s.get("img_prompt") or "")
        if not chu.strip():
            continue
        # "Da co phong cach cua kenh" khong doi CHEP Y HET: AI hay viet mot ban
        # rut gon ("...Pixar-like soft global illumination..." bo mat dau phay).
        # Doi y het thi 37/64 canh cua phim 0011 bi coi la thieu, va bi noi
        # them mot cau nua — hai cau phong cach cung dung, chong len nhau.
        # Nen chi doi MENH DE DAU, phan xac dinh net ve.
        dau = duoi.split(",")[0].strip()
        if dau and dau.lower() in chu.lower():
            continue
        i = chu.find(NEO)
        than, sau = (chu[:i], chu[i:]) if i > 0 else (chu, "")
        than = _LAC.sub("", than).rstrip().rstrip(",;.")
        s["img_prompt"] = than + ", " + duoi + (("\n" + sau) if sau else "")
        doi += 1
    if doi and emit is not None:
        emit({"type": "event", "event": "progress", "progress": 0.0,
              "message": "Ep phong cach kenh vao {0} canh thieu hoac lac "
                         "phong cach.".format(doi)})
    return doi


def _khoa_nhan_dang(scenes, characters, locations=()) -> None:
    """Rang buoc anh tham chieu vao loi nhac theo loi OpenStory.

    Hai phan, dung thu tu nay:

      1. RANG BUOC TRONG CAU — moi cho loi nhac goi ten `nv4` doi thanh
         `nv4 (Image 1)`, dung ngay tai cho no xuat hien trong cau van. So
         trong ngoac = vi tri anh trong `reference_files` = thu tu anh duoc
         gui len, nen no luon tro dung tam.
      2. CHU THICH CUOI — mot dong `_GUARD_NHAN_DANG`, cong mot dong cho
         nhung anh ma cau van KHONG nhac toi (thuong la boi canh). Anh da
         duoc nhac trong cau thi khong lap lai o duoi.

    Khac ban cu o cho quan trong nhat: KHONG con doan van liet ke tung net
    phai chep. Xem ghi chu o `_GUARD_NHAN_DANG` de biet vi sao va do duoc bao
    nhieu.
    """
    dan = {c["id"]: c for c in characters}
    noi = {l["id"]: l for l in locations}
    for s in scenes:
        try:
            ids = json.loads(s.get("reference_files") or "[]")
        except ValueError:
            ids = [x.strip() for x in str(s.get("reference_files") or "").split(",")]
        ids = [os.path.splitext(os.path.basename(str(i)))[0] for i in ids if str(i).strip()]
        ids = [i for i in ids if i in dan or i in noi]
        if not ids or "REFERENCE IMAGES are attached" in str(s.get("img_prompt") or ""):
            continue
        chu = str(s.get("img_prompt") or "")
        da_nhac = []
        for k, i in enumerate(ids, 1):
            # ═══ CHI RANG BUOC LAN NHAC DAU ═══
            #
            # Moi "(Image N)" la mot lenh dat tam anh ay vao khung. Ban cu thay
            # MOI lan nhac, nen mot loi nhac ke ten con meo hai lan thanh hai
            # lenh dat -> may ve HAI CON MEO. Chu du an xem phim 0003: *"luc
            # thi 1 con meo luc thi 2 con meo"*. Do tren chinh bang canh ay:
            # 5/29 canh rang buoc lap, canh 27 rang buoc `nv1` BA lan.
            #
            # Lan dau gan so; nhung lan sau goi bang cach nguoi ta van goi
            # ("the tabby cat") — van hieu la cung mot nhan vat, ma khong con
            # la mot lenh dat anh nua.
            goi_lai = _goi_lai(i, dan, noi)
            ngoac = re.compile(r"\s*\(\s*" + re.escape(i) + r"\s*(\("
                               + re.escape(i) + r"\.png\))?\s*\)")
            chu, n1 = ngoac.subn(" (Image {0})".format(k), chu, count=1)
            chu, _ = ngoac.subn("", chu)
            tran = re.compile(r"\b" + re.escape(i) + r"\b(?!\s*\(Image)")
            chu, n2 = tran.subn("{0} (Image {1})".format(i, k), chu, count=1)
            chu, _ = tran.subn(goi_lai, chu)
            da_nhac.append(bool(n1 or n2))
        dong = ["", "REFERENCE IMAGES are attached, in this order:"]
        for k, i in enumerate(ids, 1):
            # Anh nao cau van da rang buoc thi khong lap lai o duoi — chu
            # thich chi de cuu nhung anh khong ai nhac (thuong la boi canh).
            if not da_nhac[k - 1]:
                dong.append("- Image {0} = {1}".format(k, _mo_ta_tham_chieu(i, dan, noi)))
        dong.append(_GUARD_NHAN_DANG)
        if any(i in noi for i in ids):
            dong.append(_LUAT_DUNG_CHAN)
        # Id con lai (nhan vat thu ba tro di) khong co anh kem -> thay bang
        # vai, dung de tro cho may ve. Xem `_thay_id_khong_co_anh`.
        chu = _thay_id_khong_co_anh(chu, set(ids), dan, noi)
        s["img_prompt"] = chu + "\n".join(dong)
        s["video_prompt"] = _thay_id_khong_co_anh(
            _khoa_video(s.get("video_prompt"), ids, dan), set(ids), dan, noi)


def _bo_enrich(goi) -> Callable:
    """DUONG LUI: canh da cat theo dong ho roi, chi con nho AI viet loi nhac."""
    dem = {"value": 0}

    def enrich(batch, context):
        dem["value"] += 1
        compact = [{"scene_id": s["scene_id"], "srt_text": s["srt_text"],
                    "duration": s["duration"]} for s in batch]
        # `{{scenes:[...]}}`: hai dau ngoac vi chuoi nay di qua `.format` — mot
        # dau la KeyError 'scenes' ngay tren duong lui, tuc duong lui chua bao
        # gio chay duoc (pyflakes chi ra 24/08/2026).
        loi_nhac = "Tra JSON {{scenes:[...]}}. Giu scene_id; moi scene co img_prompt va video_prompt tieng Anh chi tiet. " \
                   "Khong doi timing. Context: {0}\nScenes: {1}".format(
                       _boi_canh_chu(context, ""),
                       json.dumps(compact, ensure_ascii=False))
        data = loc_json(goi(loi_nhac, "batch-{0}".format(dem["value"])))
        return data.get("scenes", []) if isinstance(data, dict) else data

    return enrich


def _apply_creative(batch, proposed):
    if not isinstance(proposed, list): raise ValueError("LLM scenes phai la danh sach")
    by_id = {int(item.get("scene_id")): item for item in proposed if isinstance(item, dict) and item.get("scene_id")}
    for scene in batch:
        item = by_id.get(scene["scene_id"])
        if not item: raise ValueError("LLM thieu scene_id {0}".format(scene["scene_id"]))
        for key in CREATIVE: scene[key] = str(item.get(key) or "")
        if not scene["img_prompt"] or not scene["video_prompt"]:
            raise ValueError("Scene {0} thieu img_prompt/video_prompt".format(scene["scene_id"]))
        scene["prompt_json"] = json.dumps({key: scene[key] for key in CREATIVE}, ensure_ascii=False)


#: Ke hoach dao dien la KICH BAN PHAN CANH co tham quyen, khong phai goi y.
#:
#: Chu du an 25/08/2026: *"nó như là làm phim vậy, cần phải biết là bối cảnh nào
#: nhân vật nào; không thể nhân vật này ở bối cảnh này, ảnh sau lại ở bối cảnh
#: khác"*. Buoc chia canh chay theo khuc 100 giay, moi khuc mot luot AI khong
#: nhin thay khuc khac, nen no hay tu doi cho. Sau khi chia, MA ep lai: canh
#: nao rơi vao beat nao thi lay dung `location` va `characters` cua beat do.
def _ep_theo_ke_hoach(scenes, ke_hoach) -> Dict[str, int]:
    beats = []
    for b in ke_hoach or []:
        try:
            beats.append((set(range(int(b["srt_from"]), int(b["srt_to"]) + 1)), b))
        except (KeyError, TypeError, ValueError):
            continue
    doi = {"boi_canh": 0, "nhan_vat": 0}
    if not beats:
        return doi
    for s in scenes:
        try:
            idx = {int(i) for i in (s.get("srt_indices") or [])}
        except (TypeError, ValueError):
            idx = set()
        if not idx:
            continue
        trung, b = max(((len(idx & dong), bt) for dong, bt in beats), key=lambda x: x[0])
        if trung == 0:
            continue
        loc = str(b.get("location") or "").strip()
        if loc and str(s.get("location_used") or "").strip() != loc:
            s["location_used"] = loc
            doi["boi_canh"] += 1
        # HOP nhan vat cua beat voi nhan vat canh da khai — khong THAY. Do
        # 25/08/2026 (canh 14): beat ghi thieu meo, ban cu ghi de -> canh mat
        # tham chieu meo dù prompt co meo -> meo ngau nhien. Cung mot nhan vat
        # o hai giai doan (nv4/nv4b) thi giu id cua beat (dung thoi diem).
        nv = [x for x in str(b.get("characters") or "").replace(",", " ").split() if x]
        cu = [x for x in str(s.get("characters_used") or "").replace(",", " ").split() if x]
        if nv:
            goc_beat = {_goc_cua_id(x) for x in nv}
            hop = list(nv) + [x for x in cu if x not in nv and _goc_cua_id(x) not in goc_beat]
            if set(hop) != set(cu):
                s["characters_used"] = ", ".join(hop)
                doi["nhan_vat"] += 1
    return doi


_goc_cua_id = goc_cua_id


def _so_lan_doi_boi_canh(scenes) -> int:
    """So lan bối cảnh đổi giữa hai cảnh liền nhau — con so de soi 'nhảy chỗ'."""
    truoc, n = None, 0
    for s in scenes:
        loc = str(s.get("location_used") or "").strip()
        if truoc is not None and loc != truoc:
            n += 1
        truoc = loc
    return n


def _validate_coverage(cues, scenes):
    """Moi dong phu de phai nam trong dung mot canh, khong ho khong nhay coc.

    Mot canh bi cat lam nhieu phan (vi qua tran do dai) thi cac phan cung mang
    day chi so cua canh goc — nen phai bo trung truoc khi so, khong thi canh bi
    cat lai bi bao la "phu de trung".
    """
    # Bo trung tren TOAN BO duong di, khong chi bo trung lien ke.
    #
    # Mot canh dai qua tran bi cat lam N phan, va MOI PHAN mang ca danh sach
    # chi so cua canh goc (xem `core/chia_canh.py`, khoa `_cue`). Voi canh cat
    # lam bon phan thi duong di la 1..10, 1..10, 1..10, 1..10 — kieu bo trung
    # cu chi chan duoc hai chi so canh nhau, nen tu phan thu hai tro di no dem
    # lai tu dau va bao "coverage khong dat 100%" cho mot bang canh hoan toan
    # binh thuong.
    found, da_thay = [], set()
    for scene in scenes:
        for index in scene["srt_indices"]:
            if index in da_thay:
                continue
            da_thay.add(index)
            found.append(index)
    if found != list(range(1, len(cues) + 1)):
        raise ValueError("Scene coverage khong dat 100%")
    if [s["scene_id"] for s in scenes] != list(range(1, len(scenes) + 1)):
        raise ValueError("scene_id khong lien tuc")


def render_workbook(path: Path, manifest: Mapping[str, Any]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    book = Workbook(); scenes_sheet = book.active; scenes_sheet.title = "scenes"
    for col, name in enumerate(SCENE_COLUMNS, 1):
        cell = scenes_sheet.cell(1, col, name); cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="70AD47")
    for row, scene in enumerate(manifest["scenes"], 2):
        for col, name in enumerate(SCENE_COLUMNS, 1): scenes_sheet.cell(row, col, scene.get(name, ""))
    chars_sheet = book.create_sheet("characters")
    for col, header in enumerate(CHARACTER_HEADERS, 1): chars_sheet.cell(1, col, header)
    for row, nv in enumerate(manifest.get("characters") or [], 2):
        # `image_file = <id>.png` + `status = pending` khop quy uoc tab Tu dong,
        # de khach thay ro moi nhan vat can mot anh tham chieu (chua co thi
        # pending), va VE3 mo file len la nhan ra ngay. nv1 co dinh (anh khach
        # dua) thi `done`.
        cid = str(nv.get("id") or "")
        gia_tri = {"id": cid, "role": nv.get("role", ""), "name": nv.get("name", ""),
                   "english_prompt": nv.get("english_prompt", ""),
                   "vietnamese_prompt": nv.get("vietnamese_prompt", ""),
                   "image_file": "{0}.png".format(cid) if cid else "",
                   "status": "done" if nv.get("co_dinh") else "pending",
                   "gender": nv.get("gender", ""), "age": nv.get("age", ""),
                   "notes": nv.get("notes", ""),
                   "sheet_prompt": nv.get("sheet_prompt", "")}
        for col, header in enumerate(CHARACTER_HEADERS, 1): chars_sheet.cell(row, col, gia_tri.get(header, ""))

    loc_sheet = book.create_sheet("locations")
    for col, header in enumerate(LOCATION_HEADERS, 1): loc_sheet.cell(1, col, header)
    for row, lc in enumerate(manifest.get("locations") or [], 2):
        lid = str(lc.get("id") or "")
        gia_tri = dict(lc, image_file="{0}.png".format(lid) if lid else "", status="pending")
        for col, header in enumerate(LOCATION_HEADERS, 1): loc_sheet.cell(row, col, gia_tri.get(header, ""))

    st = book.create_sheet("story")
    for col, header in enumerate(STORY_HEADERS, 1): st.cell(1, col, header)
    phim = manifest.get("story") or {}
    for row, seg in enumerate(phim.get("segments") or [], 2):
        gia_tri = dict(seg, genre=phim.get("genre", ""), arc=phim.get("arc", ""),
                       context_lock=phim.get("context_lock", ""))
        for col, header in enumerate(STORY_HEADERS, 1): st.cell(row, col, gia_tri.get(header, ""))

    plan = book.create_sheet("director_plan")
    for col, header in enumerate(PLAN_HEADERS, 1): plan.cell(1, col, header)
    for row, b in enumerate(manifest.get("director_plan") or [], 2):
        for col, header in enumerate(PLAN_HEADERS, 1): plan.cell(row, col, b.get(header, ""))

    tb = book.create_sheet("thumbnail")
    for col, header in enumerate(THUMB_HEADERS, 1): tb.cell(1, col, header)
    for row, t in enumerate(manifest.get("thumbnails") or [], 2):
        gia_tri = dict(t, thumb_text=manifest.get("thumb_text", ""), title=manifest.get("title", ""))
        for col, header in enumerate(THUMB_HEADERS, 1): tb.cell(row, col, gia_tri.get(header, ""))

    mu = book.create_sheet("music")
    for col, header in enumerate(MUSIC_HEADERS, 1): mu.cell(1, col, header)
    for row, m in enumerate(manifest.get("music") or [], 2):
        for col, header in enumerate(MUSIC_HEADERS, 1): mu.cell(row, col, m.get(header, ""))
    temp = path.with_suffix(".xlsx.tmp"); book.save(temp); os.replace(str(temp), str(path))


def _path(value, name):
    if not isinstance(value, Mapping) or not isinstance(value.get("path"), str): raise ValueError(name + " phai la artifact")
    path = Path(value["path"])
    if not path.is_file(): raise ValueError("Khong tim thay " + name)
    return path
def _optional_json(value):
    if value is None: return {}
    if isinstance(value, Mapping) and isinstance(value.get("path"), str): return json.loads(Path(value["path"]).read_text("utf-8"))
    return dict(value) if isinstance(value, Mapping) else {}
def _artifact_id(value): return str(value.get("artifact_id") or "") if isinstance(value, Mapping) else ""
def main():
    try: emit({"type":"result","output":handle(json.loads(sys.stdin.readline()))}); return 0
    except Exception as exc: sys.stderr.write(str(exc)+"\n"); return 2
if __name__ == "__main__": raise SystemExit(main())
