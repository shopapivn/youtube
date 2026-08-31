"""Tab **Prompt Visuals**: file giọng đọc → phụ đề → prompt từng cảnh → Excel.

Chủ dự án, 14/08/2026: *"từ mp3 voice ra srt rồi sau đó chạy ra các prompt trong
excel"*, làm theo lối bảng cảnh. Và 24/08/2026: *"thiết kế theo kiểu giống ở
bên tự động, kiểu đi từng bước"* — nên trang này bày BỐN BƯỚC, trên xuống
dưới theo thứ tự làm (26/08/2026 gom từ năm: *"mọi thứ hơi khó và trùng lặp
cũng như loạn quá"* — phong cách từng chọn được ở ba chỗ, nút lưu ở hai):

    Bước 1  đưa giọng đọc vào (mp3), kèm kịch bản .txt nếu có
    Bước 2  MỌI thứ về hình: dùng lại kênh/mẫu, hay chọn phong cách mới;
            nhân vật; ⚙ Nâng cao = đúng lời gửi AI (prompt phong cách, prompt
            chia cảnh với mạch chia là hai con số ở đầu) — sửa được, lưu mẫu
    Bước 3  tạo prompt — thấy rõ khâu nào đang chạy, như bảng khâu tab Tự động
    Bước 4  (hiện khi có kết quả) xem & chỉnh prompt từng cảnh, rồi THỬ thật
            1–3 cảnh (ảnh + clip) trước khi mang Excel sang tab Ảnh & Video

Cột của file Excel sinh ra **trùng tên với bảng cảnh chuẩn** (`scenes`,
`characters`, `director_plan`, `thumbnail`), nên mở thẳng bằng VE3 được.

Phần nghĩ nằm ở `core/prompt_visuals.py` (thuần tuý, test được) và ở hai tool
trong `tool-catalog/`. Tệp này chỉ dựng nút, hỏi cho rõ trước khi tiêu tiền, và
đổ kết quả ra màn hình.

LƯU Ý tiền: bước viết prompt và bước thử đều **tiêu ví ShopAPI** — trang nói
giá bằng lời TRƯỚC nút bấm và không bao giờ tự chạy khi mở tab. LƯU Ý nhịp:
việc ảnh/video của bước thử đi qua `app.start_batch` — cùng một hàng đợi và
cùng bộ giữ nhịp với mọi tab khác, KHÔNG tự hỏi máy chủ một câu nào riêng, nên
chạy thử ở đây không giành đường của tab Tự động hay Ảnh & Video đang chạy.
"""

from __future__ import annotations

import json
import os
import threading
from typing import Dict, List, Optional, Tuple

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QAbstractItemView, QComboBox, QFileDialog, QHeaderView, QLabel,
    QPlainTextEdit, QProgressBar, QScrollArea, QTableWidget, QTableWidgetItem,
    QTabWidget, QVBoxLayout, QWidget,
)

from core.anh_len import link_dung_lai_duoc, tai_len
from core.jobs import STATUS_CANCELLED, STATUS_DONE, STATUS_FAILED, JobSpec
from core.money import format_vnd
from core.pricing import KIND_IMAGE, KIND_VIDEO, hold_for_image, hold_for_video
from core.chia_canh import KHUON_MAC_DINH, nhip_tu_khuon
from core.prompt_visuals import (
    CHE_DO_CAN_ANH_NV, CHE_DO_KE, CHO_TRONG_KHUON_CHIA, DUOI_CHAN_DUNG,
    LOI_NHAC_XAY_PHONG_CACH, PhongCach, bia_de_xem, boi_canh_de_xem,
    canh_de_xem, cau_thieu_gi, chi_dan_tu_bo, chi_dan_tu_tra_loi_ai, dan_de_xem,
    doi_thiet_ke_nhan_vat, dung_boi_canh, dung_workflow, goc_cua_id,
    ke_hoach_de_xem, khuon_chia_dung_duoc, liet_ke_phong_cach,
    loi_nhac_thiet_ke_lai, man_de_xem, nhac_de_xem, tom_tat_dan,
)
from core.validate import check_image, check_video

from . import theme
# Mượn NGUYÊN bộ chọn phong cách của hộp Tạo kênh (tab Tự động): bảng 14 phong
# cách + thẻ ảnh mẫu + cửa sổ xem to (3 ảnh + video mẫu, miễn phí, ship kèm
# tool). Khách đã học cách dùng nó một lần ở Tạo kênh — ở đây gặp lại đúng nó.
from .kenh import (
    PHONG_CACH, _TheHinh, _XemPhongCach, _anh_mau_cua, _mau_xem_duoc,
    _video_mau_cua_nhieu,
)
from .thu_vien_ket_qua import ThuVienKetQua
from .widgets import (
    ChonThuMuc, HangXuongDong, mo_thu_muc, nhan, nut_chinh, nut_phu, the,
    tieu_de_trang,
)

__all__ = ["TrangPromptVisuals"]

#: Đuôi file giọng đọc nhận vào.
DUOI_TIENG = ("*.mp3", "*.wav", "*.m4a", "*.aac", "*.flac", "*.ogg")

#: Engine video DUY NHẤT của tab này. Chủ dự án 24/08/2026: *"video thì ở đây
#: là Veo 3, không dùng Seedance"*. Cảnh được cắt theo trần 8 giây của Veo 3
#: và clip thử ở Bước 4 cũng gửi Veo 3 — không có ô nào để chọn khác.
ENGINE_PV = "veo3"

#: Ba thứ trước đây nằm trong "Tuỳ chọn thêm", nay là số cố định. Chủ dự án
#: 24/08/2026: *"tiếng trong file ???? tao tưởng là mode sẽ tự đoán… chất lượng
#: prompt cũng thế mặc định đi đừng để khách hàng khó dùng"*. Máy tự đoán tiếng
#: (faster-whisper đoán từ 30 giây đầu), prompt viết bằng Sonnet, và luôn dựng
#: dàn nhân vật giữ xuyên suốt — khách không phải quyết gì ngoài bấm nút.
NGON_NGU_PV = "auto"
MO_HINH_PV = "claude-sonnet-5"
NHAT_QUAN_PV = True


class TrangPromptVisuals(QWidget):
    def __init__(self, app):
        super().__init__()
        self._app = app
        self._files: List[str] = []
        self._huy: Optional[threading.Event] = None
        self._da_xong: List[str] = []
        self._thu_muc_da_xuat = ""
        #: {tên hiển thị: đường dẫn file Excel} của các file vừa tạo xong, để
        #: khách chọn xem prompt từng cảnh ngay trong tab.
        self._ket_qua_xem: "dict[str, str]" = {}
        #: Các cảnh đang hiện trong bảng (theo đúng thứ tự dòng) — giữ để khi
        #: khách sửa prompt còn biết dòng nào là cảnh số mấy mà ghi lại vào Excel.
        self._canh_hien: List[dict] = []
        #: {mã: PhongCach} của các phong cách đang hiện trong ô chọn — giữ để
        #: khi chạy còn tra được khối chỉ dẫn tiếng Anh của mã đã chọn.
        self._phong_cach: "dict[str, PhongCach]" = {}
        #: Sổ của BƯỚC 5 (thử vài cảnh thật): khoá việc → số DÒNG trong bảng
        #: xem (bước 4). Tách ảnh/video hai sổ vì một dòng có cả hai việc.
        self._thu_dong_anh: Dict[str, int] = {}
        self._thu_dong_video: Dict[str, int] = {}
        #: Ảnh thử vừa xong, chờ nối sang clip: dòng → (đường ảnh trên đĩa,
        #: link máy chủ nếu dùng lại được). Rút ở `cuoi_nhip` — một ảnh có thể
        #: phát nhiều sự kiện liền nhau, gửi ngay trong `nhan_su_kien` là trùng.
        self._thu_cho_noi: Dict[int, Tuple[str, str]] = {}
        #: Ảnh khách tải lên đi kèm phong cách đang chọn (khi AI xây từ ảnh).
        self._anh_mau_chon: List[str] = []
        #: Ảnh tham chiếu đang tạo: khoá việc → (id nhân vật/bối cảnh, Excel, prompt).
        self._tc_dang_cho: Dict[str, Tuple[str, str, str]] = {}
        #: Id đã được thử lại một lần sau khi bị từ chối — không thử vô hạn.
        self._tc_da_thu_lai: set = set()
        #: Nhân vật (id gốc) đã được AI thiết kế lại một lần — không lặp vô hạn.
        self._tc_da_thiet_ke_lai: set = set()
        #: Id KHÔNG tạo được (bị từ chối cả hai lần) — nói thật với khách thay vì
        #: báo "đủ". Đo 25/08/2026: mèo đi hia 3D bị bộ lọc nhà cung cấp chặn 15
        #: cách viết liên tiếp; bản cũ im lặng, khách sẽ chạy cả phim thiếu mèo.
        self._tc_thieu: List[str] = []

        # ═══ MÀN HÌNH ĐẦU CHỈ CÓ BA THẺ, MỖI THẺ MỘT VIỆC ═══
        #
        # Chủ dự án 24/08/2026: *"tao cần đơn giản, hiệu quả, dễ sử dụng, tối
        # giản… tinh gọn"*; 26/08/2026: *"mọi thứ hơi khó và trùng lặp cũng như
        # loạn quá"*. Bản trước có năm thẻ và phong cách chọn được ở BA chỗ
        # (ô "đã lưu" ở Bước 1, lưới thẻ ở Bước 2, ô Kênh trong Nâng cao), nút
        # "Lưu phong cách" ở hai chỗ, còn "Nhân vật" nằm ở Bước 1 dù nó là
        # chuyện hình ảnh. Giờ: Bước 1 chỉ có file; Bước 2 là MỌI thứ về hình
        # (phong cách, nhân vật, và Nâng cao: prompt phong cách + mạch chia
        # cảnh + prompt chia cảnh); Bước 3 một nút; Bước 4 chỉ hiện khi có kết
        # quả, gồm cả xem/sửa lẫn thử vài cảnh thật.
        doc = QVBoxLayout(self)
        doc.setContentsMargins(24, 20, 24, 20)
        doc.setSpacing(14)
        doc.addWidget(tieu_de_trang(
            "Prompt Visuals",
            "mp3 vào → Excel đủ prompt ảnh + video từng cảnh ra."))
        doc.addWidget(self._the_nhap())
        doc.addWidget(self._the_phong_cach())
        doc.addWidget(self._the_chay())
        self._the_xem_w = self._the_xem()
        self._the_xem_w.hide()
        doc.addWidget(self._the_xem_w)
        doc.addStretch(1)

        # Nhật ký gấp sẵn: dòng trạng thái ở Bước 3 đã nói điều khách cần;
        # nhật ký đầy đủ chỉ mở khi có chuyện phải soi.
        self._nut_log = nut_phu("Nhật ký ▾", lambda: self._bat_tat(
            self._log, self._nut_log, "Nhật ký"), rong=120)
        doc.addWidget(self._nut_log, 0, Qt.AlignLeft)
        self._log = QPlainTextEdit()
        self._log.setReadOnly(True)
        self._log.setFixedHeight(120)
        self._log.setStyleSheet(
            "background:{0}; border:1px solid {1}; border-radius:8px;"
            " color:{2}; font-size:12px;".format(theme.THE_MO, theme.VIEN,
                                                 theme.CHU_MO))
        self._log.hide()
        doc.addWidget(self._log)
        self._ve_trang_thai()

    # ── Dựng giao diện ───────────────────────────────────────────────────────

    def _the_nhap(self) -> QWidget:
        """Bước 1: ĐẦU VÀO — mp3 bắt buộc, kịch bản .txt nếu có. Chỉ thế.

        Chủ dự án 24/08/2026 vẽ đúng thứ tự khách làm: *"đẩy mp3 rồi txt (nếu
        có) rồi đến chọn phong cách"*. Ô "Phong cách đã lưu" và ô "Nhân vật"
        từng nằm ở đây — hai chuyện về HÌNH lẫn vào bước về TIẾNG, và phong
        cách thành ra chọn được ở hai chỗ. Chúng đã dọn xuống Bước 2.
        """
        khung = the()
        v = QVBoxLayout(khung)
        v.setContentsMargins(18, 16, 18, 18)
        v.setSpacing(10)
        v.addWidget(nhan("Bước 1 — Giọng đọc", "h2"))

        # Hàng nút: chọn file, "+ kịch bản" (đường phụ, mở ra ô dán), bỏ file.
        # Tên file nằm DÒNG RIÊNG bên dưới: nhãn nhét vào hàng xuống dòng bị
        # bóp còn một chữ mỗi dòng (ảnh chụp 24/08/2026 — "Chưa / chọn / file").
        hang = HangXuongDong()
        hang.addWidget(nut_chinh("Chọn file mp3…", self._chon_file))
        self._nut_kich_ban = nut_phu("+ kịch bản .txt ▾", lambda: self._bat_tat(
            self._khoi_kich_ban, self._nut_kich_ban, "+ kịch bản .txt"),
            rong=170)
        self._nut_kich_ban.setToolTip(
            "Không bắt buộc. Có kịch bản gốc thì prompt bám đúng tên riêng, "
            "thuật ngữ trong bài hơn là chỉ nghe từ giọng đọc.")
        hang.addWidget(self._nut_kich_ban)
        self._nut_bo_file = nut_phu("Bỏ file", self._bo_file, rong=90)
        self._nut_bo_file.hide()
        hang.addWidget(self._nut_bo_file)
        v.addLayout(hang)
        self._nhan_file = self._chu_phu("Chưa chọn file nào.")
        v.addWidget(self._nhan_file)

        self._khoi_kich_ban = QWidget()
        self._khoi_kich_ban.setMinimumWidth(1)
        vk = QVBoxLayout(self._khoi_kich_ban)
        vk.setContentsMargins(0, 0, 0, 0)
        vk.setSpacing(6)
        vk.addWidget(self._chu_phu(
            "Kịch bản .txt không bắt buộc — có thì prompt bám đúng tên riêng, "
            "thuật ngữ trong bài, chuẩn xác hơn chỉ nghe từ giọng đọc."))
        hang2 = HangXuongDong()
        hang2.addWidget(nut_phu("Nạp từ file .txt…", self._nap_kich_ban,
                                rong=170))
        hang2.addWidget(nut_phu("Xoá", self._xoa_kich_ban, rong=80))
        vk.addLayout(hang2)
        self._o_kich_ban = QPlainTextEdit()
        self._o_kich_ban.setPlaceholderText("…hoặc dán kịch bản vào đây.")
        self._o_kich_ban.setFixedHeight(64)
        self._o_kich_ban.setMinimumWidth(1)
        vk.addWidget(self._o_kich_ban)
        self._khoi_kich_ban.hide()
        v.addWidget(self._khoi_kich_ban)
        return khung

    # ── Nhân vật cố định (loại 1, 2) ─────────────────────────────────────────

    def _che_do_ke(self) -> str:
        return str(self._o_che_do.currentData() or "tu_xay")

    def _ve_che_do(self) -> None:
        """Nói rõ cách kể đang chọn cần gì; ẩn/hiện nút tải ảnh cho đúng."""
        from PyQt5.QtGui import QPixmap  # noqa: PLC0415

        ma = self._che_do_ke()
        can_anh = ma in CHE_DO_CAN_ANH_NV
        self._nut_anh_nv.setVisible(can_anh)
        mo_ta = next((m for k, _t, m in CHE_DO_KE if k == ma), "")
        if can_anh and self._anh_nv and os.path.isfile(self._anh_nv):
            px = QPixmap(self._anh_nv)
            if not px.isNull():
                self._anh_nv_xem.setPixmap(px.scaled(
                    64, 64, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            self._anh_nv_xem.setVisible(True)
            self._nhan_che_do.setText("{0} Ảnh nhân vật: {1}.".format(
                mo_ta, os.path.basename(self._anh_nv)))
        else:
            self._anh_nv_xem.setVisible(False)
            self._nhan_che_do.setText(
                mo_ta + (" ⚠ Chưa có ảnh nhân vật — bấm “Tải ảnh nhân vật…”."
                         if can_anh else ""))

    def _chon_anh_nv(self) -> None:
        duong, _ = QFileDialog.getOpenFileName(
            self, "Chọn ảnh nhân vật chính (nv1)", "",
            "Ảnh (*.png *.jpg *.jpeg *.webp);;Mọi loại file (*)")
        if duong:
            self._anh_nv = duong
            self._ve_che_do()

    def _dat_anh_nv_tu_kenh(self, ma: str) -> None:
        """Chọn phong cách kênh/bộ vẽ ở Bước 1 → lấy luôn nv1.png của nó làm
        nhân vật cố định, và nếu đang để "AI tự xây" thì chuyển sang loại 1."""
        p = self._anh_nhan_vat(ma) if ma.startswith(("kenh:", "ve:")) else ""
        if not p:
            return
        self._anh_nv = p
        if self._che_do_ke() == "tu_xay":
            i = self._o_che_do.findData("mot_nhan_vat")
            if i >= 0:
                self._o_che_do.setCurrentIndex(i)
        self._ve_che_do()

    def _the_phong_cach(self) -> QWidget:
        """Bước 2: MỌI THỨ VỀ HÌNH ở một thẻ — phong cách, nhân vật, và Nâng cao.

        Chủ dự án 24/08/2026: *"phải cho khách xem demo các style và cụ thể
        prompt để khách tự có thể tối ưu"*; 26/08/2026: *"sau khi khách chọn
        phong cách, chọn all mọi thứ thì ở nâng cao có thể chỉnh được prompt,
        ví dụ là mạch chia là 3-8s hay là chia kiểu khác"*.

        Thứ tự từ trên xuống đúng như khách nghĩ:

        1. **Dùng lại** thứ đã có (kênh của tab Tự động, bộ vẽ trong khuôn,
           phong cách đã lưu) — chọn là xong, hai tab dưới gấp lại;
        2. hoặc **chọn mới**: lưới thẻ có ảnh + video mẫu / AI xây từ ảnh;
        3. **Nhân vật**: cách kể + ảnh nhân vật (chuyện hình, nên nằm đây);
        4. **⚙ Nâng cao** (gấp sẵn): prompt phong cách, **prompt chia cảnh**
           (đúng lời gửi AI, sửa được — mạch chia là hai con số ở đầu), đồng
           bộ với kênh.

        Một nút "Lưu để dùng lại" duy nhất — bản trước có hai.
        """
        khung = the()
        v = QVBoxLayout(khung)
        v.setContentsMargins(18, 16, 18, 18)
        v.setSpacing(10)
        v.addWidget(nhan("Bước 2 — Phong cách & nhân vật", "h2"))

        # ═══ 1. DÙNG LẠI THỨ ĐÃ CÓ ═══
        hang_lai = HangXuongDong()
        hang_lai.addWidget(nhan("Dùng lại:", "phu"))
        self._o_kenh = QComboBox()
        self._o_kenh.setMinimumWidth(1)
        self._o_kenh.setToolTip(
            "Kênh đã tạo ở tab Tự động, bộ vẽ trong khuôn, hay phong cách bạn "
            "đã lưu ở đây. Chọn một cái là xong Bước 2; để trống thì chọn mới "
            "ở dưới.")
        self._o_kenh.currentIndexChanged.connect(lambda _i: self._chon_tu_combo())
        hang_lai.addWidget(self._o_kenh)
        v.addLayout(hang_lai)

        # ═══ 2. CHỌN MỚI: HAI TAB ═══
        #
        # Chủ dự án 24/08/2026: *"bước 2 thì có 2 tab: 1 tab là chọn phong cách
        # sẵn và 1 tab là AI xây… cho khách tải vài ảnh và từ đó dùng API để
        # xác định được phong cách"*. Dòng "Đang chọn" và nút nâng cao nằm
        # DƯỚI hai tab, dùng chung — chọn ở tab nào cũng đổ về một chỗ.
        self._tab = QTabWidget()
        self._tab.setMinimumWidth(1)
        self._tab.addTab(self._tab_chon_san(), "Chọn phong cách có sẵn")
        self._tab.addTab(self._tab_ai_tu_anh(), "AI xây phong cách từ ảnh của bạn")
        v.addWidget(self._tab)

        # Thẻ minh hoạ phong cách đang chọn (ảnh + video mẫu / ảnh nhân vật /
        # ảnh khách tải) — luôn có hình, kể cả khi hai tab đã gấp.
        self._khoi_minh_hoa = QWidget()
        self._khoi_minh_hoa.setMinimumWidth(1)
        self._khoi_minh_hoa.setLayout(HangXuongDong(8))
        self._khoi_minh_hoa.hide()
        v.addWidget(self._khoi_minh_hoa)

        # Câu "Đang chọn" nổi hơn chữ phụ — là thứ duy nhất cho biết cú bấm vừa
        # rồi có ăn không (thẻ được chọn có thể đã cuộn khuất).
        self._nhan_phong_cach = self._chu_phu("")
        self._nhan_phong_cach.setStyleSheet(
            "color:{0}; background:{1}; border-radius:6px; padding:6px 8px;"
            .format(theme.CHU, theme.NHAN_NHAT))
        v.addWidget(self._nhan_phong_cach)

        hang_chon = HangXuongDong()
        self._nut_doi = nut_phu("Đổi phong cách khác", self._doi_phong_cach,
                                rong=180)
        self._nut_doi.hide()
        hang_chon.addWidget(self._nut_doi)
        self._nut_tu_dong = nut_phu("Để AI tự chọn", self._chon_tu_dong,
                                    rong=140)
        hang_chon.addWidget(self._nut_tu_dong)
        self._nut_luu_chung = nut_phu("💾 Lưu để dùng lại…", self._luu_mau,
                                      rong=190)
        self._nut_luu_chung.setToolTip(
            "Lưu phong cách đang chọn cùng mọi thứ trong Nâng cao (prompt phong "
            "cách, prompt chia cảnh đã sửa) — lần sau chọn ở ô “Dùng lại”.")
        hang_chon.addWidget(self._nut_luu_chung)
        self._nut_nang_cao = nut_phu("⚙ Nâng cao ▾", lambda: self._bat_tat(
            self._khoi_nang_cao, self._nut_nang_cao, "⚙ Nâng cao"), rong=140)
        self._nut_nang_cao.setToolTip(
            "Xem và sửa đúng những lời tool gửi AI: prompt phong cách và prompt "
            "chia cảnh (mạch chia 3–8 giây hay 30 giây một cảnh nằm ngay trong đó).")
        hang_chon.addWidget(self._nut_nang_cao)
        v.addLayout(hang_chon)

        # ═══ 3. NHÂN VẬT — BA CÁCH KỂ ═══
        #
        # Chủ dự án 24/08/2026: loại 1 một nhân vật cố định của kênh; loại 2
        # nhân vật cố định + nhân vật/bối cảnh tham chiếu khác; loại 3 AI tự
        # xây theo nội dung. Loại 1, 2 cần MỘT ảnh nhân vật (nv1.png) — chọn
        # kênh ở ô "Dùng lại" thì lấy luôn ảnh nhân vật của kênh.
        hang_nv = HangXuongDong()
        hang_nv.addWidget(nhan("Nhân vật:", "phu"))
        self._o_che_do = QComboBox()
        self._o_che_do.setMinimumWidth(1)
        for ma, ten, mo_ta in CHE_DO_KE:
            self._o_che_do.addItem(ten, ma)
            self._o_che_do.setItemData(self._o_che_do.count() - 1, mo_ta,
                                       Qt.ToolTipRole)
        self._o_che_do.currentIndexChanged.connect(lambda _i: self._ve_che_do())
        hang_nv.addWidget(self._o_che_do)
        self._nut_anh_nv = nut_phu("Tải ảnh nhân vật…", self._chon_anh_nv,
                                   rong=170)
        hang_nv.addWidget(self._nut_anh_nv)
        self._anh_nv_xem = QLabel()
        self._anh_nv_xem.setFixedSize(64, 64)
        self._anh_nv_xem.setAlignment(Qt.AlignCenter)
        self._anh_nv_xem.hide()
        hang_nv.addWidget(self._anh_nv_xem)
        v.addLayout(hang_nv)
        self._nhan_che_do = self._chu_phu("")
        v.addWidget(self._nhan_che_do)
        self._anh_nv = ""
        self._ve_che_do()

        # ═══ 4. NÂNG CAO (gấp sẵn): ĐÚNG NHỮNG LỜI TOOL GỬI AI ═══
        #
        # Chủ dự án: *"cụ thể prompt để khách tự có thể tối ưu"*, *"cho khách
        # xây template sẵn để lần sau tái sử dụng"*, và 26/08/2026: *"ở nâng
        # cao có thể chỉnh được prompt, ví dụ là mạch chia là 3-8s hay là chia
        # kiểu khác"*. Ba thứ, mỗi thứ một ô, đều đi vào lượt chạy:
        #   prompt phong cách  → `visual_style_directive`
        #   prompt chia cảnh   → `storyboard_template` (khuôn AI chia cảnh +
        #                        viết prompt; chỉ gửi khi khác mặc định). Mạch
        #                        chia là HAI CON SỐ ở đầu khuôn — chủ dự án
        #                        26/08/2026 bác ô chọn: *"khống chế ở prompt gốc
        #                        để khách xem được và có thể tối ưu, ví dụ họ
        #                        muốn 30s 1 cảnh thì họ tự tối ưu được"*.
        self._khoi_nang_cao = QWidget()
        self._khoi_nang_cao.setMinimumWidth(1)
        vc = QVBoxLayout(self._khoi_nang_cao)
        vc.setContentsMargins(0, 4, 0, 0)
        vc.setSpacing(6)
        vc.addWidget(nhan("Prompt phong cách", "phu"))
        vc.addWidget(self._chu_phu(
            "Khối tiếng Anh ghép vào mọi cảnh — sửa thẳng ở đây; để trống = "
            "AI tự chọn phong cách theo nội dung."))
        self._o_chi_dan = QPlainTextEdit()
        self._o_chi_dan.setPlaceholderText(
            "Để trống = AI tự chọn phong cách theo nội dung.")
        self._o_chi_dan.setFixedHeight(76)
        self._o_chi_dan.setMinimumWidth(1)
        vc.addWidget(self._o_chi_dan)
        self._nut_xoa_mau = nut_phu("Xoá phong cách đã lưu này", self._xoa_mau,
                                    rong=230)
        self._nut_xoa_mau.hide()
        vc.addWidget(self._nut_xoa_mau, 0, Qt.AlignLeft)

        hang_khuon = HangXuongDong()
        hang_khuon.addWidget(nhan("Prompt chia cảnh", "phu"))
        self._nut_khuon_mac_dinh = nut_phu("Khôi phục mặc định",
                                           self._khoi_phuc_khuon_chia, rong=170)
        hang_khuon.addWidget(self._nut_khuon_mac_dinh)
        vc.addLayout(hang_khuon)
        vc.addWidget(self._chu_phu(
            "Đây là đúng lời tool gửi AI để chia cảnh và viết prompt ảnh + video "
            "từng cảnh — sửa thẳng để tối ưu. Mạch chia nằm ở hai dòng số đầu "
            "tiên (MIN/MAX_SECONDS_PER_SCENE): muốn 30 giây một cảnh thì đổi "
            "MAX thành 30, muốn cắt dày thì 5. Giữ nguyên các chỗ <<…>> — tool "
            "điền phụ đề, nhân vật, kế hoạch vào đó; thiếu một chỗ là không chạy."))
        self._o_khuon_chia = QPlainTextEdit()
        self._o_khuon_chia.setPlainText(KHUON_MAC_DINH)
        self._o_khuon_chia.setFixedHeight(180)
        self._o_khuon_chia.setMinimumWidth(1)
        self._o_khuon_chia.setStyleSheet("font-size:12px;")
        self._o_khuon_chia.textChanged.connect(self._ve_trang_thai_khuon)
        vc.addWidget(self._o_khuon_chia)
        self._nhan_khuon = self._chu_phu("")
        vc.addWidget(self._nhan_khuon)
        self._ve_trang_thai_khuon()

        # ═══ ĐỒNG BỘ VỚI KÊNH CỦA TAB TỰ ĐỘNG ═══
        #
        # Chiều ngược của ô "Dùng lại": phong cách khách dựng ở đây (kể cả AI
        # xây từ ảnh, prompt sửa tay, ảnh nhân vật) ghi vào `style.yaml` +
        # `nv/nv1.png` của kênh — tab Tự động dùng ngay.
        from .kenh_chon import HangKenh  # noqa: PLC0415

        vc.addWidget(nhan("Đồng bộ với kênh của tab Tự động", "phu"))
        vc.addWidget(HangKenh(
            self._app, nap=self._nap_style_tu_kenh, luu=self._luu_style_vao_kenh,
            mach_nap="Dùng đúng phong cách và ảnh nhân vật của kênh.",
            mach_luu="Ghi prompt phong cách ở ô trên vào style.yaml của kênh, "
                     "và ảnh nhân vật (nếu có) thành nv/nv1.png."))
        self._khoi_nang_cao.hide()
        v.addWidget(self._khoi_nang_cao)

        self._chon_ma = "auto"
        self._nap_phong_cach()
        self._chon_tu_dong()
        return khung

    # ── Prompt chia cảnh (Nâng cao) ──────────────────────────────────────────

    def _khoi_phuc_khuon_chia(self) -> None:
        self._o_khuon_chia.setPlainText(KHUON_MAC_DINH)

    def _khuon_chia_da_sua(self) -> str:
        """Khuôn chia cảnh khách sửa; rỗng nếu vẫn là mặc định (không gửi gì)."""
        k = self._o_khuon_chia.toPlainText()
        return "" if k.strip() == KHUON_MAC_DINH.strip() else k

    def _ve_trang_thai_khuon(self) -> None:
        k = self._o_khuon_chia.toPlainText()
        if not self._khuon_chia_da_sua():
            self._nhan_khuon.setText("Đang dùng prompt mặc định.")
            self._nhan_khuon.setStyleSheet("")
            return
        thieu = [ct for ct in CHO_TRONG_KHUON_CHIA if ct not in k]
        if thieu:
            self._nhan_khuon.setText(
                "⚠ Thiếu chỗ trống {0} — tool sẽ KHÔNG dùng bản sửa này. Thêm "
                "lại hoặc bấm “Khôi phục mặc định”.".format(", ".join(thieu)))
            self._nhan_khuon.setStyleSheet("color:{0};".format(theme.VANG))
            return
        cap = nhip_tu_khuon(k)
        self._nhan_khuon.setStyleSheet("")
        if cap is None:
            self._nhan_khuon.setText(
                "✓ Prompt chia cảnh đã sửa — sẽ dùng bản này. (Không đọc được "
                "hai dòng MIN/MAX_SECONDS_PER_SCENE: mạch chia về mặc định 3–8.)")
        else:
            self._nhan_khuon.setText(
                "✓ Prompt chia cảnh đã sửa — sẽ dùng bản này. Mạch chia: "
                "{0:.0f}–{1:.0f} giây một cảnh{2}.".format(
                    cap[0], cap[1],
                    " (clip vẫn 8 giây, cảnh dài được quay nhiều góc máy)"
                    if cap[1] > 8 else ""))

    def _tab_chon_san(self) -> QWidget:
        """Tab 1: lưới thẻ có ảnh + video mẫu, xếp dòng theo bề rộng cửa sổ.

        Dùng `HangXuongDong` (xếp dòng) thay cho lưới 3 cột cố định: cửa sổ
        rộng thì 6 thẻ một hàng (chủ dự án: *"để 1 hàng 6 phong cách"*), cửa
        sổ hẹp tự xuống 4 — không còn khoảng trống giữa các cột.
        """
        w = QWidget()
        w.setMinimumWidth(1)
        v = QVBoxLayout(w)
        v.setContentsMargins(0, 8, 0, 0)
        v.setSpacing(8)
        v.addWidget(self._chu_phu(
            "Bấm một thẻ để xem ảnh + video mẫu (miễn phí) rồi bấm “Dùng phong "
            "cách này”."))
        self._the_phong: List[_TheHinh] = []
        luoi_wrap = QWidget()
        luoi = HangXuongDong(8)
        luoi_wrap.setLayout(luoi)
        for i, (ten, kv) in enumerate(PHONG_CACH):
            slug = str(kv.get("slug", ""))
            anhs = _anh_mau_cua(i, slug)
            the_ = _TheHinh(ten, str(kv.get("_mo_ta", "")),
                            anhs[0] if anhs else "", so_anh=len(anhs),
                            so_video=len(_video_mau_cua_nhieu(slug)))
            the_.bam.connect(lambda _i=i: self._xem_phong(_i))
            luoi.addWidget(the_)
            self._the_phong.append(the_)
        cuon = QScrollArea()
        cuon.setWidgetResizable(True)
        cuon.setFrameShape(QScrollArea.NoFrame)
        cuon.setWidget(luoi_wrap)
        cuon.setFixedHeight(330)
        cuon.setMinimumWidth(1)
        v.addWidget(cuon)
        return w

    def _tab_ai_tu_anh(self) -> QWidget:
        """Tab 2: khách tải 1–5 ảnh → AI đọc ảnh, viết ra phong cách → dùng luôn.

        Một lời gọi AI duy nhất, do khách chủ động bấm. Ảnh đọc + hoá data URL
        trên luồng vẽ; gọi mạng đẩy xuống `run_bg`; kết quả về luồng vẽ mới
        chạm widget. Phong cách rút ra là một khối chỉ dẫn như mọi thẻ khác —
        sửa được ở ⚙ Nâng cao và lưu thành mẫu để lần sau chọn ở Bước 1.
        """
        w = QWidget()
        w.setMinimumWidth(1)
        v = QVBoxLayout(w)
        v.setContentsMargins(0, 8, 0, 0)
        v.setSpacing(8)
        # Ba dòng đánh số, mỗi dòng một việc — không bày ba nút cạnh nhau.
        v.addWidget(self._chu_phu(
            "1. Tải 1–5 ảnh đúng kiểu bạn muốn (ảnh kênh khác, ảnh bạn thích…)."))
        hang = HangXuongDong()
        hang.addWidget(nut_phu("Tải ảnh mẫu…", self._chon_anh_mau, rong=150))
        self._nut_bo_anh = nut_phu("Bỏ ảnh", self._bo_anh_mau, rong=90)
        self._nut_bo_anh.hide()
        hang.addWidget(self._nut_bo_anh)
        v.addLayout(hang)
        self._anh_mau: List[str] = []
        self._dai_anh_mau = QWidget()
        self._dai_anh_mau.setMinimumWidth(1)
        self._dai_anh_mau.setLayout(HangXuongDong(6))
        v.addWidget(self._dai_anh_mau)

        v.addWidget(self._chu_phu(
            "2. Nhờ AI xem ảnh và viết ra phong cách — một lượt gọi AI, không "
            "tạo ảnh. Mọi cảnh về sau theo đúng kiểu đó."))
        self._nut_xay = nut_chinh("✨ Nhờ AI xây phong cách", self._xay_tu_anh)
        self._nut_xay.setFixedWidth(240)
        self._nut_xay.setEnabled(False)
        v.addWidget(self._nut_xay, 0, Qt.AlignLeft)
        self._nhan_xay = self._chu_phu("")
        v.addWidget(self._nhan_xay)

        v.addWidget(self._chu_phu(
            "3. Ưng thì bấm “💾 Lưu để dùng lại…” ở dưới — lần sau chọn ngay ở "
            "ô “Dùng lại”, không phải tải ảnh nữa."))
        v.addStretch(1)
        return w

    # ── AI xây phong cách từ ảnh ─────────────────────────────────────────────

    def _chon_anh_mau(self) -> None:
        duongs, _ = QFileDialog.getOpenFileNames(
            self, "Chọn 1–5 ảnh mẫu của bạn", "",
            "Ảnh (*.png *.jpg *.jpeg *.webp);;Mọi loại file (*)")
        if not duongs:
            return
        self._anh_mau = (self._anh_mau + list(duongs))[:5]
        self._ve_anh_mau()

    def _bo_anh_mau(self) -> None:
        self._anh_mau = []
        self._ve_anh_mau()

    def _ve_anh_mau(self) -> None:
        from PyQt5.QtGui import QPixmap  # noqa: PLC0415

        lay = self._dai_anh_mau.layout()
        while lay.count():
            muc = lay.takeAt(0)
            if muc.widget() is not None:
                muc.widget().deleteLater()
        for d in self._anh_mau:
            o = QLabel()
            o.setFixedSize(96, 54)
            o.setAlignment(Qt.AlignCenter)
            px = QPixmap(d)
            if not px.isNull():
                o.setPixmap(px.scaled(96, 54, Qt.KeepAspectRatioByExpanding,
                                      Qt.SmoothTransformation))
            o.setToolTip(os.path.basename(d))
            o.setStyleSheet("border:1px solid {0}; border-radius:6px;".format(
                theme.VIEN))
            lay.addWidget(o)
        co = bool(self._anh_mau)
        self._nut_bo_anh.setVisible(co)
        self._nut_xay.setEnabled(co)
        self._nhan_xay.setText(
            "{0} ảnh. Bấm “Nhờ AI xây phong cách”.".format(len(self._anh_mau))
            if co else "")

    def _xay_tu_anh(self) -> None:
        if not self._anh_mau:
            return
        if self._app.client is None:
            self._app.bao_can_khoa()
            return
        from core.auto_khau import _anh_thanh_data_url  # noqa: PLC0415
        from core.goi_van_ban import khoi_anh  # noqa: PLC0415

        # ═══ ẢNH PHẢI ĐI ĐÚNG ĐỊNH DẠNG CỦA CỔNG ═══
        #
        # Cổng ShopAPI mang dáng OpenAI nhưng bên dưới là Claude: gửi ảnh kiểu
        # `image_url` là cổng LẶNG LẼ BỎ ảnh, mô hình trả "không thấy ảnh" và
        # khách thấy "chưa đọc ra phong cách" (chính lỗi chủ dự án gặp
        # 24/08/2026). `khoi_anh` dựng đúng khối ảnh cổng chuyển tới mô hình.
        phan = [{"type": "text", "text": LOI_NHAC_XAY_PHONG_CACH}]
        try:
            for d in self._anh_mau:
                duoi = os.path.splitext(d)[1].lower().lstrip(".") or "jpeg"
                kieu = "image/" + ("jpeg" if duoi == "jpg" else duoi)
                with open(d, "rb") as f:
                    phan.append(khoi_anh(_anh_thanh_data_url(f.read(), kieu)))
        except OSError as loi:
            self._nhan_xay.setText("Chưa đọc được ảnh: {0}".format(str(loi)[:80]))
            return
        tin = [{"role": "user", "content": phan}]
        client = self._app.client
        so = len(self._anh_mau)
        self._nut_xay.setEnabled(False)
        self._nhan_xay.setText("Đang xem {0} ảnh của bạn… (khoảng nửa phút)".format(so))

        def viec():
            from core.goi_van_ban import goi_van_ban  # noqa: PLC0415
            return goi_van_ban(client, tin, toi_da_token=2000)

        self._app.run_bg(
            viec, on_ok=lambda tra: self._xay_xong(tra, so),
            on_err=lambda loi: self._xay_hong(loi))

    def _xay_xong(self, tra: str, so: int) -> None:
        self._nut_xay.setEnabled(True)
        chi_dan = chi_dan_tu_tra_loi_ai(tra)
        if not chi_dan:
            # Nói thật câu AI trả về (cắt ngắn) thay vì một câu chung chung —
            # khách và người hỗ trợ nhìn là biết hỏng ở đâu.
            self._nhan_xay.setText(
                "AI chưa trả về phong cách dùng được. AI nói: “{0}”. Thử ảnh "
                "rõ hơn hoặc bấm lại một lần.".format(
                    " ".join(str(tra or "").split())[:160]))
            self._ghi("AI xây phong cách trả về: {0}".format(str(tra)[:300]))
            return
        self._dat_chon("anh:", "Phong cách từ ảnh của bạn",
                       "AI rút từ {0} ảnh bạn tải lên.".format(so), chi_dan,
                       anh_mau=list(self._anh_mau))
        self._nhan_xay.setText(
            "✓ Đã xây phong cách từ {0} ảnh. Ưng thì bấm “💾 Lưu để dùng lại…” "
            "để lần sau chọn ngay ở ô “Dùng lại”.".format(so))
        self._ghi("Đã xây phong cách từ {0} ảnh.".format(so))

    def _xay_hong(self, loi: BaseException) -> None:
        self._nut_xay.setEnabled(True)
        self._nhan_xay.setText("Chưa gọi được AI — thử lại sau ít phút.")
        self._app.show_error(loi)

    @staticmethod
    def _bat_tat(khoi: QWidget, nut, nhan_nut: str) -> None:
        """Mở/gấp một khối phụ, đổi mũi tên trên nút cho khớp."""
        mo = not khoi.isVisible()
        khoi.setVisible(mo)
        nut.setText("{0} {1}".format(nhan_nut, "▴" if mo else "▾"))

    # ── Chọn phong cách: một cửa duy nhất đổi lựa chọn ───────────────────────

    def _xem_phong(self, i: int) -> None:
        """Bấm thẻ → mở cửa sổ xem to (ảnh + video mẫu), bấm “Dùng” mới chọn.

        Chưa có mẫu nào trên đĩa thì chọn luôn, không kẹt khách ở cửa sổ trống.
        """
        ten, kv = PHONG_CACH[i]
        muc = _mau_xem_duoc(i, str(kv.get("slug", "")))
        if not muc:
            self._chon_phong(i)
            return
        hop = _XemPhongCach(ten, str(kv.get("_mo_ta", "")), muc, self)
        hop.chon.connect(lambda _i=i: self._chon_phong(_i))
        hop.exec_()

    def _chon_phong(self, i: int) -> None:
        ten, kv = PHONG_CACH[i]
        self._dat_chon("pc:" + str(kv.get("slug", "")), ten,
                       str(kv.get("_mo_ta", "")), chi_dan_tu_bo(kv))

    def _chon_tu_dong(self) -> None:
        self._dat_chon("auto", "Tự động",
                       "AI tự chọn phong cách theo nội dung lời đọc.", "")

    def _chon_tu_combo(self) -> None:
        ma = str(self._o_kenh.currentData() or "")
        if not ma:
            # Khách chọn lại "(chưa chọn)": về tự động và MỞ LẠI hai tab để
            # chọn — bản trước bỏ qua mục rỗng nên tab cứ gấp mãi.
            self._chon_tu_dong()
            return
        if ma.startswith("mau:"):
            self._ap_mau(ma[4:])
            return
        p = self._phong_cach.get(ma)
        if p is not None:
            self._dat_chon(p.ma, p.ten, p.mo_ta, p.chi_dan)
            self._dat_anh_nv_tu_kenh(ma)

    def _dat_chon(self, ma: str, ten: str, mo_ta: str, chi_dan: str,
                  anh_mau: Optional[List[str]] = None) -> None:
        """Đổi phong cách: đánh dấu thẻ, đồng bộ ô Bước 1, điền prompt, VẼ MINH HOẠ.

        Mọi đường chọn (thẻ, nút AI tự chọn, ô Bước 1, mẫu đã lưu, AI từ ảnh)
        đều đi qua đây — các nơi hiển thị không bao giờ nói khác nhau.

        `anh_mau`: ảnh khách tải lên khi nhờ AI xây — giữ lại để minh hoạ và
        để lưu kèm mẫu (chủ dự án 24/08/2026: *"chọn phong cách đã lưu thì ở
        bước 2 không có gì minh hoạ như ảnh và video"*).
        """
        self._chon_ma = ma
        self._anh_mau_chon = list(anh_mau or [])
        for j, the_ in enumerate(self._the_phong):
            the_.dat_chon(ma == "pc:" + str(PHONG_CACH[j][1].get("slug", "")))
        self._o_kenh.blockSignals(True)
        i = self._o_kenh.findData(ma)
        self._o_kenh.setCurrentIndex(i if i >= 0 else 0)
        self._o_kenh.blockSignals(False)
        self._nhan_phong_cach.setText(
            "✓ Đang chọn: {0}{1}".format(ten, " — " + mo_ta if mo_ta else ""))
        self._o_chi_dan.setPlainText(chi_dan)
        self._nut_xoa_mau.setVisible(ma.startswith("mau:"))
        # Chọn từ ô "Dùng lại" (kênh / bộ vẽ / mẫu đã lưu) thì hai tab gấp
        # lại, còn thẻ minh hoạ + dòng "Đang chọn" + nút Đổi; không chọn thì
        # hai tab mở ra để chọn.
        tu_buoc_1 = ma.startswith(("mau:", "kenh:", "ve:"))
        self._tab.setVisible(not tu_buoc_1)
        self._nut_doi.setVisible(tu_buoc_1)
        self._nut_tu_dong.setVisible(not tu_buoc_1)
        self._ve_minh_hoa(ma, ten, mo_ta)

    def _ve_minh_hoa(self, ma: str, ten: str, mo_ta: str) -> None:
        """Thẻ minh hoạ cho phong cách đang chọn — khách THẤY, không chỉ đọc.

        * phong cách sẵn (`pc:`): đúng thẻ có 3 ảnh + video mẫu, bấm là xem to;
        * kênh / bộ vẽ: ảnh nhân vật mẫu `nv1.png` của kênh ấy;
        * AI xây từ ảnh: chính những ảnh khách đã tải lên;
        * tự động: không có gì để minh hoạ — ẩn khối.
        """
        from PyQt5.QtCore import QUrl  # noqa: PLC0415
        from PyQt5.QtGui import QDesktopServices  # noqa: PLC0415

        lay = self._khoi_minh_hoa.layout()
        while lay.count():
            muc = lay.takeAt(0)
            if muc.widget() is not None:
                muc.widget().deleteLater()

        def mo_tep(p: str) -> None:
            QDesktopServices.openUrl(QUrl.fromLocalFile(p))

        cac_the: List[_TheHinh] = []
        if ma.startswith("pc:"):
            for i, (ten_pc, kv) in enumerate(PHONG_CACH):
                slug = str(kv.get("slug", ""))
                if "pc:" + slug != ma:
                    continue
                anhs = _anh_mau_cua(i, slug)
                the_ = _TheHinh(ten_pc, str(kv.get("_mo_ta", "")),
                                anhs[0] if anhs else "", so_anh=len(anhs),
                                so_video=len(_video_mau_cua_nhieu(slug)))
                the_.bam.connect(lambda _i=i: self._xem_phong(_i))
                cac_the.append(the_)
        elif ma.startswith(("kenh:", "ve:")):
            p = self._anh_nhan_vat(ma)
            if p:
                the_ = _TheHinh(ten, mo_ta or "Nhân vật mẫu", p, so_anh=1)
                the_.bam.connect(lambda _p=p: mo_tep(_p))
                cac_the.append(the_)
        elif ma == "anh:":
            for so, p in enumerate([a for a in self._anh_mau_chon
                                    if os.path.isfile(a)][:5], start=1):
                the_ = _TheHinh("Ảnh bạn tải {0}".format(so),
                                os.path.basename(p), p, so_anh=1)
                the_.bam.connect(lambda _p=p: mo_tep(_p))
                cac_the.append(the_)
        for the_ in cac_the:
            lay.addWidget(the_)
        self._khoi_minh_hoa.setVisible(bool(cac_the))

    def _anh_nhan_vat(self, ma: str) -> str:
        """Đường `nv1.png` của kênh (`kenh:`) hay bộ vẽ (`ve:`), "" nếu không có."""
        from core.kenh import THU_MUC_NV, duong_kenh  # noqa: PLC0415
        from core.khuon import TEP_NV_MAU, duong_khuon  # noqa: PLC0415

        goc = self._app.base_dir
        loai, _, ten = ma.partition(":")
        p = (os.path.join(duong_kenh(goc, ten), THU_MUC_NV, TEP_NV_MAU)
             if loai == "kenh" else duong_khuon(goc, "ve", ten, TEP_NV_MAU))
        return p if os.path.isfile(p) else ""

    def _doi_phong_cach(self) -> None:
        """Bỏ phong cách chọn ở Bước 1 → hai tab Bước 2 mở lại để chọn tay."""
        self._chon_tu_dong()

    def _nap_style_tu_kenh(self, ma: str) -> None:
        self._nap_phong_cach()
        i = self._o_kenh.findData("kenh:" + ma)
        if i < 0:
            raise ValueError("Kênh “{0}” chưa có style.yaml có khoá hình.".format(ma))
        self._o_kenh.setCurrentIndex(i)

    def _luu_style_vao_kenh(self, ma: str) -> None:
        from core.dong_bo_kenh import (  # noqa: PLC0415
            chep_nhan_vat, chi_dan_thanh_khoa, ghi_style,
        )

        khoa = chi_dan_thanh_khoa(self._o_chi_dan.toPlainText())
        if not khoa:
            raise ValueError(
                "Ô prompt phong cách đang trống hoặc không có dòng “Image "
                "style: …” nào để ghi. Chọn một phong cách trước.")
        da = ghi_style(self._app.base_dir, ma, khoa)
        chu = "Đã ghi {0} khoá hình vào kênh {1}".format(len(da), ma)
        if self._anh_nv and os.path.isfile(self._anh_nv):
            chep_nhan_vat(self._app.base_dir, ma, self._anh_nv)
            chu += " + ảnh nhân vật nv1.png"
        self._ghi(chu + ".")
        self._nap_phong_cach()

    def kenh_da_doi(self) -> None:
        """Kênh vừa đổi ở tab khác → làm mới ô “Phong cách đã lưu”."""
        self._nap_phong_cach()

    # ── Mẫu đã lưu ───────────────────────────────────────────────────────────

    def _ap_mau(self, ten: str) -> None:
        """Điền lại phong cách + prompt theo mẫu `ten` — chọn trong ô là xong."""
        from core.mau_pv import doc_mau  # noqa: PLC0415

        mau = next((m for m in doc_mau(self._app.base_dir)
                    if str(m["ten"]) == ten), None)
        if mau is None:
            self._nap_phong_cach()
            return
        ma = str(mau.get("phong_cach") or "auto")
        thieu = ""
        if ma == "auto":
            self._chon_tu_dong()
        elif ma.startswith("pc:"):
            vi_tri = [i for i, (_t, kv) in enumerate(PHONG_CACH)
                      if "pc:" + str(kv.get("slug", "")) == ma]
            if vi_tri:
                self._chon_phong(vi_tri[0])
            else:
                thieu = ma
        elif ma in self._phong_cach:
            p = self._phong_cach[ma]
            self._dat_chon(p.ma, p.ten, p.mo_ta, p.chi_dan)
        elif ma == "anh:":
            # Phong cách AI xây từ ảnh: không có thẻ, không có kênh — nội dung
            # nằm ở `chi_dan` của mẫu (điền ngay dưới), minh hoạ bằng chính
            # những ảnh khách đã tải khi xây (lưu kèm mẫu).
            self._dat_chon("anh:", "Phong cách từ ảnh của bạn",
                           "AI đã xây từ ảnh bạn tải lên.", "",
                           anh_mau=[str(a) for a in (mau.get("anh_mau") or [])])
        else:
            # Kênh/bộ vẽ của mẫu đã bị xoá khỏi máy. Nói ra thay vì im lặng
            # dùng phong cách khác.
            thieu = ma
        # Prompt đã tinh chỉnh trong mẫu thắng prompt gốc của phong cách —
        # đó chính là phần công khách bỏ ra mà mẫu sinh ra để giữ.
        if str(mau.get("chi_dan") or "").strip():
            self._o_chi_dan.setPlainText(str(mau["chi_dan"]))
        self._o_khuon_chia.setPlainText(
            str(mau.get("khuon_chia") or "").strip() or KHUON_MAC_DINH)
        # Ô chọn đứng ở mục mẫu (không nhảy về phong cách gốc) để khách thấy
        # mình đang dùng mẫu nào, và nút "Xoá mẫu này" biết xoá cái gì.
        self._o_kenh.blockSignals(True)
        i = self._o_kenh.findData("mau:" + ten)
        if i >= 0:
            self._o_kenh.setCurrentIndex(i)
        self._o_kenh.blockSignals(False)
        self._nhan_phong_cach.setText(
            "✓ Đang dùng lại: {0}. Mọi cảnh sẽ theo đúng kiểu này — muốn "
            "xem/sửa chi tiết thì mở ⚙ Nâng cao.".format(ten))
        self._nut_xoa_mau.show()
        # Chọn mẫu rồi thì hai tab gấp lại: chỉ còn dòng trên và nút Đổi —
        # đúng ý "có chọn mẫu thì bước 2 thể hiện phong cách đó".
        self._tab.hide()
        self._nut_doi.show()
        if thieu:
            self._app.show_message(
                "Mẫu “{0}” thiếu phong cách".format(ten),
                "Phong cách của mẫu ({0}) không còn trên máy — có thể kênh hay "
                "bộ vẽ đó đã bị xoá. Prompt và các tuỳ chọn khác đã điền; hãy "
                "chọn lại phong cách rồi lưu đè mẫu.".format(thieu))
        else:
            self._ghi("Đã điền thiết lập theo mẫu “{0}”.".format(ten))

    def _luu_mau(self) -> None:
        from PyQt5.QtWidgets import QInputDialog  # noqa: PLC0415

        from core.mau_pv import luu_mau  # noqa: PLC0415

        if self._chon_ma == "auto" and not self._o_chi_dan.toPlainText().strip():
            self._app.show_message(
                "Chưa có gì để lưu",
                "Chọn một phong cách ở Bước 2 (hoặc nhờ AI xây từ ảnh) rồi "
                "mới lưu được.")
            return
        ma_hien = str(self._o_kenh.currentData() or "")
        goi_y = ma_hien[4:] if ma_hien.startswith("mau:") else ""
        ten, dong_y = QInputDialog.getText(
            self, "Lưu phong cách",
            "Đặt tên cho phong cách này (trùng tên cũ là cập nhật):",
            text=goi_y)
        if not dong_y or not str(ten).strip():
            return
        try:
            luu_mau(self._app.base_dir, str(ten), {
                "phong_cach": self._chon_ma,
                "chi_dan": self._o_chi_dan.toPlainText().strip(),
                # Ảnh khách tải khi nhờ AI xây — để lần sau chọn mẫu vẫn có
                # hình minh hoạ, không chỉ một khối chữ.
                "anh_mau": list(self._anh_mau_chon),
                # Nâng cao đi theo mẫu: prompt chia cảnh đã sửa (rỗng = mặc
                # định; mạch chia là hai con số nằm trong đó) — phần công
                # khách bỏ ra.
                "khuon_chia": self._khuon_chia_da_sua(),
            })
        except Exception as loi:  # noqa: BLE001 — lưu hỏng không được giết tab
            self._app.show_error(loi)
            return
        self._nap_phong_cach()
        self._ap_mau(str(ten).strip())
        self._ghi("Đã lưu “{0}” — lần sau chọn ở ô “Dùng lại” của Bước 2.".format(
            str(ten).strip()))

    def _xoa_mau(self) -> None:
        from core.mau_pv import xoa_mau  # noqa: PLC0415

        ma = str(self._o_kenh.currentData() or "")
        if not ma.startswith("mau:"):
            return
        xoa_mau(self._app.base_dir, ma[4:])
        self._chon_tu_dong()
        self._nap_phong_cach()
        self._ghi("Đã xoá mẫu “{0}”.".format(ma[4:]))

    def _nap_phong_cach(self) -> None:
        """Đọc lại kênh + bộ vẽ từ đĩa vào ô chọn phụ, giữ nguyên mục đang chọn.

        Gọi cả ở `showEvent`: khách vừa tạo một kênh mới bên tab Tự động rồi
        quay lại đây thì phong cách của kênh đó phải có mặt ngay, không bắt
        mở lại tool. Chỉ lấy mục `kenh:`/`ve:` — các mẫu gọn viết cứng đã có
        lưới thẻ (có ảnh mẫu) thay thế.
        """
        from core.mau_pv import doc_mau  # noqa: PLC0415

        try:
            ds = [p for p in liet_ke_phong_cach(self._app.base_dir)
                  if ":" in p.ma]
        except Exception:  # noqa: BLE001 — đọc khuôn hỏng không được chặn tab
            ds = []
        self._phong_cach = {p.ma: p for p in ds}
        mau = doc_mau(self._app.base_dir)
        dang_chon = str(self._o_kenh.currentData() or "") or self._chon_ma
        self._o_kenh.blockSignals(True)
        self._o_kenh.clear()
        self._o_kenh.addItem("(không — chọn mới ở dưới)"
                             if (ds or mau) else "(chưa có gì để dùng lại)", "")
        for p in ds:
            self._o_kenh.addItem(p.ten, p.ma)
            if p.mo_ta:
                self._o_kenh.setItemData(self._o_kenh.count() - 1, p.mo_ta,
                                         Qt.ToolTipRole)
        for m in mau:
            self._o_kenh.addItem("Mẫu: {0}".format(m["ten"]), "mau:" + str(m["ten"]))
        i = self._o_kenh.findData(dang_chon)
        self._o_kenh.setCurrentIndex(i if i >= 0 else 0)
        self._o_kenh.blockSignals(False)

    #: Ba khâu của một lượt chạy, vẽ thành MỘT dòng `● Nghe · ◉ Viết prompt ·
    #: ○ Excel` — mượn ý bảng khâu tab Tự động nhưng gọn bằng một hàng chữ.
    _TEN_KHAU = ("Nghe (trên máy, miễn phí)", "Viết prompt (AI)", "Excel")
    _KY_HIEU_KHAU = {"cho": "○", "chay": "◉", "xong": "●", "hong": "✕"}

    def _the_chay(self) -> QWidget:
        khung = the()
        v = QVBoxLayout(khung)
        v.setContentsMargins(18, 16, 18, 18)
        v.setSpacing(10)
        v.addWidget(nhan("Bước 3 — Tạo prompt", "h2"))

        # Không có ô nào để chọn: tiếng tự đoán, chất lượng mặc định, giữ nhân
        # vật xuyên suốt, Veo 3 — xem hằng ở đầu tệp. Chỉ một nút.
        self._canh_bao = self._chu_phu("")
        self._canh_bao.setStyleSheet("color:{0};".format(theme.VANG))
        self._canh_bao.hide()
        v.addWidget(self._canh_bao)

        hang = HangXuongDong()
        self._nut_chay = nut_chinh("▶ Tạo prompt", self._chay)
        self._nut_chay.setFixedWidth(200)
        self._nut_chay.setToolTip(
            "Máy tự nghe file (miễn phí, tự đoán tiếng) → AI dựng nhân vật & "
            "bối cảnh, chia cảnh và viết prompt, viết 3 prompt ảnh bìa và "
            "prompt nhạc Suno (tiêu ví) → file Excel. Cảnh cắt theo Veo 3, tối "
            "đa 8 giây.")
        hang.addWidget(self._nut_chay)
        # Nút Dừng THẾ CHỖ nút Tạo trong lúc chạy — một chỗ, một nút.
        self._nut_dung = nut_phu("■ Dừng", self._dung, rong=200)
        self._nut_dung.hide()
        hang.addWidget(self._nut_dung)
        v.addLayout(hang)
        # Ba khâu ở DÒNG RIÊNG, và chỉ hiện khi đã bấm chạy: lúc chưa chạy thì
        # Bước 3 đúng là một nút. (Nhét vào hàng xuống dòng thì nhãn bị bóp
        # còn một chữ mỗi dòng — lỗi trong ảnh chụp 24/08/2026.)
        self._nhan_khau = self._chu_phu("")
        self._nhan_khau.hide()
        v.addWidget(self._nhan_khau)
        self._khau_tt = ["cho", "cho", "cho"]
        self._ve_khau()

        self._thanh = QProgressBar()
        self._thanh.setRange(0, 100)
        self._thanh.setValue(0)
        self._thanh.setTextVisible(False)
        self._thanh.setFixedHeight(8)
        self._thanh.hide()
        v.addWidget(self._thanh)

        # Dòng trạng thái: câu mới nhất của nhật ký, cho khách biết đang làm gì
        # mà không phải mở ô nhật ký.
        self._dong_trang_thai = self._chu_phu("")
        v.addWidget(self._dong_trang_thai)

        # Thư mục lưu gấp sau một nút nhỏ: gần như không ai đổi.
        self._nut_thu_muc = nut_phu("Lưu vào ▾", lambda: self._bat_tat(
            self._thu_muc, self._nut_thu_muc, "Lưu vào"), rong=110)
        v.addWidget(self._nut_thu_muc, 0, Qt.AlignLeft)
        self._thu_muc = ChonThuMuc(self._app.default_output_dir("prompt-visuals"))
        self._thu_muc.hide()
        v.addWidget(self._thu_muc)
        return khung

    def _ve_khau(self) -> None:
        self._nhan_khau.setText("   ".join(
            "{0} {1}".format(self._KY_HIEU_KHAU[tt], ten)
            for ten, tt in zip(self._TEN_KHAU, self._khau_tt)))

    def _dat_khau(self, so: int, trang_thai: str) -> None:
        """Đổi trạng thái một khâu (`cho`/`chay`/`xong`/`hong`) rồi vẽ lại."""
        if 0 <= so < len(self._khau_tt):
            self._khau_tt[so] = trang_thai
            self._ve_khau()

    def _khau_su_kien(self, loai: str, node: str) -> None:
        """Đổi sự kiện của runner thành trạng thái ba khâu.

        `node` là id node trong tờ khai workflow (`nghe`/`prompt`) — cộng thêm
        mã riêng `excel` do `_chay_nen` tự phát khi chép file ra, và `reset`
        khi bắt đầu một file mới (chạy nhiều file thì mỗi file một vòng khâu).
        """
        if loai == "reset":
            self._khau_tt = ["cho", "cho", "cho"]
            self._ve_khau()
            return
        so = {"nghe": 0, "prompt": 1, "excel": 2}.get(node)
        if so is None:
            return
        if loai == "node_started":
            self._dat_khau(so, "chay")
        elif loai == "node_succeeded":
            self._dat_khau(so, "xong")
            if node == "prompt":
                self._dat_khau(2, "chay")
        elif loai in ("node_failed", "node_blocked", "node_cancelled"):
            self._dat_khau(so, "hong")

    def _khau_nen(self, loai: str, node: str) -> None:
        """Bản gọi được từ LUỒNG NỀN của `_khau_su_kien`."""
        self._app.goi_tren_luong_ve(lambda: self._khau_su_kien(loai, node))

    @staticmethod
    def _chu_phu(chu: str) -> QLabel:
        nh = nhan(chu, "phu")
        nh.setWordWrap(True)
        nh.setMinimumWidth(1)
        return nh

    # ── Xem trước prompt từng cảnh ─────────────────────────────────────────────
    #
    # Tab Tự động cho khách **thấy từng cảnh** (dải phim) thay vì mở thư mục rồi
    # bấm từng tệp. Prompt Visuals mượn đúng ý đó: chạy xong thì prompt hiện ra
    # ngay trong tab, **đánh số và xếp đúng thứ tự cảnh**, kèm một dòng tóm tắt
    # dàn nhân vật giữ xuyên suốt — để khách hiểu file Excel vừa tạo có gì mà
    # không cần mở VE3.

    def _the_xem(self) -> QWidget:
        khung = the()
        v = QVBoxLayout(khung)
        v.setContentsMargins(18, 16, 18, 18)
        v.setSpacing(10)
        v.addWidget(nhan("Bước 4 — Xem, sửa prompt và thử vài cảnh thật", "h2"))

        hang = HangXuongDong()
        self._o_xem = QComboBox()
        self._o_xem.setMinimumWidth(1)
        self._o_xem.setToolTip("Chọn file để xem prompt của nó.")
        self._o_xem.currentIndexChanged.connect(lambda _i: self._ve_xem())
        hang.addWidget(self._o_xem)
        v.addLayout(hang)

        self._nhan_dan = self._chu_phu(
            "Chạy xong, prompt từng cảnh sẽ hiện ra đây theo đúng thứ tự. Bấm "
            "vào ô prompt để sửa, rồi bấm “Lưu chỉnh sửa vào Excel”.")
        v.addWidget(self._nhan_dan)

        # ═══ BỐN TAB: CẢNH · ẢNH BÌA · NHẠC SUNO · NHÂN VẬT & BỐI CẢNH ═══
        #
        # Chủ dự án 24/08/2026: *"excel vẫn chưa có prompt tạo thumbnail và
        # nhạc suno"*. Excel giờ có đủ; mỗi thứ một tab, sửa tại chỗ rồi một
        # nút Lưu ghi cả ba sheet.
        self._bang_xem = self._bang_sua(
            ["#", "Lời đọc", "Lời nhắc ảnh", "Lời nhắc video"], co_dinh={0: 40},
            khong_sua=(0, 1))
        self._bang_bia = self._bang_sua(
            ["#", "Kiểu", "Lời nhắc ảnh bìa"], co_dinh={0: 40, 1: 130},
            khong_sua=(0, 1))
        self._bang_nhac = self._bang_sua(
            ["#", "Từ → đến (giây)", "Lời nhắc Suno", "Không khí"],
            co_dinh={0: 40, 1: 120}, khong_sua=(0, 1))
        self._bang_dan = self._bang_sua(
            ["Mã", "Vai / tên", "Mô tả cố định (tiếng Anh)"], co_dinh={0: 60, 1: 150},
            khong_sua=(0, 1, 2))
        # Kế hoạch đạo diễn (loại 2, 3): khách thấy phim được chia màn ra sao,
        # mỗi beat ai/ở đâu/cỡ cảnh gì — đọc để hiểu, không sửa ở đây.
        self._bang_dao_dien = self._bang_sua(
            ["Màn", "Beat", "Dòng", "Mục đích", "Ai", "Ở đâu", "Cỡ cảnh · máy",
             "Điều thay đổi"],
            co_dinh={0: 44, 1: 44, 2: 70, 4: 80, 5: 60, 6: 120},
            khong_sua=(0, 1, 2, 3, 4, 5, 6, 7))

        self._tab_xem = QTabWidget()
        self._tab_xem.setMinimumWidth(1)
        self._tab_xem.addTab(self._bang_xem, "Cảnh")
        khung_bia = QWidget()
        khung_bia.setMinimumWidth(1)
        vb = QVBoxLayout(khung_bia)
        vb.setContentsMargins(0, 6, 0, 0)
        vb.setSpacing(6)
        self._nhan_bia = self._chu_phu("")
        vb.addWidget(self._nhan_bia)
        vb.addWidget(self._bang_bia)
        self._tab_xem.addTab(khung_bia, "Ảnh bìa")
        self._tab_xem.addTab(self._bang_nhac, "Nhạc Suno")
        self._tab_xem.addTab(self._bang_dan, "Nhân vật & bối cảnh")
        khung_dd = QWidget()
        khung_dd.setMinimumWidth(1)
        vd = QVBoxLayout(khung_dd)
        vd.setContentsMargins(0, 6, 0, 0)
        vd.setSpacing(6)
        self._nhan_man = self._chu_phu("")
        vd.addWidget(self._nhan_man)
        vd.addWidget(self._bang_dao_dien)
        self._tab_xem.addTab(khung_dd, "Đạo diễn")
        v.addWidget(self._tab_xem)

        hang2 = HangXuongDong()
        self._nut_luu = nut_phu("Lưu chỉnh sửa vào Excel", self._luu_chinh_sua,
                                rong=220)
        self._nut_luu.setEnabled(False)
        hang2.addWidget(self._nut_luu)
        hang2.addWidget(nut_phu("Mở thư mục Excel",
                                lambda: mo_thu_muc(self._thu_muc_da_xuat),
                                rong=170))
        v.addLayout(hang2)

        # ═══ THỬ VÀI CẢNH THẬT — cùng thẻ, ngay dưới bảng ═══
        #
        # Chủ dự án 24/08/2026: khách cần *"có chế độ test và ảnh vài video
        # luôn"* — thấy phong cách thành ẢNH THẬT và CLIP THẬT trên một hai
        # cảnh trước khi mang cả trăm cảnh sang tab Ảnh & Video. Từng là Bước 5
        # riêng; gộp vào đây vì nó là việc làm NGAY SAU khi xem prompt, trên
        # chính bảng này, và bớt một thẻ là bớt một chỗ để lạc.
        #
        # Tiền và nhịp: mỗi cảnh thử = 1 ảnh + 1 clip, giá nói TRƯỚC nút bấm.
        # Mọi việc đi qua `app.start_batch` — cùng hàng đợi, cùng bộ giữ nhịp
        # với các tab khác. Ảnh xong thì nối sang clip bằng LINK máy chủ trả
        # sẵn khi còn dùng được (không đẩy lại ảnh lên mạng — CLAUDE.md luật 5).
        v.addWidget(nhan("Thử vài cảnh thật", "phu"))
        v.addWidget(self._chu_phu(
            "Tạo thử ảnh + clip cho một hai cảnh đầu để xem phong cách có ưng "
            "không. Chưa ưng thì đổi phong cách hoặc sửa prompt ở trên rồi thử "
            "lại. Ưng rồi thì mang file Excel sang tab Ảnh & Video → Hàng loạt "
            "để chạy hết."))
        hang3 = HangXuongDong()
        self._so_thu = QComboBox()
        for n in (1, 2, 3):
            self._so_thu.addItem("Thử {0} cảnh đầu".format(n), n)
        self._so_thu.setCurrentIndex(1)
        self._so_thu.setMinimumWidth(1)
        self._so_thu.currentIndexChanged.connect(lambda _i: self._ve_gia_thu())
        hang3.addWidget(self._so_thu)
        self._nut_thu = nut_chinh("Tạo thử ảnh + video", self._thu_phong_cach)
        self._nut_thu.setFixedWidth(220)
        self._nut_thu.setEnabled(False)
        hang3.addWidget(self._nut_thu)
        v.addLayout(hang3)
        # Giá nói trước, cập nhật theo số cảnh và engine đã chọn.
        self._nhan_gia_thu = self._chu_phu(
            "Tạo prompt xong (Bước 3) là thử được ngay tại đây.")
        v.addWidget(self._nhan_gia_thu)
        self._thu_vien = ThuVienKetQua(
            "Ảnh và clip thử sẽ hiện ở đây — mỗi cảnh một thẻ ảnh, xong ảnh "
            "thì clip của đúng cảnh đó tự chạy tiếp.")
        self._thu_vien.setMinimumWidth(1)
        self._thu_vien.setFixedHeight(300)
        v.addWidget(self._thu_vien)
        return khung

    def _cap_nhat_thu(self) -> None:
        """Bật/tắt nút thử theo việc đã có cảnh trong bảng Bước 4 hay chưa."""
        if not hasattr(self, "_nut_thu"):
            return  # nút thử dựng ở cuối thẻ Bước 4 — lượt vẽ đầu chưa có
        self._nut_thu.setEnabled(bool(self._canh_hien))
        self._ve_gia_thu()

    def _ve_gia_thu(self) -> None:
        if not hasattr(self, "_nhan_gia_thu"):
            return
        if not self._canh_hien:
            self._nhan_gia_thu.setText(
                "Tạo prompt xong (Bước 3) là thử được ngay tại đây.")
            return
        n = min(int(self._so_thu.currentData() or 1), len(self._canh_hien))
        gia = n * (hold_for_image(1, self._app.prices)
                   + hold_for_video(ENGINE_PV, self._app.prices))
        self._nhan_gia_thu.setText(
            "Thử {0} cảnh = {0} ảnh + {0} clip, tạm giữ khoảng {1}. Việc lỗi "
            "được hoàn 100%.".format(n, format_vnd(gia)))

    def _thu_muc_thu(self) -> str:
        """Thư mục riêng cho đồ thử, nằm cạnh file Excel — không lẫn với kết
        quả chạy cả loạt."""
        return os.path.join(self._thu_muc.value, "thu-phong-cach")

    def _thu_phong_cach(self) -> None:
        """Gửi việc tạo ẢNH cho N cảnh đầu; clip nối tiếp khi ảnh xong.

        Lấy prompt từ CHÍNH bảng Bước 4 (bản khách vừa sửa), không đọc lại
        Excel — khách sửa xong bấm thử là thử đúng chữ đang thấy trên màn hình.
        """
        if not self._canh_hien:
            self._app.show_message(
                "Chưa có cảnh để thử",
                "Chạy “Tạo prompt” ở Bước 3 xong là các cảnh hiện ra ở Bước 4, "
                "lúc đó mới có gì để thử.")
            return
        if self._app.client is None:
            self._app.show_message(
                "Chưa đăng nhập",
                "Tạo ảnh và clip cần ví ShopAPI. Mở tab Tài khoản & Cài đặt, "
                "gõ email và mật khẩu để đăng nhập trước.")
            return
        n = min(int(self._so_thu.currentData() or 1), len(self._canh_hien),
                self._bang_xem.rowCount())
        thu_muc = self._thu_muc_thu()
        viec: List[Tuple[int, str, List[str]]] = []   # (dòng, prompt, ảnh tham chiếu trên máy)
        for dong in range(n):
            mo_ta = self._o(self._bang_xem, dong, self._COT_ANH)
            if not mo_ta:
                continue
            van_de = check_image([mo_ta], n=1, aspect_ratio="16:9")
            if van_de:
                self._app.show_message(
                    "Cần sửa prompt cảnh {0}".format(dong + 1),
                    "\n".join("• " + v for v in van_de))
                return
            viec.append((dong, mo_ta, self._tham_chieu_cua_canh(dong)))
        if not viec:
            self._app.show_message(
                "Các cảnh đầu chưa có prompt ảnh",
                "Mở Bước 4, điền “Lời nhắc ảnh” cho các cảnh đầu rồi bấm lại.")
            return
        # Ảnh tham chiếu của các cảnh thử phải đi kèm — không thì con mèo thử ra
        # một con mèo khác và khách kết luận sai về phong cách. Tải lên ở luồng
        # nền (mỗi tệp một lần), rồi mới gửi.
        can_tai = sorted({p for _d, _m, ds in viec for p in ds})
        if can_tai and self._app.client is not None:
            from core.anh_len import tai_len  # noqa: PLC0415

            client = self._app.client

            def tai():
                kho = {}
                for p in can_tai:
                    try:
                        kho[p] = str(tai_len(client, p) or "")
                    except Exception:  # noqa: BLE001 — một ảnh hỏng không chặn cả lượt thử
                        kho[p] = ""
                return kho

            self._app.run_bg(tai, on_ok=lambda kho: self._gui_thu(viec, thu_muc, kho),
                             on_err=self._app.show_error)
            return
        self._gui_thu(viec, thu_muc, {})

    def _tham_chieu_cua_canh(self, dong: int) -> List[str]:
        """Đường dẫn ảnh tham chiếu CÓ THẬT trên máy của cảnh ở dòng `dong`."""
        if dong >= len(self._canh_hien):
            return []
        chu = str(self._canh_hien[dong].get("reference_files") or "")
        ra: List[str] = []
        try:
            ds = json.loads(chu) if chu.strip().startswith("[") else chu.split(",")
        except ValueError:
            ds = chu.split(",")
        for p in ds:
            p = str(p).strip().strip('"')
            if p and os.path.isfile(p):
                ra.append(p)
        return ra[:3]

    def _gui_thu(self, viec, thu_muc: str, kho_url: Dict[str, str]) -> None:
        specs: List[JobSpec] = []
        for dong, mo_ta, cuc_bo in viec:
            urls = [kho_url[p] for p in cuc_bo if kho_url.get(p)]
            spec = JobSpec(
                kind=KIND_IMAGE, content=mo_ta, label=mo_ta[:80],
                index=dong + 1,
                params={"n": 1, "aspect_ratio": "16:9",
                        "reference_images": urls or None,
                        "tham_chieu_cuc_bo": [p for p in cuc_bo if kho_url.get(p)] or None},
                out_dir=thu_muc,
                estimate_micro=hold_for_image(1, self._app.prices))
            self._thu_dong_anh[spec.idempotency_key] = dong
            self._thu_vien.them(spec.idempotency_key, mo_ta, False,
                                ty_le="16:9", so_canh=dong + 1)
            specs.append(spec)
        co_tc = sum(1 for _d, _m, c in viec if c)
        self._ghi("Thử phong cách: gửi {0} ảnh{1}, clip sẽ tự nối khi ảnh xong."
                  .format(len(specs), " ({0} cảnh kèm ảnh tham chiếu)".format(co_tc) if co_tc else ""))
        self._app.start_batch(specs, folder=thu_muc)

    # ── Nhận sự kiện job và nối ảnh thử → clip thử ───────────────────────────
    #
    # Cùng khuôn với tab Ảnh & Video: `nhan_su_kien` chỉ GHI SỔ, còn việc gửi
    # clip nằm ở `cuoi_nhip` — một ảnh có thể phát nhiều sự kiện liên tiếp,
    # gửi ngay lúc nhận là gửi trùng và trả tiền hai lần.

    def nhan_su_kien(self, loai: str, du_lieu) -> None:
        if loai != "job":
            return
        spec = getattr(du_lieu, "spec", None)
        khoa = str(getattr(spec, "idempotency_key", "") or "")
        if not khoa:
            return
        if khoa in self._tc_dang_cho:
            self._thu_vien.cap_nhat(du_lieu)
            trang_thai = str(getattr(du_lieu, "status", ""))
            if trang_thai == STATUS_DONE:
                self._nhan_anh_tham_chieu(khoa, du_lieu)
            elif trang_thai in (STATUS_FAILED, STATUS_CANCELLED):
                self._tham_chieu_hong(khoa, du_lieu)
            return
        if khoa in self._thu_dong_anh:
            self._thu_vien.cap_nhat(du_lieu)
            if str(getattr(du_lieu, "status", "")) != STATUS_DONE:
                return
            files = list(getattr(du_lieu, "files", ()) or ())
            if not files:
                return
            dong = self._thu_dong_anh[khoa]
            link = list(getattr(du_lieu, "urls", ()) or ())
            url = str(link[0]) if link and link_dung_lai_duoc(link[0]) else ""
            self._thu_cho_noi[dong] = (files[0], url)
        elif khoa in self._thu_dong_video:
            self._thu_vien.cap_nhat(du_lieu)

    def cuoi_nhip(self) -> None:
        """Ảnh thử nào xong thì đẩy tiếp thành clip. Gọi mỗi nhịp bơm cửa sổ."""
        if not self._thu_cho_noi:
            return
        sang = dict(self._thu_cho_noi)
        self._thu_cho_noi.clear()
        for dong, (duong_anh, url) in sang.items():
            mo_ta = self._o(self._bang_xem, dong, self._COT_VIDEO)
            if not mo_ta:
                self._ghi("Cảnh {0} không có prompt video — chỉ thử ảnh."
                          .format(dong + 1))
                continue
            if url:
                # Link máy chủ trả sẵn: gửi clip luôn, không đẩy ảnh lên lại.
                self._gui_video_thu(dong, mo_ta, url)
            elif self._app.client is not None and os.path.isfile(duong_anh):
                self._app.run_bg(
                    lambda d=duong_anh: tai_len(self._app.client, d),
                    on_ok=lambda u, dg=dong, mt=mo_ta: (
                        self._gui_video_thu(dg, mt, str(u)) if u else None),
                    on_err=lambda loi, dg=dong: self._ghi(
                        "Cảnh {0}: gửi ảnh lên để làm clip không được ({1})"
                        .format(dg + 1, str(loi)[:80])))

    def _gui_video_thu(self, dong: int, mo_ta: str, url: str) -> None:
        van_de = check_video([mo_ta], engine=ENGINE_PV, aspect_ratio="16:9",
                             image_url=url)
        if van_de:
            self._ghi("Cảnh {0}: prompt video chưa đạt — {1}".format(
                dong + 1, "; ".join(van_de)))
            return
        thu_muc = self._thu_muc_thu()
        spec = JobSpec(
            kind=KIND_VIDEO, content=mo_ta, label=mo_ta[:80], index=dong + 1,
            params={"engine": ENGINE_PV, "duration": 8,
                    "aspect_ratio": "16:9", "image_url": url},
            out_dir=thu_muc,
            estimate_micro=hold_for_video(ENGINE_PV, self._app.prices))
        self._thu_dong_video[spec.idempotency_key] = dong
        self._thu_vien.them(spec.idempotency_key, mo_ta, True,
                            ty_le="16:9", so_canh=dong + 1)
        self._app.start_batch([spec], folder=thu_muc)

    #: Chỉ số cột trong bảng Cảnh — bước Thử và nút Lưu đọc theo đây.
    _COT_ANH = 2
    _COT_VIDEO = 3

    def _bang_sua(self, cot: List[str], co_dinh: Dict[int, int],
                  khong_sua: Tuple[int, ...]) -> QTableWidget:
        """Một bảng sửa tại chỗ: cột `co_dinh` rộng cố định, còn lại co giãn.

        `khong_sua`: các cột chỉ đọc (số thứ tự, lời đọc…). Cột nào sửa được
        thì bấm đúp là gõ — khách tinh chỉnh prompt ngay trên màn hình.
        """
        bang = QTableWidget(0, len(cot))
        bang.setHorizontalHeaderLabels(cot)
        bang.verticalHeader().setVisible(False)
        bang.setEditTriggers(
            QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed
            | QAbstractItemView.AnyKeyPressed)
        bang.setWordWrap(True)
        bang.setMinimumWidth(1)
        bang.setFixedHeight(260)
        bang.setStyleSheet(
            "background:{0}; border:1px solid {1}; border-radius:8px;"
            " color:{2}; font-size:12px;".format(theme.THE_MO, theme.VIEN,
                                                 theme.CHU_MO))
        tieu = bang.horizontalHeader()
        for i in range(len(cot)):
            if i in co_dinh:
                tieu.setSectionResizeMode(i, QHeaderView.Fixed)
                bang.setColumnWidth(i, co_dinh[i])
            else:
                tieu.setSectionResizeMode(i, QHeaderView.Stretch)
        bang.khong_sua = tuple(khong_sua)  # type: ignore[attr-defined]
        return bang

    @staticmethod
    def _dien_bang(bang: QTableWidget, hang: List[List[str]]) -> None:
        """Đổ dữ liệu vào bảng; cột trong `bang.khong_sua` khoá lại."""
        khong_sua = getattr(bang, "khong_sua", ())
        bang.setRowCount(len(hang))
        for d, dong in enumerate(hang):
            for c, chu in enumerate(dong):
                o = QTableWidgetItem(str(chu))
                if c in khong_sua:
                    o.setFlags(o.flags() & ~Qt.ItemIsEditable)
                if c == 0:
                    o.setTextAlignment(Qt.AlignTop | Qt.AlignHCenter)
                bang.setItem(d, c, o)
        bang.resizeRowsToContents()

    @staticmethod
    def _o(bang: QTableWidget, dong: int, cot: int) -> str:
        o = bang.item(dong, cot)
        return (o.text() if o else "").strip()

    def _nap_xem(self, ds: List[str]) -> None:
        """Sau khi chạy xong: nạp danh sách file vào ô chọn rồi hiện file cuối."""
        self._ket_qua_xem = {}
        for d in ds:
            self._ket_qua_xem[os.path.basename(d)] = d
        self._o_xem.blockSignals(True)
        self._o_xem.clear()
        for ten in self._ket_qua_xem:
            self._o_xem.addItem(ten)
        self._o_xem.blockSignals(False)
        self._o_xem.setVisible(len(self._ket_qua_xem) > 1)
        # Có kết quả thì thẻ Bước 4 mới hiện ra — trước đó nó không có gì để
        # nói với khách.
        co = bool(self._ket_qua_xem)
        self._the_xem_w.setVisible(co)
        if co:
            self._o_xem.setCurrentIndex(self._o_xem.count() - 1)
            self._ve_xem()

    def _ve_xem(self) -> None:
        """Đọc file Excel đang chọn rồi vẽ prompt từng cảnh, xếp theo số cảnh."""
        ten = self._o_xem.currentText()
        duong = self._ket_qua_xem.get(ten, "")
        if not duong or not os.path.isfile(duong):
            self._canh_hien = []
            self._bang_xem.setRowCount(0)
            self._nut_luu.setEnabled(False)
            self._nhan_dan.setText(
                "Chạy xong, prompt từng cảnh sẽ hiện ra đây theo đúng thứ tự.")
            self._cap_nhat_thu()
            return
        try:
            so = self._doc_workbook(duong)
        except Exception as loi:  # noqa: BLE001 — đọc lỗi không được chặn tab
            self._canh_hien = []
            self._bang_xem.setRowCount(0)
            self._nut_luu.setEnabled(False)
            self._nhan_dan.setText("Chưa đọc được file để xem: {0}".format(
                str(loi)[:120]))
            self._cap_nhat_thu()
            return
        canh, dan = so["canh"], so["dan"]
        self._canh_hien = list(canh)
        self._bia_hien = list(so["bia"])
        self._nhac_hien = list(so["nhac"])
        self._nhan_dan.setText("{0}  ·  {1} cảnh  ·  {2} ảnh bìa  ·  {3} track nhạc".format(
            tom_tat_dan(dan), len(canh), len(so["bia"]), len(so["nhac"])))
        self._dien_bang(self._bang_xem, [
            ["#{0}".format(c["scene_id"]),
             c.get("srt_text_vi") or c.get("srt_text") or "",
             c.get("img_prompt", ""), c.get("video_prompt", "")]
            for c in canh])
        self._dien_bang(self._bang_bia, [
            [b["thumb_id"], b["version_desc"], b["img_prompt"]] for b in so["bia"]])
        if so["bia"]:
            self._nhan_bia.setText(
                "Tiêu đề đề xuất: {0}\nChữ trên ảnh bìa: {1}  (ảnh KHÔNG có chữ — "
                "chữ chèn sau bằng font thật)".format(
                    so["bia"][0].get("title") or "—",
                    so["bia"][0].get("thumb_text") or "—"))
        else:
            self._nhan_bia.setText("Chưa có prompt ảnh bìa trong file này.")
        self._dien_bang(self._bang_nhac, [
            [m["music_id"], "{0} → {1}".format(m["start_time"], m["end_time"]),
             m["suno_prompt"], m["mood"]] for m in so["nhac"]])
        self._dien_bang(self._bang_dan, [
            [d["id"], " – ".join(x for x in (d.get("role"), d.get("name")) if x),
             d.get("english_prompt", "")] for d in dan] + [
            [l["id"], l.get("name", ""), l.get("english_prompt", "")]
            for l in so["boi_canh"]])
        man = so["man"]
        if man:
            self._nhan_man.setText("Cung truyện: {0}  ·  {1} màn: {2}".format(
                man[0].get("arc") or "—", len(man),
                "  →  ".join("{0}. {1} ({2})".format(
                    m["segment_id"], m["name"], m["emotion"]) for m in man)))
        else:
            self._nhan_man.setText(
                "Không có kế hoạch đạo diễn: cách kể “Một nhân vật cố định” chạy "
                "như tab Tự động, chỉ loại 2 và 3 mới lên kế hoạch màn.")
        self._dien_bang(self._bang_dao_dien, [
            [b["segment_id"], b["beat"], "{0}–{1}".format(b["srt_from"], b["srt_to"]),
             b["purpose"], b["characters"] or "—", b["location"] or "—",
             "{0} · {1}".format(b["shot_size"], b["camera"]), b["element_motion"]]
            for b in so["ke_hoach"]])
        self._nut_luu.setEnabled(bool(canh))
        self._cap_nhat_thu()

    @staticmethod
    def _doc_workbook(duong: str) -> Dict[str, list]:
        """Đọc các sheet của file Excel → cảnh, dàn, ảnh bìa, nhạc, bối cảnh.

        Đọc chế độ chỉ-đọc: file có thể tới cả trăm cảnh, và ta chỉ cần các hàng
        thô để đưa cho các hàm `*_de_xem` lo phần ghép và sắp.
        """
        from openpyxl import load_workbook  # noqa: PLC0415

        wb = load_workbook(duong, read_only=True, data_only=True)
        try:
            def hang_cua(ten):
                if ten not in wb.sheetnames:
                    return []
                return [list(r) for r in wb[ten].iter_rows(values_only=True)]

            return {"canh": canh_de_xem(hang_cua("scenes")),
                    "dan": dan_de_xem(hang_cua("characters")),
                    "bia": bia_de_xem(hang_cua("thumbnail")),
                    "nhac": nhac_de_xem(hang_cua("music")),
                    "boi_canh": boi_canh_de_xem(hang_cua("locations")),
                    "man": man_de_xem(hang_cua("story")),
                    "ke_hoach": ke_hoach_de_xem(hang_cua("director_plan"))}
        finally:
            wb.close()

    def _luu_chinh_sua(self) -> None:
        """Ghi prompt đã sửa trong bảng ngược lại file Excel đang chọn.

        Chỉ đụng đúng file kết quả khách đang mở (bản đã chép ra thư mục của
        khách), chỉ hai cột `img_prompt`/`video_prompt`, theo số cảnh — không
        đổi timing, không xoá cột nào khác.
        """
        ten = self._o_xem.currentText()
        duong = self._ket_qua_xem.get(ten, "")
        if not duong or not os.path.isfile(duong) or not self._canh_hien:
            self._app.show_message("Chưa có gì để lưu",
                                   "Hãy chạy tạo prompt rồi mới chỉnh và lưu.")
            return
        sua: "dict[int, tuple]" = {}
        for dong, c in enumerate(self._canh_hien):
            sua[int(c["scene_id"])] = (self._o(self._bang_xem, dong, self._COT_ANH),
                                       self._o(self._bang_xem, dong, self._COT_VIDEO))
        sua_bia = {str(b["thumb_id"]): {"img_prompt": self._o(self._bang_bia, d, 2)}
                   for d, b in enumerate(getattr(self, "_bia_hien", []))}
        sua_nhac = {str(m["music_id"]): {"suno_prompt": self._o(self._bang_nhac, d, 2),
                                         "mood": self._o(self._bang_nhac, d, 3)}
                    for d, m in enumerate(getattr(self, "_nhac_hien", []))}
        try:
            so = self._ghi_prompt_vao_xlsx(duong, sua)
            if sua_bia:
                self._ghi_cot_vao_xlsx(duong, "thumbnail", "thumb_id", sua_bia)
            if sua_nhac:
                self._ghi_cot_vao_xlsx(duong, "music", "music_id", sua_nhac)
        except Exception as loi:  # noqa: BLE001 — lưu hỏng không được giết tab
            self._ghi("Lưu chỉnh sửa hỏng: {0}".format(str(loi)[:200]))
            self._app.show_error(loi)
            return
        self._ghi("Đã lưu chỉnh sửa {0} cảnh (+ ảnh bìa, nhạc) vào {1}.".format(
            so, os.path.basename(duong)))
        self._app.show_message(
            "Đã lưu",
            "Đã ghi prompt đã sửa vào:\n{0}\n\nMở lại bằng VE3 hoặc Excel là "
            "thấy bản mới.".format(duong))

    @staticmethod
    def _ghi_prompt_vao_xlsx(duong: str, sua) -> int:
        """Cập nhật `img_prompt`/`video_prompt` sheet `scenes` theo số cảnh.

        Trả về số cảnh đã ghi. Giữ dạng `{số: (ảnh, video)}` vì bài kiểm và nút
        Lưu gọi theo dạng ấy; bên dưới đi qua `_ghi_cot_vao_xlsx`.
        """
        return TrangPromptVisuals._ghi_cot_vao_xlsx(
            duong, "scenes", "scene_id",
            {str(int(so)): {"img_prompt": img, "video_prompt": video}
             for so, (img, video) in sua.items()})

    @staticmethod
    def _ghi_cot_vao_xlsx(duong: str, sheet: str, cot_khoa: str,
                          sua: Dict[str, Dict[str, str]]) -> int:
        """Ghi vài ô của một sheet theo khoá dòng. Trả về số dòng đã ghi.

        Đọc tiêu đề để tìm đúng cột — không đoán vị trí, vì lệch một cột là ghi
        đè nhầm dữ liệu khác. Khoá số (`1`, `1.0`) và chữ (`"1"`) đều khớp.
        """
        from openpyxl import load_workbook  # noqa: PLC0415

        def chuan(gia) -> str:
            chu = str(gia if gia is not None else "").strip()
            try:
                return str(int(float(chu)))
            except ValueError:
                return chu

        wb = load_workbook(duong)
        try:
            if sheet not in wb.sheetnames:
                raise ValueError("File không có sheet {0}".format(sheet))
            ws = wb[sheet]
            tieu = [str(c.value or "").strip() for c in ws[1]]
            vi_tri = {ten: i for i, ten in enumerate(tieu)}
            can = {cot_khoa} | {c for gia in sua.values() for c in gia}
            thieu = [c for c in can if c not in vi_tri]
            if thieu:
                raise ValueError("Sheet {0} thiếu cột {1}".format(
                    sheet, ", ".join(sorted(thieu))))
            sua_chuan = {chuan(k): v for k, v in sua.items()}
            dem = 0
            for hang in ws.iter_rows(min_row=2):
                khoa = chuan(hang[vi_tri[cot_khoa]].value)
                if khoa not in sua_chuan:
                    continue
                for cot, gia in sua_chuan[khoa].items():
                    hang[vi_tri[cot]].value = gia
                dem += 1
            tmp = duong + ".tmp"
            wb.save(tmp)
        finally:
            wb.close()
        os.replace(tmp, duong)
        return dem

    # ── Trạng thái máy ───────────────────────────────────────────────────────

    def _dich_vu(self):
        """`BuilderService` gắn với ví của người dùng.

        Dựng mới mỗi lần hỏi chứ không giữ sẵn: người dùng có thể vừa đăng nhập
        ở tab Tài khoản xong, mà bản giữ sẵn thì vẫn ôm cái khoá rỗng cũ.
        """
        from core.builder_service import BuilderService  # noqa: PLC0415

        def bi_mat():
            return {"SHOPAPI_API_KEY": (self._app.config.api_key or "").strip(),
                    "SHOPAPI_BASE_URL": self._app.config.base_url or ""}

        return BuilderService(self._app.base_dir, shopapi_secret=bi_mat)

    def _ve_trang_thai(self) -> None:
        """Hỏi máy còn thiếu gì, rồi nói ra bằng tiếng người."""
        try:
            dv = self._dich_vu()
            wf = dung_workflow("kiem-tra", ma_chay="prompt-visuals-kiemtra")
            thieu = cau_thieu_gi(dv.readiness(wf).issues)
        except Exception as loi:  # noqa: BLE001 — dò hỏng không được chặn tab
            thieu = ["Chưa dò được máy: {0}".format(loi)]
        if thieu:
            self._canh_bao.setText("Chưa chạy được:\n• " + "\n• ".join(thieu))
            self._canh_bao.show()
        else:
            self._canh_bao.hide()
        self._nut_chay.setEnabled(not thieu and bool(self._files))

    def showEvent(self, su_kien):  # noqa: N802 — tên do Qt quy định
        """Dò lại mỗi lần mở tab: người dùng có thể vừa cài đặt xong, vừa đăng
        nhập, hoặc vừa tạo một kênh mới (phong cách của nó phải hiện ra đây)."""
        super().showEvent(su_kien)
        self._nap_phong_cach()
        self._ve_trang_thai()

    # ── Chọn file ────────────────────────────────────────────────────────────

    def _chon_file(self) -> None:
        duong, _ = QFileDialog.getOpenFileNames(
            self, "Chọn file giọng đọc", "",
            "Âm thanh ({0});;Tất cả (*.*)".format(" ".join(DUOI_TIENG)))
        if not duong:
            return
        for d in duong:
            if d not in self._files:
                self._files.append(d)
        self._ve_danh_sach()

    def _bo_file(self) -> None:
        self._files = []
        self._ve_danh_sach()

    def _ve_danh_sach(self) -> None:
        if not self._files:
            self._nhan_file.setText("Chưa chọn file nào.")
        elif len(self._files) == 1:
            self._nhan_file.setText(os.path.basename(self._files[0]))
        else:
            ten = [os.path.basename(d) for d in self._files]
            self._nhan_file.setText("{0} file: {1}".format(
                len(ten), ", ".join(ten[:3]) + ("…" if len(ten) > 3 else "")))
        self._nut_bo_file.setVisible(bool(self._files))
        self._ve_trang_thai()

    # ── Chạy ─────────────────────────────────────────────────────────────────

    def _chay(self) -> None:
        if not self._files:
            self._app.show_message("Chưa có file",
                                   "Bấm “Chọn file…” để chọn file giọng đọc.")
            return
        engine = ENGINE_PV
        mo_hinh = MO_HINH_PV
        ngon_ngu = NGON_NGU_PV
        # Ô prompt phong cách là nguồn sự thật — khách sửa gì trong đó là gửi
        # đúng cái đó, không gửi lại prompt gốc của thẻ.
        chi_dan = self._o_chi_dan.toPlainText().strip()
        kich_ban = self._o_kich_ban.toPlainText()
        nhat_quan = NHAT_QUAN_PV
        che_do = self._che_do_ke()
        anh_nv = self._anh_nv if che_do in CHE_DO_CAN_ANH_NV else ""
        if che_do in CHE_DO_CAN_ANH_NV and not (anh_nv and os.path.isfile(anh_nv)):
            self._app.show_message(
                "Thiếu ảnh nhân vật",
                "Cách kể bạn chọn cần MỘT ảnh nhân vật chính. Bấm “Tải ảnh nhân "
                "vật…” ở Bước 2, hoặc đổi sang “AI tự xây nhân vật & bối cảnh”.")
            return
        # Nâng cao: khuôn chia cảnh khách sửa phải đủ chỗ trống — chặn TRƯỚC
        # khi tốn tiền nghe + gọi AI, và nói rõ thiếu gì thay vì lặng lẽ dùng
        # mặc định (khách tưởng bản sửa của mình đã chạy).
        khuon_chia = self._khuon_chia_da_sua()
        if khuon_chia and not khuon_chia_dung_duoc(khuon_chia):
            thieu = [ct for ct in CHO_TRONG_KHUON_CHIA if ct not in khuon_chia]
            self._app.show_message(
                "Prompt chia cảnh thiếu chỗ trống",
                "Bản bạn sửa thiếu {0}. Tool điền phụ đề, nhân vật, kế hoạch vào "
                "các chỗ ấy — thiếu là AI không có gì để chia. Mở ⚙ Nâng cao ở "
                "Bước 2, thêm lại hoặc bấm “Khôi phục mặc định”.".format(
                    ", ".join(thieu)))
            return
        files = list(self._files)

        self._huy = threading.Event()
        huy = self._huy
        self._nut_chay.hide()
        self._nut_dung.show()
        self._da_xong = []
        # Lượt mới thì bỏ các ảnh thử đang chờ nối clip: bảng cảnh sắp đổi,
        # nối theo số dòng cũ là ghép clip vào prompt của cảnh khác.
        self._thu_cho_noi.clear()
        self._thanh.setValue(0)
        self._thanh.show()
        self._nhan_khau.show()
        self._ghi("Bắt đầu — {0} file.".format(len(files)))

        def viec() -> List[str]:
            return self._chay_nen(files, engine, mo_hinh, ngon_ngu, chi_dan,
                                  kich_ban, nhat_quan, huy, che_do, anh_nv,
                                  khuon_chia=khuon_chia)

        self._app.run_bg(viec, on_ok=self._xong, on_err=self._hong)

    def _chay_nen(self, files, engine, mo_hinh, ngon_ngu, chi_dan, kich_ban,
                  nhat_quan, huy, che_do="tu_xay", anh_nv="",
                  khuon_chia="") -> List[str]:
        """LUỒNG NỀN. Không chạm widget — mọi câu chữ đi qua `_ghi_nen`."""
        import json  # noqa: PLC0415

        from core.artifacts import LocalArtifactStore  # noqa: PLC0415
        from core.workflow_runner import CancellationToken  # noqa: PLC0415

        dv = self._dich_vu()
        kho: LocalArtifactStore = dv.artifacts

        # Kịch bản + phong cách hình ảnh gói chung một `context`, dùng cho cả
        # loạt file (chúng cùng một video). Rỗng thì không tạo artifact, workflow
        # về đúng dạng cũ (chỉ có phụ đề).
        boi_canh = dung_boi_canh(kich_ban, chi_dan=chi_dan, che_do_ke=che_do,
                                 nhan_vat_co_dinh={"image_file": "nv1.png"},
                                 khuon_chia=khuon_chia)
        ma_context = ""
        if boi_canh:
            try:
                ct = kho.put_text(json.dumps(boi_canh, ensure_ascii=False),
                                  filename="kich-ban.json", kind="json",
                                  schema="content-package.v1")
                ma_context = _ma_artifact(ct)
            except Exception as loi:  # noqa: BLE001 — thiếu context không giết lượt
                self._ghi_nen("  (bỏ qua kịch bản/phong cách: {0})".format(
                    str(loi)[:120]))

        ra: List[str] = []
        for thu_tu, duong in enumerate(files, start=1):
            if huy.is_set():
                self._ghi_nen("Đã dừng — xong {0}/{1} file.".format(
                    len(ra), len(files)))
                break
            ten = os.path.basename(duong)
            self._ghi_nen("[{0}/{1}] {2}".format(thu_tu, len(files), ten))
            self._khau_nen("reset", "")
            try:
                # Nạp file vào kho artifact của Studio: runner chỉ làm việc với
                # mã artifact, không nhận đường dẫn trần.
                ma = kho.put_file(duong, kind="audio",
                                  schema="narration-audio.v1")
                ma_chay = "pv-" + _ma_an_toan(ten)
                wf = dung_workflow(_ma_artifact(ma), engine=engine,
                                   mo_hinh=mo_hinh, ngon_ngu=ngon_ngu,
                                   nhat_quan=nhat_quan,
                                   ma_artifact_context=ma_context,
                                   ma_chay=ma_chay, che_do_ke=che_do,
                                   kich_ban=kich_ban)
                huy_token = CancellationToken()
                if huy.is_set():
                    break

                def bao(su_kien, _ten=ten):
                    loai = getattr(su_kien, "event", "") or ""
                    node = getattr(su_kien, "node_id", "") or ""
                    if loai and node:
                        self._khau_nen(str(loai), str(node))
                    loi_nhan = getattr(su_kien, "message", "") or ""
                    tien = getattr(su_kien, "progress", None)
                    if loi_nhan:
                        self._ghi_nen("    " + str(loi_nhan))
                    if isinstance(tien, (int, float)):
                        self._tien_nen(float(tien))

                trang_thai = dv.run(
                    wf,
                    approved_permissions=("workspace.read", "workspace.write",
                                          "compute.local", "network.shopapi",
                                          "secret.shopapi"),
                    cancellation=huy_token, on_event=bao)
                duong_xlsx = self._tim_workbook(dv, trang_thai)
                if duong_xlsx:
                    dich = self._chep_ra(duong_xlsx, ten)
                    if anh_nv:
                        self._dat_anh_nv_canh_excel(dich, anh_nv)
                    # Loại 2, 3: ảnh tham chiếu (chân dung nv2…, bối cảnh loc…)
                    # tự tạo ngay — chủ dự án 24/08/2026: *"tự tạo luôn khi bấm
                    # Tạo prompt"*. Gửi từ luồng giao diện qua hàng đợi chung.
                    can_tao = self._nhan_vat_can_anh(dv, trang_thai)
                    if can_tao:
                        self._app.goi_tren_luong_ve(
                            lambda d=dich, ds=can_tao: self._tao_anh_tham_chieu(d, ds))
                    ra.append(dich)
                    self._khau_nen("node_succeeded", "excel")
                    self._ghi_nen("  xong: đã tạo {0}".format(os.path.basename(dich)))
                else:
                    self._khau_nen("node_failed", "excel")
                    self._ghi_nen("  không được: chạy xong nhưng không thấy file Excel đâu")
            except Exception as loi:  # noqa: BLE001 — một file hỏng không giết lượt
                self._ghi_nen("  không được: {0}".format(str(loi)[:200]))
        return ra

    @staticmethod
    def _tim_workbook(dv, trang_thai) -> str:
        """Lấy đường dẫn file Excel từ artifact mà bước prompt trả về."""
        nut = getattr(trang_thai, "nodes", {}) or {}
        buoc = nut.get("prompt")
        if buoc is None or getattr(buoc, "status", "") != "succeeded":
            return ""
        ma = (getattr(buoc, "outputs", {}) or {}).get("workbook")
        if not ma:
            return ""
        try:
            return str(dv.artifacts.path(ma if isinstance(ma, str) else ma[0]))
        except Exception:  # noqa: BLE001
            return ""

    def _chep_ra(self, nguon: str, ten_tieng: str) -> str:
        """Chép file Excel ra thư mục người dùng chọn, tên theo file giọng đọc.

        Để nguyên trong kho artifact thì tên là một chuỗi băm — mở ra không biết
        của file nào.
        """
        import shutil  # noqa: PLC0415

        thu_muc = self._thu_muc.value
        os.makedirs(thu_muc, exist_ok=True)
        goc = os.path.splitext(os.path.basename(ten_tieng))[0]
        dich = os.path.join(thu_muc, "{0} - prompts.xlsx".format(goc))
        shutil.copyfile(nguon, dich)
        self._thu_muc_da_xuat = thu_muc
        return dich

    def _dat_anh_nv_canh_excel(self, duong_xlsx: str, anh_nv: str) -> None:
        """Chép ảnh nhân vật thành `nv1.png` cạnh Excel, rồi ghi ĐƯỜNG DẪN THẬT
        vào `reference_files`.

        Tab Ảnh & Video → Hàng loạt chỉ nhận đường dẫn có thật trên đĩa (tên
        trơ `nv1.png` bị bỏ qua — xem `_tach_duong_tham_chieu`), còn VE3 muốn
        tệp `nv1.png` nằm cạnh Excel. Làm cả hai: tệp nằm cạnh, ô ghi đường
        dẫn đầy đủ tới đúng tệp đó.
        """
        import shutil  # noqa: PLC0415

        dich = os.path.join(os.path.dirname(duong_xlsx), "nv1.png")
        try:
            if os.path.abspath(anh_nv) != os.path.abspath(dich):
                shutil.copyfile(anh_nv, dich)
        except OSError as loi:
            self._ghi_nen("  (không chép được ảnh nhân vật: {0})".format(str(loi)[:80]))
            return
        loi = self._ghi_duong_tham_chieu(duong_xlsx, {"nv1.png": dich})
        if loi:
            self._ghi_nen("  (không ghi được đường ảnh nhân vật: {0})".format(loi))

    @staticmethod
    def _ghi_duong_tham_chieu(duong_xlsx: str, theo_ten: Dict[str, str]) -> str:
        """Đổi tên tệp (`nv2.png`, `loc1.png`) trong `reference_files` thành
        đường dẫn thật, ở cả sheet `scenes` lẫn `thumbnail`. Trả về câu lỗi
        hoặc "" khi xong.

        Ô có thể là danh sách JSON (`["nv1.png","loc1.png"]`) hay chuỗi cách
        nhau dấu phẩy (đã ghi đường dẫn từ lần trước). Ghi ra dạng dấu phẩy —
        đúng dạng tab Hàng loạt đọc; tên chưa có ảnh giữ nguyên để lần sau
        ảnh về còn thay tiếp.
        """
        import json  # noqa: PLC0415

        from openpyxl import load_workbook  # noqa: PLC0415

        def doi(chu: str) -> str:
            chu = str(chu or "").strip()
            if not chu:
                return chu
            try:
                muc = json.loads(chu) if chu.startswith("[") else None
            except ValueError:
                muc = None
            if not isinstance(muc, list):
                muc = [m.strip() for m in chu.split(",") if m.strip()]
            return ", ".join(theo_ten.get(os.path.basename(str(m)), str(m))
                             for m in muc)

        try:
            wb = load_workbook(duong_xlsx)
            try:
                for ten in ("scenes", "thumbnail"):
                    if ten not in wb.sheetnames:
                        continue
                    ws = wb[ten]
                    tieu = [str(c.value or "").strip() for c in ws[1]]
                    if "reference_files" not in tieu:
                        continue
                    cot = tieu.index("reference_files")
                    for hang in ws.iter_rows(min_row=2):
                        o = hang[cot]
                        moi = doi(o.value)
                        if moi != str(o.value or ""):
                            o.value = moi
                tmp = duong_xlsx + ".tmp"
                wb.save(tmp)
            finally:
                wb.close()
            os.replace(tmp, duong_xlsx)
        except Exception as loi:  # noqa: BLE001 — ghi hỏng thì Excel vẫn dùng được
            return str(loi)[:120]
        return ""

    # ── Ảnh tham chiếu tự tạo (loại 2, 3) ────────────────────────────────────
    #
    # Chân dung nv2… và ảnh bối cảnh loc… — mỗi cái một việc ảnh qua hàng đợi
    # chung; xong tấm nào thì chép thành `<id>.png` cạnh Excel và ghi đường
    # dẫn thật vào `reference_files` của mọi cảnh dùng nó. Tab Hàng loạt gắn
    # chúng làm ảnh tham chiếu → nhân vật và nơi chốn giữ nguyên qua cả phim.

    @staticmethod
    def _nhan_vat_can_anh(dv, trang_thai) -> List[Tuple[str, str]]:
        """`[(id, sheet_prompt)]` của nhân vật/bối cảnh chưa có ảnh trong manifest."""
        import json  # noqa: PLC0415

        nut = getattr(trang_thai, "nodes", {}) or {}
        buoc = nut.get("prompt")
        if buoc is None or getattr(buoc, "status", "") != "succeeded":
            return []
        ma = (getattr(buoc, "outputs", {}) or {}).get("scenes")
        if not ma:
            return []
        try:
            duong = str(dv.artifacts.path(ma if isinstance(ma, str) else ma[0]))
            with open(duong, "r", encoding="utf-8") as f:
                m = json.load(f)
        except Exception:  # noqa: BLE001 — không đọc được manifest thì thôi
            return []
        ra: List[Tuple[str, str]] = []
        for c in (m.get("characters") or []) + (m.get("locations") or []):
            if c.get("co_dinh") or not str(c.get("sheet_prompt") or "").strip():
                continue
            ra.append((str(c["id"]), str(c["sheet_prompt"])))
        return ra

    def _tao_anh_tham_chieu(self, duong_xlsx: str, ds: List[Tuple[str, str]],
                            thu_lai: bool = False) -> None:
        """LUỒNG GIAO DIỆN: gửi việc ảnh cho từng id, ghi sổ để nhận về."""
        if self._app.client is None:
            self._ghi("Không tạo được ảnh tham chiếu: chưa đăng nhập ShopAPI.")
            return
        thu_muc = os.path.join(os.path.dirname(duong_xlsx), "tham-chieu")
        specs: List[JobSpec] = []
        for so, (ma_id, prompt) in enumerate(ds, start=1):
            van_de = check_image([prompt], n=1, aspect_ratio="16:9")
            if van_de:
                self._ghi("Ảnh tham chiếu {0}: prompt chưa đạt — {1}".format(
                    ma_id, "; ".join(van_de)))
                continue
            spec = JobSpec(
                kind=KIND_IMAGE, content=prompt, label="Tham chiếu {0}".format(ma_id),
                index=so, params={"n": 1, "aspect_ratio": "16:9"},
                out_dir=thu_muc,
                estimate_micro=hold_for_image(1, self._app.prices))
            self._tc_dang_cho[spec.idempotency_key] = (ma_id, duong_xlsx, prompt)
            self._thu_vien.them(spec.idempotency_key, "Tham chiếu {0}: {1}".format(
                ma_id, prompt), False, ty_le="16:9")
            specs.append(spec)
        if not specs:
            return
        if not thu_lai:
            self._ghi("Đang tạo {0} ảnh tham chiếu ({1}) — xong tấm nào ghi vào "
                      "Excel tấm đó.".format(len(specs), ", ".join(i for i, _p in ds)))
        self._app.start_batch(specs, folder=thu_muc)

    def _tham_chieu_hong(self, khoa: str, du_lieu) -> None:
        """Ảnh tham chiếu bị từ chối/hỏng: thử lại MỘT lần, rồi nói thật.

        Bộ lọc của nhà cung cấp ảnh từ chối ngẫu nhiên (đo 25/08/2026: cùng
        một prompt, lần qua lần không); thử lại một lần cứu được ~1/3. Hỏng
        lần hai thì ghi rõ id nào thiếu và khách phải làm gì — không được để
        khách chạy cả phim mà không có nhân vật ấy.
        """
        ma_id, duong_xlsx, prompt = self._tc_dang_cho.pop(khoa)
        ly_do = str(getattr(du_lieu, "message", "") or "").strip()
        if ma_id not in self._tc_da_thu_lai:
            self._tc_da_thu_lai.add(ma_id)
            self._ghi("Ảnh tham chiếu {0} bị từ chối ({1}) — thử lại một lần…".format(
                ma_id, ly_do[:90] or "máy chủ không nói lý do"))
            self._tao_anh_tham_chieu(duong_xlsx, [(ma_id, prompt)], thu_lai=True)
            return
        # Lần hai vẫn hỏng → lỗi ở THIẾT KẾ nhân vật (dáng giống nhân vật có bản
        # quyền, chi tiết nhạy cảm), không phải ở một câu. Nhờ AI thiết kế lại
        # MỘT lần: sửa ở dàn, mã chèn lại khối khoá của mọi cảnh, tạo lại tham
        # chiếu. Chủ dự án 25/08/2026: *"chỗ api để xây prompt không nên cứng
        # mà phải chỉnh được nguyên lý… sẽ có nhiều câu chuyện cổ tích"*.
        goc = goc_cua_id(ma_id)
        if goc not in self._tc_da_thiet_ke_lai and self._app.client is not None:
            self._tc_da_thiet_ke_lai.add(goc)
            self._thiet_ke_lai_tham_chieu(duong_xlsx, ma_id, ly_do)
            return
        self._tc_thieu.append(ma_id)
        self._ghi("Ảnh tham chiếu {0} KHÔNG tạo được sau hai lần: {1}. Bạn mở Bước 4 → "
                  "“Nhân vật & bối cảnh”, sửa mô tả của {0} (bớt chi tiết nhạy cảm, tránh "
                  "giống nhân vật có bản quyền) rồi bấm “Tạo prompt” lại; hoặc tự chọn một "
                  "ảnh làm tham chiếu ở tab Ảnh & Video → Hàng loạt.".format(
                      ma_id, ly_do[:160] or "máy chủ không nói lý do"))
        self._bao_du_tham_chieu()

    def _thiet_ke_lai_tham_chieu(self, duong_xlsx: str, ma_id: str, ly_do: str) -> None:
        """LUỒNG GIAO DIỆN: nhờ AI thiết kế lại nhân vật (luồng nền) rồi tạo lại tham chiếu."""
        from core.goi_van_ban import goi_van_ban, loc_json  # noqa: PLC0415

        client = self._app.client
        self._ghi("Ảnh tham chiếu {0} bị từ chối hai lần — nhờ AI thiết kế lại nhân vật "
                  "(cùng vai, đổi món đồ đặc trưng) rồi cập nhật mọi cảnh…".format(ma_id))

        def lam():
            dan = self._doc_sheet(duong_xlsx, "characters")
            nv = next((c for c in dan if str(c.get("id")) == ma_id), None)
            if nv is None:
                raise ValueError("không thấy {0} trong sheet characters".format(ma_id))
            tra = goi_van_ban(client, [{"role": "user", "content": loi_nhac_thiet_ke_lai(nv, ly_do)}],
                              mo_hinh=MO_HINH_PV, toi_da_token=1024)
            moi = str(loc_json(tra).get("english_prompt") or "").strip()
            if len(moi) < 20:
                raise ValueError("AI không trả mô tả mới")
            return self._ap_thiet_ke_lai_xlsx(duong_xlsx, ma_id, moi)

        def xong(ket):
            ds, so_canh = ket
            self._ghi("Đã thiết kế lại {0} ({1} ảnh tham chiếu, {2} cảnh cập nhật): {3}…".format(
                ma_id, len(ds), so_canh, (ds[0][1] if ds else "")[:90]))
            self._nap_xem(self._da_xong)
            self._tao_anh_tham_chieu(duong_xlsx, ds, thu_lai=True)

        def hong(loi):
            self._tc_thieu.append(ma_id)
            self._ghi("Không thiết kế lại được {0} ({1}). Bạn mở Bước 4 → “Nhân vật & bối "
                      "cảnh”, sửa mô tả rồi bấm “Tạo prompt” lại.".format(ma_id, str(loi)[:120]))
            self._bao_du_tham_chieu()

        self._app.run_bg(lam, on_ok=xong, on_err=hong)

    @staticmethod
    def _doc_sheet(duong: str, ten: str) -> List[Dict[str, str]]:
        """Một sheet → danh sách dict theo tiêu đề cột (chỉ đọc)."""
        from openpyxl import load_workbook  # noqa: PLC0415

        wb = load_workbook(duong, read_only=True, data_only=True)
        try:
            if ten not in wb.sheetnames:
                return []
            hang = [list(r) for r in wb[ten].iter_rows(values_only=True)]
        finally:
            wb.close()
        if not hang:
            return []
        tieu = [str(c or "").strip() for c in hang[0]]
        return [{tieu[i]: ("" if v is None else str(v)) for i, v in enumerate(d) if i < len(tieu)}
                for d in hang[1:]]

    @staticmethod
    def _ap_thiet_ke_lai_xlsx(duong: str, ma_id: str, english_prompt_moi: str):
        """Áp thiết kế mới vào Excel: dàn (mọi giai đoạn của nhân vật) + khối khoá mọi cảnh.

        Trả `([(id, sheet_prompt_mới)…], số_cảnh_cập_nhật)`. Thuần Excel, không mạng.
        """
        dan = TrangPromptVisuals._doc_sheet(duong, "characters")
        canh = TrangPromptVisuals._doc_sheet(duong, "scenes")
        goc = goc_cua_id(ma_id)
        mau = next((c for c in dan if goc_cua_id(str(c.get("id"))) == goc), None)
        duoi_style = ""
        if mau and DUOI_CHAN_DUNG in str(mau.get("sheet_prompt") or ""):
            duoi_style = str(mau["sheet_prompt"]).split(DUOI_CHAN_DUNG, 1)[1]
        truoc = {str(s.get("scene_id")): str(s.get("img_prompt") or "") for s in canh}
        doi_thiet_ke_nhan_vat(canh, dan, ma_id, english_prompt_moi, duoi_style=duoi_style)
        sua_dan = {str(c["id"]): {"english_prompt": c["english_prompt"], "sheet_prompt": c["sheet_prompt"]}
                   for c in dan if goc_cua_id(str(c.get("id"))) == goc}
        TrangPromptVisuals._ghi_cot_vao_xlsx(duong, "characters", "id", sua_dan)
        sua_canh = {str(s["scene_id"]): (s["img_prompt"], str(s.get("video_prompt") or ""))
                    for s in canh if str(s.get("img_prompt") or "") != truoc.get(str(s.get("scene_id")))}
        if sua_canh:
            TrangPromptVisuals._ghi_prompt_vao_xlsx(duong, sua_canh)
        return [(i, v["sheet_prompt"]) for i, v in sua_dan.items()], len(sua_canh)

    def _bao_du_tham_chieu(self) -> None:
        if self._tc_dang_cho:
            return
        if self._tc_thieu:
            self._ghi("CHƯA ĐỦ ảnh tham chiếu — còn thiếu: {0}. Chạy phim lúc này thì các "
                      "cảnh có {0} sẽ mỗi cảnh một kiểu.".format(", ".join(self._tc_thieu)))
            return
        self._ghi("Đủ ảnh tham chiếu. Excel đã trỏ tới ảnh thật — sang tab Ảnh & "
                  "Video → Hàng loạt là nhân vật/bối cảnh giữ nguyên cả phim.")

    def _nhan_anh_tham_chieu(self, khoa: str, du_lieu) -> None:
        """Một ảnh tham chiếu xong → chép thành `<id>.png` cạnh Excel + ghi đường."""
        import shutil  # noqa: PLC0415

        ma_id, duong_xlsx, _prompt = self._tc_dang_cho.pop(khoa)
        files = list(getattr(du_lieu, "files", ()) or ())
        if not files:
            self._tc_thieu.append(ma_id)
            self._ghi("Ảnh tham chiếu {0}: máy chủ báo xong nhưng không có tệp.".format(ma_id))
            self._bao_du_tham_chieu()
            return
        dich = os.path.join(os.path.dirname(duong_xlsx), "{0}.png".format(ma_id))
        try:
            shutil.copyfile(files[0], dich)
        except OSError as loi:
            self._ghi("Ảnh tham chiếu {0}: không chép được ({1})".format(ma_id, loi))
            return
        loi = self._ghi_duong_tham_chieu(duong_xlsx, {"{0}.png".format(ma_id): dich})
        self._ghi("Ảnh tham chiếu {0} xong{1}.".format(
            ma_id, " — lỗi ghi Excel: " + loi if loi else ", đã gắn vào các cảnh dùng nó"))
        self._bao_du_tham_chieu()

    def _dung(self) -> None:
        if self._huy is not None:
            self._huy.set()
        self._ghi("Đã yêu cầu dừng…")

    def _xong(self, ds: List[str]) -> None:
        self._da_xong = list(ds)
        self._nut_dung.hide()
        self._nut_chay.show()
        self._thanh.setValue(100 if ds else 0)
        self._thanh.hide()
        self._nap_xem(ds)
        if ds:
            # Không bật hộp thoại: Bước 4 và 5 vừa hiện ra ngay dưới là câu trả
            # lời; một hộp "OK" chắn giữa chỉ bắt khách bấm thêm một lần.
            self._ghi("Xong: {0} file Excel — xem, sửa prompt và thử vài cảnh "
                      "thật ở Bước 4.".format(len(ds)))
        else:
            self._ghi("Không tạo được file nào — xem nhật ký bên dưới.")
            if not self._log.isVisible():
                self._bat_tat(self._log, self._nut_log, "Nhật ký")

    def _hong(self, loi: BaseException) -> None:
        self._nut_dung.hide()
        self._nut_chay.show()
        self._thanh.hide()
        self._app.show_error(loi)

    # ── Kịch bản ─────────────────────────────────────────────────────────────

    def _nap_kich_ban(self) -> None:
        """Nạp kịch bản từ một file .txt vào ô dán."""
        duong, _ = QFileDialog.getOpenFileName(
            self, "Chọn file kịch bản", "",
            "Văn bản (*.txt);;Tất cả (*.*)")
        if not duong:
            return
        try:
            with open(duong, "r", encoding="utf-8-sig") as f:
                self._o_kich_ban.setPlainText(f.read())
            self._ghi("Đã nạp kịch bản từ {0}.".format(os.path.basename(duong)))
        except Exception as loi:  # noqa: BLE001 — đọc file lỗi không được chặn tab
            self._ghi("Không đọc được file kịch bản: {0}".format(str(loi)[:120]))
            self._app.show_error(loi)

    def _xoa_kich_ban(self) -> None:
        self._o_kich_ban.clear()

    # ── Nhật ký ──────────────────────────────────────────────────────────────

    def _ghi(self, dong: str) -> None:
        self._log.appendPlainText(dong)
        # Câu mới nhất lên dòng trạng thái ở Bước 3 — bỏ thụt đầu dòng của log.
        self._dong_trang_thai.setText(dong.strip()[:160])

    def _ghi_nen(self, dong: str) -> None:
        self._app.goi_tren_luong_ve(lambda: self._ghi(dong))

    def _tien_nen(self, phan: float) -> None:
        gia_tri = max(0, min(100, int(round(phan * 100))))
        self._app.goi_tren_luong_ve(lambda: self._thanh.setValue(gia_tri))

    def doi_du_an(self, _ten: str) -> None:
        self._thu_muc.dat(self._app.default_output_dir("prompt-visuals"))


def _ma_artifact(gia_tri) -> str:
    """`put_file` trả về `Artifact` hoặc chuỗi tuỳ bản — nhận cả hai."""
    return str(getattr(gia_tri, "artifact_id", "") or gia_tri)


def _ma_an_toan(ten: str) -> str:
    """Đổi tên file thành mã workflow hợp lệ (chỉ chữ, số, `.`, `_`, `-`).

    Mã này cũng là tên tệp điểm dừng, nên mỗi file giọng đọc một mã riêng —
    dùng chung là file sau đè điểm dừng của file trước.

    Luật thật nằm ở `core.workflow.ma_an_toan`, cạnh chính cái biểu thức đi
    soi nó. Trước 26/08/2026 chỗ này tự lọc bằng `str.isalnum()` — mà hàm ấy
    hiểu cả Unicode, nên chữ tiếng Việt lọt qua rồi đâm vào bộ soi chỉ nhận
    ASCII: khách đặt tên tệp `001_Đoạn 1_1.mp3` là không tạo được file nào.
    """
    from core.workflow import ma_an_toan  # noqa: PLC0415

    return ma_an_toan(os.path.splitext(os.path.basename(ten))[0])
